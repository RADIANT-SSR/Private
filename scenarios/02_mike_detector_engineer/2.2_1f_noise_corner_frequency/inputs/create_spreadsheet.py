"""Create Mike's detector datasheet for 1/f noise analysis.

Three sheets:
  1. "Detector Specs"      — LWIR HgCdTe vendor datasheet
  2. "1/f Characterization" — measured 1/f noise parameters
  3. "System Configuration" — optics, spectral, geometry for NEDT evaluation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=4):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Detector Specs
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Detector Specs"
ws1["A1"] = "LWIR HgCdTe Staring Array — Vendor Datasheet"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "640x512 LWIR FPA, Lot L2025-014, cutoff 10.5 µm"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 50

row = 4
row = add_section(ws1, row, "Detector Parameters")
row = add_param(ws1, row, "Format", "640x512", "pixels", "Staring FPA")
row = add_param(ws1, row, "Pixel pitch", 24.0, "µm", "Square pixels")
row = add_param(ws1, row, "Cutoff wavelength", 10.5, "µm", "At 77 K operating temperature")
row = add_param(ws1, row, "Quantum efficiency (in-band avg)", 55.0, "%",
                "Measured at 77 K, 8.0–10.0 µm average")
row = add_param(ws1, row, "Dark current", 5.0e5, "e⁻/s",
                "At 77 K, typical for LWIR HgCdTe")
row = add_param(ws1, row, "Operating temperature", 77.0, "K",
                "Stirling cooler, ±0.05 K stability")
row = add_param(ws1, row, "Read noise (post-CDS)", 350.0, "e⁻ RMS",
                "Measured in dark, CDS mode")
row = add_param(ws1, row, "Full well capacity", 2.0e7, "e⁻",
                "90% linearity limit, large-format LWIR ROIC")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "14-bit ADC")
row = add_param(ws1, row, "System gain", 1220.0, "e⁻/DN",
                "Set for 100% FWC at ADC max")
row = add_param(ws1, row, "IPC coupling", 0.8, "%",
                "Per-neighbor, measured via single-pixel reset")

row += 1
row = add_section(ws1, row, "Timing Parameters")
row = add_param(ws1, row, "Nominal frame rate", 60.0, "Hz", "Baseline staring mode")
row = add_param(ws1, row, "Nominal integration time", 0.1, "ms",
                "100 µs — FWC-limited for 300 K LWIR scene")
row = add_param(ws1, row, "Dead time per frame", 16.57, "ms",
                "Readout + reset (16.67 ms frame − 0.1 ms int)")

# ---------------------------------------------------------------------------
# Sheet 2: 1/f Characterization
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("1-f Characterization")
ws2["A1"] = "1/f Noise Characterization — Lab Measurements"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Measured via temporal noise PSD analysis, 10,000 dark frames at 120 Hz"
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 55

row = 4
row = add_section(ws2, row, "1/f Noise Parameters")
row = add_param(ws2, row, "Flicker coefficient K", 2.5e4, "e⁻²",
                "From PSD fit: S(f) = K/f for f < f_corner")
row = add_param(ws2, row, "Corner frequency", 200.0, "Hz",
                "Where 1/f PSD meets white noise floor")
row = add_param(ws2, row, "White noise floor", 350.0, "e⁻ RMS",
                "Read noise plateau above f_corner")
row = add_param(ws2, row, "Measurement method", "PSD fit", "—",
                "Welch periodogram, 1024-pt FFT, Hann window")
row = add_param(ws2, row, "Measurement temperature", 77.0, "K",
                "Cold shutter closed (dark frames)")

row += 1
row = add_section(ws2, row, "Mike's Notes on 1/f vs. Frame Rate")
ws2.cell(row=row, column=1,
         value="For a staring array at frame rate f_frame:").border = thin_border
row += 1
ws2.cell(row=row, column=1,
         value="  f_low = 1/t_total = f_frame (lowest temporal frequency "
               "in the frame stream)").border = thin_border
row += 1
ws2.cell(row=row, column=1,
         value="  f_high = 1/(2·t_int) (Nyquist frequency within one "
               "integration)").border = thin_border
row += 1
ws2.cell(row=row, column=1,
         value="  σ_1f = √(K · ln(f_high / f_low))").border = thin_border
row += 1
ws2.cell(row=row, column=1,
         value="  At 30 Hz, more low-frequency noise is included → "
               "higher σ_1f").border = thin_border
row += 1
ws2.cell(row=row, column=1,
         value="  At 120 Hz, f_low is higher → less 1/f content → "
               "lower σ_1f").border = thin_border

row += 2
row = add_section(ws2, row, "Test Frame Rates")
headers = ["Frame Rate [Hz]", "t_frame [ms]", "t_int [ms]",
           "f_low [Hz]", "f_high [Hz]", "Notes"]
for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 30
row += 1

frame_rates = [
    (30.0, 33.33, 0.1, 30.0, 5000.0, "Slow scan — worst 1/f"),
    (60.0, 16.67, 0.1, 60.0, 5000.0, "Nominal operating mode"),
    (120.0, 8.33, 0.1, 120.0, 5000.0, "Fast scan — least 1/f"),
]
for fr, tf, ti, fl, fh, note in frame_rates:
    ws2.cell(row=row, column=1, value=fr).border = thin_border
    ws2.cell(row=row, column=2, value=tf).border = thin_border
    ws2.cell(row=row, column=3, value=ti).border = thin_border
    ws2.cell(row=row, column=4, value=fl).border = thin_border
    ws2.cell(row=row, column=5, value=fh).border = thin_border
    ws2.cell(row=row, column=6, value=note).border = thin_border
    row += 1

# ---------------------------------------------------------------------------
# Sheet 3: System Configuration
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("System Configuration")
ws3["A1"] = "System Configuration — LWIR Staring Imager"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Ground-based surveillance system, horizontal path"
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 50

row = 4
row = add_section(ws3, row, "Optics")
row = add_param(ws3, row, "Aperture diameter", 15.0, "cm",
                "Germanium objective lens")
row = add_param(ws3, row, "Focal length", 30.0, "cm", "f/2 for LWIR throughput")
row = add_param(ws3, row, "f-number", 2.0, "—", "Fast optics for LWIR sensitivity")
row = add_param(ws3, row, "Optical transmission", 85.0, "%",
                "Ge lens + ZnSe window, AR-coated")
row = add_param(ws3, row, "Optics temperature", 20.0, "°C", "Ambient ground-based")
row = add_param(ws3, row, "Number of optical elements", 3, "—",
                "Objective, field lens, window")

row += 1
row = add_section(ws3, row, "Spectral Band")
row = add_param(ws3, row, "Filter cut-on", 8000.0, "nm", "Cold filter at 77 K")
row = add_param(ws3, row, "Filter cut-off", 10000.0, "nm", "Cold filter at 77 K")
row = add_param(ws3, row, "Band center", 9000.0, "nm", "Arithmetic center")

row += 1
row = add_section(ws3, row, "Scene / Target")
row = add_param(ws3, row, "Target temperature", 300.0, "K",
                "Ambient outdoor scene")
row = add_param(ws3, row, "Target emissivity", 0.95, "—", "Painted metal surface")
row = add_param(ws3, row, "Background temperature", 295.0, "K",
                "Ambient background (sky + terrain)")
row = add_param(ws3, row, "Background emissivity", 0.97, "—", "Natural terrain")

from pathlib import Path
out = Path(__file__).parent / "mike_1f_noise_data.xlsx"
wb.save(out)
print(f"Created {out}")
