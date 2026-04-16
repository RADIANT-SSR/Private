"""Create Tom's pixel pitch trade study spreadsheet.

Tom is an optical designer evaluating detector pixel pitch for a MWIR
pushbroom imager. He has:
  - An optical design output from Zemax (units: mm for lengths, waves for WFE)
  - Candidate detector specs from three vendors (different pixel pitches)
  - System-level requirements from the mission spec

All values are in the units an optical designer typically works with:
  - Focal length in mm (Zemax convention)
  - Aperture in cm (system spec)
  - Wavelengths in nm (spectral design)
  - Angular resolution in arc-seconds
  - WFE in waves at reference wavelength
  - Pixel pitch in µm (detector vendor convention)
  - Dark current in e⁻/s (already characterized)
  - Integration time in ms
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=value).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value=unit).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ============================================================================
# Sheet 1: Optical Design Summary
# ============================================================================
# This mimics the output from Tom's Zemax session. Zemax uses mm for lengths.

ws1 = wb.active
ws1.title = "Optical Design Summary"

ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 15
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 55

ws1["A1"] = "MWIR Pushbroom Imager — Optical Design Summary"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "Tom Chen — Zemax OpticStudio Export, April 2026"
ws1["A2"].font = Font(italic=True, size=10)
ws1["A3"] = "Configuration: Ritchey-Chrétien, f/4, 30 cm aperture"
ws1["A3"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws1.cell(row=5, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 6

# Telescope parameters — Zemax convention: mm
row = add_section(ws1, row, "Telescope Design")
row = add_param(ws1, row, "Entrance pupil diameter", 300, "mm",
                "Circular, unobscured equivalent")
row = add_param(ws1, row, "Effective focal length", 1200, "mm",
                "Measured at 4.25 µm reference wavelength")
row = add_param(ws1, row, "f-number (working)", 4.0, "—",
                "f/# = EFL / EPD")
row = add_param(ws1, row, "Back focal distance", 185, "mm",
                "Pupil to focal plane")
row = add_param(ws1, row, "Field of view (full)", 2.4, "deg",
                "Cross-track, 640 pixels at 18 µm")
row = add_param(ws1, row, "Optical transmission", 72, "%",
                "Including 2 mirrors (Al + protective), cold filter")
row = add_param(ws1, row, "Optics temperature", 293, "K",
                "Ambient, uncooled barrel")
row = add_param(ws1, row, "WFE (RMS, on-axis)", 0.05, "waves",
                "At λ_ref = 4.25 µm; Zemax merit function result")
row = add_param(ws1, row, "WFE reference wavelength", 4250, "nm",
                "Design reference wavelength")
row = add_param(ws1, row, "Airy disk diameter", 41.6, "µm",
                "2.44 × λ × f/# at λ = 4.25 µm; first dark ring")

# Spectral design — nm convention from filter vendor
row = add_section(ws1, row, "Spectral Design")
row = add_param(ws1, row, "Band center wavelength", 4250, "nm",
                "MWIR thermal imaging band")
row = add_param(ws1, row, "Filter cut-on", 3500, "nm",
                "Cold filter, 50% transmission edge")
row = add_param(ws1, row, "Filter cut-off", 5000, "nm",
                "Cold filter, 50% transmission edge")

# Scene / target
row = add_section(ws1, row, "Scene (Nominal)")
row = add_param(ws1, row, "Target temperature", 310, "K",
                "Ground target, daytime")
row = add_param(ws1, row, "Target emissivity", 0.95, "—", "Vegetation/soil")
row = add_param(ws1, row, "Background temperature", 295, "K",
                "Surrounding terrain")
row = add_param(ws1, row, "Background emissivity", 0.95, "—", "")

# Geometry
row = add_section(ws1, row, "Observation Geometry")
row = add_param(ws1, row, "Orbit altitude", 500, "km", "LEO, sun-synchronous")
row = add_param(ws1, row, "Ground sample distance (goal)", "< 10", "m",
                "Mission requirement, drives pixel pitch selection")

# ============================================================================
# Sheet 2: Candidate Detectors
# ============================================================================
# Tom is comparing detectors from three vendors, each with different pixel
# pitches. This is the core of the trade study.

ws2 = wb.create_sheet("Candidate Detectors")

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 12

ws2["A1"] = "Candidate Detector Comparison"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "All values from vendor datasheets"
ws2["A2"].font = Font(italic=True, size=10)

# Column headers: Parameter + one column per candidate pitch
pitches = [8, 12, 15, 18, 24, 30]
pitch_labels = [f"{p} µm" for p in pitches]

det_headers = ["Parameter"] + pitch_labels + ["Unit"]
for col, h_text in enumerate(det_headers, 1):
    cell = ws2.cell(row=4, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# Detector parameters vary with pixel pitch (realistic vendor data)
# Smaller pixels: lower FWC, lower dark current, similar QE, lower read noise
det_params = [
    # (Parameter, [values per pitch], unit, notes)
    ("Pixel pitch", pitches, "µm"),
    ("Array format (cross-track)", [2048, 1280, 1024, 640, 512, 320], "pixels"),
    ("Quantum efficiency (in-band)", [70, 72, 74, 72, 70, 68], "%"),
    ("Dark current", [15, 25, 35, 50, 80, 120], "e⁻/s"),
    ("Full well capacity", [40000, 120000, 250000, 500000, 1200000, 2500000], "e⁻"),
    ("Read noise (CDS)", [8, 12, 16, 18, 22, 28], "e⁻ RMS"),
    ("ADC resolution", [14, 14, 14, 14, 14, 14], "bits"),
    ("System gain", [0.5, 1.0, 1.5, 2.0, 3.0, 5.0], "e⁻/DN"),
    ("Integration time (nom.)", [2, 2, 2, 2, 2, 2], "ms"),
    ("Q parameter (at 4.25 µm)", [2.13, 1.42, 1.13, 0.94, 0.71, 0.57], "—"),
]

for i, (param_name, values, unit) in enumerate(det_params, 5):
    ws2.cell(row=i, column=1, value=param_name).border = thin_border
    for j, val in enumerate(values):
        cell = ws2.cell(row=i, column=j + 2, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=len(values) + 2, value=unit).border = thin_border

# Add a note about Q
note_row = len(det_params) + 6
ws2.cell(row=note_row, column=1,
         value="Q = λ·f/# / p  (sampling parameter)").font = Font(italic=True)
ws2.cell(row=note_row + 1, column=1,
         value="Q > 1: oversampled, Q = 1: critically sampled, Q < 1: undersampled"
         ).font = Font(italic=True)
ws2.cell(row=note_row + 2, column=1,
         value="Airy disk diameter = 2.44 × λ × f/# = 41.6 µm at 4.25 µm"
         ).font = Font(italic=True)

# ============================================================================
# Sheet 3: Performance Requirements
# ============================================================================

ws3 = wb.create_sheet("Performance Requirements")

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 50

ws3["A1"] = "System Performance Requirements"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Source: Mission Requirements Document MRD-2025-017"
ws3["A2"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Requirement", "Threshold", "Unit", "Notes"], 1):
    cell = ws3.cell(row=4, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

requirements = [
    ("GSD (nadir)", "< 10", "m",
     "Ground sample distance at nadir, drives pixel pitch"),
    ("System MTF at Nyquist", "≥ 0.10", "—",
     "Minimum acceptable MTF for image interpretability"),
    ("SNR (extended, 310 K target)", "≥ 100", "—",
     "At nominal integration time"),
    ("EE 1×1 (ensquared energy)", "≥ 0.30", "—",
     "Minimum energy in single pixel for point source detection"),
    ("Q parameter", "0.5 – 2.0", "—",
     "Design guidance, not hard requirement; Q < 0.5 loses too much information"),
    ("Sampling regime", "avoid Q < 0.7", "—",
     "Strong aliasing below Q = 0.7 degrades image quality"),
]

for i, (req, thresh, unit, notes) in enumerate(requirements, 5):
    ws3.cell(row=i, column=1, value=req).border = thin_border
    ws3.cell(row=i, column=2, value=thresh).border = thin_border
    ws3.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws3.cell(row=i, column=3, value=unit).border = thin_border
    ws3.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    ws3.cell(row=i, column=4, value=notes).border = thin_border

output_path = "tom_pixel_pitch_trade.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
