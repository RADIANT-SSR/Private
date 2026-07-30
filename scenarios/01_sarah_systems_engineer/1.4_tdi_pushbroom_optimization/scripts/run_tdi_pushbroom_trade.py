"""Scenario 1.4: TDI Pushbroom Optimization — Line Rate vs. SNR

Sarah is designing a VNIR panchromatic pushbroom imager for a 500 km SSO.
Ground velocity constrains the per-line integration time to ~0.2 ms.
She needs TDI to build sufficient SNR and sweeps N_tdi from 1 to 128
to find where NIIRS peaks before well saturation clips the signal.

Note: the original concept specified MWIR, but MWIR thermal scenes produce
so much signal that the well saturates at N_tdi=1. TDI is primarily a VNIR
technology — reflected solar signal is orders of magnitude weaker per pixel
than MWIR thermal emission, making TDI essential for pushbroom VNIR imagers.

Approach:
  Run RADIANT at each N_tdi value. The readout stage handles TDI signal
  scaling (×N) and noise scaling (shot ×√N, read ×1 for analog TDI).
  Integration time per line is derived from orbital mechanics:
    v_ground = v_orbital × R_E / (R_E + h)
    line_period = GSD / v_ground

  Smear MTF (1 pixel/line) and TDI misalignment MTF are computed
  analytically and applied as corrections to RADIANT's system MTF.

Physics:
  - Analog TDI: charge accumulated on-chip, single readout → read noise
    constant regardless of N_tdi. This is the classic √N SNR advantage.
  - Signal grows as N_tdi × signal_per_line
  - Shot noise grows as √N_tdi (independent Poisson across stages)
  - SNR improves as √N_tdi (shot-limited) or N_tdi (read-limited)
  - Well saturation limits maximum useful N_tdi
  - Smear MTF = |sinc(π·f·pitch)| at Nyquist ≈ 0.637 (constant, all N_tdi)
  - TDI misalignment MTF degrades with √N_tdi accumulated registration error

Usage:
    python run_tdi_pushbroom_trade.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt

from radiant.api import Sensor


def main() -> None:
    """Run the scenario analysis."""
    # ---------------------------------------------------------------------------
    # Step 1: Read Sarah's spreadsheet
    # ---------------------------------------------------------------------------

    INPUT_FILE = Path(__file__).parent.parent / "inputs" / "sarah_tdi_pushbroom_data.xlsx"
    OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "tdi_pushbroom_results.xlsx"
    PLOT_DIR = Path(__file__).parent.parent / "outputs"

    wb_in = openpyxl.load_workbook(INPUT_FILE)

    ws_sys = wb_in["System Parameters"]
    specs: dict[str, object] = {}
    units: dict[str, str] = {}
    for row in ws_sys.iter_rows(min_row=2, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            try:
                specs[name] = float(value)
            except (ValueError, TypeError):
                specs[name] = value
            units[name] = str(row[2].value) if row[2].value else "--"

    ws_tdi = wb_in["TDI Sweep"]
    tdi_values: list[int] = []
    for row in ws_tdi.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            tdi_values.append(int(row[0]))

    print("=" * 80)
    print("SCENARIO 1.4: TDI Pushbroom Optimization — Line Rate vs. SNR")
    print("=" * 80)

    print("\n=== System Parameters (from spreadsheet) ===")
    for k, v in specs.items():
        print(f"  {k:<30s}: {v} [{units.get(k, '--')}]")

    print(f"\n=== TDI Sweep Points ===")
    print(f"  N_tdi = {tdi_values}")

    # ---------------------------------------------------------------------------
    # Step 2: Convert to RADIANT canonical units and derive orbital parameters
    # ---------------------------------------------------------------------------

    aperture_m = float(specs["Aperture diameter"]) / 100.0         # cm -> m
    focal_length_m = float(specs["Focal length"]) / 100.0          # cm -> m
    f_number = float(specs["f-number"])
    transmission = float(specs["Optical transmission"]) / 100.0    # % -> frac
    optics_temp_K = float(specs["Optics temperature"]) + 273.15    # C -> K
    wfe_waves = float(specs["WFE RMS"])
    obscuration = float(specs["Central obscuration"]) / 100.0      # % -> frac

    pixel_pitch_um = float(specs["Pixel pitch"])
    pixel_pitch_m = pixel_pitch_um * 1e-6
    qe = float(specs["Quantum efficiency"]) / 100.0                # % -> frac
    dark_rate = float(specs["Dark current"])
    read_noise = float(specs["Read noise"])
    fwc = float(specs["Full well capacity"])
    gain = float(specs["Gain"])
    adc_bits = int(specs["ADC bits"])

    band_min_nm = float(specs["Filter min"])
    band_max_nm = float(specs["Filter max"])
    band_min_um = band_min_nm / 1000.0                             # nm -> um
    band_max_um = band_max_nm / 1000.0
    band_center_um = (band_min_um + band_max_um) / 2.0

    target_refl = float(specs["Target reflectance"])
    bg_refl = float(specs["Background reflectance"])
    solar_zenith_deg = float(specs["Solar zenith angle"])
    solar_zenith_rad = solar_zenith_deg * math.pi / 180.0

    altitude_km = float(specs["Orbit altitude"])
    altitude_m = altitude_km * 1000.0
    v_orbital = float(specs["Orbital velocity"])
    visibility_km = float(specs["Visibility"])
    pwv_mm = float(specs["PWV"])

    tdi_misalign_pix = float(specs["TDI misalignment"])

    # Derived orbital parameters
    R_earth_m = 6_371_000.0
    v_ground = v_orbital * R_earth_m / (R_earth_m + altitude_m)
    gsd_m = pixel_pitch_m * altitude_m / focal_length_m
    line_period_s = gsd_m / v_ground
    ifov_urad = pixel_pitch_m / focal_length_m * 1e6
    f_nyquist = 1.0 / (2.0 * pixel_pitch_m)  # cycles/m on focal plane
    Q = band_center_um * f_number / pixel_pitch_um
    airy_diam_um = 2.44 * band_center_um * f_number

    # Smear MTF at Nyquist (1 pixel of smear per line — constant for all N_tdi)
    smear_mtf_nyquist = abs(np.sinc(f_nyquist * pixel_pitch_m))

    print(f"\n=== Converted to RADIANT Canonical Units ===")
    print(f"  {'Parameter':<30s} {'Value':>14s}  {'Unit':<10s}  {'Conversion'}")
    print(f"  {'-' * 30} {'-' * 14}  {'-' * 10}  {'-' * 20}")
    print(f"  {'Aperture diameter':<30s} {aperture_m:>14.4f}  {'m':<10s}  cm / 100")
    print(f"  {'Focal length':<30s} {focal_length_m:>14.4f}  {'m':<10s}  cm / 100")
    print(f"  {'Optical transmission':<30s} {transmission:>14.4f}  {'fraction':<10s}  % / 100")
    print(f"  {'Optics temperature':<30s} {optics_temp_K:>14.2f}  {'K':<10s}  C + 273.15")
    print(f"  {'Band':<30s} {band_min_um:>6.3f}-{band_max_um:<6.3f}  {'um':<10s}  nm / 1000")
    print(f"  {'QE':<30s} {qe:>14.4f}  {'fraction':<10s}  % / 100")
    print(f"  {'Obscuration':<30s} {obscuration:>14.4f}  {'fraction':<10s}  % / 100")
    print(f"  {'Solar zenith':<30s} {solar_zenith_rad:>14.4f}  {'rad':<10s}  deg × π/180")

    print(f"\n=== Derived Orbital / Timing Parameters ===")
    print(f"  Ground velocity:     {v_ground:.1f} [m/s]")
    print(f"  GSD:                 {gsd_m:.2f} [m]")
    print(f"  Line period:         {line_period_s*1e3:.4f} [ms]  ({line_period_s*1e6:.1f} [us])")
    print(f"  IFOV:                {ifov_urad:.1f} [urad]")
    print(f"  f_Nyquist:           {f_nyquist:.0f} [cy/m] ({f_nyquist/1000:.1f} [cy/mm])")
    print(f"  Q (sampling):        {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")
    print(f"  Airy disk:           {airy_diam_um:.1f} [um] ({airy_diam_um/pixel_pitch_um:.2f} pixels)")
    print(f"  Smear MTF@Nyquist:   {smear_mtf_nyquist:.4f} [--] (1 pixel/line, constant)")
    print(f"  TDI misalignment:    {tdi_misalign_pix:.2f} [pixels/stage]")

    # ---------------------------------------------------------------------------
    # Step 3: Build base RADIANT config (N_tdi set per sweep point)
    # ---------------------------------------------------------------------------

    # VNIR reflected solar scene — Kirchhoff: reflectance = 1 - emissivity
    # For target reflectance 0.15 → emissivity = 0.85
    # Thermal emission at 300 K is negligible in VNIR, signal is from solar reflection
    target_temp_K = 300.0
    bg_temp_K = 295.0
    target_emissivity = 1.0 - target_refl   # Kirchhoff: ε = 1 - ρ
    bg_emissivity = 1.0 - bg_refl

    base_config = {
        "source": {
            "target": {"temperature": target_temp_K, "emissivity": target_emissivity},
            "background": {"temperature": bg_temp_K, "emissivity": bg_emissivity},
        },
        "atmosphere": {
            "model": "simple",
            "standard_atmosphere": "midlat_summer",
        },
        "geometry": {
            "sensor_altitude_m": altitude_m,
            "path_zenith_rad": 0.0,
            "solar_zenith_rad": solar_zenith_rad,
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
            "wfe_rms_waves": wfe_waves,
            "obscuration_ratio": obscuration,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": 293.0,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": line_period_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "full_well_capacity_e": fwc,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "tdi_mode": "analog",
        },
        "performance": {"niirs": {"allow_extrapolated": True}},
    }

    print(f"\n=== RADIANT Configuration ===")
    print(f"  Integration time per line: {line_period_s*1e3:.4f} [ms]")
    print(f"  Band: {band_min_um:.3f}–{band_max_um:.3f} [um]")
    print(f"  TDI mode: analog (single readout after charge accumulation)")
    print(f"  Target reflectance: {target_refl}")
    print(f"  Solar zenith: {solar_zenith_deg:.0f} [deg]")

    # ---------------------------------------------------------------------------
    # Step 4: Sweep N_tdi
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  TDI SWEEP")
    print(f"{'=' * 80}")
    print(f"\n  {'N_tdi':>6s}  {'Signal':>12s}  {'Well Fill':>10s}  {'SNR':>10s}  "
          f"{'MTF@Nyq':>10s}  {'RER':>8s}  {'NIIRS':>8s}  {'Status'}")
    print(f"  {'':>6s}  {'[e-]':>12s}  {'[%]':>10s}  {'[--]':>10s}  "
          f"{'[--]':>10s}  {'[--]':>8s}  {'[--]':>8s}  ")
    print(f"  {'-' * 6}  {'-' * 12}  {'-' * 10}  {'-' * 10}  "
          f"{'-' * 10}  {'-' * 8}  {'-' * 8}  {'-' * 10}")

    results: list[dict] = []

    for n_tdi in tdi_values:
        config = {k: ({**v} if isinstance(v, dict) else v) for k, v in base_config.items()}
        config["source"] = {
            "target": {**base_config["source"]["target"]},
            "background": {**base_config["source"]["background"]},
        }
        config["readout"] = {**base_config["readout"], "n_tdi": n_tdi}

        sensor = Sensor.from_dict(config)
        r = sensor.evaluate()

        signal_e = r.stage_outputs["readout"]["signal_e_final"]
        well_fill_pct = (signal_e / fwc) * 100.0
        well_status = r.stage_outputs["readout"]["well_status"]
        snr = r.metrics["snr"]
        mtf_nyq = r.metrics.get("mtf_at_nyquist", 0.0)
        rer = r.metrics.get("rer", 0.0)
        niirs = r.metrics.get("niirs", 0.0)

        # Noise budget
        noise_terms = {n.name: n.value_e for n in r.noise_terms}

        # TDI misalignment MTF — random per-stage: total = δ × √N_tdi
        total_misalign_pix = tdi_misalign_pix * math.sqrt(n_tdi)
        total_misalign_m = total_misalign_pix * pixel_pitch_m
        tdi_misalign_mtf_nyq = abs(np.sinc(f_nyquist * total_misalign_m))

        # True system MTF = RADIANT MTF × smear MTF × TDI misalignment MTF
        true_mtf_nyq = mtf_nyq * smear_mtf_nyquist * tdi_misalign_mtf_nyq

        status = "OK"
        if well_status == "saturated":
            status = "SATURATED"
        elif well_fill_pct > 80:
            status = "NEAR-SAT"

        row = {
            "n_tdi": n_tdi,
            "signal_e": signal_e,
            "well_fill_pct": well_fill_pct,
            "well_status": well_status,
            "snr": snr,
            "mtf_nyq_radiant": mtf_nyq,
            "smear_mtf_nyq": smear_mtf_nyquist,
            "tdi_misalign_mtf_nyq": tdi_misalign_mtf_nyq,
            "true_mtf_nyq": true_mtf_nyq,
            "rer": rer,
            "niirs": niirs,
            "noise_terms": noise_terms,
            "status": status,
            "total_misalign_pix": total_misalign_pix,
        }
        results.append(row)

        print(f"  {n_tdi:>6d}  {signal_e:>12,.0f}  {well_fill_pct:>9.1f}%  {snr:>10.1f}  "
              f"{mtf_nyq:>10.4f}  {rer:>8.4f}  {niirs:>8.2f}  {status}")

    # ---------------------------------------------------------------------------
    # Step 5: Noise budget table
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  NOISE BUDGET (all values in e- RMS)")
    print(f"{'=' * 80}")

    noise_names = list(results[0]["noise_terms"].keys())
    active_terms = [n for n in noise_names
                    if any(r["noise_terms"].get(n, 0) > 0.01 for r in results)]

    header = f"  {'Noise Term':<20s}"
    for r in results:
        header += f"  {'N=' + str(r['n_tdi']):>10s}"
    print(header)
    print(f"  {'-' * 20}" + f"  {'-' * 10}" * len(results))

    for term in active_terms:
        line = f"  {term:<20s}"
        for r in results:
            val = r["noise_terms"].get(term, 0.0)
            line += f"  {val:>10.1f}"
        print(line)

    line = f"  {'TOTAL (RSS)':<20s}"
    for r in results:
        total = math.sqrt(sum(v**2 for v in r["noise_terms"].values()))
        line += f"  {total:>10.1f}"
    print(line)

    # ---------------------------------------------------------------------------
    # Step 6: SNR scaling analysis
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  SNR SCALING ANALYSIS")
    print(f"{'=' * 80}")

    snr_1 = results[0]["snr"]
    print(f"\n  Baseline SNR (N_tdi=1): {snr_1:.1f} [--]")
    print(f"\n  {'N_tdi':>6s}  {'SNR':>10s}  {'SNR/SNR_1':>10s}  {'√N_tdi':>10s}  "
          f"{'N_tdi':>10s}  {'Regime'}")
    print(f"  {'-' * 6}  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 15}")

    for r in results:
        n = r["n_tdi"]
        snr_ratio = r["snr"] / snr_1 if snr_1 > 0 else 0
        sqrt_n = math.sqrt(n)

        if n > 1 and snr_ratio > 0:
            exponent = math.log(snr_ratio) / math.log(n)
            if exponent > 0.75:
                regime = "read-limited"
            elif exponent < 0.55:
                regime = "shot-limited"
            else:
                regime = "mixed"
        else:
            regime = "baseline"

        print(f"  {n:>6d}  {r['snr']:>10.1f}  {snr_ratio:>10.2f}  {sqrt_n:>10.2f}  "
              f"{float(n):>10.1f}  {regime}")

    # ---------------------------------------------------------------------------
    # Step 7: MTF budget
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  MTF BUDGET (at Nyquist = {f_nyquist:.0f} cy/m)")
    print(f"{'=' * 80}")

    print(f"\n  {'N_tdi':>6s}  {'MTF_opt':>10s}  {'MTF_smear':>10s}  {'MTF_misalign':>12s}  "
          f"{'MTF_sys':>10s}  {'Misalign':>10s}")
    print(f"  {'':>6s}  {'[--]':>10s}  {'[--]':>10s}  {'[--]':>12s}  "
          f"{'[--]':>10s}  {'[pixels]':>10s}")
    print(f"  {'-' * 6}  {'-' * 10}  {'-' * 10}  {'-' * 12}  {'-' * 10}  {'-' * 10}")

    for r in results:
        print(f"  {r['n_tdi']:>6d}  {r['mtf_nyq_radiant']:>10.4f}  "
              f"{r['smear_mtf_nyq']:>10.4f}  {r['tdi_misalign_mtf_nyq']:>12.4f}  "
              f"{r['true_mtf_nyq']:>10.4f}  {r['total_misalign_pix']:>10.2f}")

    print(f"\n  Notes:")
    print(f"  - MTF_opt: from RADIANT (optics + detector + aberrations)")
    print(f"  - MTF_smear: |sinc(π/2)| = 2/π ≈ {smear_mtf_nyquist:.4f} (1 pixel/line, constant)")
    print(f"  - MTF_misalign: |sinc(π·f·δ·√N)| where δ = {tdi_misalign_pix:.1f} pixels/stage")
    print(f"  - MTF_sys = MTF_opt × MTF_smear × MTF_misalign")

    # ---------------------------------------------------------------------------
    # Step 8: Saturation analysis
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  SATURATION ANALYSIS")
    print(f"{'=' * 80}")
    print(f"\n  Full well capacity: {fwc:,.0f} [e-]")
    print(f"  Signal per line (N_tdi=1): {results[0]['signal_e']:,.0f} [e-]")

    sig_1 = results[0]["signal_e"]
    if sig_1 > 0:
        max_useful_ntdi = fwc / sig_1
        print(f"  Theoretical max N_tdi (100% fill): {max_useful_ntdi:.1f}")
    else:
        print(f"  Signal per line is zero — check scene configuration")

    sat_ntdi = None
    for r in results:
        if r["well_fill_pct"] >= 100.0:
            sat_ntdi = r["n_tdi"]
            break

    if sat_ntdi:
        print(f"  First saturated N_tdi in sweep: {sat_ntdi}")
    else:
        print(f"  No saturation in sweep range")

    print(f"\n  {'N_tdi':>6s}  {'Signal':>12s}  {'Well Fill':>10s}  {'Headroom':>12s}  {'Status'}")
    print(f"  {'':>6s}  {'[e-]':>12s}  {'[%]':>10s}  {'[e-]':>12s}")
    print(f"  {'-' * 6}  {'-' * 12}  {'-' * 10}  {'-' * 12}  {'-' * 10}")

    for r in results:
        headroom = fwc - r["signal_e"]
        print(f"  {r['n_tdi']:>6d}  {r['signal_e']:>12,.0f}  {r['well_fill_pct']:>9.1f}%  "
              f"{headroom:>12,.0f}  {r['status']}")

    # ---------------------------------------------------------------------------
    # Step 9: NIIRS analysis
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  NIIRS ANALYSIS")
    print(f"{'=' * 80}")

    niirs_1 = results[0]["niirs"]
    print(f"\n  Baseline NIIRS (N_tdi=1): {niirs_1:.2f}")
    print(f"\n  {'N_tdi':>6s}  {'SNR':>10s}  {'RER':>8s}  {'NIIRS':>8s}  {'dNIIRS':>8s}  {'Note'}")
    print(f"  {'-' * 6}  {'-' * 10}  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 25}")

    best_niirs = -999.0
    best_ntdi = 1

    for r in results:
        d_niirs = r["niirs"] - niirs_1
        note = ""
        if r["niirs"] > best_niirs and r["status"] != "SATURATED":
            best_niirs = r["niirs"]
            best_ntdi = r["n_tdi"]
        if r["status"] == "SATURATED":
            note = "WELL SATURATED"
        elif r["well_fill_pct"] > 80:
            note = "near saturation"

        print(f"  {r['n_tdi']:>6d}  {r['snr']:>10.1f}  {r['rer']:>8.4f}  "
              f"{r['niirs']:>8.2f}  {d_niirs:>+8.2f}  {note}")

    print(f"\n  Peak NIIRS: {best_niirs:.2f} at N_tdi = {best_ntdi}")

    # ---------------------------------------------------------------------------
    # Step 10: Plots
    # ---------------------------------------------------------------------------

    fig_w, fig_h = 10, 7

    # Plot 1: SNR vs N_tdi
    fig1, ax1 = plt.subplots(figsize=(fig_w, fig_h))
    ntdi_arr = [r["n_tdi"] for r in results]
    snr_arr = [r["snr"] for r in results]
    ax1.semilogx(ntdi_arr, snr_arr, "bo-", linewidth=2, markersize=8, label="RADIANT SNR")

    n_arr = np.array(ntdi_arr, dtype=float)
    ax1.semilogx(ntdi_arr, [snr_1 * math.sqrt(n) for n in ntdi_arr],
                 "g--", linewidth=1.5, alpha=0.7, label=f"√N scaling (shot-limited)")
    ax1.semilogx(ntdi_arr, [snr_1 * n for n in ntdi_arr],
                 "r--", linewidth=1.5, alpha=0.7, label=f"N scaling (read-limited)")

    for r in results:
        if r["status"] == "SATURATED":
            ax1.axvline(r["n_tdi"], color="red", linestyle=":", alpha=0.5)

    ax1.set_xlabel("N_tdi [stages]", fontsize=12)
    ax1.set_ylabel("SNR [--]", fontsize=12)
    ax1.set_title("SNR vs. TDI Stages — VNIR Pushbroom, 500 km LEO", fontsize=13)
    ax1.legend(fontsize=10)
    ax1.grid(True, which="both", alpha=0.3)
    ax1.set_xticks(ntdi_arr)
    ax1.set_xticklabels([str(n) for n in ntdi_arr])
    fig1.tight_layout()
    fig1.savefig(PLOT_DIR / "fig1_snr_vs_ntdi.png", dpi=150)

    # Plot 2: Well fill and signal vs N_tdi
    fig2, (ax2a, ax2b) = plt.subplots(1, 2, figsize=(fig_w * 1.3, fig_h))

    signal_arr = [r["signal_e"] for r in results]
    well_fill_arr = [r["well_fill_pct"] for r in results]

    ax2a.semilogx(ntdi_arr, [s / 1000 for s in signal_arr], "rs-", linewidth=2, markersize=8)
    ax2a.axhline(fwc / 1000, color="red", linestyle="--", linewidth=1.5,
                 label=f"FWC = {fwc/1000:.0f} ke-")
    ax2a.set_xlabel("N_tdi [stages]", fontsize=12)
    ax2a.set_ylabel("Signal [ke-]", fontsize=12)
    ax2a.set_title("Signal vs. TDI Stages", fontsize=13)
    ax2a.legend(fontsize=10)
    ax2a.grid(True, which="both", alpha=0.3)
    ax2a.set_xticks(ntdi_arr)
    ax2a.set_xticklabels([str(n) for n in ntdi_arr])

    ax2b.semilogx(ntdi_arr, well_fill_arr, "go-", linewidth=2, markersize=8)
    ax2b.axhline(100, color="red", linestyle="--", linewidth=1.5, label="100% (saturation)")
    ax2b.axhline(80, color="orange", linestyle="--", linewidth=1.5, label="80% (design margin)")
    ax2b.set_xlabel("N_tdi [stages]", fontsize=12)
    ax2b.set_ylabel("Well Fill [%]", fontsize=12)
    ax2b.set_title("Well Fill vs. TDI Stages", fontsize=13)
    ax2b.legend(fontsize=10)
    ax2b.grid(True, which="both", alpha=0.3)
    ax2b.set_xticks(ntdi_arr)
    ax2b.set_xticklabels([str(n) for n in ntdi_arr])

    fig2.tight_layout()
    fig2.savefig(PLOT_DIR / "fig2_signal_well_fill.png", dpi=150)

    # Plot 3: Noise budget breakdown
    fig3, ax3 = plt.subplots(figsize=(fig_w, fig_h))

    dominant_terms = []
    for term in active_terms:
        max_val = max(r["noise_terms"].get(term, 0) for r in results)
        if max_val > 1.0:
            dominant_terms.append(term)

    x_pos = np.arange(len(ntdi_arr))
    bar_width = 0.8 / max(len(dominant_terms), 1)
    colors = plt.cm.Set2(np.linspace(0, 1, len(dominant_terms)))

    for i, term in enumerate(dominant_terms):
        vals = [r["noise_terms"].get(term, 0) for r in results]
        ax3.bar(x_pos + i * bar_width, vals, bar_width, label=term.replace("_", " "),
                color=colors[i], edgecolor="white", linewidth=0.5)

    ax3.set_xlabel("N_tdi [stages]", fontsize=12)
    ax3.set_ylabel("Noise [e- RMS]", fontsize=12)
    ax3.set_title("Noise Budget vs. TDI Stages", fontsize=13)
    ax3.set_xticks(x_pos + bar_width * len(dominant_terms) / 2)
    ax3.set_xticklabels([str(n) for n in ntdi_arr])
    ax3.legend(fontsize=9, ncol=2)
    ax3.grid(True, axis="y", alpha=0.3)
    fig3.tight_layout()
    fig3.savefig(PLOT_DIR / "fig3_noise_budget.png", dpi=150)

    # Plot 4: NIIRS vs N_tdi
    fig4, ax4 = plt.subplots(figsize=(fig_w, fig_h))
    niirs_arr = [r["niirs"] for r in results]
    ax4.semilogx(ntdi_arr, niirs_arr, "mo-", linewidth=2, markersize=8, label="NIIRS (RADIANT)")

    ax4.axhline(best_niirs, color="green", linestyle=":", alpha=0.5,
                label=f"Peak = {best_niirs:.2f} at N_tdi={best_ntdi}")

    for r in results:
        if r["status"] == "SATURATED":
            ax4.plot(r["n_tdi"], r["niirs"], "rx", markersize=15, markeredgewidth=3)

    ax4.set_xlabel("N_tdi [stages]", fontsize=12)
    ax4.set_ylabel("NIIRS [--]", fontsize=12)
    ax4.set_title("NIIRS vs. TDI Stages — Optimal TDI Selection", fontsize=13)
    ax4.legend(fontsize=10)
    ax4.grid(True, which="both", alpha=0.3)
    ax4.set_xticks(ntdi_arr)
    ax4.set_xticklabels([str(n) for n in ntdi_arr])
    fig4.tight_layout()
    fig4.savefig(PLOT_DIR / "fig4_niirs_vs_ntdi.png", dpi=150)

    # ---------------------------------------------------------------------------
    # Step 11: Output spreadsheet
    # ---------------------------------------------------------------------------

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "TDI Trade Results"

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "N_tdi [--]", "Signal [e-]", "Well Fill [%]", "SNR [--]",
        "MTF_opt@Nyq [--]", "Smear MTF@Nyq [--]", "Misalign MTF@Nyq [--]",
        "True MTF_sys@Nyq [--]", "Misalign [pixels]",
        "RER [--]", "NIIRS [--]", "dNIIRS [--]", "Status",
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results, start=2):
        d_niirs = r["niirs"] - niirs_1
        vals = [
            r["n_tdi"], round(r["signal_e"]), round(r["well_fill_pct"], 1),
            round(r["snr"], 1),
            round(r["mtf_nyq_radiant"], 4), round(r["smear_mtf_nyq"], 4),
            round(r["tdi_misalign_mtf_nyq"], 4),
            round(r["true_mtf_nyq"], 4), round(r["total_misalign_pix"], 2),
            round(r["rer"], 4), round(r["niirs"], 2), round(d_niirs, 2),
            r["status"],
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws_out.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if r["status"] == "SATURATED":
                cell.fill = PatternFill("solid", fgColor="FFD7D7")

    # Noise budget sheet
    ws_noise = wb_out.create_sheet("Noise Budget")
    noise_headers = ["Noise Term [e- RMS]"] + [f"N_tdi={r['n_tdi']}" for r in results]
    for col_idx, h in enumerate(noise_headers, start=1):
        cell = ws_noise.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, term in enumerate(active_terms, start=2):
        ws_noise.cell(row=row_idx, column=1, value=term)
        for col_idx, r in enumerate(results, start=2):
            ws_noise.cell(row=row_idx, column=col_idx,
                          value=round(r["noise_terms"].get(term, 0), 2))

    for ws in [ws_out, ws_noise]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb_out.save(OUTPUT_FILE)
    print(f"\n  Output spreadsheet: {OUTPUT_FILE}")

    # ---------------------------------------------------------------------------
    # Step 12: Summary
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  SUMMARY")
    print(f"{'=' * 80}")
    print(f"\n  System: VNIR pushbroom, {aperture_m*100:.0f} cm aperture, f/{f_number:.0f}, "
          f"{pixel_pitch_um:.0f} um pitch")
    print(f"  Band:   {band_min_nm:.0f}–{band_max_nm:.0f} nm panchromatic")
    print(f"  Orbit:  {altitude_km:.0f} km LEO, v_ground = {v_ground:.0f} m/s")
    print(f"  GSD:    {gsd_m:.2f} m, line period = {line_period_s*1e3:.4f} ms")
    print(f"  FWC:    {fwc:,.0f} e-")
    print(f"  Signal per line: {results[0]['signal_e']:,.0f} e- "
          f"({results[0]['well_fill_pct']:.1f}% well fill)")
    print(f"\n  Optimal N_tdi: {best_ntdi} (NIIRS = {best_niirs:.2f})")
    print(f"  SNR at optimal: {[r for r in results if r['n_tdi'] == best_ntdi][0]['snr']:.1f} [--]")
    print(f"  SNR improvement from N_tdi=1: "
          f"{[r for r in results if r['n_tdi'] == best_ntdi][0]['snr'] / snr_1:.1f}×")

    if sat_ntdi:
        unsaturated = [r for r in results if r["status"] != "SATURATED"]
        if unsaturated:
            print(f"\n  Saturation onset: N_tdi = {sat_ntdi}")
            print(f"  Max N_tdi before saturation: {unsaturated[-1]['n_tdi']}")

    print(f"\n  Design recommendation:")
    print(f"    Use N_tdi = {best_ntdi} for peak NIIRS = {best_niirs:.2f}.")
    safe = [r for r in results if r["well_fill_pct"] < 80 and r["status"] != "SATURATED"]
    if safe:
        best_safe = max(safe, key=lambda r: r["niirs"])
        print(f"    With 80% well-fill margin: N_tdi = {best_safe['n_tdi']} "
              f"(NIIRS = {best_safe['niirs']:.2f}, well fill = {best_safe['well_fill_pct']:.0f}%)")

    print(f"\n  Why not MWIR?")
    print(f"    MWIR thermal scenes (300+ K) produce ~250,000+ e- per line at this")
    print(f"    aperture and integration time. The well saturates at N_tdi=1, making")
    print(f"    TDI counterproductive. TDI is essential for VNIR pushbroom where")
    print(f"    reflected solar provides only {results[0]['signal_e']:,.0f} e- per line.")

    plt.show()


if __name__ == "__main__":
    main()
