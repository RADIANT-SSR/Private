"""Scenario 5.4: Jitter Tolerance — Line-of-Sight Stability Requirements

Tom needs to derive the jitter requirement for his VNIR panchromatic imager.
He sweeps jitter RMS from 0 to 5 µrad and finds where NIIRS drops by 0.5
and 1.0 full grade.

Approach:
  Run RADIANT at each jitter level using the built-in PlatformStage.
  The platform.jitter_rms_urad parameter feeds into PlatformStage, which
  convolves a Gaussian jitter kernel into the EffectivePSF. PerformanceStage
  then computes MTF, RER, and NIIRS from the jitter-degraded PSF.

  Jitter does NOT affect SNR — it blurs the image (degrades MTF/RER) but
  doesn't add or remove photons. NIIRS changes through the RER term only.

Usage:
    python run_jitter_tolerance.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import matplotlib.pyplot as plt

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Tom's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "tom_jitter_tolerance_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "jitter_tolerance_results.xlsx"
PLOT_DIR = Path(__file__).parent.parent / "outputs"

wb = openpyxl.load_workbook(INPUT_FILE)


def _read_sheet(ws, min_row: int = 5):
    """Read parameter name->value and name->unit from a spreadsheet sheet."""
    specs: dict[str, object] = {}
    units: dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_row, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
                continue
            try:
                specs[name] = float(value)
            except (ValueError, TypeError):
                specs[name] = value
            units[name] = str(row[2].value) if row[2].value else "—"
    return specs, units


opt_specs, opt_units = _read_sheet(wb["Optical System"])
det_specs, det_units = _read_sheet(wb["Detector & Readout"])
sweep_specs, sweep_units = _read_sheet(wb["Jitter Sweep"])

print("=" * 80)
print("SCENARIO 5.4: Jitter Tolerance — Line-of-Sight Stability Requirements")
print("=" * 80)

print("\n=== Optical System ===")
for k, v in opt_specs.items():
    print(f"  {k}: {v} {opt_units.get(k, '')}")

print("\n=== Detector & Readout ===")
for k, v in det_specs.items():
    print(f"  {k}: {v} {det_units.get(k, '')}")

print("\n=== Jitter Sweep ===")
for k, v in sweep_specs.items():
    print(f"  {k}: {v} {sweep_units.get(k, '')}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

aperture_m = float(opt_specs["Aperture diameter"]) / 100.0       # cm -> m
focal_length_m = float(opt_specs["Focal length"]) / 100.0        # cm -> m
f_number = float(opt_specs["f-number"])
transmission = float(opt_specs["Optical transmission"]) / 100.0  # % -> frac
optics_temp_K = float(opt_specs["Optics temperature"]) + 273.15  # C -> K
wfe_waves = float(opt_specs["WFE RMS"])
obscuration = float(opt_specs["Central obscuration"]) / 100.0    # % -> frac

band_min_um = float(opt_specs["Filter cut-on"]) / 1000.0         # nm -> um
band_max_um = float(opt_specs["Filter cut-off"]) / 1000.0        # nm -> um
band_center_um = float(opt_specs["Band center"]) / 1000.0        # nm -> um

altitude_m = float(opt_specs["Orbit altitude"]) * 1000.0         # km -> m
offnadir_rad = float(opt_specs["Off-nadir angle"]) * math.pi / 180.0

target_temp_K = float(opt_specs["Target temperature"])
bg_temp_K = float(opt_specs["Background temperature"])
target_refl = float(opt_specs["Target reflectance"])
bg_refl = float(opt_specs["Background reflectance"])
solar_zenith_rad = float(opt_specs["Solar zenith angle"]) * math.pi / 180.0

pixel_pitch_um = float(det_specs["Pixel pitch"])
pixel_pitch_m = pixel_pitch_um * 1e-6
qe = float(det_specs["Quantum efficiency"]) / 100.0              # % -> frac
dark_rate = float(det_specs["Dark current"])
operating_temp_K = float(det_specs["Operating temperature"])
ipc_coupling = float(det_specs["IPC coupling"]) / 100.0          # % -> frac

read_noise = float(det_specs["Read noise"])
fwc = float(det_specs["Full well capacity"])
adc_bits = int(det_specs["ADC resolution"])
gain = float(det_specs["System gain"])
t_int_s = float(det_specs["Integration time"]) / 1000.0          # ms -> s

jitter_min_urad = float(sweep_specs["Jitter RMS min"])
jitter_max_urad = float(sweep_specs["Jitter RMS max"])
n_sweep = int(sweep_specs["Number of sweep points"])
delta_niirs_1 = float(sweep_specs["NIIRS degradation threshold 1"])
delta_niirs_2 = float(sweep_specs["NIIRS degradation threshold 2"])
niirs_floor = float(sweep_specs["Minimum acceptable NIIRS"])

# Derived quantities
ifov_urad = pixel_pitch_um / (focal_length_m * 1e6) * 1e6  # um/m -> urad
ifov_rad = pixel_pitch_m / focal_length_m
gsd_m = pixel_pitch_m * altitude_m / focal_length_m
airy_diam_um = 2.44 * band_center_um * f_number
Q = band_center_um * f_number / pixel_pitch_um
f_nyquist = 1.0 / (2.0 * pixel_pitch_m)  # cycles/m

print("\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm / 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm / 100")
print(f"  {'f-number':<35s} {f_number:>14.1f}  {'--':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % / 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  C + 273.15")
print(f"  {'Band':<35s} {band_min_um:>6.3f}-{band_max_um:<6.3f}  {'um':<15s}  nm / 1000")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'um':<15s}  no conversion")
print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<15s}  % / 100")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms / 1000")
print(f"  {'Sensor altitude':<35s} {altitude_m:>14.0f}  {'m':<15s}  km x 1000")
print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<15s}  % / 100")
print(f"  {'Solar zenith angle':<35s} {solar_zenith_rad:>14.4f}  {'rad':<15s}  deg x pi/180")

print(f"\n=== Derived Parameters ===")
print(f"  IFOV:             {ifov_urad:.2f} urad ({ifov_rad*1e6:.2f} urad)")
print(f"  GSD at nadir:     {gsd_m:.2f} m")
print(f"  Airy disk:        {airy_diam_um:.1f} um")
print(f"  Q (sampling):     {Q:.2f} [--] ({'well-sampled' if Q > 1 else 'undersampled'})")
print(f"  f_Nyquist:        {f_nyquist:.0f} cy/m ({f_nyquist/1000:.1f} cy/mm)")

# ---------------------------------------------------------------------------
# Step 3: Build base RADIANT config (jitter set per sweep point)
# ---------------------------------------------------------------------------

base_config = {
    "source": {
        "target": {"temperature": target_temp_K, "emissivity": target_refl},
        "background": {"temperature": bg_temp_K, "emissivity": bg_refl},
    },
    "atmosphere": {
        "model": "simple",
        "standard_atmosphere": "midlat_summer",
    },
    "geometry": {
        "sensor_altitude_m": altitude_m,
        "path_zenith_rad": offnadir_rad,
        "solar_zenith_rad": solar_zenith_rad,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "optics_temperature_K": optics_temp_K,
        "wfe_rms_waves": wfe_waves,
        "obscuration_ratio": obscuration,
    },
    "platform": {
        "jitter_rms_urad": 0.0,  # overridden per sweep point
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate,
        "detector_temperature_K": operating_temp_K,
        "ipc_coupling": ipc_coupling,
    },
    "spectral_integration": {
        "filter_min_um": band_min_um,
        "filter_max_um": band_max_um,
        "integration_time_s": t_int_s,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
        "full_well_capacity_e": fwc,
    },
}

# ---------------------------------------------------------------------------
# Step 4: Run RADIANT at zero jitter (baseline)
# ---------------------------------------------------------------------------

print(f"\n=== Running RADIANT -- Baseline (zero jitter) ===")

r = Sensor.from_dict(base_config).evaluate()

baseline_snr = r.metrics["snr"]
baseline_niirs = r.metrics.get("niirs")
baseline_rer = r.metrics.get("rer")
baseline_mtf_nyq = r.metrics.get("mtf_at_nyquist")
baseline_gsd_x = r.metrics.get("gsd_cross_track_m")
baseline_gsd_y = r.metrics.get("gsd_along_track_m")
baseline_signal = r.stage_outputs["readout"]["signal_e_final"]
baseline_well_fill = baseline_signal / fwc * 100.0

# Get baseline MTF curves from performance stage
perf_out = r.stage_outputs.get("performance", {})
mtf_freq_x = perf_out.get("mtf_freq_x")
mtf_x = perf_out.get("mtf_x")

noise_dict = {nt.name: nt.value_e for nt in r.noise_terms}
total_noise = math.sqrt(sum(v ** 2 for v in noise_dict.values()))

print(f"\n  Signal:       {baseline_signal:>12,.0f} e- ({baseline_well_fill:.1f}% well)")
print(f"  Total noise:  {total_noise:>12.1f} e- RMS")
print(f"  SNR:          {baseline_snr:>12.1f} [--]")
print(f"  GSD:          {baseline_gsd_x:>12.2f} m" if baseline_gsd_x else "  GSD: N/A")
print(f"  MTF@Nyquist:  {baseline_mtf_nyq:>12.4f} [--]" if baseline_mtf_nyq else "")
print(f"  RER:          {baseline_rer:>12.4f} [--]" if baseline_rer else "")
print(f"  NIIRS:        {baseline_niirs:>12.2f} [--]" if baseline_niirs else "  NIIRS: N/A")

print(f"\n  {'Noise Term':<25s}  {'sigma [e- RMS]':>14s}  {'Fraction [%]':>13s}")
print(f"  {'-' * 25}  {'-' * 14}  {'-' * 13}")
total_var = total_noise ** 2
for name, sigma in sorted(noise_dict.items(), key=lambda x: x[1], reverse=True):
    if sigma < 0.01:
        continue
    frac = sigma ** 2 / total_var * 100.0 if total_var > 0 else 0.0
    print(f"  {name:<25s}  {sigma:>14.1f}  {frac:>13.1f}")

# ---------------------------------------------------------------------------
# Step 5: Jitter sweep using RADIANT PlatformStage
# ---------------------------------------------------------------------------

print(f"\n=== Jitter Physics ===")
print(f"  Jitter MTF: MTF_jitter(f) = exp(-2pi^2 sigma^2 f^2)")
print(f"    where sigma = jitter_rms [rad] x focal_length [m] on the focal plane")
print(f"    This is a Gaussian blur -- kills high spatial frequencies exponentially.")
print(f"")
print(f"  At f_Nyquist = {f_nyquist:.0f} cy/m:")
print(f"    MTF_jitter = exp(-2pi^2 x sigma_fp^2 x f_Nyq^2)")
print(f"")
print(f"  Jitter affects MTF and RER but NOT SNR.")
print(f"  Jitter is random pointing wander -- it blurs the image but doesn't")
print(f"  add or remove photons. Signal and noise are unchanged.")
print(f"  NIIRS changes only through the RER term: +3.32 x log10(RER).")
print(f"")
print(f"  IMPORTANT: this analysis assumes jitter is well-sampled within the")
print(f"  integration time (many jitter cycles per frame). If jitter is slow")
print(f"  (< 1/t_int), it produces frame-to-frame pointing error, not blur.")

jitter_sweep_urad = np.linspace(jitter_min_urad, jitter_max_urad, n_sweep)

print(f"\n=== Sweeping jitter RMS: {jitter_min_urad:.0f}-{jitter_max_urad:.0f} urad"
      f" ({n_sweep} points) via RADIANT PlatformStage ===")

results = []

for i, jitter_urad in enumerate(jitter_sweep_urad):
    # Update config with this jitter value
    sweep_config = base_config.copy()
    sweep_config = {**base_config, "platform": {"jitter_rms_urad": float(jitter_urad)}}

    r = Sensor.from_dict(sweep_config).evaluate()

    # Extract metrics directly from RADIANT
    snr = r.metrics.get("snr", 0.0)
    rer = r.metrics.get("rer", 0.0)
    mtf_nyq = r.metrics.get("mtf_at_nyquist", 0.0)
    niirs = r.metrics.get("niirs", 0.0)
    delta_niirs = (niirs or 0.0) - (baseline_niirs or 0.0)

    # Get focal-plane sigma from platform stage outputs
    plat_out = r.stage_outputs.get("platform", {})
    sigma_x_m = plat_out.get("jitter_sigma_x_m", 0.0)
    sigma_fp_um = sigma_x_m * 1e6
    sigma_fp_pix = sigma_x_m / pixel_pitch_m if pixel_pitch_m > 0 else 0.0

    # Compute isolated jitter MTF at Nyquist for reference
    mtf_jitter_nyq = math.exp(-2.0 * math.pi**2 * sigma_x_m**2 * f_nyquist**2)

    results.append({
        "jitter_urad": float(jitter_urad),
        "sigma_fp_um": sigma_fp_um,
        "sigma_fp_pix": sigma_fp_pix,
        "mtf_jitter_nyq": mtf_jitter_nyq,
        "mtf_sys_nyq": mtf_nyq or 0.0,
        "rer": rer or 0.0,
        "snr": snr,
        "niirs": niirs or 0.0,
        "delta_niirs": delta_niirs,
    })

    if i % max(1, n_sweep // 10) == 0:
        print(f"  [{i+1:>3d}/{n_sweep}] jitter = {jitter_urad:>5.1f} urad"
              f"  sigma = {sigma_fp_pix:.3f} pix"
              f"  RER = {rer:.4f}  MTF@Nyq = {mtf_nyq:.4f}"
              f"  NIIRS = {niirs:.2f}  dNIIRS = {delta_niirs:+.2f}")

print(f"  Sweep complete: {n_sweep} RADIANT evaluations.")

# ---------------------------------------------------------------------------
# Step 6: Print results table
# ---------------------------------------------------------------------------

print(f"\n  {'Jitter [urad]':>13s}  {'sigma_fp [um]':>13s}  {'sigma_fp [pix]':>14s}"
      f"  {'MTF_jit [--]':>12s}  {'MTF_sys [--]':>12s}  {'RER [--]':>8s}"
      f"  {'SNR [--]':>8s}  {'NIIRS [--]':>10s}  {'dNIIRS [--]':>11s}")
print(f"  {'-' * 13}  {'-' * 13}  {'-' * 14}"
      f"  {'-' * 12}  {'-' * 12}  {'-' * 8}"
      f"  {'-' * 8}  {'-' * 10}  {'-' * 11}")

for i, res in enumerate(results):
    if i % max(1, n_sweep // 20) == 0 or i == n_sweep - 1:
        print(f"  {res['jitter_urad']:>13.1f}  {res['sigma_fp_um']:>13.2f}"
              f"  {res['sigma_fp_pix']:>14.3f}  {res['mtf_jitter_nyq']:>12.4f}"
              f"  {res['mtf_sys_nyq']:>12.4f}  {res['rer']:>8.4f}"
              f"  {res['snr']:>8.1f}  {res['niirs']:>10.2f}  {res['delta_niirs']:>+11.2f}")

# ---------------------------------------------------------------------------
# Step 7: Verify SNR is unchanged by jitter
# ---------------------------------------------------------------------------

snr_arr = [res["snr"] for res in results]
snr_min, snr_max = min(snr_arr), max(snr_arr)
snr_spread = snr_max - snr_min

print(f"\n=== SNR Invariance Check ===")
print(f"  SNR range across sweep: {snr_min:.2f} - {snr_max:.2f} [--]")
print(f"  SNR spread:             {snr_spread:.4f} [--]")
if snr_spread < 0.1:
    print(f"  CONFIRMED: SNR is unchanged by jitter (spread < 0.1 [--]).")
    print(f"  Jitter blurs the image but doesn't affect photon counts or noise.")
else:
    print(f"  WARNING: SNR varies by {snr_spread:.2f} -- investigate.")

# ---------------------------------------------------------------------------
# Step 8: Find jitter thresholds
# ---------------------------------------------------------------------------

print(f"\n=== Jitter Budget Thresholds ===")
print(f"  Baseline NIIRS (zero jitter): {baseline_niirs:.2f} [--]")
print(f"")

thresholds = [
    ("dNIIRS = -0.5", -delta_niirs_1),
    ("dNIIRS = -1.0", -delta_niirs_2),
    (f"NIIRS = {niirs_floor:.1f} floor", niirs_floor - (baseline_niirs or 0)),
]

threshold_results = {}
for label, target_delta in thresholds:
    found_jitter = None
    for i in range(len(results) - 1):
        d1 = results[i]["delta_niirs"]
        d2 = results[i + 1]["delta_niirs"]
        if d1 > target_delta >= d2:
            # Linear interpolation
            j1 = results[i]["jitter_urad"]
            j2 = results[i + 1]["jitter_urad"]
            frac = (target_delta - d1) / (d2 - d1) if (d2 - d1) != 0 else 0
            found_jitter = j1 + frac * (j2 - j1)
            break

    threshold_results[label] = found_jitter
    if found_jitter is not None:
        sigma_fp = found_jitter * 1e-6 * focal_length_m * 1e6  # um
        sigma_pix = sigma_fp / pixel_pitch_um
        print(f"  {label}: jitter = {found_jitter:.1f} urad"
              f" (sigma_fp = {sigma_fp:.1f} um = {sigma_pix:.2f} pixels)")
    else:
        print(f"  {label}: NOT reached in sweep range"
              f" (max dNIIRS = {results[-1]['delta_niirs']:+.2f} [--])")

# ---------------------------------------------------------------------------
# Step 9: Jitter in context -- pixel fractions
# ---------------------------------------------------------------------------

print(f"\n=== Jitter in Context ===")
print(f"  Pixel pitch:     {pixel_pitch_um:.0f} um = {ifov_urad:.1f} urad IFOV")
print(f"  Airy disk:       {airy_diam_um:.1f} um = {airy_diam_um/pixel_pitch_um:.2f} pixels")
print(f"  Q parameter:     {Q:.2f} [--]")
print(f"")

context_jitters = [0.2, 0.5, 1.0, 1.6, 2.0, 3.0, 5.0]
print(f"  {'Jitter [urad]':>13s}  {'Frac of IFOV':>12s}  {'sigma_fp [pix]':>14s}"
      f"  {'MTF@Nyq [--]':>12s}  {'dNIIRS [--]':>11s}")
print(f"  {'-' * 13}  {'-' * 12}  {'-' * 14}  {'-' * 12}  {'-' * 11}")

for j_urad in context_jitters:
    frac_ifov = j_urad / ifov_urad
    sigma_pix = j_urad * 1e-6 * focal_length_m / pixel_pitch_m
    # Find closest sweep result
    idx = min(range(len(results)), key=lambda i: abs(results[i]["jitter_urad"] - j_urad))
    res = results[idx]
    print(f"  {j_urad:>13.1f}  {frac_ifov:>12.2f}  {sigma_pix:>14.3f}"
          f"  {res['mtf_sys_nyq']:>12.4f}  {res['delta_niirs']:>+11.2f}")

# ---------------------------------------------------------------------------
# Step 10: MTF curves at selected jitter levels
# ---------------------------------------------------------------------------

print(f"\n=== MTF vs. Spatial Frequency at Selected Jitter Levels ===")

jitter_levels_urad = [0, 0.5, 1.0, 1.6, 2.0, 3.0, 5.0]
mtf_curves = {}

if mtf_freq_x is not None and mtf_x is not None:
    mtf_freq = np.array(mtf_freq_x)
    mtf_baseline = np.array(mtf_x)
    print(f"  Baseline system MTF available ({len(mtf_freq)} frequency points)")
    print(f"  Applying jitter MTF analytically for curve plot: MTF_sys x exp(-2pi^2 sigma^2 f^2)")

    for j_urad in jitter_levels_urad:
        sigma_m = j_urad * 1e-6 * focal_length_m
        mtf_jitter = np.exp(-2.0 * np.pi**2 * sigma_m**2 * mtf_freq**2)
        mtf_curves[j_urad] = mtf_baseline * mtf_jitter
else:
    print(f"  System MTF curve not available; generating analytic approximation.")
    mtf_freq = np.linspace(0, 2 * f_nyquist, 200)
    mtf_pixel = np.abs(np.sinc(mtf_freq * pixel_pitch_m))
    f_cutoff = aperture_m / (band_center_um * 1e-6 * focal_length_m)
    f_norm = mtf_freq / f_cutoff
    mtf_diff = np.where(
        f_norm < 1.0,
        (2.0 / np.pi) * (np.arccos(f_norm) - f_norm * np.sqrt(1 - f_norm**2)),
        0.0,
    )
    mtf_baseline = mtf_pixel * mtf_diff

    for j_urad in jitter_levels_urad:
        sigma_m = j_urad * 1e-6 * focal_length_m
        mtf_jitter = np.exp(-2.0 * np.pi**2 * sigma_m**2 * mtf_freq**2)
        mtf_curves[j_urad] = mtf_baseline * mtf_jitter

# ---------------------------------------------------------------------------
# Step 11: Summary
# ---------------------------------------------------------------------------

print(f"\n=== Summary ===")
print(f"  Sensor: VNIR Pan, {aperture_m*100:.0f} cm, f/{f_number:.0f},"
      f" {pixel_pitch_um:.0f} um, {altitude_m/1000:.0f} km orbit")
print(f"  GSD: {gsd_m:.2f} m, Q = {Q:.2f} [--]")
print(f"  Baseline NIIRS: {baseline_niirs:.2f} [--] (zero jitter)")
print(f"  Baseline RER:   {baseline_rer:.4f} [--]")
print(f"  Baseline MTF@Nyquist: {baseline_mtf_nyq:.4f} [--]")
print(f"")

for label, j_val in threshold_results.items():
    if j_val is not None:
        print(f"  {label}: jitter <= {j_val:.1f} urad")
    else:
        print(f"  {label}: not reached in 0-{jitter_max_urad:.0f} urad range")

print(f"")
print(f"  JITTER BUDGET GUIDANCE:")
print(f"    For dNIIRS <= 0.5, jitter must be allocated across all sources:")
print(f"    sigma_total^2 = sigma_RW^2 + sigma_solar^2 + sigma_cryo^2"
      f" + sigma_struct^2 + sigma_ACS^2")
print(f"    Total RSS must stay below the threshold above.")
print(f"    Typical allocation: each source gets <= 40% of the total budget.")

# ---------------------------------------------------------------------------
# Step 12: Generate plots
# ---------------------------------------------------------------------------

PLOT_DIR.mkdir(parents=True, exist_ok=True)

# ---- Plot 1: NIIRS vs. Jitter RMS ----
fig1, ax1 = plt.subplots(figsize=(10, 6))

jitter_arr = [res["jitter_urad"] for res in results]
niirs_arr = [res["niirs"] for res in results]
delta_arr = [res["delta_niirs"] for res in results]

ax1.plot(jitter_arr, niirs_arr, "b-o", markersize=3, linewidth=2, label="NIIRS")

if baseline_niirs:
    ax1.axhline(y=baseline_niirs - delta_niirs_1, color="orange", linestyle="--",
                linewidth=1.5, label=f"dNIIRS = -{delta_niirs_1:.1f}")
    ax1.axhline(y=baseline_niirs - delta_niirs_2, color="r", linestyle="--",
                linewidth=1.5, label=f"dNIIRS = -{delta_niirs_2:.1f}")
    ax1.axhline(y=niirs_floor, color="darkred", linestyle=":",
                linewidth=1.5, label=f"Floor = {niirs_floor:.1f}")

for label, j_val in threshold_results.items():
    if j_val is not None:
        ax1.axvline(x=j_val, color="gray", linestyle=":", alpha=0.5)
        ax1.annotate(f"{j_val:.1f} urad",
                     xy=(j_val, min(niirs_arr) + 0.1),
                     fontsize=8, rotation=90, ha="right")

ax1.set_xlabel("Jitter RMS [urad]", fontsize=12)
ax1.set_ylabel("NIIRS [--]", fontsize=12)
ax1.legend(loc="upper right", fontsize=10)
ax1.grid(True, alpha=0.3)
ax1.set_title("NIIRS vs. Jitter RMS (RADIANT full-chain evaluation)", fontsize=14)
fig1.tight_layout()
fig1.savefig(PLOT_DIR / "niirs_vs_jitter.png", dpi=150)
print(f"\n  Plot saved: outputs/niirs_vs_jitter.png")

# ---- Plot 2: MTF@Nyquist and RER vs. Jitter ----
fig2, ax2 = plt.subplots(figsize=(10, 6))

mtf_arr = [res["mtf_sys_nyq"] for res in results]
rer_arr = [res["rer"] for res in results]

ax2.plot(jitter_arr, mtf_arr, "b-o", markersize=3, linewidth=2,
         label="System MTF@Nyquist")
ax2.plot(jitter_arr, rer_arr, "r-s", markersize=3, linewidth=2,
         label="RER")

mtf_jit_arr = [res["mtf_jitter_nyq"] for res in results]
ax2.plot(jitter_arr, mtf_jit_arr, "g--", linewidth=1.5, alpha=0.7,
         label="Jitter MTF@Nyquist (isolated)")

ax2.set_xlabel("Jitter RMS [urad]", fontsize=12)
ax2.set_ylabel("MTF / RER [--]", fontsize=12)
ax2.legend(loc="upper right", fontsize=10)
ax2.grid(True, alpha=0.3)
ax2.set_ylim(0, 1.05)
ax2.set_title("MTF and RER Degradation vs. Jitter (RADIANT chain)", fontsize=14)
fig2.tight_layout()
fig2.savefig(PLOT_DIR / "mtf_rer_vs_jitter.png", dpi=150)
print(f"  Plot saved: outputs/mtf_rer_vs_jitter.png")

# ---- Plot 3: MTF curves at selected jitter levels ----
fig3, ax3 = plt.subplots(figsize=(10, 6))

colors = ["#1f77b4", "#2ca02c", "#ff7f0e", "#d62728", "#9467bd", "#8c564b", "#e377c2"]
freq_norm = mtf_freq / f_nyquist

for idx, j_urad in enumerate(jitter_levels_urad):
    label = f"{j_urad} urad" if j_urad > 0 else "0 urad (baseline)"
    ax3.plot(freq_norm, mtf_curves[j_urad], color=colors[idx % len(colors)],
             linewidth=2 if j_urad == 0 else 1.5,
             linestyle="-" if j_urad == 0 else "--",
             label=label)

ax3.axvline(x=1.0, color="gray", linestyle=":", alpha=0.5, label="Nyquist")
ax3.set_xlabel("Spatial Frequency [x f_Nyquist]", fontsize=12)
ax3.set_ylabel("System MTF [--]", fontsize=12)
ax3.legend(loc="upper right", fontsize=9)
ax3.grid(True, alpha=0.3)
ax3.set_xlim(0, 2.0)
ax3.set_ylim(0, 1.05)
ax3.set_title("System MTF vs. Spatial Frequency at Different Jitter Levels", fontsize=13)
fig3.tight_layout()
fig3.savefig(PLOT_DIR / "mtf_curves_vs_jitter.png", dpi=150)
print(f"  Plot saved: outputs/mtf_curves_vs_jitter.png")

# ---- Plot 4: Jitter sigma in pixel fractions ----
fig4, ax4 = plt.subplots(figsize=(10, 6))

sigma_pix_arr = [res["sigma_fp_pix"] for res in results]

ax4a = ax4
ax4a.plot(jitter_arr, sigma_pix_arr, "b-", linewidth=2, label="sigma_jitter [pixels]")
ax4a.set_xlabel("Jitter RMS [urad]", fontsize=12)
ax4a.set_ylabel("Jitter sigma on Focal Plane [pixels]", fontsize=12, color="b")
ax4a.tick_params(axis="y", labelcolor="b")
ax4a.axhline(y=0.5, color="orange", linestyle="--", alpha=0.7,
             label="0.5 pixel (rule-of-thumb limit)")
ax4a.axhline(y=1.0, color="r", linestyle="--", alpha=0.7,
             label="1.0 pixel (severe)")
ax4a.legend(loc="upper left", fontsize=9)
ax4a.grid(True, alpha=0.3)

ax4b = ax4a.twinx()
ax4b.plot(jitter_arr, delta_arr, "r-", linewidth=2, alpha=0.6, label="dNIIRS")
ax4b.set_ylabel("dNIIRS [--]", fontsize=12, color="r")
ax4b.tick_params(axis="y", labelcolor="r")

ax4.set_title("Jitter Blur (Pixel Fractions) and NIIRS Impact", fontsize=14)
fig4.tight_layout()
fig4.savefig(PLOT_DIR / "jitter_pixels_and_niirs.png", dpi=150)
print(f"  Plot saved: outputs/jitter_pixels_and_niirs.png")

print(f"  All 4 plots saved to outputs/")
plt.show()

# ---------------------------------------------------------------------------
# Step 13: Write output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

owb = openpyxl.Workbook()
hfont = Font(bold=True, size=11)
sfill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
sfont = Font(bold=True, size=11, color="FFFFFF")
tb = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: Jitter Sweep ---
ows1 = owb.active
ows1.title = "Jitter Sweep"
ows1["A1"] = "NIIRS, MTF, and RER vs. Jitter RMS (RADIANT full-chain)"
ows1["A1"].font = Font(bold=True, size=14)

headers = ["Jitter [urad]", "sigma_fp [um]", "sigma_fp [pixels]",
           "MTF_jitter@Nyq [--]", "MTF_sys@Nyq [--]", "RER [--]",
           "SNR [--]", "NIIRS [--]", "dNIIRS [--]"]
for c, h in enumerate(headers, 1):
    cell = ows1.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb
for c in range(1, len(headers) + 1):
    ows1.column_dimensions[chr(64 + c)].width = 16

for i, res in enumerate(results, 4):
    ows1.cell(row=i, column=1, value=round(res["jitter_urad"], 2)).border = tb
    ows1.cell(row=i, column=2, value=round(res["sigma_fp_um"], 3)).border = tb
    ows1.cell(row=i, column=3, value=round(res["sigma_fp_pix"], 4)).border = tb
    ows1.cell(row=i, column=4, value=round(res["mtf_jitter_nyq"], 5)).border = tb
    ows1.cell(row=i, column=5, value=round(res["mtf_sys_nyq"], 5)).border = tb
    ows1.cell(row=i, column=6, value=round(res["rer"], 5)).border = tb
    ows1.cell(row=i, column=7, value=round(res["snr"], 2)).border = tb
    ows1.cell(row=i, column=8, value=round(res["niirs"], 3)).border = tb
    ows1.cell(row=i, column=9, value=round(res["delta_niirs"], 3)).border = tb

# --- Sheet 2: Thresholds ---
ows2 = owb.create_sheet("Thresholds")
ows2["A1"] = "Jitter Budget Thresholds"
ows2["A1"].font = Font(bold=True, size=14)
ows2.column_dimensions["A"].width = 28
ows2.column_dimensions["B"].width = 18
ows2.column_dimensions["C"].width = 14
ows2.column_dimensions["D"].width = 14

r_idx = 3
ows2.cell(row=r_idx, column=1, value="Threshold").font = hfont
ows2.cell(row=r_idx, column=2, value="Jitter [urad]").font = hfont
ows2.cell(row=r_idx, column=3, value="sigma_fp [um]").font = hfont
ows2.cell(row=r_idx, column=4, value="sigma_fp [pix]").font = hfont
r_idx += 1

for label, j_val in threshold_results.items():
    ows2.cell(row=r_idx, column=1, value=label).border = tb
    if j_val is not None:
        sigma_um = j_val * 1e-6 * focal_length_m * 1e6
        sigma_pix = sigma_um / pixel_pitch_um
        ows2.cell(row=r_idx, column=2, value=round(j_val, 2)).border = tb
        ows2.cell(row=r_idx, column=3, value=round(sigma_um, 2)).border = tb
        ows2.cell(row=r_idx, column=4, value=round(sigma_pix, 3)).border = tb
    else:
        ows2.cell(row=r_idx, column=2, value="N/A").border = tb
    r_idx += 1

# --- Sheet 3: Summary ---
ows3 = owb.create_sheet("Summary")
ows3["A1"] = "Scenario 5.4 Summary"
ows3["A1"].font = Font(bold=True, size=14)
ows3.column_dimensions["A"].width = 35
ows3.column_dimensions["B"].width = 25

summaries = [
    ("System", f"VNIR Pan, {aperture_m*100:.0f} cm, f/{f_number:.0f}"),
    ("Band", f"{band_min_um*1000:.0f}-{band_max_um*1000:.0f} nm"),
    ("Pixel pitch", f"{pixel_pitch_um:.0f} um"),
    ("GSD", f"{gsd_m:.2f} m"),
    ("Q parameter", f"{Q:.2f} [--]"),
    ("Baseline NIIRS", f"{baseline_niirs:.2f} [--]"),
    ("Baseline RER", f"{baseline_rer:.4f} [--]"),
    ("Baseline MTF@Nyquist", f"{baseline_mtf_nyq:.4f} [--]"),
    ("Baseline SNR", f"{baseline_snr:.1f} [--]"),
    ("IFOV", f"{ifov_urad:.1f} urad"),
    ("Method", "RADIANT full-chain with PlatformStage jitter convolution"),
]
row = 3
for label, val in summaries:
    ows3.cell(row=row, column=1, value=label).border = tb
    ows3.cell(row=row, column=2, value=val).border = tb
    row += 1

owb.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
