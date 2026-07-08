#!/usr/bin/env python3
"""Create Karen's inputs for Scenario 7.5: performance at temperature extremes.

Emits:

- ``karen_dark_current_tvac.csv`` — measured dark current vs FPA
  temperature (``T_K, dark_current_e_per_s``), 70–95 K. The data is
  Arrhenius at low T but turns UP super-Arrhenius above ~85 K (tunnel /
  defect-assisted leakage), the deviation the scenario exposes.
- ``karen_qe_vs_temperature.csv`` — QE at three FPA temperatures
  (``T_K, QE``): a mild fall with warming.
- ``karen_env_sensor.xlsx`` — the as-built bench sensor + spec limits.

Run:  python create_spreadsheet.py
"""

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent
K_B_EV = 8.617333262e-5  # eV/K

# ---------------------------------------------------------------------------
# Measured dark current J(T) — Arrhenius + super-Arrhenius upturn
# ---------------------------------------------------------------------------

TEMPS_K = [70.0, 75.0, 77.0, 80.0, 82.0, 85.0, 88.0, 90.0, 92.0, 95.0]
EA_EV = 0.24                      # diffusion activation energy (MWIR HgCdTe)
J0_ANCHOR_T, J0_ANCHOR = 77.0, 5.0e4  # e-/s at 77 K (Arrhenius part)


def arrhenius_e_per_s(T: float) -> float:
    j0 = J0_ANCHOR / math.exp(-EA_EV / (K_B_EV * J0_ANCHOR_T))
    return j0 * math.exp(-EA_EV / (K_B_EV * T))


# Super-Arrhenius excess above ~83 K (grows sharply): a soft-onset term.
def measured_e_per_s(T: float) -> float:
    base = arrhenius_e_per_s(T)
    excess = 0.0
    if T > 83.0:
        excess = base * 0.9 * ((T - 83.0) / 5.0) ** 2.4
    return base + excess


with open(HERE / "karen_dark_current_tvac.csv", "w", encoding="utf-8") as fh:
    fh.write("# TVAC dark-current measurement, MWIR HgCdTe FPA, per-pixel mean\n")
    fh.write("# Deviates super-Arrhenius above ~85 K (defect-assisted leakage)\n")
    fh.write("T_K,dark_current_e_per_s\n")
    for T in TEMPS_K:
        fh.write(f"{T},{measured_e_per_s(T):.1f}\n")
print("Wrote karen_dark_current_tvac.csv")

# ---------------------------------------------------------------------------
# QE(T) — three measured temperatures
# ---------------------------------------------------------------------------

QE_T = [(70.0, 0.78), (80.0, 0.75), (90.0, 0.71)]
with open(HERE / "karen_qe_vs_temperature.csv", "w", encoding="utf-8") as fh:
    fh.write("# Band-average QE vs FPA temperature (3 measured points)\n")
    fh.write("T_K,QE\n")
    for T, q in QE_T:
        fh.write(f"{T},{q}\n")
print("Wrote karen_qe_vs_temperature.csv")

# ---------------------------------------------------------------------------
# As-built sensor + spec limits
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sensor and Spec"
ws["A1"] = "Scenario 7.5: Environmental test — as-built bench sensor"
ws["A1"].font = Font(bold=True, size=14)
for col, width in zip("ABCD", [34, 14, 12, 42]):
    ws.column_dimensions[col].width = width
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Aperture diameter", 12.0, "cm", ""),
    ("Focal length", 24.0, "cm", "f/2.0"),
    ("Optical transmission", 74.0, "%", ""),
    ("Optics emissivity", 26.0, "%", "1 - transmission (Kirchhoff)"),
    ("Nearfield fraction", 0.04, "—", "Cold-stop leakage"),
    ("Optics temperature", 20.0, "°C", "Bench ambient"),
    ("Pixel pitch", 18.0, "µm", ""),
    ("Cold filter passband", "3600–4900", "nm", ""),
    ("Integration time", 0.55, "ms", ""),
    ("Read noise (CDS)", 28.0, "e⁻ RMS", ""),
    ("System gain", 140.0, "e⁻/DN", ""),
    ("ADC resolution", 14, "bits", ""),
    ("Full well capacity", 1500000, "e⁻", ""),
    ("Shroud temperature", 300.0, "K", "Chamber blackbody source"),
    ("Shroud emissivity", 0.98, "—", ""),
    ("SPEC: min SNR", 750.0, "—", "Acceptance threshold"),
    ("SPEC: max NEDT", 35.0, "mK", "Acceptance threshold"),
]
r = 4
for name, value, unit, note in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "karen_env_sensor.xlsx"
wb.save(out)
print(f"Wrote {out}")
