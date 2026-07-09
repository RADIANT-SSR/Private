#!/usr/bin/env python3
"""Create Raj's inputs for Scenario 3.5: nighttime MWIR imaging feasibility.

Emits two vendor-style files:

- ``raj_scene.xlsx`` — the thermal scene (building complex vs terrain
  background), the tropical-atmosphere column, and the dual-band airborne
  sensor config (MWIR 3.5–5 µm and the LWIR 8–12 µm comparison band).
- ``noaa_lst_strip.csv`` — a stand-in for the NOAA land-surface-temperature
  GeoTIFF Raj actually has (RADIANT has no raster reader — see gaps.md).
  A 1-D strip of surface temperatures across the scene; the run script uses
  its min/mean/max as the background-temperature envelope.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="264478")

ws = wb.active
ws.title = "Scene"
ws["A1"] = "Scenario 3.5: nighttime thermal scene"
ws["A1"].font = Font(bold=True, size=14)
for i, text in enumerate(["Surface", "Temp (K)", "Emissivity", "Note"]):
    c = ws.cell(row=3, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
for col, w in zip("ABCD", [22, 10, 12, 32], strict=True):
    ws.column_dimensions[col].width = w
rows = [
    ("Building complex", 295.0, 0.92, "target — concrete/metal, retained daytime heat"),
    ("Terrain background", 288.0, 0.95, "scrub vegetation, NOAA LST map"),
]
for r, (name, t, e, note) in enumerate(rows, start=4):
    for col, val in zip("ABCD", [name, t, e, note], strict=True):
        ws[f"{col}{r}"] = val
ws["A7"] = "Scene ΔT = 7 K (target − background), nighttime (no solar illumination)"
ws["A7"].font = Font(italic=True)

ws2 = wb.create_sheet("Atmosphere")
for i, text in enumerate(["Parameter", "Value", "Unit"]):
    c = ws2.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws2.column_dimensions["A"].width = 26
for i, (name, val, unit) in enumerate(
    [
        ("Standard atmosphere", "tropical", "—"),
        ("Precipitable water", 4.1, "cm"),
        ("Sensor altitude", 3.0, "km"),
    ],
    start=2,
):
    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=unit)

ws3 = wb.create_sheet("Sensor")
for i, text in enumerate(["Parameter", "Value", "Unit"]):
    c = ws3.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws3.column_dimensions["A"].width = 26
for i, (name, val, unit) in enumerate(
    [
        ("Primary band (MWIR)", "3.5-5.0", "µm"),
        ("Comparison band (LWIR)", "8-12", "µm"),
        ("Aperture diameter", 30.0, "cm"),
        ("Focal length", 1.2, "m (f/4)"),
        ("Pixel pitch", 30.0, "µm"),
        ("QE", 75.0, "%"),
        ("Dark current", 5.0e5, "e⁻/s"),
        ("Detector temperature", 110.0, "K"),
        ("Integration time", 0.2, "ms"),
        ("Read noise", 50.0, "e⁻"),
        ("Full well", 1.0e7, "e⁻"),
    ],
    start=2,
):
    ws3.cell(row=i, column=1, value=name)
    ws3.cell(row=i, column=2, value=val)
    ws3.cell(row=i, column=3, value=unit)

out = HERE / "raj_scene.xlsx"
wb.save(out)
print(f"Wrote {out}")

# NOAA LST GeoTIFF stand-in: a 1-D surface-temperature strip (K) across the
# scene. RADIANT cannot read the real GeoTIFF (gaps.md); this CSV is the
# transcribed background-temperature envelope.
strip = [287.6, 288.1, 288.4, 288.0, 287.8, 288.3, 288.6, 288.2, 287.9, 288.0]
csv = HERE / "noaa_lst_strip.csv"
csv.write_text(
    "pixel_index,surface_temperature_K\n" + "".join(f"{i},{t}\n" for i, t in enumerate(strip))
)
print(f"Wrote {csv}")
