"""Scenario 10.2 — Air-to-air level-arm MWIR IRST (direction-general validation).

Sarah is sizing an airborne MWIR IRST.  Own-ship cruises at 10 km; the target
cruises **co-altitude** at 10 km.  That makes the line of sight a *level arm*:
scene class ``air_to_air``, matrix cell E5 of the ADR-0011 observer x target
grid, and one of the eight classes that did not exist before Geometry-Flexibility
Phases 1-4.

What this scenario validates
----------------------------
1. **Scene class + composition.**  ``GeometryStage`` derives ``air_to_air`` from
   the altitude pair alone (ADR-0011 decision 8), publishes it, and validates the
   optional ``geometry.scene_class`` assertion against it.
2. **Level-arm geometry.**  Equal altitudes with a chord entry (mode V0) resolve
   through the central-angle solution
   ``phi = 2 asin(d / 2r)``, ``theta_o = pi/2 + phi/2`` — the solution that
   subsumed the old collocated no-triangle carve-out (guardrail G4).  Both
   endpoints look slightly *down* at each other because the chord sags below the
   constant-altitude shell.
3. **The horizon guard, both verdicts, in one sweep.**  The tangent depression
   ``dh`` grows as ``L^2 / 8r``: below 100 m the guard is silent, from 100 m to
   2 km it emits a quantified ``UserWarning`` naming the excluded refraction.
   The 25-100 km sweep crosses that boundary at ~71.4 km, so the same
   configuration is clean at the short end and warned at the long end.
4. **Target kinematics (Gap 111), both doors.**  K2 (speed + heading + climb)
   derives the relative LOS angular rate; K1 (direct rate) is fed the same number
   and the resolver's V0-V4 agreement check confirms they agree; a deliberately
   wrong K1 raises.  The rate then drives the one relative-motion smear.
5. **Metric relevance flips on the target band** (guardrail G3): an air target
   has no ground plane, so GSD / ground range / swath / NIIRS default off and the
   target-plane sample distances default on.
6. **Cross-model anchor.**  The simple model's analytic level arm is compared
   against the delivered MODTRAN 6 horizontal grid at 10 km altitude
   (runs L16-L20), band by band.

Physics notes printed by the run
--------------------------------
* the radiometric regime and why;
* which parameters do *not* affect the answer here and why;
* the direction-aware path product used for a level arm, and the sky background
  the LOS termination selects;
* the quantified refraction caveat the horizon guard raises, and how big it is
  next to the band-model error the MODTRAN anchor exposes.

Usage (from anywhere; all paths are repo-relative)::

    python scenarios/10_direction_general/10.2_air_to_air_level_irst/\
        scripts/run_air_to_air_level_irst.py

Runtime: ~7 s (39 chain evaluations at ~0.1 s each).

Module-level factory for the GUI baseline registry
--------------------------------------------------
``make_sensor() -> Sensor`` builds the nominal (50 km) validated configuration
without running ``main()``; ``make_config(range_m, ...) -> dict`` is the plain
dict behind it.
"""

from __future__ import annotations

import logging
import math
import warnings
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")  # headless — figures are written, never shown

import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402

from radiant.api import Sensor  # noqa: E402
from radiant.api.scene_relevance import default_off_metrics  # noqa: E402
from radiant.core.constants import R_EARTH_M  # noqa: E402
from radiant.core.viewing_triangle import (  # noqa: E402
    GUARD_DH_CLEAN_M,
    GUARD_DH_RAISE_M,
    classify_horizon_topology,
)

# ---------------------------------------------------------------------------
# Paths — repo-relative only (Rule 30)
# ---------------------------------------------------------------------------

SCENARIO_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[4]
INPUT_XLSX = SCENARIO_DIR / "inputs" / "irst_air_to_air_vendor_data.xlsx"
OUTPUT_DIR = SCENARIO_DIR / "outputs"
RESULTS_XLSX = OUTPUT_DIR / "10.2_air_to_air_results.xlsx"
MODTRAN_RUNS = REPO_ROOT / "modtran" / "real_runs"

#: MODTRAN 6 horizontal (ITYPE=1) grid, 10 km altitude row — the L-grid cells
#: this scenario's level arm is anchored against.  ``docs/plans/modtran_run_matrix.csv``.
L_GRID_10KM: tuple[tuple[str, float], ...] = (
    ("L16", 5.0e3),
    ("L17", 1.0e4),
    ("L18", 2.5e4),
    ("L19", 5.0e4),
    ("L20", 1.0e5),
)

#: Bands the anchor compares over [um].  MWIR is this sensor's own filter band.
ANCHOR_BANDS: tuple[tuple[str, float, float], ...] = (
    ("MWIR 3.5-5.0 um (sensor band)", 3.5, 5.0),
    ("LWIR 8-12 um (reference)", 8.0, 12.0),
)

#: Effective-Earth-radius factor used ONLY for the order-of-magnitude estimate of
#: what the unmodelled refraction is worth (ADR-0011 decision 5 excludes it).
#: The standard 4/3 value; quoted as an estimate, never as a RADIANT result.
REFRACTION_K_FACTOR: float = 4.0 / 3.0

#: Density scale height near 10 km used in the same estimate [m].
DENSITY_SCALE_HEIGHT_M: float = 6500.0

#: Matrix section 7 point-source validity bound, as ``OpticsStage`` states it:
#: sqrt(A_t)/d must not exceed this multiple of the system PSF FWHM.
POINT_SOURCE_ANGULAR_LIMIT: float = 0.1


# ---------------------------------------------------------------------------
# Step 1 — read the vendor workbook (vendor units, NOT RADIANT canonical)
# ---------------------------------------------------------------------------


def _read_sheet(path: Path, sheet: str) -> dict[str, Any]:
    """Return ``{Parameter: Value}`` for one four-column vendor sheet."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    out: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=5, max_col=4, values_only=True):
        name, value = row[0], row[1]
        if name is None or value is None:
            continue
        out[str(name)] = value
    workbook.close()
    return out


def read_vendor_inputs(path: Path = INPUT_XLSX) -> dict[str, dict[str, Any]]:
    """All four vendor sheets, keyed by sheet name."""
    return {
        "optics": _read_sheet(path, "IRST Optical Head"),
        "fpa": _read_sheet(path, "FPA and ROIC"),
        "engagement": _read_sheet(path, "Engagement"),
        "atmosphere": _read_sheet(path, "Atmosphere"),
    }


# ---------------------------------------------------------------------------
# Step 2 — vendor -> RADIANT canonical, exactly once, each with a comment
# ---------------------------------------------------------------------------

KNOT_TO_M_S = 0.514444  # 1 kt = 1 nautical mile/h = 1852 m / 3600 s


def to_canonical(vendor: dict[str, dict[str, Any]]) -> dict[str, float | str]:
    """Convert every vendor quantity to RADIANT canonical units.

    Canonical: wavelength um, angles rad, length m, time s, temperature K.
    This is the ONLY place a unit conversion happens (Rule 2).
    """
    optics, fpa = vendor["optics"], vendor["fpa"]
    eng, atm = vendor["engagement"], vendor["atmosphere"]

    return {
        # --- optical head ---
        "aperture_m": float(optics["Entrance pupil diameter"]) / 1000.0,  # mm -> m
        "focal_length_m": float(optics["Effective focal length"]) / 1000.0,  # mm -> m
        "f_number": float(optics["f-number"]),  # dimensionless
        "tau_optics": float(optics["Optical transmission"]) / 100.0,  # % -> fraction
        "optics_temp_K": float(optics["Optics temperature"]) + 273.15,  # degC -> K
        "obscuration": float(optics["Central obscuration"]) / 100.0,  # % -> fraction
        "wfe_rms_waves": float(optics["Wavefront error RMS"]),  # waves (canonical)
        "filter_min_um": float(optics["Spectral filter min"]),  # um (canonical)
        "filter_max_um": float(optics["Spectral filter max"]),  # um (canonical)
        # --- FPA / ROIC ---
        "pitch_um": float(fpa["Pixel pitch"]),  # um (schema input unit is um)
        "fill_factor": float(fpa["Fill factor"]) / 100.0,  # % -> fraction
        "qe": float(fpa["Quantum efficiency"]) / 100.0,  # % -> fraction
        "dark_e_s": float(fpa["Dark current"]),  # e-/s (canonical)
        "det_temp_K": float(fpa["Operating temperature"]),  # K (canonical)
        "read_noise_e": float(fpa["Read noise"]),  # e- RMS (canonical)
        "full_well_e": float(fpa["Full well capacity"]) * 1.0e3,  # ke- -> e-
        "gain_e_dn": float(fpa["System gain"]),  # e-/DN (canonical)
        "adc_bits": int(fpa["ADC resolution"]),  # bits
        "t_int_s": float(fpa["Frame integration time"]) / 1000.0,  # ms -> s
        # --- engagement ---
        "h_sensor_m": float(eng["Own-ship altitude"]) * 1000.0,  # km -> m
        "h_target_m": float(eng["Target altitude"]) * 1000.0,  # km -> m
        "range_start_m": float(eng["Range sweep start"]) * 1000.0,  # km -> m
        "range_stop_m": float(eng["Range sweep stop"]) * 1000.0,  # km -> m
        "range_step_m": float(eng["Range sweep step"]) * 1000.0,  # km -> m
        "range_nominal_m": float(eng["Nominal range"]) * 1000.0,  # km -> m
        "target_temp_K": float(eng["Target hot-parts temperature"]) + 273.15,  # degC -> K
        "target_area_m2": float(eng["Target hot-parts area"]),  # m^2 (canonical)
        "target_emissivity": float(eng["Target emissivity"]) / 100.0,  # % -> fraction
        "own_speed_m_s": float(eng["Own-ship true airspeed"]) * KNOT_TO_M_S,  # kt -> m/s
        "target_speed_m_s": float(eng["Target true airspeed"]) * KNOT_TO_M_S,  # kt -> m/s
        "target_heading_rad": math.radians(float(eng["Target heading"])),  # deg -> rad
        "target_climb_rad": math.radians(float(eng["Target climb angle"])),  # deg -> rad
        "snr_threshold": float(eng["Detection SNR threshold"]),  # dimensionless
        # --- atmosphere ---
        "profile": str(atm["Standard atmosphere"]),
        "pwv_cm": float(atm["Precipitable water"]),  # cm (schema input unit is cm)
        "visibility_km": float(atm["Visibility"]),  # km (schema input unit is km)
        "aerosol": str(atm["Aerosol type"]),
        "illumination": str(atm["Illumination"]),
    }


# ---------------------------------------------------------------------------
# Step 3 — the config factory (module level; no side effects on import)
# ---------------------------------------------------------------------------


def make_config(
    range_m: float | None = None,
    *,
    kinematics: str = "platform",
    los_rate_rad_s: float | None = None,
) -> dict[str, Any]:
    """Build the RADIANT config dict for one point of the level-arm sweep.

    Parameters
    ----------
    range_m:
        Slant range along the level arm [m].  ``None`` uses the workbook's
        nominal range (50 km) — the point the GUI baseline is built at.
    kinematics:
        ``"platform"`` — no target-velocity input (mode K0: the published LOS
        rate is the platform-only ``ground_speed / slant``).
        ``"target"`` — mode K2: the target speed/heading/climb triple.
    los_rate_rad_s:
        When given, also sets ``geometry.los_angular_rate_rad_s`` (mode K1).
        Combined with ``kinematics="target"`` this exercises the V0-V4-style
        agreement check between the two Gap 111 doors.
    """
    c = to_canonical(read_vendor_inputs())
    slant_m = float(c["range_nominal_m"]) if range_m is None else float(range_m)

    geometry: dict[str, Any] = {
        # Equal altitudes -> level arm -> scene class air_to_air (derived).
        "sensor_altitude_m": c["h_sensor_m"],
        "target_altitude_m": c["h_target_m"],
        # Viewing mode V0: the chord fixes the central angle for a level path.
        "target_range_m": slant_m,
        "solar_illumination": c["illumination"],
        "ground_speed_m_s": c["own_speed_m_s"],
        # Hot-parts area, so the schematic and the regime guards see the extent.
        "target": {"projected_area_m2": c["target_area_m2"]},
    }
    if kinematics == "target":
        geometry["target_speed_m_s"] = c["target_speed_m_s"]
        geometry["target_heading_rad"] = c["target_heading_rad"]
        geometry["target_climb_rad"] = c["target_climb_rad"]
    if los_rate_rad_s is not None:
        geometry["los_angular_rate_rad_s"] = los_rate_rad_s

    return {
        "geometry": geometry,
        "source": {
            # IRST targets are specified as in-band radiant intensity, so the
            # T7 point-intensity door (temperature + emitting area + emissivity)
            # is the native descriptor here.
            "scene_type": "point_source",
            "target": {
                "point_intensity_temperature_K": c["target_temp_K"],
                "point_intensity_area_m2": c["target_area_m2"],
                "point_intensity_emissivity": c["target_emissivity"],
            },
        },
        "atmosphere": {
            "model": "simple",
            "standard_atmosphere": c["profile"],
            "precipitable_water_cm": c["pwv_cm"],
            "visibility_km": c["visibility_km"],
            "aerosol_type": c["aerosol"],
        },
        "optics": {
            "aperture_diameter_m": c["aperture_m"],
            "focal_length_m": c["focal_length_m"],
            "transmission_scalar": c["tau_optics"],
            "optics_temperature_K": c["optics_temp_K"],
            "obscuration_ratio": c["obscuration"],
            "wfe_rms_waves": c["wfe_rms_waves"],
        },
        "detector": {
            "pixel_pitch_x_um": c["pitch_um"],
            "pixel_pitch_y_um": c["pitch_um"],
            "qe_value": c["qe"],
            "dark_rate_e_per_s": c["dark_e_s"],
            "detector_temperature_K": c["det_temp_K"],
            "fill_factor": c["fill_factor"],
        },
        "spectral_integration": {
            "filter_min_um": c["filter_min_um"],
            "filter_max_um": c["filter_max_um"],
            "integration_time_s": c["t_int_s"],
        },
        "readout": {
            "read_noise_e_rms": c["read_noise_e"],
            "full_well_capacity_e": c["full_well_e"],
            "gain_e_per_dn": c["gain_e_dn"],
            "adc_bits": c["adc_bits"],
        },
        "performance": {"detection_snr_threshold": c["snr_threshold"]},
    }


def make_sensor() -> Sensor:
    """The nominal (50 km level arm) validated Sensor — the GUI baseline.

    Module-level factory: importing this module and calling ``make_sensor()``
    never runs ``main()``.
    """
    return Sensor.from_dict(make_config())


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def _rule(char: str = "=", width: int = 92) -> str:
    return char * width


def _band_mean(lam: np.ndarray, values: np.ndarray, lo_um: float, hi_um: float) -> float:
    """Band mean of *values* over ``[lo, hi]`` um — same units as *values*."""
    band = (lam >= lo_um) & (lam <= hi_um)
    return float(np.trapezoid(values[band], lam[band]) / (hi_um - lo_um))


def _evaluate(config: dict[str, Any]) -> tuple[Any, list[warnings.WarningMessage]]:
    """Evaluate a config with warnings captured (never suppressed)."""
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result = Sensor.from_dict(config).evaluate()
    return result, list(caught)


# ---------------------------------------------------------------------------
# Report sections
# ---------------------------------------------------------------------------


def print_inputs(vendor: dict[str, dict[str, Any]], c: dict[str, float | str]) -> None:
    print(_rule())
    print("SCENARIO 10.2 — AIR-TO-AIR LEVEL-ARM MWIR IRST (scene class air_to_air, cell E5)")
    print(_rule())
    print(f"\n  Vendor workbook: {INPUT_XLSX.relative_to(REPO_ROOT)}")

    for sheet, rows in vendor.items():
        print(f"\n  --- {sheet} (vendor units) ---")
        for name, value in rows.items():
            print(f"    {name:<34s} {value}")

    print("\n  --- vendor -> RADIANT canonical (each conversion done exactly once) ---")
    print(f"    {'Quantity':<30s} {'Value':>14s}  {'Unit':<10s} Conversion")
    print(f"    {'-' * 30} {'-' * 14}  {'-' * 10} {'-' * 22}")
    table = [
        ("Entrance pupil diameter", c["aperture_m"], "m", "mm / 1000", "{:.4f}"),
        ("Effective focal length", c["focal_length_m"], "m", "mm / 1000", "{:.4f}"),
        ("Optical transmission", c["tau_optics"], "fraction", "% / 100", "{:.4f}"),
        ("Optics temperature", c["optics_temp_K"], "K", "degC + 273.15", "{:.2f}"),
        ("Pixel pitch", c["pitch_um"], "um", "(canonical input)", "{:.1f}"),
        ("Integration time", c["t_int_s"], "s", "ms / 1000", "{:.6f}"),
        ("Full well capacity", c["full_well_e"], "e-", "ke- x 1000", "{:.3e}"),
        ("Own-ship altitude", c["h_sensor_m"], "m", "km x 1000", "{:.0f}"),
        ("Target altitude", c["h_target_m"], "m", "km x 1000", "{:.0f}"),
        ("Nominal slant range", c["range_nominal_m"], "m", "km x 1000", "{:.0f}"),
        ("Target hot-parts temp", c["target_temp_K"], "K", "degC + 273.15", "{:.2f}"),
        ("Own-ship TAS", c["own_speed_m_s"], "m/s", "kt x 0.514444", "{:.2f}"),
        ("Target TAS", c["target_speed_m_s"], "m/s", "kt x 0.514444", "{:.2f}"),
        ("Target heading", c["target_heading_rad"], "rad", "deg x pi/180", "{:.6f}"),
        ("Target climb", c["target_climb_rad"], "rad", "deg x pi/180", "{:.6f}"),
    ]
    for name, value, unit, conv, fmt in table:
        print(f"    {name:<30s} {fmt.format(value):>14s}  {unit:<10s} {conv}")

    ifov_urad = float(c["pitch_um"]) * 1e-6 / float(c["focal_length_m"]) * 1e6
    print("\n  --- derived instrument scales ---")
    print(f"    IFOV (one pixel):            {ifov_urad:.2f} urad")
    print(f"    f-number:                    {float(c['f_number']):.2f} (dimensionless)")
    band_centre_um = 0.5 * (float(c["filter_min_um"]) + float(c["filter_max_um"]))
    q_sampling = band_centre_um * float(c["f_number"]) / float(c["pitch_um"])
    print(f"    Band centre:                 {band_centre_um:.3f} um")
    print(f"    Q = lambda F / p:            {q_sampling:.3f} (dimensionless, undersampled < 1)")


def print_geometry(result: Any, c: dict[str, float | str]) -> None:
    geo = result.stage_outputs["geometry"]
    theta_o = float(geo["theta_o_rad"])
    h = float(c["h_sensor_m"])

    print(f"\n{_rule()}")
    print("  1. SCENE CLASS, LEVEL-ARM GEOMETRY, AND THE Delta-h SAG")
    print(_rule())
    print(f"\n    scene_class (derived):       {geo['scene_class']}")
    print(f"    observer_class / target_class: {geo['observer_class']} / {geo['target_class']}")
    print(f"    los_direction (derived):     {geo['los_direction']}")
    print(f"    viewing_mode:                {geo['viewing_mode']}")
    print(f"    theta_o (target-side zenith): {theta_o:.9f} rad = {math.degrees(theta_o):.5f} deg")
    print(f"    eta (sensor-side off-nadir):  {float(geo['eta_rad']):.9f} rad "
          f"= {math.degrees(float(geo['eta_rad'])):.5f} deg")
    print(f"    slant range:                 {float(geo['slant_range_m']) / 1000.0:.3f} km")
    print(f"    ground range (surface arc):  {float(geo['ground_range_m']) / 1000.0:.3f} km")
    print(f"    h_sensor / h_target:         {float(geo['h_sensor_m']) / 1000.0:.1f} / "
          f"{float(geo['h_target_m']) / 1000.0:.1f} km")

    print("\n    Physics — why theta_o is slightly MORE than 90 deg on a level arm:")
    print("      Both endpoints sit on the same shell of radius r = R_E + h, so the")
    print("      straight chord between them sags below that shell. Each endpoint")
    print("      therefore looks slightly DOWN at the other:")
    print("        phi     = 2 asin(d / 2r)         [central angle, rad]")
    print("        theta_o = pi/2 + phi/2           [both endpoints, isoceles triangle]")

    guard = classify_horizon_topology(theta_o, h, h)
    dh_m = float(guard.dh_m) if guard.dh_m is not None else float("nan")
    print("\n    Horizon topology (core.viewing_triangle.classify_horizon_topology):")
    print(f"      topology:                  {guard.topology}")
    print(f"      verdict:                   {guard.action}")
    print(f"      zeta_low (lower-endpoint zenith): {guard.zenith_low_rad:.9f} rad "
          f"= {math.degrees(guard.zenith_low_rad):.5f} deg")
    print(f"      tangent depression Delta-h: {dh_m:.2f} m  "
          f"(tangent altitude {h - dh_m:.1f} m MSL)")
    print(f"      guard thresholds:          clean < {GUARD_DH_CLEAN_M:.0f} m, "
          f"warn < {GUARD_DH_RAISE_M:.0f} m, raise beyond")
    print("\n      This is the SAME number the GUI schematic prints in its Delta-h leader")
    print("      pill for a level scene (radiant.gui.viewer.schematic_view calls this exact")
    print(f"      classifier), i.e. the pill reads 'Delta-h  {dh_m:.0f} m' at this range.")


def print_regime(result: Any, c: dict[str, float | str]) -> None:
    optics = result.stage_outputs["optics"]
    source = result.stage_outputs["source"]
    geo = result.stage_outputs["geometry"]
    epsf = optics["effective_psf"]
    focal = float(c["focal_length_m"])
    psf_fwhm_rad = float(epsf.fwhm(axis="x")) / focal
    slant_m = float(geo["slant_range_m"])
    extent_rad = math.sqrt(float(c["target_area_m2"])) / slant_m
    ifov_rad = float(c["pitch_um"]) * 1e-6 / focal

    print(f"\n{_rule()}")
    print("  2. RADIOMETRIC REGIME, BACKGROUND, AND THE METRIC-RELEVANCE FLIP")
    print(_rule())
    print(f"\n    regime (final, OpticsStage):  {optics['regime']}")
    print(f"    regime (tentative, Source):   {source['regime_tentative']}")
    print(f"    scene_type declared:          {source['scene_type_declared']}")
    print(f"    target descriptor:            {type(source['target']).__name__}")
    print(f"    background descriptor:        {type(source['background']).__name__}")

    print("\n    Why POINT_SOURCE: an IRST target is specified as in-band radiant")
    print("    intensity I [W/sr], not radiance, so the T7 point-intensity door is the")
    print("    native descriptor. RADIANT then applies EE_box once, in")
    print("    SpectralIntegrationStage, to the target term only (Rule 9).")
    print("\n    Why SkyBackground: the LOS termination classifier follows the ray past")
    print("    the target; a level arm at 10 km leaves the atmosphere rather than")
    print("    striking the ground, so the background behind the target is the")
    print("    sky-continuation radiance, NOT a ground background. This is the")
    print("    direction-general behaviour — the same config with the target on the")
    print("    ground would have selected GroundBackground.")

    print("\n    Angular-size bookkeeping (matrix section 7 point-source validity):")
    print(f"      sqrt(A_t):                  {math.sqrt(float(c['target_area_m2'])):.3f} m")
    print(f"      sqrt(A_t)/d:                {extent_rad * 1e6:.3f} urad")
    print(f"      system PSF FWHM:            {psf_fwhm_rad * 1e6:.3f} urad")
    print(f"      one pixel IFOV:             {ifov_rad * 1e6:.3f} urad")
    print(f"      ratio sqrt(A_t)/d / FWHM:   {extent_rad / psf_fwhm_rad:.3f} "
          f"(matrix bound {POINT_SOURCE_ANGULAR_LIMIT:g})")
    print(f"      target fills:               {extent_rad / ifov_rad:.3f} pixel")
    compliant_range_m = math.sqrt(float(c["target_area_m2"])) / (
        POINT_SOURCE_ANGULAR_LIMIT * psf_fwhm_rad
    )
    print(f"      bound would be met beyond:  {compliant_range_m / 1000.0:.1f} km")
    near_ratio = math.sqrt(float(c["target_area_m2"])) / 25_000.0 / psf_fwhm_rad
    far_ratio = math.sqrt(float(c["target_area_m2"])) / 100_000.0 / psf_fwhm_rad
    print("      NOTE: the target is comfortably sub-pixel but NOT small compared with")
    print("      the PSF, so it sits outside the matrix bound everywhere in this sweep")
    print(f"      ({near_ratio:.2f}x FWHM at 25 km down to {far_ratio:.2f}x at 100 km, "
          f"bound {POINT_SOURCE_ANGULAR_LIMIT:g}x). The T7")
    print("      intensity door does not enforce that bound (see gaps.md) — the T1")
    print("      radiance door does, and would have refused this configuration. The")
    print("      consequence is bounded and one-sided: pre-integrating a target of")
    print("      finite extent into a delta function over-concentrates its energy, so")
    print("      the reported EE_box (and hence SNR) is mildly optimistic — by roughly")
    print("      the quadrature broadening 1/sqrt(1 + (extent/FWHM)^2) = "
          f"{1.0 / math.hypot(1.0, extent_rad / psf_fwhm_rad):.3f} at this range.")

    print("\n    Metric relevance for an AIR target (guardrail G3 — one declarative map):")
    off = sorted(default_off_metrics("air_to_air"))
    for name in off:
        present = "present" if name in result.metrics else "absent"
        print(f"      off by default: {name:<32s} -> {present} in result.metrics")
    on_keys = sorted(k for k in result.metrics if k.startswith("target_plane_sample_distance"))
    for name in on_keys:
        print(f"      on for air target: {name:<29s} = {result.metrics[name]:.4f} m")
    print("\n      GSD, ground range, swath, access rate and NIIRS all project the")
    print("      sample footprint onto a ground plane the target does not have; the")
    print("      target-plane sample distance p*d/f replaces them.")

    print("\n    Parameters that do NOT affect this result, and why:")
    print("      - geometry.solar_zenith_rad / solar_azimuth_rad: illumination is")
    print("        'night', so no reflected-solar term exists in the source or the")
    print("        sky background.")
    print("      - source.background.temperature / emissivity: the background is")
    print("        SkyBackground (the LOS leaves the atmosphere), whose radiance comes")
    print("        from the atmosphere stage, not from a user-set surface.")
    print("      - optics.obscuration_ratio = 0 (refractive head): the pupil is a")
    print("        filled circle, so no obscuration term enters the autocorrelation.")
    print("      - geometry.circular_orbit: an aircraft is not in orbit; the platform")
    print("        speed comes from the direct ground-speed door.")


def sweep(c: dict[str, float | str]) -> list[dict[str, Any]]:
    """The 25-100 km level-arm range sweep."""
    ranges_m = np.arange(
        float(c["range_start_m"]),
        float(c["range_stop_m"]) + 0.5 * float(c["range_step_m"]),
        float(c["range_step_m"]),
    )

    print(f"\n{_rule()}")
    print("  3. LEVEL-ARM RANGE SWEEP — SNR, DETECTION RANGE, HORIZON GUARD")
    print(_rule())
    print(f"\n    {'Range':>8s} {'theta_o':>10s} {'Delta-h':>9s} {'guard':>6s} "
          f"{'tau MWIR':>9s} {'alpha_eff':>10s} {'signal':>11s} {'noise':>9s} "
          f"{'SNR':>9s} {'det range':>10s} {'well':>7s}")
    print(f"    {'[km]':>8s} {'[deg]':>10s} {'[m]':>9s} {'[--]':>6s} "
          f"{'[--]':>9s} {'[1/km]':>10s} {'[e-]':>11s} {'[e- rms]':>9s} "
          f"{'[--]':>9s} {'[km]':>10s} {'[dB]':>7s}")
    print(f"    {'-' * 8} {'-' * 10} {'-' * 9} {'-' * 6} {'-' * 9} {'-' * 10} "
          f"{'-' * 11} {'-' * 9} {'-' * 9} {'-' * 10} {'-' * 7}")

    rows: list[dict[str, Any]] = []
    for range_m in ranges_m:
        result, caught = _evaluate(make_config(float(range_m)))
        geo = result.stage_outputs["geometry"]
        theta_o = float(geo["theta_o_rad"])
        guard = classify_horizon_topology(theta_o, float(c["h_sensor_m"]), float(c["h_target_m"]))
        wavelength_um = np.asarray(result.wavelength_um, dtype=float)
        tau = np.asarray(result.stage_outputs["atmosphere"]["tau_atm"], dtype=float)
        tau_mwir = _band_mean(
            wavelength_um, tau, float(c["filter_min_um"]), float(c["filter_max_um"])
        )
        # Effective band extinction coefficient implied by the band-mean tau.
        alpha_eff_per_km = -math.log(tau_mwir) / (float(range_m) / 1000.0)
        signal_e = float(result.stage_outputs["readout"]["signal_e_final"])
        noise_e = math.sqrt(sum(term.value_e**2 for term in result.noise_terms))
        horizon_warnings = [
            str(w.message) for w in caught if "horizon guard" in str(w.message)
        ]
        other_warnings = [
            str(w.message) for w in caught if "horizon guard" not in str(w.message)
        ]
        target_free_noise_e = math.sqrt(
            sum(t.value_e**2 for t in result.noise_terms if t.name != "signal_shot")
        )
        # Second evaluation at the same range with the K2 target-velocity door,
        # so the kinematics figure plots MEASURED rates, not a 1/R extrapolation.
        result_k2, _ = _evaluate(make_config(float(range_m), kinematics="target"))
        rows.append(
            {
                "target_free_noise_e": target_free_noise_e,
                "los_rate_k0_rad_s": float(geo["los_angular_rate_rad_s"]),
                "los_rate_k2_rad_s": float(
                    result_k2.stage_outputs["geometry"]["los_angular_rate_rad_s"]
                ),
                "smear_k2_m": float(result_k2.stage_outputs["platform"]["smear_width_m"]),
                "range_m": float(range_m),
                "theta_o_rad": theta_o,
                "dh_m": float(guard.dh_m) if guard.dh_m is not None else float("nan"),
                "guard_action": guard.action,
                "tau_mwir": tau_mwir,
                "alpha_eff_per_km": alpha_eff_per_km,
                "signal_e": signal_e,
                "noise_e": noise_e,
                "snr": float(result.metrics["snr"]),
                "detection_range_m": float(result.metrics["detection_range_m"]),
                "well_margin_dB": float(result.metrics["well_margin_dB"]),
                "nedt_mK": float(result.metrics["nedt_K"]) * 1000.0,
                "tpsd_m": float(result.metrics["target_plane_sample_distance_geometric_mean_m"]),
                "horizon_warnings": horizon_warnings,
                "other_warnings": other_warnings,
            }
        )
        print(f"    {float(range_m) / 1000.0:>8.1f} {math.degrees(theta_o):>10.5f} "
              f"{rows[-1]['dh_m']:>9.1f} {guard.action:>6s} {tau_mwir:>9.4f} "
              f"{alpha_eff_per_km:>10.5f} {signal_e:>11.4e} {noise_e:>9.1f} "
              f"{rows[-1]['snr']:>9.1f} {rows[-1]['detection_range_m'] / 1000.0:>10.1f} "
              f"{rows[-1]['well_margin_dB']:>7.1f}")

    _print_detection_range_discussion(rows, c)

    unexpected = sorted({m for row in rows for m in row["other_warnings"]})
    if unexpected:
        print("\n    UNEXPECTED warnings during the sweep:")
        for message in unexpected:
            print(f"      - {message}")
    else:
        print("\n    No warnings other than the horizon guard were raised in the sweep.")
    return rows


def _print_detection_range_discussion(
    rows: list[dict[str, Any]], c: dict[str, float | str]
) -> None:
    """Why detection_range_m moves with the range it is evaluated at."""
    near, far = rows[0], rows[-1]
    threshold = float(c["snr_threshold"])
    print("\n    Non-obvious result — detection_range_m depends on WHERE it is evaluated:")
    print(f"      referenced at {near['range_m'] / 1000.0:>5.0f} km -> "
          f"{near['detection_range_m'] / 1000.0:6.1f} km")
    print(f"      referenced at {far['range_m'] / 1000.0:>5.0f} km -> "
          f"{far['detection_range_m'] / 1000.0:6.1f} km  "
          f"({far['detection_range_m'] / near['detection_range_m']:.2f}x)")
    print("      The path-aware solver scales the SIGNAL along the path")
    print("      (S(R) = S_ref (R_ref/R)^2 tau(R)/tau(R_ref)) while holding the TOTAL")
    print("      noise at its reference value. That is exact in a background-limited")
    print("      system. This one is not background limited at short range:")
    print(f"      {'range [km]':>11s} {'total noise':>12s} {'signal shot':>12s} "
          f"{'target-free':>12s}")
    print(f"      {'':>11s} {'[e- rms]':>12s} {'[e- rms]':>12s} {'[e- rms]':>12s}")
    for entry in (near, far):
        shot = math.sqrt(max(entry["noise_e"] ** 2 - entry["target_free_noise_e"] ** 2, 0.0))
        print(f"      {entry['range_m'] / 1000.0:>11.0f} {entry['noise_e']:>12.1f} "
              f"{shot:>12.1f} {entry['target_free_noise_e']:>12.1f}")
    print("      At 25 km the noise is almost entirely the TARGET'S OWN shot noise,")
    print("      which vanishes as the target recedes — so freezing it makes the")
    print("      25 km answer strongly pessimistic. The far-field answer is the")
    print("      trustworthy one, and even it is slightly pessimistic.")

    # Target-free (background + read + quantisation + dark) noise floor solve.
    noise_floor_e = far["target_free_noise_e"]
    alpha_per_m = far["alpha_eff_per_km"] / 1000.0
    signal_ref_e = far["signal_e"]
    ref_m = far["range_m"]

    def _snr(range_m: float) -> float:
        signal = signal_ref_e * (ref_m / range_m) ** 2 * math.exp(-alpha_per_m * (range_m - ref_m))
        return signal / noise_floor_e

    lo_m, hi_m = ref_m, 1.0e6
    for _ in range(200):
        mid_m = 0.5 * (lo_m + hi_m)
        if _snr(mid_m) > threshold:
            lo_m = mid_m
        else:
            hi_m = mid_m
    print(f"\n      Re-solved against the TARGET-FREE noise floor of "
          f"{noise_floor_e:.1f} e- rms")
    print("      (sky background shot + read + quantisation + dark, i.e. every noise")
    print("      term that does not vanish with the target):")
    print(f"        detection range = {0.5 * (lo_m + hi_m) / 1000.0:.1f} km "
          f"at SNR = {threshold:.0f}")
    print("      That is the number an IRST engineer would quote for this design")
    print("      against this target, on the simple model. The MODTRAN anchor in")
    print("      section 6 says the real MWIR arm is more transparent still, so even")
    print("      this is a floor, not a ceiling.")


def print_horizon_guard(rows: list[dict[str, Any]], c: dict[str, float | str]) -> None:
    print(f"\n{_rule()}")
    print("  4. THE HORIZON GUARD ACROSS THE SWEEP — THE QUANTIFIED REFRACTION CAVEAT")
    print(_rule())

    clean = [r for r in rows if r["guard_action"] == "clean"]
    warn = [r for r in rows if r["guard_action"] == "warn"]
    print(f"\n    clean arms: {len(clean)} of {len(rows)}  "
          f"({clean[0]['range_m'] / 1000.0:.0f}-{clean[-1]['range_m'] / 1000.0:.0f} km, "
          f"Delta-h {clean[0]['dh_m']:.1f}-{clean[-1]['dh_m']:.1f} m)")
    if warn:
        print(f"    warned arms: {len(warn)} of {len(rows)}  "
              f"({warn[0]['range_m'] / 1000.0:.0f}-{warn[-1]['range_m'] / 1000.0:.0f} km, "
              f"Delta-h {warn[0]['dh_m']:.1f}-{warn[-1]['dh_m']:.1f} m)")

    # Analytic crossover: Delta-h = L^2 / 8r  ->  L = sqrt(8 r dh_clean).
    r_shell_m = R_EARTH_M + float(c["h_sensor_m"])
    crossover_m = math.sqrt(8.0 * r_shell_m * GUARD_DH_CLEAN_M)
    print("\n    Analytic crossover into the warning shoulder: L = sqrt(8 r Delta-h_clean)")
    print(f"      r = R_E + h = {r_shell_m / 1000.0:.1f} km, "
          f"Delta-h_clean = {GUARD_DH_CLEAN_M:.0f} m")
    print(f"      L_crossover = {crossover_m / 1000.0:.2f} km  "
          f"(the sweep crosses between {clean[-1]['range_m'] / 1000.0:.0f} and "
          f"{warn[0]['range_m'] / 1000.0:.0f} km — consistent)")

    if warn:
        print("\n    The verbatim UserWarning RADIANT raised on the longest arm:")
        for message in warn[-1]["horizon_warnings"][:1]:
            for line in _wrap(message, 84):
                print(f"      | {line}")

        print("\n    What that warning caveats, quantified:")
        far = warn[-1]
        dh_geo = far["dh_m"]
        dh_ref = dh_geo / REFRACTION_K_FACTOR
        # Mean sag over a parabolic arc is 2/3 of the maximum sag.
        mean_shift_m = (dh_geo - dh_ref) * 2.0 / 3.0
        optical_depth = -math.log(far["tau_mwir"])
        d_tau_frac = optical_depth * mean_shift_m / DENSITY_SCALE_HEIGHT_M
        print(f"      At L = {far['range_m'] / 1000.0:.0f} km the ray's tangent point sits")
        print(f"      Delta-h = {dh_geo:.1f} m below the endpoints on a straight-ray")
        print("      spherical Earth. Refraction bends the ray downward, which is")
        print("      conventionally absorbed into an effective Earth radius k*R_E;")
        print(f"      with the standard k = {REFRACTION_K_FACTOR:.4f} the sag becomes")
        print(f"      Delta-h/k = {dh_ref:.1f} m, i.e. the modelled ray samples air an")
        print(f"      average of {mean_shift_m:.1f} m LOWER than the real one (2/3 of the")
        print("      difference, the mean of a parabolic sag).")
        print("      With a density scale height of "
              f"{DENSITY_SCALE_HEIGHT_M / 1000.0:.1f} km and a")
        print(f"      band optical depth of {optical_depth:.3f}, that altitude error is worth")
        print(f"        d(tau)/tau ~ tau_od * dz / H = {d_tau_frac * 100.0:.2f} % "
              "in band transmittance.")
        print("      So the guard is flagging a sub-percent effect. The MODTRAN anchor")
        print("      in section 6 shows the simple model's own band-model error on the")
        print("      same arm is two orders of magnitude larger. The guard is correct to")
        print("      warn (Rule 17: no silent physics), but it is not the dominant error.")
        print("      k = 4/3 is a standard-atmosphere convention used here only to size")
        print("      the excluded effect; RADIANT itself models no refraction at all.")


def _wrap(text: str, width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        if current and len(current) + 1 + len(word) > width:
            lines.append(current)
            current = word
        else:
            current = f"{current} {word}".strip()
    if current:
        lines.append(current)
    return lines


def print_kinematics(c: dict[str, float | str]) -> dict[str, Any]:
    """Gap 111 — both doors, the agreement check, and the smear consequence."""
    print(f"\n{_rule()}")
    print("  5. TARGET KINEMATICS (Gap 111) — K0 / K2 / K1 AND THE SMEAR CONSEQUENCE")
    print(_rule())

    nominal_m = float(c["range_nominal_m"])
    focal_m = float(c["focal_length_m"])
    pitch_m = float(c["pitch_um"]) * 1e-6
    t_int_s = float(c["t_int_s"])

    # --- K0: platform only (no target-velocity input at all) ---
    r_k0, _ = _evaluate(make_config(nominal_m, kinematics="platform"))
    g_k0 = r_k0.stage_outputs["geometry"]
    rate_k0 = float(g_k0["los_angular_rate_rad_s"])
    smear_k0 = float(r_k0.stage_outputs["platform"]["smear_width_m"])

    # --- K2: the target-velocity triple ---
    r_k2, _ = _evaluate(make_config(nominal_m, kinematics="target"))
    g_k2 = r_k2.stage_outputs["geometry"]
    rate_k2 = float(g_k2["los_angular_rate_rad_s"])
    smear_k2 = float(r_k2.stage_outputs["platform"]["smear_width_m"])

    # --- K1: the same rate entered directly ---
    r_k1, _ = _evaluate(make_config(nominal_m, los_rate_rad_s=rate_k2))
    g_k1 = r_k1.stage_outputs["geometry"]
    rate_k1 = float(g_k1["los_angular_rate_rad_s"])

    # --- K1 + K2 together: the V0-V4 agreement check must accept them ---
    r_both, _ = _evaluate(
        make_config(nominal_m, kinematics="target", los_rate_rad_s=rate_k2)
    )
    mode_both = r_both.stage_outputs["geometry"]["los_rate_mode"]

    print(f"\n    Nominal range: {nominal_m / 1000.0:.1f} km, integration {t_int_s * 1e6:.0f} us")
    print(f"\n    {'door':<10s} {'mode string':<52s} {'omega_LOS':>12s} {'smear':>10s}")
    print(f"    {'':<10s} {'':<52s} {'[mrad/s]':>12s} {'[um]':>10s}")
    print(f"    {'-' * 10} {'-' * 52} {'-' * 12} {'-' * 10}")
    print(f"    {'K0':<10s} {str(g_k0['los_rate_mode']):<52s} {rate_k0 * 1e3:>12.5f} "
          f"{smear_k0 * 1e6:>10.4f}")
    print(f"    {'K2':<10s} {str(g_k2['los_rate_mode']):<52s} {rate_k2 * 1e3:>12.5f} "
          f"{smear_k2 * 1e6:>10.4f}")
    print(f"    {'K1':<10s} {str(g_k1['los_rate_mode']):<52s} {rate_k1 * 1e3:>12.5f} "
          f"{'--':>10s}")
    print(f"    {'K1+K2':<10s} {str(mode_both)[:52]:<52s} "
          f"{float(r_both.stage_outputs['geometry']['los_angular_rate_rad_s']) * 1e3:>12.5f} "
          f"{'--':>10s}")

    print(f"\n    K1 vs K2 relative difference: "
          f"{abs(rate_k1 - rate_k2) / rate_k2:.3e} (agreement bound 1e-2)")

    # --- and the disagreement raises (Rule 15 / ADR-0006 rule 2) ---
    print("\n    Deliberate disagreement (K1 fed a wrong rate) must RAISE:")
    try:
        _evaluate(make_config(nominal_m, kinematics="target", los_rate_rad_s=2.0 * rate_k2))
    except Exception as exc:  # noqa: BLE001 — the raise IS the demonstrated behaviour
        print(f"      {type(exc).__name__}: {str(exc).splitlines()[0][:150]}")
    else:
        print("      NO ERROR RAISED — the agreement check did not fire (unexpected).")

    # --- hand check of the K2 rate ---
    v_t = float(c["target_speed_m_s"])
    psi = float(c["target_heading_rad"])
    gamma = float(c["target_climb_rad"])
    v_s = float(c["own_speed_m_s"])
    theta_o = float(g_k2["theta_o_rad"])
    v_par = v_t * math.cos(gamma) * math.cos(psi)
    v_perp = v_t * math.cos(gamma) * math.sin(psi) - v_s
    v_up = v_t * math.sin(gamma)
    cross = math.hypot(v_perp, v_up * math.sin(theta_o) - v_par * math.cos(theta_o))
    rate_hand = cross / nominal_m
    print("\n    Hand check of the K2 rate (module docstring frame, e_par/e_perp/e_up):")
    print(f"      v_par  = v_T cos(gamma) cos(psi)          = {v_par:>+10.3f} m/s")
    print(f"      v_perp = v_T cos(gamma) sin(psi) - v_S    = {v_perp:>+10.3f} m/s")
    print(f"      v_up   = v_T sin(gamma)                   = {v_up:>+10.3f} m/s")
    print(f"      |v_rel x u_hat|                           = {cross:>10.3f} m/s")
    print(f"      omega = |v_rel x u_hat| / R               = {rate_hand:.9f} rad/s")
    print(f"      RADIANT                                    = {rate_k2:.9f} rad/s")
    print(f"      relative difference                        = "
          f"{abs(rate_hand - rate_k2) / rate_k2:.2e}")

    print("\n    Physics — why ONE rate, not two smears:")
    print("      Platform motion and target motion are not two independent blurs. They")
    print("      are two contributions to a single focal-plane translation, so they")
    print("      compose in the VELOCITY domain (v_rel = v_T - v_S) and only then")
    print("      become a smear. Here the beam-aspect crosser moves against the")
    print(f"      platform's cross-track motion, so the rates ADD: {rate_k0 * 1e3:.3f} mrad/s")
    print(f"      platform-only becomes {rate_k2 * 1e3:.3f} mrad/s relative — a factor")
    print(f"      {rate_k2 / rate_k0:.2f}. An RSS of two smears would have given "
          f"{math.hypot(rate_k0, rate_k2 - rate_k0) * 1e3:.3f} mrad/s.")

    print("\n    Smear consequence at this frame time:")
    print(f"      smear (platform only, K0):   {smear_k0 * 1e6:.4f} um = "
          f"{smear_k0 / pitch_m:.4f} pixel")
    print(f"      smear (relative, K2):        {smear_k2 * 1e6:.4f} um = "
          f"{smear_k2 / pitch_m:.4f} pixel")
    t_1px_k0 = pitch_m / (rate_k0 * focal_m)
    t_1px_k2 = pitch_m / (rate_k2 * focal_m)
    print(f"      integration time for 1-pixel smear, K0: {t_1px_k0 * 1e3:.3f} ms")
    print(f"      integration time for 1-pixel smear, K2: {t_1px_k2 * 1e3:.3f} ms")
    print(f"      -> a crossing target cuts the usable integration budget by "
          f"{100.0 * (1.0 - t_1px_k2 / t_1px_k0):.1f} %.")
    print(f"      At the {t_int_s * 1e6:.0f} us search frame the smear is far sub-pixel in")
    print("      both cases, so target motion costs no MTF here; it becomes the")
    print("      binding constraint only for track-mode integration times.")

    return {
        "rate_k0_rad_s": rate_k0,
        "rate_k2_rad_s": rate_k2,
        "rate_k1_rad_s": rate_k1,
        "rate_hand_rad_s": rate_hand,
        "smear_k0_m": smear_k0,
        "smear_k2_m": smear_k2,
        "t_1px_k0_s": t_1px_k0,
        "t_1px_k2_s": t_1px_k2,
    }


def print_modtran_anchor(c: dict[str, float | str]) -> list[dict[str, Any]]:
    """Simple-model level arm vs the delivered MODTRAN L-grid, 10 km row."""
    print(f"\n{_rule()}")
    print("  6. CROSS-MODEL ANCHOR — LEVEL ARM vs MODTRAN 6 HORIZONTAL GRID (10 km row)")
    print(_rule())

    missing = [run for run, _ in L_GRID_10KM if not (MODTRAN_RUNS / f"{run}.tp7").exists()]
    if missing:
        print(f"\n    SKIPPED — the delivered MODTRAN runs are not staged in "
              f"{MODTRAN_RUNS.relative_to(REPO_ROOT)} (gitignored; see")
        print("    modtran/real_runs/README.md). Missing: " + ", ".join(missing))
        print("    The rest of the scenario is unaffected; only this section needs them.")
        return []

    # Imported here so the scenario still runs when the staged set is absent.
    from radiant.atmosphere.level_arm import evaluate_level_arm
    from radiant.atmosphere.modtran import Tape7Reader
    from radiant.atmosphere.segments import LevelArmSpec
    from radiant.atmosphere.simple import SimpleAtmosphere

    atmosphere = SimpleAtmosphere(
        standard_atmosphere=str(c["profile"]),
        precipitable_water_cm=float(c["pwv_cm"]),
        visibility_km=float(c["visibility_km"]),
        aerosol_type=str(c["aerosol"]),
    )

    print(f"\n    Model side:   SimpleAtmosphere({c['profile']}, PWV {c['pwv_cm']} cm, "
          f"vis {c['visibility_km']} km, {c['aerosol']} aerosol)")
    print("                  evaluate_level_arm(LevelArmSpec(h=10 km, L)) — Beer-Lambert")
    print("                  at the LOCAL extinction coefficient over the true chord.")
    print("    MODTRAN side: L16-L20, ITYPE=1 horizontal decks, same profile/aerosol/")
    print("                  visibility, H1 = H2 = 10 km, Card-3 RANGE = L.")
    print("    Both sides are band-mean transmittance [dimensionless] over the band.")

    rows: list[dict[str, Any]] = []
    for band_name, lo_um, hi_um in ANCHOR_BANDS:
        print(f"\n    --- {band_name} ---")
        print(f"    {'run':>5s} {'range':>8s} {'MODTRAN tau':>12s} {'model tau':>11s} "
              f"{'ratio':>8s} {'difference':>11s} {'a_MODTRAN':>10s} {'a_model':>9s}")
        print(f"    {'':>5s} {'[km]':>8s} {'[--]':>12s} {'[--]':>11s} "
              f"{'[--]':>8s} {'[%]':>11s} {'[1/km]':>10s} {'[1/km]':>9s}")
        print(f"    {'-' * 5} {'-' * 8} {'-' * 12} {'-' * 11} {'-' * 8} {'-' * 11} "
              f"{'-' * 10} {'-' * 9}")
        for run, range_m in L_GRID_10KM:
            native = Tape7Reader(MODTRAN_RUNS / f"{run}.tp7").parse()
            nu = native.wavenumber_cm1
            keep = nu > 0.0
            lam = 1.0e4 / nu[keep]
            order = np.argsort(lam)
            lam = lam[order]
            tau_modtran = native.total_transmittance[keep][order]
            quantities = evaluate_level_arm(
                atmosphere, lam, LevelArmSpec(float(c["h_sensor_m"]), range_m)
            )
            tau_model = np.asarray(quantities.tau, dtype=float)
            m = _band_mean(lam, tau_modtran, lo_um, hi_um)
            s = _band_mean(lam, tau_model, lo_um, hi_um)
            # Effective band extinction coefficient each side implies.
            alpha_modtran = -math.log(m) / (range_m / 1000.0)
            alpha_model = -math.log(s) / (range_m / 1000.0)
            rows.append(
                {
                    "band": band_name,
                    "run": run,
                    "range_m": range_m,
                    "tau_modtran": m,
                    "tau_model": s,
                    "ratio": s / m,
                    "diff_pct": 100.0 * (s - m) / m,
                    "alpha_modtran_per_km": alpha_modtran,
                    "alpha_model_per_km": alpha_model,
                }
            )
            print(f"    {run:>5s} {range_m / 1000.0:>8.1f} {m:>12.4f} {s:>11.4f} "
                  f"{s / m:>8.3f} {100.0 * (s - m) / m:>+11.1f} "
                  f"{alpha_modtran:>10.5f} {alpha_model:>9.5f}")

    print("\n    Expected residual, and why it is one-sided at long range:")
    print("      A band-averaged transmittance is not multiplicative in path length.")
    print("      Within a band the strong lines saturate first and flux leaks through")
    print("      the windows between them, so by Jensen's inequality")
    print("      <exp(-2kL)> >= <exp(-kL)>^2 with equality only if k(lambda) is FLAT")
    print("      across the band. Equivalently, the effective band extinction")
    print("      alpha = -ln(tau)/L must FALL with path length by exactly as much as")
    print("      k(lambda) varies inside the band. The last two columns measure that")
    print("      on both sides:")
    modtran_alpha = {}
    model_alpha = {}
    for band_name, _, _ in ANCHOR_BANDS:
        band_rows = [r for r in rows if r["band"] == band_name]
        modtran_alpha[band_name] = (
            band_rows[0]["alpha_modtran_per_km"] / band_rows[-1]["alpha_modtran_per_km"]
        )
        model_alpha[band_name] = (
            band_rows[0]["alpha_model_per_km"] / band_rows[-1]["alpha_model_per_km"]
        )
        print(f"        {band_name:<32s} alpha(5 km)/alpha(100 km): "
              f"MODTRAN {modtran_alpha[band_name]:.2f}x, model {model_alpha[band_name]:.2f}x")
    print("      In the MWIR the simple model's k(lambda) is essentially FLAT across")
    print("      3.5-5.0 um — the documented CU-161 region-flat spectral-shape")
    print("      limitation — so its band mean stays very nearly exponential and it")
    print("      cannot reproduce MODTRAN's saturation at all. In the LWIR the model")
    print("      does carry some spectral structure, so it recovers part of the effect")
    print("      but still far too little. Real MWIR line structure (the CO2 4.3 um")
    print("      band and dense H2O lines cutting the window) is what MODTRAN has and")
    print("      the model does not.")
    print("      Direction of the consequence for this scenario: the model is")
    print("      progressively TOO OPAQUE as the arm lengthens, so RADIANT's SNR and")
    print("      detection range at the long end of the sweep are PESSIMISTIC. Usable")
    print("      band for the level arm at 10 km on the LWIR evidence: within ~5 % to")
    print("      25 km; beyond ~50 km treat the MWIR result as a lower bound.")
    return rows


def print_rule4(c: dict[str, float | str]) -> dict[str, Any]:
    """Rule 4 — dual-path (PSF vs MTF-product) consistency for this scene class."""
    print(f"\n{_rule()}")
    print("  7. RULE 4 DUAL-PATH CONSISTENCY CHECK FOR THIS SCENE CLASS")
    print(_rule())

    log_records: list[logging.LogRecord] = []

    class _Collector(logging.Handler):
        def emit(self, record: logging.LogRecord) -> None:
            log_records.append(record)

    handler = _Collector(level=logging.WARNING)
    root = logging.getLogger()
    previous_level = root.level
    root.addHandler(handler)
    root.setLevel(logging.WARNING)
    try:
        result, caught = _evaluate(make_config(float(c["range_nominal_m"])))
    finally:
        root.removeHandler(handler)
        root.setLevel(previous_level)

    consistency = result.stage_outputs["performance"]["dual_path_consistency"]
    print(f"\n    passed_x / passed_y:          {consistency.passed_x} / {consistency.passed_y}")
    print(f"    max |FFT(PSF) - prod(MTF)| x: "
          f"{consistency.max_absolute_error_x:.6f} (dimensionless)")
    print(f"    max |FFT(PSF) - prod(MTF)| y: "
          f"{consistency.max_absolute_error_y:.6f} (dimensionless)")
    print(f"    tolerance:                    {consistency.tolerance:.6f} (dimensionless)")
    consistency_logs = [r for r in log_records if "consistency" in r.getMessage().lower()]
    print(f"    consistency WARNING log records emitted: {len(consistency_logs)}")
    py_warnings = [str(w.message) for w in caught if "horizon guard" not in str(w.message)]
    print(f"    non-horizon Python warnings at the nominal point: {len(py_warnings)}")
    for message in py_warnings:
        print(f"      - {message[:150]}")
    verdict = (
        "SILENT — the dual-path invariant holds for the air_to_air level arm."
        if consistency.passed_x and consistency.passed_y and not consistency_logs
        else "NOT SILENT — investigate before trusting the spatial metrics."
    )
    print(f"\n    Rule-4 verdict: {verdict}")
    print("    Both paths root in the same complex pupil; the level arm changes only")
    print("    the radiometry and the geometry, not the spatial degradations, so the")
    print("    residual here is the ordinary discretisation floor.")
    return {
        "passed_x": bool(consistency.passed_x),
        "passed_y": bool(consistency.passed_y),
        "max_err_x": float(consistency.max_absolute_error_x),
        "max_err_y": float(consistency.max_absolute_error_y),
        "tolerance": float(consistency.tolerance),
        "log_records": len(consistency_logs),
    }


def print_cross_checks(
    rows: list[dict[str, Any]], kin: dict[str, Any], c: dict[str, float | str]
) -> None:
    print(f"\n{_rule()}")
    print("  8. INDEPENDENT CROSS-CHECKS (hand calculations against RADIANT)")
    print(_rule())

    # --- Check 1: level-arm central angle / theta_o closed form ---
    row = next(r for r in rows if abs(r["range_m"] - float(c["range_nominal_m"])) < 1.0)
    r_shell_m = R_EARTH_M + float(c["h_sensor_m"])
    phi = 2.0 * math.asin(row["range_m"] / (2.0 * r_shell_m))
    theta_o_hand = math.pi / 2.0 + phi / 2.0
    print("\n    Check 1 — level-arm theta_o closed form (isoceles chord triangle)")
    print(f"      phi = 2 asin(d/2r)  = {phi:.9f} rad = {math.degrees(phi):.6f} deg")
    print(f"      theta_o = pi/2+phi/2 = {theta_o_hand:.9f} rad "
          f"= {math.degrees(theta_o_hand):.6f} deg")
    print(f"      RADIANT               = {row['theta_o_rad']:.9f} rad "
          f"= {math.degrees(row['theta_o_rad']):.6f} deg")
    print(f"      relative difference   = "
          f"{abs(theta_o_hand - row['theta_o_rad']) / row['theta_o_rad']:.2e}")

    # --- Check 2: tangent depression small-angle form ---
    print("\n    Check 2 — tangent depression small-angle form  Delta-h ~ L^2 / 8r")
    print(f"      {'range [km]':>11s} {'L^2/8r [m]':>12s} {'RADIANT [m]':>12s} {'rel diff':>10s}")
    for target_km in (25.0, 50.0, 100.0):
        entry = min(rows, key=lambda r: abs(r["range_m"] - target_km * 1000.0))
        approx = entry["range_m"] ** 2 / (8.0 * r_shell_m)
        print(f"      {entry['range_m'] / 1000.0:>11.1f} {approx:>12.2f} "
              f"{entry['dh_m']:>12.2f} {abs(approx - entry['dh_m']) / entry['dh_m']:>10.2e}")

    # --- Check 3: inverse-square x Beer-Lambert signal scaling ---
    near = min(rows, key=lambda r: abs(r["range_m"] - 25_000.0))
    far = min(rows, key=lambda r: abs(r["range_m"] - 50_000.0))
    predicted = (near["range_m"] / far["range_m"]) ** 2 * (far["tau_mwir"] / near["tau_mwir"])
    actual = far["signal_e"] / near["signal_e"]
    print("\n    Check 3 — point-source signal scaling S ~ I tau(R) / R^2")
    print(f"      predicted S(50 km)/S(25 km) = (25/50)^2 x (tau_50/tau_25) = {predicted:.6f}")
    print(f"      RADIANT   S(50 km)/S(25 km) = {actual:.6f}")
    print(f"      relative difference          = {abs(predicted - actual) / actual:.2e}")
    print("      This is the identity that proves the point-source term really is")
    print("      inverse-square in the slant range and Beer-Lambert in the band tau,")
    print("      with no hidden ground-projection factor sneaking in for an air target.")

    # --- Check 4: LOS rate (already computed in section 5) ---
    print("\n    Check 4 — relative LOS angular rate (see section 5 for the algebra)")
    print(f"      hand    = {kin['rate_hand_rad_s']:.9f} rad/s "
          f"= {kin['rate_hand_rad_s'] * 1e3:.5f} mrad/s")
    print(f"      RADIANT = {kin['rate_k2_rad_s']:.9f} rad/s "
          f"= {kin['rate_k2_rad_s'] * 1e3:.5f} mrad/s")
    print(f"      relative difference = "
          f"{abs(kin['rate_hand_rad_s'] - kin['rate_k2_rad_s']) / kin['rate_k2_rad_s']:.2e}")

    # --- Check 5: target-plane sample distance ---
    tpsd_hand = float(c["pitch_um"]) * 1e-6 * row["range_m"] / float(c["focal_length_m"])
    print("\n    Check 5 — target-plane sample distance  p d / f (the air-target GSD)")
    print(f"      hand    = {tpsd_hand:.6f} m")
    print(f"      RADIANT = {row['tpsd_m']:.6f} m")
    print(f"      relative difference = {abs(tpsd_hand - row['tpsd_m']) / row['tpsd_m']:.2e}")


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------


def make_figures(
    rows: list[dict[str, Any]],
    anchor: list[dict[str, Any]],
    kin: dict[str, Any],
    c: dict[str, float | str],
) -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    ranges_km = [r["range_m"] / 1000.0 for r in rows]
    warn_start_km = next(
        (r["range_m"] / 1000.0 for r in rows if r["guard_action"] == "warn"), None
    )

    def _shade(ax: Any) -> None:
        if warn_start_km is not None:
            ax.axvspan(
                warn_start_km,
                ranges_km[-1],
                color="orange",
                alpha=0.12,
                label="horizon-guard warning shoulder",
            )

    # Figure 1 — SNR and detection range vs range
    fig, ax_left = plt.subplots(figsize=(10, 6.5))
    ax_right = ax_left.twinx()
    _shade(ax_left)
    line_snr, = ax_left.semilogy(
        ranges_km, [r["snr"] for r in rows], "o-", color="tab:blue", label="SNR"
    )
    line_thr = ax_left.axhline(
        float(c["snr_threshold"]),
        color="tab:red",
        linestyle="--",
        label=f"detection threshold SNR = {float(c['snr_threshold']):.0f}",
    )
    line_det, = ax_right.plot(
        ranges_km,
        [r["detection_range_m"] / 1000.0 for r in rows],
        "s--",
        color="tab:green",
        label="detection range (path-aware solver)",
    )
    ax_left.set_xlabel("Level-arm slant range [km]")
    ax_left.set_ylabel("SNR [dimensionless]", color="tab:blue")
    ax_right.set_ylabel("Detection range [km]", color="tab:green")
    ax_left.tick_params(axis="y", labelcolor="tab:blue")
    ax_right.tick_params(axis="y", labelcolor="tab:green")
    ax_left.set_title(
        "10.2 air_to_air level arm at 10 km — SNR and detection range vs range\n"
        "MWIR 3.5-5.0 um, point-source target, simple atmosphere (midlat summer)"
    )
    ax_left.grid(True, alpha=0.3, which="both")
    handles = [line_snr, line_thr, line_det]
    if warn_start_km is not None:
        handles.append(ax_left.patches[0])
    ax_left.set_ylim(bottom=3.0)
    ax_left.legend(
        handles,
        [h.get_label() for h in handles],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.10),
        ncol=2,
        fontsize=9,
    )
    fig.tight_layout()
    path = OUTPUT_DIR / "10.2_snr_and_detection_range_vs_range.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # Figure 2 — transmittance vs range, model vs MODTRAN
    fig, ax = plt.subplots(figsize=(10, 6.5))
    _shade(ax)
    ax.plot(
        ranges_km,
        [r["tau_mwir"] for r in rows],
        "-",
        color="tab:blue",
        label="RADIANT level arm, MWIR 3.5-5.0 um (chain)",
    )
    styles = {"MWIR 3.5-5.0 um (sensor band)": ("tab:blue", "o"),
              "LWIR 8-12 um (reference)": ("tab:purple", "^")}
    for band_name, (colour, marker) in styles.items():
        band_rows = [r for r in anchor if r["band"] == band_name]
        if not band_rows:
            continue
        ax.plot(
            [r["range_m"] / 1000.0 for r in band_rows],
            [r["tau_modtran"] for r in band_rows],
            marker,
            color=colour,
            markersize=9,
            markerfacecolor="none",
            label=f"MODTRAN 6 L16-L20, {band_name}",
        )
        ax.plot(
            [r["range_m"] / 1000.0 for r in band_rows],
            [r["tau_model"] for r in band_rows],
            marker,
            color=colour,
            markersize=5,
            label=f"level arm, {band_name}",
        )
    ax.set_xlabel("Level-arm range [km]")
    ax.set_ylabel("Band-mean transmittance [dimensionless]")
    ax.set_title(
        "Cross-model anchor — analytic level arm vs MODTRAN 6 horizontal grid\n"
        "10 km constant altitude, midlat summer, 23 km visibility, rural aerosol"
    )
    ax.set_ylim(0.0, 1.0)
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=8, loc="lower left")
    fig.tight_layout()
    path = OUTPUT_DIR / "10.2_transmittance_vs_modtran_lgrid.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # Figure 3 — horizon guard: tangent depression vs range
    fig, ax = plt.subplots(figsize=(10, 6.5))
    ax.plot(ranges_km, [r["dh_m"] for r in rows], "o-", color="tab:brown",
            label="tangent depression Delta-h (core classifier)")
    ax.axhline(GUARD_DH_CLEAN_M, color="tab:orange", linestyle="--",
               label=f"clean / warn boundary = {GUARD_DH_CLEAN_M:.0f} m")
    ax.axhline(GUARD_DH_RAISE_M, color="tab:red", linestyle="--",
               label=f"warn / raise boundary = {GUARD_DH_RAISE_M:.0f} m")
    ax.plot(
        ranges_km,
        [(r["range_m"] ** 2) / (8.0 * (R_EARTH_M + float(c["h_sensor_m"]))) for r in rows],
        ":",
        color="black",
        label="hand check  L^2 / 8r",
    )
    ax.set_yscale("log")
    ax.set_xlabel("Level-arm slant range [km]")
    ax.set_ylabel("Tangent depression below the endpoints [m]")
    ax.set_title(
        "ADR-0011 horizon guard on a level arm — interior-tangent topology\n"
        "clean below 100 m, quantified refraction warning to 2 km, raise beyond"
    )
    ax.grid(True, alpha=0.3, which="both")
    ax.legend(fontsize=9, loc="lower right")
    fig.tight_layout()
    path = OUTPUT_DIR / "10.2_horizon_guard_tangent_depression.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # Figure 4 — LOS rate and smear budget vs range
    fig, ax_left = plt.subplots(figsize=(10, 6.5))
    ax_right = ax_left.twinx()
    rate_k0 = [r["los_rate_k0_rad_s"] for r in rows]
    rate_k2 = [r["los_rate_k2_rad_s"] for r in rows]
    l0, = ax_left.plot(ranges_km, [v * 1e3 for v in rate_k0], "o-", color="tab:gray",
                       label="omega_LOS, platform only (K0)")
    l2, = ax_left.plot(ranges_km, [v * 1e3 for v in rate_k2], "s-", color="tab:red",
                       label="omega_LOS, relative with crossing target (K2)")
    pitch_m = float(c["pitch_um"]) * 1e-6
    focal_m = float(c["focal_length_m"])
    l3, = ax_right.plot(
        ranges_km,
        [pitch_m / (v * focal_m) * 1e3 for v in rate_k2],
        "^--",
        color="tab:green",
        label="integration time for 1-pixel smear (K2)",
    )
    ax_left.set_xlabel("Level-arm slant range [km]")
    ax_left.set_ylabel("LOS angular rate [mrad/s]", color="tab:red")
    ax_right.set_ylabel("Integration time for 1-pixel smear [ms]", color="tab:green")
    ax_left.tick_params(axis="y", labelcolor="tab:red")
    ax_right.tick_params(axis="y", labelcolor="tab:green")
    ax_left.set_title(
        "Target kinematics (Gap 111) — relative LOS rate and the smear budget\n"
        "beam-aspect crossing target, own-ship and target both level at 10 km"
    )
    ax_left.grid(True, alpha=0.3)
    ax_left.legend([l0, l2, l3], [h.get_label() for h in (l0, l2, l3)], fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "10.2_los_rate_and_smear_budget.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    return written


# ---------------------------------------------------------------------------
# Results workbook (regenerate-on-demand; gitignored per Rule 26)
# ---------------------------------------------------------------------------


def write_results_workbook(
    rows: list[dict[str, Any]], anchor: list[dict[str, Any]]
) -> Path:
    workbook = openpyxl.Workbook()
    head_font = Font(bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2E75B6")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _write(ws: Any, headers: list[str], data: list[list[Any]]) -> None:
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row_idx, values in enumerate(data, start=2):
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col) + 3
            ws.column_dimensions[col[0].column_letter].width = width

    ws = workbook.active
    ws.title = "Range sweep"
    _write(
        ws,
        [
            "Range [km]", "theta_o [deg]", "Delta-h [m]", "Guard verdict",
            "tau band-mean MWIR [--]", "alpha_eff [1/km]", "Signal [e-]",
            "Noise [e- rms]", "SNR [--]", "Detection range [km]",
            "Well margin [dB]", "NEDT [mK]", "Target-plane sample distance [m]",
            "omega_LOS K0 [mrad/s]", "omega_LOS K2 [mrad/s]", "Smear K2 [um]",
        ],
        [
            [
                round(r["range_m"] / 1000.0, 3),
                round(math.degrees(r["theta_o_rad"]), 6),
                round(r["dh_m"], 2),
                r["guard_action"],
                round(r["tau_mwir"], 5),
                round(r["alpha_eff_per_km"], 6),
                round(r["signal_e"], 1),
                round(r["noise_e"], 2),
                round(r["snr"], 2),
                round(r["detection_range_m"] / 1000.0, 2),
                round(r["well_margin_dB"], 2),
                round(r["nedt_mK"], 2),
                round(r["tpsd_m"], 4),
                round(r["los_rate_k0_rad_s"] * 1e3, 5),
                round(r["los_rate_k2_rad_s"] * 1e3, 5),
                round(r["smear_k2_m"] * 1e6, 5),
            ]
            for r in rows
        ],
    )

    if anchor:
        ws2 = workbook.create_sheet("MODTRAN anchor")
        _write(
            ws2,
            ["Band", "Run", "Range [km]", "MODTRAN tau [--]", "Model tau [--]",
             "Ratio model/MODTRAN [--]", "Difference [%]"],
            [
                [
                    r["band"], r["run"], round(r["range_m"] / 1000.0, 1),
                    round(r["tau_modtran"], 5), round(r["tau_model"], 5),
                    round(r["ratio"], 4), round(r["diff_pct"], 2),
                ]
                for r in anchor
            ],
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(RESULTS_XLSX)
    return RESULTS_XLSX


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> None:
    vendor = read_vendor_inputs()
    c = to_canonical(vendor)

    print_inputs(vendor, c)

    nominal_result, nominal_warnings = _evaluate(make_config())
    print_geometry(nominal_result, c)
    print_regime(nominal_result, c)

    rows = sweep(c)
    print_horizon_guard(rows, c)
    kin = print_kinematics(c)
    anchor = print_modtran_anchor(c)
    rule4 = print_rule4(c)
    print_cross_checks(rows, kin, c)

    figures = make_figures(rows, anchor, kin, c)
    results_path = write_results_workbook(rows, anchor)

    print(f"\n{_rule()}")
    print("  SUMMARY")
    print(_rule())
    near, far = rows[0], rows[-1]
    print(f"\n    Scene class:            {nominal_result.stage_outputs['geometry']['scene_class']}"
          f" (derived, ADR-0011 decision 8)")
    print("    LOS direction:          "
          f"{nominal_result.stage_outputs['geometry']['los_direction']}")
    print(f"    Regime:                 {nominal_result.stage_outputs['optics']['regime']}")
    print(f"    SNR at {near['range_m'] / 1000.0:.0f} km:           "
          f"{near['snr']:.1f} (dimensionless)")
    print(f"    SNR at {far['range_m'] / 1000.0:.0f} km:          {far['snr']:.1f} (dimensionless)")
    print(f"    Detection range (ref {near['range_m'] / 1000.0:.0f} km): "
          f"{near['detection_range_m'] / 1000.0:.1f} km")
    print(f"    Detection range (ref {far['range_m'] / 1000.0:.0f} km): "
          f"{far['detection_range_m'] / 1000.0:.1f} km")
    print(f"    Delta-h at {near['range_m'] / 1000.0:.0f} km:        {near['dh_m']:.1f} m "
          f"({near['guard_action']})")
    print(f"    Delta-h at {far['range_m'] / 1000.0:.0f} km:       {far['dh_m']:.1f} m "
          f"({far['guard_action']})")
    print(f"    omega_LOS platform only: {kin['rate_k0_rad_s'] * 1e3:.4f} mrad/s at "
          f"{float(c['range_nominal_m']) / 1000.0:.0f} km")
    print(f"    omega_LOS with crosser:  {kin['rate_k2_rad_s'] * 1e3:.4f} mrad/s at "
          f"{float(c['range_nominal_m']) / 1000.0:.0f} km")
    print(f"    Rule-4 dual path:        passed_x={rule4['passed_x']} passed_y={rule4['passed_y']} "
          f"max_err={max(rule4['max_err_x'], rule4['max_err_y']):.2e} "
          f"(tol {rule4['tolerance']:.2e})")
    print(f"    Warnings at the nominal point: "
          f"{len(nominal_warnings)} (horizon guard silent at 50 km)")

    print("\n    Artifacts")
    for path in figures:
        print(f"      figure : {path.relative_to(REPO_ROOT)}")
    print(f"      workbook: {results_path.relative_to(REPO_ROOT)} (gitignored, "
          "regenerate-on-demand)")


if __name__ == "__main__":
    main()
