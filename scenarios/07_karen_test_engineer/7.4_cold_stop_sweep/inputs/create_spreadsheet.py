"""Create Karen's cold stop characterization spreadsheet.

Karen is a test engineer in a thermal-vacuum chamber. She has:
  - An instrument spec sheet with optics and detector parameters (vendor units)
  - Lab background measurements at several cold stop positions
  - System-level performance requirements

All values are in the units Karen actually works with in the lab:
  - Lengths in cm and mm, pixel pitch in µm
  - Percentages (not fractions) for QE, transmission, efficiency
  - Temperatures in K (standard for cryo work)
  - Integration time in ms
  - Currents in pA/pixel for dark current
  - Wavelengths in nm for band edges
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=value).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value=unit).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ============================================================================
# Sheet 1: Instrument Spec Sheet
# ============================================================================
# This mimics what Karen gets from the instrument vendor / design team.
# Units are whatever the vendor uses — not RADIANT canonical.

ws1 = wb.active
ws1.title = "Instrument Spec Sheet"

ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 15
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 55

ws1["A1"] = "MWIR Imager — Cold Stop Characterization Test"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "Karen Torres — TVAC Test Campaign, April 2026"
ws1["A2"].font = Font(italic=True, size=10)
ws1["A3"] = "Source: Instrument CDR Spec Sheet Rev D"
ws1["A3"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws1.cell(row=5, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 6

# Optics section — note vendor uses cm for aperture, mm for focal length
row = add_section(ws1, row, "Telescope / Optics")
row = add_param(ws1, row, "Primary aperture diameter", 25, "cm",
                "Cassegrain, 20% central obscuration")
row = add_param(ws1, row, "Effective focal length", 1000, "mm",
                "Measured at 4.0 µm reference wavelength")
row = add_param(ws1, row, "f-number", 4.0, "—", "Effective f/#")
row = add_param(ws1, row, "End-to-end optical transmission", 68, "%",
                "Includes all mirrors, lenses, and filter")
row = add_param(ws1, row, "Number of optical elements", 5, "—",
                "Primary, secondary, fold mirror, field lens, filter")
row = add_param(ws1, row, "Optics barrel temperature", 20, "°C",
                "Ambient temp in TVAC shroud (controlled)")
row = add_param(ws1, row, "Cold stop design efficiency", 100, "%",
                "Fully baffled cold shield, designed for 100% rejection")
row = add_param(ws1, row, "WFE (RMS)", 0.07, "waves",
                "At 4.0 µm; spec is < 0.10 waves")

# Detector section — vendor uses pA/pixel for dark current
row = add_section(ws1, row, "Detector (HgCdTe FPA)")
row = add_param(ws1, row, "Pixel pitch", 20, "µm", "Square pixels")
row = add_param(ws1, row, "Array format", "640 × 512", "pixels", "")
row = add_param(ws1, row, "Cutoff wavelength", 5200, "nm",
                "50% response at 77 K operating temp")
row = add_param(ws1, row, "Average QE (in-band)", 75, "%",
                "Vendor spec, 3.7–4.8 µm weighted average")
row = add_param(ws1, row, "Dark current (mean)", 80, "fA/pixel",
                "At 77 K; vendor datasheet value")
row = add_param(ws1, row, "Operating temperature", 77, "K",
                "Stirling cooler setpoint")
row = add_param(ws1, row, "Full well capacity", 8.5e6, "e⁻",
                "CTIA readout, large well mode")
row = add_param(ws1, row, "Read noise (CDS)", 25, "e⁻ RMS",
                "After correlated double sampling")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "")
row = add_param(ws1, row, "System gain", 2.5, "e⁻/DN", "Calibrated")

# Scene / Calibration — lab blackbody source
row = add_section(ws1, row, "Calibration Source (in TVAC)")
row = add_param(ws1, row, "Blackbody temperature", 35, "°C",
                "Extended-area blackbody, fills entire FOV")
row = add_param(ws1, row, "Blackbody emissivity", 0.98, "—",
                "CI Systems SR-800 calibration cert")
row = add_param(ws1, row, "Chamber shroud temperature", 20, "°C",
                "Painted Aeroglaze Z306 (ε ≈ 0.95)")
row = add_param(ws1, row, "Chamber shroud emissivity", 0.95, "—", "")

# Spectral / Integration
row = add_section(ws1, row, "Spectral / Integration")
row = add_param(ws1, row, "Cold filter passband", "3700 – 4800", "nm",
                "Narrowband MWIR, specified in nm on vendor datasheet")
row = add_param(ws1, row, "Integration time", 8, "ms",
                "Nominal frame time for background characterization")

# Geometry — lab test (essentially zero range)
row = add_section(ws1, row, "Test Geometry")
row = add_param(ws1, row, "Source distance", 1.5, "m",
                "Blackbody to entrance pupil")

# ============================================================================
# Sheet 2: Lab Background Measurements
# ============================================================================
# Karen measured the background signal level with the blackbody shuttered
# (cold plate blocking the aperture) at several cold stop positions.
# She misaligned the cold stop deliberately to get a range of efficiencies.
# These measurements are what Karen wants to reconcile with the model.

ws2 = wb.create_sheet("Background Measurements")

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 22
ws2.column_dimensions["F"].width = 45

ws2["A1"] = "Cold Stop Alignment — Background Signal Measurements"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "TVAC Chamber 3, 14-Apr-2026"
ws2["A2"].font = Font(italic=True, size=10)
ws2["A3"] = ("Blackbody shuttered (cold plate at 77 K blocking aperture). "
             "Background signal is from warm optics + cold stop leakage only.")
ws2["A3"].font = Font(italic=True, size=10)

meas_headers = [
    "Test Point",
    "Cold Stop Position [mm]",
    "Mean Bkg Signal [DN]",
    "Bkg Noise (σ) [DN]",
    "Mean Bkg Signal [e⁻]",
    "Notes",
]
for col, h_text in enumerate(meas_headers, 1):
    cell = ws2.cell(row=5, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Karen's measurements: she moved the cold stop in 0.5mm increments.
# Position 0.0 mm is nominal (perfectly aligned, design efficiency ≈ 100%).
# As the cold stop moves off-center, warm background leaks in.
# She records in DN (the raw number from the ADC), and converts to e⁻
# using the calibrated system gain of 2.5 e⁻/DN.
#
# The gain of 2.5 e⁻/DN means:
#   DN 14000 × 2.5 = 35,000 e⁻  (nominal)
#   DN 18000 × 2.5 = 45,000 e⁻  (suspected misalignment)

measurements = [
    ("CS-NOM",    0.0,  14200,  85,  35500,  "Nominal alignment, coldshield seated"),
    ("CS-OFF-05", 0.5,  15800, 102,  39500,  "0.5 mm lateral offset"),
    ("CS-OFF-10", 1.0,  17600, 118,  44000,  "1.0 mm offset — near expected anomaly"),
    ("CS-OFF-15", 1.5,  19100, 130,  47750,  "1.5 mm offset"),
    ("CS-OFF-20", 2.0,  20800, 143,  52000,  "2.0 mm offset — significant leakage"),
    ("CS-OFF-25", 2.5,  22300, 155,  55750,  "2.5 mm offset — approaching FPA damage threshold"),
]

for i, (test_id, pos, dn_mean, dn_sigma, e_mean, notes) in enumerate(measurements, 6):
    ws2.cell(row=i, column=1, value=test_id).border = thin_border
    ws2.cell(row=i, column=2, value=pos).border = thin_border
    ws2.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=3, value=dn_mean).border = thin_border
    ws2.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=4, value=dn_sigma).border = thin_border
    ws2.cell(row=i, column=4).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=5, value=e_mean).border = thin_border
    ws2.cell(row=i, column=5).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=6, value=notes).border = thin_border

# Add a note about how e⁻ was derived from DN
note_row = len(measurements) + 7
ws2.cell(row=note_row, column=1, value="Conversion:").font = Font(bold=True)
ws2.cell(row=note_row, column=2, value="e⁻ = DN × gain (2.5 e⁻/DN)")

# ============================================================================
# Sheet 3: Performance Requirements
# ============================================================================

ws3 = wb.create_sheet("Performance Requirements")

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 50

ws3["A1"] = "System Performance Requirements"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Source: System Requirements Document SRD-2024-042 Rev C"
ws3["A2"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Requirement", "Threshold", "Unit", "Notes"], 1):
    cell = ws3.cell(row=4, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

requirements = [
    ("Max background signal (shuttered)", 40000, "e⁻",
     "With calibration source blocked; excess indicates cold stop leakage"),
    ("SNR (extended scene, 35°C BB)", "≥ 300", "—",
     "At nominal integration time and scene temperature"),
    ("Background noise contribution", "< 50%", "of total noise",
     "Background shot noise must not dominate noise budget"),
    ("Cold stop efficiency", "≥ 95", "%",
     "Minimum acceptable cold stop efficiency after alignment"),
]

for i, (req, thresh, unit, notes) in enumerate(requirements, 5):
    ws3.cell(row=i, column=1, value=req).border = thin_border
    ws3.cell(row=i, column=2, value=thresh).border = thin_border
    ws3.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws3.cell(row=i, column=3, value=unit).border = thin_border
    ws3.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    ws3.cell(row=i, column=4, value=notes).border = thin_border

output_path = "karen_cold_stop_data.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
