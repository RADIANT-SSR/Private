#!/usr/bin/env python3
"""Create Mike's inputs for Scenario 2.4: persistence characterization.

Emits ``mike_persistence.xlsx`` — the measured persistence parameters, the
bright prior exposure, and the current-scene / detector configuration.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Persistence"
ws["A1"] = "Scenario 2.4: Type-II SLS LWIR persistence characterization"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [30, 14, 12, 42]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="7030A0")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Persistence fraction", 1.5, "%", "Residual after one frame at 60 Hz"),
    ("Decay time constant", 50.0, "ms", "Trap release τ"),
    ("Prior-frame signal", 150000.0, "e⁻", "Hot 800 K calibration source"),
    ("Frame rate", 60.0, "Hz", "→ 16.67 ms frame interval"),
    ("Current integration time", 10.0, "ms", ""),
    ("Current scene signal", 20000.0, "e⁻", "Nominal scene in the current frame"),
    ("Dark current", 5.0e5, "e⁻/s", "SLS LWIR at operating temp"),
    ("Read noise", 300.0, "e⁻ RMS", ""),
    ("System gain", 100.0, "e⁻/DN", "1 LSB = 100 e⁻"),
    ("Frames to simulate", 20, "—", ""),
]
r = 4
for name, value, unit, note in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "mike_persistence.xlsx"
wb.save(out)
print(f"Wrote {out}")
