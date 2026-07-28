"""Generate the vendor-format input workbook for scenario 10.2.

The workbook is deliberately **not** in RADIANT canonical units
(``docs/guides/scenario_testing.md`` Rule 1).  It is what an airborne-IRST
programme actually circulates:

* the optical head from a sensor-house datasheet — millimetres, per-cent,
  degrees Celsius, waves RMS;
* the FPA/ROIC from a detector datasheet — micrometres, per-cent,
  kilo-electrons, milliseconds;
* the engagement from a flight-test card — kilometres, knots, degrees;
* the atmosphere from a met brief — centimetres of precipitable water,
  kilometres of visibility.

``scripts/run_air_to_air_level_irst.py`` reads this file and performs every
vendor -> canonical conversion once, each with an explicit comment.

Run from anywhere::

    python scenarios/10_direction_general/10.2_air_to_air_level_irst/inputs/create_spreadsheet.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

OUT = Path(__file__).resolve().parent / "irst_air_to_air_vendor_data.xlsx"

_TITLE_FONT = Font(bold=True, size=13)
_SUB_FONT = Font(italic=True, size=9, color="555555")
_HEAD_FONT = Font(bold=True, size=10, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="2E75B6")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Every sheet uses the same four-column layout, starting at row 5, so the
# reader in scripts/ has exactly one parsing rule.
_COLUMNS = ("Parameter", "Value", "Unit", "Note")


def _sheet(wb: openpyxl.Workbook, title: str, subtitle: str) -> Worksheet:
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = _SUB_FONT
    for col, name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col, value=name)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
    return ws


def _rows(ws: Worksheet, rows: list[tuple[str, object, str, str]]) -> None:
    for offset, (name, value, unit, note) in enumerate(rows):
        for col, value_ in enumerate((name, value, unit, note), start=1):
            cell = ws.cell(row=5 + offset, column=col, value=value_)
            cell.border = _BORDER
            if col == 2:
                cell.alignment = Alignment(horizontal="center")
    widths = (34, 14, 16, 78)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width


def build() -> Path:
    """Write the workbook and return its path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = _sheet(
        wb,
        "IRST Optical Head",
        "Sensor-house datasheet extract — mm / % / degC / waves RMS",
    )
    _rows(
        ws,
        [
            ("Entrance pupil diameter", 150.0, "mm", "Clear aperture of the gimballed MWIR head"),
            ("Effective focal length", 450.0, "mm", "f/3 — fixed focus at infinity"),
            ("f-number", 3.0, "--", "Derived; carried for the datasheet cross-check"),
            ("Optical transmission", 75.0, "%", "Band-averaged over 3.5-5.0 um, all elements"),
            ("Optics temperature", -23.15, "degC", "Uncooled fore-optics soaked at flight OAT"),
            ("Central obscuration", 0.0, "%", "Refractive head — no obscuration"),
            ("Wavefront error RMS", 0.05, "waves", "At the 4.25 um band centre"),
            ("Spectral filter min", 3.50, "um", "Cold filter cut-on"),
            ("Spectral filter max", 5.00, "um", "Cold filter cut-off"),
        ],
    )

    ws = _sheet(
        wb,
        "FPA and ROIC",
        "Detector datasheet extract — um / % / ke- / ms",
    )
    _rows(
        ws,
        [
            ("Pixel pitch", 20.0, "um", "Square pixel, InSb"),
            ("Fill factor", 100.0, "%", "Backside-illuminated, no microlens loss quoted"),
            ("Quantum efficiency", 80.0, "%", "Band-averaged 3.5-5.0 um"),
            ("Dark current", 50000.0, "e-/s", "At the quoted operating temperature"),
            ("Operating temperature", 80.0, "K", "Stirling-cooled cold shield + FPA"),
            ("Read noise", 40.0, "e- RMS", "CTIA input cell, high-flux mode"),
            ("Full well capacity", 1000.0, "ke-", "High-flux ROIC well"),
            ("System gain", 61.0, "e-/DN", "Well / 2^14 — ADC matched to the well"),
            ("ADC resolution", 14, "bits", "Digital video output"),
            ("Frame integration time", 0.10, "ms", "100 us search-frame integration"),
        ],
    )

    ws = _sheet(
        wb,
        "Engagement",
        "Flight-test card — km / knots / degrees / degC. Level (co-altitude) arm.",
    )
    _rows(
        ws,
        [
            ("Own-ship altitude", 10.0, "km", "Pressure altitude, MSL"),
            ("Target altitude", 10.0, "km", "Co-altitude — this is the LEVEL arm"),
            ("Range sweep start", 25.0, "km", "Slant range along the level arm"),
            ("Range sweep stop", 100.0, "km", "Slant range along the level arm"),
            ("Range sweep step", 5.0, "km", "16 sweep points"),
            ("Nominal range", 50.0, "km", "The baseline point the GUI config is built at"),
            ("Target hot-parts temperature", 226.85, "degC",
             "Tail-aspect nozzle + near plume — the MWIR signature driver"),
            ("Target hot-parts area", 0.36, "m^2",
             "Effective emitting area of the nozzle/plume, tail aspect"),
            ("Target emissivity", 90.0, "%", "Nozzle hot metal + plume, band-averaged"),
            ("Own-ship true airspeed", 480.0, "kt", "Level cruise"),
            ("Target true airspeed", 580.0, "kt", "Level cruise + shallow climb"),
            ("Target heading", 270.0, "deg",
             "From the observer ground azimuth — beam-aspect crosser"),
            ("Target climb angle", 2.0, "deg", "Shallow climb, positive up"),
            ("Detection SNR threshold", 5.0, "--", "Single-frame declaration threshold"),
        ],
    )

    ws = _sheet(
        wb,
        "Atmosphere",
        "Met brief — profile name / cm of precipitable water / km visibility",
    )
    _rows(
        ws,
        [
            ("Standard atmosphere", "midlat_summer", "--", "Matches the MODTRAN L-grid runs"),
            ("Precipitable water", 2.92, "cm", "Mid-latitude summer column value"),
            ("Visibility", 23.0, "km", "Rural haze, matches the L-grid runs"),
            ("Aerosol type", "rural", "--", "Matches the L-grid runs"),
            ("Illumination", "night", "--", "Night engagement — no solar term"),
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(f"wrote {build()}")
