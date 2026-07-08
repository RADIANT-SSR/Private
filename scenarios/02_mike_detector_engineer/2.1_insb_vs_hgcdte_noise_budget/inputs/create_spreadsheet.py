#!/usr/bin/env python3
"""Create Mike's vendor inputs for Scenario 2.1: InSb vs. HgCdTe shootout.

Emits the five files a real detector trade starts from:

- ``insb_qe.csv``       — InSb QE curve, VENDOR FORMAT A: wavelength_nm, QE_pct
- ``insb_jdark.csv``    — InSb J_dark(T), columns T_K, Jdark_A_cm2
- ``hgcdte_qe.csv``     — HgCdTe QE curve, VENDOR FORMAT B: lambda_um, quantum_efficiency
- ``hgcdte_jdark.csv``  — HgCdTe J_dark(T), columns T_K, Jdark_A_cm2
- ``mike_roic_specs.xlsx`` — shared ROIC + bench-test configuration

The two QE files deliberately use different column conventions — that
mismatch is the point of the scenario (radiant.io.qe_csv resolves both).

Dark-current curves are Arrhenius-generated (J = J0·exp(−Ea/kT)) with
physically motivated parameters:

- InSb  (λc ≈ 5.4 µm): Ea = 0.225 eV (≈ InSb band gap), anchored to
  J(77 K) = 2.0e-8 A/cm² — typical diffusion-limited InSb.
- HgCdTe (λc ≈ 5.25 µm): Ea = 0.236 eV (Eg = 1.24/5.25), anchored to
  J(77 K) = 5.0e-10 A/cm² — MWIR MCT runs about 40× lower dark current
  than InSb at the same temperature (the known trade against InSb's
  better uniformity/operability).

Run:  python create_spreadsheet.py
"""

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent
K_B_EV = 8.617333262e-5  # Boltzmann constant [eV/K]

# ---------------------------------------------------------------------------
# QE curves
# ---------------------------------------------------------------------------

# InSb — vendor format A: wavelength_nm, QE_pct
INSB_QE_NM_PCT = [
    (3000, 68.0), (3250, 74.0), (3500, 79.0), (3750, 83.0),
    (4000, 86.0), (4250, 88.0), (4500, 89.0), (4750, 88.0),
    (5000, 85.0), (5200, 78.0), (5300, 65.0), (5400, 40.0),
    (5450, 15.0), (5500, 3.0),
]

with open(HERE / "insb_qe.csv", "w", encoding="utf-8") as fh:
    fh.write("# IRA-3541 InSb FPA — spectral quantum efficiency (vendor test report)\n")
    fh.write("# 77 K, per-pixel average over the full array\n")
    fh.write("wavelength_nm,QE_pct\n")
    for nm, pct in INSB_QE_NM_PCT:
        fh.write(f"{nm},{pct}\n")
print("Wrote insb_qe.csv")

# HgCdTe — vendor format B: lambda_um, quantum_efficiency (fraction)
HGCDTE_QE_UM_FRAC = [
    (3.0, 0.72), (3.3, 0.74), (3.6, 0.76), (3.9, 0.77),
    (4.2, 0.78), (4.5, 0.80), (4.8, 0.80), (5.0, 0.79),
    (5.1, 0.75), (5.2, 0.60), (5.25, 0.40), (5.3, 0.18),
]

with open(HERE / "hgcdte_qe.csv", "w", encoding="utf-8") as fh:
    fh.write("# MCT-5250 HgCdTe FPA — spectral quantum efficiency (vendor datasheet)\n")
    fh.write("# 77 K, lambda_cutoff = 5.25 um\n")
    fh.write("lambda_um,quantum_efficiency\n")
    for um, frac in HGCDTE_QE_UM_FRAC:
        fh.write(f"{um},{frac}\n")
print("Wrote hgcdte_qe.csv")

# ---------------------------------------------------------------------------
# Dark current curves — Arrhenius with vendor anchors
# ---------------------------------------------------------------------------

TEMPS_K = list(range(60, 135, 5))  # 60..130 K (covers BLIP temps ~100-120 K)


def arrhenius_series(ea_ev: float, t_anchor_k: float, j_anchor: float) -> list[tuple[int, float]]:
    j0 = j_anchor / math.exp(-ea_ev / (K_B_EV * t_anchor_k))
    return [(t, j0 * math.exp(-ea_ev / (K_B_EV * t))) for t in TEMPS_K]


INSB_JDARK = arrhenius_series(ea_ev=0.225, t_anchor_k=77.0, j_anchor=2.0e-8)
HGCDTE_JDARK = arrhenius_series(ea_ev=0.236, t_anchor_k=77.0, j_anchor=5.0e-10)

for name, series, label in [
    ("insb_jdark.csv", INSB_JDARK, "IRA-3541 InSb"),
    ("hgcdte_jdark.csv", HGCDTE_JDARK, "MCT-5250 HgCdTe"),
]:
    with open(HERE / name, "w", encoding="utf-8") as fh:
        fh.write(f"# {label} — measured dark current density vs temperature\n")
        fh.write("# Vendor acceptance test data, mean over array\n")
        fh.write("T_K,Jdark_A_cm2\n")
        for t, j in series:
            fh.write(f"{t},{j:.6e}\n")
    print(f"Wrote {name}")

# ---------------------------------------------------------------------------
# ROIC + bench configuration workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ROIC and Test Config"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
section_fill = PatternFill("solid", fgColor="D9E2F3")

ws["A1"] = "Scenario 2.1: InSb vs. HgCdTe Shootout — Shared ROIC + Bench Setup"
ws["A1"].font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 52

for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = header_font
    ws[f"{col}3"].fill = header_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

rows = [
    ("ROIC", None, None, None),
    ("Pixel pitch", 15.0, "µm", "Same ROIC for both FPAs"),
    ("Node capacitance", 33.0, "fF", "Charge integration node"),
    ("CDS mode", 1, "—", "1 = correlated double sampling ON (kTC suppressed)"),
    ("ROIC glow", 5.0, "e⁻/s", "Self-emission of the readout circuit"),
    ("Full well capacity", 5000000, "e⁻", "At nominal bias"),
    ("ADC resolution", 14, "bits", ""),
    ("System gain", 305.0, "e⁻/DN", "Well / 2^14"),
    ("Read noise (CDS) — InSb FPA", 18.0, "e⁻ RMS", "Vendor acceptance test"),
    ("Read noise (CDS) — HgCdTe FPA", 12.0, "e⁻ RMS", "Vendor acceptance test"),
    ("Bench test", None, None, None),
    ("Blackbody temperature", 300.0, "K", "Flat-plate source filling the aperture"),
    ("Blackbody emissivity", 0.995, "—", "Cavity-type calibration source"),
    ("Collimator aperture", 2.5, "cm", ""),
    ("Collimator focal length", 5.75, "cm", "f/2.3"),
    ("Optical transmission", 90.0, "%", "Collimator + window"),
    ("Optics temperature", 295.0, "K", "Bench ambient"),
    ("Cold filter passband", "3500–5000", "nm", "Common MWIR band for the trade"),
    ("Integration time", 1.0, "ms", ""),
    ("Operating temperature (nominal)", 77.0, "K", "Both FPAs; trade explores warmer set points"),
]

r = 4
for name, value, unit, note in rows:
    if value is None and unit is None:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=11)
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = section_fill
    else:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "mike_roic_specs.xlsx"
wb.save(out)
print(f"Wrote {out}")
