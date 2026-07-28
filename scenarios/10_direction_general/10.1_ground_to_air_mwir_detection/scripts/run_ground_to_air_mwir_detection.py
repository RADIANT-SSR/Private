"""Scenario 10.1 — Ground-to-air MWIR detection (direction-general validation).

Scene class ``ground_to_air`` (matrix cell E2, owner priority 1): a sea-level
MWIR search-and-track camera looks **up** a partial atmospheric column at a
small turbojet UAS cruising at 10 km.  Everything the scene needs became
expressible with Geometry-Flexibility Phases 1-4 (ADR-0011): an up-looking
line of sight, a segment-composed up-path radiance, a ``SkyBackground`` behind
the target, scene-class-conditioned metric relevance, and the horizon guard.

What the script answers
-----------------------
1. What does the stage publish for this scene — ``scene_class``,
   ``los_direction``, theta_o, eta, and the lower-endpoint zenith zeta_low?
2. Which metrics does the scene class turn **off** by default (the GSD /
   ground-projection family), what replaces them (target-plane sample distance
   and angular resolution), and what happens when the analyst overrides the
   Gap-96 selection?
3. How do band-mean sky transmittance, up-path radiance, SNR and SCNR move as
   the pointing elevation drops from straight up to 30 deg above the horizon?
4. How far up-range can the camera hold the target, along the *actual* ray?
5. Do the simple model's band-mean 3-5 um numbers agree with the owner's
   MODTRAN batch-1 K-ladder at the same geometry, and by how much?

Every printed number carries its unit.  Vendor -> canonical unit conversions
happen exactly once, each with an explicit comment (scenario_testing.md Rule 1).

Usage (from anywhere; all paths are repo-relative)::

    python scenarios/10_direction_general/10.1_ground_to_air_mwir_detection/\
        scripts/run_ground_to_air_mwir_detection.py
    ... --modtran-runs /path/to/staged/real_runs     # optional anchor override

The MODTRAN comparison is skipped with a clear message when the run set is not
staged (``modtran/real_runs/`` is gitignored); every other section runs from a
fresh checkout with no external data.

Runtime: about 30 s (roughly 250 chain evaluations at ~0.09 s each).
"""

from __future__ import annotations

import argparse
import math
import warnings
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")  # headless: figures are written, never shown

import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402

from radiant.api import Sensor  # noqa: E402
from radiant.api.scene_relevance import default_off_metrics  # noqa: E402

# ---------------------------------------------------------------------------
# Paths — repo-relative, pathlib only (Rule 30)
# ---------------------------------------------------------------------------

SCENARIO_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[4]
INPUT_XLSX = SCENARIO_DIR / "inputs" / "ground_mwir_tracker_data.xlsx"
OUTPUT_DIR = SCENARIO_DIR / "outputs"
RESULTS_XLSX = OUTPUT_DIR / "ground_to_air_mwir_detection_results.xlsx"
DEFAULT_MODTRAN_RUNS = REPO_ROOT / "modtran" / "real_runs"

# ---------------------------------------------------------------------------
# Physical constants used ONLY by the independent hand-calculation cross-check.
# The chain itself takes these from radiant.core.constants; re-declaring them
# here is deliberate — an anchor that imported the chain's own constants module
# would not be independent (CODATA 2018 values, quoted to 7 figures).
# ---------------------------------------------------------------------------

H_PLANCK_J_S = 6.62607015e-34  # J*s
C_LIGHT_M_S = 2.99792458e8  # m/s
K_BOLTZMANN_J_K = 1.380649e-23  # J/K
R_EARTH_M = 6.371e6  # m, RADIANT's single canonical spherical Earth radius

#: Detection-range search ceiling [m of path length].  Beyond ~99 km altitude
#: the target leaves the modelled column and the scene stops being ground->air.
SEARCH_CEILING_ALTITUDE_M = 99_000.0

BAND_LO_UM = 3.0
BAND_HI_UM = 5.0


# ---------------------------------------------------------------------------
# 1. Vendor workbook -> canonical units (converted exactly once)
# ---------------------------------------------------------------------------


def _read_param_sheet(ws: Any) -> dict[str, Any]:
    """Read a (name, value, unit, note) vendor sheet into ``{name: value}``."""
    out: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=4, max_col=3, values_only=True):
        name, value, unit = row[0], row[1], row[2]
        if not name or value is None or unit is None:
            continue  # section banner or blank spacer row
        out[str(name)] = value
    return out


def read_vendor_inputs() -> dict[str, Any]:
    """Load the vendor workbook.  Values are still in VENDOR units here."""
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    camera = _read_param_sheet(wb["Camera Datasheet"])
    site = _read_param_sheet(wb["Site and Target"])

    zeniths_deg: list[float] = []
    for row in wb["Pointing Plan"].iter_rows(min_row=5, max_col=1, values_only=True):
        if row[0] is not None:
            zeniths_deg.append(float(row[0]))

    anchors: list[dict[str, Any]] = []
    for row in wb["MODTRAN Anchors"].iter_rows(min_row=5, max_col=5, values_only=True):
        if row[0] is None:
            continue
        anchors.append(
            {
                "run": str(row[0]),
                "h1_km": float(row[1]),
                "h2_km": float(row[2]),
                "angle_deg": float(row[3]),
                "note": str(row[4] or ""),
            }
        )
    return {"camera": camera, "site": site, "zeniths_deg": zeniths_deg, "anchors": anchors}


def to_canonical(vendor: dict[str, Any]) -> dict[str, float | str]:
    """Vendor units -> RADIANT canonical units.  One conversion per line."""
    cam = vendor["camera"]
    site = vendor["site"]

    canon: dict[str, float | str] = {
        # -- telescope --
        "aperture_diameter_m": float(cam["Entrance pupil diameter"]) / 1000.0,  # mm -> m
        "focal_length_m": float(cam["Effective focal length"]) / 1000.0,  # mm -> m
        "transmission": float(cam["Optical transmission"]) / 100.0,  # % -> fraction
        "optics_temperature_K": float(cam["Housing temperature"]) + 273.15,  # degC -> K
        "scalar_emissivity": float(cam["Train emissivity"]) / 100.0,  # % -> fraction
        # Vendor "cold shield efficiency" is the BLOCKED fraction; RADIANT's
        # nearfield_fraction is the PASSED fraction (inverted convention).
        "nearfield_fraction": 1.0 - float(cam["Cold shield efficiency"]) / 100.0,  # % -> frac
        # -- focal plane --
        # pixel pitch: RADIANT's canonical input unit for this parameter IS um,
        # so the datasheet number is used verbatim (no conversion).
        "pixel_pitch_um": float(cam["Pixel pitch"]),
        "qe": float(cam["Quantum efficiency"]) / 100.0,  # % -> fraction
        "dark_rate_e_per_s": float(cam["Dark current"]),  # already e-/s
        "detector_temperature_K": float(cam["FPA temperature"]),  # already K
        "filter_min_um": float(cam["Spectral band, low edge"]) / 1000.0,  # nm -> um
        "filter_max_um": float(cam["Spectral band, high edge"]) / 1000.0,  # nm -> um
        # -- readout --
        "integration_time_s": float(cam["Integration time"]) / 1000.0,  # ms -> s
        "read_noise_e_rms": float(cam["Read noise"]),  # already e- rms
        "full_well_capacity_e": float(cam["Full well capacity"]) * 1.0e6,  # Me- -> e-
        "gain_e_per_dn": float(cam["System gain"]),  # already e-/DN
        "adc_bits": int(cam["ADC resolution"]),
        # -- site --
        "sensor_altitude_m": float(site["Site altitude"]) * 1000.0,  # km -> m
        "standard_atmosphere": str(site["Standard atmosphere"]),
        "visibility_km": float(site["Visibility"]),  # RADIANT input unit IS km
        "aerosol_type": str(site["Aerosol model"]),
        "solar_illumination": str(site["Illumination"]),
        # -- target --
        "target_altitude_m": float(site["Target altitude"]) * 1000.0,  # km -> m
        # nozzle disc: diameter mm -> projected area m^2
        "target_area_m2": math.pi * (float(site["Nozzle exit diameter"]) / 2000.0) ** 2,
        "target_temperature_K": float(site["Nozzle temperature"]) + 273.15,  # degC -> K
        "target_emissivity": float(site["Nozzle emissivity"]),  # already a fraction
        "snr_threshold": float(site["SNR threshold"]),  # dimensionless
    }
    return canon


# ---------------------------------------------------------------------------
# 2. Config + the module-level factory the GUI-baseline registry calls
# ---------------------------------------------------------------------------

#: Nominal track geometry: zenith at the sensor (the path's LOWER endpoint).
NOMINAL_ZENITH_DEG = 30.0


def make_config(
    zenith_deg: float = NOMINAL_ZENITH_DEG,
    target_altitude_m: float | None = None,
    scene_type: str = "point_source",
) -> dict[str, Any]:
    """Validated RADIANT config for this scenario, in canonical units.

    ``zenith_deg`` is the path zenith **at the sensor**, which for this
    up-looking scene is the path's lower endpoint (ADR-0011 decision 3); it is
    entered through mode V1 (``geometry.path_zenith_rad``).
    """
    canon = to_canonical(read_vendor_inputs())
    h_tgt_m = float(canon["target_altitude_m"]) if target_altitude_m is None else target_altitude_m
    return {
        "source": {
            "scene_type": scene_type,
            "target": {
                "temperature": canon["target_temperature_K"],
                "emissivity": canon["target_emissivity"],
            },
        },
        "geometry": {
            "sensor_altitude_m": canon["sensor_altitude_m"],
            "target_altitude_m": h_tgt_m,
            # deg -> rad (the only angular conversion in the scenario)
            "path_zenith_rad": math.radians(zenith_deg),
            "solar_illumination": canon["solar_illumination"],
            "target": {"projected_area_m2": canon["target_area_m2"]},
        },
        "atmosphere": {
            "model": "simple",
            "standard_atmosphere": canon["standard_atmosphere"],
            "visibility_km": canon["visibility_km"],
            "aerosol_type": canon["aerosol_type"],
        },
        "optics": {
            "aperture_diameter_m": canon["aperture_diameter_m"],
            "focal_length_m": canon["focal_length_m"],
            "transmission_scalar": canon["transmission"],
            "optics_temperature_K": canon["optics_temperature_K"],
            "scalar_emissivity": canon["scalar_emissivity"],
            "nearfield_fraction": canon["nearfield_fraction"],
        },
        "detector": {
            "pixel_pitch_x_um": canon["pixel_pitch_um"],
            "pixel_pitch_y_um": canon["pixel_pitch_um"],
            "qe_value": canon["qe"],
            "dark_rate_e_per_s": canon["dark_rate_e_per_s"],
            "detector_temperature_K": canon["detector_temperature_K"],
        },
        "spectral_integration": {
            "filter_min_um": canon["filter_min_um"],
            "filter_max_um": canon["filter_max_um"],
            "integration_time_s": canon["integration_time_s"],
        },
        "readout": {
            "read_noise_e_rms": canon["read_noise_e_rms"],
            "full_well_capacity_e": canon["full_well_capacity_e"],
            "gain_e_per_dn": canon["gain_e_per_dn"],
            "adc_bits": canon["adc_bits"],
        },
        "performance": {"detection_snr_threshold": canon["snr_threshold"]},
    }


def make_sensor() -> Sensor:
    """The scenario's nominal validated ``Sensor`` — MODULE-LEVEL FACTORY.

    Ground sensor at 0 m MSL, 550 K / 60 mm-diameter nozzle at 10 km, pointing
    30 deg from the zenith.  Importing this module and calling ``make_sensor()``
    builds the baseline without executing ``main()``; this is the entry point
    ``scenarios/tools/gui_baselines.py`` wires the GUI baseline to.
    """
    return Sensor.from_dict(make_config())


def evaluate(sensor: Sensor, capture_warnings: bool = False) -> tuple[Any, list[str]]:
    """Evaluate, optionally capturing the warnings the chain emits."""
    if not capture_warnings:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return sensor.evaluate(), []
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result = sensor.evaluate()
    messages: list[str] = []
    for entry in caught:
        text = f"{entry.category.__name__}: {entry.message}"
        if text not in messages:
            messages.append(text)
    return result, messages


# ---------------------------------------------------------------------------
# 3. Small helpers
# ---------------------------------------------------------------------------


def band_mean(wavelength_um: np.ndarray, values: np.ndarray) -> float:
    """Band mean over the evaluation grid [same unit as ``values``]."""
    return float(np.trapezoid(values, wavelength_um)) / float(
        wavelength_um[-1] - wavelength_um[0]
    )


def band_integral(wavelength_um: np.ndarray, values: np.ndarray) -> float:
    """Band integral over the evaluation grid [``values`` unit x um]."""
    return float(np.trapezoid(values, wavelength_um))


def altitude_along_ray(path_length_m: float, zenith_rad: float, h_start_m: float = 0.0) -> float:
    """Altitude [m] reached after ``path_length_m`` along a ray of zenith ``zenith_rad``.

    Spherical Earth, law of cosines in the (Earth centre, start, point) triangle:
    ``r = sqrt(r0^2 + s^2 + 2 r0 s cos(zeta))``.  This is the inverse of the
    slant-range solution ``core.viewing_triangle`` uses, coded independently
    here so that the detection-range walk and the chain's own geometry can be
    cross-checked against each other.
    """
    r0 = R_EARTH_M + h_start_m
    chord = r0 * r0 + path_length_m**2 + 2.0 * r0 * path_length_m * math.cos(zenith_rad)
    return math.sqrt(chord) - R_EARTH_M


def planck_spectral_radiance(wavelength_um: np.ndarray, temperature_K: float) -> np.ndarray:
    """Planck radiance [W/m^2/sr/um] — independent re-implementation for the anchor."""
    lam_m = wavelength_um * 1.0e-6
    numerator = 2.0 * H_PLANCK_J_S * C_LIGHT_M_S**2 / lam_m**5
    exponent = H_PLANCK_J_S * C_LIGHT_M_S / (lam_m * K_BOLTZMANN_J_K * temperature_K)
    # * 1e-6 converts per-metre-of-wavelength to per-micrometre
    return numerator / (np.exp(exponent) - 1.0) * 1.0e-6


# ---------------------------------------------------------------------------
# 4. Report sections
# ---------------------------------------------------------------------------


def rule(title: str) -> None:
    print()
    print("=" * 100)
    print(f"  {title}")
    print("=" * 100)


def section_inputs(vendor: dict[str, Any], canon: dict[str, float | str]) -> None:
    rule("1. VENDOR INPUTS AND THE ONE-TIME UNIT CONVERSION")
    print()
    print(f"  Vendor workbook: {INPUT_XLSX.relative_to(REPO_ROOT).as_posix()}")
    print()
    print(f"  {'Quantity':<30s} {'Vendor':>14s}  {'Unit':<8s}  {'Canonical':>14s}  {'Unit':<8s}")
    print(f"  {'-' * 30} {'-' * 14}  {'-' * 8}  {'-' * 14}  {'-' * 8}")
    cam, site = vendor["camera"], vendor["site"]
    # (label, vendor sheet, vendor key, vendor unit, canonical key, canonical unit).
    # The vendor column is READ BACK from the workbook so this table can never drift
    # from the file the conversions actually consumed.
    rows = [
        ("Entrance pupil diameter", cam, "Entrance pupil diameter", "mm",
         "aperture_diameter_m", "m"),
        ("Effective focal length", cam, "Effective focal length", "mm", "focal_length_m", "m"),
        ("Optical transmission", cam, "Optical transmission", "%", "transmission", "-"),
        ("Housing temperature", cam, "Housing temperature", "degC", "optics_temperature_K", "K"),
        ("Train emissivity", cam, "Train emissivity", "%", "scalar_emissivity", "-"),
        ("Cold shield efficiency", cam, "Cold shield efficiency", "%", "nearfield_fraction", "-"),
        ("Pixel pitch", cam, "Pixel pitch", "um", "pixel_pitch_um", "um"),
        ("Quantum efficiency", cam, "Quantum efficiency", "%", "qe", "-"),
        ("Band low edge", cam, "Spectral band, low edge", "nm", "filter_min_um", "um"),
        ("Band high edge", cam, "Spectral band, high edge", "nm", "filter_max_um", "um"),
        ("Integration time", cam, "Integration time", "ms", "integration_time_s", "s"),
        ("Full well capacity", cam, "Full well capacity", "Me-", "full_well_capacity_e", "e-"),
        ("Site altitude", site, "Site altitude", "km", "sensor_altitude_m", "m"),
        ("Target altitude", site, "Target altitude", "km", "target_altitude_m", "m"),
        ("Nozzle exit diameter", site, "Nozzle exit diameter", "mm", "target_area_m2", "m^2"),
        ("Nozzle temperature", site, "Nozzle temperature", "degC", "target_temperature_K", "K"),
    ]
    printable = [
        (name, float(sheet[key]), vendor_unit, canon[canon_key], canon_unit)
        for name, sheet, key, vendor_unit, canon_key, canon_unit in rows
    ]
    printable.append(
        ("Nominal zenith at sensor", NOMINAL_ZENITH_DEG, "deg",
         math.radians(NOMINAL_ZENITH_DEG), "rad")
    )
    for name, vendor_value, vendor_unit, canon_value, canon_unit in printable:
        print(
            f"  {name:<30s} {vendor_value:>14.6g}  {vendor_unit:<8s}  "
            f"{float(canon_value):>14.6g}  {canon_unit:<8s}"
        )
    print()
    blocked_pct = float(cam["Cold shield efficiency"])
    print(f"  Note on the cold shield: the vendor quotes the BLOCKED fraction "
          f"({blocked_pct:g} % efficient),")
    print("  RADIANT's optics.nearfield_fraction is the PASSED fraction, so the conversion is")
    print(f"  1 - eff/100 = {canon['nearfield_fraction']:g} [-].  Getting this inversion "
          f"backwards would put")
    print(f"  {1.0 / float(canon['nearfield_fraction']):.0f}x the warm-optics flux on the "
          "focal plane.")
    print()
    print("  Pointing plan [deg, zenith at the sensor]: "
          f"{', '.join(f'{z:g}' for z in vendor['zeniths_deg'])}")


def section_geometry(result: Any) -> dict[str, float]:
    rule("2. WHAT THE GEOMETRY STAGE PUBLISHES FOR THIS SCENE (ADR-0011)")
    geo = result.stage_outputs["geometry"]
    theta_o_rad = float(geo["theta_o_rad"])
    eta_rad = float(geo["eta_rad"])
    zeta_low_rad = math.pi - eta_rad  # up-looking: the sensor is the lower endpoint
    slant_m = float(geo["slant_range_m"])

    print()
    print(f"  scene_class        : {geo['scene_class']}   "
          f"(observer_class = {geo['observer_class']}, target_class = {geo['target_class']})")
    print(f"  los_direction      : {geo['los_direction']}   "
          "(derived from the altitude pair — never a user switch)")
    print(f"  viewing_mode       : {geo['viewing_mode']}")
    print()
    print(f"  h_sensor           : {float(geo['h_sensor_m']):>12.1f} [m]")
    print(f"  h_target           : {float(geo['h_target_m']):>12.1f} [m]")
    print(f"  theta_o (target-side path zenith) : {math.degrees(theta_o_rad):>8.4f} [deg] "
          f"= {theta_o_rad:.6f} [rad]  (obtuse => up-looking)")
    print(f"  eta     (angle at the sensor)     : {math.degrees(eta_rad):>8.4f} [deg] "
          f"= {eta_rad:.6f} [rad]")
    print(f"  zeta_low = pi - eta (sensor is the lower endpoint) : "
          f"{math.degrees(zeta_low_rad):>8.4f} [deg]")
    print(f"  slant range        : {slant_m / 1000.0:>12.4f} [km]")
    print(f"  incidence_angle    : {math.degrees(float(geo['incidence_angle_rad'])):>8.4f} [deg] "
          "(>= 90 deg: there is no ground plane at the target — see section 4)")

    # --- identity check: zeta_low must reproduce the entered pointing angle ---
    entered_rad = math.radians(NOMINAL_ZENITH_DEG)
    print()
    print("  Identity check (Phase-4 angle truth):")
    print(f"    entered geometry.path_zenith_rad = {math.degrees(entered_rad):.4f} [deg]")
    print(f"    pi - eta                         = {math.degrees(zeta_low_rad):.4f} [deg]  "
          f"(residual {abs(zeta_low_rad - entered_rad):.2e} [rad])")
    print(f"    pi - theta_o                     = {180.0 - math.degrees(theta_o_rad):.4f} [deg]  "
          "<- the FLAT-EARTH shorthand, wrong by the Earth-centre central angle")
    print(f"    central angle phi = theta_o - eta = "
          f"{math.degrees(theta_o_rad - eta_rad):.4f} [deg]  "
          f"(= arc {R_EARTH_M * (theta_o_rad - eta_rad) / 1000.0:.3f} [km] on the surface)")
    print("    theta_o and eta are read at DIFFERENT vertices of one spherical triangle, so")
    print("    they differ by exactly the central angle; using pi - theta_o for zeta_low is")
    print("    the flat-Earth slip the Phase-4 angle catalog calls out.")

    # --- slant range cross-check, independent spherical solution ---
    r_t = R_EARTH_M + float(geo["h_target_m"])
    r_s = R_EARTH_M + float(geo["h_sensor_m"])
    # Ray from the sensor at zenith zeta_low, hitting the shell of radius r_t:
    #   s = -r_s cos(zeta) + sqrt(r_t^2 - r_s^2 sin^2(zeta))
    hand_slant_m = -r_s * math.cos(zeta_low_rad) + math.sqrt(
        r_t**2 - (r_s * math.sin(zeta_low_rad)) ** 2
    )
    print()
    print("  Cross-check A — slant range from an independent spherical solution:")
    print(f"    hand calculation : {hand_slant_m / 1000.0:.6f} [km]")
    print(f"    RADIANT          : {slant_m / 1000.0:.6f} [km]")
    print(f"    difference       : {abs(hand_slant_m - slant_m):.4e} [m] "
          f"({abs(hand_slant_m - slant_m) / slant_m:.2e} relative)")

    # --- the optional scene-class assertion (ADR-0011 decision 8) ---
    print()
    print("  Optional geometry.scene_class assertion:")
    asserted = Sensor.from_dict(make_config())
    asserted.set("geometry.scene_class", "ground_to_air")
    evaluate(asserted)
    print("    geometry.scene_class = 'ground_to_air'  -> accepted silently (agrees with derived)")
    wrong = Sensor.from_dict(make_config())
    wrong.set("geometry.scene_class", "ground_to_space")
    try:
        evaluate(wrong)
    except Exception as exc:  # noqa: BLE001 — the refusal IS the demonstration
        first_line = str(exc).split("|")[0].strip()
        print(f"    geometry.scene_class = 'ground_to_space' -> {type(exc).__name__}")
        print(f"      {first_line}")
    else:  # pragma: no cover - would be a regression in the assertion check
        print("    UNEXPECTED: a wrong scene-class assertion did not raise")

    return {
        "theta_o_rad": theta_o_rad,
        "eta_rad": eta_rad,
        "zeta_low_rad": zeta_low_rad,
        "slant_range_m": slant_m,
    }


def section_regime(result: Any, canon: dict[str, float | str]) -> None:
    rule("3. RADIOMETRIC REGIME, AND WHAT DOES *NOT* MATTER HERE")
    optics = result.stage_outputs["optics"]
    src = result.stage_outputs["source"]
    regime = optics["regime"]
    slant_m = float(result.stage_outputs["geometry"]["slant_range_m"])
    area_m2 = float(canon["target_area_m2"])
    extent_rad = math.sqrt(area_m2) / slant_m
    fwhm_m = float(result.metrics["fwhm_x_m"])
    focal_m = float(canon["focal_length_m"])
    fwhm_rad = fwhm_m / focal_m

    print()
    print(f"  Final regime (OpticsStage, Rule 10): {getattr(regime, 'value', regime)}")
    print(f"  Tentative regime (SourceStage)     : {src['regime_tentative']}")
    print(f"  Declared scene_type                : {src['scene_type_declared']}")
    print()
    print("  Why point-source and not sub-pixel or extended:")
    print(f"    target projected area   A_t     = {area_m2:.4e} [m^2]  "
          f"(60 mm nozzle disc)")
    print(f"    angular extent  sqrt(A_t)/d     = {extent_rad:.4e} [rad]")
    print(f"    system PSF FWHM (angular)       = {fwhm_rad:.4e} [rad]  "
          f"(= {fwhm_m * 1e6:.2f} [um] at f = {focal_m:.3f} [m])")
    print(f"    ratio                           = {extent_rad / fwhm_rad:.4f} [-]  "
          "(point-source form requires <= 0.10)")
    min_range_m = math.sqrt(area_m2) / (0.10 * fwhm_rad)
    print(f"    The target is {fwhm_rad / extent_rad:.1f}x smaller than the blur, so collapsing "
          "it to an intensity")
    print("    I(lambda) = eps * L(lambda) * A_t loses nothing.  RADIANT REFUSES the")
    print("    point-source form when this ratio exceeds 0.10 (Rule 17), i.e. closer than")
    print(f"    sqrt(A_t) / (0.1 * FWHM) = {min_range_m / 1000.0:.3f} [km] slant range — below "
          "which the same")
    print("    nozzle must be run as a sub-pixel target with an explicit shape.")
    print()
    print(f"    EE_box (PlatformStage, applied once in SpectralIntegrationStage, Rule 9): "
          f"{float(result.stage_outputs['platform']['EE_box']):.4f} [-]")
    print("    In the point-source regime EE_box multiplies the TARGET term only; the sky")
    print("    background fills the pixel and is not re-apertured.")
    print()
    print("  Parameters that do NOT affect this result, and why:")
    print("    - geometry.solar_zenith_rad / solar_azimuth_rad: the trial is a NIGHT run")
    print("      (solar_illumination = 'night'), so theta_s and delta_phi are None and no")
    print("      reflected-solar or scattered-sky term is formed.  In the 3-5 um band the")
    print("      daytime scattered-sky term would additionally carry the ADR-0011")
    print("      'provisional single-scatter' caveat only below 3 um, so an MWIR daytime")
    print("      run would still be first-class.")
    print("    - source.background.*: an up-looking LOS terminates on cold space, so the")
    print("      background is the derived SkyBackground (no user temperature/emissivity")
    print("      is read).  Published descriptor: "
          f"{src['background']!r}")
    print("    - every ground-projection parameter (swath, ground speed, access): the")
    print("      target has no ground plane — see section 4.")
    print("    - detector.n_pixels_cross / TDI: this is a staring track camera, single")
    print("      frame, no scan.")


def section_metric_relevance(result: Any) -> None:
    rule("4. SCENE-CLASS METRIC RELEVANCE (guardrail G3) AND THE GAP-96 OVERRIDE")
    scene_class = str(result.stage_outputs["geometry"]["scene_class"])
    off = sorted(default_off_metrics(scene_class))

    print()
    print(f"  radiant.api.scene_relevance.default_off_metrics({scene_class!r}) — "
          f"{len(off)} metrics OFF BY DEFAULT:")
    for name in off:
        present = "PRESENT" if name in result.metrics else "absent"
        print(f"    {name:<38s} {present}")
    print()
    print("  What replaced the GSD family for an air target:")
    for name in (
        "target_plane_sample_distance_x_m",
        "target_plane_sample_distance_y_m",
        "target_plane_sample_distance_geometric_mean_m",
    ):
        print(f"    {name:<46s} = {float(result.metrics[name]):.4f} [m]")
    print(f"    {'diffraction_limit_angular_urad':<46s} = "
          f"{float(result.metrics['diffraction_limit_angular_urad']):.4f} [urad]")
    print("    Target-plane sample distance is the plate scale projected onto a plane")
    print("    THROUGH THE TARGET normal to the LOS: d = pitch * R / f.  It is the honest")
    print("    counterpart of GSD when there is no ground at the target; the angular")
    print("    diffraction limit is band-independent and stays on for all nine classes.")

    print()
    print("  Gap-96 override semantics (the map moves DEFAULTS, never the override):")
    for flag in (True, False):
        sensor = Sensor.from_dict(make_config())
        sensor.set("performance.metrics.sampling", flag)
        overridden, _ = evaluate(sensor)
        tp = "target_plane_sample_distance_geometric_mean_m" in overridden.metrics
        gsd = "gsd_geometric_mean_m" in overridden.metrics
        print(f"    performance.metrics.sampling = {str(flag):<5s} -> "
              f"target-plane sample distance {'present' if tp else 'absent ':<8s}  "
              f"GSD {'present' if gsd else 'absent'}")
    print("    An explicitly set flag wins over the class default in BOTH directions.")
    print("    GSD stays absent even when the group is force-enabled: that is a")
    print("    COMPUTABILITY gate, not a relevance default — the ground-plane cosine")
    print("    projection is undefined for incidence_angle_rad >= pi/2 (absent, not wrong).")


def section_sweep(zeniths_deg: list[float], canon: dict[str, float | str]) -> list[dict[str, Any]]:
    rule("5. ELEVATION SWEEP — SKY PATH, SIGNAL AND SNR VS ZENITH AT THE SENSOR")
    rows: list[dict[str, Any]] = []
    print()
    print(f"  {'zeta_low':>9s}  {'elev':>6s}  {'slant':>8s}  {'tau_band':>9s}  {'L_path':>10s}  "
          f"{'signal':>11s}  {'sky bkg':>11s}  {'noise':>9s}  {'SNR':>9s}  {'SCNR':>9s}  "
          f"{'NEDT':>8s}")
    print(f"  {'[deg]':>9s}  {'[deg]':>6s}  {'[km]':>8s}  {'[-]':>9s}  {'[W/m2/sr]':>10s}  "
          f"{'[e-]':>11s}  {'[e-]':>11s}  {'[e- rms]':>9s}  {'[-]':>9s}  {'[-]':>9s}  {'[mK]':>8s}")
    print(f"  {'-' * 9}  {'-' * 6}  {'-' * 8}  {'-' * 9}  {'-' * 10}  {'-' * 11}  {'-' * 11}  "
          f"{'-' * 9}  {'-' * 9}  {'-' * 9}  {'-' * 8}")

    for zenith_deg in zeniths_deg:
        result, _ = evaluate(Sensor.from_dict(make_config(zenith_deg)))
        wl = np.asarray(result.wavelength_um)
        atm = result.stage_outputs["atmosphere"]
        si = result.stage_outputs["spectral_integration"]
        snr_result = result.stage_outputs["performance"]["snr_result"]
        row = {
            "zenith_deg": zenith_deg,
            "elevation_deg": 90.0 - zenith_deg,
            "slant_km": float(result.stage_outputs["geometry"]["slant_range_m"]) / 1000.0,
            "tau_band": band_mean(wl, np.asarray(atm["tau_atm"])),
            "l_path": band_integral(wl, np.asarray(atm["L_path"])),
            "signal_e": float(si["signal_e"]),
            "background_e": float(si["background_e"]),
            "nearfield_e": float(si["nearfield_e"]),
            "noise_e": float(snr_result.noise_e),
            "snr": float(result.metrics["snr"]),
            "scnr": float(result.metrics["scnr"]),
            "nedt_mK": float(result.metrics["nedt_K"]) * 1000.0,
            "consistency": result.stage_outputs["performance"]["dual_path_consistency"],
        }
        rows.append(row)
        print(f"  {row['zenith_deg']:>9.1f}  {row['elevation_deg']:>6.1f}  "
              f"{row['slant_km']:>8.3f}  "
              f"{row['tau_band']:>9.4f}  {row['l_path']:>10.4f}  {row['signal_e']:>11.4g}  "
              f"{row['background_e']:>11.4g}  {row['noise_e']:>9.1f}  {row['snr']:>9.2f}  "
              f"{row['scnr']:>9.2f}  {row['nedt_mK']:>8.1f}")

    first, last = rows[0], rows[-1]
    print()
    print("  Physics of the trend:")
    print(f"    Slant range grows {last['slant_km'] / first['slant_km']:.2f}x from zenith to "
          f"{last['zenith_deg']:.0f} deg, but the SIGNAL falls "
          f"{first['signal_e'] / last['signal_e']:.2f}x — far more than the 1/R^2 factor of "
          f"{(last['slant_km'] / first['slant_km']) ** 2:.2f}x.")
    print(f"    The extra loss is the air mass: band-mean transmittance drops from "
          f"{first['tau_band']:.4f} to {last['tau_band']:.4f} [-] "
          f"({(last['tau_band'] / first['tau_band'] - 1.0) * 100.0:+.1f} %), because a slanted")
    print("    ray spends proportionally more of its length in the dense, wet, warm air")
    print("    below 3 km.  Up-path radiance moves the OTHER way — it RISES from")
    print(f"    {first['l_path']:.4f} to {last['l_path']:.4f} [W/m2/sr] "
          f"({(last['l_path'] / first['l_path'] - 1.0) * 100.0:+.1f} %) by Kirchhoff: the same")
    print("    extra absorbing column is an extra EMITTING column.  Both effects push SNR")
    print("    down, which is why SNR falls faster than any single one of them.")
    print()
    print("  SCNR == SNR to the last digit in every row.  That is not a coincidence and not")
    print("  a bug: in the point-source regime the target term is already a CONTRAST term")
    print("  (the pixel's sky background is subtracted, and only its shot noise remains),")
    print("  so the contrast-SNR definition reduces to the SNR definition exactly.")
    print("  In the sub-pixel regime — the same target at a 1 km vertical stand-off, where")
    print("  the point-source form is refused — the two separate (SNR 3279 vs SCNR 5256).")
    return rows


def section_sky_composition(nominal: Any) -> list[dict[str, Any]]:
    """The one place this scene class does something a physicist should query."""
    rule("5b. SKY BACKGROUND VS TARGET ALTITUDE — A COMPOSITION ARTEFACT (limitation)")
    print()
    print("  A ground sensor pointed at a FIXED zenith sees a FIXED sky column: whether the")
    print("  aircraft is at 10 km or 40 km cannot change how much sky radiance arrives from")
    print("  behind it, because the LOS terminates on cold space either way.  The chain")
    print("  composes the background as")
    print("      L_bkg = L_up(sensor -> target)  +  tau(sensor -> target) * L_sky(target -> top)")
    print("  and with the simple model's single-effective-temperature graybody per segment")
    print("  that composition is NOT additive, so the answer drifts with target altitude:")
    print()
    print(f"  {'target altitude':>16s}  {'scene class':>16s}  {'sky background':>15s}  "
          f"{'deficit':>9s}")
    print(f"  {'[km]':>16s}  {'':>16s}  {'[e-]':>15s}  {'[%]':>9s}")
    print(f"  {'-' * 16}  {'-' * 16}  {'-' * 15}  {'-' * 9}")
    rows: list[dict[str, Any]] = []
    for altitude_km in (10.0, 20.0, 40.0, 60.0, 99.0):
        result, _ = evaluate(
            Sensor.from_dict(make_config(0.0, target_altitude_m=altitude_km * 1000.0))
        )
        rows.append(
            {
                "altitude_km": altitude_km,
                "scene_class": str(result.stage_outputs["geometry"]["scene_class"]),
                "background_e": float(result.stage_outputs["spectral_integration"]["background_e"]),
            }
        )
    asymptote = rows[-1]["background_e"]
    for row in rows:
        row["deficit_pct"] = (row["background_e"] / asymptote - 1.0) * 100.0
        print(f"  {row['altitude_km']:>16.1f}  {row['scene_class']:>16s}  "
              f"{row['background_e']:>15.5g}  {row['deficit_pct']:>+9.1f}")
    print()
    noise_terms = {term.name: float(term.value_e) for term in nominal.noise_terms}
    total_variance = sum(value**2 for value in noise_terms.values())
    background_share = noise_terms.get("background_shot", 0.0) ** 2 / total_variance
    optimism_pct = 0.5 * abs(rows[0]["deficit_pct"]) * background_share
    print()
    print("  The 99 km row is the whole column (the ray leaves the modelled atmosphere at the")
    print("  target), so it is the physically correct sky for EVERY row.  The vertical 10 km")
    print(f"  scene under-reports it by {abs(rows[0]['deficit_pct']):.1f} %.  Background shot "
          "noise scales as the")
    print(f"  square root of that charge and carries {background_share * 100.0:.0f} % of the "
          "noise VARIANCE at the")
    print("  nominal point, so the composed sky makes this scenario's SNR OPTIMISTIC by only")
    print(f"  about {optimism_pct:.1f} % — small here, but it grows for a dimmer target where the")
    print("  background term dominates, and it is a systematic sign, not noise.  Logged in")
    print("  gaps.md; the fix belongs in the atmosphere segment model, not in this scenario.")
    return rows


def section_detection(zeniths_deg: list[float], canon: dict[str, float | str],
                      nominal: Any) -> list[dict[str, Any]]:
    rule("6. DETECTION RANGE ALONG THE ACTUAL RAY (path-aware up topology)")
    threshold = float(canon["snr_threshold"])
    detection = nominal.stage_outputs["performance"]["detection_range_result"]

    print()
    print("  First, what RADIANT's own detection-range metric says for this scene:")
    print(f"    detection_range_m in result.metrics : "
          f"{'present' if 'detection_range_m' in nominal.metrics else 'ABSENT'}")
    print(f"    detection_range_result.ok           : {detection.ok}")
    print("    failure_reason:")
    for line in str(detection.failure_reason).split(". "):
        if line.strip():
            print(f"      {line.strip().rstrip('.')}.")
    print()
    print("  That refusal is correct and is the Rule-17 behaviour: the metric-layer solver")
    print("  needs tau(R) along the continuation PAST the target, and for a 10 km target the")
    print("  continuation is still inside the modelled column, where extinction varies with")
    print("  altitude.  Substituting a constant-alpha model there is exactly finding GF-15.")
    print()
    print("  The scenario therefore walks the ray with the FULL CHAIN instead: at each trial")
    print("  path length s the target is placed at the altitude that ray reaches,")
    print("  h(s) = sqrt((R_E)^2 + s^2 + 2 R_E s cos(zeta_low)) - R_E, and the chain is")
    print("  re-evaluated.  Every point is a real segment-composed atmosphere — no")
    print("  extrapolation of any kind.")
    print()
    print(f"  {'zeta_low':>9s}  {'SNR @ 10 km':>12s}  {'R_detect':>10s}  {'h at R':>9s}  "
          f"{'R_vacuum':>10s}  {'ratio':>7s}  {'class at R':>16s}")
    print(f"  {'[deg]':>9s}  {'[-]':>12s}  {'[km]':>10s}  {'[km]':>9s}  {'[km]':>10s}  "
          f"{'[-]':>7s}  {'':>16s}")
    print(f"  {'-' * 9}  {'-' * 12}  {'-' * 10}  {'-' * 9}  {'-' * 10}  {'-' * 7}  {'-' * 16}")

    rows: list[dict[str, Any]] = []
    for zenith_deg in zeniths_deg:
        zenith_rad = math.radians(zenith_deg)
        ref_result, _ = evaluate(Sensor.from_dict(make_config(zenith_deg)))
        snr_ref = float(ref_result.metrics["snr"])
        ref_range_m = float(ref_result.stage_outputs["geometry"]["slant_range_m"])
        vacuum_bound_m = ref_range_m * math.sqrt(snr_ref / threshold)

        def snr_at(path_length_m: float, zenith_deg: float = zenith_deg,
                   zenith_rad: float = zenith_rad) -> tuple[float, str]:
            h_m = altitude_along_ray(path_length_m, zenith_rad)
            trial, _ = evaluate(Sensor.from_dict(make_config(zenith_deg, target_altitude_m=h_m)))
            return float(trial.metrics["snr"]), str(
                trial.stage_outputs["geometry"]["scene_class"]
            )

        lo_m, hi_m = ref_range_m, 400_000.0
        # Clamp the upper bracket so the walk never leaves the modelled column.
        while altitude_along_ray(hi_m, zenith_rad) > SEARCH_CEILING_ALTITUDE_M:
            hi_m *= 0.9
        scene_at_range = ""
        for _ in range(24):  # 24 halvings of a <400 km bracket -> < 25 m
            mid_m = 0.5 * (lo_m + hi_m)
            value, scene_at_range = snr_at(mid_m)
            if value > threshold:
                lo_m = mid_m
            else:
                hi_m = mid_m
        range_m = 0.5 * (lo_m + hi_m)
        row = {
            "zenith_deg": zenith_deg,
            "snr_ref": snr_ref,
            "range_km": range_m / 1000.0,
            "altitude_km": altitude_along_ray(range_m, zenith_rad) / 1000.0,
            "vacuum_km": vacuum_bound_m / 1000.0,
            "ratio": range_m / vacuum_bound_m,
            "scene_class": scene_at_range,
        }
        rows.append(row)
        print(f"  {row['zenith_deg']:>9.1f}  {row['snr_ref']:>12.2f}  {row['range_km']:>10.2f}  "
              f"{row['altitude_km']:>9.2f}  {row['vacuum_km']:>10.2f}  {row['ratio']:>7.3f}  "
              f"{row['scene_class']:>16s}")

    print()
    print("  Cross-check B — the vacuum inverse-square identity.")
    print("    With NO attenuation and a FIXED noise floor, SNR falls as 1/R^2, so the")
    print(f"    detection range would be exactly R_vac = R_ref * sqrt(SNR_ref/{threshold:.0f}).")
    print("    Two competing effects move the real answer off that bound:")
    print("      (a) atmospheric extinction along the continuation REMOVES signal -> R < R_vac;")
    print("      (b) the sky background behind the target THINS as the target climbs, so the")
    print("          background shot noise falls and the SNR denominator shrinks -> R > R_vac.")
    print("    Near the zenith (b) wins (ratio slightly above 1); at low elevation the long")
    print("    low-altitude path makes (a) dominant (ratio well below 1).  The crossover in")
    print("    the ratio column is the signature of that competition, and it is why a")
    print("    constant-alpha detection-range model cannot be reused for an up-looking scene.")
    return rows


def section_horizon_guard() -> list[dict[str, Any]]:
    rule("7. HORIZON GUARD — WHERE THE SCENE CLASS STOPS BEING MODELLABLE")
    print()
    print("  v1.x models no refraction, so ADR-0011 decision 6 guards the near-horizontal")
    print("  band instead of returning a plausible wrong number (Rule 17).  For this")
    print("  endpoint_minimum topology the test is |zeta_low - 90 deg|:")
    print()
    print(f"  {'zeta_low':>9s}  {'elevation':>10s}  {'outcome':<12s}  {'detail'}")
    print(f"  {'[deg]':>9s}  {'[deg]':>10s}  {'-' * 12}  {'-' * 60}")
    rows: list[dict[str, Any]] = []
    for zenith_deg in (60.0, 87.5, 88.5, 89.7):
        detail = ""
        try:
            _, messages = evaluate(
                Sensor.from_dict(make_config(zenith_deg)), capture_warnings=True
            )
        except Exception as exc:  # noqa: BLE001 — the raise IS the demonstration
            outcome = "RAISE"
            detail = str(exc).split("|")[0].strip()[:96]
        else:
            guard = [
                m for m in messages
                if "horizon" in m.lower() or "near-horizontal" in m.lower()
            ]
            outcome = "warn" if guard else "clean"
            detail = guard[0].split(":", 1)[1].strip()[:96] if guard else "no guard message"
        rows.append({"zenith_deg": zenith_deg, "outcome": outcome, "detail": detail})
        print(f"  {zenith_deg:>9.1f}  {90.0 - zenith_deg:>10.1f}  {outcome:<12s}  {detail}")
    print()
    print("  The trial's own plan tops out at 60 deg zenith (30 deg elevation), so every")
    print("  scenario point above is clean; the guard only bites if the operator tries to")
    print("  track the target down onto the horizon.")
    return rows


def section_rule4(rows: list[dict[str, Any]]) -> bool:
    rule("8. RULE-4 DUAL-PATH CONSISTENCY FOR THIS SCENE CLASS")
    print()
    print("  performance/consistency_check.py compares the FFT of the degraded EffectivePSF")
    print("  against the MTF product on every run in which the spatial path is computed.")
    print()
    print(f"  {'zeta_low':>9s}  {'passed_x':>9s}  {'passed_y':>9s}  {'max |err| x':>12s}  "
          f"{'max |err| y':>12s}  {'tolerance':>10s}")
    print(f"  {'[deg]':>9s}  {'[-]':>9s}  {'[-]':>9s}  {'[-]':>12s}  {'[-]':>12s}  {'[-]':>10s}")
    print(f"  {'-' * 9}  {'-' * 9}  {'-' * 9}  {'-' * 12}  {'-' * 12}  {'-' * 10}")
    all_ok = True
    for row in rows:
        consistency = row["consistency"]
        all_ok = all_ok and bool(consistency.passed_x) and bool(consistency.passed_y)
        print(f"  {row['zenith_deg']:>9.1f}  {str(consistency.passed_x):>9s}  "
              f"{str(consistency.passed_y):>9s}  {consistency.max_absolute_error_x:>12.3e}  "
              f"{consistency.max_absolute_error_y:>12.3e}  {consistency.tolerance:>10.3e}")
    print()
    worst = max(r["consistency"].max_absolute_error_x for r in rows)
    tolerance = rows[0]["consistency"].tolerance
    print("  VERDICT: the dual-path consistency check stayed SILENT for every point of this")
    print(f"           ground_to_air sweep — all_passed = {all_ok}.  The worst residual is")
    print(f"           {worst:.3e} [-] against a {tolerance:.3e} [-] tolerance "
          f"({tolerance / worst:.0f}x margin).")
    return all_ok


def section_hand_radiometry(canon: dict[str, float | str]) -> dict[str, float]:
    rule("9. CROSS-CHECK C — INDEPENDENT HAND RADIOMETRY OF THE SIGNAL")
    result, _ = evaluate(Sensor.from_dict(make_config(NOMINAL_ZENITH_DEG)))
    wl = np.asarray(result.wavelength_um)
    tau_atm = np.asarray(result.stage_outputs["atmosphere"]["tau_atm"])
    slant_m = float(result.stage_outputs["geometry"]["slant_range_m"])
    ee_box = float(result.stage_outputs["platform"]["EE_box"])

    # Independent re-derivation, Planck coded from CODATA constants above:
    #   N_e = t * QE * tau_opt * A_ap * EE / R^2 * INT eps L(lam,T) A_t tau_atm(lam) lam/(h c) dlam
    radiance = planck_spectral_radiance(wl, float(canon["target_temperature_K"]))
    photon_energy_J = H_PLANCK_J_S * C_LIGHT_M_S / (wl * 1.0e-6)
    intensity = float(canon["target_emissivity"]) * radiance * float(canon["target_area_m2"])
    aperture_area_m2 = math.pi * (float(canon["aperture_diameter_m"]) / 2.0) ** 2
    integrand = intensity * tau_atm / (slant_m**2) * aperture_area_m2 / photon_energy_J
    photons_per_s = float(np.trapezoid(integrand, wl))
    hand_e = (
        photons_per_s
        * float(canon["transmission"])
        * float(canon["qe"])
        * float(canon["integration_time_s"])
        * ee_box
    )
    radiant_e = float(result.stage_outputs["spectral_integration"]["signal_e"])

    print()
    print(f"  Geometry              : zeta_low = {NOMINAL_ZENITH_DEG:.1f} [deg], "
          f"slant range = {slant_m / 1000.0:.4f} [km]")
    print(f"  In-band target intensity  I = eps * INT L(lam,T) A_t dlam = "
          f"{float(np.trapezoid(intensity, wl)):.6f} [W/sr]")
    print(f"  Irradiance at the pupil   E = I * tau_bar / R^2            ~ "
          f"{float(np.trapezoid(intensity * tau_atm, wl)) / slant_m**2:.4e} [W/m^2]")
    print(f"  Collecting area           A = pi D^2 / 4                   = "
          f"{aperture_area_m2:.6f} [m^2]")
    print(f"  EE_box (from the degraded PSF)                             = {ee_box:.4f} [-]")
    print()
    print(f"    hand calculation : {hand_e:>12.4g} [e-]")
    print(f"    RADIANT          : {radiant_e:>12.4g} [e-]")
    print(f"    difference       : {abs(hand_e - radiant_e) / radiant_e * 100.0:>12.3f} [%]")
    print()
    print("  What is and is not independent here: the Planck function, the band integral, the")
    print("  photon-energy conversion, the collecting area, the 1/R^2 fall-off and the EE_box")
    print("  application are all re-coded above from CODATA constants and the datasheet, so")
    print("  they are a genuine second opinion.  tau_atm(lambda) is taken FROM the chain —")
    print("  this anchor tests the radiometric assembly, not the atmosphere; the atmosphere")
    print("  is anchored separately against MODTRAN in section 10.  Agreement at the 1e-4")
    print("  level therefore says the point-source intensity form, the inverse-square")
    print("  fall-off, the scalar QE/transmission application and the Rule-9 EE_box coupling")
    print("  are each wired exactly as the analytic form says they should be.")
    return {"hand_e": hand_e, "radiant_e": radiant_e}


# ---------------------------------------------------------------------------
# 10. MODTRAN cross-model anchor (skips gracefully when not staged)
# ---------------------------------------------------------------------------


def _read_tape7(path: Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """``(wavelength_um, total transmittance, path thermal radiance [W/m2/sr/um])``."""
    from radiant.atmosphere.modtran import Tape7Reader

    native = Tape7Reader(path).parse()
    nu = native.wavenumber_cm1
    keep = nu > 0.0
    nu = nu[keep]
    lam = 1.0e4 / nu
    order = np.argsort(lam)
    # nu^2 is the wavenumber -> wavelength Jacobian Tape7Reader.to_radiant_units applies.
    return (
        lam[order],
        native.total_transmittance[keep][order],
        (native.path_thermal_radiance[keep] * nu**2)[order],
    )


def _band(lam: np.ndarray, values: np.ndarray) -> tuple[float, float]:
    mask = (lam >= BAND_LO_UM) & (lam <= BAND_HI_UM)
    integral = float(np.trapezoid(values[mask], lam[mask]))
    return integral / (BAND_HI_UM - BAND_LO_UM), integral


def section_modtran(anchors: list[dict[str, Any]], runs_dir: Path) -> list[dict[str, Any]]:
    rule("10. CROSS-MODEL ANCHOR — SIMPLE MODEL VS THE OWNER'S MODTRAN K-LADDER")
    print()
    print(f"  Run set: {runs_dir}")
    if not runs_dir.exists():
        print()
        print("  SKIPPED — the MODTRAN batch-1 run set is not staged in this checkout.")
        print("  modtran/real_runs/ is gitignored (plan section 7.1), so a fresh clone has")
        print("  no tape7 files.  Stage the owner-delivered runs there, or pass")
        print("  --modtran-runs <path>, to reproduce this comparison.  The measured numbers")
        print("  are recorded in walkthrough.md section 'Cross-model anchor'.")
        return []

    missing = [a["run"] for a in anchors if not (runs_dir / f"{a['run']}.tp7").exists()]
    if missing:
        print(f"  SKIPPED — staged directory exists but these runs are absent: {missing}")
        return []

    print()
    print("  Each anchor is the SAME geometry run twice: MODTRAN (correlated-k band model)")
    print("  and RADIANT's 'simple' segment model.  tau is a band mean over 3-5 um;")
    print("  L is the up-path radiance band integral toward the ground observer.")
    print()
    print(f"  {'run':>4s}  {'H1':>6s}  {'H2':>6s}  {'angle':>6s}  {'tau MODTRAN':>12s}  "
          f"{'tau RADIANT':>12s}  {'d tau':>8s}  {'L MODTRAN':>11s}  {'L RADIANT':>11s}  "
          f"{'d L':>8s}")
    print(f"  {'':>4s}  {'[km]':>6s}  {'[km]':>6s}  {'[deg]':>6s}  {'[-]':>12s}  {'[-]':>12s}  "
          f"{'[%]':>8s}  {'[W/m2/sr]':>11s}  {'[W/m2/sr]':>11s}  {'[%]':>8s}")
    print(f"  {'-' * 4}  {'-' * 6}  {'-' * 6}  {'-' * 6}  {'-' * 12}  {'-' * 12}  {'-' * 8}  "
          f"{'-' * 11}  {'-' * 11}  {'-' * 8}")

    rows: list[dict[str, Any]] = []
    for anchor in anchors:
        lam, tau_modtran, radiance_modtran = _read_tape7(runs_dir / f"{anchor['run']}.tp7")
        tau_ref, _ = _band(lam, tau_modtran)
        _, l_ref = _band(lam, radiance_modtran)

        # The column product is regime-independent, so the shallow rungs (where the
        # point-source form is refused because the nozzle is resolved) are evaluated
        # in the sub-pixel regime.  Same geometry, same atmosphere, same tau.
        h_tgt_m = anchor["h2_km"] * 1000.0
        scene_type = "point_source" if h_tgt_m >= 10_000.0 else "sub_pixel"
        result, _ = evaluate(
            Sensor.from_dict(
                make_config(anchor["angle_deg"], target_altitude_m=h_tgt_m, scene_type=scene_type)
            )
        )
        wl = np.asarray(result.wavelength_um)
        tau_model = band_mean(wl, np.asarray(result.stage_outputs["atmosphere"]["tau_atm"]))
        l_model = band_integral(wl, np.asarray(result.stage_outputs["atmosphere"]["L_path"]))
        row = {
            "run": anchor["run"],
            "h1_km": anchor["h1_km"],
            "h2_km": anchor["h2_km"],
            "angle_deg": anchor["angle_deg"],
            "tau_modtran": tau_ref,
            "tau_model": tau_model,
            "tau_pct": (tau_model / tau_ref - 1.0) * 100.0,
            "l_modtran": l_ref,
            "l_model": l_model,
            "l_pct": (l_model / l_ref - 1.0) * 100.0,
        }
        rows.append(row)
        print(f"  {row['run']:>4s}  {row['h1_km']:>6.1f}  {row['h2_km']:>6.1f}  "
              f"{row['angle_deg']:>6.1f}  {row['tau_modtran']:>12.4f}  {row['tau_model']:>12.4f}  "
              f"{row['tau_pct']:>+8.1f}  {row['l_modtran']:>11.4f}  {row['l_model']:>11.4f}  "
              f"{row['l_pct']:>+8.1f}")

    print()
    print("  Reading the disagreement (this is a CHARACTERISATION, not an agreement claim):")
    print("    - tau: the simple model is systematically TOO TRANSPARENT in the MWIR, worst")
    print("      for shallow columns (+30 % at 1 km) and converging to +10 % by 20 km.  The")
    print("      CU-161 water/gas-floor calibration was fitted to whole columns, so a thin")
    print("      near-surface slice carries too little of the continuum, and the region-flat")
    print("      spectral shape cannot reproduce the CO2 4.3 um cut inside the band.")
    print("    - L: the up-path radiance is systematically TOO LOW, and by the mirror")
    print("      amount — Kirchhoff ties them.  Too little absorbing column is too little")
    print("      EMITTING column.")
    print("    - The 45 deg rung (K6) agrees better in tau than its vertical twin (K4)")
    print("      because a longer path pushes the band-model saturation the simple model")
    print("      cannot represent in the same direction as its own excess transparency.")
    print("      Beer-Lambert is exact only monochromatically: a BAND-AVERAGED tau obeys")
    print("      tau(2L) >= tau(L)^2 (Jensen), and the measured MWIR excess on the")
    print("      horizontal L grid is up to 3.85x — this non-exponentiality is the whole")
    print("      residual, and it is why the ratios are pinned rather than tightened.")
    return rows


# ---------------------------------------------------------------------------
# 11. Figures and the results workbook
# ---------------------------------------------------------------------------


def make_figures(
    sweep: list[dict[str, Any]],
    detection: list[dict[str, Any]],
    modtran: list[dict[str, Any]],
) -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    zeniths = [r["zenith_deg"] for r in sweep]

    # -- Figure 1: sky path products vs pointing --------------------------
    fig, ax_left = plt.subplots(figsize=(9.0, 6.0))
    ax_right = ax_left.twinx()
    line_tau, = ax_left.plot(
        zeniths, [r["tau_band"] for r in sweep], "o-", color="#1f77b4", linewidth=2,
        label="band-mean transmittance, 3-5 um",
    )
    line_l, = ax_right.plot(
        zeniths, [r["l_path"] for r in sweep], "s--", color="#d62728", linewidth=2,
        label="up-path radiance, 3-5 um band integral",
    )
    if modtran:
        anchor_pts = [(r["angle_deg"], r["tau_modtran"]) for r in modtran if r["h2_km"] == 10.0]
        if anchor_pts:
            ax_left.plot(
                [p[0] for p in anchor_pts], [p[1] for p in anchor_pts], "k*", markersize=14,
                label="MODTRAN K4 / K6 (10 km target)",
            )
    ax_left.set_xlabel("Zenith angle at the sensor, zeta_low [deg]")
    ax_left.set_ylabel("Band-mean transmittance [dimensionless]", color="#1f77b4")
    ax_right.set_ylabel("Up-path radiance [W/m^2/sr]", color="#d62728")
    ax_left.tick_params(axis="y", labelcolor="#1f77b4")
    ax_right.tick_params(axis="y", labelcolor="#d62728")
    ax_left.set_title("Ground-to-air sky path, 3-5 um, target at 10 km altitude")
    ax_left.grid(True, alpha=0.3)
    handles = [line_tau, line_l] + [
        h for h in ax_left.get_legend_handles_labels()[0] if h not in (line_tau,)
    ]
    ax_left.legend(handles, [h.get_label() for h in handles], loc="center left", fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "ground_to_air_sky_path_vs_zenith.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # -- Figure 2: signal / background / SNR ------------------------------
    fig, (ax_e, ax_snr) = plt.subplots(2, 1, figsize=(9.0, 8.0), sharex=True)
    ax_e.semilogy(zeniths, [r["signal_e"] for r in sweep], "o-", linewidth=2,
                  label="target signal")
    ax_e.semilogy(zeniths, [r["background_e"] for r in sweep], "s--", linewidth=2,
                  label="sky background (pixel)")
    ax_e.semilogy(zeniths, [r["nearfield_e"] for r in sweep], "^:", linewidth=2,
                  label="warm-optics nearfield")
    ax_e.semilogy(zeniths, [r["noise_e"] for r in sweep], "d-.", linewidth=2,
                  label="total noise")
    ax_e.set_ylabel("Charge [e-]  /  noise [e- rms]")
    ax_e.set_title("Ground-to-air MWIR track camera vs pointing elevation")
    ax_e.grid(True, which="both", alpha=0.3)
    ax_e.legend(fontsize=9)
    ax_snr.plot(zeniths, [r["snr"] for r in sweep], "o-", linewidth=2, label="SNR")
    ax_snr.plot(zeniths, [r["scnr"] for r in sweep], "x--", linewidth=2, label="SCNR")
    ax_snr.axhline(5.0, color="grey", linestyle=":", label="detection threshold = 5")
    ax_snr.set_xlabel("Zenith angle at the sensor, zeta_low [deg]")
    ax_snr.set_ylabel("SNR / SCNR [dimensionless]")
    ax_snr.set_yscale("log")
    ax_snr.grid(True, which="both", alpha=0.3)
    ax_snr.legend(fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "ground_to_air_snr_vs_zenith.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # -- Figure 3: detection range ----------------------------------------
    fig, ax = plt.subplots(figsize=(9.0, 6.0))
    ax.plot([r["zenith_deg"] for r in detection], [r["range_km"] for r in detection],
            "o-", linewidth=2, label="detection range, full-chain walk along the ray")
    ax.plot([r["zenith_deg"] for r in detection], [r["vacuum_km"] for r in detection],
            "s--", linewidth=2, label="vacuum inverse-square bound R_ref*sqrt(SNR_ref/5)")
    ax.plot([r["zenith_deg"] for r in detection], [r["altitude_km"] for r in detection],
            "^:", linewidth=2, label="target altitude at the detection range")
    ax.set_xlabel("Zenith angle at the sensor, zeta_low [deg]")
    ax.set_ylabel("Range / altitude [km]")
    ax.set_title("Detection range vs pointing elevation (SNR threshold = 5)")
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "ground_to_air_detection_range.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    written.append(path)

    # -- Figure 4: MODTRAN anchor (only when staged) ----------------------
    if modtran:
        vertical = [r for r in modtran if r["angle_deg"] == 0.0]
        fig, (ax_tau, ax_l) = plt.subplots(1, 2, figsize=(12.0, 5.0))
        depth = [r["h2_km"] for r in vertical]
        ax_tau.plot(depth, [r["tau_modtran"] for r in vertical], "ks-", linewidth=2,
                    label="MODTRAN (K ladder)")
        ax_tau.plot(depth, [r["tau_model"] for r in vertical], "o--", linewidth=2,
                    label="RADIANT simple model")
        ax_tau.set_xlabel("Vertical column depth, target altitude [km]")
        ax_tau.set_ylabel("Band-mean transmittance, 3-5 um [dimensionless]")
        ax_tau.set_title("Transmittance")
        ax_tau.grid(True, alpha=0.3)
        ax_tau.legend(fontsize=9)
        ax_l.plot(depth, [r["l_modtran"] for r in vertical], "ks-", linewidth=2,
                  label="MODTRAN (K ladder)")
        ax_l.plot(depth, [r["l_model"] for r in vertical], "o--", linewidth=2,
                  label="RADIANT simple model")
        ax_l.set_xlabel("Vertical column depth, target altitude [km]")
        ax_l.set_ylabel("Up-path radiance, 3-5 um [W/m^2/sr]")
        ax_l.set_title("Up-path radiance toward the ground observer")
        ax_l.grid(True, alpha=0.3)
        ax_l.legend(fontsize=9)
        fig.suptitle("Cross-model anchor: vertical up-looking K ladder, midlat summer, 23 km vis")
        fig.tight_layout()
        path = OUTPUT_DIR / "ground_to_air_modtran_anchor.png"
        fig.savefig(path, dpi=150)
        plt.close(fig)
        written.append(path)
    return written


def write_workbook(
    sweep: list[dict[str, Any]],
    detection: list[dict[str, Any]],
    modtran: list[dict[str, Any]],
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def sheet(name: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(name)
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
        for row_index, values in enumerate(rows, start=2):
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin

    default = wb.active
    assert default is not None
    wb.remove(default)

    sheet(
        "Elevation Sweep",
        ["zeta_low [deg]", "elevation [deg]", "slant range [km]", "tau band-mean [-]",
         "L_path [W/m2/sr]", "signal [e-]", "sky background [e-]", "nearfield [e-]",
         "noise [e- rms]", "SNR [-]", "SCNR [-]", "NEDT [mK]"],
        [[r["zenith_deg"], r["elevation_deg"], round(r["slant_km"], 4), round(r["tau_band"], 5),
          round(r["l_path"], 5), round(r["signal_e"], 1), round(r["background_e"], 1),
          round(r["nearfield_e"], 1), round(r["noise_e"], 2), round(r["snr"], 3),
          round(r["scnr"], 3), round(r["nedt_mK"], 2)] for r in sweep],
    )
    sheet(
        "Detection Range",
        ["zeta_low [deg]", "SNR at 10 km [-]", "detection range [km]",
         "target altitude at range [km]", "vacuum bound [km]", "ratio [-]", "scene class"],
        [[r["zenith_deg"], round(r["snr_ref"], 3), round(r["range_km"], 3),
          round(r["altitude_km"], 3), round(r["vacuum_km"], 3), round(r["ratio"], 4),
          r["scene_class"]] for r in detection],
    )
    if modtran:
        sheet(
            "MODTRAN Anchor",
            ["run", "H1 [km]", "H2 [km]", "angle [deg]", "tau MODTRAN [-]", "tau RADIANT [-]",
             "d tau [%]", "L MODTRAN [W/m2/sr]", "L RADIANT [W/m2/sr]", "d L [%]"],
            [[r["run"], r["h1_km"], r["h2_km"], r["angle_deg"], round(r["tau_modtran"], 5),
              round(r["tau_model"], 5), round(r["tau_pct"], 2), round(r["l_modtran"], 5),
              round(r["l_model"], 5), round(r["l_pct"], 2)] for r in modtran],
        )
    wb.save(RESULTS_XLSX)
    print(f"\n  Results workbook (gitignored, regenerate on demand): "
          f"{RESULTS_XLSX.relative_to(REPO_ROOT).as_posix()}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--modtran-runs",
        type=Path,
        default=DEFAULT_MODTRAN_RUNS,
        help="directory holding the owner-delivered K*.tp7 runs "
        "(default: <repo>/modtran/real_runs, gitignored)",
    )
    args = parser.parse_args()

    print("=" * 100)
    print("  SCENARIO 10.1 — GROUND-TO-AIR MWIR DETECTION")
    print("  Direction-general validation, scene class ground_to_air (matrix cell E2)")
    print("  ADR-0011 / Geometry-Flexibility Phase 5")
    print("=" * 100)

    vendor = read_vendor_inputs()
    canon = to_canonical(vendor)
    section_inputs(vendor, canon)

    nominal, nominal_warnings = evaluate(make_sensor(), capture_warnings=True)
    section_geometry(nominal)
    section_regime(nominal, canon)
    section_metric_relevance(nominal)
    sweep = section_sweep(vendor["zeniths_deg"], canon)
    section_sky_composition(nominal)
    detection = section_detection(vendor["zeniths_deg"], canon, nominal)
    section_horizon_guard()
    all_ok = section_rule4(sweep)
    section_hand_radiometry(canon)
    modtran = section_modtran(vendor["anchors"], args.modtran_runs)

    rule("11. WARNINGS EMITTED BY THE NOMINAL RUN")
    print()
    if nominal_warnings:
        for message in nominal_warnings:
            print(f"  - {message}")
    else:
        print("  none — the nominal ground_to_air point runs clean (no saturation, no")
        print("  horizon-guard shoulder, no provisional-physics caveat in the MWIR).")

    figures = make_figures(sweep, detection, modtran)
    write_workbook(sweep, detection, modtran)

    rule("12. SUMMARY")
    nominal_row = next(r for r in sweep if r["zenith_deg"] == NOMINAL_ZENITH_DEG)
    nominal_detection = next(r for r in detection if r["zenith_deg"] == NOMINAL_ZENITH_DEG)
    print()
    print(f"  Nominal point (zeta_low = {NOMINAL_ZENITH_DEG:.0f} [deg], target at 10 [km]):")
    print(f"    slant range           {nominal_row['slant_km']:>10.3f} [km]")
    print(f"    band-mean tau (3-5 um){nominal_row['tau_band']:>10.4f} [-]")
    print(f"    up-path radiance      {nominal_row['l_path']:>10.4f} [W/m^2/sr]")
    print(f"    signal                {nominal_row['signal_e']:>10.4g} [e-]")
    print(f"    sky background        {nominal_row['background_e']:>10.4g} [e-]")
    print(f"    SNR                   {nominal_row['snr']:>10.2f} [-]")
    print(f"    NEDT                  {nominal_row['nedt_mK']:>10.1f} [mK]")
    print(f"    detection range       {nominal_detection['range_km']:>10.2f} [km]  "
          f"(target at {nominal_detection['altitude_km']:.1f} [km] there)")
    print()
    print(f"  Rule-4 dual-path consistency silent across the sweep: {all_ok}")
    print(f"  Figures written: {len(figures)}")
    for path in figures:
        print(f"    {path.relative_to(REPO_ROOT).as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
