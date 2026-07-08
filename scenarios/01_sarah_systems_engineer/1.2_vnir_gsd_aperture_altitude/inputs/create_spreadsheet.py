#!/usr/bin/env python3
"""Create Sarah's inputs for Scenario 1.2: VNIR GSD vs aperture vs altitude.

Emits:

- ``silicon_ccd_qe.csv`` — front-illuminated silicon CCD QE curve (the
  datasheet plot, DIGITIZED to a 2-column CSV; a JPEG cannot be read
  directly, so digitization is a manual pre-step, noted here). Read by
  radiant.io.qe_csv.load_qe_csv.
- ``sarah_vnir_design.xlsx`` — the fixed design parameters, the
  sun-synchronous orbit (LTAN + target latitude + 4 season days), and
  the aperture/altitude sweep ranges.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# Silicon CCD QE (digitized datasheet plot), wavelength_nm / QE_pct
# ---------------------------------------------------------------------------

# Front-illuminated Si CCD: rises from UV, peaks ~600-700 nm at ~90%,
# falls into the NIR toward the 1.1 µm band edge.
SI_QE_NM_PCT = [
    (400, 28.0), (450, 52.0), (500, 72.0), (550, 84.0), (600, 90.0),
    (650, 91.0), (700, 88.0), (750, 82.0), (800, 74.0), (850, 63.0),
    (900, 48.0), (950, 33.0), (1000, 18.0),
]

with open(HERE / "silicon_ccd_qe.csv", "w", encoding="utf-8") as fh:
    fh.write("# Front-illuminated silicon CCD QE (digitized from datasheet plot)\n")
    fh.write("# Digitization is a manual pre-step — the vendor plot is a JPEG\n")
    fh.write("wavelength_nm,QE_pct\n")
    for nm, pct in SI_QE_NM_PCT:
        fh.write(f"{nm},{pct}\n")
print("Wrote silicon_ccd_qe.csv")

# ---------------------------------------------------------------------------
# Design + orbit + sweep workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Design"
ws["A1"] = "Scenario 1.2: VNIR pan-sharpened imager — design + orbit"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [34, 16, 12, 44]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Target GSD", 0.5, "m", "Panchromatic requirement"),
    ("Pixel pitch", 6.5, "µm", "Fixed; focal length derived per altitude"),
    ("Pan band min", 450.0, "nm", ""),
    ("Pan band max", 700.0, "nm", ""),
    ("Optical transmission", 85.0, "%", ""),
    ("Dark current", 50.0, "e⁻/s", "Cooled Si CCD, 280 K"),
    ("Operating temperature", 280.0, "K", ""),
    ("Read noise", 20.0, "e⁻ RMS", ""),
    ("Full well capacity", 30000.0, "e⁻", ""),
    ("System gain", 8.0, "e⁻/DN", ""),
    ("ADC resolution", 12, "bits", ""),
    ("Integration time", 0.5, "ms", "TDI / line rate set elsewhere"),
    ("Target reflectance", 0.30, "—", "Mixed urban/vegetation scene"),
    ("SPEC: min SNR", 50.0, "—", "Pan acceptance threshold"),
    ("Orbit", None, None, None),
    ("LTAN", 10.5, "hr", "10:30 AM local time of ascending node"),
    ("Target latitude", 35.0, "deg", "Mid-latitude collection site"),
    ("Aperture min", 20.0, "cm", "Sweep lower bound"),
    ("Aperture max", 80.0, "cm", "Sweep upper bound"),
    ("Altitude min", 400.0, "km", "Sweep lower bound"),
    ("Altitude max", 600.0, "km", "Sweep upper bound"),
]
r = 4
section_fill = PatternFill("solid", fgColor="D9E2F3")
for name, value, unit, note in ROWS:
    if value is None and unit is None:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True, size=11)
        for cc in range(1, 5):
            ws.cell(row=r, column=cc).fill = section_fill
    else:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
    r += 1

# Seasons sheet: 4 day-of-year values
ws2 = wb.create_sheet("Seasons")
for col, text in zip("AB", ["Season", "day_of_year"]):
    ws2[f"{col}1"] = text
    ws2[f"{col}1"].font = hdr_font
    ws2[f"{col}1"].fill = hdr_fill
for i, (season, doy) in enumerate(
    [("Spring equinox", 80), ("Summer solstice", 172),
     ("Autumn equinox", 266), ("Winter solstice", 355)], start=2
):
    ws2.cell(row=i, column=1, value=season)
    ws2.cell(row=i, column=2, value=doy)
ws2.column_dimensions["A"].width = 18

out = HERE / "sarah_vnir_design.xlsx"
wb.save(out)
print(f"Wrote {out}")
