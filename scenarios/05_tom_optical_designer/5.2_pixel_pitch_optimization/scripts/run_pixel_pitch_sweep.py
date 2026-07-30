"""Scenario 5.2: Pixel Pitch / Q-Parameter Trade Study

Tom (optical designer) wants to evaluate how pixel pitch affects image
quality and sensitivity for his f/4 MWIR pushbroom imager. He has six
candidate detectors with pixel pitches from 8 to 30 µm.

The key parameter is Q = λ·f/# / p, which governs sampling:
  Q > 1: oversampled (smooth PSF, lower MTF at Nyquist)
  Q = 1: critically sampled (Airy disk spans ~2 pixels)
  Q < 1: undersampled (aliasing, higher apparent MTF at Nyquist)

The trade-off: smaller pixels give better spatial resolution (GSD)
and higher Q (better sampling), but collect fewer photons per pixel
(lower SNR). Larger pixels give better sensitivity but coarser
resolution and risk undersampling.

This script:
  1. Reads Tom's spreadsheet (Zemax/vendor units: mm, nm, %, arc-sec)
  2. Converts to RADIANT canonical units
  3. For each candidate pixel pitch (8, 12, 15, 18, 24, 30 µm):
     - Computes Q = λ·f/# / p
     - Computes GSD = p × h / f
     - Runs RADIANT with matched detector specs from the vendor table
     - Extracts MTF at Nyquist, EE 1×1, EE 3×3, SNR
  4. Identifies the optimal pixel pitch (best trade of resolution vs. SNR)
  5. Generates plots and writes results to an output spreadsheet

Usage:
    python run_pixel_pitch_sweep.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import matplotlib

matplotlib.use("Agg")  # headless-safe: plt.show() is a no-op, so the runner completes in CI/batch
import matplotlib.pyplot as plt

# ---------------------------------------------------------------------------
# Step 4: Run RADIANT for each candidate pixel pitch
# ---------------------------------------------------------------------------

from radiant.api import Sensor


def main() -> None:
    """Run the scenario analysis."""
    # ---------------------------------------------------------------------------
    # Step 1: Read Tom's spreadsheet
    # ---------------------------------------------------------------------------
    # Three sheets:
    #   "Optical Design Summary" — telescope and scene parameters
    #   "Candidate Detectors"    — 6 pixel pitches with matched detector specs
    #   "Performance Requirements" — pass/fail thresholds

    INPUT_FILE = Path(__file__).parent.parent / "inputs" / "tom_pixel_pitch_trade.xlsx"
    OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "pixel_pitch_trade_results.xlsx"
    PLOT_DIR = Path(__file__).parent.parent / "outputs"

    wb = openpyxl.load_workbook(INPUT_FILE)

    # --- Parse Optical Design Summary ---
    ws_opt = wb["Optical Design Summary"]
    opt_specs: dict[str, object] = {}
    for row in ws_opt.iter_rows(min_row=6, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
                continue
            opt_specs[name] = value

    # --- Parse Candidate Detectors ---
    # This sheet has a header row with pitch labels, then parameter rows.
    # Layout: [Parameter, 8µm, 12µm, 15µm, 18µm, 24µm, 30µm, Unit]
    ws_det = wb["Candidate Detectors"]

    # Read the pitch values from row 4 (header), columns B-G
    pitches_um: list[float] = []
    for col in range(2, 8):
        cell_val = ws_det.cell(row=4, column=col).value
        if cell_val and "µm" in str(cell_val):
            pitches_um.append(float(str(cell_val).replace("µm", "").strip()))

    # Read parameter rows into a dict of lists
    det_table: dict[str, list] = {}
    for row in ws_det.iter_rows(min_row=5, max_col=8, values_only=True):
        param_name = row[0]
        if param_name and row[1] is not None:
            values = [row[i] for i in range(1, 7)]
            det_table[param_name] = values

    # --- Parse Performance Requirements ---
    ws_req = wb["Performance Requirements"]
    reqs: dict[str, object] = {}
    for row in ws_req.iter_rows(min_row=5, max_col=4, values_only=True):
        if row[0] and row[1] is not None:
            reqs[row[0]] = row[1]

    print("=== Optical Design Summary (Zemax output) ===")
    print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit'}")
    print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}")
    print(f"  {'Entrance pupil diameter':<35s} {opt_specs['Entrance pupil diameter']:>14}  mm")
    print(f"  {'Effective focal length':<35s} {opt_specs['Effective focal length']:>14}  mm")
    print(f"  {'f-number (working)':<35s} {opt_specs['f-number (working)']:>14}  —")
    print(f"  {'Optical transmission':<35s} {opt_specs['Optical transmission']:>14}  %")
    print(f"  {'Optics temperature':<35s} {opt_specs['Optics temperature']:>14}  K")
    print(f"  {'WFE (RMS, on-axis)':<35s} {opt_specs['WFE (RMS, on-axis)']:>14}  waves")
    print(f"  {'WFE reference wavelength':<35s} {opt_specs['WFE reference wavelength']:>14}  nm")
    print(f"  {'Airy disk diameter':<35s} {opt_specs['Airy disk diameter']:>14}  µm")
    print(f"  {'Band center wavelength':<35s} {opt_specs['Band center wavelength']:>14}  nm")
    print(f"  {'Filter cut-on':<35s} {opt_specs['Filter cut-on']:>14}  nm")
    print(f"  {'Filter cut-off':<35s} {opt_specs['Filter cut-off']:>14}  nm")
    print(f"  {'Target temperature':<35s} {opt_specs['Target temperature']:>14}  K")
    print(f"  {'Background temperature':<35s} {opt_specs['Background temperature']:>14}  K")
    print(f"  {'Orbit altitude':<35s} {opt_specs['Orbit altitude']:>14}  km")

    print(f"\n=== Candidate Detectors ===")
    print(f"  {'Parameter':<30s}", end="")
    for p in pitches_um:
        print(f"  {p:>8.0f}µm", end="")
    print()
    print(f"  {'-' * 30}", end="")
    for _ in pitches_um:
        print(f"  {'-' * 10}", end="")
    print()
    for param_name, values in det_table.items():
        print(f"  {param_name:<30s}", end="")
        for v in values:
            print(f"  {v:>10}", end="")
        print()

    # ---------------------------------------------------------------------------
    # Step 2: Convert to RADIANT canonical units
    # ---------------------------------------------------------------------------

    # Optics — Zemax uses mm
    aperture_m = float(opt_specs["Entrance pupil diameter"]) / 1000.0   # mm → m
    focal_length_m = float(opt_specs["Effective focal length"]) / 1000.0  # mm → m
    f_number = float(opt_specs["f-number (working)"])
    transmission = float(opt_specs["Optical transmission"]) / 100.0     # % → fraction
    optics_temp_K = float(opt_specs["Optics temperature"])              # already K
    wfe_waves = float(opt_specs["WFE (RMS, on-axis)"])                  # already waves

    # Spectral — nm → µm
    lambda_center_nm = float(opt_specs["Band center wavelength"])
    lambda_center_um = lambda_center_nm / 1000.0                        # nm → µm
    band_min_nm = float(opt_specs["Filter cut-on"])
    band_max_nm = float(opt_specs["Filter cut-off"])
    band_min_um = band_min_nm / 1000.0                                  # nm → µm
    band_max_um = band_max_nm / 1000.0                                  # nm → µm

    # Scene
    target_temp = float(opt_specs["Target temperature"])                # already K
    target_emiss = float(opt_specs["Target emissivity"])
    bg_temp = float(opt_specs["Background temperature"])                # already K
    bg_emiss = float(opt_specs["Background emissivity"])

    # Geometry
    altitude_km = float(opt_specs["Orbit altitude"])
    altitude_m = altitude_km * 1000.0                                   # km → m

    # Airy disk
    airy_diam_um = float(opt_specs["Airy disk diameter"])               # µm

    print("\n=== Converted to RADIANT canonical units ===")
    print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
    print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 25}")
    print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  mm ÷ 1000")
    print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  mm ÷ 1000")
    print(f"  {'f-number':<35s} {f_number:>14.1f}  {'—':<15s}  no conversion")
    print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
    print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.1f}  {'K':<15s}  no conversion")
    print(f"  {'WFE (RMS)':<35s} {wfe_waves:>14.4f}  {'waves':<15s}  no conversion")
    print(f"  {'Band center':<35s} {lambda_center_um:>14.3f}  {'µm':<15s}  nm ÷ 1000")
    print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  nm ÷ 1000")
    print(f"  {'Target temperature':<35s} {target_temp:>14.1f}  {'K':<15s}  no conversion")
    print(f"  {'Background temperature':<35s} {bg_temp:>14.1f}  {'K':<15s}  * contrast SNR only")
    print(f"  {'Orbit altitude':<35s} {altitude_m:>14.0f}  {'m':<15s}  km × 1000")
    print(f"  {'Airy disk diameter':<35s} {airy_diam_um:>14.1f}  {'µm':<15s}  no conversion")

    # ---------------------------------------------------------------------------
    # Step 3: Compute Q parameter and GSD for each pitch
    # ---------------------------------------------------------------------------

    print(f"\n=== Sampling Analysis ===")
    print(f"  Q = λ_center × f/# / p    (sampling parameter)")
    print(f"  GSD = p × h / f           (ground sample distance)")
    print(f"  Airy disk = 2.44 × λ × f/# = {airy_diam_um:.1f} µm")
    print(f"  Nyquist frequency = 1 / (2p)")
    print(f"")
    print(f"  {'Pitch [µm]':>12s}  {'Q [—]':>8s}  {'GSD [m]':>10s}  {'f_Nyq [cy/mm]':>14s}  {'Airy/pitch':>12s}  {'Sampling'}")
    print(f"  {'-' * 12}  {'-' * 8}  {'-' * 10}  {'-' * 14}  {'-' * 12}  {'-' * 15}")

    for p_um in pitches_um:
        p_m = p_um * 1e-6
        Q = (lambda_center_um * f_number) / p_um   # Q = λ·f/# / p (all in µm)
        gsd_m = p_m * altitude_m / focal_length_m   # GSD [m]
        f_nyq = 1.0 / (2.0 * p_m)                  # Nyquist [cycles/m]
        f_nyq_per_mm = f_nyq / 1000.0               # cycles/mm (optical convention)
        airy_ratio = airy_diam_um / p_um            # Airy disk / pixel pitch

        if Q > 1.5:
            regime = "oversampled"
        elif Q > 1.0:
            regime = "well-sampled"
        elif Q > 0.7:
            regime = "undersampled"
        else:
            regime = "ALIASED"

        print(f"  {p_um:>12.0f}  {Q:>8.2f}  {gsd_m:>10.1f}  {f_nyq_per_mm:>14.1f}  {airy_ratio:>12.1f}  {regime}")

    print(f"\n=== Running RADIANT for each candidate detector ===")

    # Storage for results
    results_table: list[dict] = []

    for idx, p_um in enumerate(pitches_um):
        # Get matched detector specs for this pitch from the vendor table
        qe_pct = float(det_table["Quantum efficiency (in-band)"][idx])
        dark_rate = float(det_table["Dark current"][idx])
        fwc = float(det_table["Full well capacity"][idx])
        read_noise = float(det_table["Read noise (CDS)"][idx])
        adc_bits = int(det_table["ADC resolution"][idx])
        gain = float(det_table["System gain"][idx])
        t_int_ms = float(det_table["Integration time (nom.)"][idx])

        # Convert
        qe = qe_pct / 100.0          # % → fraction
        t_int_s = t_int_ms / 1000.0   # ms → s

        config = {
            "source": {
                "target": {
                    "temperature": target_temp,
                    "emissivity": target_emiss,
                },
                "background": {
                    "temperature": bg_temp,
                    "emissivity": bg_emiss,
                },
            },
            "atmosphere": {
                "model": "simple",
            },
            "geometry": {
                "sensor_altitude_m": altitude_m,
            },
            "optics": {
                "aperture_diameter_m": aperture_m,
                "focal_length_m": focal_length_m,
                "transmission_scalar": transmission,
                "optics_temperature_K": optics_temp_K,
            },
            "detector": {
                "pixel_pitch_x_um": p_um,
                "pixel_pitch_y_um": p_um,
                "qe_value": qe,
                "dark_rate_e_per_s": dark_rate,
                "detector_temperature_K": 77.0,
            },
            "spectral_integration": {
                "filter_min_um": band_min_um,
                "filter_max_um": band_max_um,
                "integration_time_s": t_int_s,
            },
            "readout": {
                "read_noise_e_rms": read_noise,
                "gain_e_per_dn": gain,
                "adc_bits": adc_bits,
                "full_well_capacity_e": fwc,
            },
            # CU-178: config outside the GIQE-5 envelope → NIIRS N/A by default; opt into
            # the extrapolated NIIRS trend (read as relative, not calibrated).
            "performance": {"niirs": {"allow_extrapolated": True}},
        }

        sensor = Sensor.from_dict(config)
        result = sensor.evaluate()

        # Extract native metrics (Q, GSD computed by RADIANT's PerformanceStage)
        p_m = p_um * 1e-6
        f_nyq_per_mm = 1.0 / (2.0 * p_m) / 1000.0

        snr = result.metrics["snr"]
        contrast_snr = result.metrics.get("contrast_snr")
        mtf_nyq = result.metrics["mtf_at_nyquist"]
        ee_1x1 = result.metrics["ee_1x1"]
        ee_3x3 = result.metrics["ee_3x3"]
        signal_e = result.stage_outputs["readout"]["signal_e_final"]

        # Native Q and GSD from RADIANT (no manual computation needed)
        Q = result.metrics.get("q_center", (lambda_center_um * f_number) / p_um)
        gsd_m = result.metrics.get("gsd_cross_track_m", p_m * altitude_m / focal_length_m)
        rer = result.metrics.get("rer")
        fwhm_x_m = result.metrics.get("fwhm_x_m")
        nedt_K = result.metrics.get("nedt_K")
        niirs = result.metrics.get("niirs")
        strehl = result.metrics.get("strehl")

        # Noise terms
        noise_dict = {nt.name: nt.value_e for nt in result.noise_terms}
        total_noise = math.sqrt(sum(v**2 for v in noise_dict.values()))

        row_data = {
            "pitch_um": p_um,
            "Q": Q,
            "gsd_m": gsd_m,
            "f_nyq_per_mm": f_nyq_per_mm,
            "snr": snr,
            "contrast_snr": contrast_snr,
            "mtf_nyq": mtf_nyq,
            "ee_1x1": ee_1x1,
            "ee_3x3": ee_3x3,
            "rer": rer,
            "fwhm_x_m": fwhm_x_m,
            "nedt_K": nedt_K,
            "niirs": niirs,
            "strehl": strehl,
            "signal_e": signal_e,
            "total_noise": total_noise,
            "signal_shot": noise_dict.get("signal_shot", 0.0),
            "bkg_shot": noise_dict.get("background_shot", 0.0),
            "dark_shot": noise_dict.get("dark_shot", 0.0),
            "read_noise": noise_dict.get("read_noise", 0.0),
            "nf_shot": noise_dict.get("nearfield_shot", 0.0),
            "qe_pct": qe_pct,
            "dark_rate": dark_rate,
            "fwc": fwc,
            "read_noise_spec": read_noise,
        }
        results_table.append(row_data)

        niirs_str = f"NIIRS={niirs:.1f}" if niirs else "NIIRS=N/A"
        print(f"  {p_um:>5.0f} µm  Q={Q:.2f}  GSD={gsd_m:.1f} m  "
              f"SNR={snr:.1f}  MTF@Nyq={mtf_nyq:.4f}  EE1={ee_1x1:.4f}  "
              f"{niirs_str}  Signal={signal_e:.0f} e⁻")

    # ---------------------------------------------------------------------------
    # Step 5: Regime and physics notes
    # ---------------------------------------------------------------------------

    print(f"\n=== Radiometric Regime ===")
    print(f"  All configurations use EXTENDED regime (310 K ground fills the pixel).")
    print(f"  Background temperature (295 K) is used only for contrast SNR.")
    print(f"")
    print(f"  PIXEL PITCH TRADE-OFF PHYSICS:")
    print(f"    Pixel pitch affects the system through three competing mechanisms:")
    print(f"")
    print(f"    1. SPATIAL RESOLUTION — smaller pixels give finer GSD:")
    print(f"       GSD = p × h / f = p × {altitude_m:.0f} / {focal_length_m:.3f}")
    print(f"       At 8 µm:  GSD = {8e-6 * altitude_m / focal_length_m:.1f} m")
    print(f"       At 30 µm: GSD = {30e-6 * altitude_m / focal_length_m:.1f} m")
    print(f"")
    print(f"    2. SAMPLING (Q PARAMETER) — determines aliasing regime:")
    print(f"       Q = λ·f/# / p  at λ = {lambda_center_um:.2f} µm, f/# = {f_number:.0f}")
    print(f"       The Airy disk diameter is {airy_diam_um:.1f} µm.")
    print(f"       When p < Airy/2 (Q > ~1.2), the PSF is well-sampled.")
    print(f"       When p > Airy (Q < ~0.6), severe aliasing occurs.")
    print(f"")
    print(f"    3. SENSITIVITY (SNR) — larger pixels collect more photons:")
    print(f"       Signal ∝ Ω_pixel = (p/f)² → signal ∝ p²")
    print(f"       Signal shot noise ∝ √signal ∝ p")
    print(f"       SNR ∝ signal/noise. In photon-noise-limited regime, SNR ∝ p.")
    print(f"       Smaller pixels sacrifice SNR for resolution.")

    # ---------------------------------------------------------------------------
    # Step 6: Requirements compliance
    # ---------------------------------------------------------------------------

    MTF_REQ = 0.10
    SNR_REQ = 100.0
    EE_1x1_REQ = 0.30

    print(f"\n=== Requirements Compliance ===")
    print(f"  {'Pitch':>8s}  {'Q':>6s}  {'GSD [m]':>8s}  {'MTF@Nyq':>8s}  {'EE 1×1':>8s}  {'SNR':>8s}  {'GSD<10m':>8s}  {'Verdict'}")
    print(f"  {'-' * 8}  {'-' * 6}  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 10}")

    for r in results_table:
        gsd_ok = r["gsd_m"] < 10.0
        mtf_ok = r["mtf_nyq"] >= MTF_REQ
        ee_ok = r["ee_1x1"] >= EE_1x1_REQ
        snr_ok = r["snr"] >= SNR_REQ
        all_ok = gsd_ok and mtf_ok and ee_ok and snr_ok

        gsd_str = "PASS" if gsd_ok else "FAIL"
        verdict = "ALL PASS" if all_ok else "FAIL"

        # Identify which requirement(s) failed
        failures = []
        if not gsd_ok:
            failures.append("GSD")
        if not mtf_ok:
            failures.append("MTF")
        if not ee_ok:
            failures.append("EE")
        if not snr_ok:
            failures.append("SNR")
        if failures:
            verdict = f"FAIL ({', '.join(failures)})"

        print(f"  {r['pitch_um']:>8.0f}  {r['Q']:>6.2f}  {r['gsd_m']:>8.1f}  {r['mtf_nyq']:>8.4f}  "
              f"{r['ee_1x1']:>8.4f}  {r['snr']:>8.1f}  {gsd_str:>8s}  {verdict}")

    # ---------------------------------------------------------------------------
    # Step 7: Identify the optimal pixel pitch
    # ---------------------------------------------------------------------------
    # The "best" pitch is the one that passes all requirements and offers
    # the best balance of GSD and SNR. We use a simple figure of merit:
    #   FoM = SNR / GSD  (higher is better — want high SNR AND low GSD)
    # Only candidates that pass all requirements are considered.

    print(f"\n=== Optimal Pixel Pitch Selection ===")

    compliant = []
    for r in results_table:
        if (r["gsd_m"] < 10.0 and r["mtf_nyq"] >= MTF_REQ
                and r["ee_1x1"] >= EE_1x1_REQ and r["snr"] >= SNR_REQ):
            fom = r["snr"] / r["gsd_m"]
            compliant.append((r, fom))

    if compliant:
        print(f"  Compliant candidates (pass all requirements):")
        print(f"  {'Pitch [µm]':>12s}  {'Q':>6s}  {'GSD [m]':>8s}  {'SNR':>8s}  {'NIIRS':>6s}  {'NEDT [K]':>9s}  {'FoM (SNR/GSD)':>14s}")
        print(f"  {'-' * 12}  {'-' * 6}  {'-' * 8}  {'-' * 8}  {'-' * 6}  {'-' * 9}  {'-' * 14}")
        for r, fom in compliant:
            niirs_s = f"{r['niirs']:.1f}" if r.get("niirs") else "N/A"
            nedt_s = f"{r['nedt_K']:.4f}" if r.get("nedt_K") else "N/A"
            print(f"  {r['pitch_um']:>12.0f}  {r['Q']:>6.2f}  {r['gsd_m']:>8.1f}  {r['snr']:>8.1f}  {niirs_s:>6s}  {nedt_s:>9s}  {fom:>14.1f}")

        best_r, best_fom = max(compliant, key=lambda x: x[1])
        print(f"\n  RECOMMENDATION: {best_r['pitch_um']:.0f} µm pixel pitch")
        print(f"    Q = {best_r['Q']:.2f} (sampling parameter)")
        print(f"    GSD = {best_r['gsd_m']:.1f} m")
        print(f"    SNR = {best_r['snr']:.1f}")
        print(f"    MTF at Nyquist = {best_r['mtf_nyq']:.4f}")
        if best_r.get("rer"):
            print(f"    RER = {best_r['rer']:.4f}")
        if best_r.get("niirs"):
            print(f"    NIIRS = {best_r['niirs']:.1f}")
        if best_r.get("nedt_K"):
            print(f"    NEDT = {best_r['nedt_K']:.4f} K")
        if best_r.get("strehl"):
            print(f"    Strehl = {best_r['strehl']:.4f}")
        print(f"    Figure of merit (SNR/GSD) = {best_fom:.1f}")
    else:
        print(f"  WARNING: No candidate meets all requirements simultaneously!")
        print(f"  Consider relaxing GSD, MTF, EE, or SNR thresholds.")

    # ---------------------------------------------------------------------------
    # Step 8: Generate plots
    # ---------------------------------------------------------------------------

    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    pitches = np.array([r["pitch_um"] for r in results_table])
    Qs = np.array([r["Q"] for r in results_table])
    gsds = np.array([r["gsd_m"] for r in results_table])
    snrs = np.array([r["snr"] for r in results_table])
    mtfs = np.array([r["mtf_nyq"] for r in results_table])
    ee1s = np.array([r["ee_1x1"] for r in results_table])
    ee3s = np.array([r["ee_3x3"] for r in results_table])
    signals = np.array([r["signal_e"] for r in results_table])

    # ---- Plot 1: Q and GSD vs. pixel pitch -----------------------------------
    fig1, ax1a = plt.subplots(figsize=(9, 6))

    color_q = "#4472C4"
    color_gsd = "#ED7D31"

    ax1a.plot(pitches, Qs, "o-", color=color_q, linewidth=2, markersize=8, label="Q parameter")
    ax1a.set_xlabel("Pixel Pitch [µm]", fontsize=12)
    ax1a.set_ylabel("Q = λ·f/# / p  [—]", fontsize=12, color=color_q)
    ax1a.tick_params(axis="y", labelcolor=color_q)

    # Q regime bands
    ax1a.axhspan(1.0, 3.0, alpha=0.08, color="green", label="Well-sampled (Q > 1)")
    ax1a.axhspan(0.7, 1.0, alpha=0.08, color="gold")
    ax1a.axhspan(0.0, 0.7, alpha=0.08, color="red", label="Aliased (Q < 0.7)")
    ax1a.axhline(1.0, color="green", linestyle=":", linewidth=0.8)
    ax1a.axhline(0.7, color="red", linestyle=":", linewidth=0.8)

    # Annotate each point with Q value
    for p, q in zip(pitches, Qs):
        ax1a.annotate(f"Q={q:.2f}", (p, q), textcoords="offset points",
                      xytext=(8, 6), fontsize=8, color=color_q)

    ax1b = ax1a.twinx()
    ax1b.plot(pitches, gsds, "s--", color=color_gsd, linewidth=2, markersize=7, label="GSD")
    ax1b.set_ylabel("Ground Sample Distance [m]", fontsize=12, color=color_gsd)
    ax1b.tick_params(axis="y", labelcolor=color_gsd)
    ax1b.axhline(10.0, color=color_gsd, linestyle=":", linewidth=1, alpha=0.5)
    ax1b.annotate("GSD < 10 m req.", (pitches[-1], 10.0),
                  textcoords="offset points", xytext=(-80, 8), fontsize=8,
                  color=color_gsd, alpha=0.7)

    lines_a, labels_a = ax1a.get_legend_handles_labels()
    lines_b, labels_b = ax1b.get_legend_handles_labels()
    ax1a.legend(lines_a + lines_b, labels_a + labels_b, loc="center right", fontsize=9)

    ax1a.set_title("Sampling Parameter Q and GSD vs. Pixel Pitch", fontsize=13, fontweight="bold")
    ax1a.set_xlim(5, 33)
    ax1a.grid(True, alpha=0.3)
    fig1.tight_layout()
    fig1.savefig(PLOT_DIR / "fig1_Q_and_GSD_vs_pitch.png", dpi=150)
    print(f"\n  Saved {PLOT_DIR / 'fig1_Q_and_GSD_vs_pitch.png'}")

    # ---- Plot 2: MTF at Nyquist and EE vs. pixel pitch -----------------------
    fig2, ax2a = plt.subplots(figsize=(9, 6))

    ax2a.plot(pitches, mtfs, "o-", color="#4472C4", linewidth=2, markersize=8,
              label="MTF at Nyquist")
    ax2a.plot(pitches, ee1s, "s-", color="#70AD47", linewidth=2, markersize=7,
              label="EE 1×1")
    ax2a.plot(pitches, ee3s, "^-", color="#A5A5A5", linewidth=1.5, markersize=6,
              label="EE 3×3")

    # Requirements
    ax2a.axhline(MTF_REQ, color="#4472C4", linestyle="--", linewidth=1, alpha=0.5)
    ax2a.annotate(f"MTF req ≥ {MTF_REQ}", (pitches[0], MTF_REQ),
                  textcoords="offset points", xytext=(5, -12), fontsize=8,
                  color="#4472C4", alpha=0.7)
    ax2a.axhline(EE_1x1_REQ, color="#70AD47", linestyle="--", linewidth=1, alpha=0.5)
    ax2a.annotate(f"EE req ≥ {EE_1x1_REQ}", (pitches[0], EE_1x1_REQ),
                  textcoords="offset points", xytext=(5, 5), fontsize=8,
                  color="#70AD47", alpha=0.7)

    # Add Q values as secondary x-axis labels
    ax2_top = ax2a.twiny()
    ax2_top.set_xlim(ax2a.get_xlim())
    ax2_top.set_xticks(pitches)
    ax2_top.set_xticklabels([f"Q={q:.2f}" for q in Qs], fontsize=8, rotation=30)

    ax2a.set_xlabel("Pixel Pitch [µm]", fontsize=12)
    ax2a.set_ylabel("MTF / EE [—]", fontsize=12)
    ax2a.set_title("Spatial Metrics vs. Pixel Pitch", fontsize=13, fontweight="bold")
    ax2a.legend(loc="best", fontsize=9)
    ax2a.grid(True, alpha=0.3)
    ax2a.set_xlim(5, 33)
    fig2.tight_layout()
    fig2.savefig(PLOT_DIR / "fig2_MTF_EE_vs_pitch.png", dpi=150)
    print(f"  Saved {PLOT_DIR / 'fig2_MTF_EE_vs_pitch.png'}")

    # ---- Plot 3: SNR and Signal vs. pixel pitch -------------------------------
    fig3, ax3a = plt.subplots(figsize=(9, 6))

    ax3a.plot(pitches, snrs, "o-", color="#4472C4", linewidth=2, markersize=8,
              label="SNR")
    ax3a.axhline(SNR_REQ, color="red", linestyle="--", linewidth=1)
    ax3a.annotate(f"SNR req ≥ {SNR_REQ:.0f}", (pitches[-1], SNR_REQ),
                  textcoords="offset points", xytext=(-80, 8), fontsize=9,
                  color="red")

    ax3a.set_xlabel("Pixel Pitch [µm]", fontsize=12)
    ax3a.set_ylabel("SNR [—]", fontsize=12, color="#4472C4")
    ax3a.tick_params(axis="y", labelcolor="#4472C4")

    ax3b = ax3a.twinx()
    ax3b.plot(pitches, signals / 1e3, "s--", color="#ED7D31", linewidth=1.5,
              markersize=6, label="Signal [ke⁻]")
    ax3b.set_ylabel("Signal [ke⁻]", fontsize=12, color="#ED7D31")
    ax3b.tick_params(axis="y", labelcolor="#ED7D31")

    lines_a, labels_a = ax3a.get_legend_handles_labels()
    lines_b, labels_b = ax3b.get_legend_handles_labels()
    ax3a.legend(lines_a + lines_b, labels_a + labels_b, loc="center left", fontsize=9)

    # Add Q values as secondary x-axis
    ax3_top = ax3a.twiny()
    ax3_top.set_xlim(ax3a.get_xlim())
    ax3_top.set_xticks(pitches)
    ax3_top.set_xticklabels([f"Q={q:.2f}" for q in Qs], fontsize=8, rotation=30)

    ax3a.set_title("SNR and Signal vs. Pixel Pitch", fontsize=13, fontweight="bold")
    ax3a.grid(True, alpha=0.3)
    ax3a.set_xlim(5, 33)
    fig3.tight_layout()
    fig3.savefig(PLOT_DIR / "fig3_SNR_signal_vs_pitch.png", dpi=150)
    print(f"  Saved {PLOT_DIR / 'fig3_SNR_signal_vs_pitch.png'}")

    # ---- Plot 4: Trade space — SNR vs. GSD with Q color-coded ----------------
    fig4, ax4 = plt.subplots(figsize=(9, 6))

    # Color-code by Q: green for well-sampled, yellow for marginal, red for aliased
    colors = []
    for q in Qs:
        if q >= 1.0:
            colors.append("#70AD47")
        elif q >= 0.7:
            colors.append("#FFC000")
        else:
            colors.append("#FF4444")

    scatter = ax4.scatter(gsds, snrs, c=Qs, cmap="RdYlGn", s=200,
                          edgecolors="black", linewidths=1, zorder=5,
                          vmin=0.4, vmax=2.2)
    cbar = plt.colorbar(scatter, ax=ax4, label="Q parameter [—]")

    # Annotate each point
    for r in results_table:
        ax4.annotate(f"{r['pitch_um']:.0f} µm",
                     (r["gsd_m"], r["snr"]),
                     textcoords="offset points", xytext=(10, -5), fontsize=9,
                     fontweight="bold")

    # Requirement box
    ax4.axvline(10.0, color="red", linestyle="--", linewidth=1, alpha=0.5)
    ax4.axhline(SNR_REQ, color="red", linestyle="--", linewidth=1, alpha=0.5)
    ax4.fill_between([0, 10], [SNR_REQ, SNR_REQ], [max(snrs) * 1.2, max(snrs) * 1.2],
                     alpha=0.05, color="green", label="Compliant region")

    ax4.set_xlabel("Ground Sample Distance [m]", fontsize=12)
    ax4.set_ylabel("SNR [—]", fontsize=12)
    ax4.set_title("Trade Space: SNR vs. GSD (color = Q parameter)",
                  fontsize=13, fontweight="bold")
    ax4.legend(loc="upper left", fontsize=9)
    ax4.grid(True, alpha=0.3)
    ax4.set_xlim(0, max(gsds) * 1.15)
    fig4.tight_layout()
    fig4.savefig(PLOT_DIR / "fig4_trade_space.png", dpi=150)
    print(f"  Saved {PLOT_DIR / 'fig4_trade_space.png'}")

    plt.show()

    # ---------------------------------------------------------------------------
    # Step 9: Write results to output spreadsheet
    # ---------------------------------------------------------------------------

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb_out = openpyxl.Workbook()

    header_font_out = Font(bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    best_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border_out = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- Sheet 1: Trade Study Results ---
    ws1 = wb_out.active
    ws1.title = "Trade Study Results"
    ws1["A1"] = "Scenario 5.2: Pixel Pitch Trade Study"
    ws1["A1"].font = Font(bold=True, size=14)

    col_headers = ["Pitch [µm]", "Q [—]", "GSD [m]", "f_Nyq [cy/mm]",
                   "MTF@Nyq [—]", "EE 1×1 [—]", "EE 3×3 [—]", "RER [—]",
                   "SNR [—]", "NEDT [K]", "NIIRS [—]", "Strehl [—]",
                   "Signal [e⁻]", "Total Noise [e⁻]"]
    col_widths = [12, 10, 10, 14, 14, 12, 12, 10, 10, 12, 10, 10, 14, 14]

    for col, (h_text, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col, value=h_text)
        cell.font = header_font_out
        cell.border = thin_border_out
        cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[cell.column_letter].width = w

    best_pitch = best_r["pitch_um"] if compliant else None

    for i, r in enumerate(results_table, 4):
        vals = [r["pitch_um"], round(r["Q"], 2), round(r["gsd_m"], 1),
                round(r["f_nyq_per_mm"], 1), round(r["mtf_nyq"], 4),
                round(r["ee_1x1"], 4), round(r["ee_3x3"], 4),
                round(r["rer"], 4) if r.get("rer") else "",
                round(r["snr"], 1),
                round(r["nedt_K"], 4) if r.get("nedt_K") else "",
                round(r["niirs"], 1) if r.get("niirs") else "",
                round(r["strehl"], 4) if r.get("strehl") else "",
                round(r["signal_e"]), round(r["total_noise"], 1)]

        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.border = thin_border_out
            if best_pitch and r["pitch_um"] == best_pitch:
                cell.fill = best_fill

        # Color-code metrics against requirements
        mtf_cell = ws1.cell(row=i, column=5)
        mtf_cell.fill = pass_fill if r["mtf_nyq"] >= MTF_REQ else fail_fill
        ee_cell = ws1.cell(row=i, column=6)
        ee_cell.fill = pass_fill if r["ee_1x1"] >= EE_1x1_REQ else fail_fill
        snr_cell = ws1.cell(row=i, column=9)
        snr_cell.fill = pass_fill if r["snr"] >= SNR_REQ else fail_fill

    # --- Sheet 2: Noise Breakdown ---
    ws2 = wb_out.create_sheet("Noise Breakdown")
    ws2["A1"] = "Noise Budget per Pixel Pitch"
    ws2["A1"].font = Font(bold=True, size=14)

    noise_headers = ["Pitch [µm]", "Q [—]", "Signal Shot [e⁻]", "Bkg Shot [e⁻]",
                     "NF Shot [e⁻]", "Dark Shot [e⁻]", "Read Noise [e⁻]",
                     "Total Noise [e⁻]", "SNR [—]"]
    noise_widths = [12, 10, 16, 14, 14, 14, 14, 16, 10]

    for col, (h_text, w) in enumerate(zip(noise_headers, noise_widths), 1):
        cell = ws2.cell(row=3, column=col, value=h_text)
        cell.font = header_font_out
        cell.border = thin_border_out
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[cell.column_letter].width = w

    for i, r in enumerate(results_table, 4):
        vals = [r["pitch_um"], round(r["Q"], 2),
                round(r["signal_shot"], 1), round(r["bkg_shot"], 1),
                round(r["nf_shot"], 1), round(r["dark_shot"], 1),
                round(r["read_noise"], 1), round(r["total_noise"], 1),
                round(r["snr"], 1)]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = thin_border_out

    # --- Sheet 3: Summary ---
    ws3 = wb_out.create_sheet("Summary")
    ws3["A1"] = "Scenario 5.2: Pixel Pitch Trade Study Summary"
    ws3["A1"].font = Font(bold=True, size=14)

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 25

    summary_items = [
        ("Optical System", ""),
        ("Aperture", f"{aperture_m * 1000:.0f} mm ({aperture_m * 100:.0f} cm)"),
        ("Focal length", f"{focal_length_m * 1000:.0f} mm"),
        ("f-number", f"{f_number}"),
        ("Spectral band", f"{band_min_um:.1f} – {band_max_um:.1f} µm"),
        ("Band center", f"{lambda_center_um:.2f} µm"),
        ("Airy disk diameter", f"{airy_diam_um:.1f} µm"),
        ("Orbit altitude", f"{altitude_km:.0f} km"),
        ("", ""),
        ("Requirements", ""),
        ("GSD", "< 10 m"),
        ("MTF at Nyquist", f"≥ {MTF_REQ}"),
        ("EE 1×1", f"≥ {EE_1x1_REQ}"),
        ("SNR", f"≥ {SNR_REQ:.0f}"),
        ("", ""),
        ("Results", ""),
    ]

    for r in results_table:
        gsd_ok = r["gsd_m"] < 10.0
        mtf_ok = r["mtf_nyq"] >= MTF_REQ
        ee_ok = r["ee_1x1"] >= EE_1x1_REQ
        snr_ok = r["snr"] >= SNR_REQ
        status = "ALL PASS" if (gsd_ok and mtf_ok and ee_ok and snr_ok) else "FAIL"
        summary_items.append((
            f"{r['pitch_um']:.0f} µm (Q={r['Q']:.2f}, GSD={r['gsd_m']:.1f} m)",
            f"SNR={r['snr']:.0f}, MTF={r['mtf_nyq']:.3f} [{status}]"
        ))

    if compliant:
        summary_items.append(("", ""))
        summary_items.append(("Recommendation", ""))
        summary_items.append(("Optimal pixel pitch",
                              f"{best_r['pitch_um']:.0f} µm (Q={best_r['Q']:.2f})"))
        summary_items.append(("Rationale",
                              f"Best SNR/GSD figure of merit among compliant options"))

    for i, (label, value) in enumerate(summary_items, 3):
        cell_a = ws3.cell(row=i, column=1, value=label)
        cell_b = ws3.cell(row=i, column=2, value=value)
        if label and not value:
            cell_a.font = Font(bold=True, size=11)

    wb_out.save(OUTPUT_FILE)
    print(f"\nResults written to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
