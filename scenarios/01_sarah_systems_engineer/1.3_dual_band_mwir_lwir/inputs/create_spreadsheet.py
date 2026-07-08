#!/usr/bin/env python3
"""Create Sarah's inputs for Scenario 1.3: dual-band MWIR/LWIR wildfire trade.

Emits:

- ``forest_conifer_aster.txt`` — conifer-canopy spectrum in the JPL/NASA
  ASTER library TEXT FORMAT (metadata header, descending-wavelength
  two-column data, reflectance in percent) — read by
  radiant.io.aster_library.load_aster_spectrum. Values follow the
  well-known TIR behaviour of green vegetation: reflectance of a few
  percent falling through the thermal infrared (ε ≈ 0.96–0.99).
- ``sarah_detector_options.xlsx`` — the vendor comparison table (NOT
  YAML): two HgCdTe options with noise specs at their operating
  temperatures, shared optics/platform block.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# ASTER-format forest spectrum (descending wavelength, percent — native style)
# ---------------------------------------------------------------------------

# (wavelength_um, reflectance_pct) — descending, 3.0–13.0 µm
CONIFER_RHO = [
    (13.0, 1.6), (12.5, 1.5), (12.0, 1.4), (11.5, 1.3), (11.0, 1.3),
    (10.5, 1.4), (10.0, 1.5), (9.5, 1.8), (9.0, 2.2), (8.5, 2.6),
    (8.0, 3.0), (7.5, 3.4), (7.0, 3.6), (6.5, 3.5), (6.0, 3.2),
    (5.5, 3.4), (5.0, 3.8), (4.5, 4.4), (4.0, 5.0), (3.5, 5.6),
    (3.0, 6.4),
]

header = """\
Name: Conifer forest canopy (needleleaf, green)
Type: vegetation
Class: tree
Subclass: needle
Particle Size: n/a
Sample No.: veg.needle.conifer.synthetic
Owner: Program office (synthetic, ASTER-format)
Wavelength Range: TIR
Origin: Synthesized for scenario 1.3 following ASTER library conventions
Description: Directional hemispherical reflectance of a green conifer
canopy, 3.0-13.0 micrometers. Emissivity = 1 - reflectance (opaque).
Measurement: Directional Hemispherical Reflectance
First Column: X
Second Column: Y
X Units: Wavelength (micrometers)
Y Units: Reflectance (percent)
First X Value: 13.0
Last X Value: 3.0
Number of X Values: 21
Additional Information: none
"""

with open(HERE / "forest_conifer_aster.txt", "w", encoding="utf-8") as fh:
    fh.write(header)
    for wl, pct in CONIFER_RHO:
        fh.write(f"{wl:6.2f}  {pct:6.2f}\n")
print("Wrote forest_conifer_aster.txt")

# ---------------------------------------------------------------------------
# Detector options workbook (vendor comparison table, not YAML)
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Detector Options"

ws["A1"] = "Scenario 1.3: MWIR vs LWIR HgCdTe options — vendor comparison"
ws["A1"].font = Font(bold=True, size=14)
for col, width in zip("ABCD", [34, 18, 18, 12]):
    ws.column_dimensions[col].width = width

hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
for col, text in zip("ABCD", ["Parameter", "MWIR option", "LWIR option", "Unit"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Band minimum", 3.5, 8.0, "µm"),
    ("Band maximum", 5.0, 12.0, "µm"),
    ("Cutoff wavelength", 5.0, 12.5, "µm"),
    ("Quantum efficiency (band avg)", 78.0, 68.0, "%"),
    ("Dark current", 8.0e4, 2.5e6, "e⁻/s"),
    ("Operating temperature", 80.0, 60.0, "K"),
    ("Read noise (CDS)", 25.0, 45.0, "e⁻ RMS"),
    ("Full well capacity", 4.0e6, 1.2e7, "e⁻"),
    ("Pixel pitch", 20.0, 20.0, "µm"),
    ("System gain", 260.0, 780.0, "e⁻/DN"),
    ("ADC resolution", 14, 14, "bits"),
    ("Integration time (fire mode)", 0.005, 0.025, "ms"),
]
r = 4
for name, mwir, lwir, unit in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=mwir)
    ws.cell(row=r, column=3, value=lwir)
    ws.cell(row=r, column=4, value=unit)
    r += 1

ws2 = wb.create_sheet("Shared Platform")
for col, width in zip("ABC", [34, 18, 12]):
    ws2.column_dimensions[col].width = width
for col, text in zip("ABC", ["Parameter", "Value", "Unit"]):
    ws2[f"{col}1"] = text
    ws2[f"{col}1"].font = hdr_font
    ws2[f"{col}1"].fill = hdr_fill

SHARED = [
    ("Aperture diameter", 2.5, "cm"),
    ("Focal length", 5.0, "cm"),
    ("Optical transmission", 78.0, "%"),
    ("Optics temperature", 5.0, "°C"),
    ("Platform altitude", 10.0, "km"),
    ("Hotspot area", 5.0, "m²"),
    ("Hotspot temperature (nominal)", 600.0, "K"),
    ("Hotspot emissivity", 0.85, "—"),
    ("Forest background temperature", 300.0, "K"),
    ("Scene clutter sigma", 0.03, "—"),
]
r = 2
for name, value, unit in SHARED:
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=value)
    ws2.cell(row=r, column=3, value=unit)
    r += 1

out = HERE / "sarah_detector_options.xlsx"
wb.save(out)
print(f"Wrote {out}")
