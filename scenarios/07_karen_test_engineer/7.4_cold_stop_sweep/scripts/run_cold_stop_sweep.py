"""Scenario 7.4: Cold Stop Efficiency Sweep — How Misaligned Is My Cold Shield?

Karen (test engineer) is running a TVAC background characterization test.
With the calibration blackbody shuttered (cold plate blocking the aperture),
she measures the background signal level. She finds 44,000 e⁻ at one cold
stop position, but the design predicts ~35,000 e⁻ with a perfect cold stop.

She suspects the cold shield is misaligned, allowing warm optics thermal
emission to reach the FPA through gaps in the cold stop baffle.

This script:
  1. Reads Karen's spreadsheet (vendor/lab units: cm, mm, %, °C, nm, ms, fA)
  2. Converts to RADIANT canonical units (m, fractions, K, µm, s, e⁻/s),
     deriving the lumped-train emissivity ε = 1 − τ (Kirchhoff, Gap 37)
  3. Runs RADIANT at nearfield_fraction = 1.0 (no cold stop) as the
     maximum-leakage reference
  4. Sweeps nearfield_fraction from 0.00 to 1.00
  5. For each leakage value, extracts:
     - Nearfield signal (warm optics self-emission) [e⁻]
     - Background signal (scene contribution) [e⁻]
     - Total background = nearfield + scene background [e⁻]
     - Noise breakdown: nearfield shot, background shot, dark, read [e⁻ RMS]
     - SNR
  6. Inverts each lab measurement to a nearfield_fraction with
     Sensor.solve_for (Brent root-finding on the forward model, Gap 10)
  7. Compares model predictions against Karen's measurements
  8. Writes results to an output spreadsheet

Parameter-name note: RADIANT's former `optics.cold_stop_efficiency` was
renamed `optics.nearfield_fraction` (Gap 12) — the value is the fraction of
the FPA hemisphere filled by warm-emitting elements (0 = perfect cold stop,
1 = no cold stop), i.e. 1 − vendor "cold stop efficiency".

Usage:
    python run_cold_stop_sweep.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ---------------------------------------------------------------------------
# Step 1: Read Karen's spreadsheet
# ---------------------------------------------------------------------------
# Karen has three sheets:
#   "Instrument Spec Sheet"    — vendor specs in lab/vendor units
#   "Background Measurements"  — lab data at various cold stop positions
#   "Performance Requirements" — pass/fail thresholds

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "karen_cold_stop_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "cold_stop_sweep_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

# --- Parse Instrument Spec Sheet ---
# Layout: rows with [Parameter, Value, Unit, Notes], with blue section headers

ws_spec = wb["Instrument Spec Sheet"]
specs: dict[str, object] = {}
for row in ws_spec.iter_rows(min_row=6, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        # Skip section headers (blue-filled rows)
        fill = row[0].fill
        if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
            continue
        specs[name] = value

# --- Parse Background Measurements sheet ---
# Layout: [Test Point, Cold Stop Position [mm], Mean Bkg Signal [DN],
#          Bkg Noise [DN], Mean Bkg Signal [e⁻], Notes]

ws_meas = wb["Background Measurements"]
lab_data: list[dict] = []
for row in ws_meas.iter_rows(min_row=6, max_col=6, values_only=True):
    if row[0] and row[1] is not None and isinstance(row[1], (int, float)):
        lab_data.append({
            "test_id": row[0],
            "position_mm": float(row[1]),
            "dn_mean": float(row[2]),
            "dn_sigma": float(row[3]),
            "bkg_e": float(row[4]),
            "notes": row[5] or "",
        })

# --- Parse Performance Requirements sheet ---

ws_req = wb["Performance Requirements"]
reqs: dict[str, object] = {}
for row in ws_req.iter_rows(min_row=5, max_col=4, values_only=True):
    if row[0] and row[1] is not None:
        reqs[row[0]] = row[1]

print("=== Instrument Spec Sheet (vendor units) ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}")
print(f"  {'Primary aperture diameter':<35s} {specs['Primary aperture diameter']:>14}  cm")
print(f"  {'Effective focal length':<35s} {specs['Effective focal length']:>14}  mm")
print(f"  {'f-number':<35s} {specs['f-number']:>14}  —")
print(f"  {'End-to-end optical transmission':<35s} {specs['End-to-end optical transmission']:>14}  %")
print(f"  {'Optics barrel temperature':<35s} {specs['Optics barrel temperature']:>14}  °C")
print(f"  {'Cold stop design efficiency':<35s} {specs['Cold stop design efficiency']:>14}  %")
print(f"  {'Pixel pitch':<35s} {specs['Pixel pitch']:>14}  µm")
print(f"  {'Average QE (in-band)':<35s} {specs['Average QE (in-band)']:>14}  %")
print(f"  {'Dark current (mean)':<35s} {specs['Dark current (mean)']:>14}  fA/pixel")
print(f"  {'Operating temperature':<35s} {specs['Operating temperature']:>14}  K")
print(f"  {'Full well capacity':<35s} {specs['Full well capacity']:>14.0f}  e⁻")
print(f"  {'Read noise (CDS)':<35s} {specs['Read noise (CDS)']:>14}  e⁻ RMS")
print(f"  {'ADC resolution':<35s} {specs['ADC resolution']:>14}  bits")
print(f"  {'System gain':<35s} {specs['System gain']:>14}  e⁻/DN")

print(f"\n=== Lab Background Measurements ===")
print(f"  {'Test Point':<14s} {'Pos [mm]':>10s}  {'Bkg DN':>10s}  {'σ DN':>8s}  {'Bkg [e⁻]':>12s}")
print(f"  {'-' * 14} {'-' * 10}  {'-' * 10}  {'-' * 8}  {'-' * 12}")
for d in lab_data:
    print(f"  {d['test_id']:<14s} {d['position_mm']:>10.1f}  {d['dn_mean']:>10.0f}  {d['dn_sigma']:>8.0f}  {d['bkg_e']:>12.0f}")

print(f"\n=== Performance Requirements ===")
for req_name, req_val in reqs.items():
    print(f"  {req_name:<45s}  {req_val}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------
# RADIANT expects: meters, fractions, seconds, µm, K, e⁻/s
# Karen's spreadsheet has: cm, mm, %, °C, nm, ms, fA/pixel

# Optics
aperture_m = float(specs["Primary aperture diameter"]) / 100.0    # cm → m
focal_length_m = float(specs["Effective focal length"]) / 1000.0  # mm → m
f_number = float(specs["f-number"])                                # dimensionless
transmission = float(specs["End-to-end optical transmission"]) / 100.0  # % → fraction
optics_temp_K = float(specs["Optics barrel temperature"]) + 273.15     # °C → K
cold_stop_design = float(specs["Cold stop design efficiency"]) / 100.0  # % → fraction
# RADIANT convention (Gap 12): nearfield_fraction is the LEAKAGE fraction,
# i.e. 1 − vendor "cold stop efficiency" (where 100% efficient = full blocking).
nearfield_design = 1.0 - cold_stop_design                          # vendor % → leakage fraction
# Kirchhoff-derived lumped-train emissivity (Gap 37): the warm train is
# reflective, so the non-transmitted fraction is absorbed → ε = 1 − τ.
# (optics.scalar_emissivity requires ε + τ ≤ 1; equality holds here.)
optics_emissivity = 1.0 - transmission                             # derived, not an input
wfe_waves = float(specs["WFE (RMS)"])                              # already in waves

# Detector
pixel_pitch_um = float(specs["Pixel pitch"])                       # already µm
pixel_pitch_m = pixel_pitch_um * 1e-6                              # µm → m
qe = float(specs["Average QE (in-band)"]) / 100.0                 # % → fraction
# Dark current: fA/pixel → e⁻/s
# I_dark [fA] = Q_dark [e⁻/s] × q_e [C]
# Q_dark [e⁻/s] = I_dark [fA] × 1e-15 / 1.602e-19
dark_fA = float(specs["Dark current (mean)"])
dark_rate_e_per_s = dark_fA * 1e-15 / 1.602e-19                   # fA → e⁻/s
operating_temp_K = float(specs["Operating temperature"])           # already K
fwc = float(specs["Full well capacity"])                           # already e⁻
read_noise = float(specs["Read noise (CDS)"])                      # already e⁻ RMS
adc_bits = int(specs["ADC resolution"])                            # already bits
gain = float(specs["System gain"])                                 # already e⁻/DN

# Calibration source — Karen's blackbody in TVAC
bb_temp_K = float(specs["Blackbody temperature"]) + 273.15        # °C → K
bb_emiss = float(specs["Blackbody emissivity"])                    # dimensionless
shroud_temp_K = float(specs["Chamber shroud temperature"]) + 273.15  # °C → K
shroud_emiss = float(specs["Chamber shroud emissivity"])           # dimensionless

# Spectral — vendor gives band edges in nm
band_str = str(specs["Cold filter passband"])
band_parts = band_str.replace("–", "-").replace("—", "-").split("-")
band_min_nm = float(band_parts[0].strip())
band_max_nm = float(band_parts[1].strip())
band_min_um = band_min_nm / 1000.0                                # nm → µm
band_max_um = band_max_nm / 1000.0                                # nm → µm

# Integration time
t_int_s = float(specs["Integration time"]) / 1000.0               # ms → s

# Geometry — lab test, blackbody at 1.5 m
source_dist_m = float(specs["Source distance"])                    # already m

print("\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 25}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  mm ÷ 1000")
print(f"  {'f-number':<35s} {f_number:>14.1f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Cold stop design eff.':<35s} {cold_stop_design:>14.2f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Nearfield fraction (design)':<35s} {nearfield_design:>14.2f}  {'fraction':<15s}  1 − vendor efficiency")
print(f"  {'Optics emissivity (derived)':<35s} {optics_emissivity:>14.4f}  {'fraction':<15s}  ε = 1 − τ (Kirchhoff)")
print(f"  {'WFE (RMS)':<35s} {wfe_waves:>14.4f}  {'waves':<15s}  no conversion")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'Quantum efficiency':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<35s} {dark_fA:>14.1f}  {'fA/pixel':<15s}  (input)")
print(f"  {'Dark current (converted)':<35s} {dark_rate_e_per_s:>14.1f}  {'e⁻/s':<15s}  fA × 1e-15 ÷ q_e")
print(f"  {'Detector temperature':<35s} {operating_temp_K:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Read noise':<35s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'System gain':<35s} {gain:>14.1f}  {'e⁻/DN':<15s}  no conversion")
print(f"  {'Blackbody temperature':<35s} {bb_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Blackbody emissivity':<35s} {bb_emiss:>14.2f}  {'—':<15s}  dimensionless")
print(f"  {'Shroud temperature':<35s} {shroud_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Shroud emissivity':<35s} {shroud_emiss:>14.2f}  {'—':<15s}  dimensionless")
print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  nm ÷ 1000")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms ÷ 1000")

# ---------------------------------------------------------------------------
# Step 3: Run RADIANT baseline (nearfield_fraction = 1.0, no cold stop)
# ---------------------------------------------------------------------------
# This is a lab / TVAC test:
#   - Atmosphere model: "exo" (vacuum chamber, no atmospheric path)
#   - The blackbody fills the FOV → extended regime
#   - Nearfield emission is enabled (optics at 293 K radiate thermally,
#     ε = 1 − τ derived above)
#   - nearfield_fraction controls how much of that nearfield reaches the FPA

from radiant.api import Sensor
from radiant.api.solve import SolveBracketError

config = {
    "source": {
        "target": {
            "temperature": bb_temp_K,
            "emissivity": bb_emiss,
        },
        "background": {
            "temperature": shroud_temp_K,
            "emissivity": shroud_emiss,
        },
    },
    "atmosphere": {
        "model": "exo",   # Vacuum — TVAC chamber, no atmosphere
    },
    "geometry": {
        "sensor_altitude_m": 0.0,   # Lab test — not orbital
    },
    "platform": {
        # Stage-7 stop-gap: the "exo" backend routes through the
        # no_atmosphere 'space' sub-case, whose Earth-limb intercept check
        # requires a positive user-set platform.h_sensor [m above MSL].
        # 1.0 m ≈ optical-bench height; the value feeds only the
        # limb-clearance check (no radiometric effect in this lab setup).
        "h_sensor": 1.0,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "scalar_emissivity": optics_emissivity,    # ε = 1 − τ (Kirchhoff, Gap 37)
        "optics_temperature_K": optics_temp_K,
        "nearfield_fraction": 1.0,   # Baseline: no cold stop (max leakage)
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate_e_per_s,
        "detector_temperature_K": operating_temp_K,
    },
    "spectral_integration": {
        "filter_min_um": band_min_um,
        "filter_max_um": band_max_um,
        "integration_time_s": t_int_s,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
        "full_well_capacity_e": fwc,
    },
    # CU-178: config outside the GIQE-5 envelope → NIIRS N/A by default; opt into
    # the extrapolated NIIRS trend (read as relative, not calibrated). Propagates to
    # config_shuttered via the per-dict copy below.
    "performance": {"niirs": {"allow_extrapolated": True}},
}

print("\n=== Running RADIANT baseline (nearfield_fraction = 1.0, no cold stop) ===")
sensor = Sensor.from_dict(config)
result = sensor.evaluate()

# Extract baseline values
regime = result.stage_outputs["optics"]["regime"]
baseline_signal_e = result.stage_outputs["readout"]["signal_e_final"]
baseline_nearfield_e = result.stage_outputs["spectral_integration"]["nearfield_e"]
baseline_background_e = result.stage_outputs["spectral_integration"]["background_e"]
baseline_snr = result.metrics["snr"]
baseline_contrast_snr = result.metrics.get("contrast_snr")
baseline_nedt = result.metrics.get("nedt_K")
baseline_niirs = result.metrics.get("niirs")
baseline_gsd = result.metrics.get("gsd_geometric_mean_m")
baseline_q = result.metrics.get("q_center")
baseline_strehl = result.metrics.get("strehl")
baseline_mtf_nyq = result.metrics.get("mtf_at_nyquist")
baseline_well_margin = result.metrics.get("well_margin_dB")

# Extract noise terms
noise_dict: dict[str, float] = {}
for nt in result.noise_terms:
    noise_dict[nt.name] = nt.value_e

print(f"\n=== Radiometric Regime ===")
print(f"  Regime:  {regime}")
print(f"")
print(f"  This is a lab/TVAC test with an extended-area blackbody at {bb_temp_K:.1f} K")
print(f"  ({specs['Blackbody temperature']}°C) filling the sensor FOV → extended regime.")
print(f"")
print(f"  UNUSED PARAMETER NOTE: in the extended regime the target fills the")
print(f"  pixel IFOV, so RADIANT skips the separate scene-background photon")
print(f"  term (background_e = 0 e⁻ by design — matrix Decision #13). The")
print(f"  chamber-shroud temperature ({shroud_temp_K:.1f} K / "
      f"{specs['Chamber shroud temperature']}°C) stays in the")
print(f"  config for descriptor completeness but contributes no photons here;")
print(f"  the only background terms are warm-optics nearfield and dark current.")
print(f"")
print(f"  COLD STOP PHYSICS:")
print(f"    The cold stop is a cryogenic baffle (at ~77 K) that surrounds the FPA")
print(f"    and blocks thermal radiation from the warm optics barrel ({optics_temp_K:.1f} K)")
print(f"    from reaching the detector. When perfectly aligned (vendor efficiency")
print(f"    = 100%), it eliminates all warm-optics self-emission seen by the FPA.")
print(f"")
print(f"    In practice, manufacturing and alignment tolerances mean some warm")
print(f"    radiation leaks through gaps. RADIANT's nearfield_fraction parameter")
print(f"    (η_nf, formerly cold_stop_efficiency — renamed under Gap 12) is the")
print(f"    fraction of the FPA hemisphere filled by warm-emitting elements:")
print(f"")
print(f"      E_nearfield ∝ η_nf × ε_optics × B(λ, T_optics)")
print(f"")
print(f"      η_nf = 0.0 → PERFECT cold stop: all warm radiation blocked")
print(f"      η_nf = 1.0 → NO cold stop: all warm radiation reaches FPA")
print(f"      η_nf = 1 − (vendor cold stop efficiency)")
print(f"")
print(f"    WARM-OPTICS EMISSIVITY (Kirchhoff, Rule 5):")
print(f"    In scalar transmission mode the train is one lumped element. Its")
print(f"    emissivity is DERIVED, not free: treating the non-transmitted power")
print(f"    as absorbed gives ε = 1 − τ = 1 − {transmission:.2f} = {optics_emissivity:.2f}")
print(f"    (optics.scalar_emissivity, Gap 37). Without it the lump defaults to")
print(f"    the refractive assumption ε = 0 and nearfield_e would be zero for")
print(f"    every η_nf — the failure recorded as Gap 4 in this scenario's gaps.md.")
print(f"")
print(f"  Reference point:")
print(f"    At nearfield_fraction = 1.0: nearfield_e = {baseline_nearfield_e:.0f} e⁻")
print(f"    This is the maximum warm-optics contribution (no cold stop).")

print(f"\n=== Baseline Metrics (nearfield_fraction = 1.0) ===")
print(f"  {'Metric':<35s} {'Value':>14s}  {'Unit'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}")
print(f"  {'Signal (target)':<35s} {baseline_signal_e:>14.0f}  e⁻")
print(f"  {'Nearfield (warm optics)':<35s} {baseline_nearfield_e:>14.0f}  e⁻")
print(f"  {'Background (scene)':<35s} {baseline_background_e:>14.0f}  e⁻")
print(f"  {'SNR':<35s} {baseline_snr:>14.2f}  — (dimensionless)")
if baseline_contrast_snr is not None:
    print(f"  {'Contrast SNR':<35s} {baseline_contrast_snr:>14.2f}  — (dimensionless)")
if baseline_nedt is not None:
    print(f"  {'NEDT':<35s} {baseline_nedt * 1e3:>14.2f}  mK")
if baseline_niirs is not None:
    print(f"  {'NIIRS':<35s} {baseline_niirs:>14.2f}  — (dimensionless)")
if baseline_gsd is not None:
    print(f"  {'GSD (geometric mean)':<35s} {baseline_gsd:>14.4f}  m")
if baseline_q is not None:
    print(f"  {'Q (sampling parameter)':<35s} {baseline_q:>14.4f}  — (dimensionless)")
if baseline_strehl is not None:
    print(f"  {'Strehl ratio':<35s} {baseline_strehl:>14.4f}  — (dimensionless)")
if baseline_mtf_nyq is not None:
    print(f"  {'MTF at Nyquist':<35s} {baseline_mtf_nyq:>14.4f}  — (dimensionless)")
if baseline_well_margin is not None:
    print(f"  {'Well margin':<35s} {baseline_well_margin:>14.2f}  dB")

# MTF budget display
mtf_budget = result.stage_outputs.get("performance", {}).get("mtf_budget")
if mtf_budget is not None:
    per_term = mtf_budget.per_term_at_nyquist
    print(f"\n  MTF Budget (at Nyquist):")
    print(f"  {'Component':<30s} {'MTF_x':>10s}  {'MTF_y':>10s}")
    print(f"  {'-' * 30} {'-' * 10}  {'-' * 10}")
    seen: set[str] = set()
    for key in per_term:
        base = key.rsplit("_", 1)[0]
        if base in seen:
            continue
        seen.add(base)
        val_x = per_term.get(f"{base}_x", 1.0)
        val_y = per_term.get(f"{base}_y", 1.0)
        label = base.replace("mtf_", "").replace("_", " ").title()
        print(f"  {label:<30s} {val_x:>10.4f}  {val_y:>10.4f}")
    print(f"  {'System':<30s} {mtf_budget.system_mtf_at_nyquist_x:>10.4f}  {mtf_budget.system_mtf_at_nyquist_y:>10.4f}")

print(f"\n  Noise breakdown (at nearfield_fraction = 1.0):")
print(f"  {'Noise Term':<35s} {'Value [e⁻ RMS]':>14s}")
print(f"  {'-' * 35} {'-' * 14}")
for name in ["signal_shot", "background_shot", "nearfield_shot", "dark_shot",
             "read_noise", "quantization"]:
    if name in noise_dict:
        print(f"  {name:<35s} {noise_dict[name]:>14.2f}")

# ---------------------------------------------------------------------------
# Step 4: Sweep nearfield_fraction from 0.0 to 1.0
# ---------------------------------------------------------------------------
# We sweep the full range to understand the relationship, then focus on
# the region that matches Karen's measurements.
#
# For each leakage value, we re-run RADIANT and extract the nearfield
# signal and noise terms. SNR changes because nearfield shot noise changes.

eta_sweep = np.linspace(0.0, 1.0, 51)   # nearfield_fraction η_nf [fraction]

sweep_nearfield_e = np.zeros_like(eta_sweep)
sweep_background_e = np.zeros_like(eta_sweep)
sweep_total_bkg_e = np.zeros_like(eta_sweep)
sweep_snr = np.zeros_like(eta_sweep)
sweep_nf_shot = np.zeros_like(eta_sweep)
sweep_bkg_shot = np.zeros_like(eta_sweep)
sweep_dark_shot = np.zeros_like(eta_sweep)
sweep_read_noise = np.zeros_like(eta_sweep)
sweep_signal_e = np.zeros_like(eta_sweep)
sweep_nedt = np.full_like(eta_sweep, np.nan)
sweep_niirs = np.full_like(eta_sweep, np.nan)

print(f"\n=== Sweeping nearfield_fraction from 0.00 to 1.00 ===")

for i, eta in enumerate(eta_sweep):
    config["optics"]["nearfield_fraction"] = float(eta)
    sensor_i = Sensor.from_dict(config)
    result_i = sensor_i.evaluate()

    sweep_nearfield_e[i] = result_i.stage_outputs["spectral_integration"]["nearfield_e"]
    sweep_background_e[i] = result_i.stage_outputs["spectral_integration"]["background_e"]
    sweep_total_bkg_e[i] = sweep_nearfield_e[i] + sweep_background_e[i]
    sweep_snr[i] = result_i.metrics["snr"]
    sweep_signal_e[i] = result_i.stage_outputs["readout"]["signal_e_final"]
    nedt_val = result_i.metrics.get("nedt_K")
    niirs_val = result_i.metrics.get("niirs")
    if nedt_val is not None:
        sweep_nedt[i] = nedt_val
    if niirs_val is not None:
        sweep_niirs[i] = niirs_val

    nd = {}
    for nt in result_i.noise_terms:
        nd[nt.name] = nt.value_e
    sweep_nf_shot[i] = nd.get("nearfield_shot", 0.0)
    sweep_bkg_shot[i] = nd.get("background_shot", 0.0)
    sweep_dark_shot[i] = nd.get("dark_shot", 0.0)
    sweep_read_noise[i] = nd.get("read_noise", 0.0)

print(f"\n=== Nearfield Fraction Sweep Results ===")
print(f"  {'η_nf':>8s}  {'NF [e⁻]':>12s}  {'Bkg [e⁻]':>12s}  {'Total [e⁻]':>12s}  {'NF Shot':>10s}  {'SNR [—]':>10s}  {'NEDT [mK]':>10s}")
print(f"  {'-' * 8}  {'-' * 12}  {'-' * 12}  {'-' * 12}  {'-' * 10}  {'-' * 10}  {'-' * 10}")
for i in range(0, len(eta_sweep), 5):
    nedt_str = f"{sweep_nedt[i] * 1e3:>10.2f}" if not np.isnan(sweep_nedt[i]) else f"{'N/A':>10s}"
    print(f"  {eta_sweep[i]:>8.2f}  {sweep_nearfield_e[i]:>12.0f}  {sweep_background_e[i]:>12.0f}"
          f"  {sweep_total_bkg_e[i]:>12.0f}  {sweep_nf_shot[i]:>10.1f}  {sweep_snr[i]:>10.1f}  {nedt_str}")

# ---------------------------------------------------------------------------
# Step 5: Find the nearfield fraction that matches each lab measurement
# ---------------------------------------------------------------------------
# Karen's lab data gives total background signal in e⁻ at each cold stop
# position. We find the η_nf that produces the same total background.
#
# Note: Karen's "background signal" includes both scene background (from
# the shuttered cold plate — but this isn't truly zero because the chamber
# shroud is warm) and nearfield (warm optics) emission. In RADIANT terms:
#   Karen's measured bkg ≈ nearfield_e + background_e
#
# However, in Karen's test the blackbody is shuttered (cold plate at 77 K),
# so the "scene background" she sees is actually the cold plate emission
# plus chamber shroud, not the blackbody. For this analysis, we model the
# target as the scene the sensor is viewing during the shuttered test.
# Since the cold plate is at ~77 K and emits negligibly in MWIR compared
# to the warm optics at 293 K, we can approximate the shuttered case by
# using a very cold target temperature.
#
# BUT: we ran RADIANT with the blackbody as the target (308 K). For the
# background-only test (shuttered), we need a separate run with a cold
# target. Let's do that.

print(f"\n=== Shuttered-aperture model (cold plate at 77 K as target) ===")
print(f"  Re-running sweep with target = cold plate (77 K, ε = 0.98)")
print(f"  This models what Karen sees when the blackbody is shuttered.")

config_shuttered = {k: (dict(v) if isinstance(v, dict) else v) for k, v in config.items()}
config_shuttered["source"] = {
    "target": {
        "temperature": 77.0,      # Cold plate blocking the aperture [K]
        "emissivity": 0.98,
    },
    "background": {
        "temperature": 77.0,      # Cold plate also blocks the shroud [K]
        "emissivity": 0.98,       # FPA sees only cold plate + warm optics
    },
}
# NOTE: With the aperture shuttered, the FPA sees ONLY:
#   1. The cold plate at 77 K (negligible MWIR emission)
#   2. Warm optics self-emission through cold stop gaps (nearfield)
# The chamber shroud (293 K) is NOT visible — the cold plate blocks it.

# Re-sweep with shuttered aperture
shut_nearfield_e = np.zeros_like(eta_sweep)
shut_scene_bkg_e = np.zeros_like(eta_sweep)
shut_total_bkg_e = np.zeros_like(eta_sweep)
shut_snr = np.zeros_like(eta_sweep)

for i, eta in enumerate(eta_sweep):
    config_shuttered["optics"]["nearfield_fraction"] = float(eta)
    sensor_i = Sensor.from_dict(config_shuttered)
    result_i = sensor_i.evaluate()

    shut_nearfield_e[i] = result_i.stage_outputs["spectral_integration"]["nearfield_e"]
    shut_scene_bkg_e[i] = result_i.stage_outputs["spectral_integration"]["background_e"]
    shut_total_bkg_e[i] = shut_nearfield_e[i] + shut_scene_bkg_e[i]

print(f"\n=== Shuttered Sweep Results (background only) ===")
print(f"  {'η_nf':>8s}  {'NF [e⁻]':>12s}  {'Scene Bkg':>12s}  {'Total Bkg':>12s}")
print(f"  {'':>8s}  {'(warm optics)':>12s}  {'(cold plate)':>12s}  {'[e⁻]':>12s}")
print(f"  {'-' * 8}  {'-' * 12}  {'-' * 12}  {'-' * 12}")
for i in range(0, len(eta_sweep), 5):
    print(f"  {eta_sweep[i]:>8.2f}  {shut_nearfield_e[i]:>12.0f}  {shut_scene_bkg_e[i]:>12.0f}"
          f"  {shut_total_bkg_e[i]:>12.0f}")


# Invert each lab measurement to a nearfield fraction with Sensor.solve_for
# (Gap 10 — Brent root-finding on the forward model; replaces the former
# sweep + linear-interpolation workaround recorded as Gap 1 in gaps.md).
def total_background_e(r: object) -> float:
    """Model total shuttered background [e⁻]: nearfield + scene."""
    si = r.stage_outputs["spectral_integration"]
    return float(si["nearfield_e"] + si["background_e"])


sensor_shuttered = Sensor.from_dict(config_shuttered)

print(f"\n=== Matching Lab Measurements to Nearfield Fraction (Sensor.solve_for) ===")
print(f"  {'Test Point':<14s} {'Meas [e⁻]':>12s}  {'Best-fit η_nf':>13s}  {'Model [e⁻]':>12s}  {'Δ [e⁻]':>10s}  {'Evals':>6s}")
print(f"  {'-' * 14} {'-' * 12}  {'-' * 13}  {'-' * 12}  {'-' * 10}  {'-' * 6}")

matched_eta: list[float | None] = []
for d in lab_data:
    measured_bkg = d["bkg_e"]
    try:
        sol = sensor_shuttered.solve_for(
            "optics.nearfield_fraction",
            measured_bkg,
            bounds=(0.0, 1.0),
            metric=total_background_e,
        )
        matched_eta.append(sol.solution)
        print(f"  {d['test_id']:<14s} {measured_bkg:>12.0f}  {sol.solution:>13.4f}"
              f"  {sol.achieved:>12.0f}  {sol.achieved - measured_bkg:>10.1f}  {sol.n_evaluations:>6d}")
    except SolveBracketError:
        # Measurement outside the model's [0, 1] leakage range.
        matched_eta.append(None)
        if measured_bkg > shut_total_bkg_e[-1]:
            note = "(above model range: > η_nf = 1)"
        else:
            note = "(below model range: < η_nf = 0)"
        print(f"  {d['test_id']:<14s} {measured_bkg:>12.0f}  {'N/A':>13s}  {'—':>12s}  {'—':>10s}  {note}")

# ---------------------------------------------------------------------------
# Step 6: Assess requirements compliance
# ---------------------------------------------------------------------------

MAX_BKG_E = float(reqs["Max background signal (shuttered)"])  # 40,000 e⁻

print(f"\n=== Requirements Assessment ===")
print(f"  Maximum allowed background (shuttered): {MAX_BKG_E:.0f} e⁻")
print(f"")

# Find the η_nf where total background = MAX_BKG_E (Sensor.solve_for, Gap 10)
try:
    sol_req = sensor_shuttered.solve_for(
        "optics.nearfield_fraction",
        MAX_BKG_E,
        bounds=(0.0, 1.0),
        metric=total_background_e,
    )
    eta_max_bkg = sol_req.solution
    print(f"  Nearfield fraction at requirement limit: η_nf = {eta_max_bkg:.4f} [—]")
    print(f"  ({sol_req.n_evaluations} forward-model evaluations)")
    print(f"  To meet the {MAX_BKG_E:.0f} e⁻ requirement, nearfield_fraction")
    print(f"  must be ≤ {eta_max_bkg:.4f}, i.e. vendor cold stop efficiency")
    print(f"  ≥ {(1.0 - eta_max_bkg) * 100:.2f} %.")
except SolveBracketError:
    if MAX_BKG_E >= shut_total_bkg_e[-1]:
        print(f"  Requirement met across entire sweep range (η_nf 0.00–1.00).")
        eta_max_bkg = 1.0
    else:
        print(f"  Requirement not met even at η_nf = 0.00.")
        eta_max_bkg = 0.0

print(f"")
print(f"  Lab measurement compliance:")
print(f"  {'Test Point':<14s} {'Bkg [e⁻]':>12s}  {'Limit [e⁻]':>12s}  {'Status':>10s}  {'η_nf [—]':>10s}")
print(f"  {'-' * 14} {'-' * 12}  {'-' * 12}  {'-' * 10}  {'-' * 10}")
for d, eta_m in zip(lab_data, matched_eta):
    status = "PASS" if d["bkg_e"] <= MAX_BKG_E else "FAIL"
    eta_str = f"{eta_m:.4f}" if eta_m is not None else "N/A"
    print(f"  {d['test_id']:<14s} {d['bkg_e']:>12.0f}  {MAX_BKG_E:>12.0f}  {status:>10s}  {eta_str:>10s}")

# ---------------------------------------------------------------------------
# Step 7: SNR impact at nominal vs. anomalous cold stop
# ---------------------------------------------------------------------------
# Run RADIANT at the nominal measurement (CS-NOM) and at the anomalous
# measurement (CS-OFF-10, which is near the 44,000 e⁻ measurement Karen
# is concerned about).

print(f"\n=== SNR Impact: Nominal vs. Anomalous Cold Stop ===")

# Find the two key test points
nominal_data = next(d for d in lab_data if d["test_id"] == "CS-NOM")
anomaly_data = next(d for d in lab_data if d["test_id"] == "CS-OFF-10")
nominal_eta = next(e for e, d in zip(matched_eta, lab_data) if d["test_id"] == "CS-NOM")
anomaly_eta = next(e for e, d in zip(matched_eta, lab_data) if d["test_id"] == "CS-OFF-10")

for label, eta_val, meas_data in [("Nominal (CS-NOM)", nominal_eta, nominal_data),
                                    ("Anomaly (CS-OFF-10)", anomaly_eta, anomaly_data)]:
    if eta_val is None:
        print(f"\n  {label}: could not determine η_nf from measurement")
        continue

    config["optics"]["nearfield_fraction"] = float(eta_val)
    sensor_i = Sensor.from_dict(config)
    result_i = sensor_i.evaluate()

    sig_e = result_i.stage_outputs["readout"]["signal_e_final"]
    nf_e = result_i.stage_outputs["spectral_integration"]["nearfield_e"]
    bkg_e = result_i.stage_outputs["spectral_integration"]["background_e"]
    snr_i = result_i.metrics["snr"]
    nedt_i = result_i.metrics.get("nedt_K")
    niirs_i = result_i.metrics.get("niirs")
    well_margin_i = result_i.metrics.get("well_margin_dB")

    nd = {}
    for nt in result_i.noise_terms:
        nd[nt.name] = nt.value_e

    total_noise = math.sqrt(sum(v**2 for v in nd.values()))
    nf_fraction = nd.get("nearfield_shot", 0.0)**2 / total_noise**2 * 100

    print(f"\n  {label}:")
    print(f"    Matched η_nf:         {eta_val:.4f} [—]")
    print(f"    Measured bkg:         {meas_data['bkg_e']:.0f} e⁻")
    print(f"    Signal (target):      {sig_e:.0f} e⁻")
    print(f"    Nearfield (optics):   {nf_e:.0f} e⁻")
    print(f"    Scene background:     {bkg_e:.0f} e⁻")
    print(f"    SNR:                  {snr_i:.1f} — (dimensionless)")
    if nedt_i is not None:
        print(f"    NEDT:                 {nedt_i * 1e3:.2f} mK")
    if niirs_i is not None:
        print(f"    NIIRS:                {niirs_i:.2f} — (dimensionless)")
    if well_margin_i is not None:
        print(f"    Well margin:          {well_margin_i:.2f} dB")
    print(f"    Noise breakdown:")
    for name in ["signal_shot", "background_shot", "nearfield_shot", "dark_shot",
                 "read_noise", "quantization"]:
        if name in nd:
            print(f"      {name:<25s} {nd[name]:>10.2f} e⁻ RMS")
    print(f"    Nearfield noise fraction: {nf_fraction:.1f}% of total noise variance")

# ---------------------------------------------------------------------------
# Step 8: Generate plots
# ---------------------------------------------------------------------------

import matplotlib

matplotlib.use("Agg")  # headless-safe: plt.show() is a no-op, so the runner completes in CI/batch
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

PLOT_DIR = Path(__file__).parent.parent / "outputs"
PLOT_DIR.mkdir(parents=True, exist_ok=True)

# ---- Plot 1: Nearfield signal vs. cold stop efficiency (shuttered) --------
fig1, ax1 = plt.subplots(figsize=(9, 6))
ax1.plot(eta_sweep, shut_total_bkg_e / 1e3, "b-", linewidth=2,
         label="Model: total background (nearfield + scene)")

# Lab measurements overlaid
lab_etas = [e for e in matched_eta if e is not None]
lab_bkg = [d["bkg_e"] for d, e in zip(lab_data, matched_eta) if e is not None]
lab_ids = [d["test_id"] for d, e in zip(lab_data, matched_eta) if e is not None]
ax1.scatter(lab_etas, [b / 1e3 for b in lab_bkg], s=80, c="red", zorder=5,
            edgecolors="black", linewidths=0.8, label="Lab measurements")

# Annotate each lab point
for eta_pt, bkg_pt, label_pt in zip(lab_etas, lab_bkg, lab_ids):
    ax1.annotate(label_pt, (eta_pt, bkg_pt / 1e3),
                 textcoords="offset points", xytext=(8, -4), fontsize=7,
                 color="darkred")

# Requirement threshold
ax1.axhline(MAX_BKG_E / 1e3, color="red", linestyle="--", linewidth=1.2,
            label=f"Requirement: {MAX_BKG_E / 1e3:.0f} ke⁻")

# Max η annotation
if eta_max_bkg is not None and 0 < eta_max_bkg < 1:
    ax1.axvline(eta_max_bkg, color="gray", linestyle=":", linewidth=1)
    ax1.annotate(f"η_nf,max = {eta_max_bkg:.3f}",
                 xy=(eta_max_bkg, MAX_BKG_E / 1e3),
                 textcoords="offset points", xytext=(8, 8), fontsize=9,
                 arrowprops=dict(arrowstyle="->", color="gray"))

ax1.set_xlabel("Nearfield Fraction η_nf [—]  (0 = perfect cold stop, 1 = no cold stop)", fontsize=11)
ax1.set_ylabel("Background Signal [ke⁻]", fontsize=11)
ax1.set_title("Shuttered Background vs. Cold Stop Leakage", fontsize=13, fontweight="bold")
ax1.legend(loc="upper left", fontsize=9)
ax1.grid(True, alpha=0.3)
ax1.set_xlim(0, 0.12)
if lab_bkg:
    ax1.set_ylim(0, max(lab_bkg) * 1.3 / 1e3)
else:
    ax1.set_ylim(0, max(float(shut_total_bkg_e[-1]), MAX_BKG_E) * 1.3 / 1e3)
fig1.tight_layout()
fig1.savefig(PLOT_DIR / "fig1_background_vs_eta.png", dpi=150)
print(f"  Saved {PLOT_DIR / 'fig1_background_vs_eta.png'}")

# ---- Plot 2: Cold stop position vs. matched η_nf -------------------------
fig2, ax2 = plt.subplots(figsize=(8, 5))
positions = [d["position_mm"] for d, e in zip(lab_data, matched_eta) if e is not None]
etas_matched = [e for e in matched_eta if e is not None]

ax2.scatter(positions, etas_matched, s=80, c="steelblue", edgecolors="black",
            linewidths=0.8, zorder=5)

# Linear fit
if len(positions) >= 2 and max(positions) > 0:
    coeffs = np.polyfit(positions, etas_matched, 1)
    fit_x = np.linspace(0, max(positions) * 1.1, 50)
    fit_y = np.polyval(coeffs, fit_x)
    ax2.plot(fit_x, fit_y, "b--", linewidth=1, alpha=0.6,
             label=f"Linear fit: η_nf = {coeffs[0]:.4f} × pos + {coeffs[1]:.4f}")

# Requirement limit
if eta_max_bkg is not None and 0 < eta_max_bkg < 1:
    ax2.axhline(eta_max_bkg, color="red", linestyle="--", linewidth=1.2,
                label=f"Max η_nf = {eta_max_bkg:.4f} ({MAX_BKG_E / 1e3:.0f} ke⁻ limit)")

for pos, eta_pt, d in zip(positions, etas_matched, lab_data):
    color = "green" if d["bkg_e"] <= MAX_BKG_E else "red"
    ax2.annotate(d["test_id"], (pos, eta_pt),
                 textcoords="offset points", xytext=(8, 4), fontsize=8,
                 color=color)

ax2.set_xlabel("Cold Stop Offset Position [mm]", fontsize=11)
ax2.set_ylabel("Matched Nearfield Fraction η_nf [—]", fontsize=11)
ax2.set_title("Cold Stop Misalignment vs. Effective Leakage", fontsize=13, fontweight="bold")
ax2.legend(loc="upper left", fontsize=9)
ax2.grid(True, alpha=0.3)
fig2.tight_layout()
fig2.savefig(PLOT_DIR / "fig2_position_vs_eta.png", dpi=150)
print(f"  Saved {PLOT_DIR / 'fig2_position_vs_eta.png'}")

# ---- Plot 3: Noise budget comparison (nominal vs. anomaly) ---------------
fig3, axes3 = plt.subplots(1, 2, figsize=(12, 5), sharey=True)

noise_labels_short = ["Signal\nshot", "Bkg\nshot", "NF\nshot", "Dark\nshot",
                       "Read\nnoise", "Quant."]
noise_keys = ["signal_shot", "background_shot", "nearfield_shot", "dark_shot",
              "read_noise", "quantization"]
bar_colors = ["#4472C4", "#70AD47", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5"]

for ax, (label, eta_val, meas_data) in zip(
        axes3, [("Nominal (CS-NOM)", nominal_eta, nominal_data),
                ("Anomaly (CS-OFF-10)", anomaly_eta, anomaly_data)]):
    if eta_val is None:
        continue
    config["optics"]["nearfield_fraction"] = float(eta_val)
    sensor_i = Sensor.from_dict(config)
    result_i = sensor_i.evaluate()
    nd = {nt.name: nt.value_e for nt in result_i.noise_terms}

    vals = [nd.get(k, 0.0) for k in noise_keys]
    bars = ax.bar(noise_labels_short, vals, color=bar_colors, edgecolor="black",
                  linewidth=0.5)

    # Value labels on bars
    for bar, v in zip(bars, vals):
        if v > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 15,
                    f"{v:.0f}", ha="center", va="bottom", fontsize=8)

    ax.set_title(f"{label}\nη_nf = {eta_val:.4f}, Bkg = {meas_data['bkg_e']:.0f} e⁻",
                 fontsize=10, fontweight="bold")
    ax.set_ylabel("Noise [e⁻ RMS]" if ax == axes3[0] else "", fontsize=10)
    ax.grid(True, axis="y", alpha=0.3)

fig3.suptitle("Noise Budget: Nominal vs. Anomalous Cold Stop", fontsize=13, fontweight="bold")
fig3.tight_layout()
fig3.savefig(PLOT_DIR / "fig3_noise_budget.png", dpi=150)
print(f"  Saved {PLOT_DIR / 'fig3_noise_budget.png'}")

# ---- Plot 4: SNR vs. cold stop efficiency (illuminated) -------------------
fig4, ax4a = plt.subplots(figsize=(9, 5))

ax4a.plot(eta_sweep, sweep_snr, "b-", linewidth=2, label="SNR (308 K BB illuminated)")
ax4a.set_xlabel("Nearfield Fraction η_nf [—]", fontsize=11)
ax4a.set_ylabel("SNR [—]", fontsize=11, color="blue")
ax4a.tick_params(axis="y", labelcolor="blue")

# Secondary axis: nearfield signal
ax4b = ax4a.twinx()
ax4b.plot(eta_sweep, sweep_nearfield_e / 1e3, "r--", linewidth=1.5,
          label="Nearfield signal [ke⁻]")
ax4b.set_ylabel("Nearfield Signal [ke⁻]", fontsize=11, color="red")
ax4b.tick_params(axis="y", labelcolor="red")

# Mark the nominal and anomaly points on SNR curve
for eta_pt, d in [(nominal_eta, nominal_data), (anomaly_eta, anomaly_data)]:
    if eta_pt is not None:
        # Interpolate SNR at this eta
        snr_at_pt = np.interp(eta_pt, eta_sweep, sweep_snr)
        ax4a.scatter([eta_pt], [snr_at_pt], s=60, c="blue", zorder=5,
                     edgecolors="black", linewidths=0.8)
        ax4a.annotate(d["test_id"], (eta_pt, snr_at_pt),
                      textcoords="offset points", xytext=(8, -12), fontsize=8)

# Combined legend
lines_a, labels_a = ax4a.get_legend_handles_labels()
lines_b, labels_b = ax4b.get_legend_handles_labels()
ax4a.legend(lines_a + lines_b, labels_a + labels_b, loc="center right", fontsize=9)

ax4a.set_title("SNR and Nearfield Signal vs. Cold Stop Leakage (Illuminated)",
               fontsize=13, fontweight="bold")
ax4a.grid(True, alpha=0.3)
fig4.tight_layout()
fig4.savefig(PLOT_DIR / "fig4_snr_vs_eta.png", dpi=150)
print(f"  Saved {PLOT_DIR / 'fig4_snr_vs_eta.png'}")

plt.show()

# ---------------------------------------------------------------------------
# Step 9: Write results to output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb_out = openpyxl.Workbook()

header_font_out = Font(bold=True, size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: Cold Stop Sweep ---
ws1 = wb_out.active
ws1.title = "Cold Stop Sweep"
ws1["A1"] = "Scenario 7.4: Cold Stop Efficiency Sweep (Shuttered Aperture)"
ws1["A1"].font = Font(bold=True, size=14)

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 16
ws1.column_dimensions["G"].width = 16

# Columns 2–4 are the SHUTTERED sweep (what Karen measures); columns 5–7
# are from the ILLUMINATED sweep at the same η_nf (operational impact).
sweep_headers = ["η_nf [—]", "Nearfield [e⁻]", "Scene Bkg [e⁻]",
                 "Total Bkg [e⁻]", "NF Shot (illum) [e⁻]",
                 "Bkg Shot (illum) [e⁻]", "SNR (illum) [—]"]
for col, h_text in enumerate(sweep_headers, 1):
    cell = ws1.cell(row=3, column=col, value=h_text)
    cell.font = header_font_out
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

for i in range(len(eta_sweep)):
    r = i + 4
    ws1.cell(row=r, column=1, value=round(eta_sweep[i], 2)).border = thin_border
    ws1.cell(row=r, column=2, value=round(float(shut_nearfield_e[i]), 0)).border = thin_border
    ws1.cell(row=r, column=3, value=round(float(shut_scene_bkg_e[i]), 0)).border = thin_border
    ws1.cell(row=r, column=4, value=round(float(shut_total_bkg_e[i]), 0)).border = thin_border
    ws1.cell(row=r, column=5, value=round(float(sweep_nf_shot[i]), 2)).border = thin_border
    ws1.cell(row=r, column=6, value=round(float(sweep_bkg_shot[i]), 2)).border = thin_border
    ws1.cell(row=r, column=7, value=round(float(sweep_snr[i]), 2)).border = thin_border

    # Color-code total background against requirement
    bkg_cell = ws1.cell(row=r, column=4)
    bkg_cell.fill = pass_fill if shut_total_bkg_e[i] <= MAX_BKG_E else fail_fill

# --- Sheet 2: Lab Comparison ---
ws2 = wb_out.create_sheet("Lab Comparison")
ws2["A1"] = "Cold Stop Position vs. Model Efficiency Matching"
ws2["A1"].font = Font(bold=True, size=14)

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 14

lab_headers = ["Test Point", "Position [mm]", "Meas Bkg [e⁻]", "Best-fit η_nf [—]", "Status"]
for col, h_text in enumerate(lab_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h_text)
    cell.font = header_font_out
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

for i, (d, eta_m) in enumerate(zip(lab_data, matched_eta), 4):
    ws2.cell(row=i, column=1, value=d["test_id"]).border = thin_border
    ws2.cell(row=i, column=2, value=d["position_mm"]).border = thin_border
    ws2.cell(row=i, column=3, value=d["bkg_e"]).border = thin_border
    ws2.cell(row=i, column=4, value=round(eta_m, 4) if eta_m is not None else "N/A").border = thin_border
    status = "PASS" if d["bkg_e"] <= MAX_BKG_E else "FAIL"
    status_cell = ws2.cell(row=i, column=5, value=status)
    status_cell.border = thin_border
    status_cell.fill = pass_fill if status == "PASS" else fail_fill

# --- Sheet 3: Summary ---
ws3 = wb_out.create_sheet("Summary")
ws3["A1"] = "Scenario 7.4: Cold Stop Analysis Summary"
ws3["A1"].font = Font(bold=True, size=14)

ws3.column_dimensions["A"].width = 45
ws3.column_dimensions["B"].width = 25

summary_items = [
    ("Instrument", ""),
    ("Aperture", f"{aperture_m * 100:.0f} cm"),
    ("f-number", f"{f_number}"),
    ("Optics temperature", f"{optics_temp_K:.1f} K ({specs['Optics barrel temperature']}°C)"),
    ("Spectral band", f"{band_min_um:.2f} – {band_max_um:.2f} µm"),
    ("Integration time", f"{t_int_s * 1000:.0f} ms"),
    ("", ""),
    ("Baseline (η_nf = 1.0, no cold stop)", ""),
    ("Nearfield signal [e⁻]", f"{baseline_nearfield_e:.0f}"),
    ("Background signal [e⁻]", f"{baseline_background_e:.0f}"),
    ("SNR [—]", f"{baseline_snr:.1f}"),
    ("NEDT [mK]", f"{baseline_nedt * 1e3:.2f}" if baseline_nedt is not None else "N/A"),
    ("NIIRS [—]", f"{baseline_niirs:.2f}" if baseline_niirs is not None else "N/A"),
    ("MTF at Nyquist [—]", f"{baseline_mtf_nyq:.4f}" if baseline_mtf_nyq is not None else "N/A"),
    ("Strehl ratio [—]", f"{baseline_strehl:.4f}" if baseline_strehl is not None else "N/A"),
    ("Well margin [dB]", f"{baseline_well_margin:.2f}" if baseline_well_margin is not None else "N/A"),
    ("", ""),
    ("Requirements", ""),
    ("Max shuttered background [e⁻]", f"{MAX_BKG_E:.0f}"),
    ("Max η_nf to meet requirement [—]",
     f"{eta_max_bkg:.4f}" if eta_max_bkg is not None else "N/A"),
    ("", ""),
    ("Lab Results", ""),
]

for d, eta_m in zip(lab_data, matched_eta):
    eta_str = f"{eta_m:.4f}" if eta_m is not None else "N/A"
    status = "PASS" if d["bkg_e"] <= MAX_BKG_E else "FAIL"
    summary_items.append((f"{d['test_id']} ({d['position_mm']:.1f} mm offset)",
                          f"{d['bkg_e']:.0f} e⁻ → η_nf = {eta_str} [{status}]"))

for i, (label, value) in enumerate(summary_items, 3):
    cell_a = ws3.cell(row=i, column=1, value=label)
    cell_b = ws3.cell(row=i, column=2, value=value)
    if label and not value:
        cell_a.font = Font(bold=True, size=11)

wb_out.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
