#!/usr/bin/env python3
"""Create Dr. Chen's inputs for Scenario 6.5: emissivity sensitivity.

Emits ``chen_retrieval_config.xlsx`` — the true LWIR scene and the range of
*assumed* emissivities over which Dr. Chen studies temperature-retrieval
bias.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Retrieval"
ws["A1"] = "Scenario 6.5: emissivity sensitivity for LWIR temperature retrieval"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [28, 14, 12, 44]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="375623")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("True surface temperature", 300.0, "K", "The value retrieval should recover"),
    ("True emissivity", 0.95, "—", "Actual surface emissivity"),
    ("Assumed emissivity min", 0.90, "—", "Sweep lower bound"),
    ("Assumed emissivity max", 1.00, "—", "Sweep upper bound"),
    ("Assumed emissivity step", 0.01, "—", ""),
    ("Waveband", "8–12", "µm", "LWIR"),
    ("System NEDT", 50.0, "mK", "For NEDT-equivalent ε uncertainty"),
    ("Aperture diameter", 15.0, "cm", ""),
    ("Focal length", 0.5, "m", ""),
    ("Pixel pitch", 25.0, "µm", ""),
    ("Optical transmission", 80.0, "%", ""),
    ("QE", 70.0, "%", ""),
    ("Integration time", 5.0, "ms", ""),
]
r = 4
for name, value, unit, note in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "chen_retrieval_config.xlsx"
wb.save(out)
print(f"Wrote {out}")
