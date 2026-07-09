#!/usr/bin/env python3
"""Create Dr. Chen's inputs for Scenario 6.1: published-datasheet benchmark.

Emits ``chen_lwir_datasheet.xlsx`` — a published cooled-LWIR HgCdTe FPA
datasheet (specific detectivity D*, NETD, pixel, band, reference optics)
that Dr. Chen benchmarks RADIANT against.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Datasheet"
ws["A1"] = "Scenario 6.1: published cooled-LWIR HgCdTe FPA datasheet"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [30, 16, 14, 44]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="264478")
for col, text in zip("ABCD", ["Quantity", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("PUBLISHED SPECS", None, None, None),
    ("Specific detectivity D*", 2.0e11, "cm·Hz^½/W", "Peak, BLIP-class (Jones)"),
    ("NETD", 25.0, "mK", "f/2, 300 K scene, 8–12 µm"),
    ("REFERENCE CONDITIONS", None, None, None),
    ("Waveband", "8–12", "µm", "LWIR"),
    ("Pixel pitch", 30.0, "µm", "Detector element"),
    ("f-number", 2.0, "—", "Cold-shielded"),
    ("Scene temperature", 300.0, "K", "NETD reference scene"),
    ("Integration time", 30.0, "µs", "Sets the noise bandwidth Δf = 1/(2·t)"),
    ("Quantum efficiency", 75.0, "%", "In-band"),
    ("DETECTOR COMPONENTS (as-built)", None, None, None),
    ("Dark current", 1.0e5, "e⁻/s", "Cooled to 77 K"),
    ("Read noise", 60.0, "e⁻ RMS", ""),
    ("Full well", 1.0e7, "e⁻", ""),
    ("Optical transmission", 85.0, "%", ""),
    ("BENCHMARK", None, None, None),
    ("Validation tolerance", 15.0, "%", "Chain vs datasheet D* and NETD"),
]
r = 4
section_fill = PatternFill("solid", fgColor="DCE6F1")
for name, value, unit, note in ROWS:
    if value is None and unit is None:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True, size=11)
        for cc in range(1, 5):
            ws.cell(row=r, column=cc).fill = section_fill
    else:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "chen_lwir_datasheet.xlsx"
wb.save(out)
print(f"Wrote {out}")
