"""Create Raj's weather sensitivity spreadsheet.

Three sheets:
  1. "Sensor Configuration"  — MWIR sensor baseline (existing design)
  2. "Atmosphere & Geometry"  — observation geometry, atmosphere baseline
  3. "Sweep Definition"       — visibility & PWV sweep ranges, NIIRS requirement
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=4):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Sensor Configuration
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Sensor Configuration"
ws1["A1"] = "MWIR Reconnaissance Sensor — Baseline Configuration"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "500 km SSO, nadir-looking, extended-scene target"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 55

row = 4
row = add_section(ws1, row, "Optics")
row = add_param(ws1, row, "Aperture diameter", 30.0, "cm",
                "Cassegrain telescope")
row = add_param(ws1, row, "Focal length", 120.0, "cm", "f/4")
row = add_param(ws1, row, "f-number", 4.0, "—", "Moderate speed")
row = add_param(ws1, row, "Optical transmission", 75.0, "%",
                "All elements combined")
row = add_param(ws1, row, "Optics temperature", 20.0, "°C",
                "Ambient (space thermal control)")
row = add_param(ws1, row, "Number of optical elements", 5, "—",
                "Primary, secondary, fold, filter, window")

row += 1
row = add_section(ws1, row, "Detector")
row = add_param(ws1, row, "Format", "1024x1024", "pixels", "HgCdTe MWIR FPA")
row = add_param(ws1, row, "Pixel pitch", 18.0, "µm", "Square pixels")
row = add_param(ws1, row, "Cutoff wavelength", 5.0, "µm", "At 80 K")
row = add_param(ws1, row, "Quantum efficiency", 70.0, "%",
                "In-band average, 3.5–5.0 µm")
row = add_param(ws1, row, "Dark current", 150.0, "e⁻/s",
                "At 80 K operating temperature")
row = add_param(ws1, row, "Operating temperature", 80.0, "K",
                "Stirling cooler")
row = add_param(ws1, row, "Read noise (post-CDS)", 18.0, "e⁻ RMS",
                "CDS mode")
row = add_param(ws1, row, "Full well capacity", 500000, "e⁻",
                "Standard ROIC")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "14-bit ADC")
row = add_param(ws1, row, "System gain", 32.0, "e⁻/DN",
                "Set for 105% FWC at ADC max")
row = add_param(ws1, row, "IPC coupling", 1.5, "%",
                "Per-neighbor, measured")
row = add_param(ws1, row, "Integration time", 1.0, "ms",
                "LEO staring: limited by smear (7 km/s ground vel, 7.5 m GSD)")

# ---------------------------------------------------------------------------
# Sheet 2: Atmosphere & Geometry
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Atmosphere & Geometry")
ws2["A1"] = "Observation Geometry & Atmosphere Baseline"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "500 km SSO, nadir look, midlatitude summer"
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 55

row = 4
row = add_section(ws2, row, "Observation Geometry")
row = add_param(ws2, row, "Sensor altitude", 500.0, "km",
                "Sun-synchronous orbit")
row = add_param(ws2, row, "Target altitude", 0.0, "m",
                "Sea-level target")
row = add_param(ws2, row, "Off-nadir angle", 0.0, "deg",
                "Nadir look (best case)")

row += 1
row = add_section(ws2, row, "Atmosphere Baseline")
row = add_param(ws2, row, "Standard atmosphere", "midlat_summer", "—",
                "MODTRAN profile")
row = add_param(ws2, row, "Baseline visibility", 23.0, "km",
                "Koschmieder clear-day default")
row = add_param(ws2, row, "Baseline PWV", 1.4, "cm",
                "US Standard annual mean")
row = add_param(ws2, row, "Aerosol type", "rural", "—",
                "Continental/rural aerosol model")

row += 1
row = add_section(ws2, row, "Scene")
row = add_param(ws2, row, "Target temperature", 300.0, "K",
                "Standard ambient target (building, road)")
row = add_param(ws2, row, "Target emissivity", 0.95, "—",
                "Concrete/asphalt")
row = add_param(ws2, row, "Background temperature", 290.0, "K",
                "Terrain background")
row = add_param(ws2, row, "Background emissivity", 0.97, "—",
                "Vegetation/soil")

# ---------------------------------------------------------------------------
# Sheet 3: Sweep Definition
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Sweep Definition")
ws3["A1"] = "Weather Sensitivity Sweep — Requirements & Ranges"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Find visibility and PWV thresholds for NIIRS ≥ 4.0"
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 55

row = 4
row = add_section(ws3, row, "Performance Requirement")
row = add_param(ws3, row, "Minimum NIIRS", 4.0, "—",
                "Mission threshold for useful imagery")
row = add_param(ws3, row, "NIIRS goal", 5.0, "—",
                "Desired performance for vehicle ID")

row += 1
row = add_section(ws3, row, "Visibility Sweep")
row = add_param(ws3, row, "Visibility min", 2.0, "km",
                "Heavy haze / light fog")
row = add_param(ws3, row, "Visibility max", 100.0, "km",
                "Crystal clear")
row = add_param(ws3, row, "Visibility points", 25, "—",
                "Log-spaced sweep")

row += 1
row = add_section(ws3, row, "Precipitable Water Vapor Sweep")
row = add_param(ws3, row, "PWV min", 0.5, "cm",
                "Dry desert / winter arctic")
row = add_param(ws3, row, "PWV max", 5.0, "cm",
                "Tropical humid")
row = add_param(ws3, row, "PWV points", 15, "—",
                "Linear-spaced sweep")

row += 1
row = add_section(ws3, row, "Named Weather Conditions")
headers = ["Condition", "Visibility [km]", "PWV [cm]", "Description"]
for c, h in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border
row += 1

conditions = [
    ("Crystal clear", 100.0, 0.8, "Best-case winter desert"),
    ("Clear", 50.0, 1.0, "Standard clear day"),
    ("Standard", 23.0, 1.4, "Koschmieder default"),
    ("Light haze", 10.0, 2.0, "Mild aerosol loading"),
    ("Moderate haze", 5.0, 3.0, "Visible haze, summer humid"),
    ("Heavy haze", 2.0, 4.0, "Poor visibility, tropical"),
    ("Tropical humid", 10.0, 5.0, "High moisture, moderate vis"),
    ("Arctic dry", 50.0, 0.5, "Cold dry, clear"),
]
for cond, vis, pwv, desc in conditions:
    ws3.cell(row=row, column=1, value=cond).border = thin_border
    ws3.cell(row=row, column=2, value=vis).border = thin_border
    ws3.cell(row=row, column=3, value=pwv).border = thin_border
    ws3.cell(row=row, column=4, value=desc).border = thin_border
    row += 1


from pathlib import Path
out = Path(__file__).parent / "raj_weather_sensitivity_data.xlsx"
wb.save(out)
print(f"Created {out}")
