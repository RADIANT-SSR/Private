"""Create Karen's lab notebook spreadsheet for NEDT reconciliation.

Three sheets:
  1. "System Configuration" — optics, spectral, geometry
  2. "As-Built Detector"    — measured detector parameters (vendor + lab)
  3. "NEDT Measurements"    — measured NEDT at multiple blackbody temperatures
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Sheet 1: System Configuration
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "System Configuration"
ws1["A1"] = "Scenario 7.1: NEDT Reconciliation — System Configuration"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "As-built MWIR sensor, lab/TVAC test configuration"
ws1.column_dimensions["A"].width = 40
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 40

# Section: Optics
row = 4
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Optics", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

data_1 = [
    ("Primary aperture diameter", 30.0, "cm", "As-built, measured on optical bench"),
    ("Effective focal length", 121.5, "cm", "Measured via star test (nominal 120 cm)"),
    ("f-number (actual)", 4.05, "—", "Derived from D and f (nominal f/4.0)"),
    ("Optical transmission", 71.0, "%", "Measured at λ = 4.25 µm, includes all elements"),
    ("Optics temperature", 22.0, "°C", "Lab ambient (TVAC shroud at 22°C)"),
    ("Number of optical elements", 4, "—", "Primary, secondary, fold, window"),
]

for i, (param, val, unit, note) in enumerate(data_1, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# Section: Spectral
row = 12
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Spectral Band", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

data_2 = [
    ("Filter cut-on", 3500.0, "nm", "Cold filter, measured at 77 K"),
    ("Filter cut-off", 5000.0, "nm", "Cold filter, measured at 77 K"),
    ("Band center", 4250.0, "nm", "Arithmetic center"),
]

for i, (param, val, unit, note) in enumerate(data_2, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# Section: Test Setup
row = 17
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Test Setup", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

data_3 = [
    ("Blackbody emissivity", 0.995, "—", "CI Systems SR-800R extended area blackbody"),
    ("Blackbody fill", "overfills FOV", "—", "30×30 cm aperture at 50 cm working distance"),
    ("Chamber pressure", "<1e-5", "Torr", "TVAC chamber, turbomolecular pump"),
    ("Shroud temperature", 22.0, "°C", "Ambient lab temperature"),
    ("Shroud emissivity", 0.95, "—", "Anodized aluminum shroud"),
    ("Cold stop efficiency", 1.0, "—", "Designed; assumed 100% for this test"),
]

for i, (param, val, unit, note) in enumerate(data_3, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# ---------------------------------------------------------------------------
# Sheet 2: As-Built Detector
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("As-Built Detector")
ws2["A1"] = "As-Built Detector Parameters (Measured)"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "HgCdTe MWIR FPA, 640×512, Lot W2024-037"
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 45

row = 4
for c in range(1, 5):
    cell = ws2.cell(row=row, column=c,
                    value=["Parameter", "Value", "Unit", "Notes"][c - 1])
    cell.font = header_font
    cell.border = thin_border

det_data = [
    ("Pixel pitch", 18.0, "µm", "Square pixels"),
    ("Quantum efficiency (in-band avg)", 68.0, "%", "Measured at 77 K, 3.5–5.0 µm average"),
    ("Dark current (measured)", 135.0, "e⁻/s", "At 77.0 K operating temperature"),
    ("Operating temperature", 77.0, "K", "LN2 cold finger, ±0.1 K stability"),
    ("Read noise (post-CDS)", 14.2, "e⁻ RMS", "Measured in dark, 200-frame ensemble"),
    ("Full well capacity", 500000.0, "e⁻", "90% linearity limit"),
    ("ADC resolution", 14, "bits", "14-bit ADC, 2's complement"),
    ("System gain", 12.0, "e⁻/DN", "Set for 70% FWC at ADC max"),
    ("Integration time", 0.5, "ms", "Lab NEDT test (shorter than 8 ms orbital nominal to avoid well saturation)"),
    ("IPC coupling", 1.5, "%", "Measured via single-pixel reset method"),
    ("Flicker noise (1/f)", 0.0, "e⁻", "Below measurement floor at this frame rate"),
    ("ROIC glow", 5.0, "e⁻/s", "Measured with cold shutter closed"),
]

for i, (param, val, unit, note) in enumerate(det_data, 5):
    ws2.cell(row=i, column=1, value=param).border = thin_border
    ws2.cell(row=i, column=2, value=val).border = thin_border
    ws2.cell(row=i, column=3, value=unit).border = thin_border
    ws2.cell(row=i, column=4, value=note).border = thin_border

# Nominal vs. as-built comparison
row = 19
for c in range(1, 5):
    cell = ws2.cell(row=row, column=c,
                    value=["Nominal vs. As-Built", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

compare_data = [
    ("Parameter", "Nominal", "As-Built", "Delta"),
    ("QE [%]", 72.0, 68.0, -4.0),
    ("Dark current [e⁻/s]", 100.0, 135.0, 35.0),
    ("Read noise [e⁻ RMS]", 12.0, 14.2, 2.2),
    ("f-number", 4.0, 4.05, 0.05),
    ("Transmission [%]", 73.0, 71.0, -2.0),
]

for i, row_data in enumerate(compare_data, row + 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=c, value=val)
        cell.border = thin_border
        if i == row + 1:
            cell.font = header_font

# ---------------------------------------------------------------------------
# Sheet 3: NEDT Measurements
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("NEDT Measurements")
ws3["A1"] = "NEDT Measurements — TVAC Test Campaign"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Date: 2026-03-28, Operator: K. Martinez, Chamber: TVAC-3"
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 30

row = 4
headers = ["Blackbody T [°C]", "Blackbody T [K]", "Measured NEDT [mK]",
           "Std Dev [mK]", "N frames", "Notes"]
for c, h in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border

# NEDT measurements at multiple blackbody temperatures
nedt_data = [
    (15.0, 288.15, 160.0, 8.5, 500, "Low-temperature point"),
    (20.0, 293.15, 142.0, 7.2, 500, ""),
    (25.0, 298.15, 127.0, 6.0, 500, "Primary test point"),
    (30.0, 303.15, 114.0, 5.1, 500, ""),
    (35.0, 308.15, 104.0, 4.5, 500, ""),
    (40.0, 313.15, 95.0, 3.8, 500, ""),
    (50.0, 323.15, 81.0, 3.0, 500, "High-temperature point"),
]

for i, (t_c, t_k, nedt, std, nf, note) in enumerate(nedt_data, 5):
    ws3.cell(row=i, column=1, value=t_c).border = thin_border
    ws3.cell(row=i, column=2, value=t_k).border = thin_border
    ws3.cell(row=i, column=3, value=nedt).border = thin_border
    ws3.cell(row=i, column=4, value=std).border = thin_border
    ws3.cell(row=i, column=5, value=nf).border = thin_border
    ws3.cell(row=i, column=6, value=note).border = thin_border

# Notes
ws3.cell(row=13, column=1, value="Measurement Method:").font = Font(bold=True)
ws3.cell(row=14, column=1,
         value="NEDT computed as σ(frame-to-frame pixel value) / (dS/dT) "
               "from 2-point calibration at ±2°C around each test temperature.")
ws3.cell(row=15, column=1,
         value="Spatial average of NEDT across central 100×100 pixel ROI "
               "(excludes edge pixels and known bad pixels).")

from pathlib import Path
out = Path(__file__).parent / "karen_nedt_lab_data.xlsx"
wb.save(out)
print(f"Created {out}")
