"""Scenario 5.3: Monochromatic vs. Polychromatic PSF Comparison.

Tom wants to understand how much chromaticism matters for his MWIR system.
His band spans 3.5–5.0 µm — the Airy disk grows ~44% from short to long λ.
Does a single monochromatic PSF at band-center give misleading spatial metrics
compared to a properly flux-weighted polychromatic PSF?

Approach:
  1. Run RADIANT at each individual wavelength (narrow band) to get per-λ metrics
  2. Run RADIANT with monochromatic PSF (N=1, band-center) over full band
  3. Run RADIANT with polychromatic PSF (N=11, N=21) over full band
  4. Compare MTF, EE, FWHM, SNR across all cases
  5. Quantify the chromaticism error from monochromatic analysis

Input: tom_chromatic_psf_analysis.xlsx (3-sheet workbook in Zemax/vendor units)
Output: Comparison tables, 4 plots, Excel results
"""

import sys
import os
import numpy as np
import openpyxl
from openpyxl.styles import Font
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. Read Tom's spreadsheet
# ---------------------------------------------------------------------------

input_dir = Path(__file__).resolve().parent.parent / "inputs"
xlsx_path = input_dir / "tom_chromatic_psf_analysis.xlsx"

if not xlsx_path.exists():
    print(f"Spreadsheet not found at {xlsx_path}")
    print("Run create_spreadsheet.py first.")
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# --- Sheet 1: Optical Design ---
ws1 = wb["Optical Design"]
design = {}
for row in ws1.iter_rows(min_row=6, max_col=4, values_only=False):
    cell_fill = row[0].fill
    if cell_fill and cell_fill.start_color and cell_fill.start_color.rgb == "002E75B6":
        continue  # skip blue section headers
    name = row[0].value
    value = row[1].value
    if name and value is not None and isinstance(value, (int, float)):
        design[name] = value

print("=" * 80)
print("SCENARIO 5.3: Monochromatic vs. Polychromatic PSF Comparison")
print("=" * 80)
print()
print("Optical Design (from Tom's Zemax export):")
for k, v in design.items():
    print(f"  {k}: {v}")
print()

# ---------------------------------------------------------------------------
# 2. Unit conversions (Zemax/vendor → RADIANT canonical)
# ---------------------------------------------------------------------------
# Zemax convention: lengths in mm, wavelengths in nm, percentages, ms

aperture_mm = design["Entrance pupil diameter"]
focal_length_mm = design["Effective focal length"]
f_number = design["f-number (working)"]
transmission_pct = design["Optical transmission"]
optics_temp_K = design["Optics temperature"]
wfe_waves = design["WFE (RMS, on-axis)"]

lambda_center_nm = design["Band center wavelength"]
filter_min_nm = design["Filter cut-on"]
filter_max_nm = design["Filter cut-off"]

pitch_um = design["Pixel pitch"]
qe_pct = design["Quantum efficiency (in-band)"]
dark_rate_e_per_s = design["Dark current"]
fwc = design["Full well capacity"]
read_noise = design["Read noise (CDS)"]
adc_bits = design["ADC resolution"]
gain = design["System gain"]
t_int_ms = design["Integration time"]

target_temp = design["Target temperature"]
target_emiss = design["Target emissivity"]
altitude_km = design["Orbit altitude"]

# Convert
aperture_m = aperture_mm / 1000.0       # mm → m
focal_length_m = focal_length_mm / 1000.0  # mm → m
transmission = transmission_pct / 100.0  # % → fraction
qe = qe_pct / 100.0                     # % → fraction
t_int_s = t_int_ms / 1000.0             # ms → s
altitude_m = altitude_km * 1000.0        # km → m

lambda_center_um = lambda_center_nm / 1000.0  # nm → µm
band_min_um = filter_min_nm / 1000.0     # nm → µm
band_max_um = filter_max_nm / 1000.0     # nm → µm

print("Unit conversions (boundary):")
print(f"  Aperture:     {aperture_mm} mm → {aperture_m} m")
print(f"  Focal length: {focal_length_mm} mm → {focal_length_m} m")
print(f"  Transmission: {transmission_pct}% → {transmission}")
print(f"  QE:           {qe_pct}% → {qe}")
print(f"  Band:         {filter_min_nm}–{filter_max_nm} nm → {band_min_um}–{band_max_um} µm")
print(f"  Integration:  {t_int_ms} ms → {t_int_s} s")
print(f"  Altitude:     {altitude_km} km → {altitude_m} m")
print()

# ---------------------------------------------------------------------------
# 3. Compute per-wavelength Airy disk and Q (analytic, before RADIANT)
# ---------------------------------------------------------------------------

analysis_wavelengths_nm = [3500, 4000, 4250, 4500, 5000]
analysis_wavelengths_um = [w / 1000.0 for w in analysis_wavelengths_nm]

print("Per-wavelength sampling analysis (18 µm pixel, f/4):")
print(f"  {'λ [µm]':>8}  {'Airy Ø [µm]':>12}  {'Q':>6}  {'Regime':>14}")
print(f"  {'------':>8}  {'-----------':>12}  {'-----':>6}  {'------':>14}")
for lam_um in analysis_wavelengths_um:
    airy_diam = 2.44 * lam_um * f_number  # µm
    q = lam_um * f_number / pitch_um
    if q > 1.5:
        regime = "oversampled"
    elif q > 1.0:
        regime = "well-sampled"
    elif q > 0.7:
        regime = "undersampled"
    else:
        regime = "ALIASED"
    print(f"  {lam_um:>8.2f}  {airy_diam:>12.1f}  {q:>6.2f}  {regime:>14}")
print()
print("  Key observation: Q ranges from 0.78 (3.5 µm) to 1.11 (5.0 µm).")
print("  At 3.5 µm, the system is slightly undersampled; at 5.0 µm, it's well-sampled.")
print("  The Airy disk grows 44% across the band (34.2 → 48.8 µm).")
print()

# ---------------------------------------------------------------------------
# 4. Build RADIANT configs and run
# ---------------------------------------------------------------------------

try:
    from radiant.api import Sensor
except ImportError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[4] / "src"))
    from radiant.api import Sensor


def make_config(filter_min, filter_max, psf_n_wavelengths=1):
    """Build RADIANT config dict. All values already in canonical units."""
    return {
        "source": {
            "target": {
                "temperature": target_temp,
                "emissivity": target_emiss,
            },
            "background": {
                "temperature": 295.0,
                "emissivity": 0.95,
            },
        },
        "atmosphere": {
            "model": "simple",
        },
        "geometry": {
            "sensor_altitude_m": altitude_m,
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
            "psf_n_wavelengths": psf_n_wavelengths,
        },
        "detector": {
            "pixel_pitch_x_um": pitch_um,
            "pixel_pitch_y_um": pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate_e_per_s,
            "detector_temperature_K": 77.0,
        },
        "spectral_integration": {
            "filter_min_um": filter_min,
            "filter_max_um": filter_max,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
    }


def run_and_extract(config, label):
    """Run RADIANT and extract key spatial + radiometric metrics."""
    sensor = Sensor.from_dict(config)
    result = sensor.evaluate()

    metrics = {
        "label": label,
        "snr": result.metrics["snr"],
        "mtf_at_nyquist": result.metrics["mtf_at_nyquist"],
        "ee_1x1": result.metrics["ee_1x1"],
        "ee_3x3": result.metrics["ee_3x3"],
        "fwhm_x_m": result.metrics["fwhm_x_m"],
        "rer": result.metrics.get("rer"),
        "signal_e": result.stage_outputs["readout"]["signal_e_final"],
        "q_center": result.metrics.get("q_center"),
        "gsd_cross_m": result.metrics.get("gsd_cross_track_m"),
        "nedt_K": result.metrics.get("nedt_K"),
        "niirs": result.metrics.get("niirs"),
        "strehl": result.metrics.get("strehl"),
    }
    return metrics, result


# --- 4a. Per-wavelength runs (narrow-band, monochromatic PSF) ---
# Use a narrow band (±50 nm) centered at each analysis wavelength.
# This isolates the PSF behavior at each wavelength.

print("-" * 80)
print("PART 1: Per-wavelength analysis (narrow-band, monochromatic PSF)")
print("-" * 80)
print()

half_bw_um = 0.050  # ±50 nm = 100 nm total bandwidth
per_wavelength_results = []

for lam_um in analysis_wavelengths_um:
    fmin = lam_um - half_bw_um
    fmax = lam_um + half_bw_um
    label = f"λ = {lam_um:.2f} µm (mono, ±50 nm)"
    config = make_config(fmin, fmax, psf_n_wavelengths=1)
    metrics, _ = run_and_extract(config, label)
    metrics["wavelength_um"] = lam_um
    metrics["airy_diam_um"] = 2.44 * lam_um * f_number
    metrics["q_parameter"] = lam_um * f_number / pitch_um
    per_wavelength_results.append(metrics)

print(f"  {'λ [µm]':>8}  {'Airy Ø [µm]':>12}  {'Q':>6}  {'MTF@Nyq':>8}"
      f"  {'EE 1×1':>7}  {'EE 3×3':>7}  {'RER':>6}  {'FWHM [µm]':>10}  {'SNR':>7}")
print(f"  {'------':>8}  {'-----------':>12}  {'-----':>6}  {'-------':>8}"
      f"  {'------':>7}  {'------':>7}  {'-----':>6}  {'---------':>10}  {'---':>7}")
for r in per_wavelength_results:
    fwhm_um = r["fwhm_x_m"] * 1e6
    rer_s = f"{r['rer']:.3f}" if r.get("rer") else "N/A"
    print(f"  {r['wavelength_um']:>8.2f}  {r['airy_diam_um']:>12.1f}"
          f"  {r['q_parameter']:>6.2f}  {r['mtf_at_nyquist']:>8.3f}"
          f"  {r['ee_1x1']:>7.3f}  {r['ee_3x3']:>7.3f}"
          f"  {rer_s:>6}  {fwhm_um:>10.1f}  {r['snr']:>7.0f}")

print()
print("  Physics: shorter λ → smaller Airy disk → more energy in single pixel (higher EE)")
print("  but also higher Nyquist-to-cutoff ratio → MTF at Nyquist depends on where")
print("  Nyquist sits relative to the diffraction cutoff.")
print()

# --- 4b. Full-band runs: mono (N=1) vs. poly (N=11, N=21) ---

print("-" * 80)
print("PART 2: Full-band comparison — mono vs. polychromatic PSF")
print("-" * 80)
print()

psf_configs = [
    (1, "Monochromatic (N=1, band-center)"),
    (5, "Polychromatic (N=5)"),
    (11, "Polychromatic (N=11)"),
    (21, "Polychromatic (N=21)"),
]

fullband_results = []
for n_wl, label in psf_configs:
    config = make_config(band_min_um, band_max_um, psf_n_wavelengths=n_wl)
    metrics, result = run_and_extract(config, label)
    metrics["n_wavelengths"] = n_wl
    fullband_results.append(metrics)

print(f"  {'PSF Model':>35}  {'MTF@Nyq':>8}  {'EE 1×1':>7}  {'EE 3×3':>7}"
      f"  {'RER':>6}  {'FWHM [µm]':>10}  {'SNR':>7}  {'NEDT [K]':>9}  {'NIIRS':>6}")
print(f"  {'--------':>35}  {'-------':>8}  {'------':>7}  {'------':>7}"
      f"  {'-----':>6}  {'---------':>10}  {'---':>7}  {'--------':>9}  {'-----':>6}")
for r in fullband_results:
    fwhm_um = r["fwhm_x_m"] * 1e6
    rer_s = f"{r['rer']:.3f}" if r.get("rer") else "N/A"
    nedt_s = f"{r['nedt_K']:.4f}" if r.get("nedt_K") else "N/A"
    niirs_s = f"{r['niirs']:.1f}" if r.get("niirs") else "N/A"
    print(f"  {r['label']:>35}  {r['mtf_at_nyquist']:>8.3f}"
          f"  {r['ee_1x1']:>7.3f}  {r['ee_3x3']:>7.3f}"
          f"  {rer_s:>6}  {fwhm_um:>10.1f}  {r['snr']:>7.0f}"
          f"  {nedt_s:>9}  {niirs_s:>6}")

# --- 4c. Compute chromaticism error ---

mono = fullband_results[0]
poly_11 = fullband_results[2]  # N=11 as reference

print()
print("  Chromaticism error (monochromatic vs. polychromatic N=11):")

def pct_diff(a, b):
    """Percent difference: (a - b) / b × 100."""
    if abs(b) < 1e-15:
        return float("inf")
    return (a - b) / b * 100.0

mtf_err = pct_diff(mono["mtf_at_nyquist"], poly_11["mtf_at_nyquist"])
ee1_err = pct_diff(mono["ee_1x1"], poly_11["ee_1x1"])
ee3_err = pct_diff(mono["ee_3x3"], poly_11["ee_3x3"])
rer_err = pct_diff(mono["rer"], poly_11["rer"]) if mono.get("rer") and poly_11.get("rer") else None
fwhm_err = pct_diff(mono["fwhm_x_m"], poly_11["fwhm_x_m"])
snr_err = pct_diff(mono["snr"], poly_11["snr"])

print(f"    MTF at Nyquist:  {mtf_err:+.1f}%")
print(f"    EE 1×1:          {ee1_err:+.1f}%")
print(f"    EE 3×3:          {ee3_err:+.1f}%")
if rer_err is not None:
    print(f"    RER:             {rer_err:+.1f}%")
print(f"    FWHM:            {fwhm_err:+.1f}%")
print(f"    SNR:             {snr_err:+.1f}%")
print()

# --- 4d. Convergence check (N=11 vs N=21) ---
poly_21 = fullband_results[3]
conv_mtf = abs(pct_diff(poly_11["mtf_at_nyquist"], poly_21["mtf_at_nyquist"]))
conv_ee = abs(pct_diff(poly_11["ee_1x1"], poly_21["ee_1x1"]))
conv_fwhm = abs(pct_diff(poly_11["fwhm_x_m"], poly_21["fwhm_x_m"]))

print("  Polychromatic convergence (N=11 vs N=21):")
print(f"    MTF at Nyquist:  {conv_mtf:.2f}% difference")
print(f"    EE 1×1:          {conv_ee:.2f}% difference")
print(f"    FWHM:            {conv_fwhm:.2f}% difference")
if max(conv_mtf, conv_ee, conv_fwhm) < 1.0:
    print("    → N=11 is converged (< 1% difference from N=21)")
else:
    print("    → Not yet converged — consider higher N for production runs")
print()

# --- 4e. Assessment ---

print("-" * 80)
print("ASSESSMENT")
print("-" * 80)
print()

if abs(mtf_err) < 5 and abs(ee1_err) < 5:
    print("  For this f/4 MWIR system (3.5–5.0 µm, 18 µm pixel):")
    print("  Monochromatic PSF at band-center gives spatial metrics within ~5%")
    print("  of the polychromatic result. The chromaticism effect is MODEST.")
    print()
    print("  Why modest? The 300 K blackbody source spectrum strongly weights")
    print("  longer wavelengths (more photon flux at 5 µm than 3.5 µm in-band),")
    print("  and the band-center (4.25 µm) is close to the flux-weighted mean.")
else:
    print("  For this f/4 MWIR system (3.5–5.0 µm, 18 µm pixel):")
    print(f"  Monochromatic analysis has significant error: MTF {mtf_err:+.1f}%, EE {ee1_err:+.1f}%.")
    print("  Polychromatic PSF should be used for accurate spatial analysis.")

print()
print("  Recommendation: use polychromatic PSF (N=11) for final design reports.")
print("  Monochromatic is acceptable for quick trades where speed matters.")
print()

# ---------------------------------------------------------------------------
# 5. Generate plots
# ---------------------------------------------------------------------------

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

output_dir = Path(__file__).resolve().parent.parent / "outputs"
output_dir.mkdir(exist_ok=True)

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("Scenario 5.3: Monochromatic vs. Polychromatic PSF\n"
             f"f/{f_number:.0f}, D = {aperture_mm:.0f} mm, {pitch_um:.0f} µm pixel, "
             f"{band_min_um:.1f}–{band_max_um:.1f} µm",
             fontsize=13, fontweight="bold")

# --- Plot 1: Spatial metrics vs. wavelength ---
ax1 = axes[0, 0]
wls = [r["wavelength_um"] for r in per_wavelength_results]
mtfs = [r["mtf_at_nyquist"] for r in per_wavelength_results]
ee1s = [r["ee_1x1"] for r in per_wavelength_results]

ax1.plot(wls, mtfs, "o-", color="tab:blue", linewidth=2, markersize=8, label="MTF @ Nyquist")
ax1.plot(wls, ee1s, "s-", color="tab:red", linewidth=2, markersize=8, label="EE 1×1")

# Overlay polychromatic (N=11) as horizontal bands
ax1.axhline(poly_11["mtf_at_nyquist"], color="tab:blue", linestyle="--", alpha=0.5,
            label=f"Poly N=11 MTF ({poly_11['mtf_at_nyquist']:.3f})")
ax1.axhline(poly_11["ee_1x1"], color="tab:red", linestyle="--", alpha=0.5,
            label=f"Poly N=11 EE ({poly_11['ee_1x1']:.3f})")

ax1.set_xlabel("Wavelength [µm]")
ax1.set_ylabel("Metric Value")
ax1.set_title("Spatial Metrics vs. Wavelength")
ax1.legend(fontsize=8, loc="best")
ax1.grid(True, alpha=0.3)
ax1.set_xlim(3.3, 5.2)

# --- Plot 2: FWHM and Airy disk vs. wavelength ---
ax2 = axes[0, 1]
fwhms = [r["fwhm_x_m"] * 1e6 for r in per_wavelength_results]
airys = [r["airy_diam_um"] for r in per_wavelength_results]

ax2.plot(wls, fwhms, "o-", color="tab:green", linewidth=2, markersize=8, label="FWHM (RADIANT)")
ax2.plot(wls, airys, "^--", color="tab:orange", linewidth=2, markersize=8,
         label="Airy disk Ø (2.44λf/#)")

# Polychromatic FWHM
ax2.axhline(poly_11["fwhm_x_m"] * 1e6, color="tab:green", linestyle="--", alpha=0.5,
            label=f"Poly N=11 FWHM ({poly_11['fwhm_x_m']*1e6:.1f} µm)")

ax2.set_xlabel("Wavelength [µm]")
ax2.set_ylabel("Size [µm]")
ax2.set_title("PSF Size vs. Wavelength")
ax2.legend(fontsize=8, loc="best")
ax2.grid(True, alpha=0.3)
ax2.set_xlim(3.3, 5.2)

# Add pixel pitch reference
ax2.axhline(pitch_um, color="gray", linestyle=":", alpha=0.5)
ax2.text(3.35, pitch_um + 0.5, f"pixel pitch = {pitch_um} µm", fontsize=8, color="gray")

# --- Plot 3: Mono vs. poly bar chart ---
ax3 = axes[1, 0]

labels_bar = ["MTF@Nyq", "EE 1×1", "EE 3×3"]
mono_vals = [mono["mtf_at_nyquist"], mono["ee_1x1"], mono["ee_3x3"]]
poly_vals = [poly_11["mtf_at_nyquist"], poly_11["ee_1x1"], poly_11["ee_3x3"]]

x_bar = np.arange(len(labels_bar))
width = 0.35

bars1 = ax3.bar(x_bar - width/2, mono_vals, width, label="Mono (N=1)", color="tab:blue", alpha=0.8)
bars2 = ax3.bar(x_bar + width/2, poly_vals, width, label="Poly (N=11)", color="tab:orange", alpha=0.8)

# Add error labels
for i, (m, p) in enumerate(zip(mono_vals, poly_vals)):
    err = pct_diff(m, p)
    ax3.text(i, max(m, p) + 0.01, f"{err:+.1f}%", ha="center", fontsize=9, fontweight="bold")

ax3.set_xticks(x_bar)
ax3.set_xticklabels(labels_bar)
ax3.set_ylabel("Metric Value")
ax3.set_title("Mono vs. Poly: Spatial Metrics")
ax3.legend()
ax3.grid(True, alpha=0.3, axis="y")

# --- Plot 4: Convergence with N ---
ax4 = axes[1, 1]

n_vals = [r["n_wavelengths"] for r in fullband_results]
mtf_vals = [r["mtf_at_nyquist"] for r in fullband_results]
ee_vals = [r["ee_1x1"] for r in fullband_results]
fwhm_vals = [r["fwhm_x_m"] * 1e6 for r in fullband_results]

ax4_twin = ax4.twinx()

ax4.plot(n_vals, mtf_vals, "o-", color="tab:blue", linewidth=2, markersize=8, label="MTF@Nyq")
ax4.plot(n_vals, ee_vals, "s-", color="tab:red", linewidth=2, markersize=8, label="EE 1×1")
ax4_twin.plot(n_vals, fwhm_vals, "^-", color="tab:green", linewidth=2, markersize=8,
              label="FWHM [µm]")

ax4.set_xlabel("Number of PSF Wavelengths (N)")
ax4.set_ylabel("MTF / EE")
ax4_twin.set_ylabel("FWHM [µm]", color="tab:green")
ax4.set_title("Polychromatic PSF Convergence")

lines1, labels1 = ax4.get_legend_handles_labels()
lines2, labels2 = ax4_twin.get_legend_handles_labels()
ax4.legend(lines1 + lines2, labels1 + labels2, fontsize=8, loc="center right")
ax4.grid(True, alpha=0.3)
ax4.set_xticks(n_vals)

plt.tight_layout()
plot_path = output_dir / "chromatic_psf_comparison.png"
plt.savefig(plot_path, dpi=150)
print(f"Saved plot: {plot_path}")
plt.close()

# ---------------------------------------------------------------------------
# 6. Write output spreadsheet
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()

# Sheet 1: Per-wavelength results
ws_wl = wb_out.active
ws_wl.title = "Per-Wavelength Results"

headers_wl = ["Wavelength [µm]", "Airy Ø [µm]", "Q", "MTF@Nyquist",
              "EE 1×1", "EE 3×3", "RER", "FWHM [µm]", "SNR", "Signal [e⁻]"]
for col, h in enumerate(headers_wl, 1):
    ws_wl.cell(row=1, column=col, value=h).font = Font(bold=True)

for i, r in enumerate(per_wavelength_results, 2):
    ws_wl.cell(row=i, column=1, value=r["wavelength_um"])
    ws_wl.cell(row=i, column=2, value=round(r["airy_diam_um"], 1))
    ws_wl.cell(row=i, column=3, value=round(r["q_parameter"], 2))
    ws_wl.cell(row=i, column=4, value=round(r["mtf_at_nyquist"], 4))
    ws_wl.cell(row=i, column=5, value=round(r["ee_1x1"], 4))
    ws_wl.cell(row=i, column=6, value=round(r["ee_3x3"], 4))
    ws_wl.cell(row=i, column=7, value=round(r["rer"], 4) if r.get("rer") else "")
    ws_wl.cell(row=i, column=8, value=round(r["fwhm_x_m"] * 1e6, 1))
    ws_wl.cell(row=i, column=9, value=round(r["snr"], 1))
    ws_wl.cell(row=i, column=10, value=round(r["signal_e"]))

# Sheet 2: Mono vs. poly comparison
ws_mp = wb_out.create_sheet("Mono vs Poly")

headers_mp = ["PSF Model", "N wavelengths", "MTF@Nyquist", "EE 1×1", "EE 3×3",
              "RER", "FWHM [µm]", "SNR", "NEDT [K]", "NIIRS", "Signal [e⁻]"]
for col, h in enumerate(headers_mp, 1):
    ws_mp.cell(row=1, column=col, value=h).font = Font(bold=True)

for i, r in enumerate(fullband_results, 2):
    ws_mp.cell(row=i, column=1, value=r["label"])
    ws_mp.cell(row=i, column=2, value=r["n_wavelengths"])
    ws_mp.cell(row=i, column=3, value=round(r["mtf_at_nyquist"], 4))
    ws_mp.cell(row=i, column=4, value=round(r["ee_1x1"], 4))
    ws_mp.cell(row=i, column=5, value=round(r["ee_3x3"], 4))
    ws_mp.cell(row=i, column=6, value=round(r["rer"], 4) if r.get("rer") else "")
    ws_mp.cell(row=i, column=7, value=round(r["fwhm_x_m"] * 1e6, 1))
    ws_mp.cell(row=i, column=8, value=round(r["snr"], 1))
    ws_mp.cell(row=i, column=9, value=round(r["nedt_K"], 4) if r.get("nedt_K") else "")
    ws_mp.cell(row=i, column=10, value=round(r["niirs"], 1) if r.get("niirs") else "")
    ws_mp.cell(row=i, column=11, value=round(r["signal_e"]))

# Add chromaticism error row
err_row = len(fullband_results) + 3
ws_mp.cell(row=err_row, column=1, value="Chromaticism error (mono vs N=11)").font = Font(bold=True)
ws_mp.cell(row=err_row + 1, column=1, value="MTF@Nyquist")
ws_mp.cell(row=err_row + 1, column=2, value=f"{mtf_err:+.1f}%")
ws_mp.cell(row=err_row + 2, column=1, value="EE 1×1")
ws_mp.cell(row=err_row + 2, column=2, value=f"{ee1_err:+.1f}%")
ws_mp.cell(row=err_row + 3, column=1, value="EE 3×3")
ws_mp.cell(row=err_row + 3, column=2, value=f"{ee3_err:+.1f}%")
ws_mp.cell(row=err_row + 4, column=1, value="FWHM")
ws_mp.cell(row=err_row + 4, column=2, value=f"{fwhm_err:+.1f}%")
ws_mp.cell(row=err_row + 5, column=1, value="SNR")
ws_mp.cell(row=err_row + 5, column=2, value=f"{snr_err:+.1f}%")

# Sheet 3: Summary
ws_sum = wb_out.create_sheet("Summary")
ws_sum["A1"] = "Scenario 5.3: Chromatic PSF Analysis Summary"
ws_sum["A1"].font = Font(bold=True, size=14)
ws_sum["A3"] = "System"
ws_sum["B3"] = f"f/{f_number:.0f}, D = {aperture_mm:.0f} mm, {pitch_um:.0f} µm pixel"
ws_sum["A4"] = "Band"
ws_sum["B4"] = f"{band_min_um:.1f} – {band_max_um:.1f} µm (MWIR)"
ws_sum["A5"] = "Q at band center"
ws_sum["B5"] = round(lambda_center_um * f_number / pitch_um, 2)
ws_sum["A6"] = "Q range across band"
ws_sum["B6"] = (f"{analysis_wavelengths_um[0] * f_number / pitch_um:.2f} "
                f"– {analysis_wavelengths_um[-1] * f_number / pitch_um:.2f}")
ws_sum["A8"] = "Chromaticism impact"
ws_sum["A8"].font = Font(bold=True)
ws_sum["A9"] = "MTF error (mono vs poly)"
ws_sum["B9"] = f"{mtf_err:+.1f}%"
ws_sum["A10"] = "EE 1×1 error"
ws_sum["B10"] = f"{ee1_err:+.1f}%"
ws_sum["A11"] = "FWHM error"
ws_sum["B11"] = f"{fwhm_err:+.1f}%"
ws_sum["A13"] = "Recommendation"
ws_sum["A13"].font = Font(bold=True)
if abs(mtf_err) < 5 and abs(ee1_err) < 5:
    ws_sum["A14"] = ("Monochromatic PSF acceptable for quick trades. "
                     "Use polychromatic (N=11) for final design reports.")
else:
    ws_sum["A14"] = "Polychromatic PSF (N≥11) required for accurate spatial analysis."

xlsx_out = output_dir / "chromatic_psf_results.xlsx"
wb_out.save(xlsx_out)
print(f"Saved results: {xlsx_out}")
print()
print("=" * 80)
print("SCENARIO 5.3 COMPLETE")
print("=" * 80)
