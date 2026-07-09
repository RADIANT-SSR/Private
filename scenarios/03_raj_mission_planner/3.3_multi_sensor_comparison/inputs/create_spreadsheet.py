#!/usr/bin/env python3
"""Create Raj's inputs for Scenario 3.3: multi-sensor comparison.

Emits ``raj_sensor_proposals.xlsx`` — three competing MWIR sensor proposals
(the vendors' spec sheets, transcribed to a common tabular form; PDF
parsing is out of scope, so the vendor numbers are captured here) plus the
procurement requirements the compliance matrix checks against.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()

# --- Sensor proposals -------------------------------------------------
ws = wb.active
ws.title = "Proposals"
ws["A1"] = "Scenario 3.3: three competing MWIR sensor proposals"
ws["A1"].font = Font(bold=True, size=14)
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="C55A11")
cols = ["Parameter", "Vendor A", "Vendor B", "Vendor C", "Unit"]
for i, text in enumerate(cols):
    c = ws.cell(row=3, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
for col, w in zip("ABCDE", [26, 12, 12, 12, 10]):
    ws.column_dimensions[col].width = w

ROWS = [
    ("Aperture diameter", 30.0, 25.0, 35.0, "cm"),
    ("f-number", 4.0, 3.0, 5.0, "—"),
    ("Pixel pitch", 18.0, 24.0, 10.0, "µm"),
    ("Band min", 3.7, 3.0, 3.7, "µm"),
    ("Band max", 4.8, 5.0, 4.8, "µm"),
    ("QE", 70.0, 80.0, 65.0, "%"),
    ("Dark current", 200.0, 500.0, 300.0, "e⁻/s"),
    ("Read noise", 25.0, 30.0, 20.0, "e⁻"),
    ("Detector temperature", 80.0, 77.0, 80.0, "K"),
    ("Optical transmission", 82.0, 85.0, 80.0, "%"),
]
r = 4
for name, a, b, c, unit in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=a)
    ws.cell(row=r, column=3, value=b)
    ws.cell(row=r, column=4, value=c)
    ws.cell(row=r, column=5, value=unit)
    r += 1

# --- Requirements (compliance matrix) ---------------------------------
ws2 = wb.create_sheet("Requirements")
for i, text in enumerate(["Requirement", "Threshold", "Direction", "Unit"]):
    c = ws2.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws2.column_dimensions["A"].width = 22
REQS = [
    ("SNR", 50.0, ">=", "—"),
    ("NIIRS", 4.0, ">=", "—"),
    ("NEDT", 50.0, "<=", "mK"),
    ("GSD", 1.5, "<=", "m"),
    ("MTF at Nyquist", 0.25, ">=", "—"),
]
for i, (name, thr, direction, unit) in enumerate(REQS, start=2):
    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=thr)
    ws2.cell(row=i, column=3, value=direction)
    ws2.cell(row=i, column=4, value=unit)

# --- Common operating point -------------------------------------------
ws3 = wb.create_sheet("Operating")
for i, text in enumerate(["Parameter", "Value", "Unit"]):
    c = ws3.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws3.column_dimensions["A"].width = 24
for i, (name, val, unit) in enumerate(
    [
        ("Altitude", 600.0, "km"),
        ("Scene temperature", 300.0, "K"),
        ("Scene emissivity", 0.95, "—"),
        ("Integration time", 8.0, "ms"),
        ("Cross-track pixels", 4096, "—"),
    ],
    start=2,
):
    ws3.cell(row=i, column=1, value=name)
    ws3.cell(row=i, column=2, value=val)
    ws3.cell(row=i, column=3, value=unit)

out = HERE / "raj_sensor_proposals.xlsx"
wb.save(out)
print(f"Wrote {out}")
