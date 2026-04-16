"""Create Karen's lab data spreadsheet for MTF measurement vs. prediction.

Four sheets:
  1. "System Configuration" — as-built optics and detector parameters
  2. "Measured MTF"         — slanted-edge MTF (50 points, spatial_freq in cy/mm)
  3. "As-Built WFE"         — measured wavefront error
  4. "Focus Position"       — measured defocus from best focus
"""

import math

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Sheet 1: System Configuration
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "System Configuration"
ws1["A1"] = "Scenario 7.3: MTF Measurement vs. Prediction"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "As-built VNIR pushbroom sensor — lab slanted-edge MTF test"
ws1.column_dimensions["A"].width = 40
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 50

# -- Optics section
row = 4
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Optics", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

optics_data = [
    ("Aperture diameter", 20.0, "cm", "As-built, measured on optical bench"),
    ("Effective focal length", 60.0, "cm", "Measured via collimator (nominal 60 cm)"),
    ("f-number", 3.0, "—", "Derived from D and f"),
    ("Optical transmission", 82.0, "%", "Measured at 650 nm, includes all elements"),
    ("Optics temperature", 22.0, "°C", "Lab ambient"),
    ("Central obscuration", 25.0, "%", "Secondary mirror (Cassegrain)"),
    ("Number of elements", 3, "—", "Primary, secondary, corrector"),
]

for i, (param, val, unit, note) in enumerate(optics_data, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# -- Spectral section
row = 13
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Spectral Band", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

spec_data = [
    ("Filter min", 550.0, "nm", "VNIR band"),
    ("Filter max", 750.0, "nm", "VNIR band"),
    ("MTF test wavelength", 650.0, "nm", "Quasi-monochromatic collimator source"),
]

for i, (param, val, unit, note) in enumerate(spec_data, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# -- Detector section
row = 18
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Detector", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

det_data = [
    ("Pixel pitch", 10.0, "µm", "Square pixels, silicon CCD"),
    ("Fill factor", 100.0, "%", "Full-frame CCD"),
    ("Quantum efficiency", 85.0, "%", "Measured at 650 nm"),
    ("Dark current", 50.0, "e⁻/s", "At 263 K operating temperature"),
    ("Operating temperature", 263.0, "K", "TEC-cooled"),
    ("Read noise", 8.0, "e⁻ RMS", "CDS readout"),
    ("Full well capacity", 120000.0, "e⁻", "90% linearity"),
    ("ADC bits", 14, "—", ""),
    ("System gain", 8.0, "e⁻/DN", ""),
    ("IPC coupling", 1.0, "%", "Measured via single-pixel illumination"),
]

for i, (param, val, unit, note) in enumerate(det_data, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# -- Test Setup section
row = 30
for c in range(1, 5):
    cell = ws1.cell(row=row, column=c, value=["Test Setup", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

test_data = [
    ("Collimator focal length", 3.0, "m", "Off-axis parabola, λ/20 surface"),
    ("Target type", "slanted edge", "—", "4° tilt angle, knife-edge on backlit aperture"),
    ("Integration time", 2.0, "ms", "Lab MTF test frame rate"),
    ("Number of frames averaged", 100, "—", "Reduces noise in ESF measurement"),
]

for i, (param, val, unit, note) in enumerate(test_data, row + 1):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    ws1.cell(row=i, column=2, value=val).border = thin_border
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=note).border = thin_border

# ---------------------------------------------------------------------------
# Sheet 2: Measured MTF — 50 points from slanted-edge analysis
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Measured MTF")
ws2["A1"] = "Slanted-Edge MTF Measurement — Cross-Track Axis"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Method: ISO 12233 slanted-edge, 4° tilt, 100-frame average"
ws2["A3"] = "Date: 2026-04-10, Operator: K. Martinez, Setup: Optical bench #2"
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 18

row = 5
ws2.cell(row=row, column=1, value="spatial_frequency_cy_mm").font = header_font
ws2.cell(row=row, column=1).border = thin_border
ws2.cell(row=row, column=2, value="MTF_measured").font = header_font
ws2.cell(row=row, column=2).border = thin_border

# Generate realistic measured MTF data.
# System: 20 cm aperture, f/3, 10 um pixels, 650 nm, with 0.07 waves WFE + defocus.
# Nyquist = 1/(2*10e-6) = 50000 cy/m = 50 cy/mm.
# Diffraction cutoff = 1/(650e-9 * 3) = 512820 cy/m = 512.8 cy/mm.
# We'll generate 50 points from 0 to 100 cy/mm (2× Nyquist).

pixel_pitch_m = 10.0e-6
f_number = 3.0
wavelength_m = 650.0e-9
f_nyquist_cy_m = 1.0 / (2.0 * pixel_pitch_m)  # 50000 cy/m
f_cutoff_cy_m = 1.0 / (wavelength_m * f_number)  # ~512820 cy/m

n_points = 50
f_max_cy_mm = 100.0  # 2× Nyquist in cy/mm
freq_cy_mm = np.linspace(0.0, f_max_cy_mm, n_points)
freq_cy_m = freq_cy_mm * 1000.0  # convert to cy/m

# Compute "true" MTF as product of components for the as-built system:
# 1. Diffraction MTF (circular aperture with obscuration)
#    Using the Besinc approximation: MTF_diff ≈ MTF of unobscured * correction
#    For simplicity, use the standard circular aperture MTF:
#    MTF(f) = (2/pi)[arccos(f/fc) - (f/fc)*sqrt(1-(f/fc)^2)] for f < fc
f_norm = freq_cy_m / f_cutoff_cy_m
mtf_diff = np.zeros_like(f_norm)
valid = f_norm < 1.0
mtf_diff[valid] = (2.0 / math.pi) * (
    np.arccos(f_norm[valid]) - f_norm[valid] * np.sqrt(1.0 - f_norm[valid] ** 2)
)
mtf_diff[freq_cy_m == 0] = 1.0

# 2. WFE degradation (approximate as Strehl-weighted, Marechal):
#    MTF_wfe ≈ Strehl * MTF_diff + (1 - Strehl) * halo
#    For simplicity, scale by a WFE factor ~ exp(-k*f^2)
wfe_rms_waves = 0.07  # at 633 nm
opd_rms_m = wfe_rms_waves * 633.0e-9
# Shannon's approximation: MTF_aberrated/MTF_diff ≈ 1 - (2*pi*OPD/lambda)^2 * H(f)
# Use a simpler empirical factor
strehl = math.exp(-(2.0 * math.pi * opd_rms_m / wavelength_m) ** 2)
# Mix: at low freq MTF ≈ MTF_diff, at high freq degraded by WFE
wfe_factor = strehl + (1.0 - strehl) * np.exp(-0.5 * (freq_cy_m / (0.3 * f_cutoff_cy_m)) ** 2)

# 3. Pixel aperture MTF: |sinc(pi * f * p)|
mtf_pixel = np.abs(np.sinc(freq_cy_m * pixel_pitch_m))

# 4. Electronics/diffusion MTF: exp(-2*pi^2*sigma^2*f^2), sigma ~ 2 um
sigma_elec_m = 2.0e-6
mtf_elec = np.exp(-2.0 * math.pi**2 * sigma_elec_m**2 * freq_cy_m**2)

# 5. Defocus MTF: additional Gaussian blur from 5 um defocus
#    Defocus spot radius ≈ defocus / (2 * f_number^2) — geometric
defocus_um = 5.0
defocus_m = defocus_um * 1e-6
# Geometric blur radius for defocus:
spot_radius_m = defocus_m / (2.0 * f_number)
# Equivalent Gaussian sigma ≈ spot_radius / 2
sigma_defocus_m = spot_radius_m / 2.0
mtf_defocus = np.exp(-2.0 * math.pi**2 * sigma_defocus_m**2 * freq_cy_m**2)

# System MTF = product of all components
mtf_system = mtf_diff * wfe_factor * mtf_pixel * mtf_elec * mtf_defocus

# Add measurement noise (realistic: ~0.5-2% MTF noise from slanted-edge method)
rng = np.random.default_rng(seed=42)
noise_level = 0.015  # 1.5% RMS
mtf_measured = mtf_system + rng.normal(0.0, noise_level, size=n_points)
mtf_measured = np.clip(mtf_measured, 0.0, 1.0)
mtf_measured[0] = 1.0  # DC is always 1.0 by definition

for i, (f_val, m_val) in enumerate(zip(freq_cy_mm, mtf_measured), row + 1):
    ws2.cell(row=i, column=1, value=round(float(f_val), 3)).border = thin_border
    ws2.cell(row=i, column=2, value=round(float(m_val), 4)).border = thin_border

# Add notes
note_row = row + n_points + 2
ws2.cell(row=note_row, column=1, value="Notes:").font = Font(bold=True)
ws2.cell(row=note_row + 1, column=1,
         value="Frequency is in cycles/mm on the focal plane.")
ws2.cell(row=note_row + 2, column=1,
         value="MTF normalized to 1.0 at DC (zero frequency).")
ws2.cell(row=note_row + 3, column=1,
         value="Nyquist frequency = 50 cy/mm (for 10 um pixels).")
ws2.cell(row=note_row + 4, column=1,
         value="Data extends to 2× Nyquist (100 cy/mm) for aliasing analysis.")

# ---------------------------------------------------------------------------
# Sheet 3: As-Built WFE
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("As-Built WFE")
ws3["A1"] = "As-Built Wavefront Error — Interferometric Measurement"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "4D Technology PhaseCam 6110, 633 nm HeNe"
ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 50

row = 4
for c in range(1, 5):
    cell = ws3.cell(row=row, column=c,
                    value=["Parameter", "Value", "Unit", "Notes"][c - 1])
    cell.font = header_font
    cell.border = thin_border

wfe_data = [
    ("Total WFE RMS", 0.07, "waves", "At 633 nm HeNe reference"),
    ("WFE reference wavelength", 633.0, "nm", "HeNe laser"),
    ("WFE P-V", 0.35, "waves", "Peak-to-valley at 633 nm"),
    ("Dominant aberration", "trefoil", "—", "Likely from primary mirror mount stress"),
    ("Measurement uncertainty (2σ)", 0.005, "waves", "Repeatability over 10 measurements"),
    ("Measurement temperature", 22.0, "°C", "Lab ambient during interferometric test"),
    ("Measurement date", "2026-04-08", "—", ""),
]

for i, (param, val, unit, note) in enumerate(wfe_data, row + 1):
    ws3.cell(row=i, column=1, value=param).border = thin_border
    ws3.cell(row=i, column=2, value=val).border = thin_border
    ws3.cell(row=i, column=3, value=unit).border = thin_border
    ws3.cell(row=i, column=4, value=note).border = thin_border

# ---------------------------------------------------------------------------
# Sheet 4: Focus Position
# ---------------------------------------------------------------------------
ws4 = wb.create_sheet("Focus Position")
ws4["A1"] = "Focus Position Measurement"
ws4["A1"].font = Font(bold=True, size=14)
ws4["A2"] = "Through-focus MTF scan at 650 nm"
ws4.column_dimensions["A"].width = 35
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 50

row = 4
for c in range(1, 5):
    cell = ws4.cell(row=row, column=c,
                    value=["Parameter", "Value", "Unit", "Notes"][c - 1])
    cell.font = header_font
    cell.border = thin_border

focus_data = [
    ("Defocus from best focus", 5.0, "µm", "Measured via through-focus scan"),
    ("Best focus position", 600.023, "mm", "Micrometer reading at peak MTF"),
    ("Current focus position", 600.028, "mm", "Operating position (5 µm beyond best)"),
    ("Focus measurement method", "through-focus MTF", "—", "Peak of MTF@Nyquist vs. focus curve"),
    ("Focus step size", 1.0, "µm", "Micrometer resolution"),
    ("Focus uncertainty (2σ)", 2.0, "µm", "Estimated from repeated measurements"),
]

for i, (param, val, unit, note) in enumerate(focus_data, row + 1):
    ws4.cell(row=i, column=1, value=param).border = thin_border
    ws4.cell(row=i, column=2, value=val).border = thin_border
    ws4.cell(row=i, column=3, value=unit).border = thin_border
    ws4.cell(row=i, column=4, value=note).border = thin_border

# Defocus sweep for sensitivity analysis
row = 12
for c in range(1, 5):
    cell = ws4.cell(row=row, column=c,
                    value=["Defocus Sweep", "", "", ""][c - 1])
    cell.fill = section_fill
    cell.font = section_font

ws4.cell(row=row + 1, column=1, value="Defocus [µm]").font = header_font
ws4.cell(row=row + 1, column=1).border = thin_border

defocus_values = [0.0, 1.0, 2.0, 3.0, 5.0, 7.0, 10.0, 15.0, 20.0]
for i, d in enumerate(defocus_values, row + 2):
    ws4.cell(row=i, column=1, value=d).border = thin_border

from pathlib import Path
out = Path(__file__).parent / "karen_mtf_lab_data.xlsx"
wb.save(out)
print(f"Created {out}")
