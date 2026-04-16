"""Create Dr. Chen's sensor parameter spreadsheet for noise model verification."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

# --- Sheet 1: Sensor Parameters ---
ws = wb.active
ws.title = "Sensor Parameters"

header_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 45

# Title
ws["A1"] = "Noise Model Verification — Sensor Parameters"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Dr. Chen — April 2026"
ws["A2"].font = Font(italic=True, size=10)

# Headers
for col, header in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 5


def add_section(ws, row, title):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=value).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value=unit).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# Optics
row = add_section(ws, row, "Optics")
row = add_param(ws, row, "Aperture diameter", 30, "cm", "Customer spec — circular, unobscured")
row = add_param(ws, row, "Focal ratio (f/#)", 4, "—", "f/4 design")
row = add_param(ws, row, "Optical transmission", 70, "%", "End-to-end, includes all elements")
row = add_param(ws, row, "Number of optical elements", 4, "—", "2 mirrors + 2 lenses")
row = add_param(ws, row, "Optics temperature", 293, "K", "Room temperature bench test")

# Detector
row = add_section(ws, row, "Detector")
row = add_param(ws, row, "Pixel pitch", 18, "µm", "Square pixels")
row = add_param(ws, row, "Quantum efficiency", 70, "%", "Spec sheet average over band")
row = add_param(ws, row, "Dark current", 100, "e⁻/s", "At 77 K operating temp")
row = add_param(ws, row, "Operating temperature", 77, "K", "LN2 cooled")
row = add_param(ws, row, "Full well capacity", 2.0e6, "e⁻", "Large-format InSb")

# Readout
row = add_section(ws, row, "Readout Electronics")
row = add_param(ws, row, "Read noise", 20, "e⁻ RMS", "After CDS")
row = add_param(ws, row, "ADC bits", 14, "bits", "")
row = add_param(ws, row, "System gain", 1.0, "e⁻/DN", "Unity gain assumed")

# Scene
row = add_section(ws, row, "Scene / Target")
row = add_param(ws, row, "Target temperature", 300, "K", "Blackbody calibration source")
row = add_param(ws, row, "Target emissivity", 0.95, "—", "Near-ideal blackbody")
row = add_param(ws, row, "Background temperature", 290, "K", "Lab ambient wall")
row = add_param(ws, row, "Background emissivity", 0.95, "—", "Painted wall")

# Spectral
row = add_section(ws, row, "Spectral / Integration")
row = add_param(ws, row, "Band minimum", 3.5, "µm", "Cold filter cut-on")
row = add_param(ws, row, "Band maximum", 5.0, "µm", "Cold filter cut-off")
row = add_param(ws, row, "Integration time", 5, "ms", "Single frame")
row = add_param(ws, row, "Number of TDI stages", 1, "—", "Staring mode, no TDI")

# Geometry
row = add_section(ws, row, "Geometry")
row = add_param(ws, row, "Sensor altitude", 8, "km", "Airborne platform")
row = add_param(ws, row, "Look angle", 0, "deg", "Nadir")


# --- Sheet 2: Hand Calculations ---
ws2 = wb.create_sheet("Hand Calculations")

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 15
ws2.column_dimensions["F"].width = 40

ws2["A1"] = "Noise Model Verification — Hand Calculations"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Fill in 'Hand Calc' column, then compare to RADIANT output"
ws2["A2"].font = Font(italic=True, size=10)

for col, header in enumerate(
    ["Noise Term", "Equation", "Hand Calc (e⁻)", "RADIANT (e⁻)", "% Error", "Notes"], 1
):
    cell = ws2.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

noise_terms = [
    ("Signal shot noise", "sqrt(S_target)", "", "Fill from result.noise_terms"),
    ("Background shot noise", "sqrt(S_background)", "", "Scene background photons"),
    ("Dark current shot noise", "sqrt(J_dark × t_int)", "", ""),
    ("Read noise", "σ_read (given)", "", "Fixed value from spec"),
    ("Quantization noise", "gain / sqrt(12)", "", "ADC quantization"),
    ("Nearfield emission shot", "sqrt(S_nearfield)", "", "Warm optics self-emission"),
    ("1/f noise", "See noise model doc", "", "May be zero if not configured"),
    ("Clock-induced charge", "CIC × n_reads", "", "May be zero for InSb"),
    ("ADC nonlinearity", "See model", "", ""),
    ("Total noise (RSS)", "sqrt(Σ σ_i²)", "", "Root sum of squares"),
]

row = 5
for name, equation, hand_calc, notes in noise_terms:
    ws2.cell(row=row, column=1, value=name).border = thin_border
    ws2.cell(row=row, column=2, value=equation).border = thin_border
    ws2.cell(row=row, column=3, value=hand_calc).border = thin_border
    ws2.cell(row=row, column=4, value="").border = thin_border
    # % Error formula (only if both C and D have values)
    ws2.cell(
        row=row,
        column=5,
        value=f'=IF(AND(C{row}<>"",D{row}<>""), ABS(C{row}-D{row})/C{row}*100, "")',
    ).border = thin_border
    ws2.cell(row=row, column=5).number_format = "0.00"
    ws2.cell(row=row, column=6, value=notes).border = thin_border
    row += 1

# Derived quantities section
row += 1
ws2.cell(row=row, column=1, value="Derived Quantities").font = header_font
row += 1
derived = [
    ("Focal length", "f/# × D_aperture", "m"),
    ("Pixel IFOV", "p_pitch / f", "µrad"),
    ("Pixel solid angle (Ω)", "IFOV²", "sr"),
    ("Collecting area", "π/4 × D²", "m²"),
    ("f/# (Ω form)", "π / (4 × f/#²)", "sr"),
    ("Signal electrons (target)", "L_target × τ_optics × Ω × A_pix × QE × t_int / E_photon", "e⁻"),
]
for name, equation, unit in derived:
    ws2.cell(row=row, column=1, value=name).border = thin_border
    ws2.cell(row=row, column=2, value=equation).border = thin_border
    ws2.cell(row=row, column=3, value=unit).border = thin_border
    row += 1

output_path = "dr_chen_sensor_parameters.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
