#!/usr/bin/env python3
"""Create Tom's inputs for Scenario 5.5: stray-light / veiling-glare analysis.

Emits ``tom_straylight.xlsx`` — the daytime VNIR scene (two-reflectance
target/background), the FRED-derived stray-light numbers (veiling glare
index and out-of-field absolute irradiance), and the sensor config.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="264478")

ws = wb.active
ws.title = "Scene"
ws["A1"] = "Scenario 5.5: daytime VNIR scene + FRED stray light"
ws["A1"].font = Font(bold=True, size=14)
for i, text in enumerate(["Surface", "Reflectance", "Note"]):
    c = ws.cell(row=3, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
for col, w in zip("ABC", [16, 14, 34], strict=True):
    ws.column_dimensions[col].width = w
for r, (name, rho, note) in enumerate(
    [
        ("Target (rooftop)", 0.30, "bright man-made surface"),
        ("Background (veg.)", 0.15, "vegetation / soil"),
    ],
    start=4,
):
    for col, val in zip("ABC", [name, rho, note], strict=True):
        ws[f"{col}{r}"] = val

ws2 = wb.create_sheet("StrayLight")
for i, text in enumerate(["Quantity", "Value", "Unit"]):
    c = ws2.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws2.column_dimensions["A"].width = 30
for i, (name, val, unit) in enumerate(
    [
        ("Veiling glare index (FRED)", 0.03, "fraction"),
        ("Out-of-field stray irradiance", 2.5, "W/m²"),
        ("VGI sweep range", "0-10", "%"),
        ("Full stray-light PSF", "2D array (FRED)", "not ingestable — see gaps.md"),
    ],
    start=2,
):
    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=unit)

ws3 = wb.create_sheet("Sensor")
for i, text in enumerate(["Parameter", "Value", "Unit"]):
    c = ws3.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws3.column_dimensions["A"].width = 24
for i, (name, val, unit) in enumerate(
    [
        ("Waveband", "0.5-0.8", "µm (VNIR pan)"),
        ("Aperture diameter", 15.0, "cm"),
        ("Focal length", 0.9, "m (f/6)"),
        ("Pixel pitch", 15.0, "µm"),
        ("QE", 60.0, "%"),
        ("Dark current", 2.0e4, "e⁻/s"),
        ("Detector temperature", 290.0, "K"),
        ("Integration time", 5.0, "ms"),
        ("Read noise", 30.0, "e⁻"),
        ("Full well", 3.0e5, "e⁻"),
        ("Sensor altitude", 7.0, "km"),
        ("Solar zenith", 30.0, "deg"),
    ],
    start=2,
):
    ws3.cell(row=i, column=1, value=name)
    ws3.cell(row=i, column=2, value=val)
    ws3.cell(row=i, column=3, value=unit)

out = HERE / "tom_straylight.xlsx"
wb.save(out)
print(f"Wrote {out}")
