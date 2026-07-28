"""Generate the vendor-format inputs for scenario 10.3 (ground-to-space SST, visible).

Two artifacts, both in the units a real SST site would hand an analyst — NOT
RADIANT canonical units.  The runner
(``scripts/run_ground_to_space_sst_visible.py``) is the single place the
vendor → canonical conversions happen, each with an explicit comment.

1. ``sst_site_and_tasking.xlsx`` — the site/telescope datasheet plus the
   night's tasking card.  Vendor units: mm, %, µm, ms, km, degrees, ke-.

2. ``object_signature_ORB-4471.csv`` — the catalogue "signature file" for the
   tracked object: spectral radiant intensity in **W/sr/nm** against
   wavelength in **nm**, the convention an optical-signatures group publishes.

Why a signature file and not an albedo + shape entry: RADIANT's point-source
door is radiant intensity ``I(λ)`` (``source.target.user_intensity_path`` →
``T7IntensityAtSource``, ADR-0004).  There is no reflective point-source door
that takes (albedo, projected area, solar geometry) and derives ``I(λ)``
internally — see ``gaps.md`` G1.  So the reflective physics is done here, in
the "vendor tool", exactly as a signatures group would do it:

    I(λ) = ρ · A_proj · E_sun(λ) · p(α) / π            [W/sr/µm]

with

* ``ρ``      — diffuse (Lambertian) albedo of the object [dimensionless],
* ``A_proj`` — projected area toward the observer [m²],
* ``E_sun``  — top-of-atmosphere solar spectral irradiance at 1 AU
               [W/m²/µm]; the object is above the terminator shadow so it sees
               the unattenuated beam,
* ``p(α)``   — the diffuse-sphere phase function at solar phase angle α,
               ``p(α) = [sin α + (π − α) cos α] / π`` — 1 at full phase
               (α = 0) and 0 at α = π,
* ``1/π``    — the Lambertian conversion from irradiance to radiance.

``E_sun(λ)`` is taken from ``radiant.core.solar`` (a 5778 K Planck shape
normalised to the solar constant S₀ = 1361 W/m² at 1 AU).  That import is a
*data* dependency of the fake vendor tool, not a RADIANT modelling path: none
of the chain arithmetic the runner cross-checks touches it, because the
intensity door consumes the tabulated ``I(λ)`` verbatim.

Run:  python create_spreadsheet.py
"""

from __future__ import annotations

import math
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from radiant.core.solar import toa_solar_spectral_irradiance

HERE = Path(__file__).resolve().parent

WORKBOOK = HERE / "sst_site_and_tasking.xlsx"
SIGNATURE_CSV = HERE / "object_signature_ORB-4471.csv"

# --- Object definition (the signatures group's model of ORB-4471) -----------
OBJECT_NAME = "ORB-4471"
OBJECT_ALBEDO = 0.25  # dimensionless diffuse albedo (aged white-paint / MLI mix)
OBJECT_PROJECTED_AREA_M2 = 1.00  # m², bus + one deployed array face, sun-facing
SOLAR_PHASE_ANGLE_DEG = 35.0  # deg, sun-object-observer angle at the tasked pass

# Signature grid: 380–950 nm at 5 nm, comfortably outside the 400–900 nm band
# the telescope filter defines, so the interpolation never extrapolates.
SIGNATURE_LO_NM = 380.0
SIGNATURE_HI_NM = 950.0
SIGNATURE_STEP_NM = 5.0

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)


def diffuse_sphere_phase_function(alpha_rad: float) -> float:
    """Lambertian-sphere phase function ``p(α) = [sin α + (π − α) cos α] / π``.

    ``p(0) = 1`` (full phase) and ``p(π) = 0`` (new phase).  This is the
    standard photometric model an SST catalogue uses when the attitude is
    unknown; it is the only aspect model in the signature file.
    """
    return (math.sin(alpha_rad) + (math.pi - alpha_rad) * math.cos(alpha_rad)) / math.pi


def build_signature_csv() -> None:
    """Write the vendor signature file: wavelength [nm], I(λ) [W/sr/nm]."""
    lam_nm = np.arange(SIGNATURE_LO_NM, SIGNATURE_HI_NM + 0.5 * SIGNATURE_STEP_NM, SIGNATURE_STEP_NM)
    lam_um = lam_nm / 1000.0  # nm → µm, only so the solar model can be sampled
    e_sun_per_um = toa_solar_spectral_irradiance(lam_um)  # W/m²/µm at 1 AU
    phase = diffuse_sphere_phase_function(math.radians(SOLAR_PHASE_ANGLE_DEG))

    # I(λ) = ρ · A · E_sun(λ) · p(α) / π   [W/sr/µm]
    intensity_per_um = OBJECT_ALBEDO * OBJECT_PROJECTED_AREA_M2 * e_sun_per_um * phase / math.pi
    intensity_per_nm = intensity_per_um / 1000.0  # W/sr/µm → W/sr/nm (vendor unit)

    lines = [
        f"# {OBJECT_NAME} optical signature — diffuse-sphere model",
        f"# albedo = {OBJECT_ALBEDO:.3f} [dimensionless]",
        f"# projected_area = {OBJECT_PROJECTED_AREA_M2:.3f} [m^2]",
        f"# solar_phase_angle = {SOLAR_PHASE_ANGLE_DEG:.1f} [deg]",
        f"# phase_function p(alpha) = {phase:.6f} [dimensionless]",
        "# columns: wavelength [nm], spectral radiant intensity [W/sr/nm]",
        "wavelength_nm,intensity_W_per_sr_per_nm",
    ]
    lines += [f"{w:.1f},{v:.8e}" for w, v in zip(lam_nm, intensity_per_nm, strict=True)]
    SIGNATURE_CSV.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"Wrote {SIGNATURE_CSV.name}: {lam_nm.size} rows, {lam_nm[0]:.0f}–{lam_nm[-1]:.0f} nm")


def _write_table(ws, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "All values in VENDOR units — the runner converts to RADIANT canonical units."
    ws["A3"].font = Font(italic=True, size=9)
    for col, head in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=head)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=5):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    widths = [42, 16, 18, 62]
    for i, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook() -> None:
    """Write the site/telescope datasheet and the night's tasking card."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Site & Telescope"
    _write_table(
        ws,
        "SST Site 'Dry Lake' — 1 m tracking telescope, visible channel",
        ["Parameter", "Value", "Unit", "Note"],
        [
            ["Site elevation MSL", 900.0, "m", "Below the 1 km ground/air band edge"],
            ["Entrance Pupil Diameter", 1000.0, "mm", "Unobscured equivalent aperture"],
            ["Effective Focal Length", 10000.0, "mm", "f/10 Ritchey-Chretien"],
            ["Optical Transmission", 60.0, "%", "3 aluminised mirrors + window + filter"],
            ["Filter Band Low", 400.0, "nm", "Broad visible/NIR clear filter"],
            ["Filter Band High", 900.0, "nm", "Detector red cut-off"],
            ["Pixel Pitch", 15.0, "um", "Back-illuminated CCD, square pixel"],
            ["Quantum Efficiency", 80.0, "%", "Band-averaged, scalar"],
            ["Dark Current", 100.0, "e-/s", "Thermo-electrically cooled to -40 C"],
            ["Read Noise", 5.0, "e- RMS", "Slow-scan readout"],
            ["Full Well Capacity", 400.0, "ke-", "Per pixel"],
            ["Gain", 10.0, "e-/DN", ""],
            ["ADC Resolution", 16, "bits", ""],
            ["Exposure Time", 5.0, "ms", "Rate-matched track; LEO transit limits it"],
        ],
    )

    ws = wb.create_sheet("Tasking & Geometry")
    _write_table(
        ws,
        "Tasking card — pass of ORB-4471, evening terminator window",
        ["Parameter", "Value", "Unit", "Note"],
        [
            ["Object Altitude", 700.0, "km", "Catalogue mean altitude at culmination"],
            ["Pointing Zenith Angle", 20.0, "deg", "Measured at the TELESCOPE (lower endpoint)"],
            ["Solar Depression", 12.0, "deg", "Sun below the site horizon (nautical twilight)"],
            ["Solar Relative Azimuth", 45.0, "deg", "Sun azimuth minus pointing azimuth"],
            ["Meteorological Visibility", 100.0, "km", "Exceptional dry-air night"],
            ["Standard Atmosphere", "midlat_summer", "-", "Climate profile for the column"],
            ["Sun-Up Comparison Depression", -10.0, "deg", "Negative = sun ABOVE the horizon"],
        ],
    )

    ws = wb.create_sheet("Object Catalog")
    _write_table(
        ws,
        "Catalogue entry — the object being tracked",
        ["Parameter", "Value", "Unit", "Note"],
        [
            ["Object Designator", OBJECT_NAME, "-", "Signature file name key"],
            ["Diffuse Albedo", OBJECT_ALBEDO, "-", "Dimensionless, 0-1"],
            ["Projected Area", OBJECT_PROJECTED_AREA_M2, "m^2", "Toward the observer"],
            ["Solar Phase Angle", SOLAR_PHASE_ANGLE_DEG, "deg", "Sun-object-observer"],
            ["Signature File", SIGNATURE_CSV.name, "-", "wavelength [nm], I [W/sr/nm]"],
        ],
    )

    ws = wb.create_sheet("Reflective Door Object")
    _write_table(
        ws,
        "Second tasking — a GEO object, entered as shape + albedo instead of a signature file",
        ["Parameter", "Value", "Unit", "Note"],
        [
            ["Object Designator", "GEO-2210", "-", "Geostationary comsat"],
            ["Object Altitude", 35786.0, "km", "Geostationary radius minus R_E"],
            ["Diffuse Albedo", 0.20, "-", "Solar-array-dominated"],
            ["Projected Area", 10.0, "m^2", "Bus + array face toward the observer"],
            ["Pointing Zenith Angle", 30.0, "deg", "Typical GEO-belt search elevation"],
            ["Daylight Solar Zenith", 60.0, "deg", "Sun ABOVE the site horizon"],
        ],
    )

    ws = wb.create_sheet("Seeing")
    _write_table(
        ws,
        "Site seeing model — Hufnagel-Valley Cn2 profile",
        ["Parameter", "Value", "Unit", "Note"],
        [
            ["Cn2 Profile", "hufnagel_valley", "-", "HV-5/7 parameterisation"],
            ["High-Altitude Wind RMS", 21.0, "m/s", "The 'w' of HV-5/7"],
            ["Ground Turbulence Strength", 1.7e-14, "m^(-2/3)", "The 'A' of HV-5/7"],
            ["Wave Type", "plane", "-", "Distant source; astronomical convention"],
        ],
    )

    ws = wb.create_sheet("Zenith Ladder")
    _write_table(
        ws,
        "Pointing-zenith ladder — the pass from culmination to low elevation",
        ["Pointing Zenith Angle", "Unit", "Note"],
        [
            [0.0, "deg", "Culmination (object at the site zenith)"],
            [20.0, "deg", "Nominal tasking point"],
            [40.0, "deg", ""],
            [55.0, "deg", ""],
            [65.0, "deg", ""],
            [75.0, "deg", "Acquisition / loss-of-track elevation"],
        ],
    )

    ws = wb.create_sheet("Air Mass Probe")
    _write_table(
        ws,
        "Fine zenith probe across the 80 deg air-mass handover in the simple model",
        ["Pointing Zenith Angle", "Unit", "Note"],
        [
            [76.0, "deg", "Flat-Earth branch"],
            [78.0, "deg", "Flat-Earth branch"],
            [79.9, "deg", "Last sample below the switch"],
            [80.1, "deg", "First sample above the switch"],
            [82.0, "deg", "Spherical branch"],
            [85.0, "deg", "Spherical branch"],
        ],
    )

    ws = wb.create_sheet("Terminator Ladder")
    _write_table(
        ws,
        "Solar-depression ladder — the shadow-height test (GF-9)",
        ["Solar Depression", "Unit", "Note"],
        [
            [0.0, "deg", "Sun on the site horizon"],
            [3.0, "deg", ""],
            [6.0, "deg", "End of civil twilight"],
            [9.0, "deg", ""],
            [12.0, "deg", "End of nautical twilight — nominal tasking"],
            [15.0, "deg", ""],
            [18.0, "deg", "End of astronomical twilight"],
            [22.0, "deg", "Deep night"],
            [26.0, "deg", "Shadow height crosses the object altitude"],
            [30.0, "deg", "Object fully eclipsed — no reflected signal"],
        ],
    )

    wb.save(WORKBOOK)
    print(f"Wrote {WORKBOOK.name}: {len(wb.sheetnames)} sheets ({', '.join(wb.sheetnames)})")


def main() -> None:
    build_signature_csv()
    build_workbook()


if __name__ == "__main__":
    main()
