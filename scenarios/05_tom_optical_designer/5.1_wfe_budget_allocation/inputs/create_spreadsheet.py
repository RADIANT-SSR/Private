#!/usr/bin/env python3
"""Create the input spreadsheet for Scenario 5.1: WFE Budget Allocation.

Tom has a Zernike decomposition from Zemax for a 40 cm VNIR telescope.
He wants to see the impact of WFE RMS on Strehl, MTF, EE, RER, and NIIRS.

Also emits tom_zernike_zemax.txt — the same coefficients in the Zemax
"Zernike Standard Coefficients" text-export format that
radiant.io.zemax_zernike.load_zemax_zernike parses (Gap 26). The workbook
sheet is retained for human review.

Run:  python create_spreadsheet.py
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

wb = openpyxl.Workbook()

# -- Sheet 1: Optical System ------------------------------------------------
ws = wb.active
ws.title = "Optical System"
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 55

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
note_font = Font(italic=True, color="666666")

for col in ["A", "B", "C", "D"]:
    ws[f"{col}1"] = ""
    ws[f"{col}1"].font = header_font
    ws[f"{col}1"].fill = header_fill
    ws[f"{col}1"].alignment = Alignment(horizontal="center")

ws["A1"] = "Parameter"
ws["B1"] = "Value"
ws["C1"] = "Unit"
ws["D1"] = "Notes"

rows = [
    # Optics
    ("Aperture diameter",       40,      "cm",    "40 cm Cassegrain primary"),
    ("Focal length",            400,     "cm",    "f/10"),
    ("f-number",                10.0,    "--",    "Derived: f/D"),
    ("Optical transmission",    75,      "%",     "End-to-end"),
    ("Optics temperature",      20,      "C",     "Ambient"),
    ("Central obscuration",     35,      "%",     "Linear, Cassegrain secondary"),
    # WFE
    ("WFE reference wavelength", 633,    "nm",    "HeNe laser interferometry"),
    # Detector
    ("Pixel pitch",             10.0,    "um",    "Si CCD"),
    ("Quantum efficiency",      85,      "%",     "Broadband VNIR average"),
    ("Dark current",            3.0,     "e-/s",  "Cooled Si CCD"),
    ("Read noise",              5.0,     "e- RMS", "Low-noise CCD"),
    ("Full well capacity",      100000,  "e-",    "10 um pixel CCD"),
    ("Gain",                    1.0,     "e-/DN", "16-bit ADC"),
    ("ADC bits",                16,      "--",    "65535 DN max"),
    # Spectral
    ("Filter min",              500,     "nm",    "VNIR panchromatic"),
    ("Filter max",              800,     "nm",    "VNIR panchromatic"),
    # Scene
    ("Target reflectance",      0.15,    "--",    "Urban/asphalt"),
    ("Background reflectance",  0.10,    "--",    "Vegetation"),
    ("Solar zenith angle",      30,      "deg",   "Mid-morning"),
    # Orbit & geometry
    ("Orbit altitude",          500,     "km",    "LEO sun-synchronous"),
    # Atmosphere
    ("Atmosphere model",        "simple", "--",   "Simple Beer-Lambert"),
    ("Visibility",              23,      "km",    "Clear day"),
    ("PWV",                     20,      "mm",    "Standard mid-latitude"),
    # Integration time
    ("Integration time",        2.0,     "ms",    "Staring mode"),
]

for i, (param, value, unit, note) in enumerate(rows, start=2):
    ws[f"A{i}"] = param
    ws[f"B{i}"] = value
    ws[f"C{i}"] = unit
    ws[f"D{i}"] = note
    ws[f"A{i}"].font = Font(size=10)
    ws[f"D{i}"].font = note_font

# -- Sheet 2: WFE Sweep -----------------------------------------------------
ws2 = wb.create_sheet("WFE Sweep")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 40

for col in ["A", "B", "C"]:
    ws2[f"{col}1"].font = header_font
    ws2[f"{col}1"].fill = header_fill
    ws2[f"{col}1"].alignment = Alignment(horizontal="center")

ws2["A1"] = "WFE RMS [waves]"
ws2["B1"] = "Strehl (Marechal)"
ws2["C1"] = "Notes"

import math

wfe_values = [0.000, 0.020, 0.040, 0.060, 0.071, 0.080, 0.100,
              0.120, 0.140, 0.160, 0.180, 0.200, 0.250]
notes = [
    "Perfect optics",
    "Excellent (space-quality)",
    "Very good",
    "Good",
    "lambda/14 (diffraction-limited threshold)",
    "Acceptable",
    "Moderate aberration",
    "Noticeable degradation",
    "Significant degradation",
    "Large WFE",
    "Very large WFE",
    "Severe aberration",
    "Extreme — beyond Marechal validity",
]

for i, (wfe, note) in enumerate(zip(wfe_values, notes), start=2):
    ws2[f"A{i}"] = wfe
    strehl = math.exp(-(2.0 * math.pi * wfe) ** 2)
    ws2[f"B{i}"] = round(strehl, 4)
    ws2[f"C{i}"] = note

# -- Sheet 3: Tom's Zernike Coefficients (reference) ------------------------
ws3 = wb.create_sheet("Zernike Coefficients")
ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 50

for col in ["A", "B", "C", "D"]:
    ws3[f"{col}1"].font = header_font
    ws3[f"{col}1"].fill = header_fill
    ws3[f"{col}1"].alignment = Alignment(horizontal="center")

ws3["A1"] = "Zernike Index"
ws3["B1"] = "Name"
ws3["C1"] = "Coefficient [waves]"
ws3["D1"] = "Notes"

zernike_data = [
    (4,  "Defocus",              0.020, "Primary defocus"),
    (5,  "Astigmatism 0",        0.015, ""),
    (6,  "Astigmatism 45",       0.010, ""),
    (7,  "Coma Y",               0.025, "Dominant aberration"),
    (8,  "Coma X",               0.018, ""),
    (9,  "Trefoil Y",            0.005, ""),
    (10, "Trefoil X",            0.004, ""),
    (11, "Spherical",            0.030, "Secondary mirror alignment"),
    (12, "2nd Astigmatism 0",    0.003, ""),
    (13, "2nd Astigmatism 45",   0.002, ""),
    (14, "2nd Coma Y",           0.002, ""),
    (15, "2nd Coma X",           0.001, ""),
]

for i, (idx, name, coeff, note) in enumerate(zernike_data, start=2):
    ws3[f"A{i}"] = idx
    ws3[f"B{i}"] = name
    ws3[f"C{i}"] = coeff
    ws3[f"D{i}"] = note

# Total RMS from Zernike coefficients
rms = math.sqrt(sum(c**2 for _, _, c, _ in zernike_data))
ws3[f"A{len(zernike_data) + 3}"] = "Total RMS"
ws3[f"C{len(zernike_data) + 3}"] = round(rms, 4)
ws3[f"D{len(zernike_data) + 3}"] = "RSS of all coefficients (waves at 633 nm)"

from pathlib import Path

out = Path(__file__).parent / "tom_wfe_budget_data.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"  Total Zernike RMS: {rms:.4f} waves at 633 nm")

# ---------------------------------------------------------------------------
# Zemax "Zernike Standard Coefficients" text export
# (read by radiant.io.zemax_zernike.load_zemax_zernike, Gap 26)
# ---------------------------------------------------------------------------
zmx_out = Path(__file__).parent / "tom_zernike_zemax.txt"
zmx_lines = [
    "Zernike Standard Coefficients",
    "",
    "File : cassegrain_40cm_asbuilt.zmx",
    "Title: TOM 40 CM CASSEGRAIN - AS-BUILT",
    "",
    "Wavelength           :     0.6330 µm",
    "Surface              : Image",
    "Field                : 0.0000, 0.0000 (deg)",
    "",
]
for idx, name, coeff, _note in zernike_data:
    zmx_lines.append(f"Z {idx:3d}    {coeff:14.8f}   :  {name}")
zmx_out.write_text("\n".join(zmx_lines) + "\n", encoding="utf-8")
print(f"Wrote {zmx_out}")
