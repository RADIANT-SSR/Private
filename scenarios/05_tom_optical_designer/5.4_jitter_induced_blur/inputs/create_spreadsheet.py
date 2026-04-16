"""Create Tom's jitter tolerance spreadsheet.

Three sheets:
  1. "Optical System"    — VNIR panchromatic imager, f/10, 50 cm aperture
  2. "Detector & Readout" — silicon CCD, 8 µm pitch
  3. "Jitter Sweep"      — sweep definition and NIIRS requirement
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=4):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Optical System
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Optical System"
ws1["A1"] = "VNIR Panchromatic Imager — Optical Design"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "High-resolution EO satellite, 500 km SSO"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 55

row = 4
row = add_section(ws1, row, "Optics")
row = add_param(ws1, row, "Aperture diameter", 50.0, "cm",
                "Three-mirror anastigmat (TMA)")
row = add_param(ws1, row, "Focal length", 500.0, "cm",
                "f/10 — diffraction-limited design")
row = add_param(ws1, row, "f-number", 10.0, "—", "Slow optics for resolution")
row = add_param(ws1, row, "Optical transmission", 70.0, "%",
                "All elements combined (3 mirrors + filters)")
row = add_param(ws1, row, "Optics temperature", 20.0, "°C",
                "Thermal control on bus")
row = add_param(ws1, row, "WFE RMS", 0.05, "waves",
                "At 633 nm reference, diffraction-limited")
row = add_param(ws1, row, "Central obscuration", 30.0, "%",
                "Linear obscuration ratio (secondary mirror)")

row += 1
row = add_section(ws1, row, "Spectral Band")
row = add_param(ws1, row, "Filter cut-on", 450.0, "nm", "Panchromatic band")
row = add_param(ws1, row, "Filter cut-off", 700.0, "nm", "Panchromatic band")
row = add_param(ws1, row, "Band center", 575.0, "nm", "Arithmetic center")

row += 1
row = add_section(ws1, row, "Geometry")
row = add_param(ws1, row, "Orbit altitude", 500.0, "km",
                "Sun-synchronous orbit")
row = add_param(ws1, row, "Off-nadir angle", 0.0, "deg",
                "Nadir look")

row += 1
row = add_section(ws1, row, "Scene")
row = add_param(ws1, row, "Target reflectance", 0.15, "—",
                "Urban/concrete, 15% reflectance")
row = add_param(ws1, row, "Background reflectance", 0.10, "—",
                "Vegetation/mixed terrain")
row = add_param(ws1, row, "Solar zenith angle", 30.0, "deg",
                "Mid-morning, mid-latitude")
row = add_param(ws1, row, "Target temperature", 300.0, "K",
                "Ambient — thermal contribution negligible in VNIR")
row = add_param(ws1, row, "Background temperature", 295.0, "K",
                "Ambient background")

# ---------------------------------------------------------------------------
# Sheet 2: Detector & Readout
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Detector & Readout")
ws2["A1"] = "Silicon CCD — Detector Configuration"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Back-illuminated CCD, 8 µm pitch, high QE"
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 55

row = 4
row = add_section(ws2, row, "Detector")
row = add_param(ws2, row, "Format", "4096x4096", "pixels",
                "Staring full-frame CCD")
row = add_param(ws2, row, "Pixel pitch", 8.0, "µm", "Square pixels")
row = add_param(ws2, row, "Quantum efficiency", 85.0, "%",
                "Back-illuminated, in-band average")
row = add_param(ws2, row, "Dark current", 5.0, "e⁻/s",
                "At -30°C operating temperature")
row = add_param(ws2, row, "Operating temperature", 243.0, "K",
                "-30°C, thermoelectric cooler")
row = add_param(ws2, row, "IPC coupling", 0.5, "%",
                "CCD — minimal IPC")

row += 1
row = add_section(ws2, row, "Readout")
row = add_param(ws2, row, "Read noise", 5.0, "e⁻ RMS",
                "Low-noise CCD, slow readout")
row = add_param(ws2, row, "Full well capacity", 100000, "e⁻",
                "Standard well depth")
row = add_param(ws2, row, "ADC resolution", 14, "bits", "14-bit ADC")
row = add_param(ws2, row, "System gain", 6.5, "e⁻/DN",
                "Set for ~105% FWC at ADC max")
row = add_param(ws2, row, "Integration time", 0.5, "ms",
                "Limited by smear (7 km/s ground vel, 0.8 m GSD)")

# ---------------------------------------------------------------------------
# Sheet 3: Jitter Sweep
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Jitter Sweep")
ws3["A1"] = "Jitter Tolerance Study — Sweep Definition"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Find jitter RMS that degrades NIIRS by 0.5 and 1.0"
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 55

row = 4
row = add_section(ws3, row, "Jitter Sweep Definition")
row = add_param(ws3, row, "Jitter RMS min", 0.0, "µrad",
                "Zero jitter (baseline)")
row = add_param(ws3, row, "Jitter RMS max", 5.0, "µrad",
                "~3 pixels of blur at 5 m focal length")
row = add_param(ws3, row, "Number of sweep points", 51, "—",
                "Linear-spaced including 0")

row += 1
row = add_section(ws3, row, "Performance Requirements")
row = add_param(ws3, row, "Baseline NIIRS (no jitter)", "computed", "—",
                "Run RADIANT with zero jitter to establish baseline")
row = add_param(ws3, row, "NIIRS degradation threshold 1", 0.5, "—",
                "Find jitter at ΔNIIRS = −0.5 (mild degradation)")
row = add_param(ws3, row, "NIIRS degradation threshold 2", 1.0, "—",
                "Find jitter at ΔNIIRS = −1.0 (severe degradation)")
row = add_param(ws3, row, "Minimum acceptable NIIRS", 6.0, "—",
                "Absolute floor for mission utility")

row += 1
row = add_section(ws3, row, "Reference Values")
row = add_param(ws3, row, "IFOV", "computed", "µrad",
                "pixel_pitch / focal_length")
row = add_param(ws3, row, "GSD at nadir", "computed", "m",
                "pixel_pitch × altitude / focal_length")
row = add_param(ws3, row, "Airy disk diameter", "computed", "µm",
                "2.44 × λ_center × f/#")

row += 1
row = add_section(ws3, row, "Jitter Physical Context")
headers = ["Jitter Source", "Typical RMS [µrad]", "Frequency [Hz]", "Notes"]
for c, h in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border
row += 1

sources = [
    ("Reaction wheel imbalance", "1–5", "10–100",
     "Dominant on small satellites"),
    ("Solar array drive", "0.5–2", "0.01–1",
     "Step-and-stare or slow oscillation"),
    ("Cryocooler vibration", "0.1–1", "30–60",
     "Stirling cooler (if present)"),
    ("Structural modes", "1–10", "5–50",
     "Depends on spacecraft stiffness"),
    ("Attitude control residual", "0.5–5", "0.1–10",
     "Star tracker + gyro loop"),
    ("Thermal snap", "1–20", "transient",
     "Sudden thermal gradient change"),
]
for src, rms, freq, note in sources:
    ws3.cell(row=row, column=1, value=src).border = thin_border
    ws3.cell(row=row, column=2, value=rms).border = thin_border
    ws3.cell(row=row, column=3, value=freq).border = thin_border
    ws3.cell(row=row, column=4, value=note).border = thin_border
    row += 1

from pathlib import Path
out = Path(__file__).parent / "tom_jitter_tolerance_data.xlsx"
wb.save(out)
print(f"Created {out}")
