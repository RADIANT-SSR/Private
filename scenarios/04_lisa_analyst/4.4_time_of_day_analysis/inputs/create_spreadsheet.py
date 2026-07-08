#!/usr/bin/env python3
"""Create Lisa's inputs for Scenario 4.4: time-of-day (diurnal) analysis.

Emits:

- ``diurnal_thermal_profile.csv`` — a 24-hour field-measured surface-
  temperature record for a target (a painted-metal vehicle, low thermal
  inertia → large, early diurnal swing) and its background (soil/
  vegetation, higher thermal inertia → smaller, lagged swing). This is the
  non-RADIANT INPUT DATA (a thermal field campaign product); the two
  surfaces cross temperature twice a day, producing the classic morning
  and evening thermal washout.
- ``lisa_lwir_sensor.xlsx`` — the LWIR sensor configuration Lisa evaluates.

The profile is generated from two offset sinusoids (a standard first-order
diurnal surface model): T(t) = T_mean + A·sin(2π(t − t0)/24), peaking about
6 h after t0. Thermal inertia differences are encoded as different
amplitude and phase.

Run:  python create_spreadsheet.py
"""

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

# --- Diurnal surface-temperature profile ------------------------------
# Target: painted-metal vehicle — low thermal inertia, large swing, peaks
#   ~13:30 (t0 = 7.5). Background: soil — higher inertia, smaller swing,
#   lags to ~15:00 (t0 = 9.0).
TGT_MEAN_K, TGT_AMP_K, TGT_T0 = 297.0, 13.0, 7.5
BG_MEAN_K, BG_AMP_K, BG_T0 = 294.0, 7.0, 9.0


def surface_temp_k(mean: float, amp: float, t0: float, hour: float) -> float:
    return mean + amp * math.sin(2.0 * math.pi * (hour - t0) / 24.0)


rows = []
h = 0.0
while h <= 24.0 + 1e-9:
    tt = surface_temp_k(TGT_MEAN_K, TGT_AMP_K, TGT_T0, h)
    tb = surface_temp_k(BG_MEAN_K, BG_AMP_K, BG_T0, h)
    rows.append((round(h, 2), round(tt, 3), round(tb, 3)))
    h += 0.5

with open(HERE / "diurnal_thermal_profile.csv", "w", encoding="utf-8") as fh:
    fh.write("# 24-hour field-measured surface temperatures (thermal campaign)\n")
    fh.write("# target = painted-metal vehicle; background = soil/vegetation\n")
    fh.write("hour_local,T_target_K,T_background_K\n")
    for hour, tt, tb in rows:
        fh.write(f"{hour},{tt},{tb}\n")
print("Wrote diurnal_thermal_profile.csv")

# --- LWIR sensor config -----------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SensorConfig"
ws["A1"] = "Scenario 4.4: LWIR thermal sensor (time-of-day analysis)"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [30, 14, 12, 44]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="7030A0")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Waveband", "8–12", "µm", "LWIR (thermal emission dominates)"),
    ("Target emissivity", 0.92, "—", "Painted metal"),
    ("Background emissivity", 0.95, "—", "Soil/vegetation"),
    ("Detectability threshold", 6.0, "—", "|contrast SNR| for reliable detection"),
    ("Platform altitude", 3000.0, "m", "Airborne ISR"),
    ("Aperture diameter", 15.0, "cm", ""),
    ("Focal length", 0.6, "m", ""),
    ("Pixel pitch", 25.0, "µm", "LWIR MCT"),
    ("Optical transmission", 80.0, "%", ""),
    ("QE", 70.0, "%", ""),
    ("Integration time", 8.0, "ms", ""),
    ("Dark current", 5.0e6, "e⁻/s", "Cooled LWIR MCT"),
    ("Detector temperature", 77.0, "K", ""),
    ("Read noise", 300.0, "e⁻ RMS", ""),
    ("Full well", 6.0e6, "e⁻", ""),
    ("System gain", 120.0, "e⁻/DN", ""),
    ("ADC resolution", 14, "bits", ""),
]
r = 4
for name, value, unit, note in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "lisa_lwir_sensor.xlsx"
wb.save(out)
print(f"Wrote {out}")
