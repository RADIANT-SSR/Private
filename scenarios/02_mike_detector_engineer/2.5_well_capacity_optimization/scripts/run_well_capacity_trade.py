"""Scenario 2.5: Well Capacity Optimization — Integration Time vs. Dynamic Range

Mike needs to set integration time for a MWIR sensor viewing a scene with
both cold sky (200 K) and hot exhaust plume (1500 K) in the same frame.
With a 2M e⁻ FWC, what integration time gives adequate SNR on the cold target
without saturating on the hot target?

This script:
  1. Reads Mike's spreadsheet (µm, %, cm, ms)
  2. Converts to RADIANT canonical units
  3. Sweeps integration time from 1 µs to 50 ms
  4. At each t_int: computes signal, well fill, SNR for 10 scene temperatures
  5. Finds the max scene temperature before saturation at each t_int
  6. Finds the t_int that gives 70% well fill for each temperature
  7. Compares noise budgets: cold (200 K) vs. hot (400 K) at optimal t_int
  8. Writes results to an output spreadsheet

Usage:
    python run_well_capacity_trade.py
"""

import copy
import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Mike's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "mike_well_capacity_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "well_capacity_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)


def _read_sheet(ws, min_row: int = 5):
    """Read parameter name→value and name→unit from a spreadsheet sheet."""
    specs: dict[str, object] = {}
    units: dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_row, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
                continue
            try:
                specs[name] = float(value)
            except (ValueError, TypeError):
                specs[name] = value
            units[name] = str(row[2].value) if row[2].value else "—"
    return specs, units


det_specs, det_units = _read_sheet(wb["Detector Specs"])
sys_specs, sys_units = _read_sheet(wb["Optics & Scene"])
req_specs, req_units = _read_sheet(wb["Trade Requirements"])

print("=" * 80)
print("SCENARIO 2.5: Well Capacity Optimization — Integration Time vs. Dynamic Range")
print("=" * 80)

print(f"\n=== Detector Specs ===")
for k, v in det_specs.items():
    print(f"  {k}: {v} {det_units.get(k, '')}")

print(f"\n=== Optics & Scene ===")
for k, v in sys_specs.items():
    print(f"  {k}: {v} {sys_units.get(k, '')}")

print(f"\n=== Trade Requirements ===")
for k, v in req_specs.items():
    print(f"  {k}: {v} {req_units.get(k, '')}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

aperture_m = float(sys_specs["Aperture diameter"]) / 100.0
focal_length_m = float(sys_specs["Focal length"]) / 100.0
f_number = float(sys_specs["f-number"])
transmission = float(sys_specs["Optical transmission"]) / 100.0
optics_temp_K = float(sys_specs["Optics temperature"]) + 273.15

band_min_um = float(sys_specs["Filter cut-on"]) / 1000.0
band_max_um = float(sys_specs["Filter cut-off"]) / 1000.0

cold_temp_K = float(sys_specs["Cold target temperature"])
hot_temp_K = float(sys_specs["Hot target temperature"])
target_emiss = float(sys_specs["Target emissivity"])
bg_temp_K = float(sys_specs["Background temperature"])
bg_emiss = float(sys_specs["Background emissivity"])

pixel_pitch_um = float(det_specs["Pixel pitch"])
qe = float(det_specs["Quantum efficiency (in-band avg)"]) / 100.0
dark_rate = float(det_specs["Dark current"])
operating_temp_K = float(det_specs["Operating temperature"])
read_noise = float(det_specs["Read noise (post-CDS)"])
fwc = float(det_specs["Full well capacity"])
adc_bits = int(det_specs["ADC resolution"])
gain = float(det_specs["System gain"])
ipc_coupling = float(det_specs["IPC coupling"]) / 100.0

snr_req = float(req_specs["Minimum SNR (cold target)"])
max_fill_pct = float(req_specs["Maximum well fill (hot target)"])
sat_limit_pct = float(req_specs["Saturation limit"])
t_int_min_s = float(req_specs["Integration time min"]) / 1000.0
t_int_max_s = float(req_specs["Integration time max"]) / 1000.0
n_sweep = int(req_specs["Number of sweep points"])

# Scene temperatures for multi-target analysis
scene_temps_K = [200.0, 250.0, 280.0, 300.0, 350.0, 400.0, 500.0, 700.0, 1000.0, 1500.0]

print(f"\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'f-number':<35s} {f_number:>14.2f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  nm ÷ 1000")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<35s} {dark_rate:>14.0f}  {'e⁻/s':<15s}  no conversion")
print(f"  {'Read noise':<35s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'FWC':<35s} {fwc:>14.0f}  {'e⁻':<15s}  no conversion")
print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<15s}  % ÷ 100")

# ---------------------------------------------------------------------------
# Step 3: Build RADIANT config
# ---------------------------------------------------------------------------

print(f"\n=== Radiometric Regime ===")
print(f"  Atmosphere model: exo (short-range ground-based)")
print(f"  Extended regime: target fills entire pixel FOV.")
print(f"  Scene dynamic range: {cold_temp_K:.0f} K (cold sky) to"
      f" {hot_temp_K:.0f} K (exhaust plume)")
print(f"  FWC = {fwc/1e6:.1f} M e⁻")
print(f"")
print(f"  WELL FILL PHYSICS NOTE:")
print(f"    Well fill = signal_e / FWC.")
print(f"    The well accumulates signal photons + dark current + ROIC glow.")
print(f"    Background and nearfield photons also fill the well in reality,")
print(f"    but RADIANT currently clips signal_e against (FWC - dark - glow).")
print(f"    For hot targets, signal dominates; for cold targets, background")
print(f"    and nearfield from warm optics ({optics_temp_K:.0f} K) dominate.")
print(f"")
print(f"  DYNAMIC RANGE NOTE:")
print(f"    A 1500 K blackbody in MWIR (3.5–5.0 µm) produces ~10,000× more")
print(f"    radiance than a 200 K target. No single integration time can")
print(f"    give adequate SNR on 200 K without saturating on 1500 K.")
print(f"    This scenario quantifies the trade-off.")


def make_config(target_temp_K: float, t_int_s: float) -> dict:
    """Build RADIANT config for a given target temperature and integration time."""
    return {
        "source": {
            "target": {"temperature": target_temp_K, "emissivity": target_emiss},
            "background": {"temperature": bg_temp_K, "emissivity": bg_emiss},
        },
        "atmosphere": {"model": "exo"},
        "geometry": {"sensor_altitude_m": 0.0},
        # Stage-7 stop-gap (registry Gap 42): "exo" routes through the
        # no_atmosphere 'space' sub-case, whose Earth-limb check requires a
        # positive user-set platform.h_sensor [m above MSL]. 1.0 m ≈ bench
        # height; feeds only the limb check, no radiometric effect.
        "platform": {"h_sensor": 1.0},
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": operating_temp_K,
            "ipc_coupling": ipc_coupling,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
    }


# ---------------------------------------------------------------------------
# Step 4: Sweep integration time for each scene temperature
# ---------------------------------------------------------------------------

t_int_sweep_s = np.logspace(
    np.log10(t_int_min_s), np.log10(t_int_max_s), n_sweep
)

print(f"\n=== Sweeping integration time: {t_int_min_s*1e6:.0f} µs to"
      f" {t_int_max_s*1e3:.0f} ms ({n_sweep} points, log-spaced) ===")
print(f"  Evaluating {len(scene_temps_K)} scene temperatures × {n_sweep}"
      f" integration times = {len(scene_temps_K) * n_sweep} evaluations")

# Store results: sweep_data[temp_K] = list of dicts per t_int
sweep_data: dict[float, list[dict]] = {}

for t_K in scene_temps_K:
    sweep_data[t_K] = []
    for t_int in t_int_sweep_s:
        cfg = make_config(t_K, t_int)
        r = Sensor.from_dict(cfg).evaluate()

        signal_e = r.stage_outputs["readout"]["signal_e_final"]
        well_fill = signal_e / fwc * 100.0
        snr = r.metrics["snr"]
        well_status = r.stage_outputs["readout"].get("well_status", "unknown")

        sweep_data[t_K].append({
            "t_int_s": t_int,
            "signal_e": signal_e,
            "well_fill_pct": well_fill,
            "snr": snr,
            "saturated": well_fill >= 99.9,
            "well_status": well_status,
        })

# ---------------------------------------------------------------------------
# Step 5: Print well fill vs. integration time for key temperatures
# ---------------------------------------------------------------------------

key_temps = [200.0, 300.0, 400.0, 500.0, 1000.0, 1500.0]

print(f"\n=== Well Fill [%] vs. Integration Time ===")
# Header
header = f"  {'t_int':>12s}"
for t_K in key_temps:
    header += f"  {t_K:.0f} K".rjust(10)
print(header)
print(f"  {'-' * 12}" + f"  {'-' * 10}" * len(key_temps))

# Print every ~5th point for readability
for i in range(0, n_sweep, max(1, n_sweep // 12)):
    t_int = t_int_sweep_s[i]
    if t_int < 1e-3:
        t_str = f"{t_int*1e6:.1f} µs"
    elif t_int < 1.0:
        t_str = f"{t_int*1e3:.2f} ms"
    else:
        t_str = f"{t_int:.2f} s"

    row_str = f"  {t_str:>12s}"
    for t_K in key_temps:
        data = sweep_data[t_K][i]
        fill = data["well_fill_pct"]
        if fill >= 99.9:
            row_str += "       SAT".rjust(10)
        else:
            row_str += f"  {fill:>7.1f} %"
    print(row_str)

# ---------------------------------------------------------------------------
# Step 6: SNR vs. integration time for cold target (200 K)
# ---------------------------------------------------------------------------

print(f"\n=== SNR vs. Integration Time for Cold Target ({cold_temp_K:.0f} K) ===")
print(f"  Requirement: SNR ≥ {snr_req:.0f} [—]")
print(f"")
print(f"  {'t_int':>12s}  {'Signal [e⁻]':>14s}  {'Well Fill [%]':>13s}"
      f"  {'SNR [—]':>10s}  {'Status':>10s}")
print(f"  {'-' * 12}  {'-' * 14}  {'-' * 13}  {'-' * 10}  {'-' * 10}")

snr_threshold_t_int = None
for i, d in enumerate(sweep_data[cold_temp_K]):
    t_int = d["t_int_s"]
    if t_int < 1e-3:
        t_str = f"{t_int*1e6:.1f} µs"
    elif t_int < 1.0:
        t_str = f"{t_int*1e3:.2f} ms"
    else:
        t_str = f"{t_int:.2f} s"

    status = "PASS" if d["snr"] >= snr_req else "FAIL"
    if snr_threshold_t_int is None and d["snr"] >= snr_req:
        snr_threshold_t_int = t_int
        status += " ←"

    if i % max(1, n_sweep // 12) == 0 or status.startswith("PASS ←"):
        print(f"  {t_str:>12s}  {d['signal_e']:>14,.0f}  {d['well_fill_pct']:>12.1f}%"
              f"  {d['snr']:>10.1f}  {status:>10s}")

if snr_threshold_t_int is not None:
    if snr_threshold_t_int < 1e-3:
        thr_str = f"{snr_threshold_t_int*1e6:.1f} µs"
    else:
        thr_str = f"{snr_threshold_t_int*1e3:.2f} ms"
    print(f"\n  → SNR ≥ {snr_req:.0f} achieved at t_int ≥ {thr_str}")
else:
    print(f"\n  → SNR ≥ {snr_req:.0f} NOT achieved in sweep range")

# ---------------------------------------------------------------------------
# Step 7: Max scene temperature before saturation at each t_int
# ---------------------------------------------------------------------------

print(f"\n=== Scene Dynamic Range vs. Integration Time ===")
print(f"  For each t_int, find the highest scene temperature that stays below"
      f" {sat_limit_pct:.0f}% well fill.")
print(f"")
print(f"  {'t_int':>12s}  {'Max T (90% fill) [K]':>20s}  {'Max T (70% fill) [K]':>20s}"
      f"  {'Cold SNR [—]':>12s}")
print(f"  {'-' * 12}  {'-' * 20}  {'-' * 20}  {'-' * 12}")

for i in range(0, n_sweep, max(1, n_sweep // 12)):
    t_int = t_int_sweep_s[i]
    if t_int < 1e-3:
        t_str = f"{t_int*1e6:.1f} µs"
    elif t_int < 1.0:
        t_str = f"{t_int*1e3:.2f} ms"
    else:
        t_str = f"{t_int:.2f} s"

    max_t_90 = 0.0
    max_t_70 = 0.0
    for t_K in scene_temps_K:
        fill = sweep_data[t_K][i]["well_fill_pct"]
        if fill < sat_limit_pct:
            max_t_90 = t_K
        if fill < max_fill_pct:
            max_t_70 = t_K

    cold_snr = sweep_data[cold_temp_K][i]["snr"]

    t90_str = f"{max_t_90:.0f}" if max_t_90 > 0 else "< 200"
    t70_str = f"{max_t_70:.0f}" if max_t_70 > 0 else "< 200"
    print(f"  {t_str:>12s}  {t90_str:>20s}  {t70_str:>20s}  {cold_snr:>12.1f}")

# ---------------------------------------------------------------------------
# Step 8: Find optimal t_int for 70% well fill at each temperature
# ---------------------------------------------------------------------------

print(f"\n=== Integration Time for {max_fill_pct:.0f}% Well Fill at Each Temperature ===")
print(f"  {'Scene Temp [K]':>14s}  {'t_int for 70% [ms]':>20s}  {'SNR at that t_int [—]':>22s}"
      f"  {'Signal [e⁻]':>14s}")
print(f"  {'-' * 14}  {'-' * 20}  {'-' * 22}  {'-' * 14}")

target_fill = max_fill_pct / 100.0 * fwc  # e⁻

for t_K in scene_temps_K:
    # Find t_int where well fill crosses 70%
    found_t = None
    found_snr = None
    found_sig = None
    for d in sweep_data[t_K]:
        if d["signal_e"] >= target_fill:
            found_t = d["t_int_s"]
            found_snr = d["snr"]
            found_sig = d["signal_e"]
            break

    if found_t is not None:
        if found_t < 1e-3:
            t_str = f"{found_t*1e6:.1f} µs"
        else:
            t_str = f"{found_t*1e3:.3f} ms"
        print(f"  {t_K:>14.0f}  {t_str:>20s}  {found_snr:>22.1f}  {found_sig:>14,.0f}")
    else:
        # Signal never reaches 70% — extrapolate from last point
        last = sweep_data[t_K][-1]
        if last["signal_e"] > 0 and last["signal_e"] < target_fill:
            # Linear extrapolation: signal ∝ t_int
            ratio = target_fill / last["signal_e"]
            est_t = last["t_int_s"] * ratio
            if est_t < 1e-3:
                t_str = f"~{est_t*1e6:.0f} µs"
            elif est_t < 1.0:
                t_str = f"~{est_t*1e3:.1f} ms"
            else:
                t_str = f"~{est_t:.1f} s"
            print(f"  {t_K:>14.0f}  {t_str:>20s}  {'(extrapolated)':>22s}  {'—':>14s}")
        else:
            print(f"  {t_K:>14.0f}  {'never reaches 70%':>20s}  {'—':>22s}  {'—':>14s}")

# ---------------------------------------------------------------------------
# Step 9: Noise budget comparison — cold vs. warm at a practical t_int
# ---------------------------------------------------------------------------

# Pick a practical integration time: 1 ms (where cold target has some signal)
COMPARE_T_INT = 0.001  # s

print(f"\n=== Noise Budget Comparison at t_int = {COMPARE_T_INT*1e3:.1f} ms ===")
print(f"  Comparing cold ({cold_temp_K:.0f} K) vs. warm (400 K) targets")
print(f"  at the same integration time to show noise regime differences.")

compare_temps = [cold_temp_K, 400.0]
compare_results = {}

for t_K in compare_temps:
    cfg = make_config(t_K, COMPARE_T_INT)
    r = Sensor.from_dict(cfg).evaluate()
    signal_e = r.stage_outputs["readout"]["signal_e_final"]
    snr = r.metrics["snr"]
    noise_dict = {nt.name: nt.value_e for nt in r.noise_terms}
    total_noise = math.sqrt(sum(v ** 2 for v in noise_dict.values()))
    well_fill = signal_e / fwc * 100.0

    compare_results[t_K] = {
        "signal_e": signal_e,
        "snr": snr,
        "well_fill_pct": well_fill,
        "total_noise": total_noise,
        "noise_dict": noise_dict,
    }

for t_K in compare_temps:
    cr = compare_results[t_K]
    print(f"\n  --- {t_K:.0f} K Target ---")
    print(f"  Signal:     {cr['signal_e']:>14,.0f} e⁻ ({cr['well_fill_pct']:.1f}% well)")
    print(f"  SNR:        {cr['snr']:>14.1f} [—]")
    print(f"  Total noise:{cr['total_noise']:>14.1f} e⁻ RMS")
    print(f"")
    print(f"  {'Noise Term':<25s}  {'σ [e⁻ RMS]':>12s}  {'Fraction [%]':>13s}")
    print(f"  {'-' * 25}  {'-' * 12}  {'-' * 13}")

    total_var = cr["total_noise"] ** 2
    sorted_terms = sorted(cr["noise_dict"].items(), key=lambda x: x[1], reverse=True)
    for name, sigma in sorted_terms:
        if sigma < 0.01:
            continue
        frac = sigma ** 2 / total_var * 100.0 if total_var > 0 else 0.0
        print(f"  {name:<25s}  {sigma:>12.1f}  {frac:>13.1f}")

# Regime comparison
print(f"\n  Key difference:")
cold_cr = compare_results[cold_temp_K]
warm_cr = compare_results[400.0]
cold_sig_frac = (cold_cr["noise_dict"]["signal_shot"] ** 2 /
                 cold_cr["total_noise"] ** 2 * 100.0)
warm_sig_frac = (warm_cr["noise_dict"]["signal_shot"] ** 2 /
                 warm_cr["total_noise"] ** 2 * 100.0)
print(f"    At {cold_temp_K:.0f} K: signal_shot is {cold_sig_frac:.1f}% of noise"
      f" → {'BLIP (background-limited)' if cold_sig_frac < 30 else 'signal-limited'}")
print(f"    At 400 K: signal_shot is {warm_sig_frac:.1f}% of noise"
      f" → {'signal-limited' if warm_sig_frac > 50 else 'mixed regime'}")

# ---------------------------------------------------------------------------
# Step 10: Summary
# ---------------------------------------------------------------------------

print(f"\n=== Summary ===")
print(f"  Scene dynamic range requested: {cold_temp_K:.0f} K to {hot_temp_K:.0f} K")
print(f"  FWC: {fwc/1e6:.1f} M e⁻")
print(f"")

# Find practical limits
# At 1 ms: what's the max temp?
idx_1ms = min(range(n_sweep), key=lambda i: abs(t_int_sweep_s[i] - 0.001))
max_t_at_1ms = 0.0
for t_K in scene_temps_K:
    if sweep_data[t_K][idx_1ms]["well_fill_pct"] < sat_limit_pct:
        max_t_at_1ms = t_K

cold_snr_1ms = sweep_data[cold_temp_K][idx_1ms]["snr"]

print(f"  At t_int = 1 ms:")
print(f"    Cold target ({cold_temp_K:.0f} K) SNR = {cold_snr_1ms:.1f} [—]"
      f" ({'PASS' if cold_snr_1ms >= snr_req else 'FAIL'}, req ≥ {snr_req:.0f})")
print(f"    Max scene temp below {sat_limit_pct:.0f}% fill: {max_t_at_1ms:.0f} K")
print(f"    Achievable dynamic range: {cold_temp_K:.0f}–{max_t_at_1ms:.0f} K")
print(f"")
print(f"  The {cold_temp_K:.0f}–{hot_temp_K:.0f} K dynamic range is physically")
print(f"  impossible in a single frame. A {hot_temp_K:.0f} K blackbody in MWIR produces")
if max_t_at_1ms > 0:
    ratio = sweep_data[hot_temp_K][idx_1ms]["signal_e"] / max(
        sweep_data[cold_temp_K][idx_1ms]["signal_e"], 1
    )
    print(f"  ~{ratio:,.0f}× more signal than a {cold_temp_K:.0f} K target.")
print(f"")
print(f"  Solutions for wide dynamic range:")
print(f"    1. Dual-integration (short + long exposure per frame)")
print(f"    2. High-dynamic-range (HDR) readout modes")
print(f"    3. Gain switching per pixel")
print(f"    4. Spectral narrowing (reduce band to lower flux)")
print(f"    5. Accept saturation on hottest targets (flag and discard)")

# ---------------------------------------------------------------------------
# Step 11: Write output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

owb = openpyxl.Workbook()
hfont = Font(bold=True, size=11)
sfill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
sfont = Font(bold=True, size=11, color="FFFFFF")
tb = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: Well Fill Sweep ---
ows1 = owb.active
ows1.title = "Well Fill Sweep"
ows1["A1"] = "Well Fill [%] vs. Integration Time and Scene Temperature"
ows1["A1"].font = Font(bold=True, size=14)

headers = ["t_int [ms]"] + [f"{t:.0f} K [%]" for t in scene_temps_K]
for c, h in enumerate(headers, 1):
    cell = ows1.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb
ows1.column_dimensions["A"].width = 14
for c in range(2, len(headers) + 1):
    ows1.column_dimensions[chr(64 + c)].width = 12

for i, t_int in enumerate(t_int_sweep_s, 4):
    ows1.cell(row=i, column=1, value=round(t_int * 1000, 6)).border = tb
    for c, t_K in enumerate(scene_temps_K, 2):
        fill = sweep_data[t_K][i - 4]["well_fill_pct"]
        ows1.cell(row=i, column=c, value=round(fill, 2)).border = tb

# --- Sheet 2: SNR Sweep ---
ows2 = owb.create_sheet("SNR Sweep")
ows2["A1"] = "SNR [—] vs. Integration Time and Scene Temperature"
ows2["A1"].font = Font(bold=True, size=14)

headers2 = ["t_int [ms]"] + [f"{t:.0f} K [—]" for t in scene_temps_K]
for c, h in enumerate(headers2, 1):
    cell = ows2.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb
ows2.column_dimensions["A"].width = 14

for i, t_int in enumerate(t_int_sweep_s, 4):
    ows2.cell(row=i, column=1, value=round(t_int * 1000, 6)).border = tb
    for c, t_K in enumerate(scene_temps_K, 2):
        snr_val = sweep_data[t_K][i - 4]["snr"]
        ows2.cell(row=i, column=c, value=round(snr_val, 2)).border = tb

# --- Sheet 3: Noise Comparison ---
ows3 = owb.create_sheet("Noise Comparison")
ows3["A1"] = f"Noise Budget Comparison at t_int = {COMPARE_T_INT*1e3:.1f} ms"
ows3["A1"].font = Font(bold=True, size=14)
ows3.column_dimensions["A"].width = 26
ows3.column_dimensions["B"].width = 18
ows3.column_dimensions["C"].width = 18

headers3 = ["Noise Term", f"{cold_temp_K:.0f} K [e⁻ RMS]", "400 K [e⁻ RMS]"]
for c, h in enumerate(headers3, 1):
    cell = ows3.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb

all_terms = sorted(
    set(list(compare_results[cold_temp_K]["noise_dict"].keys()) +
        list(compare_results[400.0]["noise_dict"].keys()))
)
row_idx = 4
for name in all_terms:
    s_cold = compare_results[cold_temp_K]["noise_dict"].get(name, 0.0)
    s_warm = compare_results[400.0]["noise_dict"].get(name, 0.0)
    if s_cold < 0.01 and s_warm < 0.01:
        continue
    ows3.cell(row=row_idx, column=1, value=name).border = tb
    ows3.cell(row=row_idx, column=2, value=round(s_cold, 2)).border = tb
    ows3.cell(row=row_idx, column=3, value=round(s_warm, 2)).border = tb
    row_idx += 1

# --- Sheet 4: Summary ---
ows4 = owb.create_sheet("Summary")
ows4["A1"] = "Scenario 2.5 Summary"
ows4["A1"].font = Font(bold=True, size=14)
ows4.column_dimensions["A"].width = 40
ows4.column_dimensions["B"].width = 25

summaries = [
    ("System", f"MWIR HgCdTe, {pixel_pitch_um:.0f} µm, f/{f_number:.1f}"),
    ("Band", f"{band_min_um:.1f}–{band_max_um:.1f} µm"),
    ("FWC", f"{fwc/1e6:.1f} M e⁻"),
    ("Scene range requested", f"{cold_temp_K:.0f}–{hot_temp_K:.0f} K"),
    ("Achievable range (1 ms, <90% fill)", f"{cold_temp_K:.0f}–{max_t_at_1ms:.0f} K"),
    ("Cold target SNR at 1 ms", f"{cold_snr_1ms:.1f} [—]"),
    ("SNR requirement", f"≥ {snr_req:.0f} [—]"),
    ("Max well fill requirement", f"{max_fill_pct:.0f} %"),
]
row = 3
for label, val in summaries:
    ows4.cell(row=row, column=1, value=label).border = tb
    ows4.cell(row=row, column=2, value=val).border = tb
    row += 1

owb.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
