"""Scenario 7.1: Predicted vs. Measured NEDT Reconciliation

Karen measured NEDT on the as-built MWIR sensor in TVAC at multiple blackbody
temperatures.  The primary measurement at 25°C (298.15 K) gave NEDT = 22.0 mK,
but RADIANT predicts ~18 mK with the as-built parameters.  Where does the
4 mK discrepancy come from?

This script:
  1. Reads Karen's lab notebook spreadsheet (°C, %, nm, ms, e⁻/s)
  2. Converts to RADIANT canonical units
  3. Runs RADIANT in exo (vacuum/lab) mode at each measured temperature
  4. Computes predicted NEDT via finite-difference dS/dT
  5. Breaks down NEDT by individual noise term
  6. Identifies which noise term(s) explain the gap
  7. Computes sensitivity: d(NEDT)/d(each parameter)
  8. Writes results to an output spreadsheet

Usage:
    python run_nedt_reconciliation.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Karen's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "karen_nedt_lab_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "nedt_reconciliation_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

# --- System Configuration ---
ws_sys = wb["System Configuration"]
sys_specs: dict[str, object] = {}
for row in ws_sys.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        fill = row[0].fill
        if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
            continue
        try:
            sys_specs[name] = float(value)
        except (ValueError, TypeError):
            sys_specs[name] = value

# --- As-Built Detector ---
ws_det = wb["As-Built Detector"]
det_specs: dict[str, object] = {}
for row in ws_det.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        fill = row[0].fill
        if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
            continue
        try:
            det_specs[name] = float(value)
        except (ValueError, TypeError):
            det_specs[name] = value

# --- NEDT Measurements ---
ws_nedt = wb["NEDT Measurements"]
nedt_meas: list[dict] = []
for row in ws_nedt.iter_rows(min_row=5, max_col=6, values_only=True):
    if row[0] is not None and row[2] is not None:
        nedt_meas.append({
            "bb_temp_C": float(row[0]),
            "bb_temp_K": float(row[1]),
            "nedt_mK": float(row[2]),
            "std_mK": float(row[3]),
            "n_frames": int(row[4]),
            "notes": row[5] or "",
        })

print("=" * 80)
print("SCENARIO 7.1: Predicted vs. Measured NEDT Reconciliation")
print("=" * 80)

print(f"\n=== System Configuration ===")
for k, v in sys_specs.items():
    print(f"  {k}: {v}")

print(f"\n=== As-Built Detector ===")
for k, v in det_specs.items():
    print(f"  {k}: {v}")

print(f"\n=== Lab NEDT Measurements ===")
print(f"  {'BB T [°C]':>10s}  {'BB T [K]':>10s}  {'NEDT [mK]':>10s}  {'σ [mK]':>8s}  {'N':>5s}")
print(f"  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 8}  {'-' * 5}")
for m in nedt_meas:
    print(f"  {m['bb_temp_C']:>10.1f}  {m['bb_temp_K']:>10.2f}  {m['nedt_mK']:>10.1f}"
          f"  {m['std_mK']:>8.1f}  {m['n_frames']:>5d}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

# Optics
aperture_m = float(sys_specs["Primary aperture diameter"]) / 100.0  # cm → m
focal_length_m = float(sys_specs["Effective focal length"]) / 100.0  # cm → m
f_number = float(sys_specs["f-number (actual)"])
transmission = float(sys_specs["Optical transmission"]) / 100.0  # % → fraction
optics_temp_K = float(sys_specs["Optics temperature"]) + 273.15  # °C → K

# Spectral — nm → µm
band_min_um = float(sys_specs["Filter cut-on"]) / 1000.0
band_max_um = float(sys_specs["Filter cut-off"]) / 1000.0

# Test setup
bb_emiss = float(sys_specs["Blackbody emissivity"])
shroud_temp_K = float(sys_specs["Shroud temperature"]) + 273.15  # °C → K
shroud_emiss = float(sys_specs["Shroud emissivity"])

# Detector
pixel_pitch_um = float(det_specs["Pixel pitch"])
qe = float(det_specs["Quantum efficiency (in-band avg)"]) / 100.0  # % → fraction
dark_rate = float(det_specs["Dark current (measured)"])  # already e⁻/s
operating_temp_K = float(det_specs["Operating temperature"])  # already K
read_noise = float(det_specs["Read noise (post-CDS)"])  # already e⁻ RMS
fwc = float(det_specs["Full well capacity"])  # already e⁻
adc_bits = int(det_specs["ADC resolution"])
gain = float(det_specs["System gain"])  # already e⁻/DN
t_int_s = float(det_specs["Integration time"]) / 1000.0  # ms → s
ipc_coupling = float(det_specs["IPC coupling"]) / 100.0  # % → fraction

print(f"\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'f-number':<35s} {f_number:>14.2f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  nm ÷ 1000")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<35s} {dark_rate:>14.1f}  {'e⁻/s':<15s}  no conversion")
print(f"  {'Read noise':<35s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms ÷ 1000")
print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<15s}  % ÷ 100")

# ---------------------------------------------------------------------------
# Step 3: Build RADIANT config (exo = vacuum/lab)
# ---------------------------------------------------------------------------


def make_config(target_temp_K: float) -> dict:
    """Build RADIANT config for a lab blackbody at the given temperature."""
    return {
        "source": {
            "target": {
                "temperature": target_temp_K,
                "emissivity": bb_emiss,
            },
            "background": {
                "temperature": shroud_temp_K,
                "emissivity": shroud_emiss,
            },
        },
        "atmosphere": {
            "model": "exo",  # Vacuum — no atmospheric effects
        },
        "geometry": {
            "sensor_altitude_m": 0.0,  # Lab bench
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": operating_temp_K,
            "ipc_coupling": ipc_coupling,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
    }


# ---------------------------------------------------------------------------
# Step 4: Run RADIANT at each measured temperature — compute predicted NEDT
# ---------------------------------------------------------------------------
# NEDT = σ_total / (dS/dT)
# We compute dS/dT via finite difference: (S(T+δ) - S(T-δ)) / (2δ)

DELTA_T = 0.5  # K — finite difference step for dS/dT

print(f"\n=== Radiometric Regime ===")
print(f"  Atmosphere model: exo (vacuum — TVAC chamber)")
print(f"  Extended regime: blackbody fills entire pixel FOV.")
print(f"  Background (shroud at {shroud_temp_K:.1f} K) enters contrast SNR only.")
print(f"")
print(f"  NEDT PHYSICS NOTE:")
print(f"    NEDT = σ_total / (dS/dT)")
print(f"    where σ_total is the RSS of all noise terms [e⁻ RMS]")
print(f"    and dS/dT is the signal derivative with respect to target")
print(f"    temperature [e⁻/K], computed from the Planck function")
print(f"    integrated over the spectral band.")
print(f"")
print(f"    NEDT tells Karen: the minimum temperature difference this")
print(f"    sensor can resolve (at 1σ) against a uniform background.")
print(f"    Lower NEDT = better thermal sensitivity.")

print(f"\n=== Running RADIANT at each measurement temperature ===")
print(f"  dS/dT computed via finite difference with δT = {DELTA_T} K")

pred_results: list[dict] = []

for m in nedt_meas:
    T = m["bb_temp_K"]

    # Run at T, T+δ, T-δ
    result_center = Sensor.from_dict(make_config(T)).evaluate()
    result_plus = Sensor.from_dict(make_config(T + DELTA_T)).evaluate()
    result_minus = Sensor.from_dict(make_config(T - DELTA_T)).evaluate()

    signal_e = result_center.stage_outputs["readout"]["signal_e_final"]
    signal_plus = result_plus.stage_outputs["readout"]["signal_e_final"]
    signal_minus = result_minus.stage_outputs["readout"]["signal_e_final"]
    ds_dt = (signal_plus - signal_minus) / (2.0 * DELTA_T)  # e⁻/K

    # Noise budget
    noise_dict = {nt.name: nt.value_e for nt in result_center.noise_terms}
    total_noise = math.sqrt(sum(v ** 2 for v in noise_dict.values()))

    # Predicted NEDT
    nedt_pred_K = total_noise / ds_dt if ds_dt > 0 else float("inf")
    nedt_pred_mK = nedt_pred_K * 1000.0

    # Per-term NEDT contribution: NEDT_i = σ_i / (dS/dT)
    # Since NEDT_total = √(Σ NEDT_i²), each term contributes independently
    nedt_per_term = {}
    for name, sigma in noise_dict.items():
        nedt_per_term[name] = (sigma / ds_dt * 1000.0) if ds_dt > 0 else float("inf")

    # Native NEDT from RADIANT (Planck-derivative approximation)
    nedt_native_mK = result_center.metrics.get("nedt_K")
    if nedt_native_mK is not None:
        nedt_native_mK *= 1000.0

    pred_results.append({
        "bb_temp_K": T,
        "bb_temp_C": m["bb_temp_C"],
        "meas_nedt_mK": m["nedt_mK"],
        "meas_std_mK": m["std_mK"],
        "pred_nedt_mK": nedt_pred_mK,
        "native_nedt_mK": nedt_native_mK,
        "signal_e": signal_e,
        "ds_dt": ds_dt,
        "total_noise": total_noise,
        "noise_dict": noise_dict,
        "nedt_per_term": nedt_per_term,
        "snr": result_center.metrics["snr"],
    })

# Print results
print(f"\n=== Predicted vs. Measured NEDT ===")
print(f"  {'BB T [°C]':>10s}  {'BB T [K]':>10s}  {'Meas [mK]':>10s}  {'Pred [mK]':>10s}"
      f"  {'Δ [mK]':>8s}  {'Signal [e⁻]':>12s}  {'dS/dT [e⁻/K]':>14s}  {'σ_total [e⁻]':>14s}")
print(f"  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 10}"
      f"  {'-' * 8}  {'-' * 12}  {'-' * 14}  {'-' * 14}")
for r in pred_results:
    delta = r["meas_nedt_mK"] - r["pred_nedt_mK"]
    print(f"  {r['bb_temp_C']:>10.1f}  {r['bb_temp_K']:>10.2f}  {r['meas_nedt_mK']:>10.1f}"
          f"  {r['pred_nedt_mK']:>10.2f}  {delta:>8.2f}  {r['signal_e']:>12,.0f}"
          f"  {r['ds_dt']:>14.1f}  {r['total_noise']:>14.2f}")

# ---------------------------------------------------------------------------
# Step 5: NEDT breakdown by noise term at the primary test point (25°C)
# ---------------------------------------------------------------------------

primary = next(r for r in pred_results if r["bb_temp_C"] == 25.0)

print(f"\n=== NEDT Breakdown at Primary Test Point (25°C / 298.15 K) ===")
print(f"  Signal:     {primary['signal_e']:>14,.0f} e⁻")
print(f"  dS/dT:      {primary['ds_dt']:>14.1f} e⁻/K")
print(f"  Total noise:{primary['total_noise']:>14.2f} e⁻ RMS")
print(f"")
print(f"  {'Noise Term':<25s}  {'σ [e⁻ RMS]':>12s}  {'NEDT_i [mK]':>12s}  {'Fraction [%]':>13s}")
print(f"  {'-' * 25}  {'-' * 12}  {'-' * 12}  {'-' * 13}")

nedt_squared_total = sum(v ** 2 for v in primary["nedt_per_term"].values())
sorted_terms = sorted(primary["nedt_per_term"].items(), key=lambda x: x[1], reverse=True)

for name, nedt_i in sorted_terms:
    sigma = primary["noise_dict"][name]
    frac_pct = (nedt_i ** 2 / nedt_squared_total * 100.0) if nedt_squared_total > 0 else 0.0
    print(f"  {name:<25s}  {sigma:>12.2f}  {nedt_i:>12.3f}  {frac_pct:>13.1f}")

print(f"  {'─' * 25}  {'─' * 12}  {'─' * 12}  {'─' * 13}")
print(f"  {'RSS TOTAL':<25s}  {primary['total_noise']:>12.2f}  {primary['pred_nedt_mK']:>12.3f}  {'100.0':>13s}")
print(f"\n  Measured NEDT:  {primary['meas_nedt_mK']:.1f} mK")
print(f"  Predicted NEDT: {primary['pred_nedt_mK']:.2f} mK")
print(f"  Gap:            {primary['meas_nedt_mK'] - primary['pred_nedt_mK']:.2f} mK")

# ---------------------------------------------------------------------------
# Step 6: Gap analysis — which noise term explains the discrepancy?
# ---------------------------------------------------------------------------
# If NEDT_measured > NEDT_predicted, there is "missing" noise:
#   σ_missing² = σ_measured² − σ_predicted²
# where σ = NEDT × dS/dT

print(f"\n=== Gap Analysis ===")

ds_dt = primary["ds_dt"]
sigma_pred = primary["total_noise"]
sigma_meas = primary["meas_nedt_mK"] / 1000.0 * ds_dt  # Convert NEDT back to noise

sigma_missing_sq = sigma_meas ** 2 - sigma_pred ** 2
if sigma_missing_sq > 0:
    sigma_missing = math.sqrt(sigma_missing_sq)
    nedt_missing = sigma_missing / ds_dt * 1000.0
else:
    sigma_missing = 0.0
    nedt_missing = 0.0

print(f"  σ_predicted:    {sigma_pred:.2f} e⁻ RMS")
print(f"  σ_measured:     {sigma_meas:.2f} e⁻ RMS  (from NEDT = {primary['meas_nedt_mK']:.1f} mK)")
print(f"  σ_missing:      {sigma_missing:.2f} e⁻ RMS  (RSS gap)")
print(f"  NEDT_missing:   {nedt_missing:.2f} mK")
print(f"")
print(f"  How much would each noise term need to increase to close the gap?")
print(f"  (i.e., if σ_i increased by Δσ so that the new RSS includes σ_missing)")
print(f"")
print(f"  {'Noise Term':<25s}  {'Current σ [e⁻]':>15s}  {'Required σ [e⁻]':>16s}  {'Increase [%]':>13s}  {'Plausible?'}")
print(f"  {'-' * 25}  {'-' * 15}  {'-' * 16}  {'-' * 13}  {'-' * 20}")

for name, sigma_i in sorted(primary["noise_dict"].items(), key=lambda x: x[1], reverse=True):
    # If this term alone absorbed all the missing noise:
    # new_σ_i² = σ_i² + σ_missing²
    new_sigma_i = math.sqrt(sigma_i ** 2 + sigma_missing ** 2)
    if sigma_i > 0:
        increase_pct = (new_sigma_i / sigma_i - 1.0) * 100.0
    else:
        increase_pct = float("inf")

    # Plausibility heuristic
    if increase_pct < 10:
        plausibility = "Very likely"
    elif increase_pct < 30:
        plausibility = "Possible"
    elif increase_pct < 100:
        plausibility = "Unlikely alone"
    else:
        plausibility = "Cannot explain alone"

    print(f"  {name:<25s}  {sigma_i:>15.2f}  {new_sigma_i:>16.2f}  {increase_pct:>13.1f}  {plausibility}")

# ---------------------------------------------------------------------------
# Step 7: Sensitivity analysis — d(NEDT)/d(each parameter)
# ---------------------------------------------------------------------------
# Perturb each input parameter by ±1% and measure the change in NEDT.
# Sensitivity = (NEDT(p+δ) − NEDT(p−δ)) / (2δ)

print(f"\n=== NEDT Sensitivity Analysis ===")
print(f"  Perturbing each parameter by ±1% around as-built values.")
print(f"  Sensitivity reported as Δ(NEDT) in mK per 1% change in parameter.")

T_primary = 298.15  # K
PERT_FRAC = 0.01  # 1% perturbation

# Parameters to perturb: (name, config_path, base_value, unit)
params_to_sweep = [
    ("QE", "qe_value", qe, "fraction"),
    ("Dark current", "dark_rate_e_per_s", dark_rate, "e⁻/s"),
    ("Read noise", "read_noise_e_rms", read_noise, "e⁻ RMS"),
    ("Integration time", "integration_time_s", t_int_s, "s"),
    ("Optical transmission", "transmission_scalar", transmission, "fraction"),
    ("f-number (via focal length)", "focal_length_m", focal_length_m, "m"),
]

# Map parameter names to config dict paths
PARAM_SECTION = {
    "qe_value": "detector",
    "dark_rate_e_per_s": "detector",
    "read_noise_e_rms": "readout",
    "integration_time_s": "spectral_integration",
    "transmission_scalar": "optics",
    "focal_length_m": "optics",
}


def compute_nedt_for_config(config: dict) -> float:
    """Run RADIANT and compute NEDT via finite difference."""
    T = T_primary
    r_c = Sensor.from_dict(config).evaluate()
    # Perturb target temperature for dS/dT
    cfg_p = config.copy()
    cfg_p = {k: (dict(v) if isinstance(v, dict) else v) for k, v in config.items()}
    cfg_p["source"] = dict(config["source"])
    cfg_p["source"]["target"] = dict(config["source"]["target"])
    cfg_p["source"]["target"]["temperature"] = T + DELTA_T
    r_p = Sensor.from_dict(cfg_p).evaluate()

    cfg_m = {k: (dict(v) if isinstance(v, dict) else v) for k, v in config.items()}
    cfg_m["source"] = dict(config["source"])
    cfg_m["source"]["target"] = dict(config["source"]["target"])
    cfg_m["source"]["target"]["temperature"] = T - DELTA_T
    r_m = Sensor.from_dict(cfg_m).evaluate()

    s_c = r_c.stage_outputs["readout"]["signal_e_final"]
    s_p = r_p.stage_outputs["readout"]["signal_e_final"]
    s_m = r_m.stage_outputs["readout"]["signal_e_final"]
    ds_dt_local = (s_p - s_m) / (2.0 * DELTA_T)

    noise_total = math.sqrt(sum(nt.value_e ** 2 for nt in r_c.noise_terms))
    return (noise_total / ds_dt_local * 1000.0) if ds_dt_local > 0 else float("inf")


print(f"\n  {'Parameter':<30s}  {'Base Value':>14s}  {'NEDT- [mK]':>11s}  {'NEDT+ [mK]':>11s}"
      f"  {'Δ/1% [mK]':>10s}  {'Direction'}")
print(f"  {'-' * 30}  {'-' * 14}  {'-' * 11}  {'-' * 11}"
      f"  {'-' * 10}  {'-' * 20}")

sensitivities: list[dict] = []

for param_name, config_key, base_val, unit in params_to_sweep:
    section = PARAM_SECTION[config_key]
    delta = base_val * PERT_FRAC

    # Config with -1%
    cfg_lo = make_config(T_primary)
    cfg_lo[section] = dict(cfg_lo[section])
    cfg_lo[section][config_key] = base_val - delta
    nedt_lo = compute_nedt_for_config(cfg_lo)

    # Config with +1%
    cfg_hi = make_config(T_primary)
    cfg_hi[section] = dict(cfg_hi[section])
    cfg_hi[section][config_key] = base_val + delta
    nedt_hi = compute_nedt_for_config(cfg_hi)

    sensitivity = (nedt_hi - nedt_lo) / 2.0  # mK per 1% change
    direction = "↑ param → ↑ NEDT" if sensitivity > 0 else "↑ param → ↓ NEDT"

    sensitivities.append({
        "param": param_name,
        "base_val": base_val,
        "unit": unit,
        "nedt_lo": nedt_lo,
        "nedt_hi": nedt_hi,
        "sensitivity_mK_per_pct": sensitivity,
    })

    print(f"  {param_name:<30s}  {base_val:>14.4f}  {nedt_lo:>11.3f}  {nedt_hi:>11.3f}"
          f"  {sensitivity:>10.4f}  {direction}")

# Rank by sensitivity magnitude
print(f"\n  Ranked by |sensitivity| (most impactful first):")
ranked = sorted(sensitivities, key=lambda x: abs(x["sensitivity_mK_per_pct"]), reverse=True)
for i, s in enumerate(ranked, 1):
    print(f"    {i}. {s['param']}: {abs(s['sensitivity_mK_per_pct']):.4f} mK per 1% change")

# ---------------------------------------------------------------------------
# Step 8: Nominal vs. as-built comparison
# ---------------------------------------------------------------------------
# Run with nominal parameters to see how the as-built deltas affect NEDT.

print(f"\n=== Nominal vs. As-Built NEDT ===")

# Nominal parameters (from Sheet 2 comparison table)
nominal_qe = 0.72
nominal_dark = 100.0
nominal_read = 12.0
nominal_f_number = 4.0
nominal_focal_length_m = aperture_m * nominal_f_number
nominal_transmission = 0.73

cfg_nominal = make_config(T_primary)
cfg_nominal["detector"] = dict(cfg_nominal["detector"])
cfg_nominal["detector"]["qe_value"] = nominal_qe
cfg_nominal["detector"]["dark_rate_e_per_s"] = nominal_dark
cfg_nominal["optics"] = dict(cfg_nominal["optics"])
cfg_nominal["optics"]["focal_length_m"] = nominal_focal_length_m
cfg_nominal["optics"]["transmission_scalar"] = nominal_transmission
cfg_nominal["readout"] = dict(cfg_nominal["readout"])
cfg_nominal["readout"]["read_noise_e_rms"] = nominal_read

nedt_nominal = compute_nedt_for_config(cfg_nominal)

print(f"  NEDT (nominal params):   {nedt_nominal:.2f} mK")
print(f"  NEDT (as-built params):  {primary['pred_nedt_mK']:.2f} mK")
print(f"  NEDT (measured):         {primary['meas_nedt_mK']:.1f} mK")
print(f"")
print(f"  Nominal → As-built degradation: {primary['pred_nedt_mK'] - nedt_nominal:.2f} mK")
print(f"  As-built → Measured gap:        {primary['meas_nedt_mK'] - primary['pred_nedt_mK']:.2f} mK")
print(f"")
print(f"  The as-built parameters (lower QE, higher dark/read noise, slower optics)")
print(f"  explain {primary['pred_nedt_mK'] - nedt_nominal:.2f} mK of degradation from nominal.")
print(f"  The remaining {primary['meas_nedt_mK'] - primary['pred_nedt_mK']:.2f} mK gap is likely due to:")
print(f"    - Unmodeled noise sources (e.g., ROIC glow, substrate luminescence)")
print(f"    - IPC-related noise redistribution effects")
print(f"    - Blackbody temperature calibration uncertainty")
print(f"    - Spatial non-uniformity (NEDT measured as spatial σ across ROI)")

# ---------------------------------------------------------------------------
# Step 9: Write results to output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb_out = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: NEDT Comparison ---
ws1 = wb_out.active
ws1.title = "NEDT Comparison"
ws1["A1"] = "Scenario 7.1: Predicted vs. Measured NEDT"
ws1["A1"].font = Font(bold=True, size=14)

headers = ["BB T [°C]", "BB T [K]", "Measured NEDT [mK]", "Predicted NEDT [mK]",
           "Δ [mK]", "Signal [e⁻]", "dS/dT [e⁻/K]", "σ_total [e⁻]", "SNR"]
col_widths = [12, 12, 20, 20, 10, 14, 16, 14, 10]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    ws1.column_dimensions[cell.column_letter].width = w

for i, r in enumerate(pred_results, 4):
    delta = r["meas_nedt_mK"] - r["pred_nedt_mK"]
    vals = [r["bb_temp_C"], round(r["bb_temp_K"], 2), r["meas_nedt_mK"],
            round(r["pred_nedt_mK"], 2), round(delta, 2),
            round(r["signal_e"]), round(r["ds_dt"], 1),
            round(r["total_noise"], 2), round(r["snr"], 1)]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=i, column=col, value=v)
        cell.border = thin_border

    # Color the delta column
    delta_cell = ws1.cell(row=i, column=5)
    if abs(delta) < 2.0:
        delta_cell.fill = pass_fill
    elif abs(delta) < 5.0:
        delta_cell.fill = warn_fill
    else:
        delta_cell.fill = fail_fill

# --- Sheet 2: Noise Breakdown ---
ws2 = wb_out.create_sheet("Noise Breakdown")
ws2["A1"] = "NEDT Breakdown by Noise Term (25°C / 298.15 K)"
ws2["A1"].font = Font(bold=True, size=14)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16

noise_headers = ["Noise Term", "σ [e⁻ RMS]", "NEDT_i [mK]", "Fraction [%]"]
for col, h in enumerate(noise_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border

for i, (name, nedt_i) in enumerate(sorted_terms, 4):
    sigma = primary["noise_dict"][name]
    frac_pct = (nedt_i ** 2 / nedt_squared_total * 100.0) if nedt_squared_total > 0 else 0.0
    ws2.cell(row=i, column=1, value=name).border = thin_border
    ws2.cell(row=i, column=2, value=round(sigma, 2)).border = thin_border
    ws2.cell(row=i, column=3, value=round(nedt_i, 3)).border = thin_border
    ws2.cell(row=i, column=4, value=round(frac_pct, 1)).border = thin_border

# --- Sheet 3: Sensitivity ---
ws3 = wb_out.create_sheet("Sensitivity")
ws3["A1"] = "NEDT Sensitivity to Input Parameters"
ws3["A1"].font = Font(bold=True, size=14)

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 16

sens_headers = ["Parameter", "Base Value", "Unit", "NEDT- [mK]", "NEDT+ [mK]", "Δ/1% [mK]"]
for col, h in enumerate(sens_headers, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border

for i, s in enumerate(sensitivities, 4):
    ws3.cell(row=i, column=1, value=s["param"]).border = thin_border
    ws3.cell(row=i, column=2, value=round(s["base_val"], 4)).border = thin_border
    ws3.cell(row=i, column=3, value=s["unit"]).border = thin_border
    ws3.cell(row=i, column=4, value=round(s["nedt_lo"], 3)).border = thin_border
    ws3.cell(row=i, column=5, value=round(s["nedt_hi"], 3)).border = thin_border
    ws3.cell(row=i, column=6, value=round(s["sensitivity_mK_per_pct"], 4)).border = thin_border

# --- Sheet 4: Summary ---
ws4 = wb_out.create_sheet("Summary")
ws4["A1"] = "Scenario 7.1: NEDT Reconciliation Summary"
ws4["A1"].font = Font(bold=True, size=14)
ws4.column_dimensions["A"].width = 45
ws4.column_dimensions["B"].width = 25

summary = [
    ("Test Conditions", ""),
    ("Blackbody temperature", f"{T_primary:.2f} K (25°C)"),
    ("Blackbody emissivity", f"{bb_emiss}"),
    ("Atmosphere", "exo (vacuum / TVAC)"),
    ("Detector temperature", f"{operating_temp_K:.1f} K"),
    ("Integration time", f"{t_int_s * 1000:.1f} ms"),
    ("", ""),
    ("Results", ""),
    ("Measured NEDT [mK]", f"{primary['meas_nedt_mK']:.1f}"),
    ("Predicted NEDT (as-built) [mK]", f"{primary['pred_nedt_mK']:.2f}"),
    ("Predicted NEDT (nominal) [mK]", f"{nedt_nominal:.2f}"),
    ("Gap (measured − predicted) [mK]", f"{primary['meas_nedt_mK'] - primary['pred_nedt_mK']:.2f}"),
    ("", ""),
    ("Signal at 298.15 K [e⁻]", f"{primary['signal_e']:,.0f}"),
    ("dS/dT [e⁻/K]", f"{primary['ds_dt']:.1f}"),
    ("Total noise [e⁻ RMS]", f"{primary['total_noise']:.2f}"),
    ("SNR", f"{primary['snr']:.1f}"),
    ("Missing noise σ [e⁻ RMS]", f"{sigma_missing:.2f}"),
    ("", ""),
    ("Dominant Noise Term", sorted_terms[0][0] if sorted_terms else "N/A"),
    ("Most Sensitive Parameter", ranked[0]["param"] if ranked else "N/A"),
]

for i, (label, value) in enumerate(summary, 3):
    cell_a = ws4.cell(row=i, column=1, value=label)
    ws4.cell(row=i, column=2, value=value)
    if label and not value:
        cell_a.font = Font(bold=True, size=11)

wb_out.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
