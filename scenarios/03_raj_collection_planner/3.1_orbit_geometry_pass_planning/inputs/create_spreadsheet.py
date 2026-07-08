#!/usr/bin/env python3
"""Create Raj's inputs for Scenario 3.1: orbit → geometry / pass planning.

Emits ``raj_orbit_sensor.xlsx`` — the orbit definition (a circular
sun-synchronous LEO), the sensor optical/detector configuration, and the
collection constraints (max slew angle, NIIRS floor). Raj thinks in orbit
altitude and pointing angle, not in RADIANT canonical units; the run
script converts.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MissionConfig"
ws["A1"] = "Scenario 3.1: orbit + sensor + collection constraints"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [30, 14, 12, 46]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="C55A11")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

section_fill = PatternFill("solid", fgColor="FCE4D6")
ROWS = [
    ("ORBIT", None, None, None),
    ("Altitude", 600.0, "km", "Circular sun-synchronous LEO"),
    ("COLLECTION", None, None, None),
    ("Max slew (off-nadir)", 45.0, "deg", "Spacecraft agility limit"),
    ("NIIRS floor", 6.0, "—", "Minimum acceptable interpretability"),
    ("SENSOR", None, None, None),
    ("Aperture diameter", 50.0, "cm", ""),
    ("Focal length", 6.0, "m", ""),
    ("Pixel pitch", 6.5, "µm", ""),
    ("Cross-track pixels", 8000, "—", "Push-broom array width"),
    ("Optical transmission", 85.0, "%", ""),
    ("QE (pan band)", 85.0, "%", ""),
    ("Pan band min", 450.0, "nm", ""),
    ("Pan band max", 700.0, "nm", ""),
    ("Integration time", 0.5, "ms", ""),
    ("Dark current", 50.0, "e⁻/s", ""),
    ("Detector temperature", 280.0, "K", ""),
    ("Read noise", 20.0, "e⁻ RMS", ""),
    ("Full well", 30000.0, "e⁻", ""),
    ("System gain", 8.0, "e⁻/DN", ""),
    ("ADC resolution", 12, "bits", ""),
    ("Target reflectance", 0.30, "—", "Mixed scene"),
    ("Solar zenith", 35.0, "deg", "Representative daytime illumination"),
]
r = 4
for name, value, unit, note in ROWS:
    if value is None and unit is None:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=11)
        for cc in range(1, 5):
            ws.cell(row=r, column=cc).fill = section_fill
    else:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "raj_orbit_sensor.xlsx"
wb.save(out)
print(f"Wrote {out}")
