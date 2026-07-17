"""XLSX workbook export (Tier-2 GT-4, owner decision D2 — openpyxl in the gui extra).

One workbook, three sheets built purely from public API surfaces:

- **Config** — every resolved parameter (dot-path, input value, input unit) via
  ``parameter_defs()`` + ``get_input()``.
- **Metrics** — the last result's ``to_records()`` (name / value / unit / description).
- **Sweep** — the last sweep, when one exists: 1-D (param + metrics columns, mirroring
  ``SweepResult.to_csv``) or 2-D long form.

Every numeric cell keeps full precision (numbers, not formatted strings); units get
their own column (R-UNITS). Rule 30: openpyxl owns the file encoding.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from radiant.core.exceptions import RadiantError

if TYPE_CHECKING:
    from radiant.api.sensor import Sensor
    from radiant.io.results import ChainResult


def export_workbook(
    path: str | Path,
    sensor: Sensor,
    result: ChainResult | None,
    sweep: Any | None = None,
) -> Path:
    """Write the config/metrics/sweep workbook to *path* and return it."""
    from openpyxl import Workbook

    book = Workbook()

    config_sheet = book.active
    config_sheet.title = "Config"
    config_sheet.append(["parameter", "value", "unit"])
    for dotpath, pdef in sorted(sensor.parameter_defs().items()):
        try:
            value = sensor.get_input(dotpath)
        except (KeyError, RadiantError):
            value = None
        config_sheet.append([dotpath, value, pdef.input_unit])

    if result is not None:
        metrics_sheet = book.create_sheet("Metrics")
        metrics_sheet.append(["name", "value", "unit", "description"])
        for record in result.to_records():
            metrics_sheet.append(
                [record["name"], record["value"], record["unit"], record["description"]]
            )

    if sweep is not None:
        sweep_sheet = book.create_sheet("Sweep")
        if hasattr(sweep, "grid"):  # Sweep2DResult — long form
            sweep_sheet.append([sweep.param1_name, sweep.param2_name, sweep.metric_name])
            for i, v1 in enumerate(sweep.values1):
                for j, v2 in enumerate(sweep.values2):
                    sweep_sheet.append([float(v1), float(v2), float(sweep.grid[i, j])])
        else:  # SweepResult
            extra = []
            if sweep.results:
                extra = sorted(
                    set().union(*(set(r.metrics) for r in sweep.results)) - {sweep.metric_name}
                )
            sweep_sheet.append([sweep.param_name, sweep.metric_name, *extra])
            for i, (v, m) in enumerate(zip(sweep.values, sweep.metric_values, strict=True)):
                extras = (
                    [float(sweep.results[i].metrics.get(name, float("nan"))) for name in extra]
                    if sweep.results
                    else []
                )
                sweep_sheet.append([float(v), float(m), *extras])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(out)
    return out


__all__ = ["export_workbook"]
