#!/usr/bin/env python3
"""Create Karen's inputs for Scenario 7.2: radiometric calibration verification.

Emits:

- ``karen_asbuilt_sensor.xlsx`` — the as-built bench sensor spec in
  vendor/lab units (cm, %, °C, nm, ms, e⁻/DN).
- ``karen_calibration_dn.csv``  — the measured calibration run: mean DN
  at five calibrated blackbody set points (``T_K, DN_measured``), read
  by ``radiant.io.measurement.load_measured_curve``.

The "measured" DN values are synthesized from an INDEPENDENT first-
principles model of the real instrument (Planck band radiance × a
radiometric gain), deliberately including the imperfections a real
calibration uncovers:

- the real system gain is 1.8% higher than the as-built spec claims,
- a constant instrument offset (self-emission + dark) of ~46 DN,
- a mild quadratic non-linearity (−0.6% of full scale at the top point),
- 0.1% RMS measurement noise (100-frame averages).

RADIANT never sees these truth parameters — the scenario's job is to
expose them as calibration residuals.

Run:  python create_spreadsheet.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# As-built sensor workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "As-Built Sensor"

ws["A1"] = "Scenario 7.2: Radiometric Calibration — As-Built Bench Sensor"
ws["A1"].font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 52

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = header_font
    ws[f"{col}3"].fill = header_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Aperture diameter", 15.0, "cm", ""),
    ("Focal length", 30.0, "cm", "f/2.0"),
    ("Optical transmission", 72.0, "%", "Witness-sample measurement"),
    ("Optics emissivity", 28.0, "%", "1 - transmission (Kirchhoff, reflective train)"),
    ("Nearfield fraction", 0.05, "—", "Cold-stop leakage (from scenario 7.4 campaign)"),
    ("Optics temperature", 20.0, "°C", "Bench ambient"),
    ("Pixel pitch", 15.0, "µm", ""),
    ("Quantum efficiency", 75.0, "%", "Band-average"),
    ("Dark current", 50000.0, "e⁻/s", "At 77 K"),
    ("Operating temperature", 77.0, "K", ""),
    ("Cold filter passband", "3700–4900", "nm", ""),
    ("Integration time", 0.25, "ms", "Calibration mode"),
    ("Read noise (CDS)", 30.0, "e⁻ RMS", ""),
    ("System gain", 125.0, "e⁻/DN", "As-built spec value (nominal)"),
    ("ADC resolution", 14, "bits", ""),
    ("Full well capacity", 2000000, "e⁻", ""),
    ("Lab ambient temperature", 22.0, "°C", "Chamber background"),
    ("Lab ambient emissivity", 0.95, "—", ""),
    ("Blackbody emissivity", 0.998, "—", "NIST-traceable cavity source"),
]
for r, (name, value, unit, note) in enumerate(ROWS, start=4):
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)

out = HERE / "karen_asbuilt_sensor.xlsx"
wb.save(out)
print(f"Wrote {out}")

# ---------------------------------------------------------------------------
# Measured calibration DN — independent truth model
# ---------------------------------------------------------------------------

H = 6.62607015e-34   # J·s
C = 2.99792458e8     # m/s
K_B = 1.380649e-23   # J/K

BB_TEMPS_K = [280.0, 300.0, 320.0, 340.0, 360.0]

# As-built radiometric constants (matching the workbook above)
TAU = 0.72
OMEGA = math.pi / (4.0 * 2.0**2)         # f/2 pixel solid angle [sr]
A_PIX = (15.0e-6) ** 2                    # [m²]
QE = 0.75
T_INT = 0.25e-3                           # [s]
BB_EMISS = 0.998

# Truth-model imperfections (RADIANT never sees these)
GAIN_SPEC = 125.0                         # e⁻/DN, as-built spec
GAIN_TRUE = GAIN_SPEC / 1.018             # real gain 1.8% higher responsivity in DN
OFFSET_DN = 46.0                          # instrument self-emission + dark [DN]
NONLIN_C2 = -3.5e-9                       # quadratic term [DN per DN²]
NOISE_FRAC = 0.001                        # 0.1% RMS (100-frame average)

wl_um = np.linspace(3.7, 4.9, 2000)
wl_m = wl_um * 1e-6


def photon_signal_e(T: float) -> float:
    """Photon-integral signal [e⁻] for the bench geometry at BB temp T."""
    B = 2.0 * H * C**2 / wl_m**5 / (np.expm1(H * C / (wl_m * K_B * T))) * 1e-6
    phot = np.trapezoid(BB_EMISS * B * wl_m / (H * C), wl_um)
    return float(phot * TAU * OMEGA * A_PIX * QE * T_INT)


rng = np.random.default_rng(seed=72)
lines = [
    "# Radiometric calibration run 2026-06-12 — bench MWIR sensor",
    "# 100-frame mean DN per set point, NIST-traceable blackbody",
    "T_K,DN_measured",
]
for T in BB_TEMPS_K:
    dn_lin = photon_signal_e(T) / GAIN_TRUE + OFFSET_DN
    dn = dn_lin + NONLIN_C2 * dn_lin**2
    dn *= 1.0 + rng.normal(0.0, NOISE_FRAC)
    lines.append(f"{T},{dn:.1f}")

csv_out = HERE / "karen_calibration_dn.csv"
csv_out.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"Wrote {csv_out}")
for line in lines[3:]:
    print("  ", line)
