"""Tests for the GT-4 export enablement (FW-B surfaces + the D2 XLSX workbook)."""

from __future__ import annotations

import warnings
from pathlib import Path

import numpy as np
import pytest

pytest.importorskip("PySide6", reason="GUI tests require the optional 'gui' extra")
pytest.importorskip("openpyxl", reason="XLSX export requires openpyxl (gui extra, D2)")

from radiant.api.sensor import Sensor  # noqa: E402
from radiant.gui.main_window import RADIANTMainWindow  # noqa: E402
from radiant.gui.xlsx_export import export_workbook  # noqa: E402

_REPO = Path(__file__).resolve()
while not (_REPO / "pyproject.toml").exists():
    _REPO = _REPO.parent
_EXAMPLE = _REPO / "examples" / "mwir_leo_minimal.yaml"
_WAIT_MS = 20000


def _window(qtbot) -> RADIANTMainWindow:  # type: ignore[no-untyped-def]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        window = RADIANTMainWindow(Sensor.load(_EXAMPLE))
    qtbot.addWidget(window)
    with qtbot.waitSignal(window.evaluationFinished, timeout=_WAIT_MS):
        pass
    return window


class TestEnablement:
    def test_export_actions_arm_with_sensor_and_result(self, qtbot) -> None:  # type: ignore[no-untyped-def]
        window = _window(qtbot)
        assert window.action("file.export_resolved_yaml").isEnabled()
        assert window.action("file.export_metrics_csv").isEnabled()
        assert window.action("file.export_xlsx").isEnabled()
        # No sweep yet — its export stays off.
        assert not window.action("file.export_sweep_csv").isEnabled()


class TestExports:
    def test_resolved_yaml_export(self, qtbot, tmp_path, monkeypatch) -> None:  # type: ignore[no-untyped-def]
        window = _window(qtbot)
        dest = tmp_path / "resolved.yaml"
        from radiant.gui import main_window as mw

        monkeypatch.setattr(
            mw.QFileDialog, "getSaveFileName", staticmethod(lambda *a, **k: (str(dest), ""))
        )
        window.action("file.export_resolved_yaml").trigger()
        text = dest.read_text(encoding="utf-8")
        # Resolved scope carries defaults the inputs scope omits.
        assert "jitter_rms_urad" in text

    def test_metrics_csv_export(self, qtbot, tmp_path, monkeypatch) -> None:  # type: ignore[no-untyped-def]
        window = _window(qtbot)
        dest = tmp_path / "metrics.csv"
        from radiant.gui import main_window as mw

        monkeypatch.setattr(
            mw.QFileDialog, "getSaveFileName", staticmethod(lambda *a, **k: (str(dest), ""))
        )
        window.action("file.export_metrics_csv").trigger()
        assert dest.read_text(encoding="utf-8").startswith("name,value,unit")

    def test_workbook_sheets_and_values(self, qtbot, tmp_path) -> None:  # type: ignore[no-untyped-def]
        import openpyxl

        sensor = Sensor.from_yaml(_EXAMPLE)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            result = sensor.evaluate()
            sweep = sensor.sweep(
                "optics.aperture_diameter_m", np.linspace(0.25, 0.35, 3), keep_results=False
            )
        out = export_workbook(tmp_path / "wb.xlsx", sensor, result, sweep)
        book = openpyxl.load_workbook(out)
        assert set(book.sheetnames) == {"Config", "Metrics", "Sweep"}
        metrics = {row[0].value: row[1].value for row in book["Metrics"].iter_rows(min_row=2)}
        assert metrics["snr"] == pytest.approx(result.metrics["snr"], rel=1e-12)
        sweep_rows = list(book["Sweep"].iter_rows(min_row=2))
        assert len(sweep_rows) == 3
