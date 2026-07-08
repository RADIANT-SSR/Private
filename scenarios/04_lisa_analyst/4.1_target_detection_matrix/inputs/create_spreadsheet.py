#!/usr/bin/env python3
"""Create Lisa's inputs for Scenario 4.1: target detection matrix.

Emits:

- ``lisa_target_library.xlsx`` — 12-row program-office target list
  (columns: target_name, length_m, width_m, height_m, temperature_K,
  emissivity, material). projected_area_m2 is NOT a column — the reader
  derives it (length × width).
- ``sensor_a_mwir_smallsat.yaml``  — 18 cm MWIR smallsat imager
- ``sensor_b_mwir_flagship.yaml``  — 50 cm MWIR flagship imager
- ``sensor_c_lwir_wide.yaml``      — 35 cm LWIR wide-area imager;
  deliberately carries the DEPRECATED ``optics.cold_stop_efficiency``
  name (the catalog says the sensor library "may have outdated parameter
  names" — RADIANT's deprecated-alias machinery must absorb it)

Atmosphere conditions are defined in the run script (they are per-run
settings, not sensor properties).

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# Target library — 12 targets
# ---------------------------------------------------------------------------

TARGETS = [
    # name,                length, width, height, T_K,   eps,  material
    ("MBT tank",              7.9,   3.6,   2.4, 310.0, 0.90, "painted steel"),
    ("APC",                   6.5,   2.8,   2.5, 306.0, 0.90, "painted steel"),
    ("Cargo truck",           8.0,   2.5,   3.2, 301.0, 0.92, "painted steel"),
    ("Technical (pickup)",    5.3,   1.9,   1.8, 303.0, 0.88, "painted steel"),
    ("SAM TEL",              11.5,   3.1,   3.4, 305.0, 0.90, "painted steel"),
    ("Towed artillery",       9.5,   2.8,   2.0, 296.0, 0.85, "painted steel"),
    ("Patrol boat",          25.0,   5.8,   4.5, 299.0, 0.85, "painted steel"),
    ("Fast attack craft",    56.0,   8.5,   6.0, 302.0, 0.85, "painted steel"),
    ("Transport aircraft",   40.0,  40.0,  11.8, 295.0, 0.30, "bare aluminum"),
    ("Fighter aircraft",     15.0,  10.0,   4.8, 297.0, 0.35, "low-e coating"),
    ("Fuel bladder farm",    30.0,  20.0,   2.0, 298.0, 0.95, "rubberized fabric"),
    ("Small UAV (parked)",    3.0,   4.0,   1.0, 294.0, 0.80, "composite"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Target Library"

headers = ["target_name", "length_m", "width_m", "height_m",
           "temperature_K", "emissivity", "material"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
for row in TARGETS:
    ws.append(list(row))
for col_letter, width in zip("ABCDEFG", [22, 10, 10, 10, 14, 11, 20]):
    ws.column_dimensions[col_letter].width = width

out = HERE / "lisa_target_library.xlsx"
wb.save(out)
print(f"Wrote {out} ({len(TARGETS)} targets)")

# ---------------------------------------------------------------------------
# Sensor library — 3 YAML configs
# ---------------------------------------------------------------------------

SENSOR_A = """\
# Sensor A: MWIR smallsat imager (18 cm) — Lisa's constellation workhorse
source:
  target:
    temperature: 300.0        # K (overridden per target by the batch)
    emissivity: 0.90

atmosphere:
  model: simple
  standard_atmosphere: midlat_summer

geometry:
  sensor_altitude_m: 500000.0  # m (500 km)
  path_zenith_rad: 0.0

optics:
  aperture_diameter_m: 0.18   # m
  focal_length_m: 0.45        # m  (f/2.5)
  transmission_scalar: 0.70

detector:
  pixel_pitch_x_um: 15.0      # um
  pixel_pitch_y_um: 15.0      # um
  qe_value: 0.72
  dark_rate_e_per_s: 5000.0   # e-/s (80 K MWIR HgCdTe)
  detector_temperature_K: 80.0

spectral_integration:
  filter_min_um: 3.6           # um
  filter_max_um: 4.9           # um
  integration_time_s: 0.004    # s

readout:
  read_noise_e_rms: 35.0      # e- RMS
  gain_e_per_dn: 12.0         # e-/DN
  adc_bits: 14
  full_well_capacity_e: 800000.0
"""

SENSOR_B = """\
# Sensor B: MWIR flagship imager (50 cm) — the program's exquisite asset
source:
  target:
    temperature: 300.0        # K (overridden per target by the batch)
    emissivity: 0.90

atmosphere:
  model: simple
  standard_atmosphere: midlat_summer

geometry:
  sensor_altitude_m: 500000.0  # m (500 km)
  path_zenith_rad: 0.0

optics:
  aperture_diameter_m: 0.50   # m
  focal_length_m: 1.50        # m  (f/3.0)
  transmission_scalar: 0.75

detector:
  pixel_pitch_x_um: 12.0      # um
  pixel_pitch_y_um: 12.0      # um
  qe_value: 0.78
  dark_rate_e_per_s: 2000.0   # e-/s (75 K MWIR HgCdTe)
  detector_temperature_K: 75.0

spectral_integration:
  filter_min_um: 3.6           # um
  filter_max_um: 4.9           # um
  integration_time_s: 0.004    # s

readout:
  read_noise_e_rms: 25.0      # e- RMS
  gain_e_per_dn: 10.0         # e-/DN
  adc_bits: 14
  full_well_capacity_e: 900000.0
"""

SENSOR_C = """\
# Sensor C: LWIR wide-area imager (35 cm)
# NOTE: this config predates the Gap 12 rename and still uses the
# OUTDATED parameter name optics.cold_stop_efficiency — RADIANT accepts
# it through the deprecated-alias mechanism (with a DeprecationWarning).
source:
  target:
    temperature: 300.0        # K (overridden per target by the batch)
    emissivity: 0.90

atmosphere:
  model: simple
  standard_atmosphere: midlat_summer

geometry:
  sensor_altitude_m: 500000.0  # m (500 km)
  path_zenith_rad: 0.0

optics:
  aperture_diameter_m: 0.35   # m
  focal_length_m: 0.70        # m  (f/2.0)
  transmission_scalar: 0.65
  cold_stop_efficiency: 1.0   # OUTDATED NAME (now optics.nearfield_fraction)

detector:
  pixel_pitch_x_um: 17.0      # um
  pixel_pitch_y_um: 17.0      # um
  qe_value: 0.65
  dark_rate_e_per_s: 200000.0 # e-/s (60 K LWIR HgCdTe)
  detector_temperature_K: 60.0

spectral_integration:
  filter_min_um: 8.0           # um
  filter_max_um: 11.5          # um
  integration_time_s: 0.002    # s

readout:
  read_noise_e_rms: 40.0      # e- RMS
  gain_e_per_dn: 40.0         # e-/DN
  adc_bits: 14
  full_well_capacity_e: 3000000.0
"""

for name, text in [
    ("sensor_a_mwir_smallsat.yaml", SENSOR_A),
    ("sensor_b_mwir_flagship.yaml", SENSOR_B),
    ("sensor_c_lwir_wide.yaml", SENSOR_C),
]:
    (HERE / name).write_text(text, encoding="utf-8")
    print(f"Wrote {name}")
