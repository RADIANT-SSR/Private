"""Scenario 1.3: Dual-Band MWIR/LWIR Comparison — Wildfire Detection Trade.

Sarah must choose between a MWIR (3.5–5.0 µm) and a LWIR (8–12 µm)
HgCdTe option for a wildfire mission: 10 km airborne platform, a 5 m²
hotspot at ~600 K against a 300 K conifer forest. The vendor data
arrives as an Excel comparison table (not YAML), and the forest
emissivity as a JPL/NASA ASTER-library text file.

This script:
  1. Loads the forest spectrum with radiant.io.aster_library
     (ε(λ) = 1 − ρ(λ), per-band averages: the SAME canopy has different
     emissivity in the two bands — 0.953 MWIR vs 0.982 LWIR).
  2. Loads the two detector options + shared platform from the workbook.
  3. Runs both bands through the chain: sub-pixel hotspot (regime
     override, as in scenario 4.1, so in-pixel background photons and
     scene clutter are counted) at nadir from 10 km through a
     mid-latitude summer atmosphere.
  4. Side-by-side: SNR, contrast SNR, SCNR (script-side, clutter
     included), NEDT (with the Gap 43 caveat), full noise budgets, well
     fill and ADC state.
  5. Spectral contrast ΔL(λ) = ε_t·B(λ,T_t) − ε_bg(λ)·B(λ,300 K)
     (hand Planck) with both bands shaded — the classic band-trade plot.
  6. Hotspot-temperature sweep 400–1200 K: SCNR and detection
     probability per band (Gaussian threshold model at P_fa = 1e-6),
     with well/ADC saturation flagged — MWIR fire products saturate on
     hot fires; LWIR keeps dynamic range. That trade IS the answer.

Usage:
    python run_dual_band_trade.py
"""

import math
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from radiant.api import Sensor
from radiant.core.constants import c, h, k_B
from radiant.io.aster_library import load_aster_spectrum

INPUTS = Path(__file__).parent.parent / "inputs"
OUTPUTS = Path(__file__).parent.parent / "outputs"
OUTPUT_FILE = OUTPUTS / "dual_band_results.xlsx"

SCNR_THRESHOLD = 4.75  # Gaussian threshold for P_fa = 1e-6 [dimensionless]

print("=" * 95)
print("  SCENARIO 1.3: Dual-Band MWIR/LWIR Wildfire Detection Trade")
print("=" * 95)

# ---------------------------------------------------------------------------
# Step 1: Forest emissivity from the ASTER library file
# ---------------------------------------------------------------------------

forest = load_aster_spectrum(INPUTS / "forest_conifer_aster.txt")
print(f"\n=== Forest background spectrum (radiant.io.aster_library) ===")
print(f"  Material: {forest.name}")
print(f"  Range: {forest.wavelength_um[0]:.1f}–{forest.wavelength_um[-1]:.1f} µm, "
      f"{len(forest.wavelength_um)} points, reflectance in "
      f"{'percent' if forest.y_units_percent else 'fraction'} (converted)")

# ---------------------------------------------------------------------------
# Step 2: Detector options + shared platform (Excel comparison table)
# ---------------------------------------------------------------------------

wb_in = openpyxl.load_workbook(INPUTS / "sarah_detector_options.xlsx")
ws_det = wb_in["Detector Options"]
det: dict[str, dict[str, float]] = {"MWIR": {}, "LWIR": {}}
for row in ws_det.iter_rows(min_row=4, max_col=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        det["MWIR"][str(row[0])] = float(row[1])
        det["LWIR"][str(row[0])] = float(row[2])

ws_shared = wb_in["Shared Platform"]
shared: dict[str, float] = {}
for row in ws_shared.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[0] is not None and row[1] is not None:
        shared[str(row[0])] = float(row[1])

bands = {
    "MWIR": (det["MWIR"]["Band minimum"], det["MWIR"]["Band maximum"]),
    "LWIR": (det["LWIR"]["Band minimum"], det["LWIR"]["Band maximum"]),
}
eps_forest = {
    b: forest.band_averaged_emissivity(*bands[b]) for b in bands
}

print(f"\n=== Detector options (vendor comparison table) ===")
print(f"  {'Parameter':<32s} {'MWIR':>12s}  {'LWIR':>12s}")
print(f"  {'-' * 32} {'-' * 12}  {'-' * 12}")
for name in det["MWIR"]:
    print(f"  {name:<32s} {det['MWIR'][name]:>12g}  {det['LWIR'][name]:>12g}")
print(f"\n=== Shared platform ===")
for name, value in shared.items():
    print(f"  {name:<32s}: {value:g}")
print(f"\n  Forest band-averaged emissivity (from ASTER curve, ε = 1 − ρ):")
print(f"    MWIR {bands['MWIR'][0]:.1f}–{bands['MWIR'][1]:.1f} µm: "
      f"ε = {eps_forest['MWIR']:.4f} [--]")
print(f"    LWIR {bands['LWIR'][0]:.1f}–{bands['LWIR'][1]:.1f} µm: "
      f"ε = {eps_forest['LWIR']:.4f} [--]")

altitude_m = shared["Platform altitude"] * 1000.0        # km → m
aperture_m = shared["Aperture diameter"] / 100.0         # cm → m
focal_m = shared["Focal length"] / 100.0                 # cm → m
gsd_m = det["MWIR"]["Pixel pitch"] * 1e-6 * altitude_m / focal_m
footprint_m2 = gsd_m**2
fill = shared["Hotspot area"] / footprint_m2

print(f"\n  Geometry: GSD = {gsd_m:.1f} m at {altitude_m / 1e3:.0f} km → pixel "
      f"footprint {footprint_m2:.0f} m²; the {shared['Hotspot area']:.0f} m² "
      f"hotspot fills {fill * 100:.0f}% of a pixel → SUB-PIXEL regime.")


# ---------------------------------------------------------------------------
# Step 3: Build and run both bands
# ---------------------------------------------------------------------------


def build_sensor(band: str, hotspot_T: float) -> Sensor:
    d = det[band]
    s = Sensor()
    s.set("source.scene_type", "sub_pixel")
    s.set("source.regime_override", "sub_pixel")  # keep in-pixel background
    s.set("source.target.temperature", hotspot_T)
    s.set("source.target.emissivity", shared["Hotspot emissivity"])
    s.set("geometry.target.projected_area_m2", shared["Hotspot area"])
    # CU-060: the sub-pixel regime weights the target by fill_fraction, NOT
    # by projected_area_m2 (which drives only the point-source A/R² path).
    # Leaving it at the default 1.0 models the hotspot as pixel-filling and
    # overstates its signal by 1/fill (~3× here).
    s.set("source.target.fill_fraction", min(1.0, fill))
    s.set("geometry.target_range_m", altitude_m)  # nadir
    s.set("source.background.temperature", shared["Forest background temperature"])
    s.set("source.background.emissivity", eps_forest[band])
    s.set("detector.clutter_sigma", shared["Scene clutter sigma"])
    s.set("atmosphere.model", "simple")
    s.set("atmosphere.standard_atmosphere", "midlat_summer")
    s.set("geometry.sensor_altitude_m", altitude_m)
    s.set("geometry.path_zenith_rad", 0.0)
    s.set("optics.aperture_diameter_m", shared["Aperture diameter"], unit="cm")
    s.set("optics.focal_length_m", shared["Focal length"], unit="cm")
    s.set("optics.transmission_scalar", shared["Optical transmission"], unit="%")
    s.set("optics.optics_temperature_K", shared["Optics temperature"] + 273.15)
    s.set("detector.pixel_pitch_x_um", d["Pixel pitch"])
    s.set("detector.pixel_pitch_y_um", d["Pixel pitch"])
    s.set("detector.qe_value", d["Quantum efficiency (band avg)"], unit="%")
    s.set("detector.dark_rate_e_per_s", d["Dark current"])
    s.set("detector.detector_temperature_K", d["Operating temperature"])
    s.set("spectral_integration.filter_min_um", d["Band minimum"])
    s.set("spectral_integration.filter_max_um", d["Band maximum"])
    s.set("spectral_integration.integration_time_s",
          d["Integration time (fire mode)"], unit="ms")
    s.set("readout.read_noise_e_rms", d["Read noise (CDS)"])
    s.set("readout.gain_e_per_dn", d["System gain"])
    s.set("readout.adc_bits", int(d["ADC resolution"]))
    s.set("readout.full_well_capacity_e", d["Full well capacity"])
    return s


def scnr_of(result) -> float:
    """|contrast| / RSS(all noise terms incl. clutter) — detection SCNR."""
    contrast_e = abs(float(result.stage_outputs["spectral_integration"]["contrast_e"]))
    total = math.sqrt(sum(nt.value_e**2 for nt in result.noise_terms))
    return contrast_e / total if total > 0 else 0.0


warnings.filterwarnings("ignore")

T_nominal = shared["Hotspot temperature (nominal)"]
results = {}
for band in ("MWIR", "LWIR"):
    print(f"\n=== Running RADIANT — {band} option at T_hotspot = {T_nominal:.0f} K ===")
    results[band] = build_sensor(band, T_nominal).evaluate()

# ---------------------------------------------------------------------------
# Step 4: Side-by-side comparison at the nominal 600 K hotspot
# ---------------------------------------------------------------------------

noise = {b: {nt.name: nt.value_e for nt in results[b].noise_terms} for b in results}
total_noise = {b: math.sqrt(sum(v**2 for v in noise[b].values())) for b in results}
contrast_e = {b: results[b].stage_outputs["spectral_integration"]["contrast_e"]
              for b in results}
signal_e = {b: results[b].stage_outputs["readout"]["signal_e_final"] for b in results}
scnr = {b: scnr_of(results[b]) for b in results}
well = {b: signal_e[b] / det[b]["Full well capacity"] * 100 for b in results}

print(f"\n{'=' * 95}")
print(f"  SIDE-BY-SIDE AT T_HOTSPOT = {T_nominal:.0f} K "
      f"(fill = {fill * 100:.0f}%, clutter σ = {shared['Scene clutter sigma']:.2f})")
print(f"{'=' * 95}")
print(f"  {'Quantity':<34s} {'MWIR':>14s}  {'LWIR':>14s}")
print(f"  {'-' * 34} {'-' * 14}  {'-' * 14}")
rows = [
    ("Pixel signal [e⁻]", *(f"{signal_e[b]:,.0f}" for b in ("MWIR", "LWIR"))),
    ("Contrast (fire − forest) [e⁻]", *(f"{contrast_e[b]:,.0f}" for b in ("MWIR", "LWIR"))),
    ("Well fill [%]", *(f"{well[b]:.1f}" for b in ("MWIR", "LWIR"))),
    ("Total noise [e⁻ RMS]", *(f"{total_noise[b]:,.1f}" for b in ("MWIR", "LWIR"))),
    ("SNR [--]", *(f"{results[b].metrics['snr']:.1f}" for b in ("MWIR", "LWIR"))),
    ("Contrast SNR [--]", *(f"{results[b].metrics.get('contrast_snr', float('nan')):.1f}"
                            for b in ("MWIR", "LWIR"))),
    ("SCNR (incl. clutter) [--]", *(f"{scnr[b]:.1f}" for b in ("MWIR", "LWIR"))),
    ("NEDT [mK] (Gap 43 approx.)", *(f"{results[b].metrics.get('nedt_K', float('nan')) * 1e3:.1f}"
                                     for b in ("MWIR", "LWIR"))),
]
for label, m, lw in rows:
    print(f"  {label:<34s} {m:>14s}  {lw:>14s}")

print(f"\n  Noise budgets [e⁻ RMS]:")
all_terms = sorted(set(noise["MWIR"]) | set(noise["LWIR"]),
                   key=lambda k: -(noise["MWIR"].get(k, 0) + noise["LWIR"].get(k, 0)))
print(f"  {'Term':<24s} {'MWIR':>12s}  {'LWIR':>12s}")
print(f"  {'-' * 24} {'-' * 12}  {'-' * 12}")
for term in all_terms:
    vm, vl = noise["MWIR"].get(term, 0.0), noise["LWIR"].get(term, 0.0)
    if vm > 0.01 or vl > 0.01:
        print(f"  {term:<24s} {vm:>12.1f}  {vl:>12.1f}")

# ---------------------------------------------------------------------------
# Step 5: Spectral contrast ΔL(λ) — hand Planck, the band-trade picture
# ---------------------------------------------------------------------------

wl_um = np.linspace(3.0, 13.0, 800)
wl_m = wl_um * 1e-6


def planck(T: float) -> np.ndarray:
    return 2.0 * h * c**2 / wl_m**5 / np.expm1(h * c / (wl_m * k_B * T)) * 1e-6


eps_bg_spectral = np.interp(wl_um, forest.wavelength_um, forest.emissivity())
dL_600 = shared["Hotspot emissivity"] * planck(600.0) - eps_bg_spectral * planck(300.0)

print(f"\n=== Spectral contrast ΔL(λ) = ε_t·B(λ,T_t) − ε_bg(λ)·B(λ,300 K) ===")
for b in ("MWIR", "LWIR"):
    lo, hi = bands[b]
    mask = (wl_um >= lo) & (wl_um <= hi)
    dl_band = float(np.trapezoid(dL_600[mask], wl_um[mask]))
    print(f"  {b}: band-integrated ΔL(600 K) = {dl_band:8.2f} W/m²/sr over "
          f"{lo:.1f}–{hi:.1f} µm")
print(f"  The 600 K Planck peak sits at {2898.0 / 600.0:.1f} µm — inside MWIR.")
print(f"  Per-pixel contrast also scales with the photon energy and the")
print(f"  detector parameters, which is why the chain comparison above, not")
print(f"  ΔL alone, decides the trade.")

# ---------------------------------------------------------------------------
# Step 6: Hotspot temperature sweep 400–1200 K — SCNR, P_d, saturation
# ---------------------------------------------------------------------------

T_sweep = np.arange(400.0, 1250.0, 100.0)
sweep_scnr: dict[str, list[float]] = {"MWIR": [], "LWIR": []}
sweep_sat: dict[str, list[bool]] = {"MWIR": [], "LWIR": []}

print(f"\n=== Hotspot temperature sweep (fire detection) ===")
print(f"  Detection: P_d from the Gaussian threshold model at P_fa = 1e-6")
print(f"  (threshold = {SCNR_THRESHOLD:.2f}σ): P_d = Q(threshold − SCNR).")
print(f"\n  {'T_fire [K]':>10s}  {'MWIR SCNR':>10s} {'sat?':>5s}  "
      f"{'LWIR SCNR':>10s} {'sat?':>5s}  {'P_d MWIR':>9s}  {'P_d LWIR':>9s}")
print(f"  {'-' * 10}  {'-' * 10} {'-' * 5}  {'-' * 10} {'-' * 5}  {'-' * 9}  {'-' * 9}")


def p_detect(scnr_val: float) -> float:
    return 0.5 * math.erfc((SCNR_THRESHOLD - scnr_val) / math.sqrt(2.0))


for T in T_sweep:
    row = []
    for band in ("MWIR", "LWIR"):
        r = build_sensor(band, float(T)).evaluate()
        s_val = scnr_of(r)
        saturated = (r.stage_outputs["readout"]["signal_e_final"]
                     >= 0.98 * det[band]["Full well capacity"])
        sweep_scnr[band].append(s_val)
        sweep_sat[band].append(saturated)
        row.append((s_val, saturated))
    print(f"  {T:>10.0f}  {row[0][0]:>10.1f} {'YES' if row[0][1] else 'no':>5s}  "
          f"{row[1][0]:>10.1f} {'YES' if row[1][1] else 'no':>5s}  "
          f"{p_detect(row[0][0]):>9.3f}  {p_detect(row[1][0]):>9.3f}")

sat_note_bands = [b for b in ("MWIR", "LWIR") if any(sweep_sat[b])]
if sat_note_bands:
    for b in sat_note_bands:
        t_sat = T_sweep[sweep_sat[b].index(True)]
        print(f"\n  {b} SATURATES from T_fire ≈ {t_sat:.0f} K (signal ≥ 98% of the")
        print(f"  {det[b]['Full well capacity']:,.0f} e⁻ well even in fire mode) — beyond that the")
        print(f"  radiometry clips and fire temperature cannot be retrieved in-band.")

# ---------------------------------------------------------------------------
# Step 7: Plots
# ---------------------------------------------------------------------------

OUTPUTS.mkdir(parents=True, exist_ok=True)

fig1, ax1 = plt.subplots(figsize=(10, 6))
ax1.semilogy(wl_um, np.maximum(dL_600, 1e-4), "k-", linewidth=2,
             label="ΔL(λ), T_fire = 600 K")
for b, color in [("MWIR", "tab:orange"), ("LWIR", "tab:red")]:
    lo, hi = bands[b]
    ax1.axvspan(lo, hi, alpha=0.15, color=color, label=f"{b} {lo:.1f}–{hi:.1f} µm")
ax1.axvline(2898.0 / 600.0, color="gray", linestyle=":",
            label=f"600 K Planck peak ({2898.0 / 600.0:.1f} µm)")
ax1.set_xlabel("Wavelength [µm]", fontsize=12)
ax1.set_ylabel("Spectral contrast ΔL [W/m²/sr/µm]", fontsize=12)
ax1.set_title("Fire-to-Forest Spectral Contrast (hand Planck, ASTER ε_bg(λ))",
              fontsize=13)
ax1.legend(fontsize=10)
ax1.grid(True, which="both", alpha=0.3)
fig1.tight_layout()
fig1.savefig(OUTPUTS / "fig1_spectral_contrast.png", dpi=150)
print(f"\n  Saved {OUTPUTS / 'fig1_spectral_contrast.png'}")

fig2, (ax2a, ax2b) = plt.subplots(1, 2, figsize=(13, 6))
for b, color in [("MWIR", "tab:orange"), ("LWIR", "tab:red")]:
    scnr_arr = np.array(sweep_scnr[b])
    sat_arr = np.array(sweep_sat[b])
    ax2a.semilogy(T_sweep, scnr_arr, "o-", color=color, linewidth=2, label=b)
    if sat_arr.any():
        ax2a.semilogy(T_sweep[sat_arr], scnr_arr[sat_arr], "x", color="black",
                      markersize=11, markeredgewidth=2,
                      label=f"{b} saturated" if b == "MWIR" else None)
    ax2b.plot(T_sweep, [p_detect(v) for v in scnr_arr], "o-", color=color,
              linewidth=2, label=b)
ax2a.axhline(SCNR_THRESHOLD, color="gray", linestyle="--",
             label=f"P_fa = 1e-6 threshold ({SCNR_THRESHOLD:.2f})")
ax2a.set_xlabel("Hotspot Temperature [K]", fontsize=11)
ax2a.set_ylabel("SCNR [dimensionless]", fontsize=11)
ax2a.set_title("Detection SCNR vs Fire Temperature", fontsize=12)
ax2a.legend(fontsize=9)
ax2a.grid(True, which="both", alpha=0.3)
ax2b.set_xlabel("Hotspot Temperature [K]", fontsize=11)
ax2b.set_ylabel("Detection Probability P_d [--]", fontsize=11)
ax2b.set_ylim(0, 1.05)
ax2b.set_title("P_d at P_fa = 1e-6", fontsize=12)
ax2b.legend(fontsize=9)
ax2b.grid(True, alpha=0.3)
fig2.tight_layout()
fig2.savefig(OUTPUTS / "fig2_detection_vs_temperature.png", dpi=150)
print(f"  Saved {OUTPUTS / 'fig2_detection_vs_temperature.png'}")

fig3, ax3 = plt.subplots(figsize=(10, 6))
terms_plot = [t for t in all_terms
              if noise["MWIR"].get(t, 0) > 0.01 or noise["LWIR"].get(t, 0) > 0.01]
x = np.arange(len(terms_plot))
w = 0.38
ax3.bar(x - w / 2, [noise["MWIR"].get(t, 0) for t in terms_plot], w,
        color="tab:orange", edgecolor="black", linewidth=0.5, label="MWIR")
ax3.bar(x + w / 2, [noise["LWIR"].get(t, 0) for t in terms_plot], w,
        color="tab:red", edgecolor="black", linewidth=0.5, label="LWIR")
ax3.set_xticks(x)
ax3.set_xticklabels(terms_plot, rotation=30, ha="right", fontsize=9)
ax3.set_yscale("log")
ax3.set_ylabel("Noise [e⁻ RMS]", fontsize=11)
ax3.set_title(f"Noise Budgets at T_fire = {T_nominal:.0f} K (log scale)", fontsize=13)
ax3.legend(fontsize=10)
ax3.grid(True, axis="y", which="both", alpha=0.3)
fig3.tight_layout()
fig3.savefig(OUTPUTS / "fig3_noise_budgets.png", dpi=150)
print(f"  Saved {OUTPUTS / 'fig3_noise_budgets.png'}")

# ---------------------------------------------------------------------------
# Step 8: Output workbook
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()
ws1 = wb_out.active
ws1.title = "Band Trade"
hdr_font = Font(bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
ws1["A1"] = f"Scenario 1.3 — Dual-band trade at T_fire = {T_nominal:.0f} K"
ws1["A1"].font = Font(bold=True, size=14)
for col, htext in enumerate(["Quantity", "MWIR", "LWIR"], 1):
    cell = ws1.cell(row=3, column=col, value=htext)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
for i, (label, m, lw) in enumerate(rows, 4):
    ws1.cell(row=i, column=1, value=label).border = border
    ws1.cell(row=i, column=2, value=m).border = border
    ws1.cell(row=i, column=3, value=lw).border = border
ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 16

ws2 = wb_out.create_sheet("T Sweep")
for col, htext in enumerate(["T_fire [K]", "MWIR SCNR", "MWIR saturated",
                             "LWIR SCNR", "LWIR saturated",
                             "P_d MWIR", "P_d LWIR"], 1):
    cell = ws2.cell(row=1, column=col, value=htext)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = border
for i, T in enumerate(T_sweep, 2):
    vals = [float(T),
            round(sweep_scnr["MWIR"][i - 2], 2), str(sweep_sat["MWIR"][i - 2]),
            round(sweep_scnr["LWIR"][i - 2], 2), str(sweep_sat["LWIR"][i - 2]),
            round(p_detect(sweep_scnr["MWIR"][i - 2]), 4),
            round(p_detect(sweep_scnr["LWIR"][i - 2]), 4)]
    for col, v in enumerate(vals, 1):
        ws2.cell(row=i, column=col, value=v).border = border
for col_letter in "ABCDEFG":
    ws2.column_dimensions[col_letter].width = 15

wb_out.save(OUTPUT_FILE)
print(f"  Output workbook: {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# Step 9: Summary and recommendation
# ---------------------------------------------------------------------------

print(f"\n{'=' * 95}")
print(f"  SUMMARY AND RECOMMENDATION")
print(f"{'=' * 95}")
print(f"\n  Mission: 5 m² hotspot vs 300 K conifer forest, 10 km airborne, "
      f"{gsd_m:.0f} m GSD (fill {fill * 100:.0f}%)")
print(f"  Forest ε from ASTER: {eps_forest['MWIR']:.3f} (MWIR) / "
      f"{eps_forest['LWIR']:.3f} (LWIR)")
print(f"\n  At the nominal 600 K fire:")
print(f"    MWIR: SCNR = {scnr['MWIR']:.1f}, well fill {well['MWIR']:.0f}%, "
      f"P_d = {p_detect(scnr['MWIR']):.3f}")
print(f"    LWIR: SCNR = {scnr['LWIR']:.1f}, well fill {well['LWIR']:.0f}%, "
      f"P_d = {p_detect(scnr['LWIR']):.3f}")
print(f"\n  Physics of the trade:")
print(f"    1. The 600 K Planck peak ({2898.0 / 600.0:.1f} µm) sits inside the MWIR band,")
print(f"       and Planck contrast at fixed ΔT grows steeply at short wavelengths")
print(f"       (Wien side) — MWIR fire contrast is enormous.")
print(f"    2. LWIR sees the 300 K background ~10× brighter (background photon")
print(f"       noise + clutter up) while its fire contrast grows only ~8× from")
print(f"       300→600 K — LWIR is the mapping band, not the detection band.")
print(f"    3. Dynamic range flips the argument at high T: MWIR saturates first")
print(f"       (see sweep) — a real fire product pairs MWIR detection with LWIR")
print(f"       (or sub-frame integrations) for temperature retrieval.")
print(f"    4. Same canopy, different ε per band (ASTER: "
      f"{eps_forest['MWIR']:.3f} vs {eps_forest['LWIR']:.3f}) — a scalar")
print(f"       emissivity shared across bands would bias the background by ~3%.")
print(f"\n  RECOMMENDATION: MWIR for detection (higher SCNR at all fire")
print(f"  temperatures below saturation); add LWIR or sub-frame MWIR modes if")
print(f"  fire-temperature retrieval above ~{T_sweep[sweep_sat['MWIR'].index(True)] if any(sweep_sat['MWIR']) else 1200:.0f} K matters.")
