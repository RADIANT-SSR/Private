#!/usr/bin/env python3
"""Create Dr. Chen's inputs for Scenario 6.4: synthetic scene generation.

Emits ``chen_scene.xlsx`` — the multi-target scene (5 targets at different
ranges, temperatures, emissivities, and sizes), the uniform background, and
the LWIR sensor configuration for the per-pixel strip + ROC study.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Targets"
ws["A1"] = "Scenario 6.4: synthetic multi-target scene"
ws["A1"].font = Font(bold=True, size=14)
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="264478")
cols = ["Target", "Range (km)", "Temp (K)", "Emissivity", "Size (m)"]
for i, text in enumerate(cols):
    c = ws.cell(row=3, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
for col, w in zip("ABCDE", [12, 12, 10, 12, 10], strict=True):
    ws.column_dimensions[col].width = w
TARGETS = [
    ("T1", 10.0, 330.0, 0.95, 3.0),
    ("T2", 20.0, 322.0, 0.93, 3.0),
    ("T3", 50.0, 316.0, 0.92, 3.0),
    ("T4", 100.0, 310.0, 0.95, 3.0),
    ("T5", 200.0, 305.0, 0.93, 3.0),
]
r = 4
for name, rng, t, e, size in TARGETS:
    for col, val in zip("ABCDE", [name, rng, t, e, size], strict=True):
        ws[f"{col}{r}"] = val
    r += 1

ws2 = wb.create_sheet("Sensor")
for i, text in enumerate(["Parameter", "Value", "Unit"]):
    c = ws2.cell(row=1, column=i + 1, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
ws2.column_dimensions["A"].width = 26
for i, (name, val, unit) in enumerate(
    [
        ("Background temperature", 290.0, "K"),
        ("Background emissivity", 0.95, "—"),
        ("Waveband", "8-12", "µm"),
        ("Aperture diameter", 5.0, "cm"),
        ("Focal length", 1.0, "m"),
        ("Pixel pitch", 25.0, "µm"),
        ("QE", 70.0, "%"),
        ("Dark current", 1.0e6, "e⁻/s"),
        ("Read noise", 100.0, "e⁻"),
        ("Integration time", 0.5, "ms"),
        ("Detection threshold P_fa", 1.0e-4, "—"),
    ],
    start=2,
):
    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=unit)

out = HERE / "chen_scene.xlsx"
wb.save(out)
print(f"Wrote {out}")
