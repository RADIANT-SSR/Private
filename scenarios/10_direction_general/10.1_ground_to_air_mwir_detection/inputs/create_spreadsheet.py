"""Generate the vendor-format input workbook for scenario 10.1.

The workbook is deliberately **not** in RADIANT-internal units.  It is what a
range test engineer actually receives:

* a camera vendor datasheet (mm, %, ms, µm, °C, e-),
* a range/target sheet from the IR-signature group (km, mm, °C),
* a pointing plan (degrees),
* the owner-delivered MODTRAN batch-1 run list (km, degrees).

``scripts/run_ground_to_air_mwir_detection.py`` performs every unit conversion
once, with an explicit comment per conversion (scenario_testing.md Rule 1).

Run from anywhere::

    python scenarios/10_direction_general/10.1_ground_to_air_mwir_detection/\
        inputs/create_spreadsheet.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

OUT_PATH = Path(__file__).resolve().parent / "ground_mwir_tracker_data.xlsx"

TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="00375E97", end_color="00375E97", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def add_section(ws: Worksheet, row: int, title: str, n_cols: int = 4) -> int:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN
    return row + 1


def add_param(ws: Worksheet, row: int, name: str, value: object, unit: str, note: str = "") -> int:
    ws.cell(row=row, column=1, value=name).border = THIN
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.border = THIN
    value_cell.alignment = Alignment(horizontal="center")
    unit_cell = ws.cell(row=row, column=3, value=unit)
    unit_cell.border = THIN
    unit_cell.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=note).border = THIN
    return row + 1


def add_table_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN
    return row + 1


def add_table_row(ws: Worksheet, row: int, values: list[object]) -> int:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN
    return row + 1


def build() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------
    # Sheet 1 — camera vendor datasheet (mm / % / µm / ms / °C / e-)
    # -----------------------------------------------------------------
    ws = wb.active
    assert ws is not None
    ws.title = "Camera Datasheet"
    ws["A1"] = "GT-100 ground MWIR search-and-track camera — vendor datasheet"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Cooled InSb FPA, all-reflective f/2 telescope, tripod-mounted, uncooled housing"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 62

    row = 4
    row = add_section(ws, row, "Telescope")
    row = add_param(ws, row, "Entrance pupil diameter", 100.0, "mm", "Clear aperture, unobscured")
    row = add_param(ws, row, "Effective focal length", 200.0, "mm", "f/2.0")
    row = add_param(ws, row, "Optical transmission", 75.0, "%", "Band-average, 3-5 um")
    row = add_param(ws, row, "Housing temperature", 20.0, "degC", "Ambient; NOT cooled")
    row = add_param(
        ws,
        row,
        "Train emissivity",
        25.0,
        "%",
        "All-reflective train: eps = 1 - tau (Kirchhoff); vendor states it explicitly",
    )
    row = add_param(
        ws,
        row,
        "Cold shield efficiency",
        90.0,
        "%",
        "Fraction of the FPA hemisphere BLOCKED from warm optics (vendor convention)",
    )

    row += 1
    row = add_section(ws, row, "Focal plane")
    row = add_param(ws, row, "Pixel pitch", 15.0, "um", "Square pixels")
    row = add_param(ws, row, "Quantum efficiency", 75.0, "%", "Band-average")
    row = add_param(ws, row, "Dark current", 50000.0, "e-/s", "At the stated FPA temperature")
    row = add_param(ws, row, "FPA temperature", 80.0, "K", "Stirling cooler")
    row = add_param(ws, row, "Spectral band, low edge", 3000.0, "nm", "Cold filter")
    row = add_param(ws, row, "Spectral band, high edge", 5000.0, "nm", "Cold filter")

    row += 1
    row = add_section(ws, row, "Readout")
    row = add_param(ws, row, "Integration time", 0.5, "ms", "Track mode, 1 kHz frame rate")
    row = add_param(ws, row, "Read noise", 250.0, "e- rms", "CDS")
    row = add_param(ws, row, "Full well capacity", 11.0, "Me-", "High-flux mode")
    row = add_param(ws, row, "System gain", 200.0, "e-/DN", "")
    row = add_param(ws, row, "ADC resolution", 16, "bits", "")

    # -----------------------------------------------------------------
    # Sheet 2 — range site + target signature (km / mm / degC / %)
    # -----------------------------------------------------------------
    ws2 = wb.create_sheet("Site and Target")
    ws2["A1"] = "Desert range trial — site conditions and target signature"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Night trial: no solar illumination on any part of the scene"
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 62

    row = 4
    row = add_section(ws2, row, "Observation site")
    row = add_param(ws2, row, "Site altitude", 0.0, "km", "Sea-level range; sensor on the ground")
    row = add_param(ws2, row, "Standard atmosphere", "midlat_summer", "-", "Radiosonde match")
    row = add_param(ws2, row, "Visibility", 23.0, "km", "Meteorological range, clear")
    row = add_param(ws2, row, "Aerosol model", "rural", "-", "")
    row = add_param(ws2, row, "Illumination", "night", "-", "Trial runs 0100-0400 local")

    row += 1
    row = add_section(ws2, row, "Target — small turbojet UAS, aft aspect")
    row = add_param(ws2, row, "Target altitude", 10.0, "km", "Level cruise leg")
    row = add_param(ws2, row, "Nozzle exit diameter", 60.0, "mm", "Projected disc, aft aspect")
    row = add_param(ws2, row, "Nozzle temperature", 276.85, "degC", "= 550 K, IR-signature group")
    row = add_param(ws2, row, "Nozzle emissivity", 0.90, "-", "Oxidised metal, material property")

    row += 1
    row = add_section(ws2, row, "Detection criterion")
    row = add_param(ws2, row, "SNR threshold", 5.0, "-", "Range acceptance criterion")

    # -----------------------------------------------------------------
    # Sheet 3 — pointing plan (degrees)
    # -----------------------------------------------------------------
    ws3 = wb.create_sheet("Pointing Plan")
    ws3["A1"] = "Elevation pointing plan — zenith angle at the sensor (the path's lower endpoint)"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = "0 deg = straight up; 60 deg = 30 deg above the horizon"
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 60

    row = 4
    row = add_table_header(ws3, row, ["Zenith angle [deg]", "Elevation [deg]", "Note"])
    for zenith_deg, note in (
        (0.0, "Overhead pass"),
        (10.0, ""),
        (20.0, ""),
        (30.0, "Nominal track geometry"),
        (40.0, ""),
        (45.0, "Matches MODTRAN run K6"),
        (50.0, ""),
        (60.0, "Lowest elevation in the trial plan"),
    ):
        row = add_table_row(ws3, row, [zenith_deg, 90.0 - zenith_deg, note])

    # -----------------------------------------------------------------
    # Sheet 4 — MODTRAN anchor runs (owner-delivered batch 1)
    # -----------------------------------------------------------------
    ws4 = wb.create_sheet("MODTRAN Anchors")
    ws4["A1"] = "Owner-delivered MODTRAN batch-1 up-looking runs used as truth anchors"
    ws4["A1"].font = TITLE_FONT
    ws4["A2"] = "Files: modtran/real_runs/ (gitignored); see docs/plans/modtran_run_matrix.csv"
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 46

    row = 4
    row = add_table_header(
        ws4, row, ["Run", "H1 [km]", "H2 [km]", "Path angle [deg]", "Note"]
    )
    for run, h1_km, h2_km, angle_deg, note in (
        ("K1", 0.0, 1.0, 0.0, "Vertical, 1 km column"),
        ("K2", 0.0, 3.0, 0.0, "Vertical, 3 km column"),
        ("K3", 0.0, 5.0, 0.0, "Vertical, 5 km column"),
        ("K4", 0.0, 10.0, 0.0, "Vertical, 10 km column — the nominal target altitude"),
        ("K5", 0.0, 20.0, 0.0, "Vertical, 20 km column"),
        ("K6", 0.0, 10.0, 45.0, "45 deg slant twin of K4"),
    ):
        row = add_table_row(ws4, row, [run, h1_km, h2_km, angle_deg, note])

    return wb


def main() -> None:
    wb = build()
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
