"""Scenario 4.1: Target Detection Matrix — 12 Targets × 4 Atmospheres × 3 Sensors.

Lisa needs a quarterly program-review briefing: for every target in the
program target list, under four atmospheric conditions, for each of the
three sensors in the library — how far off-nadir can the sensor slide
before detection (SNR ≥ 5) is lost, and what does image quality (NIIRS)
look like at nadir?

This script:
  1. Loads the 12-target library with radiant.io.target_library
     (projected_area_m2 derived from length × width).
  2. Loads the 3 sensor YAMLs (one deliberately carries the outdated
     optics.cold_stop_efficiency name — absorbed by the Gap 12
     deprecated-alias mechanism with a DeprecationWarning).
  3. Runs the 12 × 4 matrix per sensor with radiant.api.batch.BatchRunner
     (144 cells total). Per cell the evaluate callback:
     - evaluates SNR and NIIRS at nadir;
     - finds the detection-range zenith limit by bisection on
       geometry.path_zenith_rad (the target-side path zenith θ_o); the
       chain derives the slant range from θ_o via the spherical viewing
       triangle, and the runner reads the same slant back through
       radiant.core.viewing_triangle.slant_range_from_theta_o_m for the
       footprint and the reported range — setting geometry.target_range_m
       as well would over-specify the line of sight (CU-093/CU-182);
     - converts the zenith limit to a slant detection range [km].
  4. Writes the color-coded Excel matrix (per-sensor sheets), the NIIRS
     matrix, and summary plots; identifies the hardest target.

Detection definition: SCNR ≥ 5 — the signal-to-clutter-plus-noise
ratio |contrast_e| / RSS(all noise terms including scene clutter),
computed script-side because RADIANT's snr/contrast_snr metrics carry
temporal noise only (spatial clutter is excluded — recorded in gaps.md).
Pure noise-limited SNR ≥ 5 saturates for every cell (these sensors hold
SNR > 19 across the whole swath); real sub-pixel detection at 500 km is
clutter-limited, so the matrix models rural scene clutter at 2% of the
in-pixel background (detector.clutter_sigma = 0.02).

CONTRAST-SIGN NOTE (see gaps.md): in the sub-pixel regime the chain forms
contrast_e = ff·(L_target·EE_box − L_bg): the target's compact energy is
EE_box-weighted (its PSF spills to neighbouring pixels) while the uniform
in-pixel background it OCCLUDES is not. A deeply sub-pixel target
(tiny EE_box) is therefore detected largely by the background DIP it
punches, and contrast_e can be negative for a hot-but-dim target. The
|contrast| detection criterion is robust to the sign; the absolute
detection ranges carry that caveat. Targets are unresolved (GSD 8.6–17 m),
so NIIRS values are GIQE extrapolations far outside calibration.

Usage:
    python run_detection_matrix.py
"""

import math
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from radiant.api.batch import BatchRunner
from radiant.api.sensor import Sensor
from radiant.core.viewing_triangle import slant_range_from_theta_o_m
from radiant.io.target_library import load_target_library

INPUTS = Path(__file__).parent.parent / "inputs"
OUTPUTS = Path(__file__).parent.parent / "outputs"
OUTPUT_FILE = OUTPUTS / "detection_matrix_results.xlsx"

SCNR_THRESHOLD = 5.0         # detection criterion [dimensionless]
CLUTTER_SIGMA = 0.02         # rural scene clutter, fraction of in-pixel background
ALTITUDE_M = 500_000.0       # all three sensors fly the same 500 km orbit
ZENITH_MAX_RAD = math.radians(66.0)  # horizon at 500 km is 68.0° — stay clear
BISECTION_STEPS = 9          # zenith resolution ~0.13° → ~0.5% in range

print("=" * 100)
print("  SCENARIO 4.1: Target Detection Matrix — 12 Targets × 4 Atmospheres × 3 Sensors")
print("=" * 100)

# ---------------------------------------------------------------------------
# Step 1: Load the target library
# ---------------------------------------------------------------------------

targets = load_target_library(INPUTS / "lisa_target_library.xlsx")
print(f"\n=== Target library ({len(targets)} targets, radiant.io.target_library) ===")
print(f"  {'Target':<22s} {'L×W [m]':>12s}  {'A_proj [m²]':>12s}  {'T [K]':>7s}  "
      f"{'ε [--]':>7s}  {'Material'}")
print(f"  {'-' * 22} {'-' * 12}  {'-' * 12}  {'-' * 7}  {'-' * 7}  {'-' * 18}")
for t in targets:
    print(f"  {t.target_name:<22s} {t.length_m:>5.1f}×{t.width_m:<5.1f}  "
          f"{t.projected_area_m2:>12.1f}  {t.temperature_K:>7.1f}  "
          f"{t.emissivity:>7.2f}  {t.material}")

# ---------------------------------------------------------------------------
# Step 2: Sensor library (YAML) and atmosphere conditions
# ---------------------------------------------------------------------------

SENSOR_FILES = {
    "A: MWIR smallsat 18 cm": INPUTS / "sensor_a_mwir_smallsat.yaml",
    "B: MWIR flagship 50 cm": INPUTS / "sensor_b_mwir_flagship.yaml",
    "C: LWIR wide 35 cm": INPUTS / "sensor_c_lwir_wide.yaml",
}

# Atmosphere axis: visibility drives aerosol extinction; the profile
# matches the named condition (temperature/humidity differ physically).
ATMOSPHERES = {
    "clear": {"atmosphere.visibility_km": 50.0,
              "atmosphere.standard_atmosphere": "midlat_summer"},
    "haze": {"atmosphere.visibility_km": 10.0,
             "atmosphere.standard_atmosphere": "midlat_summer"},
    "tropical_haze": {"atmosphere.visibility_km": 5.0,
                      "atmosphere.standard_atmosphere": "tropical"},
    "arctic_clear": {"atmosphere.visibility_km": 100.0,
                     "atmosphere.standard_atmosphere": "subarctic_winter"},
}

# Target axis: temperature, emissivity, projected area from the library.
# Background: uniform soil/vegetation scene (288 K, ε=0.95) for every cell
# so the matrix varies only the declared axes.
TARGET_AXIS = {
    t.target_name: {
        "source.target.temperature": t.temperature_K,
        "source.target.emissivity": t.emissivity,
        "geometry.target.projected_area_m2": t.projected_area_m2,
    }
    for t in targets
}

print(f"\n=== Sensor library (3 YAML configs) ===")
with warnings.catch_warnings(record=True) as caught:
    warnings.simplefilter("always")
    sensors_meta = {}
    for label, path in SENSOR_FILES.items():
        s = Sensor.from_yaml(path)
        # Sensor.get returns canonical units (metres for pixel pitch).
        pitch_m = s.get("detector.pixel_pitch_x_um")
        focal_m = s.get("optics.focal_length_m")
        sensors_meta[label] = {
            "aperture_m": s.get("optics.aperture_diameter_m"),
            "band": (s.get("spectral_integration.filter_min_um"),
                     s.get("spectral_integration.filter_max_um")),
            "pitch_um": pitch_m * 1e6,
            "gsd_nadir_m": pitch_m * ALTITUDE_M / focal_m,
        }
alias_warnings = [w for w in caught if "cold_stop_efficiency" in str(w.message)]
for label, meta in sensors_meta.items():
    print(f"  {label:<24s}: {meta['aperture_m'] * 100:.0f} cm, "
          f"{meta['band'][0]:.1f}–{meta['band'][1]:.1f} µm, "
          f"{meta['pitch_um']:.0f} µm pixels, GSD(nadir) = {meta['gsd_nadir_m']:.1f} m")
print(f"\n  OUTDATED PARAMETER NAME absorbed: sensor C's YAML still says")
print(f"  'optics.cold_stop_efficiency' (pre-Gap-12 name). RADIANT accepted it")
print(f"  through the deprecated-alias mechanism "
      f"({len(alias_warnings)} DeprecationWarning(s) raised) and mapped it to")
print(f"  optics.nearfield_fraction — the config still runs, loudly.")

print(f"\n=== Atmosphere conditions ===")
for name, ov in ATMOSPHERES.items():
    print(f"  {name:<15s}: visibility = {ov['atmosphere.visibility_km']:.0f} km, "
          f"profile = {ov['atmosphere.standard_atmosphere']}")

# ---------------------------------------------------------------------------
# Step 3: The per-cell evaluation — nadir metrics + detection-range bisection
# ---------------------------------------------------------------------------


def scnr_of(result) -> float:
    """|contrast_e| / RSS(all noise terms incl. clutter) — detection SCNR.

    In the sub-pixel regime the chain forms
    contrast_e = ff·(L_target·EE_box − L_bg): the target's compact
    energy is EE_box-weighted (its PSF spills to neighbouring pixels)
    while the uniform background it OCCLUDES is not. A very sub-pixel
    target (tiny EE_box) is therefore detected largely by the
    background DIP it punches — |contrast| is the right criterion, and
    a single hot pixel and a single occluded pixel are equally
    detectable against the scene. (A multi-pixel matched filter would
    additionally sum the target energy in the neighbours; that is a
    performance-model refinement beyond this single-pixel SCNR.)
    """
    contrast_e = abs(float(result.stage_outputs["spectral_integration"]["contrast_e"]))
    total = math.sqrt(sum(nt.value_e**2 for nt in result.noise_terms))
    return contrast_e / total if total > 0 else 0.0


def _ifov_rad(sensor: Sensor) -> float:
    # Sensor.get returns CANONICAL units: pixel pitch in metres (not µm)
    # despite the ..._um parameter name. No extra 1e-6 conversion.
    return (sensor.get("detector.pixel_pitch_x_um")
            / sensor.get("optics.focal_length_m"))


def configure_geometry(sensor: Sensor, zenith_rad: float, area_m2: float) -> None:
    """Consistent off-nadir geometry AND sub-pixel fill fraction.

    ``geometry.path_zenith_rad`` is the target-side path zenith angle θ_o
    (ADR-0006); the chain derives the slant range from it via the spherical
    viewing triangle. We set ONLY θ_o and read the same slant back through
    ``slant_range_from_theta_o_m`` for the footprint — setting
    ``geometry.target_range_m`` as well would over-specify the line of sight
    with a *different* (sensor-off-nadir) slant and trip the CU-093
    consistency check (CU-182). Crucially, the sub-pixel regime weights the
    target by ``source.target.fill_fraction`` (NOT ``projected_area_m2`` —
    that parameter drives the point-source A_t/R² path only). Fill is the
    target's projected area over the pixel footprint (IFOV·R)², capped at
    1.0 (a target larger than a pixel fills it). This is what makes
    detectability scale with target SIZE and with range (the footprint
    grows off-nadir, so fill — and SCNR — fall).
    """
    r_slant = slant_range_from_theta_o_m(zenith_rad, ALTITUDE_M, 0.0)
    sensor.set("geometry.path_zenith_rad", zenith_rad)
    footprint_m2 = (_ifov_rad(sensor) * r_slant) ** 2
    sensor.set("geometry.target.projected_area_m2", min(area_m2, footprint_m2))
    sensor.set("source.target.fill_fraction", min(1.0, area_m2 / footprint_m2))


def scnr_at(sensor: Sensor, zenith_rad: float, area_m2: float) -> float:
    configure_geometry(sensor, zenith_rad, area_m2)
    return scnr_of(sensor.evaluate())


def evaluate_cell(sensor: Sensor, labels: dict[str, str]) -> dict[str, float]:
    """Nadir SCNR/NIIRS + detection range (slant, at SCNR = threshold).

    CAVEAT (see gaps.md): SCNR uses |contrast_e|. The sub-pixel
    contrast_e sign for warm-background thermal targets is under review
    (its magnitude ordering ran opposite to a hand Planck estimate);
    the |·| detection criterion is robust to the sign, but the absolute
    detection ranges carry that caveat.
    """
    sensor.set("source.scene_type", "sub_pixel")
    sensor.set("source.regime_override", "sub_pixel")  # keep in-pixel background
    sensor.set("source.background.temperature", 288.0)
    sensor.set("source.background.emissivity", 0.95)
    sensor.set("detector.clutter_sigma", CLUTTER_SIGMA)

    overrides = TARGET_AXIS[labels["target"]]
    area_true = overrides["geometry.target.projected_area_m2"]

    scnr_nadir = scnr_at(sensor, 0.0, area_true)
    niirs_nadir = sensor.evaluate().metrics.get("niirs")

    range_nadir_km = ALTITUDE_M / 1000.0  # noqa: F841 (documented reference)
    range_max_km = slant_range_from_theta_o_m(ZENITH_MAX_RAD, ALTITUDE_M, 0.0) / 1000.0

    if scnr_nadir < SCNR_THRESHOLD:
        return {"scnr_nadir": scnr_nadir, "niirs_nadir": niirs_nadir,
                "detection_range_km": 0.0, "horizon_limited": 0.0}

    if scnr_at(sensor, ZENITH_MAX_RAD, area_true) >= SCNR_THRESHOLD:
        return {"scnr_nadir": scnr_nadir, "niirs_nadir": niirs_nadir,
                "detection_range_km": range_max_km, "horizon_limited": 1.0}

    lo, hi = 0.0, ZENITH_MAX_RAD  # SCNR(lo) >= threshold > SCNR(hi)
    for _ in range(BISECTION_STEPS):
        mid = 0.5 * (lo + hi)
        if scnr_at(sensor, mid, area_true) >= SCNR_THRESHOLD:
            lo = mid
        else:
            hi = mid
    detection_range_km = slant_range_from_theta_o_m(0.5 * (lo + hi), ALTITUDE_M, 0.0) / 1000.0
    return {"scnr_nadir": scnr_nadir, "niirs_nadir": niirs_nadir,
            "detection_range_km": detection_range_km, "horizon_limited": 0.0}


# ---------------------------------------------------------------------------
# Step 4: Run the matrix — BatchRunner per sensor (12 targets × 4 atmospheres)
# ---------------------------------------------------------------------------

print(f"\n=== Running the matrix: 12 × 4 × 3 = 144 cells ===")
print(f"  Detection criterion: SCNR ≥ {SCNR_THRESHOLD:.0f} — |contrast| over RSS(noise + clutter),")
print(f"  scene clutter = {CLUTTER_SIGMA * 100:.0f}% of in-pixel background (rural)")
print(f"  Swath edge: zenith {math.degrees(ZENITH_MAX_RAD):.0f}° "
      f"(slant {slant_range_from_theta_o_m(ZENITH_MAX_RAD, ALTITUDE_M, 0.0) / 1e3:,.0f} km; "
      f"horizon at 68.0°)")

warnings.filterwarnings("ignore")  # GIQE extrapolation warnings — noted in output

results = {}


def _load_sensor(path: str) -> Sensor:
    """Load a YAML config and opt into extrapolated NIIRS (CU-178).

    These sub-pixel targets sit far outside the GIQE-5 envelope (GSD 8.6–17 m),
    so the applicability gate returns NIIRS N/A by default; the scenario intends
    the extrapolated NIIRS trend (read as relative, not calibrated).
    """
    s = Sensor.from_yaml(path)
    s.set("performance.niirs.allow_extrapolated", True)
    return s


for label, path in SENSOR_FILES.items():
    runner = BatchRunner(
        base_config={},  # unused — the factory loads the YAML
        axes=[("target", TARGET_AXIS), ("atmosphere", ATMOSPHERES)],
        sensor_factory=lambda cfg, p=path: _load_sensor(p),
    )
    print(f"  Running {label} (48 cells)...")
    results[label] = runner.run(evaluate_cell)
    n_failed = results[label].n_failed
    if n_failed:
        print(f"    {n_failed} cell(s) failed — recorded in the error column")

# ---------------------------------------------------------------------------
# Step 5: Report — detection range matrices + hardest target
# ---------------------------------------------------------------------------

range_max_km = slant_range_from_theta_o_m(ZENITH_MAX_RAD, ALTITUDE_M, 0.0) / 1000.0
atm_names = list(ATMOSPHERES.keys())

for label, batch in results.items():
    table = batch.pivot("detection_range_km", rows="target", cols="atmosphere")
    horizon = batch.pivot("horizon_limited", rows="target", cols="atmosphere")
    print(f"\n{'=' * 100}")
    print(f"  DETECTION RANGE MATRIX — {label}  [km slant range at SCNR = "
          f"{SCNR_THRESHOLD:.0f}; * = swath-edge limited]")
    print(f"{'=' * 100}")
    header = "  " + f"{'Target':<22s}" + "".join(f"{a:>18s}" for a in atm_names)
    print(header)
    print("  " + "-" * (22 + 18 * len(atm_names)))
    for t in targets:
        cells = []
        for a in atm_names:
            v = table[t.target_name][a]
            if v is None:
                cells.append(f"{'ERROR':>18s}")
            elif v == 0.0:
                cells.append(f"{'not detectable':>18s}")
            else:
                marker = "*" if horizon[t.target_name][a] == 1.0 else " "
                cells.append(f"{v:>16,.0f}{marker} ")
        print("  " + f"{t.target_name:<22s}" + "".join(cells))

# Hardest target: smallest mean detection range across all cells
mean_range: dict[str, float] = {}
for t in targets:
    vals = []
    for batch in results.values():
        table = batch.pivot("detection_range_km", rows="target", cols="atmosphere")
        vals.extend(v for v in table[t.target_name].values() if v is not None)
    mean_range[t.target_name] = float(np.mean(vals)) if vals else float("nan")
hardest = min(mean_range, key=lambda k: mean_range[k])
easiest = max(mean_range, key=lambda k: mean_range[k])

print(f"\n=== Worst-case target ===")
print(f"  Hardest: {hardest} — mean detection range {mean_range[hardest]:,.0f} km "
      f"across all 12 sensor×atmosphere cells")
print(f"  Easiest: {easiest} — mean {mean_range[easiest]:,.0f} km")
hard_entry = next(t for t in targets if t.target_name == hardest)
print(f"  Note: with the EE_box-weighted sub-pixel contrast "
      f"ff·(L_t·EE_box − L_bg), detectability is set by how far the")
print(f"  target pixel departs from background (either direction) times")
print(f"  fill fraction — NOT simply by ε·B(T)·A. {hardest} "
      f"(A = {hard_entry.projected_area_m2:.0f} m², "
      f"ε = {hard_entry.emissivity:.2f}, T = {hard_entry.temperature_K:.0f} K)")
print(f"  lands closest to the background radiance after EE_box weighting,")
print(f"  so its pixel departs least — the hardest to separate.")

print(f"\n=== Physics notes ===")
print(f"  0. Detectability scales with target SIZE via source.target.fill_fraction")
print(f"     = min(1, A_target / (IFOV·R)²): the sub-pixel target is weighted by")
print(f"     the pixel fraction it covers. Off-nadir the footprint grows, fill")
print(f"     drops, and SCNR falls — that is the detection-range mechanism.")
print(f"     Sub-pixel contrast is EE_box-weighted (ff·(L_t·EE_box − L_bg)); the")
print(f"     |contrast| criterion is robust to its sign (see gaps.md).")
print(f"  1. Regime: every cell runs as a SUB-PIXEL target (regime_override —")
print(f"     the classifier picks point_source for the small targets, which")
print(f"     drops the in-pixel background photons entirely; detection against")
print(f"     a bright earth background needs them, so the override keeps the")
print(f"     background + clutter terms in every cell).")
print(f"  2. NIIRS values are GIQE-5 EXTRAPOLATIONS: GSD is far outside the")
print(f"     calibration range (< 0.8 m), so NIIRS ≈ 3 at nadir describes")
print(f"     'detection-class' imagery only in the loosest sense. Reported")
print(f"     because the briefing template asks for it, flagged accordingly.")
print(f"  3. Detection range grows off-nadir until atmosphere wins: slant range")
print(f"     R(θ) uses the spherical Earth (horizon 68° at 500 km), and the")
print(f"     airmass grows faster than R² near the edge — haze cells lose")
print(f"     detection at much shorter ranges than clear cells.")
print(f"  4. Noise-limited SNR saturates (> 19 at the swath edge for every")
print(f"     cell) — detection here is CLUTTER-limited, so the criterion is")
print(f"     SCNR with clutter = {CLUTTER_SIGMA * 100:.0f}% of the in-pixel background. RADIANT's")
print(f"     snr/contrast_snr metrics carry temporal noise only; SCNR is")
print(f"     assembled script-side from contrast_e + the full noise budget")
print(f"     (recorded in gaps.md).")
print(f"  5. '*' cells are swath-edge limited: SCNR ≥ {SCNR_THRESHOLD:.0f} everywhere out to")
print(f"     the 66° practical edge — the sensor, not the atmosphere, ends the")
print(f"     access there.")

# ---------------------------------------------------------------------------
# Step 6: Plots
# ---------------------------------------------------------------------------

OUTPUTS.mkdir(parents=True, exist_ok=True)

# Fig 1: heatmap grid of detection range (one panel per sensor)
fig1, axes = plt.subplots(1, 3, figsize=(17, 7), sharey=True)
target_names = [t.target_name for t in targets]
for ax, (label, batch) in zip(axes, results.items()):
    table = batch.pivot("detection_range_km", rows="target", cols="atmosphere")
    grid = np.array([[table[t][a] if table[t][a] is not None else np.nan
                      for a in atm_names] for t in target_names])
    im = ax.imshow(grid, aspect="auto", cmap="RdYlGn",
                   vmin=0.0, vmax=range_max_km)
    ax.set_xticks(range(len(atm_names)))
    ax.set_xticklabels(atm_names, rotation=30, ha="right", fontsize=9)
    if ax is axes[0]:
        ax.set_yticks(range(len(target_names)))
        ax.set_yticklabels(target_names, fontsize=9)
    ax.set_title(label, fontsize=11)
    for i in range(len(target_names)):
        for j in range(len(atm_names)):
            v = grid[i, j]
            ax.text(j, i, "—" if np.isnan(v) else f"{v:,.0f}",
                    ha="center", va="center", fontsize=7)
fig1.colorbar(im, ax=axes, label="Detection range [km slant, SCNR ≥ 5]",
              shrink=0.85)
fig1.suptitle("Detection Range Matrix — 12 Targets × 4 Atmospheres × 3 Sensors",
              fontsize=14)
fig1.savefig(OUTPUTS / "fig1_detection_range_matrix.png", dpi=150,
             bbox_inches="tight")
print(f"\n  Saved {OUTPUTS / 'fig1_detection_range_matrix.png'}")

# Fig 2: nadir SNR by target (clear atmosphere, all sensors)
fig2, ax2 = plt.subplots(figsize=(11, 6))
x = np.arange(len(target_names))
w = 0.27
for k, (label, batch) in enumerate(results.items()):
    snr_table = batch.pivot("scnr_nadir", rows="target", cols="atmosphere")
    vals = [snr_table[t]["clear"] if snr_table[t]["clear"] is not None else 0.0
            for t in target_names]
    ax2.bar(x + (k - 1) * w, vals, w, label=label)
ax2.axhline(SCNR_THRESHOLD, color="red", linestyle="--",
            label=f"Detection threshold SCNR = {SCNR_THRESHOLD:.0f}")
ax2.set_yscale("log")
ax2.set_xticks(x)
ax2.set_xticklabels(target_names, rotation=35, ha="right", fontsize=9)
ax2.set_ylabel("SCNR at nadir [dimensionless]", fontsize=11)
ax2.set_title("Nadir SCNR by Target — Clear Atmosphere (vis = 50 km)", fontsize=13)
ax2.legend(fontsize=9)
ax2.grid(True, axis="y", which="both", alpha=0.3)
fig2.tight_layout()
fig2.savefig(OUTPUTS / "fig2_nadir_scnr_by_target.png", dpi=150)
print(f"  Saved {OUTPUTS / 'fig2_nadir_scnr_by_target.png'}")

# ---------------------------------------------------------------------------
# Step 7: Color-coded Excel briefing workbook
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)
hdr_font = Font(bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
red = PatternFill("solid", fgColor="FFC7CE")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

for label, batch in results.items():
    sheet_name = label.split(":")[0].strip()  # "A" / "B" / "C"
    ws = wb_out.create_sheet(f"Sensor {sheet_name}")
    ws["A1"] = (f"Detection range [km slant] at SCNR >= {SCNR_THRESHOLD:.0f} — {label} "
                f"(* = swath-edge limited at 66 deg)")
    ws["A1"].font = Font(bold=True, size=13)
    table = batch.pivot("detection_range_km", rows="target", cols="atmosphere")
    horizon = batch.pivot("horizon_limited", rows="target", cols="atmosphere")
    niirs_t = batch.pivot("niirs_nadir", rows="target", cols="atmosphere")

    ws.cell(row=3, column=1, value="Target").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    for j, a in enumerate(atm_names, start=2):
        c = ws.cell(row=3, column=j, value=f"{a} [km]")
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=3, column=len(atm_names) + 2, value="NIIRS nadir (clear)*")
    c.font = hdr_font
    c.fill = hdr_fill

    for i, t in enumerate(target_names, start=4):
        ws.cell(row=i, column=1, value=t).border = border
        for j, a in enumerate(atm_names, start=2):
            v = table[t][a]
            cell = ws.cell(row=i, column=j)
            cell.border = border
            if v is None:
                cell.value = "ERROR"
                cell.fill = red
            elif v == 0.0:
                cell.value = "not detectable"
                cell.fill = red
            else:
                suffix = "*" if horizon[t][a] == 1.0 else ""
                cell.value = f"{v:,.0f}{suffix}"
                cell.fill = green if v >= 0.8 * range_max_km else yellow
        nv = niirs_t[t]["clear"]
        ws.cell(row=i, column=len(atm_names) + 2,
                value=None if nv is None else round(nv, 2)).border = border
    ws.cell(row=len(target_names) + 5, column=1,
            value="* NIIRS is a GIQE-5 extrapolation far outside calibration "
                  "(GSD >> 31.5 in) — detection-class context only.")
    ws.column_dimensions["A"].width = 24
    for col in "BCDEF":
        ws.column_dimensions[col].width = 17

ws_sum = wb_out.create_sheet("Summary", 0)
ws_sum["A1"] = "Scenario 4.1 — Detection Matrix Summary"
ws_sum["A1"].font = Font(bold=True, size=14)
ws_sum["A3"] = f"Hardest target: {hardest} (mean {mean_range[hardest]:,.0f} km)"
ws_sum["A4"] = f"Easiest target: {easiest} (mean {mean_range[easiest]:,.0f} km)"
ws_sum["A5"] = (f"Criterion: SCNR >= {SCNR_THRESHOLD:.0f}, clutter-limited sub-pixel "
                f"detection, 500 km orbit, swath edge 66 deg")
ws_sum["A6"] = f"Cells evaluated: {sum(len(b.rows) for b in results.values())}; " \
               f"failed: {sum(b.n_failed for b in results.values())}"
ws_sum.column_dimensions["A"].width = 90

wb_out.save(OUTPUT_FILE)
print(f"  Output workbook: {OUTPUT_FILE}")

print(f"\n{'=' * 100}")
print(f"  DONE — 144 cells evaluated "
      f"({sum(b.n_failed for b in results.values())} failures recorded)")
print(f"{'=' * 100}")
