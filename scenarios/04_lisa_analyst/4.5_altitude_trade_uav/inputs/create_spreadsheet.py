#!/usr/bin/env python3
"""Create Lisa's inputs for Scenario 4.5: microbolometer UAV altitude trade.

Emits ``lisa_microbolometer_uav.xlsx`` — an uncooled microbolometer
specified the way vendors quote it (by NETD, not by component noise), the
UAV optics, and the ground target Lisa wants to detect.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UAV_IR"
ws["A1"] = "Scenario 4.5: uncooled microbolometer UAV (NETD-specified)"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [30, 14, 12, 46]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="843C0C")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

section = PatternFill("solid", fgColor="FBE4D5")
ROWS = [
    ("MICROBOLOMETER (vendor spec)", None, None, None),
    ("NETD", 50.0, "mK", "Vendor spec at f/1, 300 K, 8–14 µm"),
    ("Waveband", "8–14", "µm", "LWIR, uncooled"),
    ("Pixel pitch", 17.0, "µm", "Microbolometer element"),
    ("Quantum efficiency", 70.0, "%", "Effective (absorption)"),
    ("Frame integration", 16.0, "ms", "Sets noise bandwidth"),
    ("UAV OPTICS", None, None, None),
    ("Aperture diameter", 35.0, "mm", "f/1"),
    ("Focal length", 35.0, "mm", "IFOV = pitch/focal = 486 µrad"),
    ("Optical transmission", 90.0, "%", ""),
    ("GROUND TARGET", None, None, None),
    ("Target size", 1.0, "m", "Critical dimension (small vehicle)"),
    ("Target ΔT over background", 4.0, "K", "Thermal contrast at ground"),
    ("Background temperature", 295.0, "K", ""),
    ("TRADE", None, None, None),
    ("Detection threshold", 4.0, "×NETD", "Apparent ΔT ≥ threshold·NETD"),
    ("Altitude min", 1.0, "km", "Sweep lower bound"),
    ("Altitude max", 12.0, "km", "Sweep upper bound"),
]
r = 4
for name, value, unit, note in ROWS:
    if value is None and unit is None:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True, size=11)
        for cc in range(1, 5):
            ws.cell(row=r, column=cc).fill = section
    else:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "lisa_microbolometer_uav.xlsx"
wb.save(out)
print(f"Wrote {out}")
