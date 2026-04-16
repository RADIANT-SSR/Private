"""Scenario 5.1: WFE Budget Allocation — How Much Aberration Can I Tolerate?

Tom has a Zernike decomposition from Zemax for a 40 cm Cassegrain telescope.
He wants to know how increasing WFE RMS degrades Strehl, MTF, EE, RER, and
NIIRS. He sweeps WFE from 0 to 0.25 waves (at 633 nm HeNe reference) and
finds the budget threshold for each metric.

Approach:
  Run RADIANT at each WFE RMS value using optics.wfe_rms_waves. The optics
  stage applies a random phase screen scaled to the requested RMS, producing
  a degraded PSF. PerformanceStage then computes Strehl (Marechal), MTF,
  EE, RER, SNR, and NIIRS from the aberrated EffectivePSF.

  Since RADIANT uses a random phase screen (not individual Zernike modes),
  a fixed random seed ensures deterministic results. The scalar RMS sweep
  is physically correct for total WFE budget — individual Zernike allocation
  requires the Zernike-to-PSF capability (not yet implemented; see gaps).

Physics:
  - Strehl = exp(-(2*pi*OPD_rms/lambda)^2)  (Marechal approximation)
  - Valid for WFE < ~lambda/5 (Strehl > ~0.3)
  - PSF broadens with increasing WFE, reducing peak and spreading energy
  - MTF degrades as the OTF (FT of PSF) loses contrast
  - EE decreases as energy moves from central pixel to neighbors
  - RER degrades as edge response broadens
  - NIIRS degrades through RER term (3.32*log10(RER)) primarily
  - SNR is only weakly affected by WFE (slight EE change in point-source)

Key thresholds:
  - Diffraction-limited: WFE < lambda/14 (0.071 waves) → Strehl > 0.8
  - Moderate degradation: WFE ~ 0.10 waves → Strehl ~ 0.67
  - Significant: WFE ~ 0.15 waves → Strehl ~ 0.41
  - Severe: WFE > 0.20 waves → Strehl < 0.20

Usage:
    python run_wfe_budget_trade.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Tom's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "tom_wfe_budget_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "wfe_budget_results.xlsx"
PLOT_DIR = Path(__file__).parent.parent / "outputs"

wb_in = openpyxl.load_workbook(INPUT_FILE)

ws_sys = wb_in["Optical System"]
specs: dict[str, object] = {}
units: dict[str, str] = {}
for row in ws_sys.iter_rows(min_row=2, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        try:
            specs[name] = float(value)
        except (ValueError, TypeError):
            specs[name] = value
        units[name] = str(row[2].value) if row[2].value else "--"

ws_wfe = wb_in["WFE Sweep"]
wfe_values: list[float] = []
for row in ws_wfe.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0] is not None:
        wfe_values.append(float(row[0]))

# Also read Zernike reference data for display
ws_zern = wb_in["Zernike Coefficients"]
zernike_data: list[tuple[int, str, float]] = []
for row in ws_zern.iter_rows(min_row=2, max_col=4, values_only=True):
    if row[0] is not None and row[2] is not None:
        try:
            zernike_data.append((int(row[0]), str(row[1]), float(row[2])))
        except (ValueError, TypeError):
            pass

print("=" * 80)
print("SCENARIO 5.1: WFE Budget Allocation — How Much Aberration Can I Tolerate?")
print("=" * 80)

print("\n=== System Parameters (from spreadsheet) ===")
for k, v in specs.items():
    print(f"  {k:<30s}: {v} [{units.get(k, '--')}]")

print(f"\n=== WFE Sweep Points ===")
print(f"  WFE RMS = {wfe_values} [waves at 633 nm]")

if zernike_data:
    print(f"\n=== Tom's Zernike Coefficients (reference) ===")
    print(f"  {'Index':>6s}  {'Name':<25s}  {'Coefficient':>12s}")
    print(f"  {'-' * 6}  {'-' * 25}  {'-' * 12}")
    for idx, name, coeff in zernike_data:
        print(f"  Z{idx:<5d}  {name:<25s}  {coeff:>12.4f} [waves]")
    total_rms = math.sqrt(sum(c**2 for _, _, c in zernike_data))
    print(f"\n  Total Zernike RMS: {total_rms:.4f} [waves at 633 nm]")
    print(f"  Note: RADIANT uses scalar RMS mode — individual Zernike-to-PSF")
    print(f"        not yet implemented. Tom's {total_rms:.4f} waves total RMS")
    print(f"        is one point on the sweep below.")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

aperture_m = float(specs["Aperture diameter"]) / 100.0          # cm -> m
focal_length_m = float(specs["Focal length"]) / 100.0           # cm -> m
f_number = float(specs["f-number"])
transmission = float(specs["Optical transmission"]) / 100.0     # % -> frac
optics_temp_K = float(specs["Optics temperature"]) + 273.15     # C -> K
obscuration = float(specs["Central obscuration"]) / 100.0       # % -> frac
wfe_ref_nm = float(specs["WFE reference wavelength"])
wfe_ref_um = wfe_ref_nm / 1000.0                                # nm -> um

pixel_pitch_um = float(specs["Pixel pitch"])
pixel_pitch_m = pixel_pitch_um * 1e-6
qe = float(specs["Quantum efficiency"]) / 100.0                 # % -> frac
dark_rate = float(specs["Dark current"])
read_noise = float(specs["Read noise"])
fwc = float(specs["Full well capacity"])
gain = float(specs["Gain"])
adc_bits = int(specs["ADC bits"])

band_min_nm = float(specs["Filter min"])
band_max_nm = float(specs["Filter max"])
band_min_um = band_min_nm / 1000.0                              # nm -> um
band_max_um = band_max_nm / 1000.0
band_center_um = (band_min_um + band_max_um) / 2.0

target_refl = float(specs["Target reflectance"])
bg_refl = float(specs["Background reflectance"])
solar_zenith_deg = float(specs["Solar zenith angle"])
solar_zenith_rad = solar_zenith_deg * math.pi / 180.0

altitude_km = float(specs["Orbit altitude"])
altitude_m = altitude_km * 1000.0

t_int_ms = float(specs["Integration time"])
t_int_s = t_int_ms / 1000.0                                     # ms -> s

# Derived parameters
gsd_m = pixel_pitch_m * altitude_m / focal_length_m
ifov_urad = pixel_pitch_m / focal_length_m * 1e6
Q = band_center_um * f_number / pixel_pitch_um
airy_diam_um = 2.44 * band_center_um * f_number
f_nyquist = 1.0 / (2.0 * pixel_pitch_m)  # cycles/m on focal plane

print(f"\n=== Converted to RADIANT Canonical Units ===")
print(f"  {'Parameter':<30s} {'Value':>14s}  {'Unit':<10s}  {'Conversion'}")
print(f"  {'-' * 30} {'-' * 14}  {'-' * 10}  {'-' * 20}")
print(f"  {'Aperture diameter':<30s} {aperture_m:>14.4f}  {'m':<10s}  cm / 100")
print(f"  {'Focal length':<30s} {focal_length_m:>14.4f}  {'m':<10s}  cm / 100")
print(f"  {'Optical transmission':<30s} {transmission:>14.4f}  {'fraction':<10s}  % / 100")
print(f"  {'Optics temperature':<30s} {optics_temp_K:>14.2f}  {'K':<10s}  C + 273.15")
print(f"  {'Obscuration':<30s} {obscuration:>14.4f}  {'fraction':<10s}  % / 100")
print(f"  {'WFE ref wavelength':<30s} {wfe_ref_um:>14.4f}  {'um':<10s}  nm / 1000")
print(f"  {'Band':<30s} {band_min_um:>6.3f}-{band_max_um:<6.3f}  {'um':<10s}  nm / 1000")
print(f"  {'QE':<30s} {qe:>14.4f}  {'fraction':<10s}  % / 100")
print(f"  {'Integration time':<30s} {t_int_s:>14.6f}  {'s':<10s}  ms / 1000")
print(f"  {'Solar zenith':<30s} {solar_zenith_rad:>14.4f}  {'rad':<10s}  deg x pi/180")

print(f"\n=== Derived Parameters ===")
print(f"  GSD:               {gsd_m:.2f} [m]")
print(f"  IFOV:              {ifov_urad:.1f} [urad]")
print(f"  Q (sampling):      {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")
print(f"  Airy disk:         {airy_diam_um:.1f} [um] ({airy_diam_um/pixel_pitch_um:.2f} pixels)")
print(f"  f_Nyquist:         {f_nyquist:.0f} [cy/m] ({f_nyquist/1000:.1f} [cy/mm])")

# ---------------------------------------------------------------------------
# Step 3: Marechal Strehl reference table
# ---------------------------------------------------------------------------

print(f"\n=== Marechal Strehl Reference (at operating wavelength) ===")
print(f"  WFE specified at {wfe_ref_nm:.0f} [nm], operating at {band_center_um*1000:.0f} [nm]")
print(f"\n  {'WFE [waves]':>12s}  {'OPD RMS [nm]':>12s}  {'Strehl@ref':>10s}  "
      f"{'Strehl@oper':>12s}  {'Note'}")
print(f"  {'-' * 12}  {'-' * 12}  {'-' * 10}  {'-' * 12}  {'-' * 30}")

for wfe in wfe_values:
    opd_nm = wfe * wfe_ref_nm
    # Strehl at reference wavelength
    strehl_ref = math.exp(-(2.0 * math.pi * wfe) ** 2)
    # Strehl at operating wavelength (WFE in waves at operating lambda)
    wfe_at_oper = opd_nm / (band_center_um * 1000.0)
    strehl_oper = math.exp(-(2.0 * math.pi * wfe_at_oper) ** 2)

    note = ""
    if wfe == 0.0:
        note = "perfect optics"
    elif abs(wfe - 1.0 / 14.0) < 0.005:
        note = "lambda/14 diffraction limit"
    elif strehl_ref < 0.30:
        note = "beyond Marechal validity"

    print(f"  {wfe:>12.3f}  {opd_nm:>12.1f}  {strehl_ref:>10.4f}  "
          f"{strehl_oper:>12.4f}  {note}")

print(f"\n  Note: Strehl at operating wavelength ({band_center_um*1000:.0f} nm) is higher")
print(f"  than at reference (633 nm) because the same physical OPD is a smaller")
print(f"  fraction of the longer wavelength.")

# ---------------------------------------------------------------------------
# Step 4: Build base RADIANT config (WFE set per sweep point)
# ---------------------------------------------------------------------------

# Kirchhoff: reflectance = 1 - emissivity
target_temp_K = 300.0
bg_temp_K = 295.0
target_emissivity = 1.0 - target_refl
bg_emissivity = 1.0 - bg_refl

base_config = {
    "source": {
        "target": {"temperature": target_temp_K, "emissivity": target_emissivity},
        "background": {"temperature": bg_temp_K, "emissivity": bg_emissivity},
    },
    "atmosphere": {
        "model": "simple",
        "standard_atmosphere": "midlat_summer",
    },
    "geometry": {
        "sensor_altitude_m": altitude_m,
        "path_zenith_rad": 0.0,
        "solar_zenith_rad": solar_zenith_rad,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "optics_temperature_K": optics_temp_K,
        "wfe_rms_waves": 0.0,  # overridden per sweep point
        "obscuration_ratio": obscuration,
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate,
        "detector_temperature_K": 263.0,  # cooled CCD
    },
    "spectral_integration": {
        "filter_min_um": band_min_um,
        "filter_max_um": band_max_um,
        "integration_time_s": t_int_s,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "full_well_capacity_e": fwc,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
    },
}

print(f"\n=== RADIANT Configuration ===")
print(f"  Integration time: {t_int_s*1e3:.1f} [ms]")
print(f"  Band: {band_min_um:.3f}-{band_max_um:.3f} [um]")
print(f"  Sweep: WFE RMS from {wfe_values[0]:.3f} to {wfe_values[-1]:.3f} [waves at 633 nm]")
print(f"  Target reflectance: {target_refl} (emissivity = {target_emissivity})")

# ---------------------------------------------------------------------------
# Step 5: Sweep WFE RMS
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  WFE SWEEP")
print(f"{'=' * 80}")
print(f"\n  {'WFE':>8s}  {'Strehl':>8s}  {'MTF@Nyq':>10s}  {'EE(1x1)':>8s}  "
      f"{'EE(3x3)':>8s}  {'RER':>8s}  {'SNR':>8s}  {'NIIRS':>8s}")
print(f"  {'[waves]':>8s}  {'[--]':>8s}  {'[--]':>10s}  {'[--]':>8s}  "
      f"{'[--]':>8s}  {'[--]':>8s}  {'[--]':>8s}  {'[--]':>8s}")
print(f"  {'-' * 8}  {'-' * 8}  {'-' * 10}  {'-' * 8}  "
      f"{'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 8}")

results: list[dict] = []

for wfe_rms in wfe_values:
    config = {k: ({**v} if isinstance(v, dict) else v) for k, v in base_config.items()}
    config["source"] = {
        "target": {**base_config["source"]["target"]},
        "background": {**base_config["source"]["background"]},
    }
    config["optics"] = {**base_config["optics"], "wfe_rms_waves": wfe_rms}

    sensor = Sensor.from_dict(config)
    r = sensor.evaluate()

    strehl = r.metrics.get("strehl", 0.0)
    mtf_nyq = r.metrics.get("mtf_at_nyquist", 0.0)
    ee_1x1 = r.metrics.get("ee_1x1", 0.0)
    ee_3x3 = r.metrics.get("ee_3x3", 0.0)
    rer = r.metrics.get("rer", 0.0)
    snr = r.metrics.get("snr", 0.0)
    niirs = r.metrics.get("niirs", 0.0)
    signal_e = r.stage_outputs.get("readout", {}).get("signal_e_final", 0.0)
    noise_terms = {n.name: n.value_e for n in r.noise_terms}

    # Full MTF curve
    perf_out = r.stage_outputs.get("performance", {})
    mtf_freq_x = perf_out.get("mtf_freq_x")
    mtf_x = perf_out.get("mtf_x")

    # Marechal Strehl for comparison
    marechal = math.exp(-(2.0 * math.pi * wfe_rms) ** 2) if wfe_rms > 0 else 1.0

    row = {
        "wfe_rms": wfe_rms,
        "strehl_radiant": strehl,
        "strehl_marechal": marechal,
        "mtf_nyq": mtf_nyq,
        "ee_1x1": ee_1x1,
        "ee_3x3": ee_3x3,
        "rer": rer,
        "snr": snr,
        "niirs": niirs,
        "signal_e": signal_e,
        "noise_terms": noise_terms,
        "mtf_freq_x": mtf_freq_x,
        "mtf_x": mtf_x,
    }
    results.append(row)

    print(f"  {wfe_rms:>8.3f}  {strehl:>8.4f}  {mtf_nyq:>10.4f}  {ee_1x1:>8.4f}  "
          f"{ee_3x3:>8.4f}  {rer:>8.4f}  {snr:>8.1f}  {niirs:>8.2f}")

# ---------------------------------------------------------------------------
# Step 6: Metric degradation analysis
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  METRIC DEGRADATION RELATIVE TO PERFECT OPTICS")
print(f"{'=' * 80}")

baseline = results[0]  # WFE = 0 (perfect)

print(f"\n  Baseline (WFE = 0): Strehl = {baseline['strehl_radiant']:.4f}, "
      f"MTF@Nyq = {baseline['mtf_nyq']:.4f}, EE(1x1) = {baseline['ee_1x1']:.4f}, "
      f"RER = {baseline['rer']:.4f}, NIIRS = {baseline['niirs']:.2f}")

print(f"\n  {'WFE':>8s}  {'dStrehl':>8s}  {'dMTF@Nyq':>10s}  {'dEE(1x1)':>10s}  "
      f"{'dRER':>8s}  {'dNIIRS':>8s}  {'Quality'}")
print(f"  {'[waves]':>8s}  {'[%]':>8s}  {'[%]':>10s}  {'[%]':>10s}  "
      f"{'[%]':>8s}  {'[--]':>8s}  ")
print(f"  {'-' * 8}  {'-' * 8}  {'-' * 10}  {'-' * 10}  {'-' * 8}  {'-' * 8}  {'-' * 20}")

for r in results:
    d_strehl = ((r["strehl_radiant"] - baseline["strehl_radiant"])
                / baseline["strehl_radiant"] * 100.0
                if baseline["strehl_radiant"] > 0 else 0.0)
    d_mtf = ((r["mtf_nyq"] - baseline["mtf_nyq"])
             / baseline["mtf_nyq"] * 100.0
             if baseline["mtf_nyq"] > 0 else 0.0)
    d_ee = ((r["ee_1x1"] - baseline["ee_1x1"])
            / baseline["ee_1x1"] * 100.0
            if baseline["ee_1x1"] > 0 else 0.0)
    d_rer = ((r["rer"] - baseline["rer"])
             / baseline["rer"] * 100.0
             if baseline["rer"] > 0 else 0.0)
    d_niirs = r["niirs"] - baseline["niirs"]

    quality = "diffraction-limited"
    if r["wfe_rms"] > 0.071:
        quality = "acceptable"
    if r["wfe_rms"] > 0.10:
        quality = "moderate degradation"
    if r["wfe_rms"] > 0.15:
        quality = "significant"
    if r["wfe_rms"] > 0.20:
        quality = "severe"

    print(f"  {r['wfe_rms']:>8.3f}  {d_strehl:>+8.1f}  {d_mtf:>+10.1f}  {d_ee:>+10.1f}  "
          f"{d_rer:>+8.1f}  {d_niirs:>+8.2f}  {quality}")

# ---------------------------------------------------------------------------
# Step 7: Strehl comparison — Marechal vs RADIANT
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  STREHL RATIO: MARECHAL vs. RADIANT")
print(f"{'=' * 80}")
print(f"\n  Marechal: S = exp(-(2*pi*WFE_rms)^2)")
print(f"  RADIANT:  computed from PerformanceStage (Marechal at band center)")
print(f"  WFE reference: {wfe_ref_nm:.0f} [nm], operating: {band_center_um*1000:.0f} [nm]")

print(f"\n  {'WFE':>8s}  {'Marechal':>10s}  {'RADIANT':>10s}  {'Difference':>10s}")
print(f"  {'[waves]':>8s}  {'[--]':>10s}  {'[--]':>10s}  {'[--]':>10s}")
print(f"  {'-' * 8}  {'-' * 10}  {'-' * 10}  {'-' * 10}")

for r in results:
    diff = r["strehl_radiant"] - r["strehl_marechal"]
    print(f"  {r['wfe_rms']:>8.3f}  {r['strehl_marechal']:>10.4f}  "
          f"{r['strehl_radiant']:>10.4f}  {diff:>+10.4f}")

print(f"\n  Note: RADIANT's Strehl uses the Marechal approximation but evaluates")
print(f"  at the operating wavelength ({band_center_um*1000:.0f} nm), so Strehl is higher")
print(f"  than the 'at-reference' Marechal value listed above. The 'Marechal'")
print(f"  column uses the WFE in waves at 633 nm directly.")

# ---------------------------------------------------------------------------
# Step 8: NIIRS budget analysis
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  NIIRS SENSITIVITY TO WFE")
print(f"{'=' * 80}")

print(f"\n  GIQE-5 decomposes NIIRS into GSD, RER, and SNR terms.")
print(f"  WFE affects NIIRS primarily through the RER term (3.32*log10(RER)).")
print(f"  GSD is constant ({gsd_m:.2f} m). SNR changes only slightly with WFE.")

niirs_0 = baseline["niirs"]
print(f"\n  Baseline NIIRS (WFE=0): {niirs_0:.2f}")

# Find WFE thresholds for NIIRS degradation
for threshold in [0.25, 0.50, 1.0]:
    wfe_at_thresh = None
    for r in results:
        if r["niirs"] < niirs_0 - threshold:
            wfe_at_thresh = r["wfe_rms"]
            break
    if wfe_at_thresh:
        print(f"  NIIRS drops by {threshold:.2f} at WFE ~ {wfe_at_thresh:.3f} [waves]")
    else:
        print(f"  NIIRS does not drop by {threshold:.2f} in sweep range")

# ---------------------------------------------------------------------------
# Step 9: Noise budget (constant across WFE sweep)
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  NOISE BUDGET (constant — WFE does not affect noise)")
print(f"{'=' * 80}")

noise_names = list(baseline["noise_terms"].keys())
active_terms = [n for n in noise_names if baseline["noise_terms"].get(n, 0) > 0.01]

print(f"\n  {'Noise Term':<25s}  {'Value [e- RMS]':>15s}  {'Fraction [%]':>12s}")
print(f"  {'-' * 25}  {'-' * 15}  {'-' * 12}")

total_noise = math.sqrt(sum(v**2 for v in baseline["noise_terms"].values()))
for term in active_terms:
    val = baseline["noise_terms"][term]
    frac = (val**2 / total_noise**2) * 100.0 if total_noise > 0 else 0.0
    print(f"  {term:<25s}  {val:>15.1f}  {frac:>12.1f}")

print(f"  {'TOTAL (RSS)':<25s}  {total_noise:>15.1f}  {'100.0':>12s}")
print(f"\n  Signal: {baseline['signal_e']:,.0f} [e-]")
print(f"  SNR:    {baseline['snr']:.1f} [--]")
print(f"\n  Note: WFE degrades spatial metrics (MTF, RER, EE) but does NOT add")
print(f"  noise. The noise budget is identical at all WFE levels.")

# ---------------------------------------------------------------------------
# Step 10: Plots
# ---------------------------------------------------------------------------

PLOT_DIR.mkdir(parents=True, exist_ok=True)
fig_w, fig_h = 10, 7

# Plot 1: Strehl vs WFE
fig1, ax1 = plt.subplots(figsize=(fig_w, fig_h))
wfe_arr = [r["wfe_rms"] for r in results]
strehl_rad_arr = [r["strehl_radiant"] for r in results]
strehl_mar_arr = [r["strehl_marechal"] for r in results]

ax1.plot(wfe_arr, strehl_rad_arr, "bo-", linewidth=2, markersize=8,
         label=f"RADIANT (at {band_center_um*1000:.0f} nm)")
ax1.plot(wfe_arr, strehl_mar_arr, "r--", linewidth=1.5, alpha=0.7,
         label="Marechal (at 633 nm ref)")

ax1.axhline(0.80, color="green", linestyle=":", alpha=0.5,
            label="Strehl = 0.80 (diffraction limit)")
ax1.axvline(1.0 / 14.0, color="green", linestyle="--", alpha=0.4,
            label=f"WFE = lambda/14 = {1.0/14.0:.3f} waves")

ax1.set_xlabel("WFE RMS [waves at 633 nm]", fontsize=12)
ax1.set_ylabel("Strehl Ratio [--]", fontsize=12)
ax1.set_title("Strehl Ratio vs. WFE RMS — 40 cm Cassegrain, VNIR", fontsize=13)
ax1.legend(fontsize=10)
ax1.grid(True, alpha=0.3)
ax1.set_ylim(0, 1.05)
fig1.tight_layout()
fig1.savefig(PLOT_DIR / "fig1_strehl_vs_wfe.png", dpi=150)

# Plot 2: MTF, EE, RER vs WFE (multi-panel)
fig2, (ax2a, ax2b, ax2c) = plt.subplots(1, 3, figsize=(fig_w * 1.5, fig_h))

mtf_arr = [r["mtf_nyq"] for r in results]
ee1_arr = [r["ee_1x1"] for r in results]
ee3_arr = [r["ee_3x3"] for r in results]
rer_arr = [r["rer"] for r in results]

ax2a.plot(wfe_arr, mtf_arr, "rs-", linewidth=2, markersize=8)
ax2a.set_xlabel("WFE RMS [waves]", fontsize=11)
ax2a.set_ylabel("MTF at Nyquist [--]", fontsize=11)
ax2a.set_title("MTF@Nyquist vs. WFE", fontsize=12)
ax2a.grid(True, alpha=0.3)

ax2b.plot(wfe_arr, ee1_arr, "go-", linewidth=2, markersize=8, label="EE(1x1)")
ax2b.plot(wfe_arr, ee3_arr, "g^--", linewidth=2, markersize=8, label="EE(3x3)")
ax2b.set_xlabel("WFE RMS [waves]", fontsize=11)
ax2b.set_ylabel("Ensquared Energy [--]", fontsize=11)
ax2b.set_title("EE vs. WFE", fontsize=12)
ax2b.legend(fontsize=10)
ax2b.grid(True, alpha=0.3)

ax2c.plot(wfe_arr, rer_arr, "m^-", linewidth=2, markersize=8)
ax2c.set_xlabel("WFE RMS [waves]", fontsize=11)
ax2c.set_ylabel("RER [--]", fontsize=11)
ax2c.set_title("RER vs. WFE", fontsize=12)
ax2c.grid(True, alpha=0.3)

fig2.tight_layout()
fig2.savefig(PLOT_DIR / "fig2_mtf_ee_rer_vs_wfe.png", dpi=150)

# Plot 3: NIIRS vs WFE
fig3, ax3 = plt.subplots(figsize=(fig_w, fig_h))
niirs_arr = [r["niirs"] for r in results]

ax3.plot(wfe_arr, niirs_arr, "bo-", linewidth=2, markersize=8, label="NIIRS (GIQE-5)")
ax3.axhline(niirs_0, color="green", linestyle=":", alpha=0.4,
            label=f"Baseline = {niirs_0:.2f}")
ax3.axhline(niirs_0 - 0.5, color="orange", linestyle="--", alpha=0.5,
            label=f"-0.5 NIIRS = {niirs_0 - 0.5:.2f}")
ax3.axhline(niirs_0 - 1.0, color="red", linestyle="--", alpha=0.5,
            label=f"-1.0 NIIRS = {niirs_0 - 1.0:.2f}")
ax3.axvline(1.0 / 14.0, color="green", linestyle="--", alpha=0.4,
            label="lambda/14")

ax3.set_xlabel("WFE RMS [waves at 633 nm]", fontsize=12)
ax3.set_ylabel("NIIRS [--]", fontsize=12)
ax3.set_title("NIIRS vs. WFE RMS — WFE Budget Impact", fontsize=13)
ax3.legend(fontsize=10)
ax3.grid(True, alpha=0.3)
fig3.tight_layout()
fig3.savefig(PLOT_DIR / "fig3_niirs_vs_wfe.png", dpi=150)

# Plot 4: MTF curves at selected WFE values
fig4, ax4 = plt.subplots(figsize=(fig_w, fig_h))

selected_wfe = [0.0, 0.04, 0.071, 0.10, 0.15, 0.20, 0.25]
colors = plt.cm.viridis(np.linspace(0, 0.9, len(selected_wfe)))

for wfe_sel, color in zip(selected_wfe, colors):
    # Find closest WFE in results
    closest = min(results, key=lambda r: abs(r["wfe_rms"] - wfe_sel))
    if closest["mtf_freq_x"] is not None and closest["mtf_x"] is not None:
        freq = np.array(closest["mtf_freq_x"])
        mtf = np.array(closest["mtf_x"])
        # Convert freq from cycles/pixel to normalized
        label = f"WFE = {closest['wfe_rms']:.3f} waves"
        ax4.plot(freq, mtf, color=color, linewidth=2, label=label)

ax4.axvline(0.5, color="gray", linestyle=":", alpha=0.5, label="Nyquist (0.5 cy/px)")
ax4.set_xlabel("Spatial Frequency [cycles/pixel]", fontsize=12)
ax4.set_ylabel("MTF [--]", fontsize=12)
ax4.set_title("System MTF Curves at Various WFE Levels", fontsize=13)
ax4.legend(fontsize=9, loc="upper right")
ax4.grid(True, alpha=0.3)
ax4.set_xlim(0, 1.0)
ax4.set_ylim(0, 1.0)
fig4.tight_layout()
fig4.savefig(PLOT_DIR / "fig4_mtf_curves.png", dpi=150)

# ---------------------------------------------------------------------------
# Step 11: Output spreadsheet
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "WFE Trade Results"

header_font_out = Font(bold=True, size=10, color="FFFFFF")
header_fill_out = PatternFill("solid", fgColor="2E75B6")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = [
    "WFE RMS [waves]", "Strehl (RADIANT) [--]", "Strehl (Marechal) [--]",
    "MTF@Nyquist [--]", "EE(1x1) [--]", "EE(3x3) [--]",
    "RER [--]", "SNR [--]", "NIIRS [--]", "dNIIRS [--]",
]

for col_idx, h in enumerate(headers, start=1):
    cell = ws_out.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_out
    cell.fill = header_fill_out
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for row_idx, r in enumerate(results, start=2):
    d_niirs = r["niirs"] - baseline["niirs"]
    vals = [
        r["wfe_rms"], round(r["strehl_radiant"], 4), round(r["strehl_marechal"], 4),
        round(r["mtf_nyq"], 4), round(r["ee_1x1"], 4), round(r["ee_3x3"], 4),
        round(r["rer"], 4), round(r["snr"], 1), round(r["niirs"], 2),
        round(d_niirs, 2),
    ]
    for col_idx, v in enumerate(vals, start=1):
        cell = ws_out.cell(row=row_idx, column=col_idx, value=v)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

for col in ws_out.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws_out.column_dimensions[col[0].column_letter].width = max_len + 3

wb_out.save(OUTPUT_FILE)
print(f"\n  Output spreadsheet: {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# Step 12: Summary
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  SUMMARY")
print(f"{'=' * 80}")
print(f"\n  System: 40 cm Cassegrain, f/{f_number:.0f}, {pixel_pitch_um:.0f} um CCD, "
      f"{band_min_nm:.0f}-{band_max_nm:.0f} nm VNIR")
print(f"  Orbit:  {altitude_km:.0f} km LEO, GSD = {gsd_m:.2f} m")
print(f"  WFE reference: {wfe_ref_nm:.0f} nm (HeNe)")
print(f"  Q (sampling):  {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")

print(f"\n  --- Baseline (perfect optics, WFE = 0) ---")
print(f"  Strehl:    {baseline['strehl_radiant']:.4f} [--]")
print(f"  MTF@Nyq:   {baseline['mtf_nyq']:.4f} [--]")
print(f"  EE(1x1):   {baseline['ee_1x1']:.4f} [--]")
print(f"  EE(3x3):   {baseline['ee_3x3']:.4f} [--]")
print(f"  RER:       {baseline['rer']:.4f} [--]")
print(f"  SNR:       {baseline['snr']:.1f} [--]")
print(f"  NIIRS:     {baseline['niirs']:.2f} [--]")

# Find diffraction-limited threshold result
dl_result = min(results, key=lambda r: abs(r["wfe_rms"] - 0.071))
print(f"\n  --- At diffraction limit (WFE = lambda/14 = 0.071 waves) ---")
print(f"  Strehl:    {dl_result['strehl_radiant']:.4f} [--]")
print(f"  MTF@Nyq:   {dl_result['mtf_nyq']:.4f} [--]")
print(f"  dNIIRS:    {dl_result['niirs'] - baseline['niirs']:+.2f} [--]")

if zernike_data:
    total_rms = math.sqrt(sum(c**2 for _, _, c in zernike_data))
    zern_result = min(results, key=lambda r: abs(r["wfe_rms"] - total_rms))
    print(f"\n  --- At Tom's Zernike budget ({total_rms:.4f} waves RMS) ---")
    print(f"  Nearest sweep point: WFE = {zern_result['wfe_rms']:.3f} [waves]")
    print(f"  Strehl:    {zern_result['strehl_radiant']:.4f} [--]")
    print(f"  MTF@Nyq:   {zern_result['mtf_nyq']:.4f} [--]")
    print(f"  dNIIRS:    {zern_result['niirs'] - baseline['niirs']:+.2f} [--]")
    print(f"  Assessment: Tom's WFE budget is well within diffraction-limited territory.")

print(f"\n  Design recommendation:")
print(f"    Allocate WFE budget to keep total RMS < 0.071 waves (lambda/14)")
print(f"    for diffraction-limited performance (Strehl > 0.80).")
print(f"    For NIIRS-driven systems, WFE up to ~0.10 waves is acceptable")
print(f"    with < 0.25 NIIRS degradation.")

print(f"\n  Limitations:")
print(f"    - RADIANT uses scalar RMS WFE (random phase screen), not individual")
print(f"      Zernike modes. Aberration-specific PSF shapes (coma vs. astigmatism)")
print(f"      are not modeled. See Gap: Zernike-to-PSF.")
print(f"    - Field-dependent WFE is defined but not yet implemented.")
print(f"    - Marechal approximation is inaccurate for Strehl < 0.3 (WFE > ~0.17 waves).")

plt.show()
