"""Scenario 2.1: InSb vs. HgCdTe Noise Budget Shootout at 77 K.

Mike is choosing between an InSb and an HgCdTe MWIR FPA for a space
mission. He has vendor QE(λ) and J_dark(T) data for both — in two
DIFFERENT CSV conventions — plus a shared ROIC spec. He wants a
side-by-side noise budget on a common 3.5–5.0 µm bench test, and the
three trade numbers that decide the cooler budget:

  1. Dark-current crossover temperature — where dark shot noise equals
     read noise (dark current stops being negligible).
  2. BLIP temperature — where dark current equals the photon-generated
     rate (detector stops being background-limited).
  3. NEI — noise-equivalent irradiance at the focal plane.

This script:
  1. Loads the vendor QE curves with radiant.io.qe_csv.load_qe_csv
     (InSb: wavelength_nm/QE_pct; HgCdTe: lambda_um/quantum_efficiency —
     the loader resolves both conventions from the headers).
  2. Loads J_dark(T) with radiant.io.dark_current_csv.load_dark_current_csv
     and converts A/cm² → e⁻/s at 77 K for the 15 µm pixel.
  3. Runs the full chain per detector with the SPECTRAL QE curve injected
     (stage_outputs["spectral_integration"]["qe_curve"], Rule 6 route) and
     compares against a band-averaged scalar-QE run.
  4. Prints side-by-side noise budgets (every chain noise term).
  5. Computes crossover and BLIP temperatures via the loader's exact
     Arrhenius inverse (temperature_at_rate).
  6. Computes photon-flux NEI from the total noise.
  7. Writes plots (J_dark(T), QE curves, noise budgets) and a workbook.

Usage:
    python run_detector_shootout.py
"""

import math
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from radiant.api.session import RadiantSession
from radiant.core.constants import h, c
from radiant.io.config import load_config
from radiant.io.dark_current_csv import load_dark_current_csv
from radiant.io.qe_csv import load_qe_csv

# ---------------------------------------------------------------------------
# Step 1: Load vendor inputs
# ---------------------------------------------------------------------------

INPUTS = Path(__file__).parent.parent / "inputs"

# --- Vendor QE curves: two different CSV conventions, one loader ---
qe_insb = load_qe_csv(INPUTS / "insb_qe.csv")        # wavelength_nm, QE_pct
qe_hgcdte = load_qe_csv(INPUTS / "hgcdte_qe.csv")    # lambda_um, quantum_efficiency

# --- Vendor dark-current curves ---
jd_insb = load_dark_current_csv(INPUTS / "insb_jdark.csv")
jd_hgcdte = load_dark_current_csv(INPUTS / "hgcdte_jdark.csv")

# --- ROIC + bench workbook ---
wb_in = openpyxl.load_workbook(INPUTS / "mike_roic_specs.xlsx")
ws_in = wb_in["ROIC and Test Config"]
specs: dict[str, object] = {}
for row in ws_in.iter_rows(min_row=4, max_col=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        specs[str(row[0])] = row[1]

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

pixel_pitch_um = float(specs["Pixel pitch"])                    # already µm
pixel_pitch_m = pixel_pitch_um * 1e-6                           # µm → m
node_cap_fF = float(specs["Node capacitance"])
node_cap_F = node_cap_fF * 1e-15                                # fF → F
cds_enabled = int(specs["CDS mode"])
glow_e_per_s = float(specs["ROIC glow"])                        # already e⁻/s
fwc = float(specs["Full well capacity"])                        # already e⁻
adc_bits = int(specs["ADC resolution"])
gain = float(specs["System gain"])                              # already e⁻/DN
read_noise = {
    "InSb": float(specs["Read noise (CDS) — InSb FPA"]),        # e⁻ RMS
    "HgCdTe": float(specs["Read noise (CDS) — HgCdTe FPA"]),
}
bb_temp_K = float(specs["Blackbody temperature"])               # already K
bb_emiss = float(specs["Blackbody emissivity"])
aperture_m = float(specs["Collimator aperture"]) / 100.0        # cm → m
focal_length_m = float(specs["Collimator focal length"]) / 100.0  # cm → m
transmission = float(specs["Optical transmission"]) / 100.0     # % → fraction
optics_temp_K = float(specs["Optics temperature"])              # already K
band_parts = str(specs["Cold filter passband"]).replace("–", "-").split("-")
band_min_um = float(band_parts[0]) / 1000.0                     # nm → µm
band_max_um = float(band_parts[1]) / 1000.0
t_int_s = float(specs["Integration time"]) / 1000.0             # ms → s
t_op_K = float(specs["Operating temperature (nominal)"])        # already K

# Dark rates at the nominal operating temperature (A/cm² → e⁻/s, Rule 2
# conversion inside the loader).
dark_rate = {
    "InSb": jd_insb.dark_rate_e_per_s(t_op_K, pixel_pitch_m=pixel_pitch_m),
    "HgCdTe": jd_hgcdte.dark_rate_e_per_s(t_op_K, pixel_pitch_m=pixel_pitch_m),
}
qe_band = {
    "InSb": qe_insb.band_averaged_qe(band_min_um, band_max_um),
    "HgCdTe": qe_hgcdte.band_averaged_qe(band_min_um, band_max_um),
}

# ---------------------------------------------------------------------------
# Step 3: Run the chain per detector with spectral QE injected
# ---------------------------------------------------------------------------
# The 300 K flat-plate blackbody fills the aperture → extended regime,
# "exo" vacuum path (bench test). Two architecture notes:
#   - Gap 42: "exo" routes through the no_atmosphere 'space' sub-case, so
#     the bench needs the positive platform.h_sensor placeholder (1.0 m).
#   - Spectral QE has no config path (detector.qe_table_path is schema-only,
#     unwired — recorded in gaps.md): the curve is injected via
#     stage_outputs["spectral_integration"]["qe_curve"] (Rule 6 route),
#     evaluated on the chain wavelength grid by the loader.

N_WL = 500
wl_grid = np.linspace(band_min_um, band_max_um, N_WL)


def build_config(det: str) -> dict:
    return {
        "source": {
            "target": {"temperature": bb_temp_K, "emissivity": bb_emiss},
            "background": {"temperature": optics_temp_K, "emissivity": 0.95},
        },
        "atmosphere": {"model": "exo"},  # bench vacuum path
        "geometry": {"sensor_altitude_m": 0.0},
        # Stage-7 stop-gap (registry Gap 42): 'space' sub-case Earth-limb
        # check needs a positive user-set altitude; 1.0 m ≈ bench height.
        "platform": {"h_sensor": 1.0},
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe_band[det],  # scalar fallback; spectral curve injected below
            "dark_rate_e_per_s": dark_rate[det],
            "detector_temperature_K": t_op_K,
            "glow_e_per_s": glow_e_per_s,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise[det],
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
            "cds_enabled": cds_enabled,
            "node_capacitance_F": node_cap_F,
        },
    }


def run_chain(det: str, qe_curve_obj, *, spectral_qe: bool):
    session = RadiantSession(wavelength_um=wl_grid)
    params = session.default_params()
    load_config(build_config(det), params)
    params.resolve()
    extra = None
    if spectral_qe:
        qe_on_grid = qe_curve_obj.evaluate(wl_grid, out_of_range="error")
        extra = {"spectral_integration": {"qe_curve": qe_on_grid}}
    return session.run(params, extra_stage_outputs=extra)


def main() -> None:
    """Run the scenario analysis."""
    OUTPUTS = Path(__file__).parent.parent / "outputs"
    OUTPUT_FILE = OUTPUTS / "detector_shootout_results.xlsx"

    print("=" * 95)
    print("  SCENARIO 2.1: InSb vs. HgCdTe Noise Budget Shootout at 77 K")
    print("=" * 95)

    print("\n=== Vendor QE curves (radiant.io.qe_csv, canonical µm/fraction) ===")
    for label, curve in [("InSb IRA-3541", qe_insb), ("HgCdTe MCT-5250", qe_hgcdte)]:
        print(f"  {label:<18s}: {curve.n_points} points, "
              f"{curve.wavelength_um[0]:.2f}–{curve.wavelength_um[-1]:.2f} µm, "
              f"peak QE {curve.qe.max():.2f} [--]  (from {Path(curve.source_file).name})")

    print("\n=== Vendor J_dark(T) curves (radiant.io.dark_current_csv) ===")
    for label, curve in [("InSb", jd_insb), ("HgCdTe", jd_hgcdte)]:
        print(f"  {label:<8s}: {curve.n_points} points, "
              f"{curve.temperature_K[0]:.0f}–{curve.temperature_K[-1]:.0f} K, "
              f"J(77 K) = {curve.j_dark_at(77.0):.3e} A/cm²")

    print("\n=== ROIC + bench configuration (vendor units) ===")
    for k, v in specs.items():
        print(f"  {k:<38s}: {v}")
    f_number = focal_length_m / aperture_m

    print("\n=== Converted to RADIANT canonical units ===")
    print(f"  {'Parameter':<30s} {'Value':>14s}  {'Unit':<10s}  {'Conversion'}")
    print(f"  {'-' * 30} {'-' * 14}  {'-' * 10}  {'-' * 22}")
    print(f"  {'Pixel pitch':<30s} {pixel_pitch_m:>14.2e}  {'m':<10s}  µm × 1e-6")
    print(f"  {'Node capacitance':<30s} {node_cap_F:>14.2e}  {'F':<10s}  fF × 1e-15")
    print(f"  {'Collimator aperture':<30s} {aperture_m:>14.4f}  {'m':<10s}  cm / 100")
    print(f"  {'Collimator focal length':<30s} {focal_length_m:>14.4f}  {'m':<10s}  cm / 100")
    print(f"  {'f-number (derived)':<30s} {f_number:>14.2f}  {'--':<10s}  f / D")
    print(f"  {'Optical transmission':<30s} {transmission:>14.3f}  {'fraction':<10s}  % / 100")
    print(f"  {'Band':<30s} {band_min_um:>6.2f}-{band_max_um:<6.2f}  {'µm':<10s}  nm / 1000")
    print(f"  {'Integration time':<30s} {t_int_s:>14.4f}  {'s':<10s}  ms / 1000")

    print(f"\n  Derived per-detector quantities at T = {t_op_K:.0f} K:")
    print(f"  {'Detector':<10s} {'J_dark [A/cm²]':>16s}  {'Dark rate [e⁻/s]':>17s}  "
          f"{'Band-avg QE [--]':>17s}  {'Read noise [e⁻]':>16s}")
    print(f"  {'-' * 10} {'-' * 16}  {'-' * 17}  {'-' * 17}  {'-' * 16}")
    for det, jd in [("InSb", jd_insb), ("HgCdTe", jd_hgcdte)]:
        print(f"  {det:<10s} {jd.j_dark_at(t_op_K):>16.3e}  {dark_rate[det]:>17,.0f}  "
              f"{qe_band[det]:>17.4f}  {read_noise[det]:>16.1f}")


    results = {}
    results_scalar = {}
    for det, qec in [("InSb", qe_insb), ("HgCdTe", qe_hgcdte)]:
        print(f"\n=== Running RADIANT — {det} (spectral QE injected) ===")
        results[det] = run_chain(det, qec, spectral_qe=True)
        results_scalar[det] = run_chain(det, qec, spectral_qe=False)

    regime = results["InSb"].stage_outputs["optics"]["regime"]
    print(f"\n=== Radiometric Regime ===")
    print(f"  Regime: {regime}")
    print(f"  The 300 K flat-plate blackbody fills the aperture and the pixel")
    print(f"  IFOV → extended regime. UNUSED PARAMETER NOTE: in this regime")
    print(f"  RADIANT skips the separate scene-background photon term (matrix")
    print(f"  Decision #13) — background_shot = 0 by design; the bench ambient")
    print(f"  temperature in source.background feeds only the contrast scene.")

    # ---------------------------------------------------------------------------
    # Step 4: Side-by-side noise budgets
    # ---------------------------------------------------------------------------

    noise = {det: {nt.name: nt.value_e for nt in results[det].noise_terms} for det in results}
    # The name is the tie-breaker, not decoration (CU-292): the primary key ties at
    # -0.0 for every zero-valued term (ktc_reset, background_shot, nearfield_shot are
    # all legitimately 0 in this config), and the input is a *set*, whose iteration
    # order varies with PYTHONHASHSEED. A stable sort then preserves that random order,
    # so three consecutive runs printed the zero rows three different ways.
    all_terms = sorted(
        set(noise["InSb"]) | set(noise["HgCdTe"]),
        key=lambda k: (-(noise["InSb"].get(k, 0.0) + noise["HgCdTe"].get(k, 0.0)), k),
    )

    signal_e = {det: results[det].stage_outputs["readout"]["signal_e_final"] for det in results}
    total_noise = {
        det: math.sqrt(sum(v**2 for v in noise[det].values())) for det in results
    }

    print(f"\n{'=' * 95}")
    print(f"  SIDE-BY-SIDE NOISE BUDGET AT {t_op_K:.0f} K, t_int = {t_int_s * 1e3:.1f} ms "
          f"(spectral QE)")
    print(f"{'=' * 95}")
    print(f"  {'Noise Term':<24s} | {'InSb [e⁻ RMS]':>14s} | {'HgCdTe [e⁻ RMS]':>16s} | "
          f"{'Comment':<28s}")
    print("-" * 95)
    comments = {
        "signal_shot": "√(photon e⁻) — dominant",
        "dark_shot": "√(dark rate × t_int)",
        "read_noise": "vendor CDS value",
        "glow_shot": "√(ROIC glow × t_int)",
        "ktc_reset": "0 with CDS on",
        "quantization": "gain/√12",
        "background_shot": "0 by design (extended)",
        "nearfield_shot": "0 (scalar mode, ε = 0)",
    }
    for term in all_terms:
        vi = noise["InSb"].get(term, 0.0)
        vh = noise["HgCdTe"].get(term, 0.0)
        if vi == 0.0 and vh == 0.0 and term not in comments:
            continue
        print(f"  {term:<24s} | {vi:>14.2f} | {vh:>16.2f} | {comments.get(term, ''):<28s}")
    print("-" * 95)
    print(f"  {'TOTAL (RSS)':<24s} | {total_noise['InSb']:>14.2f} | "
          f"{total_noise['HgCdTe']:>16.2f} |")
    print(f"  {'Signal':<24s} | {signal_e['InSb']:>14,.0f} | {signal_e['HgCdTe']:>16,.0f} | e⁻")
    print(f"  {'SNR':<24s} | {results['InSb'].metrics['snr']:>14.1f} | "
          f"{results['HgCdTe'].metrics['snr']:>16.1f} | dimensionless")

    # kTC cross-check (Rule: validation against hand calculation)
    ktc_hand_e = math.sqrt(1.380649e-23 * t_op_K * node_cap_F) / 1.602176634e-19
    print(f"\n  kTC cross-check: with CDS OFF the reset noise would be "
          f"√(k_B·T·C)/q = {ktc_hand_e:.1f} e⁻ RMS")
    print(f"  (33 fF at {t_op_K:.0f} K). RADIANT reports ktc_reset = "
          f"{noise['InSb'].get('ktc_reset', 0.0):.2f} e⁻ with CDS ON — suppressed, as configured.")

    # Spectral vs scalar QE comparison
    print(f"\n=== Spectral QE vs band-averaged scalar QE ===")
    print(f"  {'Detector':<10s} {'Signal (spectral) [e⁻]':>23s}  "
          f"{'Signal (scalar) [e⁻]':>21s}  {'Δ [%]':>7s}")
    print(f"  {'-' * 10} {'-' * 23}  {'-' * 21}  {'-' * 7}")
    for det in results:
        s_sp = signal_e[det]
        s_sc = results_scalar[det].stage_outputs["readout"]["signal_e_final"]
        print(f"  {det:<10s} {s_sp:>23,.0f}  {s_sc:>21,.0f}  "
              f"{(s_sp / s_sc - 1) * 100:>+7.2f}")
    print(f"  The spectral run photon-weights QE(λ) against the 300 K Planck")
    print(f"  spectrum (more photons at the long end of 3.5–5.0 µm); the flat")
    print(f"  scalar average cannot capture that correlation.")

    # ---------------------------------------------------------------------------
    # Step 5: Dark-current crossover and BLIP temperatures
    # ---------------------------------------------------------------------------
    # Crossover: dark_shot = read noise → rate = RN² / t_int.
    # BLIP: dark rate = photon-generated rate → rate = signal_e / t_int.
    # Both inverted EXACTLY on the vendor Arrhenius curve via
    # DarkCurrentCurve.temperature_at_rate (no sweep needed).

    print(f"\n{'=' * 95}")
    print(f"  COOLER-BUDGET TRADE: CROSSOVER AND BLIP TEMPERATURES")
    print(f"{'=' * 95}")
    print(f"\n  Definitions:")
    print(f"    Crossover T: dark shot noise √(rate·t_int) equals read noise —")
    print(f"                 rate = RN²/t_int. Below this T, dark current is a")
    print(f"                 second-order term; above it, it competes with the ROIC.")
    print(f"    BLIP T:      dark rate equals the photon-generated rate — above")
    print(f"                 this T the detector is no longer background-limited")
    print(f"                 (photon shot noise stops dominating dark noise).")

    trade_rows = []
    for det, jd in [("InSb", jd_insb), ("HgCdTe", jd_hgcdte)]:
        rn = read_noise[det]
        rate_crossover = rn**2 / t_int_s                      # e⁻/s
        rate_blip = signal_e[det] / t_int_s                   # photon-generated e⁻/s
        t_cross = jd.temperature_at_rate(rate_crossover, pixel_pitch_m=pixel_pitch_m)
        t_blip = jd.temperature_at_rate(rate_blip, pixel_pitch_m=pixel_pitch_m)
        trade_rows.append((det, rate_crossover, t_cross, rate_blip, t_blip))

    print(f"\n  {'Detector':<10s} {'RN²/t [e⁻/s]':>14s}  {'Crossover T [K]':>16s}  "
          f"{'Photon rate [e⁻/s]':>19s}  {'BLIP T [K]':>11s}")
    print(f"  {'-' * 10} {'-' * 14}  {'-' * 16}  {'-' * 19}  {'-' * 11}")
    for det, rc, tc, rb, tb in trade_rows:
        print(f"  {det:<10s} {rc:>14,.0f}  {tc:>16.1f}  {rb:>19,.0f}  {tb:>11.1f}")

    t_cross_insb = trade_rows[0][2]
    t_cross_mct = trade_rows[1][2]
    t_blip_insb = trade_rows[0][4]
    t_blip_mct = trade_rows[1][4]
    print(f"\n  Interpretation: at the same 77 K set point, HgCdTe can warm to")
    print(f"  {t_cross_mct:.1f} K before dark shot competes with its read noise, vs")
    print(f"  {t_cross_insb:.1f} K for InSb — a {t_cross_mct - t_cross_insb:.1f} K cooler margin. BLIP holds to")
    print(f"  {t_blip_mct:.1f} K (HgCdTe) vs {t_blip_insb:.1f} K (InSb). Each Kelvin of set-point")
    print(f"  margin is cooler mass/power at the mission level.")

    # ---------------------------------------------------------------------------
    # Step 6: Noise-equivalent irradiance (NEI)
    # ---------------------------------------------------------------------------
    # Photon-flux NEI: the focal-plane photon irradiance that produces
    # SNR = 1: NEI = σ_total / (QE_band · A_pix · t_int) [photons/s/cm²].
    # The W/cm² form uses the band-center photon energy (labeled approximate).

    print(f"\n{'=' * 95}")
    print(f"  NOISE-EQUIVALENT IRRADIANCE (NEI)")
    print(f"{'=' * 95}")

    a_pix_cm2 = (pixel_pitch_m * 100.0) ** 2
    lam_center_um = 0.5 * (band_min_um + band_max_um)
    e_photon_J = h * c / (lam_center_um * 1e-6)

    print(f"\n  {'Detector':<10s} {'σ_total [e⁻]':>13s}  {'NEI [photons/s/cm²]':>21s}  "
          f"{'NEI [W/cm²] (approx)':>21s}")
    print(f"  {'-' * 10} {'-' * 13}  {'-' * 21}  {'-' * 21}")
    nei_rows = []
    for det in results:
        nei_ph = total_noise[det] / (qe_band[det] * a_pix_cm2 * t_int_s)
        nei_w = nei_ph * e_photon_J
        nei_rows.append((det, total_noise[det], nei_ph, nei_w))
        print(f"  {det:<10s} {total_noise[det]:>13.1f}  {nei_ph:>21.3e}  {nei_w:>21.3e}")
    print(f"  (W/cm² uses E_photon at band center {lam_center_um:.2f} µm — an")
    print(f"  approximation; the photon-flux NEI is exact for this budget.)")

    # ---------------------------------------------------------------------------
    # Step 7: Plots
    # ---------------------------------------------------------------------------

    OUTPUTS.mkdir(parents=True, exist_ok=True)

    # Fig 1: J_dark(T) with crossover/BLIP markers
    fig1, ax1 = plt.subplots(figsize=(9, 6))
    for jd, det, color in [(jd_insb, "InSb", "tab:blue"), (jd_hgcdte, "HgCdTe", "tab:red")]:
        ax1.semilogy(jd.temperature_K, jd.j_dark_A_cm2, "o-", color=color,
                     linewidth=2, markersize=5, label=f"{det} (vendor data)")
    for (det, _rc, tc, _rb, tb), jd, color in zip(
        trade_rows, [jd_insb, jd_hgcdte], ["tab:blue", "tab:red"]
    ):
        ax1.axvline(tc, color=color, linestyle=":", alpha=0.7)
        ax1.annotate(f"{det} crossover\n{tc:.1f} K", xy=(tc, jd.j_dark_at(tc)),
                     textcoords="offset points", xytext=(8, -18), fontsize=8, color=color)
        ax1.axvline(tb, color=color, linestyle="--", alpha=0.5)
        ax1.annotate(f"{det} BLIP\n{tb:.1f} K", xy=(tb, jd.j_dark_at(tb)),
                     textcoords="offset points", xytext=(8, 8), fontsize=8, color=color)
    ax1.set_xlabel("Detector Temperature [K]", fontsize=12)
    ax1.set_ylabel("Dark Current Density [A/cm²]", fontsize=12)
    ax1.set_title("Vendor J_dark(T) with Crossover and BLIP Temperatures", fontsize=13)
    ax1.legend(fontsize=10)
    ax1.grid(True, which="both", alpha=0.3)
    fig1.tight_layout()
    fig1.savefig(OUTPUTS / "fig1_jdark_vs_temperature.png", dpi=150)
    print(f"\n  Saved {OUTPUTS / 'fig1_jdark_vs_temperature.png'}")

    # Fig 2: QE curves + band
    fig2, ax2 = plt.subplots(figsize=(9, 6))
    ax2.plot(qe_insb.wavelength_um, qe_insb.qe, "o-", color="tab:blue",
             linewidth=2, markersize=4, label="InSb (vendor: nm / %)")
    ax2.plot(qe_hgcdte.wavelength_um, qe_hgcdte.qe, "s-", color="tab:red",
             linewidth=2, markersize=4, label="HgCdTe (vendor: µm / fraction)")
    ax2.axvspan(band_min_um, band_max_um, alpha=0.12, color="green",
                label=f"Cold filter {band_min_um:.1f}–{band_max_um:.1f} µm")
    ax2.set_xlabel("Wavelength [µm]", fontsize=12)
    ax2.set_ylabel("Quantum Efficiency [fraction]", fontsize=12)
    ax2.set_title("Vendor QE Curves in Canonical Units (load_qe_csv)", fontsize=13)
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(0, 1.0)
    fig2.tight_layout()
    fig2.savefig(OUTPUTS / "fig2_qe_curves.png", dpi=150)
    print(f"  Saved {OUTPUTS / 'fig2_qe_curves.png'}")

    # Fig 3: side-by-side noise budget bars
    fig3, ax3 = plt.subplots(figsize=(10, 6))
    plot_terms = [t for t in all_terms
                  if noise["InSb"].get(t, 0) > 0.01 or noise["HgCdTe"].get(t, 0) > 0.01]
    x = np.arange(len(plot_terms))
    w = 0.38
    ax3.bar(x - w / 2, [noise["InSb"].get(t, 0.0) for t in plot_terms], w,
            color="tab:blue", edgecolor="black", linewidth=0.5, label="InSb")
    ax3.bar(x + w / 2, [noise["HgCdTe"].get(t, 0.0) for t in plot_terms], w,
            color="tab:red", edgecolor="black", linewidth=0.5, label="HgCdTe")
    ax3.set_xticks(x)
    ax3.set_xticklabels(plot_terms, rotation=30, ha="right", fontsize=9)
    ax3.set_ylabel("Noise [e⁻ RMS]", fontsize=12)
    ax3.set_yscale("log")
    ax3.set_title(f"Noise Budget at {t_op_K:.0f} K, t_int = {t_int_s * 1e3:.0f} ms "
                  f"(log scale)", fontsize=13)
    ax3.legend(fontsize=10)
    ax3.grid(True, axis="y", which="both", alpha=0.3)
    fig3.tight_layout()
    fig3.savefig(OUTPUTS / "fig3_noise_budget.png", dpi=150)
    print(f"  Saved {OUTPUTS / 'fig3_noise_budget.png'}")

    # Fig 4: dark shot noise vs T against read-noise floors
    fig4, ax4 = plt.subplots(figsize=(9, 6))
    t_grid = np.linspace(60.0, 110.0, 200)
    for jd, det, color in [(jd_insb, "InSb", "tab:blue"), (jd_hgcdte, "HgCdTe", "tab:red")]:
        dark_shot = [math.sqrt(jd.dark_rate_e_per_s(t, pixel_pitch_m=pixel_pitch_m) * t_int_s)
                     for t in t_grid]
        ax4.semilogy(t_grid, dark_shot, "-", color=color, linewidth=2,
                     label=f"{det} dark shot √(rate·t)")
        ax4.axhline(read_noise[det], color=color, linestyle=":",
                    label=f"{det} read noise {read_noise[det]:.0f} e⁻")
    ax4.set_xlabel("Detector Temperature [K]", fontsize=12)
    ax4.set_ylabel("Noise [e⁻ RMS]", fontsize=12)
    ax4.set_title(f"Dark Shot Noise vs Temperature (t_int = {t_int_s * 1e3:.0f} ms)",
                  fontsize=13)
    ax4.legend(fontsize=9)
    ax4.grid(True, which="both", alpha=0.3)
    fig4.tight_layout()
    fig4.savefig(OUTPUTS / "fig4_dark_shot_vs_temperature.png", dpi=150)
    print(f"  Saved {OUTPUTS / 'fig4_dark_shot_vs_temperature.png'}")

    # ---------------------------------------------------------------------------
    # Step 8: Output workbook
    # ---------------------------------------------------------------------------

    wb_out = openpyxl.Workbook()
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    ws1 = wb_out.active
    ws1.title = "Noise Budget"
    ws1["A1"] = f"Scenario 2.1 — Noise Budget at {t_op_K:.0f} K (spectral QE)"
    ws1["A1"].font = Font(bold=True, size=14)
    for col, htext in enumerate(["Noise Term", "InSb [e- RMS]", "HgCdTe [e- RMS]"], 1):
        cell = ws1.cell(row=3, column=col, value=htext)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    r = 4
    for term in all_terms:
        ws1.cell(row=r, column=1, value=term).border = border
        ws1.cell(row=r, column=2, value=round(noise["InSb"].get(term, 0.0), 3)).border = border
        ws1.cell(row=r, column=3, value=round(noise["HgCdTe"].get(term, 0.0), 3)).border = border
        r += 1
    ws1.cell(row=r, column=1, value="TOTAL (RSS)").border = border
    ws1.cell(row=r, column=2, value=round(total_noise["InSb"], 2)).border = border
    ws1.cell(row=r, column=3, value=round(total_noise["HgCdTe"], 2)).border = border
    for col_letter, width in [("A", 24), ("B", 16), ("C", 16)]:
        ws1.column_dimensions[col_letter].width = width

    ws2 = wb_out.create_sheet("Cooler Trade")
    ws2["A1"] = "Crossover / BLIP temperatures and NEI"
    ws2["A1"].font = Font(bold=True, size=14)
    headers2 = ["Detector", "Crossover T [K]", "BLIP T [K]",
                "NEI [ph/s/cm2]", "NEI [W/cm2] approx", "SNR [-]"]
    for col, htext in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=htext)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
    for i, ((det, _rc, tc, _rb, tb), (_d2, _sig, nei_ph, nei_w)) in enumerate(
        zip(trade_rows, nei_rows), 4
    ):
        ws2.cell(row=i, column=1, value=det).border = border
        ws2.cell(row=i, column=2, value=round(tc, 1)).border = border
        ws2.cell(row=i, column=3, value=round(tb, 1)).border = border
        ws2.cell(row=i, column=4, value=float(f"{nei_ph:.3e}")).border = border
        ws2.cell(row=i, column=5, value=float(f"{nei_w:.3e}")).border = border
        ws2.cell(row=i, column=6, value=round(results[det].metrics["snr"], 1)).border = border
    for col_letter in "ABCDEF":
        ws2.column_dimensions[col_letter].width = 18

    wb_out.save(OUTPUT_FILE)
    print(f"\n  Output workbook: {OUTPUT_FILE}")

    # ---------------------------------------------------------------------------
    # Step 9: Summary
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 95}")
    print(f"  SUMMARY")
    print(f"{'=' * 95}")
    print(f"\n  Bench: {bb_temp_K:.0f} K blackbody, {band_min_um:.1f}–{band_max_um:.1f} µm, "
          f"f/{f_number:.1f}, {pixel_pitch_um:.0f} µm pixels, t_int = {t_int_s * 1e3:.0f} ms, "
          f"T_FPA = {t_op_K:.0f} K")
    print(f"\n  {'Metric':<38s} {'InSb':>14s}  {'HgCdTe':>14s}")
    print(f"  {'-' * 38} {'-' * 14}  {'-' * 14}")
    print(f"  {'Band-averaged QE [--]':<38s} {qe_band['InSb']:>14.3f}  {qe_band['HgCdTe']:>14.3f}")
    print(f"  {'Dark rate at 77 K [e⁻/s]':<38s} {dark_rate['InSb']:>14,.0f}  "
          f"{dark_rate['HgCdTe']:>14,.0f}")
    print(f"  {'Dark e⁻ in t_int [e⁻]':<38s} {dark_rate['InSb'] * t_int_s:>14.1f}  "
          f"{dark_rate['HgCdTe'] * t_int_s:>14.1f}")
    print(f"  {'Total noise [e⁻ RMS]':<38s} {total_noise['InSb']:>14.1f}  "
          f"{total_noise['HgCdTe']:>14.1f}")
    print(f"  {'SNR [--]':<38s} {results['InSb'].metrics['snr']:>14.1f}  "
          f"{results['HgCdTe'].metrics['snr']:>14.1f}")
    print(f"  {'Crossover T [K]':<38s} {t_cross_insb:>14.1f}  {t_cross_mct:>14.1f}")
    print(f"  {'BLIP T [K]':<38s} {t_blip_insb:>14.1f}  {t_blip_mct:>14.1f}")

    print(f"\n  Key findings:")
    print(f"    1. Both FPAs are photon-noise-dominated at 77 K on this bench —")
    print(f"       signal shot noise dwarfs everything else, so SNR differences")
    print(f"       track the QE ratio, not the dark-current ratio.")
    print(f"    2. The trade separates at WARMER set points: HgCdTe holds its")
    print(f"       read-noise crossover to {t_cross_mct:.1f} K vs {t_cross_insb:.1f} K for InSb, and")
    print(f"       stays BLIP to {t_blip_mct:.1f} K vs {t_blip_insb:.1f} K — roughly "
          f"{t_cross_mct - t_cross_insb:.0f} K of cooler margin.")
    print(f"    3. InSb's higher, flatter in-band QE ({qe_band['InSb']:.2f} vs "
          f"{qe_band['HgCdTe']:.2f}) buys ~{(qe_band['InSb'] / qe_band['HgCdTe'] - 1) * 100:.0f}% more signal —")
    print(f"       if the mission can afford 77 K, InSb wins SNR; if the cooler")
    print(f"       is the constraint, HgCdTe wins operability margin.")
    print(f"    4. kTC (37 e⁻ if uncorrelated) is fully suppressed by CDS; the")
    print(f"       5 e⁻/s ROIC glow contributes negligibly at 1 ms.")


if __name__ == "__main__":
    main()
