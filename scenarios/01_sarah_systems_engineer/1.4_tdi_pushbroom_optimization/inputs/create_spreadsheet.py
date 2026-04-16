#!/usr/bin/env python3
"""Create the input spreadsheet for Scenario 1.4: TDI Pushbroom Optimization.

Sarah is designing a VNIR panchromatic pushbroom imager for a 500 km SSO.
Ground velocity constrains the line period (very short: ~0.2 ms).
She needs TDI to build sufficient SNR.

Note: the original scenario concept specified MWIR, but MWIR thermal scenes
generate so many photons per pixel per line that well saturation occurs at
N_tdi=1. TDI is counterproductive for MWIR Earth observation — it's a VNIR
technology where reflected solar signal is much weaker.

Run:  python create_spreadsheet.py
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

wb = openpyxl.Workbook()

# ── Sheet 1: System Parameters ──────────────────────────────────────
ws = wb.active
ws.title = "System Parameters"
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 55

header_font = Font(bold=True, size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
note_font = Font(italic=True, color="666666")

for col in ["A", "B", "C", "D"]:
    ws[f"{col}1"].font = header_font_white
    ws[f"{col}1"].fill = header_fill
    ws[f"{col}1"].alignment = Alignment(horizontal="center")

ws["A1"] = "Parameter"
ws["B1"] = "Value"
ws["C1"] = "Unit"
ws["D1"] = "Notes"

rows = [
    # Optics
    ("Aperture diameter",       25,      "cm",    "25 cm primary mirror"),
    ("Focal length",            250,     "cm",    "f/10"),
    ("f-number",                10.0,    "--",    "Derived: f/D"),
    ("Optical transmission",    70,      "%",     "End-to-end"),
    ("Optics temperature",      20,      "C",     "Ambient (not cooled)"),
    ("WFE RMS",                 0.05,    "waves", "Diffraction-limited design"),
    ("Central obscuration",     30,      "%",     "Cassegrain"),
    # Detector
    ("Pixel pitch",             7.0,     "um",    "Si CCD"),
    ("Quantum efficiency",      80,      "%",     "Broadband VNIR average"),
    ("Dark current",            5.0,     "e-/s",  "Room temp Si CCD"),
    ("Read noise",              15.0,    "e- RMS", "Per read, analog TDI: single readout"),
    ("Full well capacity",      60000,   "e-",    "Small pixel CCD"),
    ("Gain",                    1.5,     "e-/DN", "14-bit ADC"),
    ("ADC bits",                14,      "--",    "16,383 DN max"),
    # Spectral
    ("Filter min",              500,     "nm",    "VNIR panchromatic"),
    ("Filter max",              850,     "nm",    "VNIR panchromatic"),
    # Scene
    ("Target reflectance",      0.15,    "--",    "Gray asphalt / urban"),
    ("Background reflectance",  0.10,    "--",    "Vegetation background"),
    ("Solar zenith angle",      30,      "deg",   "Mid-morning"),
    # Orbit & geometry
    ("Orbit altitude",          500,     "km",    "LEO sun-synchronous"),
    ("Orbital velocity",        7500,    "m/s",   "Circular orbit"),
    # Atmosphere
    ("Atmosphere model",        "simple", "--",   "Simple Beer-Lambert"),
    ("Visibility",              23,      "km",    "Clear day"),
    ("PWV",                     20,      "mm",    "Standard mid-latitude"),
    # TDI
    ("TDI mode",                "analog", "--",   "Charge accumulation, single readout"),
    ("TDI misalignment",        0.1,     "pixels", "Cross-track registration error per stage"),
]

for i, (param, value, unit, note) in enumerate(rows, start=2):
    ws[f"A{i}"] = param
    ws[f"B{i}"] = value
    ws[f"C{i}"] = unit
    ws[f"D{i}"] = note
    ws[f"A{i}"].font = Font(size=10)
    ws[f"D{i}"].font = note_font

# ── Sheet 2: TDI Sweep ──────────────────────────────────────────────
ws2 = wb.create_sheet("TDI Sweep")
ws2.column_dimensions["A"].width = 15
ws2.column_dimensions["B"].width = 18

for col in ["A", "B"]:
    ws2[f"{col}1"].font = header_font_white
    ws2[f"{col}1"].fill = header_fill
    ws2[f"{col}1"].alignment = Alignment(horizontal="center")

ws2["A1"] = "N_tdi"
ws2["B1"] = "Notes"

tdi_values = [1, 2, 4, 8, 16, 32, 64, 96, 128]
notes = [
    "No TDI (staring)",
    "Minimal TDI",
    "Low TDI",
    "Moderate TDI",
    "Common pushbroom",
    "Standard high-res EO",
    "High TDI (Pleiades-class)",
    "Very high TDI",
    "Maximum TDI",
]

for i, (n, note) in enumerate(zip(tdi_values, notes), start=2):
    ws2[f"A{i}"] = n
    ws2[f"B{i}"] = note

# ── Sheet 3: Derived Parameters ─────────────────────────────────────
ws3 = wb.create_sheet("Derived Parameters")
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 55

for col in ["A", "B", "C", "D"]:
    ws3[f"{col}1"].font = header_font_white
    ws3[f"{col}1"].fill = header_fill
    ws3[f"{col}1"].alignment = Alignment(horizontal="center")

ws3["A1"] = "Parameter"
ws3["B1"] = "Value"
ws3["C1"] = "Unit"
ws3["D1"] = "Derivation"

R_earth = 6371000  # m
h = 500000  # m
v_orb = 7500  # m/s
pitch = 7e-6  # m
f = 2.50  # m
D = 0.25  # m
band_center = 0.675  # um

v_ground = v_orb * R_earth / (R_earth + h)
gsd = pitch * h / f
line_period = gsd / v_ground
ifov = pitch / f
Q = band_center * (f / D) / (pitch * 1e6)

derived = [
    ("Ground velocity",   round(v_ground, 1),         "m/s",  f"v_orb × R_E/(R_E+h) = {v_orb} × {R_earth}/{R_earth+h}"),
    ("GSD",               round(gsd, 2),              "m",    f"pitch × alt / f = {pitch*1e6:.0f}e-6 × {h/1e3:.0f}e3 / {f}"),
    ("Line period",       round(line_period * 1e3, 4), "ms",  f"GSD / v_ground = {gsd:.2f} / {v_ground:.1f}"),
    ("IFOV",              round(ifov * 1e6, 1),       "urad", f"pitch / f = {pitch*1e6:.0f}e-6 / {f}"),
    ("Q (sampling)",      round(Q, 3),                "--",   f"λ_center × f/# / pitch = {band_center} × {f/D:.0f} / {pitch*1e6:.0f}"),
    ("Smear per line",    1.0,                        "pixels", "By design: line rate matches ground velocity"),
    ("Smear MTF@Nyquist", round(2 / 3.14159, 4),     "--",   "|sinc(π/2)| = 2/π ≈ 0.6366 (constant for all N_tdi)"),
]

for i, (param, value, unit, deriv) in enumerate(derived, start=2):
    ws3[f"A{i}"] = param
    ws3[f"B{i}"] = value
    ws3[f"C{i}"] = unit
    ws3[f"D{i}"] = deriv

out = "sarah_tdi_pushbroom_data.xlsx"
wb.save(out)
print(f"Wrote {out}")
