"""Generate the vendor-format input workbook for scenario 10.4 (LEO -> GEO SDA).

The workbook is deliberately **not** in RADIANT-internal units: it is what the
programme office actually circulates — a telescope vendor datasheet in mm and
percent, an FPA datasheet in um / ke- / ms, a mission sheet in km and degrees,
and a target-signature sheet in m^2 and degrees Celsius.  The runner
(``scripts/run_leo_to_geo_exo.py``) performs every vendor -> canonical
conversion exactly once, each with an explicit comment (scenario_testing.md
Rule 1).

Run from anywhere::

    python scenarios/10_direction_general/10.4_leo_to_geo_exo/inputs/create_spreadsheet.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

OUT = Path(__file__).resolve().parent / "sda_leo_to_geo_sensor_data.xlsx"

TITLE_FONT = Font(bold=True, size=13)
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
NOTE_FONT = Font(italic=True, size=9)


def _sheet(wb: openpyxl.Workbook, name: str, title: str, subtitle: str) -> Worksheet:
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_FONT
    for col, head in enumerate(("Parameter", "Value", "Unit", "Notes"), start=1):
        cell = ws.cell(row=4, column=col, value=head)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    return ws


def _rows(ws: Worksheet, rows: list[tuple[str, object, str, str]]) -> None:
    for r, (name, value, unit, note) in enumerate(rows, start=5):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)


def _autosize(ws: Worksheet) -> None:
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(width + 3, 72)


def build() -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # -- 1. Telescope vendor datasheet (mm, %, deg C) ------------------------
    ws = _sheet(
        wb,
        "Telescope Datasheet",
        "SDA-500 MWIR space-surveillance telescope",
        "Vendor datasheet extract - dimensions in mm, transmission in %, temperatures in deg C",
    )
    _rows(
        ws,
        [
            ("Entrance pupil diameter", 350.0, "mm", "Cassegrain primary, clear aperture"),
            ("Effective focal length", 2100.0, "mm", "f/6.0"),
            ("f-number", 6.0, "--", "EFL / EPD"),
            ("Central obscuration", 25.0, "%", "Secondary + baffle, linear diameter ratio"),
            ("Optical transmission (in band)", 60.0, "%", "6 surfaces + cold filter, 3.5-5.0 um"),
            ("WFE RMS", 0.05, "waves", "At 4.25 um, on-axis, end of life"),
            ("Optical bench temperature", -93.15, "deg C", "Radiatively cooled aft optics"),
        ],
    )
    _autosize(ws)

    # -- 2. FPA datasheet (um, %, e-, ke-, ms) -------------------------------
    ws = _sheet(
        wb,
        "FPA Datasheet",
        "MWIR HgCdTe FPA, 1024 x 1024",
        "FPA vendor datasheet - pitch in um, QE in %, well in ke-, integration time in ms",
    )
    _rows(
        ws,
        [
            ("Pixel pitch", 18.0, "um", "Square pixel"),
            ("Array format (cross x along)", "1024 x 1024", "--", "Staring format"),
            ("Quantum efficiency", 75.0, "%", "Band-averaged 3.5-5.0 um"),
            ("Dark current", 1000.0, "e-/s", "At 80 K operating temperature"),
            ("Operating temperature", 80.0, "K", "Stirling cooler, cold shield f/6 matched"),
            ("Read noise", 25.0, "e- rms", "CDS, 1 MHz pixel rate"),
            ("Full well capacity", 100.0, "ke-", "High-flux integration capacitor"),
            ("ADC resolution", 14, "bits", "On-FPA ADC"),
            ("System gain", 6.1, "e-/DN", "Matched: full well / 2^bits"),
            ("Nominal integration time", 500.0, "ms", "Rate-tracked stare"),
            ("Spectral band, min", 3500.0, "nm", "Cold filter cut-on"),
            ("Spectral band, max", 5000.0, "nm", "Cold filter cut-off"),
        ],
    )
    _autosize(ws)

    # -- 3. Mission geometry (km, deg) --------------------------------------
    ws = _sheet(
        wb,
        "Mission Geometry",
        "LEO -> GEO space-domain-awareness pass",
        "Mission planning sheet - altitudes in km, angles in deg",
    )
    _rows(
        ws,
        [
            ("Sensor orbit altitude", 500.0, "km", "Circular LEO, host platform"),
            ("Target orbit altitude", 35786.0, "km", "Geostationary belt"),
            (
                "Sensor-side path zenith (nominal)",
                0.0,
                "deg",
                "Zeta_low: at the LOWER endpoint = the LEO sensor. 0 deg = target at zenith",
            ),
            (
                "Illumination state",
                "night",
                "--",
                "Target in Earth eclipse: thermal-only signature, no solar reflection",
            ),
            (
                "Declared scene class",
                "space_to_space",
                "--",
                "Optional ADR-0011 assertion, validated against the derived class",
            ),
            ("Rate-track residual", 1.0, "%", "Uncompensated fraction of the open-loop LOS rate"),
            ("Detection SNR threshold", 5.0, "--", "Programme-standard single-frame detection"),
        ],
    )
    _autosize(ws)

    # -- 4. Target signature model (m^2, deg C) ------------------------------
    ws = _sheet(
        wb,
        "Target Signature",
        "Reference GEO communications-satellite bus",
        "Signature working group estimate - area in m^2, temperature in deg C",
    )
    _rows(
        ws,
        [
            (
                "Projected area toward sensor",
                20.0,
                "m^2",
                "3.2 x 2.4 m bus body + edge-on wing structure, nadir aspect",
            ),
            (
                "Mean surface temperature",
                6.85,
                "deg C",
                "= 280 K. Eclipse-cooled MLI/radiator mix, mid-eclipse",
            ),
            (
                "Broadband emissivity",
                0.85,
                "--",
                "Area-weighted: MLI outer Kapton ~0.8, OSR radiator ~0.85",
            ),
        ],
    )
    _autosize(ws)

    # -- 5. Integration-time sweep (ms) -------------------------------------
    ws = wb.create_sheet("Integration Sweep")
    ws["A1"] = "Integration-time trade points"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Smear / SNR trade for the rate-tracked and open-loop (untracked) cases"
    ws["A2"].font = NOTE_FONT
    cell = ws.cell(row=4, column=1, value="Integration time [ms]")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    for r, t_ms in enumerate([5.0, 10.0, 25.0, 50.0, 100.0, 250.0, 500.0, 1000.0], start=5):
        ws.cell(row=r, column=1, value=t_ms)
    _autosize(ws)

    # -- 6. Sensor-side zenith sweep (deg) ----------------------------------
    ws = wb.create_sheet("Zenith Sweep")
    ws["A1"] = "Sensor-side path-zenith trade points"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Zeta_low measured at the LEO sensor (the path's lower endpoint), in deg"
    ws["A2"].font = NOTE_FONT
    cell = ws.cell(row=4, column=1, value="Sensor-side path zenith [deg]")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    for r, z_deg in enumerate([0.0, 5.0, 10.0, 20.0, 30.0, 45.0, 60.0], start=5):
        ws.cell(row=r, column=1, value=z_deg)
    _autosize(ws)

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote vendor input workbook: {path}")
