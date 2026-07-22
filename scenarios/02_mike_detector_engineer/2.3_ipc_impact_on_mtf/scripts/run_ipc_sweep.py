"""Scenario 2.3: IPC Characterization — How Much MTF Loss Can I Tolerate?

Mike (detector engineer) wants to sweep IPC coupling from 0% to 5% and see
how it affects system MTF at Nyquist, ensquared energy, and SNR. He has:
  - A detector vendor datasheet (Excel) with specs in vendor units
  - Lab measurements of IPC on 5 sample detectors
  - System requirements: MTF at Nyquist >= 0.15, SNR >= 100, EE_1x1 >= 0.60

This script:
  1. Reads Mike's spreadsheet (vendor units: cm, %, µm, km, ms)
  2. Converts to RADIANT canonical units (m, fractions, s)
  3. Runs a RADIANT evaluation at each IPC coupling value (0–5%)
     - RADIANT natively wires IPC into the EffectivePSF via FFT convolution
     - All spatial metrics (MTF, EE, RER, FWHM) include IPC effects
  4. Extracts MTF, EE, RER, FWHM, SNR, GSD, Q, NEDT, NIIRS from each run
  5. Finds the maximum tolerable IPC coupling against requirements
  6. Compares predictions against Mike's lab measurements
  7. Validates native results against analytic IPC MTF formula
  8. Writes results to an output spreadsheet

Usage:
    python run_ipc_sweep.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ---------------------------------------------------------------------------
# Step 1: Read Mike's spreadsheet
# ---------------------------------------------------------------------------
# Mike has three sheets:
#   "Detector Specs"      — vendor datasheet values (vendor units)
#   "IPC Measurements"    — his lab bench data (5 samples)
#   "System Requirements" — optical system + performance thresholds

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "mike_detector_ipc_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "ipc_sweep_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

# --- Parse Detector Specs sheet ---
# Layout: rows with [Parameter, Value, Unit, Notes], with section headers
# Section headers have a blue fill and should be skipped.

ws_det = wb["Detector Specs"]
det_specs: dict[str, object] = {}
for row in ws_det.iter_rows(min_row=6, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        # Skip section headers (blue-filled rows)
        fill = row[0].fill
        if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
            continue
        det_specs[name] = value

# --- Parse IPC Measurements sheet ---
# Layout: rows with [Sample ID, IPC Coupling [%], Measured MTF @ Nyq,
#                     Measured EE 1x1, Measured EE 3x3, Notes]

ws_ipc = wb["IPC Measurements"]
lab_data: list[dict] = []
for row in ws_ipc.iter_rows(min_row=6, max_col=6, values_only=True):
    if row[0] and row[1] is not None:
        lab_data.append({
            "sample_id": row[0],
            "ipc_pct": float(row[1]),         # IPC coupling [%]
            "mtf_nyq": float(row[2]),         # Measured MTF at Nyquist [—]
            "ee_1x1": float(row[3]),          # Measured EE in 1×1 pixel [—]
            "ee_3x3": float(row[4]),          # Measured EE in 3×3 pixels [—]
            "notes": row[5] or "",
        })

# --- Parse System Requirements sheet ---
# Same layout as Detector Specs: [Parameter, Value, Unit, Notes]

ws_req = wb["System Requirements"]
sys_reqs: dict[str, object] = {}
for row in ws_req.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        fill = row[0].fill
        if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
            continue
        sys_reqs[name] = value

print("=== Detector Specs (from vendor datasheet) ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}")
print(f"  {'Pixel pitch':<35s} {det_specs['Pixel pitch']:>14}  µm")
print(f"  {'Average QE (in-band)':<35s} {det_specs['Average QE (in-band)']:>14}  %")
print(f"  {'Dark current (mean)':<35s} {det_specs['Dark current (mean)']:>14}  e⁻/s")
print(f"  {'Operating temperature':<35s} {det_specs['Operating temperature']:>14}  K")
print(f"  {'Full well capacity':<35s} {det_specs['Full well capacity']:>14.0f}  e⁻")
print(f"  {'IPC coupling (typical)':<35s} {det_specs['IPC coupling (typical)']:>14}  %")
print(f"  {'IPC coupling (max)':<35s} {det_specs['IPC coupling (max)']:>14}  %")
print(f"  {'Read noise (CDS)':<35s} {det_specs['Read noise (CDS)']:>14}  e⁻ RMS")
print(f"  {'ADC resolution':<35s} {det_specs['ADC resolution']:>14}  bits")
print(f"  {'System gain':<35s} {det_specs['System gain']:>14}  e⁻/DN")

print(f"\n=== Lab IPC Measurements ===")
print(f"  {'Sample':<14s} {'IPC [%]':>8s}  {'MTF@Nyq [—]':>12s}  {'EE 1×1 [—]':>11s}  {'EE 3×3 [—]':>11s}")
print(f"  {'-' * 14} {'-' * 8}  {'-' * 12}  {'-' * 11}  {'-' * 11}")
for d in lab_data:
    print(f"  {d['sample_id']:<14s} {d['ipc_pct']:>8.1f}  {d['mtf_nyq']:>12.3f}  {d['ee_1x1']:>11.3f}  {d['ee_3x3']:>11.3f}")

print(f"\n=== System Requirements ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}")
print(f"  {'Aperture diameter':<35s} {sys_reqs['Aperture diameter']:>14}  cm")
print(f"  {'Focal length':<35s} {sys_reqs['Focal length']:>14}  cm")
print(f"  {'f-number':<35s} {sys_reqs['f-number']:>14}  —")
print(f"  {'Optical transmission':<35s} {sys_reqs['Optical transmission']:>14}  %")
print(f"  {'Target temperature':<35s} {sys_reqs['Target temperature']:>14}  K")
print(f"  {'Background temperature':<35s} {sys_reqs['Background temperature']:>14}  K")
print(f"  {'Integration time':<35s} {sys_reqs['Integration time']:>14}  ms")
print(f"  {'Orbit altitude':<35s} {sys_reqs['Orbit altitude']:>14}  km")
print(f"  {'System MTF at Nyquist':<35s} {str(sys_reqs['System MTF at Nyquist']):>14s}  —")
print(f"  {'SNR':<35s} {str(sys_reqs['SNR']):>14s}  —")
print(f"  {'EE 1×1 (ensquared energy)':<35s} {str(sys_reqs['EE 1×1 (ensquared energy)']):>14s}  —")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------
# RADIANT expects: meters, fractions (not %), seconds, µm for pixel pitch
# and wavelengths.

# Optics — from System Requirements sheet
aperture_m = float(sys_reqs["Aperture diameter"]) / 100.0       # cm → m
focal_length_m = float(sys_reqs["Focal length"]) / 100.0        # cm → m
f_number = float(sys_reqs["f-number"])                           # dimensionless
transmission = float(sys_reqs["Optical transmission"]) / 100.0  # % → fraction
optics_temp_K = float(sys_reqs["Optics temperature"])            # already K

# Detector — from Detector Specs sheet
pixel_pitch_um = float(det_specs["Pixel pitch"])                 # already µm (RADIANT input unit)
pixel_pitch_m = pixel_pitch_um * 1e-6                            # µm → m (for hand calculations)
qe = float(det_specs["Average QE (in-band)"]) / 100.0           # % → fraction
dark_rate = float(det_specs["Dark current (mean)"])              # already e⁻/s
operating_temp_K = float(det_specs["Operating temperature"])     # already K
fwc = float(det_specs["Full well capacity"])                     # already e⁻
read_noise = float(det_specs["Read noise (CDS)"])                # already e⁻ RMS
adc_bits = int(det_specs["ADC resolution"])                      # already bits
gain = float(det_specs["System gain"])                           # already e⁻/DN
ipc_nominal = float(det_specs["IPC coupling (typical)"]) / 100.0  # % → fraction

# Scene — from System Requirements sheet
target_temp = float(sys_reqs["Target temperature"])              # already K
target_emiss = float(sys_reqs["Target emissivity"])              # dimensionless
bg_temp = float(sys_reqs["Background temperature"])              # already K
bg_emiss = float(sys_reqs["Background emissivity"])              # dimensionless

# Spectral — from System Requirements sheet
# The band is given as a string "3.5 – 5.0"; parse it.
band_str = str(sys_reqs["Spectral band"])
band_parts = band_str.replace("–", "-").replace("—", "-").split("-")
band_min_um = float(band_parts[0].strip())                      # already µm
band_max_um = float(band_parts[1].strip())                      # already µm
t_int_s = float(sys_reqs["Integration time"]) / 1000.0          # ms → s

# Geometry — from System Requirements sheet
altitude_m = float(sys_reqs["Orbit altitude"]) * 1000.0          # km → m

# WFE — from System Requirements sheet
wfe_waves = float(sys_reqs["WFE (RMS)"])                         # already in waves

print("\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'f-number':<35s} {f_number:>14.1f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'Quantum efficiency':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<35s} {dark_rate:>14.1f}  {'e⁻/s':<15s}  no conversion")
print(f"  {'Detector temperature':<35s} {operating_temp_K:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Read noise':<35s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'IPC coupling (nominal)':<35s} {ipc_nominal:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Target temperature':<35s} {target_temp:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Background temperature':<35s} {bg_temp:>14.1f}  {'K':<15s}  * contrast SNR only")
print(f"  {'Band':<35s} {band_min_um:>6.1f}–{band_max_um:<6.1f}  {'µm':<15s}  no conversion")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms ÷ 1000")
print(f"  {'Orbit altitude':<35s} {altitude_m:>14.0f}  {'m':<15s}  km × 1000")

# ---------------------------------------------------------------------------
# Step 3: Run RADIANT baseline evaluation (IPC = 0)
# ---------------------------------------------------------------------------
# We run RADIANT without IPC first to get the baseline system MTF, SNR,
# and EE. This gives us the optical + detector contribution without IPC.
#
# RADIANT natively wires IPC into the signal chain: DetectorStage generates
# the 3×3 IPC kernel and stores it in stage_outputs, then PerformanceStage
# applies it to the EffectivePSF via FFT convolution. All spatial metrics
# (MTF, EE, RER, FWHM) computed from the IPC-convolved PSF are consistent.

from radiant.api import Sensor


def make_ipc_config(ipc_fraction: float) -> dict:
    """Build a RADIANT config dict with a given IPC coupling fraction."""
    return {
        "source": {
            "target": {
                "temperature": target_temp,
                "emissivity": target_emiss,
            },
            "background": {
                "temperature": bg_temp,
                "emissivity": bg_emiss,
            },
        },
        "atmosphere": {
            "model": "simple",    # LEO through atmosphere
        },
        "geometry": {
            "sensor_altitude_m": altitude_m,
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": operating_temp_K,
            "ipc_coupling": ipc_fraction,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
        "performance": {"niirs": {"allow_extrapolated": True}},
    }


print("\n=== Running RADIANT baseline evaluation (IPC = 0) ===")
sensor = Sensor.from_dict(make_ipc_config(0.0))
result = sensor.evaluate()

# Extract baseline metrics
baseline_mtf_nyq = result.metrics["mtf_at_nyquist"]      # System MTF at Nyquist [—]
baseline_snr = result.metrics["snr"]                       # Signal-to-noise ratio [—]
baseline_ee_1x1 = result.metrics["ee_1x1"]                 # Ensquared energy, 1×1 pixel [—]
baseline_ee_3x3 = result.metrics["ee_3x3"]                 # Ensquared energy, 3×3 pixels [—]
baseline_contrast_snr = result.metrics.get("contrast_snr") # Contrast SNR [—]
baseline_rer = result.metrics.get("rer")                   # Relative edge response [—]
baseline_fwhm_x = result.metrics.get("fwhm_x_m")          # FWHM cross-track [m]
baseline_gsd = result.metrics.get("gsd_cross_track_m")     # Ground sample distance [m]
baseline_q = result.metrics.get("q_center")                # Sampling parameter Q [—]
baseline_nedt = result.metrics.get("nedt_K")               # NEDT [K]
baseline_niirs = result.metrics.get("niirs")               # NIIRS [—]
baseline_strehl = result.metrics.get("strehl")             # Strehl ratio [—]

# ---------------------------------------------------------------------------
# Step 3b: Regime and physics notes
# ---------------------------------------------------------------------------

regime = result.stage_outputs["optics"]["regime"]
print(f"\n=== Radiometric Regime ===")
print(f"  Regime:  {regime}")
print(f"")
print(f"  Extended regime: the {target_temp:.0f} K target fills the pixel entirely.")
print(f"  Background temperature ({bg_temp:.0f} K) is used only for contrast SNR.")
print(f"")
print(f"  IPC PHYSICS NOTE:")
print(f"    Inter-pixel capacitance is an electrical coupling effect in the detector")
print(f"    readout. When a pixel accumulates charge, a fraction α leaks to each of")
print(f"    its 4 nearest neighbors via parasitic capacitance. This:")
print(f"      - Degrades MTF (blurs the image electrically, not optically)")
print(f"      - Reduces EE_1x1 (energy spreads from center pixel to neighbors)")
print(f"      - Increases EE_3x3 (neighbors capture the leaked energy)")
print(f"      - Does NOT change total signal or SNR (charge is redistributed, not lost)")
print(f"    The IPC MTF at Nyquist frequency is:")
print(f"      MTF_IPC(f_Nyq) = (1 - 4α) + 2α·cos(2π·f_Nyq·p) + 2α·cos(2π·f_Nyq·p)")
print(f"      At Nyquist (f = 1/2p): cos(π) = -1, so MTF_IPC = 1 - 4α")
print(f"")
print(f"  RADIANT wires IPC natively: setting detector.ipc_coupling causes the IPC")
print(f"  3×3 kernel to be convolved with the EffectivePSF via FFT. All spatial")
print(f"  metrics (MTF, EE, RER, FWHM) are computed from this combined PSF,")
print(f"  ensuring consistency (Rule 4: single PSF for all spatial metrics).")

print(f"\n=== Baseline Metrics (without IPC) ===")
print(f"  {'Metric':<30s} {'Value':>12s}  {'Unit'}")
print(f"  {'-' * 30} {'-' * 12}  {'-' * 15}")
print(f"  {'System MTF at Nyquist':<30s} {baseline_mtf_nyq:>12.4f}  — (dimensionless)")
print(f"  {'SNR':<30s} {baseline_snr:>12.2f}  — (dimensionless)")
print(f"  {'Contrast SNR':<30s} {baseline_contrast_snr:>12.2f}  — (dimensionless)")
print(f"  {'EE 1×1':<30s} {baseline_ee_1x1:>12.4f}  — (dimensionless)")
print(f"  {'EE 3×3':<30s} {baseline_ee_3x3:>12.4f}  — (dimensionless)")
print(f"  {'RER':<30s} {baseline_rer:>12.4f}  — (dimensionless)")
print(f"  {'FWHM (cross-track)':<30s} {baseline_fwhm_x * 1e6:>12.1f}  µm")
if baseline_gsd is not None:
    print(f"  {'GSD (cross-track)':<30s} {baseline_gsd:>12.2f}  m")
if baseline_q is not None:
    print(f"  {'Q (sampling parameter)':<30s} {baseline_q:>12.3f}  — (dimensionless)")
if baseline_nedt is not None:
    print(f"  {'NEDT':<30s} {baseline_nedt:>12.4f}  K")
if baseline_niirs is not None:
    print(f"  {'NIIRS':<30s} {baseline_niirs:>12.2f}  — (dimensionless)")
if baseline_strehl is not None:
    print(f"  {'Strehl ratio':<30s} {baseline_strehl:>12.4f}  — (dimensionless)")

# ---------------------------------------------------------------------------
# Step 4: Analytic IPC MTF (for validation cross-check)
# ---------------------------------------------------------------------------
# The analytic IPC 3×3 kernel is:
#     [[0,   α, 0],
#      [α, 1-4α, α],
#      [0,   α, 0]]
#
# Its 1D MTF along x at the Nyquist frequency (f_Nyq = 1/(2p)):
#     MTF_IPC(f_Nyq) = (1-4α) + 2α·cos(π) + 2α·cos(0) = 1 - 4α
#
# We use this analytic formula to validate RADIANT's native FFT-based
# IPC convolution results at representative IPC values.

from radiant.detector.ipc import ipc_mtf_1d

f_nyquist = 1.0 / (2.0 * pixel_pitch_m)  # Nyquist frequency [cycles/m]

print(f"\n=== Analytic IPC MTF (validation reference) ===")
print(f"  Nyquist frequency:  {f_nyquist:.2f} cycles/m  ({f_nyquist * 1e-3:.2f} cycles/mm)")
print(f"  Pixel pitch:        {pixel_pitch_m * 1e6:.1f} µm  ({pixel_pitch_m:.2e} m)")
print(f"")
print(f"  Analytic: MTF_IPC at Nyquist = 1 - 4α")
print(f"    At α = 0.018 (1.8%): MTF_IPC = {1.0 - 4.0 * 0.018:.4f}")
print(f"    At α = 0.050 (5.0%): MTF_IPC = {1.0 - 4.0 * 0.050:.4f}")

# ---------------------------------------------------------------------------
# Step 5: Sweep IPC coupling from 0% to 5% using native RADIANT
# ---------------------------------------------------------------------------
# For each IPC value, RADIANT:
#   - Generates the 3×3 IPC kernel in DetectorStage
#   - Convolves it with the EffectivePSF in PerformanceStage (FFT)
#   - Computes all spatial metrics from the combined PSF
# This is physically exact — no approximations needed.

ipc_sweep_pct = np.linspace(0.0, 5.0, 51)    # IPC coupling [%]
ipc_sweep_frac = ipc_sweep_pct / 100.0         # IPC coupling [fraction]

# Storage for sweep results
sweep_results: list[dict] = []

print(f"\n=== Running IPC Sweep (51 evaluations) ===")
for i, alpha in enumerate(ipc_sweep_frac):
    config = make_ipc_config(alpha)
    sensor = Sensor.from_dict(config)
    result = sensor.evaluate()

    row = {
        "ipc_pct": ipc_sweep_pct[i],
        "ipc_frac": alpha,
        "mtf_nyq": result.metrics["mtf_at_nyquist"],
        "ee_1x1": result.metrics["ee_1x1"],
        "ee_3x3": result.metrics["ee_3x3"],
        "rer": result.metrics.get("rer"),
        "fwhm_x_m": result.metrics.get("fwhm_x_m"),
        "snr": result.metrics["snr"],
        "contrast_snr": result.metrics.get("contrast_snr"),
        "nedt_K": result.metrics.get("nedt_K"),
        "niirs": result.metrics.get("niirs"),
    }
    sweep_results.append(row)

# Extract arrays for convenience
sweep_mtf_system = np.array([r["mtf_nyq"] for r in sweep_results])
sweep_ee_1x1 = np.array([r["ee_1x1"] for r in sweep_results])
sweep_ee_3x3 = np.array([r["ee_3x3"] for r in sweep_results])
sweep_rer = np.array([r["rer"] for r in sweep_results])
sweep_fwhm = np.array([r["fwhm_x_m"] for r in sweep_results])
sweep_snr = np.array([r["snr"] for r in sweep_results])

print(f"\n=== IPC Sweep Results ===")
print(f"  {'IPC [%]':>8s}  {'Sys MTF [—]':>12s}  {'EE 1×1 [—]':>12s}  {'EE 3×3 [—]':>12s}"
      f"  {'RER [—]':>10s}  {'FWHM [µm]':>10s}  {'SNR [—]':>10s}")
print(f"  {'-' * 8}  {'-' * 12}  {'-' * 12}  {'-' * 12}"
      f"  {'-' * 10}  {'-' * 10}  {'-' * 10}")
# Print every 5th point to keep the table readable
for i in range(0, len(ipc_sweep_pct), 5):
    r = sweep_results[i]
    print(f"  {r['ipc_pct']:>8.1f}  {r['mtf_nyq']:>12.4f}  {r['ee_1x1']:>12.4f}"
          f"  {r['ee_3x3']:>12.4f}  {r['rer']:>10.4f}  {r['fwhm_x_m'] * 1e6:>10.1f}"
          f"  {r['snr']:>10.2f}")

# Analytic cross-check at selected IPC values
print(f"\n=== Analytic Cross-Check (RADIANT native vs. baseline × analytic IPC MTF) ===")
print(f"  {'IPC [%]':>8s}  {'RADIANT MTF':>12s}  {'Analytic MTF':>13s}  {'Δ':>8s}")
print(f"  {'-' * 8}  {'-' * 12}  {'-' * 13}  {'-' * 8}")
for i in range(0, len(ipc_sweep_pct), 10):
    alpha = ipc_sweep_frac[i]
    radiant_mtf = sweep_results[i]["mtf_nyq"]
    analytic_mtf = baseline_mtf_nyq * float(ipc_mtf_1d(np.array([f_nyquist]), alpha, pixel_pitch_m, axis="x")[0])
    delta = radiant_mtf - analytic_mtf
    print(f"  {ipc_sweep_pct[i]:>8.1f}  {radiant_mtf:>12.4f}  {analytic_mtf:>13.4f}  {delta:>8.4f}")

# ---------------------------------------------------------------------------
# Step 6: Find maximum tolerable IPC
# ---------------------------------------------------------------------------
# Requirements from Mike's system requirements sheet:
#   - System MTF at Nyquist >= 0.15
#   - EE 1×1 >= 0.60
#   - SNR >= 100 (IPC doesn't affect SNR, so this is always met if baseline meets it)

MTF_REQ = 0.15       # System MTF at Nyquist threshold [—]
EE_1x1_REQ = 0.60    # Ensquared energy threshold [—]
SNR_REQ = 100.0      # SNR threshold [—]

# Find the IPC value where system MTF crosses below the requirement
mtf_limit_ipc = None
for i in range(len(ipc_sweep_pct)):
    if sweep_mtf_system[i] < MTF_REQ:
        # Linear interpolation between this point and the previous one
        if i > 0:
            # Solve: MTF_REQ = mtf[i-1] + (mtf[i] - mtf[i-1]) * (x - ipc[i-1]) / (ipc[i] - ipc[i-1])
            frac = (MTF_REQ - sweep_mtf_system[i - 1]) / (sweep_mtf_system[i] - sweep_mtf_system[i - 1])
            mtf_limit_ipc = ipc_sweep_pct[i - 1] + frac * (ipc_sweep_pct[i] - ipc_sweep_pct[i - 1])
        else:
            mtf_limit_ipc = 0.0  # baseline already below requirement
        break

# Find the IPC value where EE_1x1 crosses below the requirement
ee_limit_ipc = None
for i in range(len(ipc_sweep_pct)):
    if sweep_ee_1x1[i] < EE_1x1_REQ:
        if i > 0:
            frac = (EE_1x1_REQ - sweep_ee_1x1[i - 1]) / (sweep_ee_1x1[i] - sweep_ee_1x1[i - 1])
            ee_limit_ipc = ipc_sweep_pct[i - 1] + frac * (ipc_sweep_pct[i] - ipc_sweep_pct[i - 1])
        else:
            ee_limit_ipc = 0.0
        break

# The binding constraint is whichever limit is hit first (lower IPC)
if mtf_limit_ipc is not None and ee_limit_ipc is not None:
    binding_limit = min(mtf_limit_ipc, ee_limit_ipc)
    binding_constraint = "MTF" if mtf_limit_ipc < ee_limit_ipc else "EE 1×1"
elif mtf_limit_ipc is not None:
    binding_limit = mtf_limit_ipc
    binding_constraint = "MTF"
elif ee_limit_ipc is not None:
    binding_limit = ee_limit_ipc
    binding_constraint = "EE 1×1"
else:
    binding_limit = None
    binding_constraint = "None (all requirements met across full sweep range)"

print(f"\n=== Maximum Tolerable IPC ===")
print(f"  {'Requirement':<30s} {'Threshold':>10s}  {'Max IPC [%]':>12s}  {'Status'}")
print(f"  {'-' * 30} {'-' * 10}  {'-' * 12}  {'-' * 20}")
if mtf_limit_ipc is not None:
    print(f"  {'System MTF at Nyquist ≥ 0.15':<30s} {MTF_REQ:>10.2f}  {mtf_limit_ipc:>12.2f}  {'BINDING' if binding_constraint == 'MTF' else ''}")
else:
    print(f"  {'System MTF at Nyquist ≥ 0.15':<30s} {MTF_REQ:>10.2f}  {'>5.0':>12s}  met across full range")
if ee_limit_ipc is not None:
    print(f"  {'EE 1×1 ≥ 0.60':<30s} {EE_1x1_REQ:>10.2f}  {ee_limit_ipc:>12.2f}  {'BINDING' if binding_constraint == 'EE 1×1' else ''}")
else:
    print(f"  {'EE 1×1 ≥ 0.60':<30s} {EE_1x1_REQ:>10.2f}  {'>5.0':>12s}  met across full range")
snr_status = "met (baseline)" if baseline_snr >= SNR_REQ else f"FAILED at baseline ({baseline_snr:.1f})"
print(f"  {'SNR ≥ 100':<30s} {SNR_REQ:>10.0f}  {'N/A':>12s}  {snr_status}")
print(f"")
if binding_limit is not None:
    print(f"  CONCLUSION: Maximum tolerable IPC = {binding_limit:.2f} %")
    print(f"              Binding constraint: {binding_constraint}")
    print(f"              Vendor typical (1.8%): {'PASS' if 1.8 < binding_limit else 'FAIL'}")
    print(f"              Vendor max (2.5%):     {'PASS' if 2.5 < binding_limit else 'FAIL'}")
else:
    print(f"  CONCLUSION: All requirements met across the full 0–5% IPC range.")

# ---------------------------------------------------------------------------
# Step 7: Compare RADIANT predictions against lab measurements
# ---------------------------------------------------------------------------
# Mike measured MTF at Nyquist and EE for 5 detector samples with known IPC.
# We compare his measurements to RADIANT's native predictions (with IPC
# convolved into the EffectivePSF) to see if the model is consistent.
#
# NOTE: Exact agreement is not expected because:
#   - Lab MTF measurements include all system effects (optical bench, etc.)
#   - Lab EE measurements depend on the actual PSF shape, not just IPC
#   - Our model uses a symmetric 4-neighbor IPC kernel; real devices may
#     have diagonal coupling or asymmetric IPC
# But the trend (MTF decreasing with IPC) should match, and the magnitude
# should be in the right ballpark.

print(f"\n=== Lab Measurement vs. RADIANT Prediction ===")
print(f"  {'Sample':<14s} {'IPC [%]':>8s}  {'Meas MTF [—]':>13s}  {'RADIANT MTF':>12s}  {'Δ MTF [—]':>10s}  {'Meas EE1 [—]':>13s}  {'RADIANT EE1':>12s}")
print(f"  {'-' * 14} {'-' * 8}  {'-' * 13}  {'-' * 12}  {'-' * 10}  {'-' * 13}  {'-' * 12}")

lab_model_results: list[dict] = []
for d in lab_data:
    alpha = d["ipc_pct"] / 100.0
    # Run RADIANT natively at this IPC value
    config = make_ipc_config(alpha)
    sensor = Sensor.from_dict(config)
    result = sensor.evaluate()
    model_mtf = result.metrics["mtf_at_nyquist"]
    model_ee1 = result.metrics["ee_1x1"]
    delta_mtf = d["mtf_nyq"] - model_mtf

    lab_model_results.append({
        "sample_id": d["sample_id"],
        "ipc_pct": d["ipc_pct"],
        "meas_mtf": d["mtf_nyq"],
        "model_mtf": model_mtf,
        "delta_mtf": delta_mtf,
        "meas_ee1": d["ee_1x1"],
        "model_ee1": model_ee1,
    })

    print(f"  {d['sample_id']:<14s} {d['ipc_pct']:>8.1f}  {d['mtf_nyq']:>13.3f}  {model_mtf:>12.3f}  {delta_mtf:>10.3f}"
          f"  {d['ee_1x1']:>13.3f}  {model_ee1:>12.3f}")

print(f"\n  Note: Model = RADIANT with native IPC convolution (FFT-based PSF kernel).")
print(f"  Differences arise because lab measurements include optical bench effects,")
print(f"  diffraction from test setup optics, and detector-specific variations not")
print(f"  captured by the simple 4-neighbor IPC model (e.g., diagonal coupling).")

# ---------------------------------------------------------------------------
# Step 8: Write results to output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

wb_out = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: IPC Sweep ---
ws1 = wb_out.active
ws1.title = "IPC Sweep"
ws1["A1"] = "Scenario 2.3: IPC Coupling Sweep Results"
ws1["A1"].font = Font(bold=True, size=14)

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 14
ws1.column_dimensions["H"].width = 14

sweep_headers = ["IPC [%]", "System MTF@Nyq [—]", "EE 1×1 [—]", "EE 3×3 [—]",
                 "RER [—]", "FWHM [µm]", "SNR [—]", "NEDT [K]"]
for col, h_text in enumerate(sweep_headers, 1):
    cell = ws1.cell(row=3, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

for i, sr in enumerate(sweep_results):
    r = i + 4
    ws1.cell(row=r, column=1, value=round(sr["ipc_pct"], 1)).border = thin_border
    ws1.cell(row=r, column=2, value=round(sr["mtf_nyq"], 4)).border = thin_border
    ws1.cell(row=r, column=3, value=round(sr["ee_1x1"], 4)).border = thin_border
    ws1.cell(row=r, column=4, value=round(sr["ee_3x3"], 4)).border = thin_border
    ws1.cell(row=r, column=5, value=round(sr["rer"], 4) if sr["rer"] else "").border = thin_border
    ws1.cell(row=r, column=6, value=round(sr["fwhm_x_m"] * 1e6, 1) if sr["fwhm_x_m"] else "").border = thin_border
    ws1.cell(row=r, column=7, value=round(sr["snr"], 2)).border = thin_border
    ws1.cell(row=r, column=8, value=round(sr["nedt_K"], 4) if sr["nedt_K"] else "").border = thin_border

    # Color-code system MTF against requirement
    mtf_cell = ws1.cell(row=r, column=2)
    mtf_cell.fill = pass_fill if sr["mtf_nyq"] >= MTF_REQ else fail_fill
    ee_cell = ws1.cell(row=r, column=3)
    ee_cell.fill = pass_fill if sr["ee_1x1"] >= EE_1x1_REQ else fail_fill

# --- Sheet 2: Lab Comparison ---
ws2 = wb_out.create_sheet("Lab Comparison")
ws2["A1"] = "IPC Lab Measurement vs. Model Prediction"
ws2["A1"].font = Font(bold=True, size=14)

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 20
ws2.column_dimensions["G"].width = 20

lab_headers = ["Sample", "IPC [%]", "Meas MTF@Nyq [—]", "RADIANT MTF@Nyq [—]",
               "Δ MTF [—]", "Meas EE 1×1 [—]", "RADIANT EE 1×1 [—]"]
for col, h_text in enumerate(lab_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

for i, lmr in enumerate(lab_model_results, 4):
    ws2.cell(row=i, column=1, value=lmr["sample_id"]).border = thin_border
    ws2.cell(row=i, column=2, value=lmr["ipc_pct"]).border = thin_border
    ws2.cell(row=i, column=3, value=lmr["meas_mtf"]).border = thin_border
    ws2.cell(row=i, column=4, value=round(lmr["model_mtf"], 4)).border = thin_border
    ws2.cell(row=i, column=5, value=round(lmr["delta_mtf"], 4)).border = thin_border
    ws2.cell(row=i, column=6, value=lmr["meas_ee1"]).border = thin_border
    ws2.cell(row=i, column=7, value=round(lmr["model_ee1"], 4)).border = thin_border

# --- Sheet 3: Summary ---
ws3 = wb_out.create_sheet("Summary")
ws3["A1"] = "Scenario 2.3: IPC Analysis Summary"
ws3["A1"].font = Font(bold=True, size=14)

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 20

summary_items = [
    ("Baseline MTF at Nyquist (no IPC) [—]", f"{baseline_mtf_nyq:.4f}"),
    ("Baseline SNR [—]", f"{baseline_snr:.2f}"),
    ("Baseline EE 1×1 [—]", f"{baseline_ee_1x1:.4f}"),
    ("Baseline RER [—]", f"{baseline_rer:.4f}" if baseline_rer else "N/A"),
    ("Baseline FWHM [µm]", f"{baseline_fwhm_x * 1e6:.1f}" if baseline_fwhm_x else "N/A"),
    ("GSD (cross-track) [m]", f"{baseline_gsd:.2f}" if baseline_gsd else "N/A"),
    ("Q (sampling parameter) [—]", f"{baseline_q:.3f}" if baseline_q else "N/A"),
    ("NEDT [K]", f"{baseline_nedt:.4f}" if baseline_nedt else "N/A"),
    ("NIIRS [—]", f"{baseline_niirs:.2f}" if baseline_niirs else "N/A"),
    ("Strehl ratio [—]", f"{baseline_strehl:.4f}" if baseline_strehl else "N/A"),
    ("", ""),
    ("MTF requirement [—]", f"≥ {MTF_REQ}"),
    ("EE 1×1 requirement [—]", f"≥ {EE_1x1_REQ}"),
    ("SNR requirement [—]", f"≥ {SNR_REQ:.0f}"),
    ("", ""),
    ("Max IPC for MTF requirement [%]", f"{mtf_limit_ipc:.2f}" if mtf_limit_ipc else "> 5.0"),
    ("Max IPC for EE requirement [%]", f"{ee_limit_ipc:.2f}" if ee_limit_ipc else "> 5.0"),
    ("Binding constraint", binding_constraint),
    ("Max tolerable IPC [%]", f"{binding_limit:.2f}" if binding_limit else "> 5.0"),
    ("", ""),
    ("Vendor typical IPC [%]", f"{det_specs['IPC coupling (typical)']}"),
    ("Vendor typical vs. limit", "PASS" if binding_limit and 1.8 < binding_limit else "FAIL"),
    ("Vendor max IPC [%]", f"{det_specs['IPC coupling (max)']}"),
    ("Vendor max vs. limit", "PASS" if binding_limit and 2.5 < binding_limit else "FAIL"),
]

for i, (label, value) in enumerate(summary_items, 3):
    ws3.cell(row=i, column=1, value=label).font = Font(bold=True) if label else Font()
    ws3.cell(row=i, column=2, value=value)

wb_out.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
