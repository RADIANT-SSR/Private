"""Create Mike's detector characterization spreadsheet for IPC analysis."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=5):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes="", n_cols=4):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# =========================================================================
# Sheet 1: Detector Vendor Datasheet
# =========================================================================
ws1 = wb.active
ws1.title = "Detector Specs"

ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 50

ws1["A1"] = "HgCdTe MWIR Detector — Vendor Datasheet Extract"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "Vendor: Teledyne FLIR (fictional example)"
ws1["A2"].font = Font(italic=True, size=10)
ws1["A3"] = "Part No: TDI-MW-2048-18"
ws1["A3"].font = Font(italic=True, size=10)

for col, hdr in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws1.cell(row=5, column=col, value=hdr)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 6

# FPA geometry
row = add_section(ws1, row, "FPA Geometry")
row = add_param(ws1, row, "Array format", "2048 x 1024", "pixels", "TDI-capable pushbroom")
row = add_param(ws1, row, "Pixel pitch", 18, "µm", "Square pixels")
row = add_param(ws1, row, "Active area", "36.86 x 18.43", "mm", "Derived from format × pitch")
row = add_param(ws1, row, "Fill factor", 95, "%", "With microlens array")

# Electro-optical performance
row = add_section(ws1, row, "Electro-Optical Performance")
row = add_param(ws1, row, "Spectral range", "3.4 – 5.1", "µm", "50% cutoff wavelengths")
row = add_param(ws1, row, "Peak QE", 78, "%", "At 4.2 µm, 77 K")
row = add_param(ws1, row, "Average QE (in-band)", 72, "%", "Weighted average 3.5–5.0 µm")
row = add_param(ws1, row, "Dark current (mean)", 50, "e⁻/s", "At 77 K operating temp")
row = add_param(ws1, row, "Dark current (max)", 150, "e⁻/s", "Worst pixel, 77 K")
row = add_param(ws1, row, "Operating temperature", 77, "K", "LN₂ or Stirling cooler")
row = add_param(ws1, row, "Full well capacity", 1.8e6, "e⁻", "High-gain mode")
row = add_param(ws1, row, "Operability", 99.8, "%", "Pixels meeting all specs")

# IPC
row = add_section(ws1, row, "Inter-Pixel Capacitance (IPC)")
row = add_param(ws1, row, "IPC coupling (typical)", 1.8, "%", "Per nearest neighbor (4-connected)")
row = add_param(ws1, row, "IPC coupling (max)", 2.5, "%", "Worst-case across FPA")
row = add_param(ws1, row, "IPC coupling (lot range)", "1.2 – 2.8", "%", "Range across 5 wafers")
row = add_param(ws1, row, "Diagonal coupling", 0.3, "%", "Corner neighbors (typically ~1/6 of NN)")

# Noise
row = add_section(ws1, row, "Noise Performance")
row = add_param(ws1, row, "Read noise (CDS)", 18, "e⁻ RMS", "After correlated double sampling")
row = add_param(ws1, row, "Read noise (Fowler-16)", 8, "e⁻ RMS", "16-sample Fowler averaging")
row = add_param(ws1, row, "PRNU", 0.5, "%", "After 2-point NUC")
row = add_param(ws1, row, "DSNU", 5, "e⁻ RMS", "After 2-point NUC")

# ROIC
row = add_section(ws1, row, "ROIC")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "On-chip ADC")
row = add_param(ws1, row, "System gain", 1.2, "e⁻/DN", "High-gain mode")
row = add_param(ws1, row, "Max frame rate", 120, "Hz", "Full frame readout")
row = add_param(ws1, row, "TDI stages available", "1, 4, 8, 16, 32", "—", "Selectable")

# =========================================================================
# Sheet 2: IPC Lab Measurements
# =========================================================================
ws2 = wb.create_sheet("IPC Measurements")

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 40

ws2["A1"] = "IPC Lab Characterization — Knife-Edge MTF Test"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Test date: 2026-03-20    Operator: M. Chen    Dewar SN: DW-0042"
ws2["A2"].font = Font(italic=True, size=10)
ws2["A3"] = "Method: single-pixel illumination, 4-neighbor coupling fraction"
ws2["A3"].font = Font(italic=True, size=10)

for col, hdr in enumerate(
    ["Sample ID", "IPC Coupling [%]", "Measured MTF @ Nyq", "Measured EE 1x1", "Measured EE 3x3", "Notes"], 1
):
    cell = ws2.cell(row=5, column=col, value=hdr)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# Lab data — 5 samples with different IPC levels
lab_data = [
    ("DW-0042-A", 1.2, 0.38, 0.82, 0.98, "Best unit, center of array"),
    ("DW-0042-B", 1.8, 0.34, 0.78, 0.97, "Typical unit, center of array"),
    ("DW-0042-C", 2.3, 0.30, 0.74, 0.96, "Marginal unit, center of array"),
    ("DW-0042-D", 2.8, 0.26, 0.70, 0.95, "High IPC, edge of array"),
    ("DW-0042-E", 3.5, 0.21, 0.65, 0.93, "Reject unit — over spec"),
]

for i, (sample, ipc, mtf, ee1, ee3, note) in enumerate(lab_data, 6):
    ws2.cell(row=i, column=1, value=sample).border = thin_border
    ws2.cell(row=i, column=2, value=ipc).border = thin_border
    ws2.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=3, value=mtf).border = thin_border
    ws2.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=4, value=ee1).border = thin_border
    ws2.cell(row=i, column=4).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=5, value=ee3).border = thin_border
    ws2.cell(row=i, column=5).alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=6, value=note).border = thin_border

# =========================================================================
# Sheet 3: System Requirements
# =========================================================================
ws3 = wb.create_sheet("System Requirements")

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 50

ws3["A1"] = "MWIR LEO Pushbroom — System-Level Requirements"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Source: SRD-2026-042 Rev C, Section 4.3"
ws3["A2"].font = Font(italic=True, size=10)

for col, hdr in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws3.cell(row=4, column=col, value=hdr)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 5

# Optical system
row = add_section(ws3, row, "Optical System")
row = add_param(ws3, row, "Aperture diameter", 30, "cm", "Circular, unobscured")
row = add_param(ws3, row, "Focal length", 120, "cm", "Derived: f/4 × 30 cm")
row = add_param(ws3, row, "f-number", 4.0, "—", "System f/#")
row = add_param(ws3, row, "Optical transmission", 72, "%", "End-to-end, all elements")
row = add_param(ws3, row, "Optics temperature", 293, "K", "Ambient")
row = add_param(ws3, row, "WFE (RMS)", 0.07, "waves", "At 4.0 µm reference wavelength")

# Scene
row = add_section(ws3, row, "Scene / Target")
row = add_param(ws3, row, "Target temperature", 310, "K", "Warm ground target")
row = add_param(ws3, row, "Target emissivity", 0.90, "—", "Painted metal")
row = add_param(ws3, row, "Background temperature", 295, "K", "Ambient terrain")
row = add_param(ws3, row, "Background emissivity", 0.95, "—", "Vegetation/soil")

# Spectral
row = add_section(ws3, row, "Spectral / Integration")
row = add_param(ws3, row, "Spectral band", "3.5 – 5.0", "µm", "MWIR bandpass filter")
row = add_param(ws3, row, "Integration time", 2, "ms", "Pushbroom line rate constrained")

# Geometry
row = add_section(ws3, row, "Geometry")
row = add_param(ws3, row, "Orbit altitude", 500, "km", "LEO sun-synchronous")
row = add_param(ws3, row, "Look angle", 0, "deg", "Nadir")

# Performance requirements
row = add_section(ws3, row, "Performance Requirements (thresholds)")
row = add_param(ws3, row, "System MTF at Nyquist", "≥ 0.15", "—", "KPP — shall meet")
row = add_param(ws3, row, "SNR", "≥ 100", "—", "At reference scene, single frame")
row = add_param(ws3, row, "EE 1×1 (ensquared energy)", "≥ 0.60", "—", "Energy in central pixel")

output_path = "mike_detector_ipc_data.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
