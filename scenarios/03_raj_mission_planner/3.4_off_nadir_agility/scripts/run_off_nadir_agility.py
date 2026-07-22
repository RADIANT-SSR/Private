"""Scenario 3.4: Off-Nadir Performance Degradation

Raj needs to understand how performance degrades as look angle increases from
nadir to 45 deg off-nadir from a 600 km LEO.  He sweeps off-nadir angle and
tracks SNR, GSD, NIIRS, atmospheric transmission, and slant range.

Approach:
  Sweep geometry.path_zenith_rad from 0 to 45 deg (converted to radians).
  At each angle, RADIANT evaluates the full signal chain with increased
  atmospheric path length and (where applicable) degraded geometry.

  GSD at off-nadir is computed analytically in this script because RADIANT's
  GSD metric only computes nadir GSD (Gap 33).  The formula:
    GSD_cross = p × slant_range / f
    GSD_along = p × slant_range / (f × cos(theta))
  where slant_range = H / cos(theta) for flat-Earth (valid to ~60 deg).

  Swath width is computed geometrically:
    W = 2 × H × tan(half_FOV) / cos(theta)  (simplified)

Physics:
  - Slant range: R = H / cos(theta)  (flat-Earth; spherical correction at >80 deg)
  - Air mass:    M = R / H = 1/cos(theta)  → transmission = tau_nadir^M
  - GSD scales as slant_range / focal_length → grows with 1/cos(theta)
  - Along-track GSD has an additional 1/cos(theta) factor from ground projection
  - SNR decreases: longer path → lower transmission → fewer photons
  - NIIRS degrades from both GSD increase and SNR decrease

Key thresholds:
  - 15 deg: minor degradation (~3% GSD increase)
  - 30 deg: moderate (15% GSD, noticeable SNR drop)
  - 45 deg: significant (41% GSD, substantial SNR and NIIRS loss)

Usage:
    python run_off_nadir_agility.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EARTH_RADIUS_M = 6_371_000.0

# ---------------------------------------------------------------------------
# Step 1: Read Raj's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "raj_off_nadir_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "off_nadir_results.xlsx"
PLOT_DIR = Path(__file__).parent.parent / "outputs"

wb_in = openpyxl.load_workbook(INPUT_FILE)

# --- Sensor Configuration ---
ws_sys = wb_in["Sensor Configuration"]
specs: dict[str, object] = {}
units: dict[str, str] = {}
for row in ws_sys.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None and not isinstance(value, str) or (
        isinstance(value, str) and value not in ("", "—")
    ):
        try:
            specs[name] = float(value)
        except (ValueError, TypeError):
            specs[name] = value
        units[name] = str(row[2].value) if row[2].value else "—"

# --- Orbit & Geometry ---
ws_geo = wb_in["Orbit & Geometry"]
geo_specs: dict[str, object] = {}
for row in ws_geo.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        try:
            geo_specs[name] = float(value)
        except (ValueError, TypeError):
            geo_specs[name] = value

# --- Off-Nadir Sweep ---
ws_sweep = wb_in["Off-Nadir Sweep"]
angles_deg: list[float] = []
for row in ws_sweep.iter_rows(min_row=5, max_col=1, values_only=True):
    if row[0] is not None:
        try:
            angles_deg.append(float(row[0]))
        except (ValueError, TypeError):
            pass

print("=" * 80)
print("SCENARIO 3.4: Off-Nadir Performance Degradation")
print("=" * 80)

print("\n=== Sensor Parameters (from spreadsheet) ===")
for k, v in specs.items():
    print(f"  {k:<35s}: {v} [{units.get(k, '—')}]")

print(f"\n=== Orbit & Geometry ===")
for k, v in geo_specs.items():
    print(f"  {k:<35s}: {v}")

print(f"\n=== Off-Nadir Sweep ===")
print(f"  Angles: {angles_deg} [deg]")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

aperture_m = float(specs["Aperture diameter"]) / 100.0       # cm → m
focal_length_m = float(specs["Focal length"]) / 100.0        # cm → m
f_number = float(specs["f-number"])
transmission = float(specs["Optical transmission"]) / 100.0  # % → fraction
optics_temp_K = float(specs["Optics temperature"]) + 273.15  # °C → K
obscuration = float(specs["Central obscuration"]) / 100.0    # % → fraction
wfe_rms = float(specs["WFE RMS"])

filter_min_nm = float(specs["Filter min"])
filter_max_nm = float(specs["Filter max"])
filter_min_um = filter_min_nm / 1000.0                       # nm → µm
filter_max_um = filter_max_nm / 1000.0
band_center_um = (filter_min_um + filter_max_um) / 2.0

pixel_pitch_um = float(specs["Pixel pitch"])
pixel_pitch_m = pixel_pitch_um * 1e-6                        # µm → m
fill_factor = float(specs["Fill factor"]) / 100.0            # % → fraction
qe = float(specs["Quantum efficiency"]) / 100.0              # % → fraction
dark_rate = float(specs["Dark current"])
det_temp_K = float(specs["Operating temperature"])
read_noise = float(specs["Read noise"])
fwc = float(specs["Full well capacity"])
adc_bits = int(specs["ADC bits"])
gain = float(specs["System gain"])
ipc_coupling = float(specs["IPC coupling"]) / 100.0          # % → fraction
t_int_ms = float(specs["Integration time"])
t_int_s = t_int_ms / 1000.0                                  # ms → s

altitude_km = float(geo_specs["Altitude"])
altitude_m = altitude_km * 1000.0                            # km → m
solar_zenith_deg = float(geo_specs["Solar zenith angle"])
solar_zenith_rad = solar_zenith_deg * math.pi / 180.0        # deg → rad

target_refl = float(geo_specs["Target reflectance"])
bg_refl = float(geo_specs["Background reflectance"])
target_temp_K = float(geo_specs["Target temperature"])
bg_temp_K = float(geo_specs["Background temperature"])
target_emissivity = 1.0 - target_refl
bg_emissivity = 1.0 - bg_refl

# Derived nadir parameters
gsd_nadir_m = pixel_pitch_m * altitude_m / focal_length_m
ifov_urad = pixel_pitch_m / focal_length_m * 1e6
Q = band_center_um * f_number / pixel_pitch_um
f_nyquist_cy_m = 1.0 / (2.0 * pixel_pitch_m)

print(f"\n=== Converted to RADIANT Canonical Units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<12s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<12s}  cm / 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<12s}  cm / 100")
print(f"  {'Transmission':<35s} {transmission:>14.4f}  {'fraction':<12s}  % / 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<12s}  °C + 273.15")
print(f"  {'Obscuration ratio':<35s} {obscuration:>14.4f}  {'fraction':<12s}  % / 100")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_m:>14.2e}  {'m':<12s}  µm × 1e-6")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<12s}  ms / 1000")
print(f"  {'Altitude':<35s} {altitude_m:>14.0f}  {'m':<12s}  km × 1000")
print(f"  {'Solar zenith':<35s} {solar_zenith_rad:>14.4f}  {'rad':<12s}  deg × π/180")
print(f"  {'Band':<35s} {filter_min_um:>6.3f}-{filter_max_um:<6.3f}  {'µm':<12s}  nm / 1000")

print(f"\n=== Derived Nadir Parameters ===")
print(f"  GSD (nadir):       {gsd_nadir_m:.2f} [m]")
print(f"  IFOV:              {ifov_urad:.1f} [µrad]")
print(f"  Q (sampling):      {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")
print(f"  f_Nyquist:         {f_nyquist_cy_m:.0f} [cy/m]")

# ---------------------------------------------------------------------------
# Step 3: Geometry helper functions
# ---------------------------------------------------------------------------


def slant_range_flat(altitude_m: float, theta_rad: float) -> float:
    """Flat-Earth slant range: R = H / cos(theta) [m]."""
    return altitude_m / math.cos(theta_rad)


def slant_range_spherical(altitude_m: float, theta_rad: float) -> float:
    """Spherical-Earth slant range [m].

    Uses the exact geometric relation for a sensor at altitude H above
    a spherical Earth of radius R_E, looking at angle theta from nadir.

    R = sqrt((R_E + H)^2 - R_E^2 * cos^2(theta_nadir)) - R_E * sin(pi/2 - theta_nadir)

    For off-nadir angle theta (from nadir), the earth-center angle and
    slant range are:
        R = R_E * [-cos(theta + phi) + sqrt(cos^2(theta + phi) ... )]
    Simplified for the sensor zenith angle = theta:
        R = -R_E*cos(theta) + sqrt(R_E^2*cos^2(theta) + 2*R_E*H + H^2)
    """
    R_E = EARTH_RADIUS_M
    cos_t = math.cos(theta_rad)
    return -R_E * cos_t + math.sqrt(R_E**2 * cos_t**2 + 2 * R_E * altitude_m + altitude_m**2)


def ground_range(altitude_m: float, theta_rad: float) -> float:
    """Ground range (arc distance) for spherical Earth [m]."""
    R_E = EARTH_RADIUS_M
    sin_t = math.sin(theta_rad)
    # Earth-center angle: sin(gamma) = (R_E + H) * sin(theta) / R_E
    # But for near-nadir, use simpler formula
    sr = slant_range_spherical(altitude_m, theta_rad)
    # Use law of cosines to get Earth-center angle
    r_sensor = R_E + altitude_m
    cos_gamma = (r_sensor**2 + R_E**2 - sr**2) / (2 * r_sensor * R_E)
    cos_gamma = max(-1.0, min(1.0, cos_gamma))
    gamma = math.acos(cos_gamma)
    return R_E * gamma


def gsd_off_nadir(pixel_pitch_m: float, altitude_m: float,
                  focal_length_m: float, theta_rad: float) -> tuple[float, float]:
    """Compute cross-track and along-track GSD at off-nadir angle.

    Returns (gsd_cross_m, gsd_along_m).

    Cross-track GSD scales with slant range:
        GSD_cross = p × R / f

    Along-track GSD has additional ground projection factor:
        GSD_along = p × R / (f × cos(incidence_angle))

    For flat-Earth, incidence angle ≈ theta, so:
        GSD_along = p × R / (f × cos(theta)) = p × H / (f × cos^2(theta))
    """
    sr = slant_range_spherical(altitude_m, theta_rad)
    gsd_cross = pixel_pitch_m * sr / focal_length_m

    # Incidence angle at ground (differs from look angle due to Earth curvature)
    # For small angles, incidence ≈ look angle
    R_E = EARTH_RADIUS_M
    r_sensor = R_E + altitude_m
    sin_inc = r_sensor * math.sin(theta_rad) / R_E
    sin_inc = min(sin_inc, 1.0)
    inc_angle = math.asin(sin_inc)
    cos_inc = math.cos(inc_angle)

    gsd_along = pixel_pitch_m * sr / (focal_length_m * cos_inc) if cos_inc > 0.01 else float("inf")

    return gsd_cross, gsd_along


def swath_width(altitude_m: float, theta_rad: float,
                n_pixels_cross: int, pixel_pitch_m: float,
                focal_length_m: float) -> float:
    """Approximate swath width at ground [m].

    For a pushbroom scanner with n_pixels_cross pixels:
        W ≈ n_pixels × GSD_cross
    """
    gsd_x, _ = gsd_off_nadir(pixel_pitch_m, altitude_m, focal_length_m, theta_rad)
    return n_pixels_cross * gsd_x


print(f"\n{'=' * 80}")
print(f"  GEOMETRY REFERENCE TABLE")
print(f"{'=' * 80}")

print(f"\n  {'Angle':>8s}  {'Slant Range':>14s}  {'Air Mass':>10s}  {'Ground Range':>14s}  "
      f"{'GSD_cross':>10s}  {'GSD_along':>10s}")
print(f"  {'[deg]':>8s}  {'[km]':>14s}  {'[--]':>10s}  {'[km]':>14s}  "
      f"{'[m]':>10s}  {'[m]':>10s}")
print(f"  {'-' * 8}  {'-' * 14}  {'-' * 10}  {'-' * 14}  {'-' * 10}  {'-' * 10}")

for angle_deg in angles_deg:
    theta_rad = angle_deg * math.pi / 180.0
    sr = slant_range_spherical(altitude_m, theta_rad)
    am = sr / altitude_m if altitude_m > 0 else 1.0
    gr = ground_range(altitude_m, theta_rad)
    gsd_x, gsd_y = gsd_off_nadir(pixel_pitch_m, altitude_m, focal_length_m, theta_rad)

    print(f"  {angle_deg:>8.0f}  {sr / 1000:>14.1f}  {am:>10.4f}  {gr / 1000:>14.1f}  "
          f"{gsd_x:>10.2f}  {gsd_y:>10.2f}")

# ---------------------------------------------------------------------------
# Step 4: Build base RADIANT config
# ---------------------------------------------------------------------------

base_config = {
    "source": {
        "target": {"temperature": target_temp_K, "emissivity": target_emissivity},
        "background": {"temperature": bg_temp_K, "emissivity": bg_emissivity},
    },
    "atmosphere": {
        "model": "simple",
        "standard_atmosphere": "midlat_summer",
    },
    "geometry": {
        "sensor_altitude_m": altitude_m,
        "path_zenith_rad": 0.0,  # overridden per sweep point
        "solar_zenith_rad": solar_zenith_rad,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "optics_temperature_K": optics_temp_K,
        "wfe_rms_waves": wfe_rms,
        "obscuration_ratio": obscuration,
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate,
        "detector_temperature_K": det_temp_K,
        "fill_factor": fill_factor,
        "ipc_coupling": ipc_coupling,
    },
    "spectral_integration": {
        "filter_min_um": filter_min_um,
        "filter_max_um": filter_max_um,
        "integration_time_s": t_int_s,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "full_well_capacity_e": fwc,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
    },
    # CU-178: this MWIR recon config sits above the GIQE-5 SNR envelope, so the
    # CU-166 applicability gate returns NIIRS N/A by default. This scenario's
    # whole story is the NIIRS-vs-off-nadir trend, so opt into the extrapolated
    # GIQE-5-form rating (documented as a relative trend in the walkthrough).
    "performance": {
        "niirs": {"allow_extrapolated": True},
    },
}

# ---------------------------------------------------------------------------
# Step 5: Sweep off-nadir angle
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  OFF-NADIR SWEEP")
print(f"{'=' * 80}")

print(f"\n  {'Angle':>8s}  {'Slant':>10s}  {'AirMass':>8s}  {'tau_mean':>8s}  "
      f"{'GSD_x':>8s}  {'GSD_y':>8s}  {'SNR':>8s}  {'NIIRS':>8s}  {'NEDT':>10s}")
print(f"  {'[deg]':>8s}  {'[km]':>10s}  {'[--]':>8s}  {'[--]':>8s}  "
      f"{'[m]':>8s}  {'[m]':>8s}  {'[--]':>8s}  {'[--]':>8s}  {'[mK]':>10s}")
print(f"  {'-' * 8}  {'-' * 10}  {'-' * 8}  {'-' * 8}  "
      f"{'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 10}")

results: list[dict] = []

for angle_deg in angles_deg:
    theta_rad = angle_deg * math.pi / 180.0

    config = {k: ({**v} if isinstance(v, dict) else v) for k, v in base_config.items()}
    config["source"] = {
        "target": {**base_config["source"]["target"]},
        "background": {**base_config["source"]["background"]},
    }
    config["geometry"] = {**base_config["geometry"], "path_zenith_rad": theta_rad}

    sensor = Sensor.from_dict(config)
    r = sensor.evaluate()

    # Extract RADIANT metrics
    snr = r.metrics.get("snr", 0.0)
    mtf_nyq = r.metrics.get("mtf_at_nyquist", 0.0)
    rer = r.metrics.get("rer", 0.0)
    strehl = r.metrics.get("strehl", 0.0)
    ee_1x1 = r.metrics.get("ee_1x1", 0.0)
    nedt_K = r.metrics.get("nedt_K", 0.0)
    gsd_radiant_x = r.metrics.get("gsd_cross_track_m", 0.0)
    gsd_radiant_y = r.metrics.get("gsd_along_track_m", 0.0)

    # NIIRS — RADIANT computes from its own GSD (nadir only), so we note the gap
    niirs_radiant = r.metrics.get("niirs", 0.0)

    # Atmospheric transmission (mean over band)
    atm_out = r.stage_outputs.get("atmosphere", {})
    tau_arr = atm_out.get("tau_atm")
    tau_mean = float(np.mean(tau_arr)) if tau_arr is not None else 0.0

    # Signal
    ro_out = r.stage_outputs.get("readout", {})
    signal_e = ro_out.get("signal_e_final", 0.0)

    # Geometry computations (outside RADIANT)
    sr = slant_range_spherical(altitude_m, theta_rad)
    am = sr / altitude_m if altitude_m > 0 else 1.0
    gr = ground_range(altitude_m, theta_rad)
    gsd_x, gsd_y = gsd_off_nadir(pixel_pitch_m, altitude_m, focal_length_m, theta_rad)

    # Corrected NIIRS using true off-nadir GSD
    # GIQE-5 (simplified): NIIRS = c0 + c1*log10(GSD_GM) + c2*log10(RER) + c3*SNR/G
    # We'll recompute using the off-nadir GSD instead of RADIANT's nadir GSD
    gsd_gm = math.sqrt(gsd_x * gsd_y)  # geometric mean of cross/along GSD

    # GIQE-5 for VIS/NIR: a simplified approximation
    # NIIRS = 10.251 - 3.32*log10(GSD_GM) + 0.656*log10(RER_GM) - 0.334*(H/p)
    # Using RADIANT's RER and SNR but corrected GSD
    if gsd_gm > 0 and rer > 0 and snr > 0:
        niirs_corrected = (10.251
                           - 3.32 * math.log10(gsd_gm)
                           + 0.656 * math.log10(rer * rer)  # RER_GM ≈ RER^2 (both axes)
                           - 0.334 * (gsd_gm / snr * 10.0))  # simplified noise/SNR term
        # Use a more practical approximation: scale RADIANT NIIRS by GSD ratio
        if niirs_radiant > 0 and gsd_radiant_x > 0:
            gsd_ratio = gsd_gm / gsd_radiant_x
            # NIIRS change ≈ -3.32 * log10(GSD_new/GSD_old)
            niirs_corrected = niirs_radiant - 3.32 * math.log10(gsd_ratio)
        else:
            niirs_corrected = 0.0
    else:
        niirs_corrected = 0.0

    nedt_mK = nedt_K * 1000.0 if nedt_K > 0 else 0.0
    well_margin_dB = r.metrics.get("well_margin_dB", 0.0)
    dynamic_range_dB = r.metrics.get("dynamic_range_dB", 0.0)
    q_center = r.metrics.get("q_center", 0.0)
    fwhm_x_m = r.metrics.get("fwhm_x_m", 0.0)

    row_data = {
        "angle_deg": angle_deg,
        "theta_rad": theta_rad,
        "slant_range_km": sr / 1000.0,
        "air_mass": am,
        "ground_range_km": gr / 1000.0,
        "tau_mean": tau_mean,
        "gsd_cross_m": gsd_x,
        "gsd_along_m": gsd_y,
        "gsd_gm_m": gsd_gm,
        "gsd_radiant_x": gsd_radiant_x,
        "snr": snr,
        "mtf_nyq": mtf_nyq,
        "rer": rer,
        "strehl": strehl,
        "ee_1x1": ee_1x1,
        "niirs_radiant": niirs_radiant,
        "niirs_corrected": niirs_corrected,
        "nedt_mK": nedt_mK,
        "signal_e": signal_e,
        "well_margin_dB": well_margin_dB,
        "dynamic_range_dB": dynamic_range_dB,
        "q_center": q_center,
        "fwhm_x_m": fwhm_x_m,
    }
    results.append(row_data)

    print(f"  {angle_deg:>8.0f}  {sr / 1000:>10.1f}  {am:>8.4f}  {tau_mean:>8.4f}  "
          f"{gsd_x:>8.2f}  {gsd_y:>8.2f}  {snr:>8.1f}  {niirs_corrected:>8.2f}  "
          f"{nedt_mK:>10.1f}")

# ---------------------------------------------------------------------------
# Step 6: Degradation analysis
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  PERFORMANCE DEGRADATION RELATIVE TO NADIR")
print(f"{'=' * 80}")

baseline = results[0]

print(f"\n  Nadir baseline:")
print(f"    Slant range:     {baseline['slant_range_km']:.1f} [km]")
print(f"    GSD (cross):     {baseline['gsd_cross_m']:.2f} [m]")
print(f"    GSD (along):     {baseline['gsd_along_m']:.2f} [m]")
print(f"    Transmission:    {baseline['tau_mean']:.4f} [--]")
print(f"    SNR:             {baseline['snr']:.1f} [--]")
print(f"    NIIRS:           {baseline['niirs_corrected']:.2f} [--]")

print(f"\n  {'Angle':>8s}  {'dRange':>8s}  {'dGSD_x':>8s}  {'dGSD_y':>8s}  "
      f"{'dTau':>8s}  {'dSNR':>8s}  {'dNIIRS':>8s}")
print(f"  {'[deg]':>8s}  {'[%]':>8s}  {'[%]':>8s}  {'[%]':>8s}  "
      f"{'[%]':>8s}  {'[%]':>8s}  {'[--]':>8s}")
print(f"  {'-' * 8}  {'-' * 8}  {'-' * 8}  {'-' * 8}  "
      f"{'-' * 8}  {'-' * 8}  {'-' * 8}")

for rd in results:
    d_range = ((rd["slant_range_km"] - baseline["slant_range_km"])
               / baseline["slant_range_km"] * 100.0)
    d_gsd_x = ((rd["gsd_cross_m"] - baseline["gsd_cross_m"])
               / baseline["gsd_cross_m"] * 100.0)
    d_gsd_y = ((rd["gsd_along_m"] - baseline["gsd_along_m"])
               / baseline["gsd_along_m"] * 100.0)
    d_tau = ((rd["tau_mean"] - baseline["tau_mean"])
             / baseline["tau_mean"] * 100.0 if baseline["tau_mean"] > 0 else 0.0)
    d_snr = ((rd["snr"] - baseline["snr"])
             / baseline["snr"] * 100.0 if baseline["snr"] > 0 else 0.0)
    d_niirs = rd["niirs_corrected"] - baseline["niirs_corrected"]

    print(f"  {rd['angle_deg']:>8.0f}  {d_range:>+8.1f}  {d_gsd_x:>+8.1f}  {d_gsd_y:>+8.1f}  "
          f"{d_tau:>+8.1f}  {d_snr:>+8.1f}  {d_niirs:>+8.2f}")

# ---------------------------------------------------------------------------
# Step 7: NIIRS threshold analysis
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  NIIRS THRESHOLD ANALYSIS")
print(f"{'=' * 80}")

niirs_0 = baseline["niirs_corrected"]
print(f"\n  Nadir NIIRS: {niirs_0:.2f} [--]")
print(f"  GIQE-5 NIIRS degrades primarily through the GSD term:")
print(f"    dNIIRS = -3.32 × log10(GSD_offnadir / GSD_nadir)")
print(f"  Secondary effects: reduced SNR from lower atmospheric transmission.")

for threshold in [0.5, 1.0, 1.5, 2.0]:
    angle_at_thresh = None
    for rd in results:
        if rd["niirs_corrected"] < niirs_0 - threshold:
            angle_at_thresh = rd["angle_deg"]
            break
    if angle_at_thresh is not None:
        print(f"\n  NIIRS drops by {threshold:.1f} at ~{angle_at_thresh:.0f} [deg] off-nadir")
    else:
        print(f"\n  NIIRS does not drop by {threshold:.1f} within 0-45 deg range")

# ---------------------------------------------------------------------------
# Step 8: Swath width analysis
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  ACCESS AND SWATH GEOMETRY")
print(f"{'=' * 80}")

n_pixels = 12000  # typical pushbroom array
print(f"\n  Assuming {n_pixels}-pixel cross-track array:")

print(f"\n  {'Angle':>8s}  {'Swath Width':>14s}  {'Ground Range':>14s}  {'Access Area':>14s}")
print(f"  {'[deg]':>8s}  {'[km]':>14s}  {'[km]':>14s}  {'[km^2/s]':>14s}")
print(f"  {'-' * 8}  {'-' * 14}  {'-' * 14}  {'-' * 14}")

ground_speed_km_s = float(geo_specs.get("Ground speed", 6.9))

for rd in results:
    sw = swath_width(altitude_m, rd["theta_rad"], n_pixels,
                     pixel_pitch_m, focal_length_m)
    sw_km = sw / 1000.0
    gr_km = rd["ground_range_km"]
    # Access area rate ≈ swath_width × ground_speed
    access_rate = sw_km * ground_speed_km_s

    print(f"  {rd['angle_deg']:>8.0f}  {sw_km:>14.1f}  {gr_km:>14.1f}  {access_rate:>14.0f}")

print(f"\n  Note: Agile pointing trades image quality (NIIRS) for access area.")
print(f"  At 45 deg off-nadir, swath is ~{swath_width(altitude_m, 45 * math.pi / 180, n_pixels, pixel_pitch_m, focal_length_m) / 1000:.0f} km wide")
print(f"  but GSD degrades by ~{(results[-1]['gsd_gm_m'] / results[0]['gsd_gm_m'] - 1) * 100:.0f}%.")

# ---------------------------------------------------------------------------
# Step 9: RADIANT GSD vs. true off-nadir GSD comparison
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  RADIANT GSD vs. TRUE OFF-NADIR GSD")
print(f"{'=' * 80}")

print(f"\n  RADIANT's GSD metric uses nadir formula: GSD = p × H / f")
print(f"  It does NOT account for path_zenith_rad (Gap 33).")
print(f"  True GSD scales with slant range: GSD = p × R / f")

print(f"\n  {'Angle':>8s}  {'RADIANT GSD':>12s}  {'True GSD_x':>12s}  {'True GSD_y':>12s}  "
      f"{'Error':>8s}")
print(f"  {'[deg]':>8s}  {'[m]':>12s}  {'[m]':>12s}  {'[m]':>12s}  {'[%]':>8s}")
print(f"  {'-' * 8}  {'-' * 12}  {'-' * 12}  {'-' * 12}  {'-' * 8}")

for rd in results:
    error_pct = ((rd["gsd_radiant_x"] - rd["gsd_cross_m"])
                 / rd["gsd_cross_m"] * 100.0 if rd["gsd_cross_m"] > 0 else 0.0)
    print(f"  {rd['angle_deg']:>8.0f}  {rd['gsd_radiant_x']:>12.2f}  {rd['gsd_cross_m']:>12.2f}  "
          f"{rd['gsd_along_m']:>12.2f}  {error_pct:>+8.1f}")

# ---------------------------------------------------------------------------
# Step 9b: RADIANT MTF budget and performance metrics (nadir baseline)
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  RADIANT PERFORMANCE METRICS (Nadir Baseline)")
print(f"{'=' * 80}")

# Re-run nadir to get detailed outputs
nadir_config = {k: ({**v} if isinstance(v, dict) else v) for k, v in base_config.items()}
nadir_config["source"] = {
    "target": {**base_config["source"]["target"]},
    "background": {**base_config["source"]["background"]},
}
nadir_sensor = Sensor.from_dict(nadir_config)
r_nadir = nadir_sensor.evaluate()

print(f"\n  --- Spatial Metrics ---")
print(f"  Strehl:            {r_nadir.metrics.get('strehl', 0.0):.4f} [--]")
print(f"  RER:               {r_nadir.metrics.get('rer', 0.0):.4f} [--]")
print(f"  FWHM_x:            {r_nadir.metrics.get('fwhm_x_m', 0.0) * 1e6:.2f} [µm]")
print(f"  EE(1x1):           {r_nadir.metrics.get('ee_1x1', 0.0):.4f} [--]")
print(f"  Q (center):        {r_nadir.metrics.get('q_center', 0.0):.3f} [--]")
print(f"  Q (min/max):       {r_nadir.metrics.get('q_min', 0.0):.3f} / {r_nadir.metrics.get('q_max', 0.0):.3f} [--]")

print(f"\n  --- Radiometric Metrics ---")
print(f"  SNR:               {r_nadir.metrics.get('snr', 0.0):.1f} [--]")
print(f"  NEDT:              {r_nadir.metrics.get('nedt_K', 0.0) * 1000:.1f} [mK]")
print(f"  Well margin:       {r_nadir.metrics.get('well_margin_dB', 0.0):.1f} [dB]")
print(f"  Dynamic range:     {r_nadir.metrics.get('dynamic_range_dB', 0.0):.1f} [dB]")

print(f"\n  --- Image Quality ---")
gsd_val = r_nadir.metrics.get("gsd_geometric_mean_m")
niirs_val = r_nadir.metrics.get("niirs")
print(f"  GSD (cross):       {r_nadir.metrics.get('gsd_cross_track_m', 0.0):.2f} [m]")
print(f"  GSD (along):       {r_nadir.metrics.get('gsd_along_track_m', 0.0):.2f} [m]")
print(f"  GSD (GM):          {gsd_val:.2f} [m]" if gsd_val else "  GSD (GM):          N/A")
print(f"  NIIRS:             {niirs_val:.2f} [--]" if niirs_val else "  NIIRS:             N/A")

# MTF budget
mtf_budget = r_nadir.stage_outputs.get("performance", {}).get("mtf_budget")
if mtf_budget is not None:
    print(f"\n  --- RADIANT MTF Budget at Nyquist ---")
    per_term = mtf_budget.per_term_at_nyquist
    print(f"  {'Component':<30s}  {'MTF@Ny_x':>10s}  {'MTF@Ny_y':>10s}")
    print(f"  {'-' * 30}  {'-' * 10}  {'-' * 10}")

    seen: set[str] = set()
    for key in per_term:
        base = key.rsplit("_", 1)[0]
        if base in seen:
            continue
        seen.add(base)
        val_x = per_term.get(f"{base}_x", 1.0)
        val_y = per_term.get(f"{base}_y", 1.0)
        label = base.replace("mtf_", "").replace("_", " ").title()
        print(f"  {label:<30s}  {val_x:>10.4f}  {val_y:>10.4f}")

    print(f"  {'─' * 30}  {'─' * 10}  {'─' * 10}")
    print(f"  {'System (product)':30s}  {mtf_budget.system_mtf_at_nyquist_x:>10.4f}  "
          f"{mtf_budget.system_mtf_at_nyquist_y:>10.4f}")

# Noise breakdown
print(f"\n  --- Noise Breakdown ---")
print(f"  {'Source':<30s}  {'Value':>10s}")
print(f"  {'-' * 30}  {'-' * 10}")
for nt in r_nadir.noise_terms:
    if nt.value_e > 0.001:
        print(f"  {nt.name:<30s}  {nt.value_e:>10.4f} [e-]")

# Folded MTF
folded_mtf_result = r_nadir.stage_outputs.get("performance", {}).get("folded_mtf_x")
if folded_mtf_result is not None and hasattr(folded_mtf_result, 'mtf_folded'):
    folded_at_ny = r_nadir.metrics.get("mtf_folded_at_nyquist", 0.0)
    print(f"\n  --- Folded MTF ---")
    print(f"  Folded MTF at Nyquist: {folded_at_ny:.4f} [--]")
    print(f"  Note: Folded MTF > 1.0 indicates aliasing is significant.")
    print(f"  Alias fraction:        {r_nadir.metrics.get('alias_fraction_at_nyquist', 0.0):.4f} [--]")

# ---------------------------------------------------------------------------
# Step 10: Plots
# ---------------------------------------------------------------------------

PLOT_DIR.mkdir(parents=True, exist_ok=True)
fig_w, fig_h = 10, 7

angles_arr = [rd["angle_deg"] for rd in results]

# Plot 1: SNR and Transmission vs. Off-Nadir Angle
fig1, ax1a = plt.subplots(figsize=(fig_w, fig_h))
ax1b = ax1a.twinx()

snr_arr = [rd["snr"] for rd in results]
tau_arr_plot = [rd["tau_mean"] for rd in results]

l1, = ax1a.plot(angles_arr, snr_arr, "bo-", linewidth=2, markersize=8, label="SNR")
l2, = ax1b.plot(angles_arr, tau_arr_plot, "rs--", linewidth=2, markersize=8,
                label="Atm. Transmission (band mean)")

ax1a.set_xlabel("Off-Nadir Angle [deg]", fontsize=12)
ax1a.set_ylabel("SNR [--]", fontsize=12, color="blue")
ax1b.set_ylabel("Atmospheric Transmission [--]", fontsize=12, color="red")
ax1a.set_title("SNR and Atmospheric Transmission vs. Off-Nadir Angle", fontsize=13)
ax1a.tick_params(axis="y", labelcolor="blue")
ax1b.tick_params(axis="y", labelcolor="red")
lines = [l1, l2]
ax1a.legend(lines, [l.get_label() for l in lines], fontsize=10, loc="center left")
ax1a.grid(True, alpha=0.3)
fig1.tight_layout()
fig1.savefig(PLOT_DIR / "fig1_snr_transmission_vs_angle.png", dpi=150)

# Plot 2: GSD (cross and along) vs. Off-Nadir Angle
fig2, ax2 = plt.subplots(figsize=(fig_w, fig_h))

gsd_x_arr = [rd["gsd_cross_m"] for rd in results]
gsd_y_arr = [rd["gsd_along_m"] for rd in results]
gsd_gm_arr = [rd["gsd_gm_m"] for rd in results]
gsd_rad_arr = [rd["gsd_radiant_x"] for rd in results]

ax2.plot(angles_arr, gsd_x_arr, "bo-", linewidth=2, markersize=8,
         label="GSD cross-track (true)")
ax2.plot(angles_arr, gsd_y_arr, "r^-", linewidth=2, markersize=8,
         label="GSD along-track (true)")
ax2.plot(angles_arr, gsd_gm_arr, "gs--", linewidth=2, markersize=6,
         label="GSD geometric mean")
ax2.plot(angles_arr, gsd_rad_arr, "k:", linewidth=1.5, alpha=0.5,
         label="RADIANT GSD (nadir only)")

ax2.axhline(gsd_nadir_m, color="gray", linestyle=":", alpha=0.3)

ax2.set_xlabel("Off-Nadir Angle [deg]", fontsize=12)
ax2.set_ylabel("Ground Sample Distance [m]", fontsize=12)
ax2.set_title("GSD vs. Off-Nadir Angle — Cross-Track and Along-Track", fontsize=13)
ax2.legend(fontsize=10)
ax2.grid(True, alpha=0.3)
fig2.tight_layout()
fig2.savefig(PLOT_DIR / "fig2_gsd_vs_angle.png", dpi=150)

# Plot 3: NIIRS vs. Off-Nadir Angle
fig3, ax3 = plt.subplots(figsize=(fig_w, fig_h))

niirs_corr_arr = [rd["niirs_corrected"] for rd in results]
niirs_rad_arr = [rd["niirs_radiant"] for rd in results]

ax3.plot(angles_arr, niirs_corr_arr, "bo-", linewidth=2, markersize=8,
         label="NIIRS (corrected for off-nadir GSD)")
ax3.plot(angles_arr, niirs_rad_arr, "r^--", linewidth=1.5, markersize=6,
         label="NIIRS (RADIANT — nadir GSD only)")

ax3.axhline(niirs_corr_arr[0], color="green", linestyle=":", alpha=0.4,
            label=f"Nadir baseline = {niirs_corr_arr[0]:.2f}")
ax3.axhline(niirs_corr_arr[0] - 1.0, color="orange", linestyle="--", alpha=0.5,
            label=f"-1.0 NIIRS = {niirs_corr_arr[0] - 1.0:.2f}")

ax3.set_xlabel("Off-Nadir Angle [deg]", fontsize=12)
ax3.set_ylabel("NIIRS [--]", fontsize=12)
ax3.set_title("NIIRS vs. Off-Nadir Angle", fontsize=13)
ax3.legend(fontsize=10)
ax3.grid(True, alpha=0.3)
fig3.tight_layout()
fig3.savefig(PLOT_DIR / "fig3_niirs_vs_angle.png", dpi=150)

# Plot 4: Multi-panel summary
fig4, ((ax4a, ax4b), (ax4c, ax4d)) = plt.subplots(2, 2, figsize=(fig_w * 1.3, fig_h * 1.3))

# Panel a: Slant range
sr_arr = [rd["slant_range_km"] for rd in results]
ax4a.plot(angles_arr, sr_arr, "ko-", linewidth=2, markersize=6)
ax4a.set_ylabel("Slant Range [km]", fontsize=10)
ax4a.set_title("Slant Range", fontsize=11)
ax4a.grid(True, alpha=0.3)

# Panel b: Air mass
am_arr = [rd["air_mass"] for rd in results]
ax4b.plot(angles_arr, am_arr, "ko-", linewidth=2, markersize=6)
ax4b.set_ylabel("Air Mass [--]", fontsize=10)
ax4b.set_title("Air Mass (sec θ)", fontsize=11)
ax4b.grid(True, alpha=0.3)

# Panel c: SNR
ax4c.plot(angles_arr, snr_arr, "bo-", linewidth=2, markersize=6)
ax4c.set_xlabel("Off-Nadir Angle [deg]", fontsize=10)
ax4c.set_ylabel("SNR [--]", fontsize=10)
ax4c.set_title("Signal-to-Noise Ratio", fontsize=11)
ax4c.grid(True, alpha=0.3)

# Panel d: GSD geometric mean
ax4d.plot(angles_arr, gsd_gm_arr, "go-", linewidth=2, markersize=6)
ax4d.set_xlabel("Off-Nadir Angle [deg]", fontsize=10)
ax4d.set_ylabel("GSD (geometric mean) [m]", fontsize=10)
ax4d.set_title("Ground Sample Distance", fontsize=11)
ax4d.grid(True, alpha=0.3)

fig4.suptitle("Off-Nadir Performance Summary — 600 km LEO, VNIR PAN", fontsize=13)
fig4.tight_layout()
fig4.savefig(PLOT_DIR / "fig4_summary_panels.png", dpi=150)

# ---------------------------------------------------------------------------
# Step 11: Output spreadsheet
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()

header_font_out = Font(bold=True, size=10, color="FFFFFF")
header_fill_out = PatternFill("solid", fgColor="2E75B6")
thin_border_out = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ws_out = wb_out.active
ws_out.title = "Off-Nadir Results"

headers_out = [
    "Angle [deg]", "Slant Range [km]", "Air Mass [--]", "Ground Range [km]",
    "Tau (band mean) [--]", "GSD Cross [m]", "GSD Along [m]", "GSD GM [m]",
    "SNR [--]", "MTF@Ny [--]", "RER [--]", "NIIRS (corrected) [--]",
    "NIIRS (RADIANT) [--]", "Signal [e-]",
]

for col_idx, h in enumerate(headers_out, start=1):
    cell = ws_out.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_out
    cell.fill = header_fill_out
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border_out

for row_idx, rd in enumerate(results, start=2):
    vals = [
        rd["angle_deg"], round(rd["slant_range_km"], 1), round(rd["air_mass"], 4),
        round(rd["ground_range_km"], 1), round(rd["tau_mean"], 4),
        round(rd["gsd_cross_m"], 2), round(rd["gsd_along_m"], 2),
        round(rd["gsd_gm_m"], 2), round(rd["snr"], 1),
        round(rd["mtf_nyq"], 4), round(rd["rer"], 4),
        round(rd["niirs_corrected"], 2), round(rd["niirs_radiant"], 2),
        round(rd["signal_e"], 0),
    ]
    for col_idx, v in enumerate(vals, start=1):
        cell = ws_out.cell(row=row_idx, column=col_idx, value=v)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border_out

for col in ws_out.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws_out.column_dimensions[col[0].column_letter].width = max_len + 3

wb_out.save(OUTPUT_FILE)
print(f"\n  Output spreadsheet: {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# Step 12: Summary
# ---------------------------------------------------------------------------

print(f"\n{'=' * 80}")
print(f"  SUMMARY")
print(f"{'=' * 80}")

print(f"\n  System: {aperture_m * 100:.0f} cm TMA, f/{f_number:.0f}, "
      f"{pixel_pitch_um:.0f} µm pixels, {filter_min_nm:.0f}-{filter_max_nm:.0f} nm PAN")
print(f"  Orbit:  {altitude_km:.0f} km SSO, GSD (nadir) = {gsd_nadir_m:.2f} [m]")
print(f"  Q:      {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")

print(f"\n  --- Nadir Baseline ---")
print(f"  Slant range:       {baseline['slant_range_km']:.1f} [km]")
print(f"  GSD (cross/along): {baseline['gsd_cross_m']:.2f} / {baseline['gsd_along_m']:.2f} [m]")
print(f"  Transmission:      {baseline['tau_mean']:.4f} [--]")
print(f"  SNR:               {baseline['snr']:.1f} [--]")
print(f"  NIIRS:             {baseline['niirs_corrected']:.2f} [--]")

print(f"\n  --- 30 deg Off-Nadir ---")
r30 = next((rd for rd in results if abs(rd["angle_deg"] - 30) < 0.1), None)
if r30:
    print(f"  Slant range:       {r30['slant_range_km']:.1f} [km] "
          f"(+{(r30['slant_range_km'] / baseline['slant_range_km'] - 1) * 100:.0f}%)")
    print(f"  GSD (cross/along): {r30['gsd_cross_m']:.2f} / {r30['gsd_along_m']:.2f} [m] "
          f"(+{(r30['gsd_cross_m'] / baseline['gsd_cross_m'] - 1) * 100:.0f}% / "
          f"+{(r30['gsd_along_m'] / baseline['gsd_along_m'] - 1) * 100:.0f}%)")
    print(f"  Transmission:      {r30['tau_mean']:.4f} [--] "
          f"({(r30['tau_mean'] / baseline['tau_mean'] - 1) * 100:+.1f}%)")
    print(f"  SNR:               {r30['snr']:.1f} [--] "
          f"({(r30['snr'] / baseline['snr'] - 1) * 100:+.1f}%)")
    print(f"  dNIIRS:            {r30['niirs_corrected'] - baseline['niirs_corrected']:+.2f} [--]")

print(f"\n  --- 45 deg Off-Nadir ---")
r45 = results[-1]
print(f"  Slant range:       {r45['slant_range_km']:.1f} [km] "
      f"(+{(r45['slant_range_km'] / baseline['slant_range_km'] - 1) * 100:.0f}%)")
print(f"  GSD (cross/along): {r45['gsd_cross_m']:.2f} / {r45['gsd_along_m']:.2f} [m] "
      f"(+{(r45['gsd_cross_m'] / baseline['gsd_cross_m'] - 1) * 100:.0f}% / "
      f"+{(r45['gsd_along_m'] / baseline['gsd_along_m'] - 1) * 100:.0f}%)")
print(f"  Transmission:      {r45['tau_mean']:.4f} [--] "
      f"({(r45['tau_mean'] / baseline['tau_mean'] - 1) * 100:+.1f}%)")
print(f"  SNR:               {r45['snr']:.1f} [--] "
      f"({(r45['snr'] / baseline['snr'] - 1) * 100:+.1f}%)")
print(f"  dNIIRS:            {r45['niirs_corrected'] - baseline['niirs_corrected']:+.2f} [--]")

print(f"\n  Key findings:")
print(f"    1. GSD is the dominant degradation driver at off-nadir angles.")
print(f"       Cross-track GSD scales as 1/cos(theta); along-track scales faster")
print(f"       due to the ground projection foreshortening effect.")
print(f"    2. Atmospheric transmission decreases with off-nadir angle because")
print(f"       the optical path through the atmosphere lengthens (air mass = sec(theta)).")
print(f"    3. NIIRS loss is dominated by the GSD term (-3.32 × log10(GSD_ratio)).")
print(f"       SNR degradation from lower transmission is secondary.")

print(f"\n  --- Newly Available Metrics (Gap Closures) ---")
print(f"  NEDT:              {baseline['nedt_mK']:.1f} [mK] (nadir)")
print(f"  NIIRS:             {baseline['niirs_corrected']:.2f} [--] (nadir, from RADIANT)")
print(f"  GSD (RADIANT):     {baseline['gsd_radiant_x']:.2f} [m] (nadir)")
print(f"  Q:                 {baseline.get('q_center', Q):.3f} [--]")
print(f"  Strehl:            {baseline['strehl']:.4f} [--]")
print(f"  RER:               {baseline['rer']:.4f} [--]")
print(f"  Well margin:       {baseline.get('well_margin_dB', 0.0):.1f} [dB]")

print(f"\n  Limitations:")
print(f"    - RADIANT GSD metric does not account for off-nadir angle (Gap 33)")
print(f"    - NIIRS corrected using GSD scaling only — does not re-run GIQE-5")
print(f"      with true off-nadir GSD (Gap 34)")
print(f"    - No along-track vs cross-track GSD distinction in RADIANT (Gap 35)")
print(f"    - No swath width or access geometry calculator (Gap 36)")
print(f"    - Earth curvature correction is applied to slant range but not")
print(f"      to RADIANT's internal atmospheric path calculation (it uses its own)")
