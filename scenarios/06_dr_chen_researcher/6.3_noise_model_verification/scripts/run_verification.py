"""Scenario 6.3: Noise Model Verification — Analytic vs. RADIANT.

Dr. Chen's workflow:
  1. Read sensor parameters from her Excel spreadsheet (non-RADIANT units)
  2. Convert to RADIANT canonical units (m, fractions, seconds)
  3. Run RADIANT evaluation
  4. Extract each noise term
  5. Compute hand-calc values for the major noise terms
  6. Compare in a formatted table with % error
  7. Report all performance metrics (SNR, NEDT, NIIRS, GSD, MTF, etc.)
  8. Write results back to the spreadsheet

Usage:
    python run_verification.py
"""

import math
import sys
from pathlib import Path

import numpy as np
import openpyxl

# ---------------------------------------------------------------------------
# Step 3: Run RADIANT — parameters set in VENDOR units (Gap 6 unit-aware set)
# ---------------------------------------------------------------------------
# The Step 2 conversions above feed the HAND-CALCULATION anchors only.
# RADIANT itself now receives the raw vendor values with their units;
# Sensor.set(..., unit=...) converts once at the parameter boundary (Rule 2),
# and Step 3a cross-checks RADIANT's conversion against the script's own.

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 5: Hand calculations for major noise terms
# ---------------------------------------------------------------------------
# We compute the terms Dr. Chen can verify analytically.

from radiant.core.constants import h, c, k_B


def planck_spectral_radiance(lam_m: np.ndarray, T: float) -> np.ndarray:
    """Planck function B(lam,T) in W/m^2/sr/um."""
    # Compute in SI (W/m^2/sr/m) then convert to per-um
    num = 2.0 * h * c ** 2 / lam_m ** 5
    denom = np.exp(h * c / (lam_m * k_B * T)) - 1.0
    return num / denom * 1e-6  # W/m^2/sr/m -> W/m^2/sr/um

# --- Reflected-solar term (Kirchhoff: ρ = 1 − ε) ---
# The no_atmosphere 'space' sub-case illuminates the target with the
# unattenuated TOA solar spectrum at the default solar zenith (0.5 rad).
# A grey (ε, T) target therefore also REFLECTS ρ·E_sun·cosθ/π (Lambertian)
# — ~9% of the in-band signal for ρ = 0.05 in the MWIR. The solar spectrum
# is shared input data (radiant.core.solar), not the physics under test.
from radiant.core.solar import toa_solar_spectral_irradiance

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ---------------------------------------------------------------------------
# Step 1: Read the spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "dr_chen_sensor_parameters.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb["Sensor Parameters"]

# Parse the parameter table into a dict keyed by parameter name
params_raw: dict[str, float] = {}
for row in ws.iter_rows(min_row=5, max_col=4, values_only=False):
    name_cell = row[0].value
    value_cell = row[1].value
    if name_cell and value_cell is not None:
        # Skip section headers (blue-filled rows)
        if row[0].fill and row[0].fill.start_color and row[0].fill.start_color.rgb == "004472C4":
            continue
        params_raw[name_cell] = float(value_cell)


# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

# Optics
aperture_m = params_raw["Aperture diameter"] / 100.0          # cm -> m
f_number = params_raw["Focal ratio (f/#)"]
focal_length_m = f_number * aperture_m                         # derived: f = f/# x D
transmission = params_raw["Optical transmission"] / 100.0      # % -> fraction
optics_temp_K = params_raw["Optics temperature"]

# Detector
pixel_pitch_m = params_raw["Pixel pitch"] * 1e-6              # um -> m
pixel_pitch_um = params_raw["Pixel pitch"]                     # keep for RADIANT input
qe = params_raw["Quantum efficiency"] / 100.0                 # % -> fraction
dark_rate = params_raw["Dark current"]                         # already e-/s
operating_temp = params_raw["Operating temperature"]           # already K
fwc = params_raw["Full well capacity"]                         # already e-

# Readout
read_noise = params_raw["Read noise"]                          # already e- RMS
adc_bits = int(params_raw["ADC bits"])
gain = params_raw["System gain"]                               # already e-/DN

# Scene
target_temp = params_raw["Target temperature"]                 # already K
target_emiss = params_raw["Target emissivity"]                 # dimensionless
bg_temp = params_raw["Background temperature"]                 # already K
bg_emiss = params_raw["Background emissivity"]                 # dimensionless

# Spectral
band_min = params_raw["Band minimum"]                          # already um
band_max = params_raw["Band maximum"]                          # already um
t_int = params_raw["Integration time"] / 1000.0                # ms -> s

# Geometry
altitude_m = params_raw["Sensor altitude"] * 1000.0            # km -> m


sensor = Sensor()
# Vendor/lab units — converted by the unit-aware boundary (Gap 6):
sensor.set("optics.aperture_diameter_m", params_raw["Aperture diameter"], unit="cm")
sensor.set("optics.transmission_scalar", params_raw["Optical transmission"], unit="%")
sensor.set("detector.qe_value", params_raw["Quantum efficiency"], unit="%")
sensor.set("spectral_integration.integration_time_s",
           params_raw["Integration time"], unit="ms")
sensor.set("geometry.sensor_altitude_m", params_raw["Sensor altitude"], unit="km")
# Stage-7 precondition: exo routes through the no_atmosphere 'space'
# sub-case, whose Earth-limb check needs the sensor altitude. This sensor
# has a genuine platform altitude (8 km — the exo model treats the path as
# vacuum), so h_sensor is the real value, not a placeholder (unlike the
# 7.x bench scenarios — registry Gap 42).
# Already-canonical values — set without a unit tag:
sensor.set("optics.focal_length_m", focal_length_m)  # derived: f/# × D
sensor.set("optics.optics_temperature_K", optics_temp_K)
sensor.set("detector.pixel_pitch_x_um", pixel_pitch_um)
sensor.set("detector.pixel_pitch_y_um", pixel_pitch_um)
sensor.set("detector.dark_rate_e_per_s", dark_rate)
sensor.set("detector.detector_temperature_K", operating_temp)
sensor.set("source.target.temperature", target_temp)
sensor.set("source.target.emissivity", target_emiss)
sensor.set("source.background.temperature", bg_temp)
sensor.set("source.background.emissivity", bg_emiss)
sensor.set("atmosphere.model", "exo")  # Above the atmosphere — vacuum path
sensor.set("spectral_integration.filter_min_um", band_min)
sensor.set("spectral_integration.filter_max_um", band_max)
sensor.set("readout.read_noise_e_rms", read_noise)
sensor.set("readout.gain_e_per_dn", gain)
sensor.set("readout.adc_bits", adc_bits)
sensor.set("readout.full_well_capacity_e", fwc)
sensor.set("performance.niirs.allow_extrapolated", True)


def main() -> None:
    """Run the scenario analysis."""
    OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "verification_results.xlsx"

    print("=" * 95)
    print("  SCENARIO 6.3: Noise Model Verification — Analytic vs. RADIANT")
    print("=" * 95)

    print("\n=== Step 1: Raw parameters from spreadsheet ===")
    print(f"  {'Parameter':<30s} {'Value':>12s}  {'Unit'}")
    print(f"  {'-' * 30} {'-' * 12}  {'-' * 10}")
    print(f"  {'Aperture diameter':<30s} {params_raw['Aperture diameter']:>12.1f}  cm")
    print(f"  {'Focal ratio (f/#)':<30s} {params_raw['Focal ratio (f/#)']:>12.1f}  —")
    print(f"  {'Optical transmission':<30s} {params_raw['Optical transmission']:>12.1f}  %")
    print(f"  {'Number of optical elements':<30s} {params_raw['Number of optical elements']:>12.0f}  —")
    print(f"  {'Optics temperature':<30s} {params_raw['Optics temperature']:>12.1f}  K")
    print(f"  {'Pixel pitch':<30s} {params_raw['Pixel pitch']:>12.1f}  \u00b5m")
    print(f"  {'Quantum efficiency':<30s} {params_raw['Quantum efficiency']:>12.1f}  %")
    print(f"  {'Dark current':<30s} {params_raw['Dark current']:>12.1f}  e\u207b/s")
    print(f"  {'Operating temperature':<30s} {params_raw['Operating temperature']:>12.1f}  K")
    print(f"  {'Full well capacity':<30s} {params_raw['Full well capacity']:>12.0f}  e\u207b")
    print(f"  {'Read noise':<30s} {params_raw['Read noise']:>12.1f}  e\u207b RMS")
    print(f"  {'ADC bits':<30s} {params_raw['ADC bits']:>12.0f}  bits")
    print(f"  {'System gain':<30s} {params_raw['System gain']:>12.1f}  e\u207b/DN")
    print(f"  {'Target temperature':<30s} {params_raw['Target temperature']:>12.1f}  K")
    print(f"  {'Target emissivity':<30s} {params_raw['Target emissivity']:>12.2f}  —")
    print(f"  {'Background temperature':<30s} {params_raw['Background temperature']:>12.1f}  K       * contrast SNR only")
    print(f"  {'Background emissivity':<30s} {params_raw['Background emissivity']:>12.2f}  —       * contrast SNR only")
    print(f"  {'Band minimum':<30s} {params_raw['Band minimum']:>12.1f}  \u00b5m")
    print(f"  {'Band maximum':<30s} {params_raw['Band maximum']:>12.1f}  \u00b5m")
    print(f"  {'Integration time':<30s} {params_raw['Integration time']:>12.1f}  ms")
    print(f"  {'Number of TDI stages':<30s} {params_raw['Number of TDI stages']:>12.0f}  —")
    print(f"  {'Sensor altitude':<30s} {params_raw['Sensor altitude']:>12.1f}  km")
    print(f"  {'Look angle':<30s} {params_raw['Look angle']:>12.1f}  deg")


    print("\n=== Step 2: Converted to RADIANT canonical units ===")
    print(f"  {'Parameter':<30s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
    print(f"  {'-' * 30} {'-' * 14}  {'-' * 15}  {'-' * 20}")
    print(f"  {'Aperture diameter':<30s} {aperture_m:>14.4f}  {'m':<15s}  cm / 100")
    print(f"  {'Focal length (derived)':<30s} {focal_length_m:>14.4f}  {'m':<15s}  f/# x D")
    print(f"  {'f-number':<30s} {f_number:>14.1f}  {'--':<15s}  no conversion")
    print(f"  {'Optical transmission':<30s} {transmission:>14.4f}  {'fraction':<15s}  % / 100")
    print(f"  {'Optics temperature':<30s} {optics_temp_K:>14.1f}  {'K':<15s}  no conversion")
    print(f"  {'Pixel pitch':<30s} {pixel_pitch_um:>14.1f}  {'\u00b5m':<15s}  no conversion")
    print(f"  {'Pixel pitch':<30s} {pixel_pitch_m:>14.2e}  {'m':<15s}  \u00b5m x 1e-6")
    print(f"  {'Quantum efficiency':<30s} {qe:>14.4f}  {'fraction':<15s}  % / 100")
    print(f"  {'Dark current':<30s} {dark_rate:>14.1f}  {'e\u207b/s':<15s}  no conversion")
    print(f"  {'Detector temperature':<30s} {operating_temp:>14.1f}  {'K':<15s}  no conversion")
    print(f"  {'Full well capacity':<30s} {fwc:>14.0f}  {'e\u207b':<15s}  no conversion")
    print(f"  {'Read noise':<30s} {read_noise:>14.1f}  {'e\u207b RMS':<15s}  no conversion")
    print(f"  {'ADC bits':<30s} {adc_bits:>14d}  {'bits':<15s}  no conversion")
    print(f"  {'System gain':<30s} {gain:>14.1f}  {'e\u207b/DN':<15s}  no conversion")
    print(f"  {'Target temperature':<30s} {target_temp:>14.1f}  {'K':<15s}  no conversion")
    print(f"  {'Target emissivity':<30s} {target_emiss:>14.2f}  {'--':<15s}  dimensionless")
    print(f"  {'Background temperature':<30s} {bg_temp:>14.1f}  {'K':<15s}  * contrast SNR only")
    print(f"  {'Background emissivity':<30s} {bg_emiss:>14.2f}  {'--':<15s}  * contrast SNR only")
    print(f"  {'Band minimum':<30s} {band_min:>14.1f}  {'\u00b5m':<15s}  no conversion")
    print(f"  {'Band maximum':<30s} {band_max:>14.1f}  {'\u00b5m':<15s}  no conversion")
    print(f"  {'Integration time':<30s} {t_int:>14.6f}  {'s':<15s}  ms / 1000")
    print(f"  {'Sensor altitude':<30s} {altitude_m:>14.1f}  {'m':<15s}  km x 1000")


    # --- Step 3a: cross-check RADIANT's boundary conversion vs the script's ---
    print("\n=== Step 3a: Unit-aware boundary conversion cross-check (Gap 6) ===")
    print(f"  {'Parameter':<38s} {'Vendor value':>14s}  {'RADIANT':>14s}  {'Script':>14s}  {'Match'}")
    print(f"  {'-' * 38} {'-' * 14}  {'-' * 14}  {'-' * 14}  {'-' * 5}")
    _checks = [
        ("optics.aperture_diameter_m", f"{params_raw['Aperture diameter']:.1f} cm", aperture_m),
        ("optics.transmission_scalar", f"{params_raw['Optical transmission']:.1f} %", transmission),
        ("detector.qe_value", f"{params_raw['Quantum efficiency']:.1f} %", qe),
        ("spectral_integration.integration_time_s",
         f"{params_raw['Integration time']:.1f} ms", t_int),
        ("geometry.sensor_altitude_m", f"{params_raw['Sensor altitude']:.1f} km", altitude_m),
    ]
    for pname, vendor_str, script_val in _checks:
        radiant_val = float(sensor.get(pname))
        ok = abs(radiant_val - script_val) <= 1e-12 * max(1.0, abs(script_val))
        print(f"  {pname:<38s} {vendor_str:>14s}  {radiant_val:>14.6g}  "
              f"{script_val:>14.6g}  {'OK' if ok else 'MISMATCH'}")
        if not ok:
            raise ValueError(
                f"Unit-aware set() disagrees with the script conversion for "
                f"{pname}: RADIANT={radiant_val!r}, script={script_val!r}"
            )

    print("\n=== Step 3: Running RADIANT evaluation ===")
    result = sensor.evaluate()

    # ---------------------------------------------------------------------------
    # Step 3b: Regime and signal chain notes
    # ---------------------------------------------------------------------------

    regime = result.stage_outputs["optics"]["regime"]
    print(f"\n=== Radiometric Regime ===")
    print(f"  Regime:  {regime}")
    print()
    print(f"  This scenario uses EXTENDED regime (fill_fraction = 1.0, no target")
    print(f"  geometry specified). The pixel is entirely filled by the target.")
    print()
    print(f"  In extended regime:")
    print(f"    - Signal = L_target x tau_atm x tau_opt x A_collect x Omega_pixel x QE x t_int / E_photon")
    print(f"    - The {target_temp:.0f} K target fills the pixel. No background mixing occurs.")
    print(f"    - Background temperature ({bg_temp:.0f} K) and emissivity ({bg_emiss}) are used")
    print(f"      ONLY for contrast SNR (target-minus-background divided by noise).")
    print()
    print(f"  Nearfield emission note:")
    print(f"    In scalar transmission mode, RADIANT models the lumped optical train as a")
    print(f"    refractive element where loss (1 - tau) is treated as reflection, not absorption.")
    print(f"    By Kirchhoff's law: T + R = 1, so emissivity eps = 1 - T - R = 0.")
    print(f"    This is physically correct for lenses/windows. For reflective systems")
    print(f"    (mirrors with eps = 1 - R), use key_elements or full_prescription mode")
    print(f"    to specify individual elements with reflectances.")
    print(f"    Nearfield_shot = 0 in this configuration is by design.")

    # ---------------------------------------------------------------------------
    # Step 4: Extract RADIANT noise terms
    # ---------------------------------------------------------------------------

    radiant_noise: dict[str, float] = {}
    for nt in result.noise_terms:
        radiant_noise[nt.name] = nt.value_e

    print("\n=== Step 4: RADIANT noise terms ===")
    print(f"  {'Term':<25s} {'Value':>12s}  {'Unit'}")
    print(f"  {'-' * 25} {'-' * 12}  {'-' * 10}")
    for name, value in sorted(radiant_noise.items(), key=lambda x: -x[1]):
        print(f"  {name:<25s} {value:>12.4f}  e\u207b RMS")

    # -- Derived quantities --
    pixel_area_m2 = pixel_pitch_m ** 2
    f_num = f_number
    omega_pixel = math.pi / (4.0 * f_num ** 2)  # pixel solid angle [sr]

    # -- Planck spectral radiance integral --
    # L_target = integral of eps * B(lam,T) * dlam over the band
    # B(lam,T) = 2hc^2/lam^5 * 1/(exp(hc/lam*kT) - 1)  [W/m^2/sr/um]
    N_WAVE = 1000
    wavelengths_um = np.linspace(band_min, band_max, N_WAVE)
    wavelengths_m = wavelengths_um * 1e-6


    L_target_spectral = target_emiss * planck_spectral_radiance(wavelengths_m, target_temp)
    L_bg_spectral = bg_emiss * planck_spectral_radiance(wavelengths_m, bg_temp)

    # Integrate over band [W/m^2/sr]
    L_target_band = float(np.trapezoid(L_target_spectral, wavelengths_um))
    L_bg_band = float(np.trapezoid(L_bg_spectral, wavelengths_um))

    # Mean photon energy over the band (textbook first-pass reference only)
    lam_center_m = np.mean(wavelengths_m)
    E_photon = h * c / lam_center_m  # [J]

    # --- Signal electrons: PHOTON-WEIGHTED spectral integral (exact) ---
    # Photon flux = ∫ L(λ)·λ/(hc) dλ. The band-center shortcut
    # (L_band / E_photon) underestimates by ~5% here because a 300 K source
    # emits its in-band photons preferentially at the long end of 3.5–5 µm.
    phot_target = float(np.trapezoid(L_target_spectral * wavelengths_m / (h * c), wavelengths_um))
    phot_bg = float(np.trapezoid(L_bg_spectral * wavelengths_m / (h * c), wavelengths_um))

    signal_e_thermal = phot_target * transmission * omega_pixel * pixel_area_m2 * qe * t_int
    signal_e_center_approx = (
        L_target_band * transmission * omega_pixel * pixel_area_m2 * qe * t_int / E_photon
    )
    bg_e_hand = phot_bg * transmission * omega_pixel * pixel_area_m2 * qe * t_int

    SOLAR_ZENITH_RAD = 0.5  # geometry.solar_zenith_rad default
    rho_target = 1.0 - target_emiss
    E_sun_spectral = np.asarray(toa_solar_spectral_irradiance(wavelengths_um))  # W/m²/µm
    L_refl_spectral = rho_target * E_sun_spectral * math.cos(SOLAR_ZENITH_RAD) / math.pi
    phot_refl = float(np.trapezoid(L_refl_spectral * wavelengths_m / (h * c), wavelengths_um))
    signal_e_refl = phot_refl * transmission * omega_pixel * pixel_area_m2 * qe * t_int

    signal_e_hand = signal_e_thermal + signal_e_refl

    # Dark electrons
    dark_e_hand = dark_rate * t_int

    # Noise terms (hand calc)
    # Notes:
    #  - nearfield_shot is correctly zero in scalar transmission mode (eps = 0
    #    unless optics.scalar_emissivity is set).
    #  - background_shot is ZERO BY DESIGN in the extended regime: the target
    #    fills the pixel IFOV, so there is no separate scene-background photon
    #    stream (matrix Decision #13). The background temperature/emissivity
    #    inputs feed only the contrast-SNR/NEDT scene definition. bg_e_hand is
    #    retained above as a reference number only.
    hand_calc: dict[str, float] = {
        "signal_shot": math.sqrt(signal_e_hand),
        "background_shot": 0.0,  # extended regime — no separate background stream
        "nearfield_shot": 0.0,  # eps = 0 for scalar refractive lumped element
        "dark_shot": math.sqrt(dark_e_hand),
        "read_noise": read_noise,
        "quantization": gain / math.sqrt(12.0),
    }

    print("\n=== Step 5: Hand-calculated noise terms ===")
    print(f"  {'Term':<25s} {'Value':>12s}  {'Unit'}")
    print(f"  {'-' * 25} {'-' * 12}  {'-' * 10}")
    for name, value in hand_calc.items():
        print(f"  {name:<25s} {value:>12.4f}  e\u207b RMS")

    print("\n=== Derived quantities ===")
    print(f"  {'Quantity':<30s} {'Value':>14s}  {'Unit'}")
    print(f"  {'-' * 30} {'-' * 14}  {'-' * 15}")
    print(f"  {'Focal length':<30s} {focal_length_m:>14.4f}  m")
    print(f"  {'Pixel solid angle (Omega)':<30s} {omega_pixel:>14.6e}  sr")
    print(f"  {'Pixel area':<30s} {pixel_area_m2:>14.6e}  m\u00b2")
    print(f"  {'L_target (band-integrated)':<30s} {L_target_band:>14.6f}  W/m\u00b2/sr")
    print(f"  {'L_background (band-integ.)':<30s} {L_bg_band:>14.6f}  W/m\u00b2/sr")
    print(f"  {'E_photon (band center)':<30s} {E_photon:>14.6e}  J/photon")
    print(f"  {'Signal e\u207b (thermal, photon integral)':<38s} {signal_e_thermal:>14.2f}  e\u207b")
    print(f"  {'Signal e\u207b (band-center approx.)':<38s} {signal_e_center_approx:>14.2f}  e\u207b  "
          f"({(signal_e_center_approx / signal_e_thermal - 1) * 100:+.1f}% vs integral)")
    print(f"  {'Signal e\u207b (reflected solar, \u03c1={:.2f})'.format(rho_target):<38s} "
          f"{signal_e_refl:>14.2f}  e\u207b")
    print(f"  {'Signal e\u207b (total hand)':<38s} {signal_e_hand:>14.2f}  e\u207b")
    print(f"  {'Background e\u207b (reference only \u2014 not a':<38s}")
    print(f"  {'  photon term in extended regime)':<38s} {bg_e_hand:>14.2f}  e\u207b")
    print(f"  {'Dark electrons':<38s} {dark_e_hand:>14.4f}  e\u207b")

    # ---------------------------------------------------------------------------
    # Step 6: Comparison table
    # ---------------------------------------------------------------------------

    print("\n" + "=" * 95)
    print("  NOISE TERM COMPARISON: Hand Calculation vs. RADIANT")
    print("=" * 95)
    print(f"  {'Noise Term':<23s} | {'Hand Calc':>12s} | {'RADIANT':>12s} | {'Unit':<10s} | {'% Error':>8s} | {'Status'}")
    print("-" * 95)

    comparison_rows: list[dict] = []
    for name in hand_calc:
        hc_val = hand_calc[name]
        rd_val = radiant_noise.get(name, 0.0)
        if hc_val > 0:
            pct_err = abs(hc_val - rd_val) / hc_val * 100.0
        elif rd_val == 0.0:
            pct_err = 0.0
        else:
            pct_err = float("inf")
        status = "PASS" if pct_err < 5.0 else "CHECK" if pct_err < 20.0 else "FAIL"
        print(f"  {name:<23s} | {hc_val:>12.4f} | {rd_val:>12.4f} | {'e\u207b RMS':<10s} | {pct_err:>7.2f}% | {status}")
        comparison_rows.append({
            "name": name, "hand_calc": hc_val, "radiant": rd_val,
            "pct_error": pct_err, "status": status,
        })

    # Also show terms that RADIANT computes but we didn't hand-calc
    print("-" * 95)
    print("  Terms not hand-calculated (shown for reference):")
    for name, rd_val in sorted(radiant_noise.items(), key=lambda x: -x[1]):
        if name not in hand_calc:
            print(f"  {name:<23s} |     {'\u2014':>8s} | {rd_val:>12.4f} | {'e\u207b RMS':<10s} |    {'\u2014':>4s} |")

    # Total noise comparison
    hand_total = math.sqrt(sum(v ** 2 for v in hand_calc.values()))
    radiant_total_temporal = math.sqrt(
        sum(v ** 2 for k, v in radiant_noise.items()
            if k in {"signal_shot", "background_shot", "nearfield_shot", "straylight_shot",
                      "dark_shot", "gr_noise", "johnson_noise", "flicker_1f",
                      "read_noise", "ktc_reset", "quantization", "persistence_noise", "glow_shot"})
    )

    total_pct_err = abs(hand_total - radiant_total_temporal) / hand_total * 100.0
    total_status = "PASS" if total_pct_err < 5.0 else "CHECK" if total_pct_err < 20.0 else "FAIL"
    print("-" * 95)
    print(f"  {'TOTAL (RSS)':<23s} | {hand_total:>12.4f} | {radiant_total_temporal:>12.4f} | {'e\u207b RMS':<10s} | {total_pct_err:>7.2f}% | {total_status}")
    print("=" * 95)

    # ---------------------------------------------------------------------------
    # Step 7: Performance metrics — full report
    # ---------------------------------------------------------------------------

    snr = result.metrics["snr"]
    contrast_snr = result.metrics.get("contrast_snr", None)
    mtf_nyq = result.metrics.get("mtf_at_nyquist", None)
    nedt_K = result.metrics.get("nedt_K", None)
    niirs = result.metrics.get("niirs", None)
    gsd_ct = result.metrics.get("gsd_cross_track_m", None)
    gsd_at = result.metrics.get("gsd_along_track_m", None)
    gsd_geo = result.metrics.get("gsd_geometric_mean_m", None)
    strehl = result.metrics.get("strehl", None)
    q_center = result.metrics.get("q_center", None)
    ee_1x1 = result.metrics.get("ee_1x1", None)
    ee_3x3 = result.metrics.get("ee_3x3", None)
    rer = result.metrics.get("rer", None)
    well_margin = result.metrics.get("well_margin_dB", None)
    dynamic_range = result.metrics.get("dynamic_range_dB", None)

    # Hand-calc SNR for comparison
    hand_snr = signal_e_hand / hand_total
    snr_err = abs(hand_snr - snr) / hand_snr * 100.0 if hand_snr > 0 else 0.0

    # Hand-calc NEDT for comparison (if RADIANT reports it)
    # NEDT = sigma_total / (dS/dT), where dS/dT is the derivative of signal w.r.t. target temp.
    # Approximate dS/dT by finite difference at target_temp
    if nedt_K is not None:
        delta_T = 0.1  # K
        L_plus = target_emiss * planck_spectral_radiance(wavelengths_m, target_temp + delta_T)
        L_minus = target_emiss * planck_spectral_radiance(wavelengths_m, target_temp - delta_T)
        dL_dT = (L_plus - L_minus) / (2.0 * delta_T)  # W/m^2/sr/um per K
        # Photon-weighted derivative (same integral as the signal term; the
        # reflected-solar component does not depend on target temperature).
        dphot_dT = float(np.trapezoid(dL_dT * wavelengths_m / (h * c), wavelengths_um))
        dS_dT = dphot_dT * transmission * omega_pixel * pixel_area_m2 * qe * t_int  # e-/K
        hand_nedt = hand_total / dS_dT  # K
        nedt_err = abs(hand_nedt - nedt_K) / hand_nedt * 100.0 if hand_nedt > 0 else 0.0

    print(f"\n=== Step 7: Performance Metrics ===")
    print(f"  {'Metric':<30s} {'RADIANT':>12s}  {'Hand Calc':>12s}  {'Unit':<15s}  {'% Error':>8s}")
    print(f"  {'-' * 30} {'-' * 12}  {'-' * 12}  {'-' * 15}  {'-' * 8}")

    print(f"  {'SNR':<30s} {snr:>12.2f}  {hand_snr:>12.2f}  {'--':<15s}  {snr_err:>7.2f}%")

    if contrast_snr is not None:
        print(f"  {'Contrast SNR':<30s} {contrast_snr:>12.2f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if nedt_K is not None:
        print(f"  {'NEDT':<30s} {nedt_K * 1000:>12.2f}  {hand_nedt * 1000:>12.2f}  {'mK':<15s}  {nedt_err:>7.2f}%")
        print(f"    NEDT difference explained: RADIANT's performance stage uses the")
        print(f"    single-wavelength Planck-factor approximation")
        print(f"    NEDT = T / (SNR · x·eˣ/(eˣ−1)) at the band-effective wavelength")
        print(f"    (nedt.compute_nedt_from_snr), while the hand value uses the exact")
        print(f"    band-integrated photon-weighted derivative dS/dT. The dominant")
        print(f"    bias: the SNR numerator includes the reflected-solar signal (~9%),")
        print(f"    which does NOT vary with target temperature — inflated SNR →")
        print(f"    RADIANT's NEDT reads LOW (optimistic thermal sensitivity).")
        print(f"    An exact-dS/dT chain path exists (nedt.compute_nedt) but is not")
        print(f"    wired to the stage — recorded in gaps.md (registry Gap 43).")

    if niirs is not None:
        print(f"  {'NIIRS':<30s} {niirs:>12.2f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if gsd_geo is not None:
        # Hand-calc GSD: p * H / f
        hand_gsd = pixel_pitch_m * altitude_m / focal_length_m
        gsd_err = abs(hand_gsd - gsd_geo) / hand_gsd * 100.0 if hand_gsd > 0 else 0.0
        print(f"  {'GSD (geometric mean)':<30s} {gsd_geo:>12.4f}  {hand_gsd:>12.4f}  {'m':<15s}  {gsd_err:>7.2f}%")

    if gsd_ct is not None and gsd_at is not None:
        print(f"  {'GSD (cross-track)':<30s} {gsd_ct:>12.4f}  {'\u2014':>12s}  {'m':<15s}  {'\u2014':>8s}")
        print(f"  {'GSD (along-track)':<30s} {gsd_at:>12.4f}  {'\u2014':>12s}  {'m':<15s}  {'\u2014':>8s}")

    if mtf_nyq is not None:
        print(f"  {'MTF at Nyquist':<30s} {mtf_nyq:>12.4f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if strehl is not None:
        print(f"  {'Strehl ratio':<30s} {strehl:>12.4f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if q_center is not None:
        # Hand-calc Q: lam_center * f/# / pitch
        lam_center_um = (band_min + band_max) / 2.0
        hand_q = lam_center_um * 1e-6 * f_number / pixel_pitch_m
        q_err = abs(hand_q - q_center) / hand_q * 100.0 if hand_q > 0 else 0.0
        print(f"  {'Q (sampling parameter)':<30s} {q_center:>12.4f}  {hand_q:>12.4f}  {'--':<15s}  {q_err:>7.2f}%")

    if ee_1x1 is not None:
        print(f"  {'EE (1x1 pixel)':<30s} {ee_1x1:>12.4f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if ee_3x3 is not None:
        print(f"  {'EE (3x3 pixel)':<30s} {ee_3x3:>12.4f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if rer is not None:
        print(f"  {'RER':<30s} {rer:>12.4f}  {'\u2014':>12s}  {'--':<15s}  {'\u2014':>8s}")

    if well_margin is not None:
        print(f"  {'Well margin':<30s} {well_margin:>12.2f}  {'\u2014':>12s}  {'dB':<15s}  {'\u2014':>8s}")

    if dynamic_range is not None:
        print(f"  {'Dynamic range':<30s} {dynamic_range:>12.2f}  {'\u2014':>12s}  {'dB':<15s}  {'\u2014':>8s}")

    # ---------------------------------------------------------------------------
    # Step 7b: MTF budget breakdown
    # ---------------------------------------------------------------------------

    mtf_budget = result.stage_outputs.get("performance", {}).get("mtf_budget", None)
    if mtf_budget is not None:
        print(f"\n=== MTF Budget (at Nyquist) ===")
        print(f"  {'Component':<30s} {'MTF_x':>10s}  {'MTF_y':>10s}")
        print(f"  {'-' * 30} {'-' * 10}  {'-' * 10}")
        per_term = mtf_budget.per_term_at_nyquist
        # Group x/y pairs by stripping axis suffix
        seen: set[str] = set()
        for key in per_term:
            base = key.rsplit("_", 1)[0]  # e.g., "mtf_optics"
            if base in seen:
                continue
            seen.add(base)
            val_x = per_term.get(f"{base}_x", 1.0)
            val_y = per_term.get(f"{base}_y", 1.0)
            label = base.replace("mtf_", "").replace("_", " ").title()
            print(f"  {label:<30s} {val_x:>10.4f}  {val_y:>10.4f}")
        print(f"  {'-' * 30} {'-' * 10}  {'-' * 10}")
        print(f"  {'SYSTEM MTF':<30s} {mtf_budget.system_mtf_at_nyquist_x:>10.4f}  {mtf_budget.system_mtf_at_nyquist_y:>10.4f}")

    print(f"\n=== Physics Notes ===")
    print(f"  1. signal_shot now matches to <0.01%: the hand model integrates photons")
    print(f"     spectrally (∫L·λ/hc dλ, not the band-center E_photon shortcut) AND")
    print(f"     includes the Kirchhoff reflected-solar term ρ·E_TOA·cosθ/π that the")
    print(f"     'space' sub-case adds to a grey (ε, T) target — ~9% of the in-band")
    print(f"     signal here. A textbook thermal-only model reads ~9% low, not because")
    print(f"     RADIANT is wrong but because the scene includes daytime illumination.")
    print(f"  2. background_shot = 0 BY DESIGN in the extended regime: the target")
    print(f"     fills the pixel IFOV, so there is no separate scene-background photon")
    print(f"     stream (matrix Decision #13). The background temperature/emissivity")
    print(f"     inputs define the contrast scene only.")
    print(f"  3. Deterministic terms (dark_shot, read_noise, quantization) match exactly.")
    print(f"  4. nearfield_shot = 0 because scalar transmission mode models the optical")
    print(f"     train as a refractive element (eps = 0 unless optics.scalar_emissivity")
    print(f"     is set — Gap 37). To model mirror self-emission, set scalar_emissivity")
    print(f"     or use key_elements / full_prescription mode.")
    print(f"  5. NEDT differs ~13%: RADIANT uses the single-λ Planck-factor")
    print(f"     approximation (compute_nedt_from_snr); the exact band-integrated")
    print(f"     dS/dT path (compute_nedt) exists but is not wired to the stage.")

    # ---------------------------------------------------------------------------
    # Step 8: Write results to output spreadsheet
    # ---------------------------------------------------------------------------

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Noise Comparison"

    header_font = Font(bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    check_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_out.column_dimensions["A"].width = 28
    ws_out.column_dimensions["B"].width = 22
    ws_out.column_dimensions["C"].width = 22
    ws_out.column_dimensions["D"].width = 14
    ws_out.column_dimensions["E"].width = 14
    ws_out.column_dimensions["F"].width = 10

    ws_out["A1"] = "Scenario 6.3: Noise Model Verification"
    ws_out["A1"].font = Font(bold=True, size=14)

    headers = ["Noise Term", "Hand Calc [e- RMS]", "RADIANT [e- RMS]", "Unit", "% Error", "Status"]
    for col, h_text in enumerate(headers, 1):
        cell = ws_out.cell(row=3, column=col, value=h_text)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, row_data in enumerate(comparison_rows, 4):
        ws_out.cell(row=i, column=1, value=row_data["name"]).border = thin_border
        ws_out.cell(row=i, column=2, value=round(row_data["hand_calc"], 4)).border = thin_border
        ws_out.cell(row=i, column=3, value=round(row_data["radiant"], 4)).border = thin_border
        ws_out.cell(row=i, column=4, value="e- RMS").border = thin_border
        ws_out.cell(row=i, column=5, value=round(row_data["pct_error"], 2)).border = thin_border
        ws_out.cell(row=i, column=5).number_format = "0.00"
        status_cell = ws_out.cell(row=i, column=6, value=row_data["status"])
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        if row_data["status"] == "PASS":
            status_cell.fill = pass_fill
        elif row_data["status"] == "CHECK":
            status_cell.fill = check_fill
        else:
            status_cell.fill = fail_fill

    # Total row
    total_row = len(comparison_rows) + 4
    ws_out.cell(row=total_row, column=1, value="TOTAL (RSS)").border = thin_border
    ws_out.cell(row=total_row, column=1).font = header_font
    ws_out.cell(row=total_row, column=2, value=round(hand_total, 4)).border = thin_border
    ws_out.cell(row=total_row, column=3, value=round(radiant_total_temporal, 4)).border = thin_border
    ws_out.cell(row=total_row, column=4, value="e- RMS").border = thin_border
    ws_out.cell(row=total_row, column=5, value=round(total_pct_err, 2)).border = thin_border

    # Summary section — performance metrics
    summary_row = total_row + 2
    ws_out.cell(row=summary_row, column=1, value="Performance Metrics").font = Font(bold=True, size=12)

    metrics_headers = ["Metric", "RADIANT", "Hand Calc", "Unit", "% Error"]
    for col, h_text in enumerate(metrics_headers, 1):
        cell = ws_out.cell(row=summary_row + 1, column=col, value=h_text)
        cell.font = header_font
        cell.border = thin_border

    r = summary_row + 2
    ws_out.cell(row=r, column=1, value="SNR").border = thin_border
    ws_out.cell(row=r, column=2, value=round(snr, 2)).border = thin_border
    ws_out.cell(row=r, column=3, value=round(hand_snr, 2)).border = thin_border
    ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border
    ws_out.cell(row=r, column=5, value=round(snr_err, 2)).border = thin_border

    if contrast_snr is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="Contrast SNR").border = thin_border
        ws_out.cell(row=r, column=2, value=round(contrast_snr, 2)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border

    if nedt_K is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="NEDT").border = thin_border
        ws_out.cell(row=r, column=2, value=round(nedt_K * 1000, 2)).border = thin_border
        ws_out.cell(row=r, column=3, value=round(hand_nedt * 1000, 2)).border = thin_border
        ws_out.cell(row=r, column=4, value="mK").border = thin_border
        ws_out.cell(row=r, column=5, value=round(nedt_err, 2)).border = thin_border

    if niirs is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="NIIRS").border = thin_border
        ws_out.cell(row=r, column=2, value=round(niirs, 2)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border

    if gsd_geo is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="GSD (geometric mean)").border = thin_border
        ws_out.cell(row=r, column=2, value=round(gsd_geo, 4)).border = thin_border
        ws_out.cell(row=r, column=3, value=round(hand_gsd, 4)).border = thin_border
        ws_out.cell(row=r, column=4, value="m").border = thin_border
        ws_out.cell(row=r, column=5, value=round(gsd_err, 2)).border = thin_border

    if mtf_nyq is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="MTF at Nyquist").border = thin_border
        ws_out.cell(row=r, column=2, value=round(mtf_nyq, 4)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border

    if strehl is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="Strehl ratio").border = thin_border
        ws_out.cell(row=r, column=2, value=round(strehl, 4)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border

    if q_center is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="Q (sampling parameter)").border = thin_border
        ws_out.cell(row=r, column=2, value=round(q_center, 4)).border = thin_border
        ws_out.cell(row=r, column=3, value=round(hand_q, 4)).border = thin_border
        ws_out.cell(row=r, column=4, value="-- (dimensionless)").border = thin_border
        ws_out.cell(row=r, column=5, value=round(q_err, 2)).border = thin_border

    if ee_1x1 is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="EE (1x1 pixel)").border = thin_border
        ws_out.cell(row=r, column=2, value=round(ee_1x1, 4)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="-- (fraction)").border = thin_border

    if well_margin is not None:
        r += 1
        ws_out.cell(row=r, column=1, value="Well margin").border = thin_border
        ws_out.cell(row=r, column=2, value=round(well_margin, 2)).border = thin_border
        ws_out.cell(row=r, column=3, value="--").border = thin_border
        ws_out.cell(row=r, column=4, value="dB").border = thin_border

    wb_out.save(OUTPUT_FILE)
    print(f"\nResults written to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
