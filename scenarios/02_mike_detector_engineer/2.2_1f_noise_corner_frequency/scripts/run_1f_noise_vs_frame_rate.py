"""Scenario 2.2: 1/f Noise Corner Frequency Impact on LWIR Staring Array

Mike has a 640x512 LWIR HgCdTe staring array with measured 1/f noise
(K = 2.5e4 e⁻², corner frequency 200 Hz).  He wants to know how 1/f noise
impacts NEDT at three frame rates: 30, 60, and 120 Hz.

This script:
  1. Reads Mike's detector spreadsheet (µm, %, ms, Hz)
  2. Converts to RADIANT canonical units
  3. Computes analytic 1/f noise at each frame rate: σ = √(K·ln(f_high/f_low))
  4. Runs RADIANT with and without 1/f at each frame rate
  5. Breaks down noise budget and NEDT contribution
  6. Sweeps f_low from 1 Hz to 500 Hz to show full 1/f impact curve
  7. Writes results to an output spreadsheet

Usage:
    python run_1f_noise_vs_frame_rate.py
"""

import copy
import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Mike's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "mike_1f_noise_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "1f_noise_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

def _read_sheet(ws, min_row: int = 5) -> tuple[dict[str, object], dict[str, str]]:
    """Read parameter name→value and name→unit from a spreadsheet sheet."""
    specs: dict[str, object] = {}
    units: dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_row, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
                continue
            try:
                specs[name] = float(value)
            except (ValueError, TypeError):
                specs[name] = value
            units[name] = str(row[2].value) if row[2].value else "—"
    return specs, units


# --- Detector Specs ---
det_specs, det_units = _read_sheet(wb["Detector Specs"])

# --- 1/f Characterization ---
flicker_specs, flicker_units = _read_sheet(wb["1-f Characterization"])

# --- System Configuration ---
sys_specs, sys_units = _read_sheet(wb["System Configuration"])

print("=" * 80)
print("SCENARIO 2.2: 1/f Noise Corner Frequency Impact on LWIR Staring Array")
print("=" * 80)

print(f"\n=== Detector Specs ===")
for k, v in det_specs.items():
    print(f"  {k}: {v} {det_units.get(k, '')}")

print(f"\n=== 1/f Characterization ===")
for k, v in flicker_specs.items():
    print(f"  {k}: {v} {flicker_units.get(k, '')}")

print(f"\n=== System Configuration ===")
for k, v in sys_specs.items():
    print(f"  {k}: {v} {sys_units.get(k, '')}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

# Optics
aperture_m = float(sys_specs["Aperture diameter"]) / 100.0       # cm → m
focal_length_m = float(sys_specs["Focal length"]) / 100.0        # cm → m
f_number = float(sys_specs["f-number"])
transmission = float(sys_specs["Optical transmission"]) / 100.0  # % → fraction
optics_temp_K = float(sys_specs["Optics temperature"]) + 273.15  # °C → K

# Spectral — nm → µm
band_min_um = float(sys_specs["Filter cut-on"]) / 1000.0
band_max_um = float(sys_specs["Filter cut-off"]) / 1000.0

# Scene
target_temp_K = float(sys_specs["Target temperature"])
target_emiss = float(sys_specs["Target emissivity"])
bg_temp_K = float(sys_specs["Background temperature"])
bg_emiss = float(sys_specs["Background emissivity"])

# Detector
pixel_pitch_um = float(det_specs["Pixel pitch"])
qe = float(det_specs["Quantum efficiency (in-band avg)"]) / 100.0  # % → frac
dark_rate = float(det_specs["Dark current"])              # already e⁻/s
operating_temp_K = float(det_specs["Operating temperature"])  # already K
read_noise = float(det_specs["Read noise (post-CDS)"])       # already e⁻
fwc = float(det_specs["Full well capacity"])                 # already e⁻
adc_bits = int(det_specs["ADC resolution"])
gain = float(det_specs["System gain"])                       # already e⁻/DN
ipc_coupling = float(det_specs["IPC coupling"]) / 100.0     # % → fraction
t_int_s = float(det_specs["Nominal integration time"]) / 1000.0  # ms → s

# 1/f parameters
flicker_K = float(flicker_specs["Flicker coefficient K"])   # already e⁻²
f_corner = float(flicker_specs["Corner frequency"])          # already Hz

print(f"\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'f-number':<35s} {f_number:>14.2f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  nm ÷ 1000")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<35s} {dark_rate:>14.0f}  {'e⁻/s':<15s}  no conversion")
print(f"  {'Read noise':<35s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms ÷ 1000")
print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Flicker K':<35s} {flicker_K:>14.1f}  {'e⁻²':<15s}  no conversion")
print(f"  {'Corner frequency':<35s} {f_corner:>14.1f}  {'Hz':<15s}  no conversion")

# ---------------------------------------------------------------------------
# Step 3: 1/f noise physics — analytic calculation
# ---------------------------------------------------------------------------

print(f"\n=== 1/f Noise Physics ===")
print(f"  The 1/f noise PSD is S(f) = K/f for f below the corner frequency.")
print(f"  Integrating over [f_low, f_high]:")
print(f"    σ_1f = √(K · ln(f_high / f_low))  [e⁻ RMS]")
print(f"")
print(f"  For a staring array:")
print(f"    f_low  = frame rate (lowest frequency in the temporal stream)")
print(f"    f_high = 1 / (2 × t_int) = Nyquist within one integration")
print(f"           = 1 / (2 × {t_int_s*1000:.1f} ms) = {1/(2*t_int_s):.0f} Hz")
print(f"")
print(f"  The corner frequency ({f_corner:.0f} Hz) is where 1/f meets the white")
print(f"  noise floor. Above f_corner, noise is white (flat PSD). Below, 1/f.")
print(f"  Since f_high = {1/(2*t_int_s):.0f} Hz > f_corner = {f_corner:.0f} Hz,")
print(f"  the 1/f integration should technically be capped at f_corner.")
print(f"  But RADIANT integrates over the full [f_low, f_high] band as K/f,")
print(f"  which slightly overestimates 1/f for f > f_corner. We will compute")
print(f"  both the RADIANT model and a corner-limited analytic result.")

f_high_nyq = 1.0 / (2.0 * t_int_s)  # Hz

frame_rates = [30.0, 60.0, 120.0]

print(f"\n  Analytic 1/f noise at each frame rate:")
print(f"  {'Frame Rate [Hz]':>16s}  {'f_low [Hz]':>10s}  {'f_high [Hz]':>11s}"
      f"  {'ln ratio [—]':>12s}  {'σ_1f [e⁻]':>10s}  {'σ_1f (capped) [e⁻]':>20s}")
print(f"  {'-' * 16}  {'-' * 10}  {'-' * 11}  {'-' * 12}  {'-' * 10}  {'-' * 20}")

for fr in frame_rates:
    f_low = fr
    # Full band (RADIANT model)
    ln_ratio = math.log(f_high_nyq / f_low)
    sigma_1f = math.sqrt(flicker_K * ln_ratio)
    # Corner-limited (more accurate physically)
    f_eff_high = min(f_high_nyq, f_corner)
    ln_ratio_capped = math.log(f_eff_high / f_low)
    sigma_1f_capped = math.sqrt(flicker_K * ln_ratio_capped)
    print(f"  {fr:>16.0f}  {f_low:>10.0f}  {f_high_nyq:>11.0f}"
          f"  {ln_ratio:>12.3f}  {sigma_1f:>10.1f}  {sigma_1f_capped:>20.1f}")

# ---------------------------------------------------------------------------
# Step 4: Build RADIANT config and run at each frame rate
# ---------------------------------------------------------------------------


def make_config(
    t_int: float = t_int_s,
    flicker_k: float = 0.0,
    f_low: float = 60.0,
    f_high: float = 5000.0,
) -> dict:
    """Build RADIANT config for the LWIR staring system."""
    cfg: dict = {
        "source": {
            "target": {
                "temperature": target_temp_K,
                "emissivity": target_emiss,
            },
            "background": {
                "temperature": bg_temp_K,
                "emissivity": bg_emiss,
            },
        },
        "atmosphere": {
            "model": "exo",  # Ground-based, short range — treat as no atm for now
        },
        "geometry": {
            "sensor_altitude_m": 0.0,
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": operating_temp_K,
            "ipc_coupling": ipc_coupling,
            "flicker_K": flicker_k,
            "flicker_f_low_hz": f_low,
            "flicker_f_high_hz": f_high,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
    }
    return cfg


def compute_nedt(cfg: dict) -> tuple[float, float, float, dict[str, float]]:
    """Run RADIANT at T, T±δ and return (nedt_mK, ds_dt, total_noise, noise_dict)."""
    DELTA_T = 0.5  # K
    T = cfg["source"]["target"]["temperature"]

    r_center = Sensor.from_dict(cfg).evaluate()
    signal_e = r_center.stage_outputs["readout"]["signal_e_final"]

    cfg_p = copy.deepcopy(cfg)
    cfg_p["source"]["target"]["temperature"] = T + DELTA_T
    cfg_m = copy.deepcopy(cfg)
    cfg_m["source"]["target"]["temperature"] = T - DELTA_T

    sp = Sensor.from_dict(cfg_p).evaluate().stage_outputs["readout"]["signal_e_final"]
    sm = Sensor.from_dict(cfg_m).evaluate().stage_outputs["readout"]["signal_e_final"]
    ds_dt = (sp - sm) / (2.0 * DELTA_T)

    noise_dict = {nt.name: nt.value_e for nt in r_center.noise_terms}
    total_noise = math.sqrt(sum(v ** 2 for v in noise_dict.values()))

    nedt_mK = (total_noise / ds_dt * 1000.0) if ds_dt > 0 else float("inf")
    return nedt_mK, ds_dt, total_noise, noise_dict


print(f"\n=== Radiometric Regime ===")
print(f"  Atmosphere model: exo (short-range, negligible atmospheric path)")
print(f"  Extended regime: target fills entire pixel FOV.")
print(f"  LWIR system: background flux dominates the well charge.")
print(f"  Integration time: {t_int_s*1e6:.0f} µs (FWC-limited for LWIR)")
print(f"")
print(f"  1/f NOISE NOTE:")
print(f"    The 1/f noise contribution depends on the frequency band [f_low, f_high].")
print(f"    For a staring array, f_low = frame rate and f_high = 1/(2·t_int).")
print(f"    Lower frame rates include more low-frequency noise → higher σ_1f.")
print(f"    The key insight: 1/f noise scales as √ln(f_high/f_low), which is")
print(f"    a weak (logarithmic) dependence — it takes a large change in frame")
print(f"    rate to significantly change 1/f noise.")

# ---------------------------------------------------------------------------
# Step 5: Run RADIANT at each frame rate — with and without 1/f
# ---------------------------------------------------------------------------

print(f"\n=== Running RADIANT at each frame rate ===")
print(f"  Integration time fixed at {t_int_s*1e6:.0f} µs for all frame rates")
print(f"  (FWC-limited, not frame-rate-limited)")

results: list[dict] = []

for fr in frame_rates:
    f_low = fr
    f_high = f_high_nyq

    # Without 1/f
    cfg_no1f = make_config(flicker_k=0.0, f_low=f_low, f_high=f_high)
    nedt_no1f, ds_dt, noise_no1f, ndict_no1f = compute_nedt(cfg_no1f)

    # With 1/f
    cfg_1f = make_config(flicker_k=flicker_K, f_low=f_low, f_high=f_high)
    nedt_1f, _, noise_1f, ndict_1f = compute_nedt(cfg_1f)

    # Signal from the no-1/f run
    r_check = Sensor.from_dict(cfg_no1f).evaluate()
    signal_e = r_check.stage_outputs["readout"]["signal_e_final"]
    well_fill = signal_e / fwc * 100.0

    results.append({
        "frame_rate": fr,
        "f_low": f_low,
        "f_high": f_high,
        "signal_e": signal_e,
        "well_fill_pct": well_fill,
        "ds_dt": ds_dt,
        "nedt_no1f_mK": nedt_no1f,
        "nedt_1f_mK": nedt_1f,
        "noise_no1f": noise_no1f,
        "noise_1f": noise_1f,
        "ndict_no1f": ndict_no1f,
        "ndict_1f": ndict_1f,
        "sigma_1f": ndict_1f.get("flicker_1f", 0.0),
    })

# Print comparison table
print(f"\n=== NEDT: Without vs. With 1/f Noise ===")
print(f"  {'Frame Rate [Hz]':>15s}  {'f_low [Hz]':>10s}  {'f_high [Hz]':>11s}  {'σ_1f [e⁻]':>10s}"
      f"  {'σ_total (no 1/f) [e⁻]':>22s}  {'σ_total (w/ 1/f) [e⁻]':>22s}"
      f"  {'NEDT no-1/f [mK]':>17s}  {'NEDT w/ 1/f [mK]':>17s}  {'Δ NEDT [mK]':>12s}")
print(f"  {'-' * 15}  {'-' * 10}  {'-' * 11}  {'-' * 10}"
      f"  {'-' * 22}  {'-' * 22}  {'-' * 17}  {'-' * 17}  {'-' * 12}")

for r in results:
    delta = r["nedt_1f_mK"] - r["nedt_no1f_mK"]
    print(f"  {r['frame_rate']:>15.0f}  {r['f_low']:>10.0f}  {r['f_high']:>11.0f}"
          f"  {r['sigma_1f']:>10.1f}  {r['noise_no1f']:>22.1f}"
          f"  {r['noise_1f']:>22.1f}  {r['nedt_no1f_mK']:>17.1f}"
          f"  {r['nedt_1f_mK']:>17.1f}  {delta:>12.1f}")

# ---------------------------------------------------------------------------
# Step 6: Detailed noise breakdown at nominal frame rate (60 Hz)
# ---------------------------------------------------------------------------

nominal = next(r for r in results if r["frame_rate"] == 60.0)

print(f"\n=== Noise Breakdown at 60 Hz (With 1/f) ===")
print(f"  Signal:     {nominal['signal_e']:>14,.0f} e⁻ ({nominal['well_fill_pct']:.1f}% well)")
print(f"  dS/dT:      {nominal['ds_dt']:>14.0f} e⁻/K")
print(f"  Total noise: {nominal['noise_1f']:>13.1f} e⁻ RMS")
print(f"")

ds_dt = nominal["ds_dt"]
ndict = nominal["ndict_1f"]
nedt_per_term = {
    name: (sigma / ds_dt * 1000.0) if ds_dt > 0 else float("inf")
    for name, sigma in ndict.items()
}
nedt_sq_total = sum(v ** 2 for v in nedt_per_term.values())

print(f"  {'Noise Term':<25s}  {'σ [e⁻ RMS]':>12s}  {'NEDT_i [mK]':>12s}"
      f"  {'Fraction [%]':>13s}")
print(f"  {'-' * 25}  {'-' * 12}  {'-' * 12}  {'-' * 13}")

sorted_terms = sorted(nedt_per_term.items(), key=lambda x: abs(x[1]), reverse=True)
for name, nedt_i in sorted_terms:
    sigma = ndict[name]
    if sigma < 0.001:
        continue
    frac = nedt_i ** 2 / nedt_sq_total * 100.0 if nedt_sq_total > 0 else 0.0
    print(f"  {name:<25s}  {sigma:>12.1f}  {nedt_i:>12.2f}  {frac:>13.1f}")

print(f"  {'─' * 25}  {'─' * 12}  {'─' * 12}  {'─' * 13}")
print(f"  {'RSS TOTAL':<25s}  {nominal['noise_1f']:>12.1f}"
      f"  {nominal['nedt_1f_mK']:>12.2f}  {'100.0':>13s}")

# Highlight 1/f contribution
sigma_1f = nominal["sigma_1f"]
nedt_1f_contrib = (sigma_1f / ds_dt * 1000.0) if ds_dt > 0 else 0.0
frac_1f = (sigma_1f ** 2 / nominal["noise_1f"] ** 2 * 100.0)

print(f"\n  1/f contribution at 60 Hz:")
print(f"    σ_1f = {sigma_1f:.1f} e⁻ RMS")
print(f"    NEDT_1f = {nedt_1f_contrib:.2f} mK")
print(f"    Fraction of total noise variance: {frac_1f:.1f}%")
print(f"    NEDT increase due to 1/f: {nominal['nedt_1f_mK'] - nominal['nedt_no1f_mK']:.1f} mK"
      f" ({(nominal['nedt_1f_mK']/nominal['nedt_no1f_mK'] - 1)*100:.1f}%)")

# ---------------------------------------------------------------------------
# Step 7: Sweep f_low (frame rate) from 1 Hz to 500 Hz
# ---------------------------------------------------------------------------

print(f"\n=== 1/f Noise vs. Frame Rate (Sweep) ===")
print(f"  Sweeping f_low from 1 Hz to 500 Hz in 50 steps")
print(f"  f_high fixed at {f_high_nyq:.0f} Hz (= 1 / (2 × {t_int_s*1e6:.0f} µs))")
print(f"  Using analytic formula: σ_1f = √(K · ln(f_high / f_low))")
print(f"")

sweep_f_low = np.logspace(0, np.log10(500), 50)  # 1 to 500 Hz, log-spaced

print(f"  {'f_low [Hz]':>10s}  {'σ_1f [e⁻]':>10s}  {'NEDT_1f [mK]':>12s}"
      f"  {'Total σ [e⁻]':>12s}  {'NEDT total [mK]':>15s}")
print(f"  {'-' * 10}  {'-' * 10}  {'-' * 12}  {'-' * 12}  {'-' * 15}")

# Get baseline noise (no 1/f) for RSS calculation
noise_no1f_sq = nominal["noise_no1f"] ** 2

# Print every 5th point for readability
for i, fl in enumerate(sweep_f_low):
    sigma_1f_analytic = math.sqrt(flicker_K * math.log(f_high_nyq / fl))
    nedt_1f_only = sigma_1f_analytic / ds_dt * 1000.0
    total_sigma = math.sqrt(noise_no1f_sq + sigma_1f_analytic ** 2)
    total_nedt = total_sigma / ds_dt * 1000.0
    if i % 5 == 0 or fl >= 499:
        print(f"  {fl:>10.1f}  {sigma_1f_analytic:>10.1f}  {nedt_1f_only:>12.2f}"
              f"  {total_sigma:>12.1f}  {total_nedt:>15.2f}")

# ---------------------------------------------------------------------------
# Step 8: Corner frequency analysis
# ---------------------------------------------------------------------------

print(f"\n=== Corner Frequency Analysis ===")
print(f"  The corner frequency f_c = {f_corner:.0f} Hz is where 1/f PSD meets")
print(f"  the white noise floor. For f > f_c, the noise PSD is flat (white).")
print(f"")
print(f"  RADIANT's model: σ_1f = √(K · ln(f_high / f_low))")
print(f"  This integrates 1/f over the FULL band [f_low, f_high].")
print(f"  But physically, the 1/f PSD only applies below f_c = {f_corner:.0f} Hz.")
print(f"  Above f_c, noise is already captured as read noise (white floor).")
print(f"")
print(f"  Corner-limited model: σ_1f = √(K · ln(min(f_high, f_c) / f_low))")
print(f"")

print(f"  Comparison at each frame rate:")
print(f"  {'Frame Rate [Hz]':>15s}  {'σ_1f (full) [e⁻]':>17s}  {'σ_1f (capped) [e⁻]':>20s}"
      f"  {'Overestimate [%]':>16s}")
print(f"  {'-' * 15}  {'-' * 17}  {'-' * 20}  {'-' * 16}")

for fr in frame_rates:
    f_low = fr
    sigma_full = math.sqrt(flicker_K * math.log(f_high_nyq / f_low))
    f_eff = min(f_high_nyq, f_corner)
    sigma_capped = math.sqrt(flicker_K * math.log(f_eff / f_low))
    overest_pct = (sigma_full / sigma_capped - 1.0) * 100.0
    print(f"  {fr:>15.0f}  {sigma_full:>17.1f}  {sigma_capped:>20.1f}"
          f"  {overest_pct:>15.1f}%")

print(f"\n  Since f_high ({f_high_nyq:.0f} Hz) >> f_corner ({f_corner:.0f} Hz),")
print(f"  the RADIANT full-band model overestimates σ_1f by")
print(f"  ~{(math.sqrt(math.log(f_high_nyq/60)/math.log(f_corner/60))-1)*100:.0f}%"
      f" at 60 Hz. This is a known limitation:")
print(f"  RADIANT does not model the corner frequency transition.")

# ---------------------------------------------------------------------------
# Step 9: Write output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

owb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(
    start_color="002E75B6", end_color="002E75B6", fill_type="solid"
)
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# --- Sheet 1: NEDT vs Frame Rate ---
ows1 = owb.active
ows1.title = "NEDT vs Frame Rate"
ows1["A1"] = "NEDT vs. Frame Rate — 1/f Noise Impact"
ows1["A1"].font = Font(bold=True, size=14)
ows1.column_dimensions["A"].width = 16
ows1.column_dimensions["B"].width = 14
ows1.column_dimensions["C"].width = 14
ows1.column_dimensions["D"].width = 18
ows1.column_dimensions["E"].width = 18
ows1.column_dimensions["F"].width = 18
ows1.column_dimensions["G"].width = 18
ows1.column_dimensions["H"].width = 14
ows1.column_dimensions["I"].width = 14

headers = ["Frame Rate [Hz]", "f_low [Hz]", "f_high [Hz]",
           "σ_1f [e⁻ RMS]", "σ_total (no 1/f)", "σ_total (w/ 1/f)",
           "NEDT no-1/f [mK]", "NEDT w/ 1/f [mK]", "Δ NEDT [mK]"]
for c, h in enumerate(headers, 1):
    cell = ows1.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border

for i, r in enumerate(results, 4):
    delta = r["nedt_1f_mK"] - r["nedt_no1f_mK"]
    vals = [r["frame_rate"], r["f_low"], r["f_high"], r["sigma_1f"],
            r["noise_no1f"], r["noise_1f"], r["nedt_no1f_mK"],
            r["nedt_1f_mK"], delta]
    for c, v in enumerate(vals, 1):
        cell = ows1.cell(row=i, column=c, value=round(v, 2))
        cell.border = thin_border

# --- Sheet 2: Noise Breakdown at 60 Hz ---
ows2 = owb.create_sheet("Noise Breakdown 60 Hz")
ows2["A1"] = "Noise Breakdown at 60 Hz Frame Rate (With 1/f)"
ows2["A1"].font = Font(bold=True, size=14)
ows2.column_dimensions["A"].width = 26
ows2.column_dimensions["B"].width = 16
ows2.column_dimensions["C"].width = 16
ows2.column_dimensions["D"].width = 14

headers2 = ["Noise Term", "σ [e⁻ RMS]", "NEDT_i [mK]", "Fraction [%]"]
for c, h in enumerate(headers2, 1):
    cell = ows2.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border

row_idx = 4
for name, nedt_i in sorted_terms:
    sigma = ndict[name]
    if sigma < 0.001:
        continue
    frac = nedt_i ** 2 / nedt_sq_total * 100.0 if nedt_sq_total > 0 else 0.0
    ows2.cell(row=row_idx, column=1, value=name).border = thin_border
    ows2.cell(row=row_idx, column=2, value=round(sigma, 2)).border = thin_border
    ows2.cell(row=row_idx, column=3, value=round(nedt_i, 2)).border = thin_border
    ows2.cell(row=row_idx, column=4, value=round(frac, 1)).border = thin_border
    row_idx += 1

# --- Sheet 3: f_low Sweep ---
ows3 = owb.create_sheet("Frame Rate Sweep")
ows3["A1"] = "1/f Noise vs. Frame Rate (Analytic Sweep)"
ows3["A1"].font = Font(bold=True, size=14)
ows3.column_dimensions["A"].width = 14
ows3.column_dimensions["B"].width = 16
ows3.column_dimensions["C"].width = 16
ows3.column_dimensions["D"].width = 16
ows3.column_dimensions["E"].width = 18

headers3 = ["f_low [Hz]", "σ_1f [e⁻ RMS]", "NEDT_1f [mK]",
            "Total σ [e⁻]", "NEDT total [mK]"]
for c, h in enumerate(headers3, 1):
    cell = ows3.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border

for i, fl in enumerate(sweep_f_low, 4):
    sigma_1f_val = math.sqrt(flicker_K * math.log(f_high_nyq / fl))
    nedt_1f_val = sigma_1f_val / ds_dt * 1000.0
    total_sig = math.sqrt(noise_no1f_sq + sigma_1f_val ** 2)
    total_nedt_val = total_sig / ds_dt * 1000.0
    ows3.cell(row=i, column=1, value=round(fl, 2)).border = thin_border
    ows3.cell(row=i, column=2, value=round(sigma_1f_val, 2)).border = thin_border
    ows3.cell(row=i, column=3, value=round(nedt_1f_val, 2)).border = thin_border
    ows3.cell(row=i, column=4, value=round(total_sig, 2)).border = thin_border
    ows3.cell(row=i, column=5, value=round(total_nedt_val, 2)).border = thin_border

# --- Sheet 4: Summary ---
ows4 = owb.create_sheet("Summary")
ows4["A1"] = "Scenario 2.2 Summary"
ows4["A1"].font = Font(bold=True, size=14)
ows4.column_dimensions["A"].width = 40
ows4.column_dimensions["B"].width = 20

row = 3
summaries = [
    ("System", "LWIR HgCdTe staring array, 640×512"),
    ("Band", f"{band_min_um:.1f}–{band_max_um:.1f} µm"),
    ("Optics", f"{aperture_m*100:.0f} cm, f/{f_number:.1f}, τ = {transmission*100:.0f}%"),
    ("Pixel pitch", f"{pixel_pitch_um:.0f} µm"),
    ("Integration time", f"{t_int_s*1e6:.0f} µs"),
    ("FWC", f"{fwc/1e6:.0f} M e⁻"),
    ("Flicker K", f"{flicker_K:.0f} e⁻²"),
    ("Corner frequency", f"{f_corner:.0f} Hz"),
    ("", ""),
    ("Signal at 300 K", f"{nominal['signal_e']:,.0f} e⁻"),
    ("Well fill", f"{nominal['well_fill_pct']:.1f}%"),
    ("dS/dT", f"{nominal['ds_dt']:,.0f} e⁻/K"),
    ("", ""),
    ("NEDT (60 Hz, no 1/f)", f"{nominal['nedt_no1f_mK']:.1f} mK"),
    ("NEDT (60 Hz, with 1/f)", f"{nominal['nedt_1f_mK']:.1f} mK"),
    ("1/f NEDT penalty", f"{nominal['nedt_1f_mK'] - nominal['nedt_no1f_mK']:.1f} mK"),
]
for label, val in summaries:
    ows4.cell(row=row, column=1, value=label).border = thin_border
    ows4.cell(row=row, column=2, value=val).border = thin_border
    row += 1

owb.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
