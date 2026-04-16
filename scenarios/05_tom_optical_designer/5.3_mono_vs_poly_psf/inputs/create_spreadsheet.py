"""Create Tom's mono vs. polychromatic PSF comparison spreadsheet.

Tom is an optical designer who wants to understand how much chromaticism
matters for PSF / MTF / EE analysis across his MWIR band.

He has:
  - The optical design from scenario 5.2 (f/4, 30 cm aperture, 18 µm pixel)
  - Wants to compare PSF behavior at 3.5, 4.0, 4.5, and 5.0 µm
  - Wants to know if monochromatic (band-center) analysis gives misleading
    results compared to polychromatic (flux-weighted) analysis

All values are in the units an optical designer typically works with:
  - Focal length in mm (Zemax convention)
  - Aperture in mm (Zemax convention)
  - Wavelengths in nm (spectral design)
  - Pixel pitch in µm (detector vendor convention)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=value).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value=unit).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ============================================================================
# Sheet 1: Optical Design (same as 5.2, selected 18 µm pixel)
# ============================================================================

ws1 = wb.active
ws1.title = "Optical Design"

ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 15
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 55

ws1["A1"] = "MWIR Pushbroom Imager — Chromatic PSF Analysis"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "Tom Chen — Post-PDR Design, April 2026"
ws1["A2"].font = Font(italic=True, size=10)
ws1["A3"] = "Baseline: f/4, 30 cm aperture, 18 µm pixel (selected in trade 5.2)"
ws1["A3"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    cell = ws1.cell(row=5, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

row = 6

row = add_section(ws1, row, "Telescope Design")
row = add_param(ws1, row, "Entrance pupil diameter", 300, "mm",
                "Circular, unobscured equivalent")
row = add_param(ws1, row, "Effective focal length", 1200, "mm",
                "Measured at 4.25 µm reference wavelength")
row = add_param(ws1, row, "f-number (working)", 4.0, "—", "f/# = EFL / EPD")
row = add_param(ws1, row, "Optical transmission", 72, "%",
                "Including 2 mirrors (Al + protective), cold filter")
row = add_param(ws1, row, "Optics temperature", 293, "K", "Ambient, uncooled barrel")
row = add_param(ws1, row, "WFE (RMS, on-axis)", 0.05, "waves",
                "At λ_ref = 4.25 µm; diffraction-limited")

row = add_section(ws1, row, "Spectral Design")
row = add_param(ws1, row, "Band center wavelength", 4250, "nm",
                "MWIR thermal imaging band")
row = add_param(ws1, row, "Filter cut-on", 3500, "nm",
                "Cold filter, 50% transmission edge")
row = add_param(ws1, row, "Filter cut-off", 5000, "nm",
                "Cold filter, 50% transmission edge")

row = add_section(ws1, row, "Detector (Selected from Trade 5.2)")
row = add_param(ws1, row, "Pixel pitch", 18, "µm",
                "Selected: best SNR/GSD among compliant options")
row = add_param(ws1, row, "Quantum efficiency (in-band)", 72, "%", "Vendor spec")
row = add_param(ws1, row, "Dark current", 50, "e⁻/s", "At 77 K operating temp")
row = add_param(ws1, row, "Full well capacity", 500000, "e⁻", "")
row = add_param(ws1, row, "Read noise (CDS)", 18, "e⁻ RMS", "")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "")
row = add_param(ws1, row, "System gain", 2.0, "e⁻/DN", "")
row = add_param(ws1, row, "Integration time", 2, "ms", "Nominal")

row = add_section(ws1, row, "Scene / Geometry")
row = add_param(ws1, row, "Target temperature", 310, "K", "Daytime ground")
row = add_param(ws1, row, "Target emissivity", 0.95, "—", "Vegetation/soil")
row = add_param(ws1, row, "Orbit altitude", 500, "km", "LEO, sun-synchronous")

# ============================================================================
# Sheet 2: Chromatic Analysis Wavelengths
# ============================================================================

ws2 = wb.create_sheet("Analysis Wavelengths")

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12

ws2["A1"] = "Monochromatic Analysis Wavelengths"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Compare PSF at each wavelength across the MWIR band"
ws2["A2"].font = Font(italic=True, size=10)

wavelengths_nm = [3500, 4000, 4250, 4500, 5000]
wl_labels = [f"{w} nm" for w in wavelengths_nm]

headers = ["Parameter"] + wl_labels + ["Unit"]
for col, h_text in enumerate(headers, 1):
    cell = ws2.cell(row=4, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# Computed values Tom would put in his design notes
airy_diameters = [2.44 * (w / 1000) * 4.0 for w in wavelengths_nm]  # µm
q_values = [(w / 1000) * 4.0 / 18.0 for w in wavelengths_nm]
nyquist_freqs = [1 / (2 * 18e-3) for _ in wavelengths_nm]  # cy/mm
cutoff_freqs = [1 / ((w / 1000e3) * 4.0) for w in wavelengths_nm]  # cy/mm

analysis_params = [
    ("Wavelength", wavelengths_nm, "nm"),
    ("Airy disk diameter (2.44 λ f/#)",
     [round(a, 1) for a in airy_diameters], "µm"),
    ("Q parameter (λ f/# / p)",
     [round(q, 2) for q in q_values], "—"),
    ("Nyquist frequency (1/2p)",
     [round(nyquist_freqs[0], 1)] * 5, "cy/mm"),
    ("Diffraction cutoff (1/λf/#)",
     [round(c, 1) for c in cutoff_freqs], "cy/mm"),
]

for i, (param_name, values, unit) in enumerate(analysis_params, 5):
    ws2.cell(row=i, column=1, value=param_name).border = thin_border
    for j, val in enumerate(values):
        cell = ws2.cell(row=i, column=j + 2, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    ws2.cell(row=i, column=len(values) + 2, value=unit).border = thin_border

note_row = len(analysis_params) + 6
ws2.cell(row=note_row, column=1,
         value="Airy disk grows linearly with λ: 34.2 µm at 3.5 µm → 48.8 µm at 5.0 µm"
         ).font = Font(italic=True)
ws2.cell(row=note_row + 1, column=1,
         value="Q ranges from 0.78 (3.5 µm, slightly undersampled) to 1.11 (5.0 µm, well-sampled)"
         ).font = Font(italic=True)
ws2.cell(row=note_row + 2, column=1,
         value="Polychromatic PSF is flux-weighted average — 300 K blackbody peaks near 9.7 µm"
         ).font = Font(italic=True)
ws2.cell(row=note_row + 3, column=1,
         value="but in-band (3.5–5.0 µm), longer wavelengths see more photon flux"
         ).font = Font(italic=True)

# ============================================================================
# Sheet 3: Comparison Metrics
# ============================================================================

ws3 = wb.create_sheet("Comparison Metrics")

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 50

ws3["A1"] = "Metrics to Compare Across PSF Models"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Tom wants to quantify the impact of chromaticism"
ws3["A2"].font = Font(italic=True, size=10)

for col, h_text in enumerate(["Metric", "Why It Matters"], 1):
    cell = ws3.cell(row=4, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

metrics = [
    ("MTF at Nyquist",
     "Core image quality metric. Does chromatic averaging lower it?"),
    ("EE 1×1 (ensquared energy, single pixel)",
     "Point source detection sensitivity. Chromatic spread loses energy."),
    ("EE 3×3 (ensquared energy, 3×3 box)",
     "Extended point source metric. Less sensitive to chromatic spread."),
    ("FWHM (full width at half maximum)",
     "PSF core size. Polychromatic should be broader than shortest-λ."),
    ("SNR (signal-to-noise ratio)",
     "Does the PSF model choice actually affect SNR in extended scene?"),
    ("Airy disk diameter",
     "Fundamental: 2.44 × λ × f/#. Varies 34–49 µm across band."),
    ("Q parameter (λ f/# / p)",
     "Sampling adequacy. Varies 0.78–1.11 across band at 18 µm pitch."),
]

for i, (metric, why) in enumerate(metrics, 5):
    ws3.cell(row=i, column=1, value=metric).border = thin_border
    ws3.cell(row=i, column=2, value=why).border = thin_border

output_path = "tom_chromatic_psf_analysis.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
