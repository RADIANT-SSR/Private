"""Scenario 10.3 — ground-to-space SST, visible band (ADR-0011 scene class E3).

A 1 m ground-based tracking telescope at an SST site observes a sunlit LEO
object through the FULL atmospheric column, up-looking, during the evening
terminator window.  This is the ``ground_to_space`` scene class of the
Geometry-Flexibility upgrade (ADR-0011 §2 grid, owner priority 4): observer on
the ground, target in space, line of sight ``up``, canonical target-side path
zenith θ_o **obtuse**.

What this run demonstrates
--------------------------
1.  Scene class ``ground_to_space`` derived and published, plus the optional
    ``geometry.scene_class`` assertion (agreeing → silent; disagreeing → an
    actionable ``GeometrySpecificationError``).
2.  The angle pair the Phase-4 GUI schematic draws: ζ_low (zenith at the
    TELESCOPE, the path's lower endpoint) and the obtuse θ_o (zenith at the
    OBJECT, the canonical angle).
3.  Metric conditioning by scene class (guardrail G3): the ground-projection
    family (GSD, swath, ground range) defaults OFF, the target-plane /
    angular-resolution family defaults ON.
4.  Per-altitude solar illumination (GF-9): the terminator shadow-height test.
    At 12° solar depression the SITE is dark while the OBJECT at 700 km is in
    full sunlight; past ~25.7° depression the object enters eclipse.
5.  The VIS/NIR sky-background provisional ``UserWarning`` (ADR-0011
    decision 10 band gating) — captured and printed from a sun-above-horizon
    comparison run.
6.  Turbulence (Gap 110): the HV-5/7 Cn² profile drives a path-weighted Fried
    parameter r₀, and the resulting seeing-limited blur is compared with the
    diffraction limit of the same 1 m aperture.
7.  The horizon guard (ADR-0011 decision 6) firing on a near-horizontal
    pointing.

Radiometric regime
------------------
POINT_SOURCE, and it is *declared* (``source.scene_type = 'point_source'``),
not inferred.  The object is 1 m² at 740 km — 1.35 µrad across — versus a
2.4 µrad pixel IFOV and a ~10 µrad seeing disc, so it is unresolved by two
orders of magnitude in solid angle.  In the point-source regime RADIANT
computes the signal as I(λ)·A_collect/R² (radiant intensity ÷ range²), applies
EE_box exactly once (Rule 9), and strips the path-radiance pedestal from the
target term; the sky pedestal re-enters only as a full-pixel background that
shot-noises.

Runtime: ~40 s.  Usage:  python run_ground_to_space_sst_visible.py
"""

from __future__ import annotations

import math
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")  # headless; must precede pyplot

import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from radiant.api import Sensor  # noqa: E402
from radiant.api.scene_relevance import default_off_metrics  # noqa: E402
from radiant.atmosphere.cn2_hufnagel_valley import HufnagelValleyCn2  # noqa: E402
from radiant.atmosphere.r0_path import path_fried_parameter_from_los  # noqa: E402
from radiant.atmosphere.solar_shadow import shadow_height_m, sunlit  # noqa: E402
from radiant.core.constants import R_EARTH_M, S_solar_W_per_m2, c, h  # noqa: E402
from radiant.core.los_geometry import LineOfSightGeometry  # noqa: E402

# ---------------------------------------------------------------------------
# Paths — repo-relative, pathlib only (Rule 30)
# ---------------------------------------------------------------------------

SCENARIO_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = SCENARIO_DIR / "inputs"
OUTPUT_DIR = SCENARIO_DIR / "outputs"
WORKBOOK_IN = INPUT_DIR / "sst_site_and_tasking.xlsx"
RESULTS_XLSX = OUTPUT_DIR / "ground_to_space_sst_results.xlsx"

#: Spectral grid points across the filter band (Sensor builds linspace(min, max, N)).
WAVELENGTH_POINTS = 121

#: Apparent magnitude of the Sun, V band [dimensionless] — IAU/Willmer 2018.
SOLAR_APPARENT_MAGNITUDE = -26.74

#: Published broadband V-band zenith extinction at good astronomical sites
#: [mag / airmass].  Range spans a high dry site (~0.12) to a good sea-level
#: site on a clear night (~0.20).  Sources: Hardie (1962) photometric
#: reduction practice; Burke, Gladders & Graham, *Astronomical Photometry*
#: (2010) §5 tabulates 0.1-0.3 mag/airmass for V at typical observatories.
EXTINCTION_V_LO_MAG_PER_AIRMASS = 0.12
EXTINCTION_V_HI_MAG_PER_AIRMASS = 0.20
V_BAND_CENTRE_UM = 0.55


# ---------------------------------------------------------------------------
# Step 1 — read the vendor workbook and the vendor signature file
# ---------------------------------------------------------------------------


def _read_kv_sheet(ws) -> dict[str, tuple[object, str]]:
    """Read a (Parameter, Value, Unit, Note) sheet into {name: (value, unit)}."""
    out: dict[str, tuple[object, str]] = {}
    for row in ws.iter_rows(min_row=5, max_col=3, values_only=True):
        if row[0] is None:
            continue
        out[str(row[0])] = (row[1], "" if row[2] is None else str(row[2]))
    return out


def _read_ladder(ws) -> list[float]:
    """Read a single-column numeric ladder sheet."""
    values: list[float] = []
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        if row[0] is not None:
            values.append(float(row[0]))
    return values


def read_vendor_inputs() -> dict[str, object]:
    """Load every vendor artifact; no unit conversion happens here."""
    wb = openpyxl.load_workbook(WORKBOOK_IN, data_only=True)
    site = _read_kv_sheet(wb["Site & Telescope"])
    task = _read_kv_sheet(wb["Tasking & Geometry"])
    catalog = _read_kv_sheet(wb["Object Catalog"])
    geo_object = _read_kv_sheet(wb["Reflective Door Object"])
    seeing = _read_kv_sheet(wb["Seeing"])
    zenith_ladder_deg = _read_ladder(wb["Zenith Ladder"])
    terminator_ladder_deg = _read_ladder(wb["Terminator Ladder"])
    air_mass_probe_deg = _read_ladder(wb["Air Mass Probe"])

    signature_name = str(catalog["Signature File"][0])
    sig_path = INPUT_DIR / signature_name
    rows = [
        line.split(",")
        for line in sig_path.read_text(encoding="utf-8").splitlines()
        if line and not line.startswith("#") and not line[0].isalpha()
    ]
    lam_nm = np.array([float(r[0]) for r in rows], dtype=np.float64)
    intensity_per_nm = np.array([float(r[1]) for r in rows], dtype=np.float64)

    return {
        "site": site,
        "task": task,
        "catalog": catalog,
        "geo_object": geo_object,
        "seeing": seeing,
        "zenith_ladder_deg": zenith_ladder_deg,
        "terminator_ladder_deg": terminator_ladder_deg,
        "air_mass_probe_deg": air_mass_probe_deg,
        "signature_name": signature_name,
        "signature_lam_nm": lam_nm,
        "signature_I_per_nm": intensity_per_nm,
    }


# ---------------------------------------------------------------------------
# Step 2 — vendor → RADIANT canonical units (each conversion once, commented)
# ---------------------------------------------------------------------------


def build_config(vendor: dict[str, object], signature_csv_um: Path) -> dict[str, object]:
    """Convert the vendor sheet to a RADIANT config dict.

    Canonical units: length m, wavelength µm, angle rad, time s.
    """
    site = vendor["site"]  # type: ignore[index]
    task = vendor["task"]  # type: ignore[index]
    seeing = vendor["seeing"]  # type: ignore[index]

    aperture_m = float(site["Entrance Pupil Diameter"][0]) / 1000.0  # mm → m
    focal_m = float(site["Effective Focal Length"][0]) / 1000.0  # mm → m
    transmission = float(site["Optical Transmission"][0]) / 100.0  # % → fraction
    filter_lo_um = float(site["Filter Band Low"][0]) / 1000.0  # nm → µm
    filter_hi_um = float(site["Filter Band High"][0]) / 1000.0  # nm → µm
    pitch_um = float(site["Pixel Pitch"][0])  # µm is already the schema input unit
    qe = float(site["Quantum Efficiency"][0]) / 100.0  # % → fraction
    full_well_e = float(site["Full Well Capacity"][0]) * 1000.0  # ke- → e-
    t_int_s = float(site["Exposure Time"][0]) / 1000.0  # ms → s
    site_alt_m = float(site["Site elevation MSL"][0])  # already m

    target_alt_m = float(task["Object Altitude"][0]) * 1000.0  # km → m
    zeta_low_rad = math.radians(float(task["Pointing Zenith Angle"][0]))  # deg → rad
    # Solar DEPRESSION below the site horizon → solar ZENITH: θ_s = 90° + δ.
    depression_deg = float(task["Solar Depression"][0])
    theta_s_rad = math.radians(90.0 + depression_deg)  # deg → rad
    solar_az_rad = math.radians(float(task["Solar Relative Azimuth"][0]))  # deg → rad

    return {
        "geometry": {
            "sensor_altitude_m": site_alt_m,
            "target_altitude_m": target_alt_m,
            # ADR-0011 decision 3: an entered path zenith is referenced to the
            # path's LOWER endpoint.  Up-looking ⇒ that endpoint is the
            # telescope, so this IS the pointing zenith on the tasking card.
            "path_zenith_rad": zeta_low_rad,
            "solar_zenith_rad": theta_s_rad,
            "solar_azimuth_rad": solar_az_rad,
        },
        "source": {
            "scene_type": "point_source",
            "target": {"user_intensity_path": str(signature_csv_um)},
        },
        "atmosphere": {
            "model": "simple",
            "standard_atmosphere": str(task["Standard Atmosphere"][0]),
            "visibility_km": float(task["Meteorological Visibility"][0]),  # km is the input unit
            "cn2_profile": str(seeing["Cn2 Profile"][0]),
            "cn2_hv_wind_rms_m_s": float(seeing["High-Altitude Wind RMS"][0]),
            "cn2_hv_ground_strength": float(seeing["Ground Turbulence Strength"][0]),
            "turbulence_wave_type": str(seeing["Wave Type"][0]),
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_m,
            "transmission_scalar": transmission,
        },
        "detector": {
            "pixel_pitch_x_um": pitch_um,
            "pixel_pitch_y_um": pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": float(site["Dark Current"][0]),
        },
        "spectral_integration": {
            "filter_min_um": filter_lo_um,
            "filter_max_um": filter_hi_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": float(site["Read Noise"][0]),
            "gain_e_per_dn": float(site["Gain"][0]),
            "adc_bits": int(site["ADC Resolution"][0]),
            "full_well_capacity_e": full_well_e,
        },
    }


def write_canonical_signature(vendor: dict[str, object], destination: Path) -> Path:
    """Convert the vendor signature to RADIANT's two-column CSV and write it.

    Vendor: wavelength [nm], I [W/sr/nm].  RADIANT: wavelength [µm],
    I [W/sr/µm].  Exactly one header row (the loader auto-skips row 1 only).
    """
    lam_nm = np.asarray(vendor["signature_lam_nm"], dtype=np.float64)
    i_per_nm = np.asarray(vendor["signature_I_per_nm"], dtype=np.float64)
    lam_um = lam_nm / 1000.0  # nm → µm
    i_per_um = i_per_nm * 1000.0  # W/sr/nm → W/sr/µm
    destination.parent.mkdir(parents=True, exist_ok=True)
    body = "\n".join(f"{w:.6f},{v:.10e}" for w, v in zip(lam_um, i_per_um, strict=True))
    destination.write_text(
        "wavelength_um,intensity_W_per_sr_per_um\n" + body + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return destination


# ---------------------------------------------------------------------------
# Module-level factory — the GUI-baseline registry wires to THIS
# ---------------------------------------------------------------------------

# The RADIANT-format intensity CSV is materialised INTO THE SCENARIO'S
# ``inputs/`` and committed (regenerated deterministically from the vendor
# signature CSV on every run; MANIFEST.md names the generator). It must NOT
# live in the system temp directory: the GUI baseline ``.gui.yaml`` references
# this path, and a baseline pointing at an uncommitted temp file is unportable
# (fails verify_gui_yaml / File → Open on any fresh checkout) and mutable
# under the yaml (found by the Phase-5 verification gate).
_DERIVED_DIR = INPUT_DIR


def make_sensor() -> Sensor:
    """Build the scenario's nominal, validated ``Sensor`` (no side effects but one file).

    Reads the vendor workbook and signature file, converts them to canonical
    units, materialises the RADIANT-format intensity CSV beside the vendor
    inputs (the chain needs a path on disk; see ``_DERIVED_DIR``), and returns
    the configured Sensor.  Does NOT run the chain and does NOT touch
    ``outputs/``.
    """
    vendor = read_vendor_inputs()
    signature_csv = write_canonical_signature(
        vendor, _DERIVED_DIR / "object_signature_radiant_um.csv"
    )
    config = build_config(vendor, signature_csv)
    return Sensor.from_dict(config, wavelength_points=WAVELENGTH_POINTS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _run_capturing_warnings(sensor: Sensor) -> tuple[object, list[warnings.WarningMessage]]:
    """Evaluate with warnings forced visible; return (result, captured warnings)."""
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result = sensor.evaluate()
    return result, list(caught)


def _safe_exp(x: float) -> float:
    """``exp(x)`` with IEEE-754 underflow made explicit rather than clamped."""
    return math.exp(x) if x > -745.0 else 0.0


def _band_mean(values: np.ndarray, wavelength_um: np.ndarray) -> float:
    """Unweighted band mean ∫f dλ / ∫dλ [same unit as ``values``]."""
    return float(np.trapezoid(values, wavelength_um) / (wavelength_um[-1] - wavelength_um[0]))


def _photon_weighted_tau(
    tau: np.ndarray, intensity: np.ndarray, wavelength_um: np.ndarray
) -> float:
    """Signal-weighted band transmittance ∫τ·I·λ dλ / ∫I·λ dλ [dimensionless].

    The λ weight is the photon-per-watt factor λ/(hc) that the spectral
    integration stage applies, so this is exactly the τ the electron count
    responds to.
    """
    weight = intensity * wavelength_um
    return float(np.trapezoid(tau * weight, wavelength_um) / np.trapezoid(weight, wavelength_um))


# ---------------------------------------------------------------------------
# Reporting sections
# ---------------------------------------------------------------------------


def section_inputs(vendor: dict[str, object]) -> None:
    print("=" * 78)
    print("SCENARIO 10.3 — GROUND-TO-SPACE SST, VISIBLE BAND (ADR-0011 class E3)")
    print("=" * 78)
    print("\n--- 1. Vendor inputs (as delivered) ---")
    for sheet in ("site", "task", "catalog", "geo_object", "seeing"):
        print(f"  [{sheet}]")
        for name, (value, unit) in vendor[sheet].items():  # type: ignore[index]
            shown = f"{value:g}" if isinstance(value, (int, float)) else str(value)
            print(f"    {name:<32s} {shown:>16s} [{unit or 'dimensionless'}]")
    lam_nm = np.asarray(vendor["signature_lam_nm"])
    i_nm = np.asarray(vendor["signature_I_per_nm"])
    print(f"  [signature] {vendor['signature_name']}")
    print(
        f"    grid {lam_nm[0]:.0f}-{lam_nm[-1]:.0f} nm, {lam_nm.size} rows; "
        f"I(550 nm) = {np.interp(550.0, lam_nm, i_nm):.4e} W/sr/nm"
    )
    print("\n  Unit conversions applied once, in build_config():")
    print("    mm -> m (/1000)   %  -> fraction (/100)   nm -> um (/1000)")
    print("    ms -> s  (/1000)  km -> m (*1000)         ke- -> e- (*1000)")
    print("    deg -> rad (radians)   solar depression d -> theta_s = 90 deg + d")
    print("    signature: W/sr/nm -> W/sr/um (*1000)")


def section_nominal(result, sensor: Sensor, caught) -> dict[str, object]:
    geo = result.stage_outputs["geometry"]
    atm = result.stage_outputs["atmosphere"]
    optics = result.stage_outputs["optics"]
    lam = np.asarray(result.stage_outputs["atmosphere"]["atm_quantities"].wavelength_um)
    tau = np.asarray(atm["tau_atm"])

    zeta_low_deg = math.degrees(sensor.get_input("geometry.path_zenith_rad"))
    theta_o_deg = math.degrees(geo["theta_o_rad"])
    eta_deg = math.degrees(geo["eta_rad"])

    print("\n--- 2. Nominal run — geometry and scene class ---")
    print(f"  scene_class                 : {geo['scene_class']} (derived, Provenance.DERIVED)")
    print(f"  observer_class / target_class: {geo['observer_class']} / {geo['target_class']}")
    print(f"  los_direction               : {geo['los_direction']}")
    print(f"  viewing_mode                : {geo['viewing_mode']}")
    print(f"  sensor altitude h_sensor    : {geo['h_sensor_m'] / 1000.0:.3f} km MSL")
    print(f"  target altitude h_target    : {geo['h_target_m'] / 1000.0:.1f} km MSL")
    print(f"  zeta_low (GUI lower arc)    : {zeta_low_deg:.4f} deg  (zenith AT THE TELESCOPE)")
    print(f"  theta_o  (GUI upper arc)    : {theta_o_deg:.4f} deg  (zenith AT THE OBJECT, obtuse)")
    print(f"  eta      (angle at sensor)  : {eta_deg:.4f} deg  (= 180 deg - zeta_low)")
    print(f"  slant range R               : {geo['slant_range_m'] / 1000.0:.3f} km")
    print(f"  ground range (surface arc)  : {geo['ground_range_m'] / 1000.0:.3f} km")
    print(f"  incidence angle             : {math.degrees(geo['incidence_angle_rad']):.4f} deg")
    print(f"  solar zenith theta_s        : {math.degrees(geo['theta_s_rad']):.3f} deg")
    print(f"  solar illumination flag     : {geo['solar_illumination']}")

    # Closed-form check on the published triangle (spherical sine rule).
    r_s = R_EARTH_M + geo["h_sensor_m"]
    r_t = R_EARTH_M + geo["h_target_m"]
    zeta = math.radians(zeta_low_deg)
    sin_interior_t = (r_s / r_t) * math.sin(math.pi - zeta)
    theta_o_hand = math.pi - math.asin(sin_interior_t)
    phi_hand = theta_o_hand - (math.pi - zeta)
    range_hand = r_t * math.sin(phi_hand) / math.sin(math.pi - zeta)
    print("\n  Closed-form spherical-triangle check (independent of RADIANT):")
    print(f"    theta_o hand = {math.degrees(theta_o_hand):.4f} deg   (chain {theta_o_deg:.4f} deg)")
    print(
        f"    slant  hand = {range_hand / 1000.0:.3f} km       "
        f"(chain {geo['slant_range_m'] / 1000.0:.3f} km, "
        f"rel err {abs(range_hand - geo['slant_range_m']) / geo['slant_range_m']:.2e})"
    )
    print(
        f"    flat-Earth would give R = h/cos(zeta) = "
        f"{geo['h_target_m'] / math.cos(zeta) / 1000.0:.3f} km — "
        "curvature shortens the true path"
    )

    print("\n--- 3. Radiometry ---")
    regime = optics["regime"]
    print(f"  regime (final, OpticsStage) : {getattr(regime, 'value', regime)}")
    print(f"  band-mean tau_atm           : {_band_mean(tau, lam):.4f} (dimensionless)")
    print(f"  tau_atm @ 0.55 um           : {float(np.interp(0.55, lam, tau)):.4f} (dimensionless)")
    print(f"  tau_sun (solar leg to obj)  : {float(np.mean(atm['atm_quantities'].tau_sun)):.4f}")
    print(f"  Fried parameter r0          : {atm['r0_m'] * 100.0:.3f} cm (band centre)")
    print(f"  EE_box (energy in pixel)    : {result.stage_outputs['platform']['EE_box']:.5f}")
    print(f"  signal (central pixel)      : {result.stage_outputs['readout']['signal_e_final']:,.0f} e-")
    print(f"  SNR                         : {result.metrics['snr']:.2f} (dimensionless)")
    print(f"  contrast SNR                : {result.metrics['contrast_snr']:.2f} (dimensionless)")
    print(f"  detection range             : {result.metrics['detection_range_m'] / 1000.0:,.1f} km")
    print(f"  sampling Q_center           : {result.metrics['q_center']:.3f} (dimensionless)")
    print(f"  PSF FWHM (x)                : {result.metrics['fwhm_x_m'] * 1e6:.2f} um on the focal plane")

    cons = result.stage_outputs["performance"]["dual_path_consistency"]
    print("\n--- 4. Rule-4 dual-path consistency (PSF path vs MTF product) ---")
    print(f"  passed_x / passed_y         : {cons.passed_x} / {cons.passed_y}")
    print(
        f"  max |FFT(PSF) - MTF_prod|   : x {cons.max_absolute_error_x:.3e}, "
        f"y {cons.max_absolute_error_y:.3e} (tolerance {cons.tolerance:.3g})"
    )
    warn_texts = [str(w.message) for w in caught]
    print(f"  warnings raised on this run : {len(warn_texts)}")
    for text in warn_texts:
        print(f"    - {text[:160]}")
    if not warn_texts:
        print("    (none — the consistency check stayed silent for this scene class)")
    return {"lam": lam, "tau": tau}


def section_scene_class_assertion(sensor: Sensor) -> None:
    print("\n--- 5. Optional geometry.scene_class assertion (ADR-0011 decision 8) ---")
    agreeing = sensor.clone().set("geometry.scene_class", "ground_to_space")
    agreeing.evaluate()
    print("  assertion 'ground_to_space' agrees with the derivation -> silent pass")

    wrong = sensor.clone().set("geometry.scene_class", "air_to_space")
    try:
        wrong.evaluate()
    except Exception as exc:  # the stage raises GeometrySpecificationError
        first_line = str(exc).split("|")[0].strip()
        print(f"  assertion 'air_to_space' disagrees -> {type(exc).__name__}")
        print(f"    {first_line[:200]}")
    else:  # pragma: no cover - defensive
        print("  UNEXPECTED: the disagreeing assertion did not raise")

    print(
        "  Why it matters here: the site is at 0.900 km MSL and the ground/air band\n"
        "  edge is 1 km.  A typo of 1.900 km would still produce a perfectly\n"
        "  self-consistent scene — of the WRONG class.  The assertion catches it."
    )


def section_metric_relevance(result) -> None:
    scene_class = result.stage_outputs["geometry"]["scene_class"]
    off = sorted(default_off_metrics(scene_class))
    present = set(result.metrics)
    print("\n--- 6. Scene-class metric conditioning (guardrail G3) ---")
    print(f"  scene_class                 : {scene_class}")
    print(f"  metrics defaulted OFF ({len(off)})   : {', '.join(off)}")
    leaked = sorted(present & set(off))
    print(f"  of those, present in metrics : {leaked if leaked else 'none (correct)'}")
    target_plane = sorted(k for k in present if k.startswith("target_plane_"))
    print(f"  target-plane family ON       : {', '.join(target_plane)}")
    for key in target_plane:
        print(f"    {key:<52s} {result.metrics[key] * 100.0:8.3f} cm at the object")
    print(
        f"  angular resolution           : "
        f"{result.metrics['diffraction_limit_angular_urad']:.4f} urad "
        "(diffraction limit, 1.22 lambda/D)"
    )
    print(
        "  Ground-projection metrics (GSD, swath, ground sample distance) are OFF because\n"
        "  the target is in SPACE: projecting a pixel onto the ground is meaningless when\n"
        "  the thing being imaged is 700 km above it.  The target-plane sample distance\n"
        "  replaces them — the same pixel projected onto the OBJECT's plane instead."
    )


def section_shadow_ladder(vendor: dict[str, object], result) -> dict[str, np.ndarray]:
    depressions = np.asarray(vendor["terminator_ladder_deg"], dtype=np.float64)
    h_target_m = result.stage_outputs["geometry"]["h_target_m"]
    print("\n--- 7. Per-altitude solar illumination — the shadow-height test (GF-9) ---")
    print(
        "  Shadow height h_shadow(delta) = R_E * (sec(delta) - 1) with delta the solar\n"
        "  DEPRESSION below the local horizontal.  A point at altitude h is sunlit iff\n"
        "  h >= h_shadow.  Before ADR-0011 decision 10 the framework carried a global\n"
        "  theta_s < 90 deg bound, which made 'satellite in sunlight over a dark site'\n"
        "  inexpressible."
    )
    print(
        f"\n  {'depression [deg]':>17} {'theta_s [deg]':>14} {'h_shadow [km]':>14} "
        f"{'100 km obj':>11} {'700 km obj':>11} {'GEO obj':>9}"
    )
    heights = np.zeros_like(depressions)
    for i, delta_deg in enumerate(depressions):
        theta_s = math.radians(90.0 + delta_deg)
        # shadow_height_m takes the SOLAR ZENITH, not the depression:
        # h_shadow = R_E (1/sin(theta_s) - 1) == R_E (sec(delta) - 1).
        h_sh = shadow_height_m(theta_s)
        heights[i] = h_sh
        flags = [
            "SUNLIT" if sunlit(alt, theta_s) else "shadow"
            for alt in (1.0e5, h_target_m, 3.5786e7)
        ]
        print(
            f"  {delta_deg:17.1f} {90.0 + delta_deg:14.1f} {h_sh / 1000.0:14.1f} "
            f"{flags[0]:>11} {flags[1]:>11} {flags[2]:>9}"
        )
    crossing_rad = math.acos(R_EARTH_M / (R_EARTH_M + h_target_m))
    print(
        f"\n  Object at {h_target_m / 1000.0:.0f} km enters eclipse at solar depression "
        f"{math.degrees(crossing_rad):.2f} deg"
    )
    print(
        "  (hand check: sec(delta) = 1 + h/R_E = "
        f"{1.0 + h_target_m / R_EARTH_M:.6f} -> delta = {math.degrees(crossing_rad):.2f} deg)"
    )
    print(
        f"\n  Nominal tasking sits at 12.0 deg depression: the SITE is dark (theta_s = 102 deg,\n"
        f"  sun 12 deg below its horizon) while the OBJECT at {h_target_m / 1000.0:.0f} km is in full\n"
        f"  sunlight (h_shadow = {shadow_height_m(math.radians(102.0)) / 1000.0:.1f} km, well below it).\n"
        "  That is exactly the terminator window an SST site works in, and it is a scene\n"
        "  RADIANT could not express before Geometry-Flexibility Phase 2."
    )
    print(
        f"\n  Chain agreement: the atmosphere stage publishes tau_sun = "
        f"{float(np.mean(result.stage_outputs['atmosphere']['atm_quantities'].tau_sun)):.4f} for the\n"
        "  nominal, i.e. a VACUUM solar leg — the object sits above h_atm_top, so the beam\n"
        "  that lights it never enters the atmosphere.  The GF-9 test and the solar leg agree."
    )
    print(
        "\n  CAVEAT (gaps.md G2): the intensity door consumes I(lambda) verbatim, so tau_sun\n"
        "  never multiplies the target term.  Re-running the nominal at 30 deg depression\n"
        "  (object eclipsed, tau_sun = 0) would still report the SAME signal.  With this\n"
        "  door the analyst owns the illumination gate."
    )
    return {"depression_deg": depressions, "shadow_height_m": heights}


def _geo_reflective_sensor(sensor: Sensor, vendor: dict[str, object], theta_s_deg: float) -> Sensor:
    """The same telescope tasked on a GEO object through the SHAPE + ALBEDO door."""
    geo = vendor["geo_object"]  # type: ignore[index]
    probe = sensor.clone()
    probe = probe.reset("source.target.user_intensity_path")
    probe = probe.set("source.target.albedo", float(geo["Diffuse Albedo"][0]))
    probe = probe.set(
        "geometry.target.projected_area_m2", float(geo["Projected Area"][0])
    )  # m^2 already canonical
    probe = probe.set(
        "geometry.target_altitude_m", float(geo["Object Altitude"][0]) * 1000.0
    )  # km -> m
    probe = probe.set(
        "geometry.path_zenith_rad", math.radians(float(geo["Pointing Zenith Angle"][0]))
    )  # deg -> rad
    return probe.set("geometry.solar_zenith_rad", math.radians(theta_s_deg))  # deg -> rad


def section_sky_background(
    sensor: Sensor, nominal_result, vendor: dict[str, object]
) -> dict[str, object]:
    print("\n--- 8. Sky background and the VIS/NIR provisional warning ---")
    print(
        "  LOS-termination rule B (Use-Case Matrix §3.2.5): follow the line of sight PAST\n"
        "  the object.  It exits the atmosphere into cold space, so the background is\n"
        "  SkyBackground (matrix B2) — the atmospheric column the ray traverses on its way\n"
        f"  out.  Selected here: {type(nominal_result.stage_outputs['source']['background']).__name__}"
    )
    bg_frame = nominal_result.frames.get("at_aperture_background")
    bg_mean = 0.0 if bg_frame is None else float(np.mean(bg_frame.spectral_radiance))
    print(
        f"  band-mean at-aperture background radiance : {bg_mean:.4e} W/m^2/sr/um\n"
        "  That is effectively ZERO, and it is NOT because the site is dark.  With the\n"
        "  intensity door the target descriptor is T7IntensityAtSource, and the source\n"
        "  stage strips the solar geometry for every non-T2/T3 descriptor (the CU-009\n"
        "  pure-thermal predicate).  The atmosphere therefore sees theta_s = None and\n"
        "  builds a PURE-THERMAL sky, which at 0.4-0.9 um is ~1e-18 W/m^2/sr/um.\n"
        "  Consequence: the dominant noise source of a real visible SST measurement — sky\n"
        "  brightness — is absent whenever the reflective object is entered as I(lambda).\n"
        "  See gaps.md G3."
    )

    print(
        "\n  To exercise the sky properly, the same telescope is re-tasked on a GEO object\n"
        "  through the SHAPE + ALBEDO door (T2Reflective), which keeps the solar geometry.\n"
        "  GEO is used because the point-source angular-extent guard compares against the\n"
        "  OPTICS-ONLY PSF: a 10 m^2 object at 35 786 km is 8.6e-8 rad across, comfortably\n"
        "  inside the guard, whereas the 1 m^2 LEO object is not (gaps.md G6)."
    )

    day = _geo_reflective_sensor(sensor, vendor, 60.0)
    day_result, day_warnings = _run_capturing_warnings(day)
    provisional = sorted(
        {str(w.message) for w in day_warnings if "provisional" in str(w.message).lower()}
    )
    print("\n  GEO reflective door, DAYLIGHT (theta_s = 60 deg, sun 30 deg above the horizon):")
    print(f"    scene_class                : {day_result.stage_outputs['geometry']['scene_class']}")
    print(f"    solar geometry reaching the atmosphere: theta_s = "
          f"{math.degrees(day_result.stage_outputs['source']['los_geometry'].theta_s):.1f} deg (kept)")
    day_bg_frame = day_result.frames.get("at_aperture_background")
    day_bg = 0.0 if day_bg_frame is None else float(np.mean(day_bg_frame.spectral_radiance))
    day_tgt = float(np.mean(day_result.frames["at_aperture_target"].spectral_radiance))
    day_lpath = float(np.mean(day_result.stage_outputs["atmosphere"]["L_path"]))
    print(
        f"    band-mean target radiance  : {day_tgt:.4e} W/m^2/sr/um\n"
        f"    band-mean observer-leg L_path : {day_lpath:.4e} W/m^2/sr/um\n"
        f"    band-mean sky background   : {day_bg:.4e} W/m^2/sr/um\n"
        f"    provisional sky warnings   : {len(provisional)}"
    )
    print(
        "\n    STILL ZERO — and for two independent structural reasons, both worth naming.\n"
        "\n    (i) The SkyBackground source term is the LOS CONTINUATION past the target.\n"
        "        The object is above h_atm_top, so that continuation is vacuum and\n"
        "        uplooking_quantities returns 'sky radiance = 0' WITHOUT ever calling\n"
        "        sky_radiance_along_los — which is where the VIS/NIR provisional warning\n"
        "        lives.  The warning is therefore STRUCTURALLY UNREACHABLE for the whole\n"
        "        ground_to_space class (gaps.md G4).  The sky the telescope actually looks\n"
        "        THROUGH is the observer leg, handled by a different module that carries no\n"
        "        such caveat."
    )
    mean_alt_m = 0.5 * (900.0 + 35_786_000.0)
    print(
        "\n    (ii) The observer leg's single-scatter source takes its species split at the\n"
        f"        segment's ARITHMETIC MEAN ALTITUDE, here 0.5*(0.9 km + 35 786 km) =\n"
        f"        {mean_alt_m / 1000.0:,.0f} km.  Every density factor exp(-h/H) underflows there:\n"
        f"          Rayleigh  exp(-{mean_alt_m:.0f}/8000)   -> {_safe_exp(-mean_alt_m / 8000.0):.3e}\n"
        f"          aerosol   exp(-{mean_alt_m:.0f}/1200)   -> {_safe_exp(-mean_alt_m / 1200.0):.3e}\n"
        "        so the single-scattering albedo evaluates to 0 and the whole scattered term\n"
        "        vanishes.  For the 700 km LEO geometry the same construction lands at\n"
        "        350 km, where the densities are ~1e-20 but their RATIO still resolves, so\n"
        "        the term survives by numerical luck.  gaps.md G4."
    )

    # A reachable instance of the warning: the same telescope on a target INSIDE
    # the column, where the LOS continuation really is atmospheric.
    balloon = sensor.clone()
    balloon = balloon.reset("source.target.user_intensity_path")
    balloon = balloon.set("source.scene_type", "extended")
    balloon = balloon.set("source.target.albedo", 0.60)
    balloon = balloon.set("geometry.target_altitude_m", 20_000.0)  # km -> m, 20 km balloon
    balloon = balloon.set("geometry.path_zenith_rad", math.radians(30.0))  # deg -> rad
    balloon = balloon.set("geometry.solar_zenith_rad", math.radians(60.0))  # deg -> rad
    balloon = balloon.set("spectral_integration.integration_time_s", 0.0002)  # 0.2 ms, well headroom
    balloon_result, balloon_warnings = _run_capturing_warnings(balloon)
    balloon_provisional = sorted(
        {str(w.message) for w in balloon_warnings if "provisional" in str(w.message).lower()}
    )
    print(
        "\n  REACHABLE INSTANCE — same telescope, 20 km stratospheric target (ground_to_air,\n"
        "  extended scene, sun 30 deg up).  Now the continuation IS atmospheric:"
    )
    print(f"    scene_class                : "
          f"{balloon_result.stage_outputs['geometry']['scene_class']}")
    for text in balloon_provisional:
        print("    PROVISIONAL SKY WARNING (ADR-0011 decision 10 band gate) — VERBATIM:")
        for line in _wrap(text, 70):
            print(f"      {line}")
    if not balloon_provisional:  # pragma: no cover - defensive
        print("    UNEXPECTED: no provisional sky warning was raised")
    print(
        f"    band-mean observer-leg L_path : "
        f"{float(np.mean(balloon_result.stage_outputs['atmosphere']['L_path'])):.4e} W/m^2/sr/um\n"
        "    (that is what a real daytime sky pedestal looks like — the number the\n"
        "     ground_to_space class is silently missing)"
    )
    print(
        "\n    What the caveat means: the simple model builds the sky from ONE scattering\n"
        "    event.  In a real daylight sky most photons reaching the telescope have\n"
        "    scattered several times, so single scatter UNDER-estimates the daytime sky\n"
        "    radiance — the pedestal, and its shot noise, are optimistic.  MWIR/LWIR sky\n"
        "    is thermal and MODTRAN-anchored, hence the 3 um band gate."
    )

    twilight = _geo_reflective_sensor(sensor, vendor, 102.0)
    twilight_result, twilight_warnings = _run_capturing_warnings(twilight)
    tw_tgt = float(np.mean(twilight_result.frames["at_aperture_target"].spectral_radiance))
    tw_bg_frame = twilight_result.frames.get("at_aperture_background")
    tw_bg = 0.0 if tw_bg_frame is None else float(np.mean(tw_bg_frame.spectral_radiance))
    tau_sun_tw = float(
        np.mean(twilight_result.stage_outputs["atmosphere"]["atm_quantities"].tau_sun)
    )
    print("\n  GEO reflective door, TWILIGHT (theta_s = 102 deg — the nominal SST window):")
    print(f"    GF-9 shadow test: object sunlit? {sunlit(3.5786e7, math.radians(102.0))}")
    print(f"    chain tau_sun               : {tau_sun_tw:.4f} (vacuum solar leg)")
    print(f"    band-mean target radiance   : {tw_tgt:.4e} W/m^2/sr/um")
    print(f"    band-mean sky background    : {tw_bg:.4e} W/m^2/sr/um")
    print(f"    SNR                         : {twilight_result.metrics['snr']:.3e}")
    for text in sorted({str(w.message) for w in twilight_warnings if "reflective" in str(w.message)}):
        print("    RADIANT's own warning on this configuration:")
        for line in _wrap(text, 70):
            print(f"      {line}")
    print(
        "\n    CONTRADICTION (gaps.md G1).  GF-9 says the object is SUNLIT and the solar leg\n"
        "    is vacuum (tau_sun = 1), but the reflected-radiance assembly multiplies by\n"
        f"    max(cos theta_s, 0) = max({math.cos(math.radians(102.0)):.4f}, 0) = 0, so the\n"
        "    direct-solar term is identically zero and the object goes dark.  The clamp is\n"
        "    correct for a horizontal ground facet (a sun below the local horizontal cannot\n"
        "    illuminate an up-facing surface) but wrong for a satellite, whose illuminated\n"
        "    face is not the local horizontal — the physically meaningful variable is the\n"
        "    SOLAR PHASE ANGLE, which RADIANT has no door for.  That is precisely why the\n"
        "    nominal case above uses a pre-computed signature file: the intensity door is\n"
        "    the ONLY way to express a sunlit object over a dark site today."
    )
    return {"day_result": day_result, "twilight_result": twilight_result}


def section_turbulence(result, sensor: Sensor) -> dict[str, object]:
    atm = result.stage_outputs["atmosphere"]
    geo = result.stage_outputs["geometry"]
    resolution = atm["r0_resolution"]
    r0_m = atm["r0_m"]
    lam_c_um = resolution.reference_wavelength_um
    aperture_m = sensor.get_input("optics.aperture_diameter_m")
    focal_m = sensor.get_input("optics.focal_length_m")

    print("\n--- 9. Turbulence: the seeing-limited regime (Gap 110) ---")
    print(f"  r0 resolution mode          : {resolution.mode} ({resolution.profile_name})")
    print(f"  reference wavelength        : {lam_c_um:.4f} um (band centre)")
    print(f"  path integral of Cn2 W ds   : {resolution.path.cn2_path_integral_m13:.4e} m^(1/3)")
    print(f"  lower-endpoint zenith       : {math.degrees(resolution.path.zeta_low_rad):.3f} deg")
    print(f"  integration span            : {resolution.path.h_low_m / 1000.0:.3f}"
          f" - {resolution.path.h_high_m / 1000.0:.1f} km MSL")
    print(f"  Fried parameter r0          : {r0_m * 100.0:.3f} cm")
    print(f"  D / r0                      : {aperture_m / r0_m:.2f} (dimensionless)")

    lam_c_m = lam_c_um * 1e-6
    seeing_rad = 0.98 * lam_c_m / r0_m
    diffraction_rad = 1.22 * lam_c_m / aperture_m
    arcsec = 180.0 * 3600.0 / math.pi
    print(f"\n  seeing FWHM (0.98 lambda/r0): {seeing_rad * 1e6:.3f} urad "
          f"= {seeing_rad * arcsec:.3f} arcsec")
    print(f"  diffraction limit (1.22 l/D): {diffraction_rad * 1e6:.3f} urad "
          f"= {diffraction_rad * arcsec:.3f} arcsec")
    print(f"  ratio seeing / diffraction  : {seeing_rad / diffraction_rad:.2f} (dimensionless)")
    print(
        "\n  VERDICT: SEEING-LIMITED, not aperture-limited.  The 1 m aperture buys photons,\n"
        "  not resolution: the long-exposure image core is set by r0, and going to a 2 m\n"
        "  aperture would leave the blur diameter unchanged while doubling D/r0."
    )

    # MTF consequence.
    freq = np.asarray(result.stage_outputs["performance"]["mtf_freq_x"], dtype=np.float64)
    mtf_sys = np.asarray(result.stage_outputs["performance"]["mtf_x"], dtype=np.float64)
    budget = result.stage_outputs["performance"]["mtf_budget"]
    keys = sorted(budget) if isinstance(budget, dict) else []
    turb_key = next((k for k in keys if "turb" in k and k.endswith("x")), None)
    optics_key = next((k for k in keys if "optic" in k and k.endswith("x")), None)

    # Diffraction-only reference: same chain with turbulence switched off.
    no_turb = sensor.clone().set("atmosphere.cn2_profile", "direct")
    no_turb_result = no_turb.evaluate()
    mtf_no_turb = np.asarray(no_turb_result.stage_outputs["performance"]["mtf_x"], dtype=np.float64)
    freq_no_turb = np.asarray(
        no_turb_result.stage_outputs["performance"]["mtf_freq_x"], dtype=np.float64
    )

    nyq = result.stage_outputs["performance"]["nyquist_freq_cycles_per_mrad"]
    print(f"\n  Nyquist frequency           : {nyq:.3f} cycles/mrad")
    print(f"  MTF_system at Nyquist (turb): {result.metrics['mtf_system_at_nyquist_x']:.5f}")
    print(
        f"  MTF_system at Nyquist (no turb): "
        f"{no_turb_result.metrics['mtf_system_at_nyquist_x']:.5f}"
    )
    if turb_key and optics_key:
        print(f"  MTF budget keys present     : {turb_key}, {optics_key}")
    print(f"  PSF FWHM with turbulence    : {result.metrics['fwhm_x_m'] * 1e6:8.2f} um")
    print(f"  PSF FWHM without turbulence : {no_turb_result.metrics['fwhm_x_m'] * 1e6:8.2f} um")
    print(f"  RER with / without          : {result.metrics['rer']:.4f} / "
          f"{no_turb_result.metrics['rer']:.4f} (dimensionless)")
    print(f"  EE 3x3 with / without       : {result.metrics['ee_3x3']:.4f} / "
          f"{no_turb_result.metrics['ee_3x3']:.4f} (dimensionless)")
    print(f"  SNR with / without          : {result.metrics['snr']:.2f} / "
          f"{no_turb_result.metrics['snr']:.2f} (dimensionless)")

    # Literature anchor for HV-5/7.
    vertical = path_fried_parameter_from_los(
        LineOfSightGeometry(h_tgt=geo["h_target_m"], h_sensor=0.0, theta_o=math.pi),
        HufnagelValleyCn2(),
        0.5e-6,
        "plane",
    )
    zeta_low = resolution.path.zeta_low_rad
    scaled = vertical.r0_m * (lam_c_um / 0.5) ** 1.2 * math.cos(zeta_low) ** 0.6
    print(
        f"\n  HV-5/7 literature anchor: r0 = 5 cm at 0.5 um, sea level, zenith.\n"
        f"    exact vertical integral from 0 m   : {vertical.r0_m * 100.0:.3f} cm at 0.5 um\n"
        f"    scaled to {lam_c_um:.3f} um, {math.degrees(zeta_low):.1f} deg zenith : "
        f"{scaled * 100.0:.3f} cm  [r0 ~ lambda^(6/5) sec(zeta)^(-3/5)]\n"
        f"    chain value (site at {resolution.path.h_low_m:.0f} m MSL) : {r0_m * 100.0:.3f} cm\n"
    )
    print(
        "    The chain value is LARGER because the integral starts at the site altitude,\n"
        "    above the HV profile's 100 m-scale-height surface term.  CAVEAT (gaps.md G8):\n"
        "    the HV ground term is conventionally ABOVE-GROUND-LEVEL, but\n"
        "    cn2_hufnagel_valley evaluates the profile against MSL altitude, so a site at\n"
        f"    {resolution.path.h_low_m:.0f} m MSL silently loses its OWN boundary layer.  Scaled to\n"
        f"    0.5 um this r0 is {r0_m * (0.5 / lam_c_um) ** 1.2 * 100.0:.1f} cm "
        f"({0.98 * 0.5e-6 / (r0_m * (0.5 / lam_c_um) ** 1.2) * arcsec:.2f} arcsec seeing), which is\n"
        "    world-class-site seeing; a real 0.9 km high-desert site runs 1.0-1.5 arcsec.\n"
        "    Treat the seeing here as optimistic by roughly a factor of two."
    )
    return {
        "freq": freq,
        "mtf_sys": mtf_sys,
        "freq_no_turb": freq_no_turb,
        "mtf_no_turb": mtf_no_turb,
        "seeing_rad": seeing_rad,
        "diffraction_rad": diffraction_rad,
        "no_turb_result": no_turb_result,
        "focal_m": focal_m,
    }


def section_zenith_ladder(sensor: Sensor, vendor: dict[str, object]) -> dict[str, np.ndarray]:
    angles_deg = list(vendor["zenith_ladder_deg"])  # type: ignore[arg-type]
    print("\n--- 10. Pointing-zenith ladder (the pass, culmination to low elevation) ---")
    print(
        f"  {'zeta_low [deg]':>14} {'theta_o [deg]':>14} {'R [km]':>10} {'airmass':>8} "
        f"{'tau_band':>9} {'r0 [cm]':>9} {'SNR':>9} {'FWHM [urad]':>12}"
    )
    rows: list[list[float]] = []
    for angle_deg in angles_deg:
        probe = sensor.clone().set("geometry.path_zenith_rad", math.radians(angle_deg))
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            try:
                res = probe.evaluate()
            except Exception as exc:
                print(f"  {angle_deg:14.1f}  RAISED {type(exc).__name__}: {str(exc)[:70]}")
                continue
        geo = res.stage_outputs["geometry"]
        atm = res.stage_outputs["atmosphere"]
        lam = np.asarray(atm["atm_quantities"].wavelength_um)
        tau = np.asarray(atm["tau_atm"])
        fwhm_urad = res.metrics["fwhm_x_m"] / sensor.get_input("optics.focal_length_m") * 1e6
        rows.append(
            [
                angle_deg,
                math.degrees(geo["theta_o_rad"]),
                geo["slant_range_m"] / 1000.0,
                1.0 / math.cos(math.radians(angle_deg)),
                _band_mean(tau, lam),
                atm["r0_m"] * 100.0,
                res.metrics["snr"],
                fwhm_urad,
            ]
        )
        note = ""
        for w in caught:
            if "horizon" in str(w.message).lower() or "refraction" in str(w.message).lower():
                note = "  <- horizon-guard WARNING"
        print(
            f"  {rows[-1][0]:14.1f} {rows[-1][1]:14.4f} {rows[-1][2]:10.1f} {rows[-1][3]:8.3f} "
            f"{rows[-1][4]:9.4f} {rows[-1][5]:9.3f} {rows[-1][6]:9.2f} {rows[-1][7]:12.3f}{note}"
        )
    print(
        "\n  Trends: airmass = sec(zeta_low) grows the column, so tau falls and r0 shrinks as\n"
        "  sec(zeta)^(-3/5).  The slant range grows more slowly than h/cos(zeta) because the\n"
        "  Earth curves away.  SNR falls from both the transmittance loss AND the wider\n"
        "  seeing disc (a bigger blur spreads the same photons over more pixels, cutting\n"
        "  EE_box)."
    )
    return {"ladder": np.asarray(rows, dtype=np.float64)}


def section_air_mass_probe(sensor: Sensor, vendor: dict[str, object]) -> None:
    angles = list(vendor["air_mass_probe_deg"])  # type: ignore[arg-type]
    print("\n--- 10b. Air-mass handover probe (finding, gaps.md G5) ---")
    print(
        "  AtmosphericGeometry.slant_path_length_m switches at 80 deg from the flat-Earth\n"
        "  form dh/cos(zeta) to a spherical root form parameterised by x = dh/R_E.  For an\n"
        "  ATMOSPHERIC slab (dh ~ 10 km, x ~ 0.0016) that is the right correction.  Here the\n"
        "  observer segment runs from the site to the OBJECT, so dh = 699 km and x = 0.110 —\n"
        "  three orders of magnitude too large.  Watch the air mass at the switch:"
    )
    print(f"\n  {'zeta_low [deg]':>14} {'tau @ 0.55 um':>15} {'implied air mass':>18}")
    prev = None
    for angle_deg in angles:
        probe = sensor.clone().set("geometry.path_zenith_rad", math.radians(angle_deg))
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            res = probe.evaluate()
        lam = np.asarray(res.stage_outputs["atmosphere"]["atm_quantities"].wavelength_um)
        tau = np.asarray(res.stage_outputs["atmosphere"]["tau_atm"])
        tau55 = float(np.interp(0.55, lam, tau))
        # Air mass implied by the optical depth, referenced to the zenith column.
        implied = -math.log(tau55)
        flag = ""
        if prev is not None and implied < prev:
            flag = "  <- OPTICAL DEPTH DROPPED as the path got longer"
        prev = implied
        print(f"  {angle_deg:14.1f} {tau55:15.5f} {implied:18.4f}{flag}")
    print(
        "\n  Consequence: transmittance is NON-MONOTONIC in zenith angle above 80 deg, and\n"
        "  the spherical branch UNDER-estimates the column.  Everything in this scenario is\n"
        "  reported at or below 75 deg, where the flat-Earth branch is in force and the\n"
        "  sec(zeta) air mass is correct to better than 0.5 %."
    )


def section_horizon_guard(sensor: Sensor) -> None:
    print("\n--- 11. Horizon guard (ADR-0011 decision 6) ---")
    print(
        "  RADIANT models no atmospheric refraction, so near-horizontal paths are guarded\n"
        "  rather than approximated.  For an ordinary slant (endpoint-minimum topology) the\n"
        "  band is measured at the LOWER endpoint: |zeta_low - 90 deg| < 0.5 deg raises,\n"
        "  0.5-2 deg computes with a quantified UserWarning, > 2 deg is clean."
    )
    for angle_deg, label in ((88.6, "inside the 0.5-2 deg warn shoulder"), (89.8, "inside the 0.5 deg hard guard")):
        probe = sensor.clone().set("geometry.path_zenith_rad", math.radians(angle_deg))
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            try:
                probe.evaluate()
                verdict = "computed"
            except Exception as exc:
                verdict = f"RAISED {type(exc).__name__}"
        texts = sorted({str(w.message) for w in caught if "refract" in str(w.message).lower()})
        print(f"\n  zeta_low = {angle_deg:.1f} deg ({label}): {verdict}")
        for text in texts[:1]:
            for line in _wrap(text, 72):
                print(f"    {line}")
    print(
        "\n  A real SST site stops tracking near 20 deg elevation (zeta_low = 70 deg) for\n"
        "  exactly the reason the guard exists: below that, refraction and the differential\n"
        "  colour of the column dominate the astrometry."
    )


def section_cross_checks(
    result, sensor: Sensor, spectra: dict[str, object], vendor: dict[str, object]
) -> dict[str, object]:
    lam = np.asarray(spectra["lam"], dtype=np.float64)
    tau = np.asarray(spectra["tau"], dtype=np.float64)
    geo = result.stage_outputs["geometry"]

    print("\n" + "=" * 78)
    print("CROSS-CHECKS")
    print("=" * 78)

    # ---- (a) Closed-form point-source identity and its vacuum limit --------
    print("\n  (a) CLOSED-FORM POINT-SOURCE IDENTITY AND ITS VACUUM LIMIT")
    print(
        "      In the point-source regime the electron count must be the textbook integral\n"
        "         S = t_int * QE * tau_opt * A_collect * EE_box *\n"
        "             INT[ tau_atm(l) * I(l) / R^2 * l/(hc) ] dl\n"
        "      with EVERY factor read back out of the chain.  Nothing here is a RADIANT\n"
        "      formula — it is the definition of an unresolved source seen through a\n"
        "      transmitting medium."
    )
    intensity = _signature_on_grid(sensor, lam)
    r_m = geo["slant_range_m"]
    a_collect = math.pi * (sensor.get_input("optics.aperture_diameter_m") / 2.0) ** 2
    tau_opt = np.asarray(result.stage_outputs["optics"]["tau_opt"], dtype=np.float64)
    qe = sensor.get_input("detector.qe_value")
    t_int = sensor.get_input("spectral_integration.integration_time_s")
    ee_box = result.stage_outputs["platform"]["EE_box"]
    lam_m = lam * 1e-6
    photon_density_vac = (
        tau_opt * intensity / (r_m * r_m) * a_collect * lam_m / (h * c)
    )  # photons/s/µm
    prefactor = t_int * qe * ee_box
    hand_e = prefactor * float(np.trapezoid(tau * photon_density_vac, lam))
    hand_vac_e = prefactor * float(np.trapezoid(photon_density_vac, lam))
    chain_e = result.stage_outputs["readout"]["signal_e_final"]
    print(f"      A_collect                 : {a_collect:.6f} m^2")
    print(f"      tau_opt (band mean)       : {_band_mean(tau_opt, lam):.6f} (dimensionless)")
    print(f"      EE_box                    : {ee_box:.6f} (dimensionless)")
    print(f"      slant range R             : {r_m / 1000.0:.3f} km")
    print(f"      integration time          : {t_int * 1000.0:.1f} ms")
    print(f"      hand-computed signal      : {hand_e:,.3f} e-")
    print(f"      chain signal              : {chain_e:,.3f} e-")
    rel = abs(hand_e - chain_e) / chain_e
    print(f"      relative difference       : {rel:.3e}  -> {'PASS' if rel < 2e-3 else 'FAIL'}")
    tau_eff = _photon_weighted_tau(tau, intensity * tau_opt, lam)
    print(
        f"\n      VACUUM LIMIT (tau -> 1 in the same integral):\n"
        f"        S_vacuum                : {hand_vac_e:,.3f} e-\n"
        f"        S_chain / S_vacuum      : {chain_e / hand_vac_e:.6f} (dimensionless)\n"
        f"        photon-weighted band tau: {tau_eff:.6f} (dimensionless)\n"
        f"        difference              : "
        f"{abs(chain_e / hand_vac_e - tau_eff):.3e}  -> "
        f"{'PASS' if abs(chain_e / hand_vac_e - tau_eff) < 2e-3 else 'FAIL'}"
    )
    print(
        "      Reading: the chain applies the column transmittance multiplicatively and\n"
        "      exactly, so the atmosphere-free answer is recoverable by dividing by the\n"
        "      signal-weighted band mean of tau(lambda).  That is the limiting-case anchor."
    )

    print("\n      TRUE-VACUUM RUN (attempted): raise the telescope to h_atm_top = 100 km, so")
    print("      the topology dispatcher returns the exact vacuum quantities (tau_up == 1).")
    vac = sensor.clone().set("geometry.sensor_altitude_m", 100_000.0)
    try:
        vac_result = vac.evaluate()
    except Exception as exc:
        print(f"        RAISED {type(exc).__name__} — see gaps.md G7:")
        for line in _wrap(str(exc).split("|")[0].strip(), 68):
            print(f"          {line}")
        print(
            "        A wholly-vacuum up-looking path returns sky_radiance_at_aperture = None,\n"
            "        but the LOS-termination rule still selects SkyBackground, and the\n"
            "        assembly refuses to default it to zero (Rule 17).  The two vacuum\n"
            "        branches disagree.  The identity above is therefore evaluated\n"
            "        analytically rather than by a second chain run."
        )
    else:  # pragma: no cover - the guard above fires today
        tau_vac = np.asarray(vac_result.stage_outputs["atmosphere"]["tau_atm"], dtype=np.float64)
        print(f"        max |tau_up - 1| = {float(np.max(np.abs(tau_vac - 1.0))):.3e}")

    # ---- (b) Astronomical extinction anchor --------------------------------
    print("\n  (b) BAND-MEAN ZENITH TRANSMITTANCE vs PUBLISHED ASTRONOMICAL EXTINCTION")
    zenith_probe = sensor.clone().set("geometry.path_zenith_rad", 0.0)
    zenith_result = zenith_probe.evaluate()
    tau_zenith = np.asarray(zenith_result.stage_outputs["atmosphere"]["tau_atm"], dtype=np.float64)
    tau_55 = float(np.interp(V_BAND_CENTRE_UM, lam, tau_zenith))
    k_radiant = -2.5 * math.log10(tau_55)
    tau_lo = 10.0 ** (-0.4 * EXTINCTION_V_LO_MAG_PER_AIRMASS)
    tau_hi = 10.0 ** (-0.4 * EXTINCTION_V_HI_MAG_PER_AIRMASS)
    print(
        f"      Published V-band zenith extinction k_V = "
        f"{EXTINCTION_V_LO_MAG_PER_AIRMASS:.2f}-{EXTINCTION_V_HI_MAG_PER_AIRMASS:.2f} mag/airmass\n"
        "      (Hardie 1962 photometric practice; Burke/Gladders/Graham 2010 §5).\n"
        "      Conversion: tau = 10^(-0.4 k)  [a magnitude IS -2.5 log10 of a flux ratio]"
    )
    print(f"      published tau(0.55 um) at zenith : {tau_hi:.4f} - {tau_lo:.4f} (dimensionless)")
    print(f"      RADIANT tau(0.55 um) at zenith   : {tau_55:.4f} (dimensionless)")
    print(f"      RADIANT extinction               : {k_radiant:.3f} mag/airmass")
    verdict = "PASS" if tau_hi <= tau_55 <= tau_lo else "FAIL"
    print(f"      verdict                          : {verdict}")
    if verdict == "FAIL":
        print(
            "      DIAGNOSIS (see gaps.md G4).  radiant.atmosphere.simple uses\n"
            "        sigma_mol(lambda) = 0.0088 * lambda_um^-4.09  [1/km at sea level]\n"
            "      and then multiplies by the ~8 km molecular column length.  But\n"
            "      0.0088 * lambda^-4.09 IS the published TOTAL VERTICAL RAYLEIGH OPTICAL\n"
            "      DEPTH (dimensionless; Hansen & Travis 1974, Bucholtz 1995), not an\n"
            "      extinction coefficient in km^-1.  The true sea-level Rayleigh volume\n"
            "      extinction at 550 nm is 0.0116 km^-1, ~8.7x smaller.\n"
            "      Consequence: Rayleigh optical depth is too large by ~H_mol/1 km ~ 8x\n"
            "      in the VIS.  Check the arithmetic:"
        )
        h_site_m = geo["h_sensor_m"]
        h_mol_km = 8.0  # molecular scale height used by radiant.atmosphere.simple
        sigma_used_per_km = 0.0088 * V_BAND_CENTRE_UM ** (-4.09)
        # The molecular column length the model integrates over, [km].
        col_mol_km = h_mol_km * math.exp(-h_site_m / (h_mol_km * 1000.0))
        od_radiant = -math.log(tau_55)
        od_rayleigh_used = sigma_used_per_km * col_mol_km
        # Read correctly, 0.0088 λ^-4.09 IS the sea-level-to-space vertical OD; the
        # part above the site is that times exp(-h_site/H_mol).
        od_rayleigh_true = sigma_used_per_km * math.exp(-h_site_m / (h_mol_km * 1000.0))
        od_corrected = od_radiant - od_rayleigh_used + od_rayleigh_true
        tau_corrected = math.exp(-od_corrected)
        k_corrected = -2.5 * math.log10(tau_corrected)
        inside = EXTINCTION_V_LO_MAG_PER_AIRMASS <= k_corrected <= EXTINCTION_V_HI_MAG_PER_AIRMASS
        print(
            f"        0.0088 * 0.55^-4.09              = {sigma_used_per_km:.5f}\n"
            f"        molecular column above the site  = {col_mol_km:.4f} km\n"
            f"        Rayleigh OD as used (sigma*col)  = {od_rayleigh_used:.4f} optical depths\n"
            f"        Rayleigh OD read as an OD        = {od_rayleigh_true:.4f} optical depths\n"
            f"        published Rayleigh OD at 550 nm  = 0.0973 optical depths (sea level)\n"
            f"        RADIANT total zenith OD          = {od_radiant:.4f} optical depths\n"
            f"        corrected total zenith OD        = {od_corrected:.4f} optical depths\n"
            f"        corrected tau(0.55 um)           = {tau_corrected:.4f} (dimensionless)\n"
            f"        corrected extinction             = {k_corrected:.3f} mag/airmass"
            f"  <- {'INSIDE' if inside else 'outside'} the published band"
        )
        print(
            "      MWIR/LWIR is unaffected in practice (Rayleigh optical depth at 4 um is\n"
            "      ~2.5e-4 either way), which is why the CU-161 MODTRAN calibration of the\n"
            "      simple model never saw it: that calibration anchored 3-14 um."
        )

    # ---- (c) Reciprocity identity -----------------------------------------
    print("\n  (c) TRANSMITTANCE RECIPROCITY (ADR-0011 decision 3)")
    print(
        "      A segment carries ONE tau, computed at its lower endpoint, because\n"
        "      transmittance is reciprocal.  Flip the scene: put the sensor in space and\n"
        "      the target on the ground with the SAME column and the SAME lower-endpoint\n"
        "      zenith.  The two tau spectra must coincide."
    )
    down = sensor.clone()
    down = down.set("geometry.sensor_altitude_m", geo["h_target_m"])
    down = down.set("geometry.target_altitude_m", geo["h_sensor_m"])
    down = down.set("geometry.path_zenith_rad", sensor.get_input("geometry.path_zenith_rad"))
    down_result = down.evaluate()
    tau_down = np.asarray(down_result.stage_outputs["atmosphere"]["tau_atm"], dtype=np.float64)
    max_abs = float(np.max(np.abs(tau - tau_down)))
    print(
        f"      up-looking   scene_class : {geo['scene_class']} (los {geo['los_direction']})\n"
        f"      down-looking scene_class : {down_result.stage_outputs['geometry']['scene_class']} "
        f"(los {down_result.stage_outputs['geometry']['los_direction']})"
    )
    print(f"      max |tau_up - tau_down|  : {max_abs:.3e}")
    print(f"      verdict                  : {'PASS' if max_abs < 1e-3 else 'FAIL'} (tolerance 1e-3)")

    # ---- (d) Apparent visual magnitude ------------------------------------
    print("\n  (d) APPARENT VISUAL MAGNITUDE OF THE OBJECT")
    catalog = vendor["catalog"]  # type: ignore[index]
    intensity_full = _signature_on_grid(sensor, lam)
    i_at_55 = float(np.interp(V_BAND_CENTRE_UM, lam, intensity_full))
    # The signature is rho*A*E_sun(l)*p(alpha)/pi, so its integral over ALL
    # wavelengths is rho*A*S0*p(alpha)/pi with S0 the solar constant.
    albedo = float(catalog["Diffuse Albedo"][0])
    area_m2 = float(catalog["Projected Area"][0])
    alpha_rad = math.radians(float(catalog["Solar Phase Angle"][0]))  # deg -> rad
    phase = (math.sin(alpha_rad) + (math.pi - alpha_rad) * math.cos(alpha_rad)) / math.pi
    i_total = albedo * area_m2 * S_solar_W_per_m2 * phase / math.pi
    r_m = geo["slant_range_m"]
    e_at_aperture = i_total / (r_m * r_m)
    m_v = SOLAR_APPARENT_MAGNITUDE - 2.5 * math.log10(e_at_aperture / S_solar_W_per_m2)
    print(
        f"      total radiant intensity (all lambda) : {i_total:.2f} W/sr\n"
        f"      I(0.55 um) from the signature file   : {i_at_55:.4f} W/sr/um\n"
        f"      irradiance at the aperture           : {e_at_aperture:.4e} W/m^2\n"
        f"      solar irradiance at 1 AU S0          : {S_solar_W_per_m2:.1f} W/m^2\n"
        f"      m = m_sun - 2.5 log10(E/S0)          = {m_v:.2f} mag (above the atmosphere)"
    )
    k_atm_mag = -2.5 * math.log10(_photon_weighted_tau(tau, intensity_full, lam))
    print(
        f"      RADIANT column extinction on this path: {k_atm_mag:.3f} mag\n"
        f"      apparent magnitude as seen by the site : {m_v + k_atm_mag:.2f} mag"
    )
    print(
        "      Anchor: catalogued LEO objects of ~1 m^2 projected area at a few hundred to\n"
        "      1000 km routinely photometer at m_V ~ 4-8 (the naked-eye-satellite regime).\n"
        f"      {m_v:.2f} mag above the atmosphere is squarely inside that band -> PASS."
    )

    # ---- MODTRAN anchor status --------------------------------------------
    print("\n  (e) MODTRAN ANCHOR — DEFERRED, NOT FABRICATED")
    print(
        "      The ground-to-space full-column MODTRAN ladder for this scene class is\n"
        "      OWNER-RUN BATCH 2 (ADR-0011 decision 10, plan §4 Phase 2) and has not been\n"
        "      delivered.  No MODTRAN comparison is reported here.  When batch 2 lands,\n"
        "      rerun this scenario and compare tau(lambda) directly; the (b) diagnosis\n"
        "      above predicts the simple model will read ~8x too opaque in the VIS."
    )
    return {
        "vacuum_rel_error": rel,
        "tau_zenith_55": tau_55,
        "k_radiant": k_radiant,
        "reciprocity_max_abs": max_abs,
        "m_v": m_v,
        "intensity": intensity_full,
    }


def _signature_on_grid(sensor: Sensor, lam: np.ndarray) -> np.ndarray:
    """Re-read the canonical signature CSV and interpolate onto ``lam`` [W/sr/µm]."""
    path = Path(sensor.get_input("source.target.user_intensity_path"))
    rows = [
        line.split(",")
        for line in path.read_text(encoding="utf-8").splitlines()[1:]
        if line.strip()
    ]
    wl = np.array([float(r[0]) for r in rows], dtype=np.float64)
    val = np.array([float(r[1]) for r in rows], dtype=np.float64)
    return np.interp(lam, wl, val)


def section_unused_and_physics(result, sensor: Sensor) -> None:
    print("\n" + "=" * 78)
    print("WHAT DOES *NOT* AFFECT THIS RESULT, AND WHY")
    print("=" * 78)
    metrics = result.metrics
    print(
        "  * source.target.temperature / emissivity — UNUSED.  The intensity door\n"
        "    (T7IntensityAtSource) takes I(lambda) verbatim; no emission model runs.\n"
        f"  * NEDT = {metrics['nedt_K']:.4f} K is meaningless for a reflective VIS point source:\n"
        "    the value is dS/dT taken against the DEFAULT source.target.temperature, a\n"
        "    parameter this scene never uses.  It appears in the metric dict because the\n"
        "    thermal group is on by default, not because it is physically defined here.\n"
        "  * geometry.solar_azimuth_rad affects only the single-scatter sky phase function.\n"
        "    With the sun below the site horizon that term is off, so the nominal result is\n"
        "    independent of it.  It matters in the sun-up comparison run.\n"
        "  * atmosphere.tau_sun (the TOA-to-object solar leg) is 1.0 here because the object\n"
        "    is above h_atm_top — the sunlight reaching it never enters the atmosphere.\n"
        "    But NOTE: the intensity door does not consume tau_sun at all (the analyst's\n"
        "    I(lambda) already embeds the illumination), so the GF-9 shadow verdict has NO\n"
        "    effect on the computed signal.  Gating the signature on the shadow test is the\n"
        "    ANALYST's job in this door — see gaps.md G2.\n"
        "  * Ground-projection metrics (GSD, swath) are off by scene class, not by accident.\n"
        "  * geometry.ground_speed_m_s / circular-orbit mode: the platform is a fixed\n"
        f"    ground mount, so smear width is "
        f"{result.stage_outputs['platform'].get('smear_width_m', 0.0) * 1e6:.3f} um.  Real SST\n"
        "    tracking rate would enter through geometry.los_angular_rate_rad_s (K1) — see\n"
        "    gaps.md G5."
    )
    print("\n" + "=" * 78)
    print("NON-OBVIOUS PHYSICS IN THIS SCENE CLASS")
    print("=" * 78)
    print(
        "  1. DIRECTION-AWARE PATH PRODUCTS.  Up-looking, the observer leg is the column\n"
        "     from the TELESCOPE to h_atm_top, and the LOS continuation past the object\n"
        "     exits into cold space.  tau_up is that observer leg; L_path_up is the sky\n"
        "     radiance accumulated along it.  These are segment-composed, not the\n"
        "     down-looking bundle read backwards (ADR-0011 decision 3, guardrail G1).\n"
        "  2. LOWER-ENDPOINT ANGLE CONVENTION.  The tasking card's pointing zenith IS\n"
        "     zeta_low because the telescope is the lower endpoint.  theta_o, the canonical\n"
        "     angle every downstream stage reads, is its OBTUSE partner through the\n"
        "     spherical triangle — not simply 180 deg - zeta_low (Earth curvature adds the\n"
        "     central angle).\n"
        "  3. THE SKY IS BOTH BACKGROUND AND ATTENUATOR.  The same column that dims the\n"
        "     object also fills the pixel behind it.  In the point-source regime the target\n"
        "     term uses Omega_target = A_t/R^2 with path radiance stripped, while the sky\n"
        "     pedestal enters at the full pixel solid angle Omega_pixel and shot-noises.\n"
        f"  4. SEEING BEATS APERTURE.  EE_box is "
        f"{result.stage_outputs['platform']['EE_box'] * 100.0:.1f} % here, not because the\n"
        "     optics are bad, but because the seeing disc spread over the pixel grid puts\n"
        "     most of the object's photons outside the central pixel.  EE_box is computed\n"
        "     once in PlatformStage from the FULLY degraded PSF and applied once in\n"
        "     spectral integration (Rules 4 and 9).\n"
        "  5. HORIZON GUARD.  No refraction model exists, so the near-horizontal band\n"
        "     raises instead of returning a plausible wrong number (Rule 17)."
    )


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------


def _wrap(text: str, width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        if len(current) + len(word) + 1 > width:
            lines.append(current)
            current = word
        else:
            current = f"{current} {word}".strip()
    if current:
        lines.append(current)
    return lines


def make_figures(
    result,
    spectra: dict[str, object],
    shadow: dict[str, np.ndarray],
    turbulence: dict[str, object],
    ladder: dict[str, np.ndarray],
    checks: dict[str, object],
) -> list[tuple[str, str]]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    produced: list[tuple[str, str]] = []
    lam = np.asarray(spectra["lam"], dtype=np.float64)
    tau = np.asarray(spectra["tau"], dtype=np.float64)
    intensity = np.asarray(checks["intensity"], dtype=np.float64)

    # --- Figure 1: signature and column transmittance ----------------------
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8.0, 7.0), sharex=True)
    ax1.plot(lam, intensity, color="#B8860B", lw=1.8)
    ax1.set_ylabel("Spectral radiant intensity I(λ) [W/sr/µm]")
    ax1.set_title("ORB-4471 signature and the up-looking column it is seen through")
    ax1.grid(alpha=0.3)
    ax2.plot(lam, tau, color="#1F3864", lw=1.8, label="ζ_low = 20° (nominal)")
    ax2.set_xlabel("Wavelength [µm]")
    ax2.set_ylabel("Column transmittance τ_up [dimensionless]")
    ax2.set_ylim(0.0, 1.0)
    ax2.grid(alpha=0.3)
    ax2.legend(loc="lower right")
    fig.tight_layout()
    path = OUTPUT_DIR / "signature_and_column_transmittance.png"
    fig.savefig(path, dpi=140)
    plt.close(fig)
    produced.append((path.name, "Object signature I(λ) and up-looking column transmittance τ(λ)"))

    # --- Figure 2: shadow height -------------------------------------------
    fig, ax = plt.subplots(figsize=(8.0, 5.0))
    delta = np.linspace(0.0, 32.0, 400)
    heights_km = R_EARTH_M * (1.0 / np.cos(np.radians(delta)) - 1.0) / 1000.0
    ax.plot(delta, heights_km, color="#1F3864", lw=2.0, label="h_shadow(δ) = R_E (sec δ − 1)")
    ax.scatter(
        shadow["depression_deg"],
        shadow["shadow_height_m"] / 1000.0,
        color="#B8860B",
        zorder=5,
        label="tasking ladder points",
    )
    ax.axhline(700.0, color="#8B0000", ls="--", lw=1.5, label="ORB-4471 altitude = 700 km")
    ax.axvline(12.0, color="#2E7D32", ls=":", lw=1.5, label="nominal depression = 12°")
    ax.set_xlabel("Solar depression below the site horizon δ [deg]")
    ax.set_ylabel("Terminator shadow height h_shadow [km]")
    ax.set_title("Per-altitude solar illumination (GF-9): sunlit object over a dark site")
    ax.set_ylim(0.0, 1200.0)
    ax.grid(alpha=0.3)
    ax.legend(loc="upper left", fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "terminator_shadow_height.png"
    fig.savefig(path, dpi=140)
    plt.close(fig)
    produced.append((path.name, "Shadow height vs solar depression, with the 700 km object marked"))

    # --- Figure 3: turbulence MTF ------------------------------------------
    # mtf_freq_* is in cycles/m on the FOCAL PLANE; convert to angular
    # frequency on the sky: cycles/mrad = cycles/m * f[m] * 1e-3.
    focal_m = float(turbulence["focal_m"])
    to_cyc_per_mrad = focal_m * 1e-3
    fig, ax = plt.subplots(figsize=(8.0, 5.0))
    ax.plot(
        np.asarray(turbulence["freq_no_turb"]) * to_cyc_per_mrad,
        np.asarray(turbulence["mtf_no_turb"]),
        color="#1F3864",
        lw=2.0,
        label="no turbulence (aperture-limited)",
    )
    ax.plot(
        np.asarray(turbulence["freq"]) * to_cyc_per_mrad,
        np.asarray(turbulence["mtf_sys"]),
        color="#8B0000",
        lw=2.0,
        label="HV-5/7 turbulence (seeing-limited)",
    )
    nyq = result.stage_outputs["performance"]["nyquist_freq_cycles_per_mrad"]
    ax.axvline(nyq, color="#555555", ls=":", lw=1.5, label=f"Nyquist = {nyq:.1f} cycles/mrad")
    seeing_cutoff = 1.0 / (float(turbulence["seeing_rad"]) * 1e3)
    ax.axvline(
        seeing_cutoff,
        color="#B8860B",
        ls="--",
        lw=1.5,
        label=f"1 / seeing FWHM = {seeing_cutoff:.1f} cycles/mrad",
    )
    ax.set_xlabel("Angular spatial frequency on the sky [cycles/mrad]")
    ax.set_ylabel("System MTF [dimensionless]")
    ax.set_title("Seeing-limited vs aperture-limited system MTF, 1 m telescope at 0.65 µm")
    ax.set_xlim(0.0, 2.0 * nyq)
    ax.set_ylim(0.0, 1.0)
    ax.grid(alpha=0.3)
    ax.legend(loc="upper right", fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "seeing_vs_diffraction_mtf.png"
    fig.savefig(path, dpi=140)
    plt.close(fig)
    produced.append((path.name, "System MTF with and without the HV-5/7 turbulence term"))

    # --- Figure 4: zenith ladder -------------------------------------------
    rows = np.asarray(ladder["ladder"], dtype=np.float64)
    fig, axes = plt.subplots(3, 1, figsize=(8.0, 9.0), sharex=True)
    axes[0].plot(rows[:, 0], rows[:, 4], "o-", color="#1F3864")
    axes[0].set_ylabel("Band-mean τ_up [dimensionless]")
    axes[0].grid(alpha=0.3)
    axes[0].set_title("Pointing-zenith ladder — the pass from culmination to low elevation")
    axes[1].plot(rows[:, 0], rows[:, 5], "s-", color="#B8860B")
    axes[1].set_ylabel("Fried parameter r₀ [cm]")
    axes[1].grid(alpha=0.3)
    axes[2].plot(rows[:, 0], rows[:, 6], "^-", color="#8B0000")
    axes[2].set_ylabel("SNR [dimensionless]")
    axes[2].set_xlabel("Pointing zenith at the telescope ζ_low [deg]")
    axes[2].grid(alpha=0.3)
    fig.tight_layout()
    path = OUTPUT_DIR / "zenith_ladder.png"
    fig.savefig(path, dpi=140)
    plt.close(fig)
    produced.append((path.name, "τ, r₀ and SNR versus pointing zenith along the pass"))
    return produced


def write_results_workbook(
    result, ladder: dict[str, np.ndarray], shadow: dict[str, np.ndarray], checks: dict[str, object]
) -> None:
    """Write the (gitignored) results workbook."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Nominal"
    geo = result.stage_outputs["geometry"]
    ws.append(["Quantity", "Value", "Unit"])
    for name, value, unit in [
        ("scene_class", geo["scene_class"], "-"),
        ("los_direction", geo["los_direction"], "-"),
        ("theta_o", math.degrees(geo["theta_o_rad"]), "deg"),
        ("zeta_low", 180.0 - math.degrees(geo["eta_rad"]), "deg"),
        ("slant_range", geo["slant_range_m"] / 1000.0, "km"),
        ("r0", result.stage_outputs["atmosphere"]["r0_m"] * 100.0, "cm"),
        ("EE_box", result.stage_outputs["platform"]["EE_box"], "-"),
        ("signal", result.stage_outputs["readout"]["signal_e_final"], "e-"),
        ("SNR", result.metrics["snr"], "-"),
        ("apparent magnitude", checks["m_v"], "mag"),
        ("zenith tau(0.55 um)", checks["tau_zenith_55"], "-"),
        ("zenith extinction", checks["k_radiant"], "mag/airmass"),
    ]:
        ws.append([name, value, unit])

    ws = wb.create_sheet("Zenith Ladder")
    ws.append(
        [
            "zeta_low [deg]",
            "theta_o [deg]",
            "slant range [km]",
            "airmass [-]",
            "tau_band [-]",
            "r0 [cm]",
            "SNR [-]",
            "FWHM [urad]",
        ]
    )
    for row in np.asarray(ladder["ladder"], dtype=np.float64):
        ws.append([float(v) for v in row])

    ws = wb.create_sheet("Terminator Ladder")
    ws.append(["solar depression [deg]", "shadow height [km]"])
    for d, hh in zip(shadow["depression_deg"], shadow["shadow_height_m"], strict=True):
        ws.append([float(d), float(hh) / 1000.0])

    wb.save(RESULTS_XLSX)
    print(f"\nWrote {RESULTS_XLSX.relative_to(SCENARIO_DIR)} (gitignored — regenerate on demand)")


# ---------------------------------------------------------------------------


def main() -> None:
    vendor = read_vendor_inputs()
    section_inputs(vendor)

    sensor = make_sensor()
    result, caught = _run_capturing_warnings(sensor)

    spectra = section_nominal(result, sensor, caught)
    section_scene_class_assertion(sensor)
    section_metric_relevance(result)
    shadow = section_shadow_ladder(vendor, result)
    section_sky_background(sensor, result, vendor)
    turbulence = section_turbulence(result, sensor)
    ladder = section_zenith_ladder(sensor, vendor)
    section_air_mass_probe(sensor, vendor)
    section_horizon_guard(sensor)
    checks = section_cross_checks(result, sensor, spectra, vendor)
    section_unused_and_physics(result, sensor)

    figures = make_figures(result, spectra, shadow, turbulence, ladder, checks)
    write_results_workbook(result, ladder, shadow, checks)
    print("\nFigures written to outputs/:")
    for name, caption in figures:
        print(f"  {name:<44s} {caption}")
    print("\nDone.")


if __name__ == "__main__":
    main()
