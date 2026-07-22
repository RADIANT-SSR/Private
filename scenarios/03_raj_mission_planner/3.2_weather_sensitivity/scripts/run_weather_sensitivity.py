"""Scenario 3.2: Weather Sensitivity — How Bad Can the Weather Get?

Raj needs a "go/no-go" weather threshold for his MWIR reconnaissance sensor.
At what visibility does NIIRS drop below the mission requirement (NIIRS ≥ 4)?
How does precipitable water vapor affect performance?

This script:
  1. Reads Raj's spreadsheet (cm, %, km, ms, °C, deg)
  2. Converts to RADIANT canonical units
  3. Sweeps visibility from 2–100 km → NIIRS, SNR, transmittance at each
  4. Sweeps PWV from 0.5–5.0 cm → NIIRS, SNR, transmittance at each
  5. Evaluates 8 named weather conditions (crystal clear → heavy haze)
  6. Finds critical visibility threshold (NIIRS = 4.0)
  7. 2D grid: NIIRS as function of visibility × PWV (go/no-go table)
  8. Writes results to output spreadsheet

Usage:
    python run_weather_sensitivity.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from radiant.api import Sensor

# ---------------------------------------------------------------------------
# Step 1: Read Raj's spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "raj_weather_sensitivity_data.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "weather_sensitivity_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)


def _read_sheet(ws, min_row: int = 5):
    """Read parameter name→value and name→unit from a spreadsheet sheet."""
    specs: dict[str, object] = {}
    units: dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_row, max_col=4, values_only=False):
        name = row[0].value
        value = row[1].value
        if name and value is not None:
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb == "002E75B6":
                continue
            try:
                specs[name] = float(value)
            except (ValueError, TypeError):
                specs[name] = value
            units[name] = str(row[2].value) if row[2].value else "—"
    return specs, units


sensor_specs, sensor_units = _read_sheet(wb["Sensor Configuration"])
atmo_specs, atmo_units = _read_sheet(wb["Atmosphere & Geometry"])
sweep_specs, sweep_units = _read_sheet(wb["Sweep Definition"])

print("=" * 80)
print("SCENARIO 3.2: Weather Sensitivity — How Bad Can the Weather Get?")
print("=" * 80)

print("\n=== Sensor Configuration ===")
for k, v in sensor_specs.items():
    print(f"  {k}: {v} {sensor_units.get(k, '')}")

print("\n=== Atmosphere & Geometry ===")
for k, v in atmo_specs.items():
    print(f"  {k}: {v} {atmo_units.get(k, '')}")

print("\n=== Sweep Definition ===")
for k, v in sweep_specs.items():
    print(f"  {k}: {v} {sweep_units.get(k, '')}")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

# Optics
aperture_m = float(sensor_specs["Aperture diameter"]) / 100.0       # cm → m
focal_length_m = float(sensor_specs["Focal length"]) / 100.0        # cm → m
f_number = float(sensor_specs["f-number"])
transmission = float(sensor_specs["Optical transmission"]) / 100.0  # % → frac
optics_temp_K = float(sensor_specs["Optics temperature"]) + 273.15  # °C → K

# Detector
pixel_pitch_um = float(sensor_specs["Pixel pitch"])
qe = float(sensor_specs["Quantum efficiency"]) / 100.0              # % → frac
dark_rate = float(sensor_specs["Dark current"])
operating_temp_K = float(sensor_specs["Operating temperature"])
read_noise = float(sensor_specs["Read noise (post-CDS)"])
fwc = float(sensor_specs["Full well capacity"])
adc_bits = int(sensor_specs["ADC resolution"])
gain = float(sensor_specs["System gain"])
ipc_coupling = float(sensor_specs["IPC coupling"]) / 100.0          # % → frac
t_int_s = float(sensor_specs["Integration time"]) / 1000.0          # ms → s

# Geometry
altitude_m = float(atmo_specs["Sensor altitude"]) * 1000.0          # km → m
target_alt_m = float(atmo_specs["Target altitude"])
offnadir_deg = float(atmo_specs["Off-nadir angle"])
offnadir_rad = offnadir_deg * math.pi / 180.0                       # deg → rad

# Atmosphere baseline
std_atmo = str(atmo_specs["Standard atmosphere"])
baseline_vis_km = float(atmo_specs["Baseline visibility"])
baseline_pwv_cm = float(atmo_specs["Baseline PWV"])
aerosol_type = str(atmo_specs["Aerosol type"])

# Scene
target_temp_K = float(atmo_specs["Target temperature"])
target_emiss = float(atmo_specs["Target emissivity"])
bg_temp_K = float(atmo_specs["Background temperature"])
bg_emiss = float(atmo_specs["Background emissivity"])

# Sweep parameters
niirs_req = float(sweep_specs["Minimum NIIRS"])
niirs_goal = float(sweep_specs["NIIRS goal"])
vis_min_km = float(sweep_specs["Visibility min"])
vis_max_km = float(sweep_specs["Visibility max"])
n_vis = int(sweep_specs["Visibility points"])
pwv_min_cm = float(sweep_specs["PWV min"])
pwv_max_cm = float(sweep_specs["PWV max"])
n_pwv = int(sweep_specs["PWV points"])

# Named conditions
named_conditions = [
    ("Crystal clear", 100.0, 0.8),
    ("Clear", 50.0, 1.0),
    ("Standard", 23.0, 1.4),
    ("Light haze", 10.0, 2.0),
    ("Moderate haze", 5.0, 3.0),
    ("Heavy haze", 2.0, 4.0),
    ("Tropical humid", 10.0, 5.0),
    ("Arctic dry", 50.0, 0.5),
]

# Spectral band — read from sensor or use MWIR defaults
band_min_um = 3.5
band_max_um = 5.0

print("\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 35} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'f-number':<35s} {f_number:>14.1f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<15s}  °C + 273.15")
print(f"  {'Band':<35s} {band_min_um:>6.2f}–{band_max_um:<6.2f}  {'µm':<15s}  hardcoded MWIR")
print(f"  {'Pixel pitch':<35s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<15s}  ms ÷ 1000")
print(f"  {'Sensor altitude':<35s} {altitude_m:>14.0f}  {'m':<15s}  km × 1000")
print(f"  {'Off-nadir angle':<35s} {offnadir_rad:>14.4f}  {'rad':<15s}  deg × π/180")
print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<15s}  % ÷ 100")

# ---------------------------------------------------------------------------
# Step 3: Build RADIANT config
# ---------------------------------------------------------------------------

print("\n=== Radiometric Regime ===")
gsd_m = pixel_pitch_um * 1e-6 * altitude_m / focal_length_m
print(f"  Atmosphere model: simple (Beer-Lambert parametric)")
print(f"  Extended regime: target fills pixel FOV")
print(f"  GSD = {pixel_pitch_um:.0f} µm × {altitude_m/1000:.0f} km / {focal_length_m:.2f} m"
      f" = {gsd_m:.1f} m")
print(f"  NIIRS requirement: ≥ {niirs_req:.1f} [—] (threshold)")
print(f"  NIIRS goal:        ≥ {niirs_goal:.1f} [—] (desired)")
print(f"")
print(f"  WEATHER SENSITIVITY NOTE:")
print(f"    Visibility controls aerosol scattering (Koschmieder: σ = 3.912/V).")
print(f"    In MWIR (3.5–5.0 µm), aerosol scattering is less dominant than in")
print(f"    visible bands because particles scatter less at longer wavelengths")
print(f"    (Angstrom exponent α ≈ 1.3 for rural aerosol).")
print(f"    Water vapor absorption has strong bands at 2.7 µm (edge of MWIR)")
print(f"    and a continuum contribution across 3.5–5.0 µm.")
print(f"    For a 500 km vertical path, these effects are significant.")


def make_config(vis_km: float, pwv_cm: float) -> dict:
    """Build RADIANT config for a given visibility and PWV."""
    return {
        "source": {
            "target": {"temperature": target_temp_K, "emissivity": target_emiss},
            "background": {"temperature": bg_temp_K, "emissivity": bg_emiss},
        },
        "atmosphere": {
            "model": "simple",
            "visibility_km": vis_km,
            "precipitable_water_cm": pwv_cm,
            "standard_atmosphere": std_atmo,
            "aerosol_type": aerosol_type,
        },
        "geometry": {
            "sensor_altitude_m": altitude_m,
            "target_altitude_m": target_alt_m,
            "path_zenith_rad": offnadir_rad,
        },
        "optics": {
            "aperture_diameter_m": aperture_m,
            "focal_length_m": focal_length_m,
            "transmission_scalar": transmission,
            "optics_temperature_K": optics_temp_K,
        },
        "detector": {
            "pixel_pitch_x_um": pixel_pitch_um,
            "pixel_pitch_y_um": pixel_pitch_um,
            "qe_value": qe,
            "dark_rate_e_per_s": dark_rate,
            "detector_temperature_K": operating_temp_K,
            "ipc_coupling": ipc_coupling,
        },
        "spectral_integration": {
            "filter_min_um": band_min_um,
            "filter_max_um": band_max_um,
            "integration_time_s": t_int_s,
        },
        "readout": {
            "read_noise_e_rms": read_noise,
            "gain_e_per_dn": gain,
            "adc_bits": adc_bits,
            "full_well_capacity_e": fwc,
        },
        "performance": {"niirs": {"allow_extrapolated": True}},
    }


def evaluate_weather(vis_km: float, pwv_cm: float) -> dict:
    """Run RADIANT at given visibility and PWV, return key metrics."""
    cfg = make_config(vis_km, pwv_cm)
    r = Sensor.from_dict(cfg).evaluate()

    signal_e = r.stage_outputs["readout"]["signal_e_final"]
    snr = r.metrics["snr"]
    niirs = r.metrics.get("niirs")
    gsd_x = r.metrics.get("gsd_cross_track_m")
    rer = r.metrics.get("rer")
    mtf_nyq = r.metrics.get("mtf_at_nyquist")
    tau_atm = r.stage_outputs["atmosphere"].get("tau_atm")
    tau_band = float(np.mean(tau_atm)) if tau_atm is not None else None
    contrast_snr = r.metrics.get("contrast_snr")

    noise_dict = {nt.name: nt.value_e for nt in r.noise_terms}
    total_noise = math.sqrt(sum(v ** 2 for v in noise_dict.values()))

    return {
        "vis_km": vis_km,
        "pwv_cm": pwv_cm,
        "signal_e": signal_e,
        "snr": snr,
        "niirs": niirs,
        "gsd_m": gsd_x,
        "rer": rer,
        "mtf_nyq": mtf_nyq,
        "tau_band": tau_band,
        "contrast_snr": contrast_snr,
        "total_noise": total_noise,
        "noise_dict": noise_dict,
    }


# ---------------------------------------------------------------------------
# Step 4: Visibility sweep (at baseline PWV)
# ---------------------------------------------------------------------------

vis_sweep_km = np.logspace(np.log10(vis_min_km), np.log10(vis_max_km), n_vis)

print(f"\n=== Visibility Sweep: {vis_min_km:.0f}–{vis_max_km:.0f} km"
      f" ({n_vis} points, log-spaced) at PWV = {baseline_pwv_cm:.1f} cm ===")

vis_results = []
for vis in vis_sweep_km:
    vis_results.append(evaluate_weather(vis, baseline_pwv_cm))

print(f"\n  {'Visibility [km]':>15s}  {'τ_band [—]':>10s}  {'Signal [e⁻]':>12s}"
      f"  {'SNR [—]':>8s}  {'NIIRS [—]':>9s}  {'Status'}")
print(f"  {'-' * 15}  {'-' * 10}  {'-' * 12}  {'-' * 8}  {'-' * 9}  {'-' * 12}")

for vr in vis_results:
    niirs_val = vr["niirs"]
    niirs_str = f"{niirs_val:.2f}" if niirs_val is not None else "N/A"
    tau_str = f"{vr['tau_band']:.4f}" if vr["tau_band"] else "—"

    if niirs_val is not None:
        if niirs_val >= niirs_goal:
            status = "GOAL"
        elif niirs_val >= niirs_req:
            status = "PASS"
        else:
            status = "FAIL"
    else:
        status = "N/A"

    print(f"  {vr['vis_km']:>15.1f}  {tau_str:>10s}  {vr['signal_e']:>12,.0f}"
          f"  {vr['snr']:>8.1f}  {niirs_str:>9s}  {status:>12s}")

# Find critical visibility threshold (NIIRS = requirement)
threshold_vis_km = None
for i in range(len(vis_results) - 1):
    n1 = vis_results[i]["niirs"]
    n2 = vis_results[i + 1]["niirs"]
    if n1 is not None and n2 is not None:
        if n1 < niirs_req <= n2:
            # Linear interpolation
            v1 = vis_results[i]["vis_km"]
            v2 = vis_results[i + 1]["vis_km"]
            frac = (niirs_req - n1) / (n2 - n1)
            threshold_vis_km = v1 + frac * (v2 - v1)
            break

if threshold_vis_km:
    print(f"\n  → Critical visibility threshold: NIIRS = {niirs_req:.1f}"
          f" at visibility ≈ {threshold_vis_km:.1f} km")
else:
    # Check if all pass or all fail
    all_niirs = [vr["niirs"] for vr in vis_results if vr["niirs"] is not None]
    if all_niirs and min(all_niirs) >= niirs_req:
        print(f"\n  → NIIRS ≥ {niirs_req:.1f} at ALL visibilities in sweep range"
              f" (min NIIRS = {min(all_niirs):.2f} [—])")
    elif all_niirs:
        print(f"\n  → NIIRS < {niirs_req:.1f} at ALL visibilities in sweep range"
              f" (max NIIRS = {max(all_niirs):.2f} [—])")

# Find goal threshold too
threshold_goal_km = None
for i in range(len(vis_results) - 1):
    n1 = vis_results[i]["niirs"]
    n2 = vis_results[i + 1]["niirs"]
    if n1 is not None and n2 is not None:
        if n1 < niirs_goal <= n2:
            v1 = vis_results[i]["vis_km"]
            v2 = vis_results[i + 1]["vis_km"]
            frac = (niirs_goal - n1) / (n2 - n1)
            threshold_goal_km = v1 + frac * (v2 - v1)
            break

if threshold_goal_km:
    print(f"  → NIIRS goal threshold: NIIRS = {niirs_goal:.1f}"
          f" at visibility ≈ {threshold_goal_km:.1f} km")

# ---------------------------------------------------------------------------
# Step 5: PWV sweep (at baseline visibility)
# ---------------------------------------------------------------------------

pwv_sweep_cm = np.linspace(pwv_min_cm, pwv_max_cm, n_pwv)

print(f"\n=== PWV Sweep: {pwv_min_cm:.1f}–{pwv_max_cm:.1f} cm"
      f" ({n_pwv} points, linear) at visibility = {baseline_vis_km:.0f} km ===")

pwv_results = []
for pwv in pwv_sweep_cm:
    pwv_results.append(evaluate_weather(baseline_vis_km, pwv))

print(f"\n  {'PWV [cm]':>10s}  {'τ_band [—]':>10s}  {'Signal [e⁻]':>12s}"
      f"  {'SNR [—]':>8s}  {'NIIRS [—]':>9s}  {'ΔNIIRS [—]':>10s}")
print(f"  {'-' * 10}  {'-' * 10}  {'-' * 12}  {'-' * 8}  {'-' * 9}  {'-' * 10}")

baseline_niirs = None
for pr in pwv_results:
    niirs_val = pr["niirs"]
    niirs_str = f"{niirs_val:.2f}" if niirs_val is not None else "N/A"
    tau_str = f"{pr['tau_band']:.4f}" if pr["tau_band"] else "—"

    if baseline_niirs is None and niirs_val is not None:
        baseline_niirs = niirs_val
    delta = f"{niirs_val - baseline_niirs:+.2f}" if (niirs_val and baseline_niirs) else "—"

    print(f"  {pr['pwv_cm']:>10.2f}  {tau_str:>10s}  {pr['signal_e']:>12,.0f}"
          f"  {pr['snr']:>8.1f}  {niirs_str:>9s}  {delta:>10s}")

# ---------------------------------------------------------------------------
# Step 6: Named weather conditions
# ---------------------------------------------------------------------------

print(f"\n=== Named Weather Conditions — Go/No-Go Assessment ===")
print(f"  Requirement: NIIRS ≥ {niirs_req:.1f} [—] (threshold)")
print(f"  Goal:        NIIRS ≥ {niirs_goal:.1f} [—] (desired)")
print(f"")
print(f"  {'Condition':<20s}  {'Vis [km]':>8s}  {'PWV [cm]':>8s}  {'τ_band [—]':>10s}"
      f"  {'SNR [—]':>8s}  {'NIIRS [—]':>9s}  {'Go/No-Go':>10s}")
print(f"  {'-' * 20}  {'-' * 8}  {'-' * 8}  {'-' * 10}  {'-' * 8}  {'-' * 9}  {'-' * 10}")

named_results = []
for cond_name, vis, pwv in named_conditions:
    nr = evaluate_weather(vis, pwv)
    nr["condition"] = cond_name
    named_results.append(nr)

    niirs_val = nr["niirs"]
    niirs_str = f"{niirs_val:.2f}" if niirs_val is not None else "N/A"
    tau_str = f"{nr['tau_band']:.4f}" if nr["tau_band"] else "—"

    if niirs_val is not None:
        if niirs_val >= niirs_goal:
            go = "GO (goal)"
        elif niirs_val >= niirs_req:
            go = "GO"
        else:
            go = "NO-GO"
    else:
        go = "N/A"

    print(f"  {cond_name:<20s}  {vis:>8.1f}  {pwv:>8.1f}  {tau_str:>10s}"
          f"  {nr['snr']:>8.1f}  {niirs_str:>9s}  {go:>10s}")

# ---------------------------------------------------------------------------
# Step 7: 2D grid — NIIRS as function of visibility × PWV
# ---------------------------------------------------------------------------

# Use a coarser grid for the 2D sweep
vis_2d = np.array([2, 5, 10, 23, 50, 100], dtype=float)
pwv_2d = np.array([0.5, 1.0, 1.4, 2.0, 3.0, 5.0], dtype=float)

print(f"\n=== 2D NIIRS Grid: Visibility × PWV ===")
print(f"  {len(vis_2d)} × {len(pwv_2d)} = {len(vis_2d) * len(pwv_2d)} evaluations")
print(f"  Green = NIIRS ≥ {niirs_goal:.1f} (goal), Yellow = NIIRS ≥ {niirs_req:.1f} (pass),"
      f" Red = NIIRS < {niirs_req:.1f} (fail)")

grid_results: dict[tuple[float, float], dict] = {}
for vis in vis_2d:
    for pwv in pwv_2d:
        grid_results[(vis, pwv)] = evaluate_weather(vis, pwv)

# Print as table
header = f"  {'Vis \\ PWV':>12s}"
for pwv in pwv_2d:
    header += f"  {pwv:.1f} cm".rjust(9)
print(f"\n{header}")
print(f"  {'-' * 12}" + f"  {'-' * 9}" * len(pwv_2d))

for vis in vis_2d:
    row_str = f"  {vis:>10.0f} km"
    for pwv in pwv_2d:
        gr = grid_results[(vis, pwv)]
        n = gr["niirs"]
        if n is not None:
            row_str += f"  {n:>7.2f}  "
        else:
            row_str += f"  {'N/A':>7s}  "
    print(row_str)

# Count go/no-go
n_go = sum(1 for gr in grid_results.values()
           if gr["niirs"] is not None and gr["niirs"] >= niirs_req)
n_total = len(grid_results)
print(f"\n  Go/No-Go summary: {n_go}/{n_total} conditions meet"
      f" NIIRS ≥ {niirs_req:.1f} ({n_go/n_total*100:.0f}%)")

# ---------------------------------------------------------------------------
# Step 8: Noise budget at baseline conditions
# ---------------------------------------------------------------------------

print(f"\n=== Noise Budget at Baseline Conditions ===")
print(f"  Visibility = {baseline_vis_km:.0f} km, PWV = {baseline_pwv_cm:.1f} cm")

baseline_result = evaluate_weather(baseline_vis_km, baseline_pwv_cm)
br = baseline_result

print(f"\n  Signal:       {br['signal_e']:>12,.0f} e⁻")
print(f"  Total noise:  {br['total_noise']:>12.1f} e⁻ RMS")
print(f"  SNR:          {br['snr']:>12.1f} [—]")
niirs_b = br["niirs"]
print(f"  NIIRS:        {niirs_b:>12.2f} [—]" if niirs_b else "  NIIRS:              N/A")
print(f"  GSD:          {br['gsd_m']:>12.1f} m" if br["gsd_m"] else "")
print(f"  RER:          {br['rer']:>12.4f} [—]" if br["rer"] else "")
print(f"  MTF@Nyquist:  {br['mtf_nyq']:>12.4f} [—]" if br["mtf_nyq"] else "")
tau_str = f"{br['tau_band']:.4f}" if br["tau_band"] else "N/A"
print(f"  τ_band:       {tau_str:>12s} [—]")

print(f"\n  {'Noise Term':<25s}  {'σ [e⁻ RMS]':>12s}  {'Fraction [%]':>13s}")
print(f"  {'-' * 25}  {'-' * 12}  {'-' * 13}")

total_var = br["total_noise"] ** 2
sorted_terms = sorted(br["noise_dict"].items(), key=lambda x: x[1], reverse=True)
for name, sigma in sorted_terms:
    if sigma < 0.01:
        continue
    frac = sigma ** 2 / total_var * 100.0 if total_var > 0 else 0.0
    print(f"  {name:<25s}  {sigma:>12.1f}  {frac:>13.1f}")

# ---------------------------------------------------------------------------
# Step 9: NIIRS sensitivity — what drives NIIRS most?
# ---------------------------------------------------------------------------

print(f"\n=== NIIRS Sensitivity — What Drives Performance? ===")
print(f"  GIQE-5: NIIRS = 9.57 − 3.32·log₁₀(GSD_inch) + 3.32·log₁₀(RER)")
print(f"          + 1.559·log₁₀(SNR) − 0.334·H − 0.01·G")
print(f"")
print(f"  At baseline ({baseline_vis_km:.0f} km, {baseline_pwv_cm:.1f} cm):")
if br["gsd_m"] and br["rer"]:
    gsd_inch = br["gsd_m"] * 39.3701
    print(f"    GSD = {br['gsd_m']:.1f} m = {gsd_inch:.1f} inches")
    print(f"    RER = {br['rer']:.4f} [—]")
    print(f"    SNR = {br['snr']:.1f} [—]")
    print(f"")
    print(f"  NIIRS is dominated by GSD (the −3.32·log₁₀(GSD) term).")
    print(f"  Visibility affects SNR through atmospheric transmission.")
    print(f"  Since NIIRS ∝ 1.559·log₁₀(SNR), halving SNR costs ~0.47 NIIRS.")
    print(f"  GSD is fixed by geometry — weather cannot change it.")
    print(f"  RER is fixed by optics/detector — weather cannot change it.")
    print(f"  So weather impact on NIIRS comes entirely through SNR.")

# ---------------------------------------------------------------------------
# Step 10: Summary
# ---------------------------------------------------------------------------

print(f"\n=== Summary ===")
print(f"  Sensor: MWIR, {aperture_m*100:.0f} cm, f/{f_number:.0f},"
      f" {pixel_pitch_um:.0f} µm, {altitude_m/1000:.0f} km orbit")
print(f"  GSD: {gsd_m:.1f} m (fixed by geometry)")
print(f"  NIIRS requirement: ≥ {niirs_req:.1f} [—]")
print(f"")

if threshold_vis_km:
    print(f"  Critical visibility threshold: {threshold_vis_km:.1f} km"
          f" (NIIRS = {niirs_req:.1f})")
else:
    all_niirs_vis = [vr["niirs"] for vr in vis_results if vr["niirs"] is not None]
    if all_niirs_vis and min(all_niirs_vis) >= niirs_req:
        print(f"  NIIRS ≥ {niirs_req:.1f} at all visibilities ({vis_min_km:.0f}–{vis_max_km:.0f} km)")
        print(f"  This MWIR sensor is robust to visibility degradation!")

if threshold_goal_km:
    print(f"  NIIRS goal visibility: {threshold_goal_km:.1f} km (NIIRS = {niirs_goal:.1f})")

best_named = max(named_results, key=lambda x: x["niirs"] if x["niirs"] else 0)
worst_named = min(named_results, key=lambda x: x["niirs"] if x["niirs"] else 99)
print(f"")
print(f"  Best condition:  {best_named['condition']}"
      f" (NIIRS = {best_named['niirs']:.2f} [—])")
print(f"  Worst condition: {worst_named['condition']}"
      f" (NIIRS = {worst_named['niirs']:.2f} [—])")
niirs_range = (best_named["niirs"] or 0) - (worst_named["niirs"] or 0)
print(f"  Weather-induced NIIRS variation: {niirs_range:.2f} [—]")

print(f"")
print(f"  MWIR WEATHER NOTE:")
print(f"    MWIR (3.5–5.0 µm) is less affected by aerosol scattering than")
print(f"    visible-band sensors because aerosol extinction scales as λ^(-α)")
print(f"    where α ≈ 1.3 (rural). At 4.2 µm vs. 0.55 µm, aerosol extinction")
print(f"    is reduced by a factor of (4.2/0.55)^1.3 ≈ 13×.")
print(f"    Water vapor absorption is the primary weather concern in MWIR,")
print(f"    especially near the 4.3 µm CO₂ band and the H₂O continuum.")

# ---------------------------------------------------------------------------
# Step 11: Generate plots
# ---------------------------------------------------------------------------

import matplotlib.pyplot as plt
from matplotlib.colors import BoundaryNorm

PLOT_DIR = Path(__file__).parent.parent / "outputs"
PLOT_DIR.mkdir(parents=True, exist_ok=True)

# ---- Plot 1: NIIRS & SNR vs. Visibility (dual y-axis) ----
fig1, ax1 = plt.subplots(figsize=(10, 6))

vis_arr = [vr["vis_km"] for vr in vis_results]
niirs_arr = [vr["niirs"] for vr in vis_results if vr["niirs"] is not None]
snr_arr = [vr["snr"] for vr in vis_results]
tau_arr = [vr["tau_band"] for vr in vis_results if vr["tau_band"] is not None]

ax1.semilogx(vis_arr, niirs_arr, "b-o", markersize=4, linewidth=2, label="NIIRS")
ax1.axhline(y=niirs_req, color="r", linestyle="--", linewidth=1.5,
            label=f"Threshold (NIIRS = {niirs_req:.1f})")
ax1.axhline(y=niirs_goal, color="g", linestyle="--", linewidth=1.5,
            label=f"Goal (NIIRS = {niirs_goal:.1f})")
if threshold_vis_km:
    ax1.axvline(x=threshold_vis_km, color="r", linestyle=":", alpha=0.7)
    ax1.annotate(f"Critical: {threshold_vis_km:.1f} km",
                 xy=(threshold_vis_km, niirs_req), fontsize=9,
                 xytext=(threshold_vis_km * 1.5, niirs_req - 0.05),
                 arrowprops=dict(arrowstyle="->", color="r"))

ax1.set_xlabel("Visibility [km]", fontsize=12)
ax1.set_ylabel("NIIRS [—]", fontsize=12, color="b")
ax1.tick_params(axis="y", labelcolor="b")
ax1.set_ylim(min(niirs_arr) - 0.1, max(niirs_arr) + 0.2)
ax1.grid(True, alpha=0.3)

ax1b = ax1.twinx()
ax1b.semilogx(vis_arr, snr_arr, "r-s", markersize=3, linewidth=1.5,
              alpha=0.6, label="SNR")
ax1b.set_ylabel("SNR [—]", fontsize=12, color="r")
ax1b.tick_params(axis="y", labelcolor="r")

lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax1b.get_legend_handles_labels()
ax1.legend(lines1 + lines2, labels1 + labels2, loc="lower right", fontsize=9)
ax1.set_title(f"NIIRS & SNR vs. Visibility (PWV = {baseline_pwv_cm:.1f} cm)", fontsize=14)
fig1.tight_layout()
fig1.savefig(PLOT_DIR / "niirs_vs_visibility.png", dpi=150)
print(f"\n  Plot saved: outputs/niirs_vs_visibility.png")

# ---- Plot 2: NIIRS & Transmittance vs. PWV (dual y-axis) ----
fig2, ax2 = plt.subplots(figsize=(10, 6))

pwv_arr = [pr["pwv_cm"] for pr in pwv_results]
niirs_pwv = [pr["niirs"] for pr in pwv_results if pr["niirs"] is not None]
tau_pwv = [pr["tau_band"] for pr in pwv_results if pr["tau_band"] is not None]

ax2.plot(pwv_arr, niirs_pwv, "b-o", markersize=5, linewidth=2, label="NIIRS")
ax2.axhline(y=niirs_req, color="r", linestyle="--", linewidth=1.5,
            label=f"Threshold (NIIRS = {niirs_req:.1f})")
ax2.axhline(y=niirs_goal, color="g", linestyle="--", linewidth=1.5,
            label=f"Goal (NIIRS = {niirs_goal:.1f})")

# Add named condition markers
for cond_name, vis, pwv in named_conditions:
    if abs(vis - baseline_vis_km) < 1:  # only show conditions near baseline vis
        ax2.axvline(x=pwv, color="gray", linestyle=":", alpha=0.4)
        ax2.text(pwv, max(niirs_pwv) + 0.01, cond_name, rotation=45,
                 fontsize=7, ha="left", va="bottom")

ax2.set_xlabel("Precipitable Water Vapor [cm]", fontsize=12)
ax2.set_ylabel("NIIRS [—]", fontsize=12, color="b")
ax2.tick_params(axis="y", labelcolor="b")
ax2.set_ylim(min(niirs_pwv) - 0.05, max(niirs_pwv) + 0.15)
ax2.grid(True, alpha=0.3)

ax2b = ax2.twinx()
ax2b.plot(pwv_arr, tau_pwv, "m-^", markersize=4, linewidth=1.5,
          alpha=0.7, label="τ_band")
ax2b.set_ylabel("Band-Mean Transmittance τ [—]", fontsize=12, color="m")
ax2b.tick_params(axis="y", labelcolor="m")
ax2b.set_ylim(0, 1.0)

lines1, labels1 = ax2.get_legend_handles_labels()
lines2, labels2 = ax2b.get_legend_handles_labels()
ax2.legend(lines1 + lines2, labels1 + labels2, loc="upper right", fontsize=9)
ax2.set_title(f"NIIRS & Transmittance vs. PWV (Visibility = {baseline_vis_km:.0f} km)",
              fontsize=14)
fig2.tight_layout()
fig2.savefig(PLOT_DIR / "niirs_vs_pwv.png", dpi=150)
print(f"  Plot saved: outputs/niirs_vs_pwv.png")

# ---- Plot 3: 2D NIIRS Heatmap (Visibility × PWV) ----
fig3, ax3 = plt.subplots(figsize=(9, 7))

niirs_grid = np.zeros((len(vis_2d), len(pwv_2d)))
for i, vis in enumerate(vis_2d):
    for j, pwv in enumerate(pwv_2d):
        n = grid_results[(vis, pwv)]["niirs"]
        niirs_grid[i, j] = n if n is not None else np.nan

# Custom colormap boundaries
bounds = [3.5, 4.0, 4.5, 5.0, 5.5]
cmap = plt.cm.RdYlGn
norm = BoundaryNorm(bounds, cmap.N, clip=True)

im = ax3.imshow(niirs_grid, cmap=cmap, norm=norm, aspect="auto",
                origin="lower", interpolation="nearest")

ax3.set_xticks(range(len(pwv_2d)))
ax3.set_xticklabels([f"{p:.1f}" for p in pwv_2d])
ax3.set_yticks(range(len(vis_2d)))
ax3.set_yticklabels([f"{v:.0f}" for v in vis_2d])
ax3.set_xlabel("Precipitable Water Vapor [cm]", fontsize=12)
ax3.set_ylabel("Visibility [km]", fontsize=12)

# Annotate each cell with NIIRS value
for i in range(len(vis_2d)):
    for j in range(len(pwv_2d)):
        val = niirs_grid[i, j]
        color = "black" if val >= niirs_req else "white"
        ax3.text(j, i, f"{val:.2f}", ha="center", va="center",
                 fontsize=10, fontweight="bold", color=color)

cbar = fig3.colorbar(im, ax=ax3, label="NIIRS [—]", shrink=0.8)
cbar.ax.axhline(y=niirs_req, color="r", linewidth=2)
cbar.ax.axhline(y=niirs_goal, color="darkgreen", linewidth=2)

ax3.set_title("Go/No-Go: NIIRS vs. Visibility & PWV", fontsize=14)
fig3.tight_layout()
fig3.savefig(PLOT_DIR / "niirs_2d_heatmap.png", dpi=150)
print(f"  Plot saved: outputs/niirs_2d_heatmap.png")

# ---- Plot 4: Named Conditions Bar Chart ----
fig4, ax4 = plt.subplots(figsize=(10, 6))

cond_names = [nr["condition"] for nr in named_results]
cond_niirs = [nr["niirs"] if nr["niirs"] else 0 for nr in named_results]
cond_colors = []
for n in cond_niirs:
    if n >= niirs_goal:
        cond_colors.append("#92D050")
    elif n >= niirs_req:
        cond_colors.append("#FFC000")
    else:
        cond_colors.append("#FF4444")

bars = ax4.barh(cond_names, cond_niirs, color=cond_colors, edgecolor="gray", height=0.6)
ax4.axvline(x=niirs_req, color="r", linestyle="--", linewidth=2,
            label=f"Threshold ({niirs_req:.1f})")
ax4.axvline(x=niirs_goal, color="g", linestyle="--", linewidth=2,
            label=f"Goal ({niirs_goal:.1f})")

for bar, n in zip(bars, cond_niirs):
    ax4.text(bar.get_width() + 0.01, bar.get_y() + bar.get_height() / 2,
             f"{n:.2f}", va="center", fontsize=10, fontweight="bold")

ax4.set_xlabel("NIIRS [—]", fontsize=12)
ax4.set_xlim(min(cond_niirs) - 0.1, max(cond_niirs) + 0.15)
ax4.legend(loc="lower right", fontsize=10)
ax4.grid(True, axis="x", alpha=0.3)
ax4.set_title("Go/No-Go by Weather Condition", fontsize=14)
fig4.tight_layout()
fig4.savefig(PLOT_DIR / "named_conditions_gonogo.png", dpi=150)
print(f"  Plot saved: outputs/named_conditions_gonogo.png")

# ---- Plot 5: Noise Budget Stacked Bar ----
fig5, ax5 = plt.subplots(figsize=(8, 6))

noise_names = [name for name, sigma in sorted_terms if sigma >= 0.01]
noise_vals = [sigma for _, sigma in sorted_terms if sigma >= 0.01]
noise_fracs = [sigma ** 2 / total_var * 100.0 for sigma in noise_vals]

colors5 = plt.cm.Set2(np.linspace(0, 1, len(noise_names)))
wedges, texts, autotexts = ax5.pie(
    noise_fracs, labels=noise_names, autopct="%1.1f%%",
    colors=colors5, startangle=90, pctdistance=0.75,
    textprops={"fontsize": 10},
)
for t in autotexts:
    t.set_fontsize(9)
    t.set_fontweight("bold")

ax5.set_title(f"Noise Budget at Baseline\n"
              f"(Vis = {baseline_vis_km:.0f} km, PWV = {baseline_pwv_cm:.1f} cm,"
              f" σ_total = {br['total_noise']:.1f} e⁻ RMS)",
              fontsize=13)
fig5.tight_layout()
fig5.savefig(PLOT_DIR / "noise_budget_baseline.png", dpi=150)
print(f"  Plot saved: outputs/noise_budget_baseline.png")

print(f"  All 5 plots saved to outputs/")
plt.show()

# ---------------------------------------------------------------------------
# Step 12: Write output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

owb = openpyxl.Workbook()
hfont = Font(bold=True, size=11)
sfill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
sfont = Font(bold=True, size=11, color="FFFFFF")
tb = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
green_fill = PatternFill(start_color="0092D050", end_color="0092D050", fill_type="solid")
yellow_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")

# --- Sheet 1: Visibility Sweep ---
ows1 = owb.active
ows1.title = "Visibility Sweep"
ows1["A1"] = f"NIIRS vs. Visibility (PWV = {baseline_pwv_cm:.1f} cm)"
ows1["A1"].font = Font(bold=True, size=14)
ows1.column_dimensions["A"].width = 16
ows1.column_dimensions["B"].width = 14
ows1.column_dimensions["C"].width = 14
ows1.column_dimensions["D"].width = 12
ows1.column_dimensions["E"].width = 12

headers_vis = ["Visibility [km]", "τ_band [—]", "Signal [e⁻]", "SNR [—]", "NIIRS [—]"]
for c, h in enumerate(headers_vis, 1):
    cell = ows1.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb

for i, vr in enumerate(vis_results, 4):
    ows1.cell(row=i, column=1, value=round(vr["vis_km"], 2)).border = tb
    ows1.cell(row=i, column=2,
              value=round(vr["tau_band"], 4) if vr["tau_band"] else "").border = tb
    ows1.cell(row=i, column=3, value=round(vr["signal_e"], 0)).border = tb
    ows1.cell(row=i, column=4, value=round(vr["snr"], 2)).border = tb
    niirs_cell = ows1.cell(row=i, column=5,
                           value=round(vr["niirs"], 2) if vr["niirs"] else "")
    niirs_cell.border = tb
    if vr["niirs"] is not None:
        if vr["niirs"] >= niirs_goal:
            niirs_cell.fill = green_fill
        elif vr["niirs"] >= niirs_req:
            niirs_cell.fill = yellow_fill
        else:
            niirs_cell.fill = red_fill

# --- Sheet 2: PWV Sweep ---
ows2 = owb.create_sheet("PWV Sweep")
ows2["A1"] = f"NIIRS vs. Precipitable Water Vapor (Vis = {baseline_vis_km:.0f} km)"
ows2["A1"].font = Font(bold=True, size=14)
ows2.column_dimensions["A"].width = 14
ows2.column_dimensions["B"].width = 14
ows2.column_dimensions["C"].width = 14
ows2.column_dimensions["D"].width = 12
ows2.column_dimensions["E"].width = 12

headers_pwv = ["PWV [cm]", "τ_band [—]", "Signal [e⁻]", "SNR [—]", "NIIRS [—]"]
for c, h in enumerate(headers_pwv, 1):
    cell = ows2.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb

for i, pr in enumerate(pwv_results, 4):
    ows2.cell(row=i, column=1, value=round(pr["pwv_cm"], 2)).border = tb
    ows2.cell(row=i, column=2,
              value=round(pr["tau_band"], 4) if pr["tau_band"] else "").border = tb
    ows2.cell(row=i, column=3, value=round(pr["signal_e"], 0)).border = tb
    ows2.cell(row=i, column=4, value=round(pr["snr"], 2)).border = tb
    niirs_cell = ows2.cell(row=i, column=5,
                           value=round(pr["niirs"], 2) if pr["niirs"] else "")
    niirs_cell.border = tb

# --- Sheet 3: Named Conditions ---
ows3 = owb.create_sheet("Named Conditions")
ows3["A1"] = "Go/No-Go Assessment by Weather Condition"
ows3["A1"].font = Font(bold=True, size=14)
ows3.column_dimensions["A"].width = 22
ows3.column_dimensions["B"].width = 14
ows3.column_dimensions["C"].width = 12
ows3.column_dimensions["D"].width = 12
ows3.column_dimensions["E"].width = 12
ows3.column_dimensions["F"].width = 12

headers_named = ["Condition", "Visibility [km]", "PWV [cm]", "SNR [—]",
                 "NIIRS [—]", "Go/No-Go"]
for c, h in enumerate(headers_named, 1):
    cell = ows3.cell(row=3, column=c, value=h)
    cell.font = hfont
    cell.border = tb

for i, nr in enumerate(named_results, 4):
    ows3.cell(row=i, column=1, value=nr["condition"]).border = tb
    ows3.cell(row=i, column=2, value=nr["vis_km"]).border = tb
    ows3.cell(row=i, column=3, value=nr["pwv_cm"]).border = tb
    ows3.cell(row=i, column=4, value=round(nr["snr"], 2)).border = tb

    niirs_val = nr["niirs"]
    niirs_cell = ows3.cell(row=i, column=5,
                           value=round(niirs_val, 2) if niirs_val else "")
    niirs_cell.border = tb

    if niirs_val is not None:
        if niirs_val >= niirs_goal:
            go = "GO (goal)"
            niirs_cell.fill = green_fill
        elif niirs_val >= niirs_req:
            go = "GO"
            niirs_cell.fill = yellow_fill
        else:
            go = "NO-GO"
            niirs_cell.fill = red_fill
    else:
        go = "N/A"
    ows3.cell(row=i, column=6, value=go).border = tb

# --- Sheet 4: 2D Grid ---
ows4 = owb.create_sheet("2D NIIRS Grid")
ows4["A1"] = "NIIRS vs. Visibility and PWV (2D Grid)"
ows4["A1"].font = Font(bold=True, size=14)
ows4.column_dimensions["A"].width = 14

header_row = 3
ows4.cell(row=header_row, column=1, value="Vis [km] \\ PWV [cm]").font = hfont
for c, pwv in enumerate(pwv_2d, 2):
    cell = ows4.cell(row=header_row, column=c, value=f"{pwv:.1f}")
    cell.font = hfont
    cell.border = tb
    ows4.column_dimensions[chr(64 + c)].width = 10

for r_idx, vis in enumerate(vis_2d, 4):
    ows4.cell(row=r_idx, column=1, value=vis).border = tb
    for c_idx, pwv in enumerate(pwv_2d, 2):
        gr = grid_results[(vis, pwv)]
        n = gr["niirs"]
        cell = ows4.cell(row=r_idx, column=c_idx,
                         value=round(n, 2) if n else "")
        cell.border = tb
        cell.alignment = Alignment(horizontal="center")
        if n is not None:
            if n >= niirs_goal:
                cell.fill = green_fill
            elif n >= niirs_req:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

owb.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
