#!/usr/bin/env python3
"""Create Lisa's inputs for Scenario 4.2: maritime ship classification.

Emits ``lisa_ship_classes.xlsx`` — a table of ship classes with their
length and height (the vendor/intelligence product Lisa works from), plus
a second sheet with the airborne MWIR sensor configuration. The critical
dimension for the Johnson criteria is derived in the run script as
√(length · height) (the standard 2-D-target convention).

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()

# --- Ship classes -----------------------------------------------------
ws = wb.active
ws.title = "ShipClasses"
ws["A1"] = "Scenario 4.2: maritime ship classes"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [24, 14, 14, 34]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F6F8B")
for col, text in zip("ABCD", ["Ship class", "Length (m)", "Height (m)", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

SHIPS = [
    ("Small boat", 6.0, 2.0, "RHIB / skiff"),
    ("Patrol craft", 30.0, 6.0, "Fast attack / coastal"),
    ("Corvette", 90.0, 10.0, "Light combatant"),
    ("Frigate", 130.0, 15.0, "Surface combatant"),
    ("Destroyer", 155.0, 18.0, "Guided-missile"),
    ("Container ship", 300.0, 30.0, "Large commercial"),
]
r = 4
for name, length, height, note in SHIPS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=length)
    ws.cell(row=r, column=3, value=height)
    ws.cell(row=r, column=4, value=note)
    r += 1

# --- Sensor config ----------------------------------------------------
ws2 = wb.create_sheet("SensorConfig")
for col, w in zip("ABCD", [28, 14, 12, 40]):
    ws2.column_dimensions[col].width = w
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws2[f"{col}1"] = text
    ws2[f"{col}1"].font = hdr_font
    ws2[f"{col}1"].fill = hdr_fill
CFG = [
    ("Platform", "MWIR UAV", "", "Airborne maritime ISR"),
    ("Platform altitude", 5.0, "km", "Sets the horizon range"),
    ("Aperture diameter", 20.0, "cm", ""),
    ("Focal length", 1.2, "m", ""),
    ("Pixel pitch", 15.0, "µm", "MWIR MCT"),
    ("Waveband", "3–5", "µm", "MWIR"),
    ("Detection N50", 1.0, "cycles", "Johnson: something is there"),
    ("Recognition N50", 4.0, "cycles", "Johnson: ship class"),
    ("Identification N50", 6.4, "cycles", "Johnson: specific hull"),
]
r = 2
for name, value, unit, note in CFG:
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=value)
    ws2.cell(row=r, column=3, value=unit)
    ws2.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "lisa_ship_classes.xlsx"
wb.save(out)
print(f"Wrote {out}")
