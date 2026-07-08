#!/usr/bin/env python3
"""Create Lisa's inputs for Scenario 4.3: camouflage effectiveness.

Emits:

- ``steel_oxidized_aster.txt`` — bare-vehicle surface (oxidized steel) in
  the ASTER library text format (read by radiant.io.aster_library).
- ``camo_net_a.csv``  — measured ε(λ), 8–14 µm, 100 points (broadband
  low-ε net).
- ``camo_net_b.csv``  — measured ε(λ), spectrally shaped: LOW in 8–10 µm,
  HIGH in 10–12 µm (a net tuned against 8–10 µm sensors).
- ``camo_net_c.csv``  — ε at only THREE wavelengths (8.0, 10.5, 14.0 µm)
  — sparse vendor data that needs interpolation.
- ``lisa_flir_sensor.xlsx`` — the LWIR FLIR + scene/platform parameters.

All ε files are 2-column CSVs read by radiant.io.measurement.
load_measured_curve (x_unit="um"); the ASTER file exercises the Gap-1.3
importer.

Run:  python create_spreadsheet.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# Bare vehicle: oxidized steel, ASTER text format (descending, percent)
# ---------------------------------------------------------------------------

wl_steel = np.arange(14.0, 7.4, -0.5)  # descending, 14.0 → 7.5 µm
# Oxidized steel: high-ish ε (0.78–0.82), mild slope
eps_steel = 0.82 - 0.006 * (wl_steel - 7.5)
rho_steel_pct = (1.0 - eps_steel) * 100.0

header = """\
Name: Steel, oxidized (weathered plate)
Type: manmade
Class: metal
Subclass: steel
Particle Size: n/a
Sample No.: manmade.metal.steel.oxidized.synthetic
Owner: Program office (synthetic, ASTER-format)
Wavelength Range: TIR
Origin: Synthesized for scenario 4.3 following ASTER library conventions
Description: Directional hemispherical reflectance of weathered oxidized
steel plate, 7.5-14.0 micrometers. Emissivity = 1 - reflectance (opaque).
Measurement: Directional Hemispherical Reflectance
First Column: X
Second Column: Y
X Units: Wavelength (micrometers)
Y Units: Reflectance (percent)
First X Value: 14.0
Last X Value: 7.5
Number of X Values: 14
Additional Information: none
"""

with open(HERE / "steel_oxidized_aster.txt", "w", encoding="utf-8") as fh:
    fh.write(header)
    for wl, rho in zip(wl_steel, rho_steel_pct):
        fh.write(f"{wl:6.2f}  {rho:6.2f}\n")
print("Wrote steel_oxidized_aster.txt")

# ---------------------------------------------------------------------------
# Camo net emissivity CSVs (vendor measurements)
# ---------------------------------------------------------------------------

wl_100 = np.linspace(8.0, 14.0, 100)

# Net A: broadband LOW-ε metalized weave (mean ~0.60) — the intuitive
# "reduce emission" choice; against a WARM ground background it reads cold
# and can be worse than bare (a real, instructive failure mode).
eps_a = 0.58 + 0.010 * (wl_100 - 8.0) + 0.01 * np.sin(2.0 * math.pi * (wl_100 - 8.0) / 3.0)

# Net B: spectrally shaped — low in 8–10 µm, high in 10–12 µm (tuned against
# an 8–10 µm sensor); band-mean ~0.78.
eps_b = 0.55 + 0.34 / (1.0 + np.exp(-(wl_100 - 10.0) / 0.35))

for name, eps in [("camo_net_a.csv", eps_a), ("camo_net_b.csv", eps_b)]:
    with open(HERE / name, "w", encoding="utf-8") as fh:
        fh.write("# Vendor emissivity measurement, hemispherical, 300 K sample\n")
        fh.write("wavelength_um,emissivity\n")
        for wl, e in zip(wl_100, eps):
            fh.write(f"{wl:.4f},{e:.4f}\n")
    print(f"Wrote {name}")

# Net C: sparse — three wavelengths only
with open(HERE / "camo_net_c.csv", "w", encoding="utf-8") as fh:
    fh.write("# Vendor quote sheet — spot emissivity values only\n")
    fh.write("wavelength_um,emissivity\n")
    fh.write("8.0,0.93\n10.5,0.94\n14.0,0.92\n")
print("Wrote camo_net_c.csv")

# ---------------------------------------------------------------------------
# Sensor + scene workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FLIR and Scene"
ws["A1"] = "Scenario 4.3: LWIR FLIR + camouflage scene"
ws["A1"].font = Font(bold=True, size=14)
for col, width in zip("ABC", [36, 16, 12]):
    ws.column_dimensions[col].width = width
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
for col, text in zip("ABC", ["Parameter", "Value", "Unit"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Aperture diameter", 10.0, "cm"),
    ("Focal length", 20.0, "cm"),
    ("Optical transmission", 75.0, "%"),
    ("Optics temperature", 15.0, "°C"),
    ("Pixel pitch", 15.0, "µm"),
    ("Quantum efficiency", 65.0, "%"),
    ("Dark current", 2.0e6, "e⁻/s"),
    ("Operating temperature", 60.0, "K"),
    ("Band minimum", 8.0, "µm"),
    ("Band maximum", 12.0, "µm"),
    ("Integration time", 0.05, "ms"),
    ("Read noise (CDS)", 45.0, "e⁻ RMS"),
    ("System gain", 780.0, "e⁻/DN"),
    ("ADC resolution", 14, "bits"),
    ("Full well capacity", 1.2e7, "e⁻"),
    ("Platform altitude", 3.0, "km"),
    ("Vehicle projected area", 18.0, "m²"),
    ("Bare vehicle temperature", 380.0, "K"),
    ("Camo net temperature", 310.0, "K"),
    ("Background temperature (scrub)", 305.0, "K"),
    ("Background emissivity (scrub)", 0.96, "—"),
    ("Scene clutter sigma", 0.03, "—"),
]
r = 4
for name, value, unit in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    r += 1

out = HERE / "lisa_flir_sensor.xlsx"
wb.save(out)
print(f"Wrote {out}")
