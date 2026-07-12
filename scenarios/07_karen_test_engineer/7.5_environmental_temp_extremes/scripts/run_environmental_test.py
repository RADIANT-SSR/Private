"""Scenario 7.5: Environmental Test — Performance at Temperature Extremes.

Karen runs a thermal-vacuum sweep of the FPA operating temperature
(70–95 K), imaging the 300 K chamber shroud. She measured dark current
at each temperature (it deviates super-Arrhenius above ~85 K) and QE at
three temperatures. She needs SNR / NEDT vs FPA temperature, the noise
budget with dark current growing, the measured-vs-Arrhenius dark story,
the QE(T)-vs-dark-current impact split, and an operating-temperature
recommendation with margin against the acceptance spec.

This script:
  1. Loads the measured J(T) with radiant.io.dark_current_csv... no — the
     TVAC data is in e⁻/s (not A/cm²), so it loads via
     radiant.io.measurement.load_measured_curve (the A/cm² importer from
     scenario 2.1 is the right reader for vendor J(T); here the lab
     already reduced to e⁻/s). QE(T) loads the same way.
  2. Fits an Arrhenius line to the LOW-temperature points (ln J vs 1/T)
     and extrapolates — the measured data's high-T excess is the
     deviation the test exists to catch.
  3. Sweeps detector_temperature_K, setting the MEASURED dark rate and
     the QE(T)-interpolated QE at each point (co-varying parameters).
  4. SNR / NEDT vs T; noise budget with dark_shot growing; QE(T) impact
     isolated by re-running with QE held at its 77 K value.
  5. Spec checker: SNR ≥ min, NEDT ≤ max → the warmest compliant T and
     the margin, plus the recommended set point.

Usage:
    python run_environmental_test.py
"""

import math
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from radiant.api import Sensor
from radiant.io.measurement import load_measured_curve

INPUTS = Path(__file__).parent.parent / "inputs"
OUTPUTS = Path(__file__).parent.parent / "outputs"
OUTPUT_FILE = OUTPUTS / "environmental_test_results.xlsx"

print("=" * 95)
print("  SCENARIO 7.5: Environmental Test — Performance at Temperature Extremes")
print("=" * 95)

# ---------------------------------------------------------------------------
# Step 1: Load inputs
# ---------------------------------------------------------------------------

wb_in = openpyxl.load_workbook(INPUTS / "karen_env_sensor.xlsx")
ws_in = wb_in["Sensor and Spec"]
spec: dict[str, object] = {}
for row in ws_in.iter_rows(min_row=4, max_col=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        spec[str(row[0])] = row[1]

jdark = load_measured_curve(INPUTS / "karen_dark_current_tvac.csv", x_unit="K")
qe_t = load_measured_curve(INPUTS / "karen_qe_vs_temperature.csv", x_unit="K")

temps_K = jdark.x
dark_meas = jdark.y

print("\n=== Measured TVAC data (load_measured_curve, Gap 30) ===")
print(f"  Dark current J(T): {jdark.n_points} points, "
      f"{temps_K[0]:.0f}–{temps_K[-1]:.0f} K")
print(f"  QE(T): {qe_t.n_points} points "
      f"({', '.join(f'{t:.0f}K→{q:.2f}' for t, q in zip(qe_t.x, qe_t.y))})")
print("\n  NOTE: J(T) is in e⁻/s (lab-reduced), so it reads through the")
print("  generic measured-curve loader. Vendor A/cm² datasheets use the")
print("  scenario-2.1 importer radiant.io.dark_current_csv instead.")

# ---------------------------------------------------------------------------
# Step 2: Arrhenius fit to the low-T points → predicted vs measured
# ---------------------------------------------------------------------------
# ln J = ln J0 − Ea/(k_B·T): linear in 1/T. Fit the three LOWEST
# temperatures (below the super-Arrhenius onset) and extrapolate.

N_FIT = 3
inv_T = 1.0 / temps_K
slope, intercept = np.polyfit(inv_T[:N_FIT], np.log(dark_meas[:N_FIT]), 1)
K_B_EV = 8.617333262e-5
Ea_fit_eV = -slope * K_B_EV
dark_arrhenius = np.exp(intercept + slope * inv_T)

print(f"\n=== Dark current: measured vs Arrhenius fit ===")
print(f"  Arrhenius fit to the {N_FIT} lowest points → Ea = {Ea_fit_eV:.3f} eV")
print(f"  {'T [K]':>7s}  {'Measured [e⁻/s]':>16s}  {'Arrhenius [e⁻/s]':>17s}  "
      f"{'Excess [%]':>10s}")
print(f"  {'-' * 7}  {'-' * 16}  {'-' * 17}  {'-' * 10}")
for i, T in enumerate(temps_K):
    excess = (dark_meas[i] / dark_arrhenius[i] - 1.0) * 100.0
    flag = "  ← super-Arrhenius" if excess > 15.0 else ""
    print(f"  {T:>7.0f}  {dark_meas[i]:>16,.0f}  {dark_arrhenius[i]:>17,.0f}  "
          f"{excess:>+10.1f}{flag}")
onset = next((temps_K[i] for i in range(len(temps_K))
              if dark_meas[i] / dark_arrhenius[i] - 1.0 > 0.15), None)
print(f"  Super-Arrhenius onset near {onset:.0f} K — above it, an Arrhenius")
print(f"  model UNDER-predicts dark current (defect-assisted leakage); a")
print(f"  Rule-07 Arrhenius extrapolation would be optimistic there.")

# ---------------------------------------------------------------------------
# Step 3: Build the sensor and sweep FPA temperature (co-varying J and QE)
# ---------------------------------------------------------------------------

band_parts = str(spec["Cold filter passband"]).replace("–", "-").split("-")
band_min_um = float(band_parts[0]) / 1000.0
band_max_um = float(band_parts[1]) / 1000.0
t_int_s = float(spec["Integration time"]) / 1000.0
gain = float(spec["System gain"])
fwc = float(spec["Full well capacity"])
qe_77 = float(np.interp(77.0, qe_t.x, qe_t.y))


def build_sensor(det_T: float, dark_e_per_s: float, qe: float) -> Sensor:
    s = Sensor()
    s.set("source.target.temperature", float(spec["Shroud temperature"]))
    s.set("source.target.emissivity", float(spec["Shroud emissivity"]))
    s.set("source.background.temperature", float(spec["Optics temperature"]) + 273.15)
    s.set("source.background.emissivity", 0.95)
    s.set("atmosphere.model", "exo")
    # 1.0 m ≈ bench height (limb-check floor; h_sensor alias folded — CU-090)
    s.set("geometry.sensor_altitude_m", 1.0)
    s.set("optics.aperture_diameter_m", float(spec["Aperture diameter"]), unit="cm")
    s.set("optics.focal_length_m", float(spec["Focal length"]), unit="cm")
    s.set("optics.transmission_scalar", float(spec["Optical transmission"]), unit="%")
    s.set("optics.scalar_emissivity", float(spec["Optics emissivity"]) / 100.0)
    s.set("optics.nearfield_fraction", float(spec["Nearfield fraction"]))
    s.set("optics.optics_temperature_K", float(spec["Optics temperature"]) + 273.15)
    s.set("detector.pixel_pitch_x_um", float(spec["Pixel pitch"]))
    s.set("detector.pixel_pitch_y_um", float(spec["Pixel pitch"]))
    s.set("detector.qe_value", qe)
    s.set("detector.dark_rate_e_per_s", dark_e_per_s)
    s.set("detector.detector_temperature_K", det_T)
    s.set("spectral_integration.filter_min_um", band_min_um)
    s.set("spectral_integration.filter_max_um", band_max_um)
    s.set("spectral_integration.integration_time_s", t_int_s)
    s.set("readout.read_noise_e_rms", float(spec["Read noise (CDS)"]))
    s.set("readout.gain_e_per_dn", gain)
    s.set("readout.adc_bits", int(spec["ADC resolution"]))
    s.set("readout.full_well_capacity_e", fwc)
    return s


warnings.filterwarnings("ignore")

qe_at = lambda T: float(np.interp(T, qe_t.x, qe_t.y))
rows_out = []
for i, T in enumerate(temps_K):
    r = build_sensor(T, float(dark_meas[i]), qe_at(T)).evaluate()
    r_fixed_qe = build_sensor(T, float(dark_meas[i]), qe_77).evaluate()
    nd = {nt.name: nt.value_e for nt in r.noise_terms}
    rows_out.append({
        "T": T,
        "snr": r.metrics["snr"],
        "nedt_mK": (r.metrics.get("nedt_K") or float("nan")) * 1e3,
        "nedt_fixed_qe_mK": (r_fixed_qe.metrics.get("nedt_K") or float("nan")) * 1e3,
        "signal_e": r.stage_outputs["readout"]["signal_e_final"],
        "well_pct": r.stage_outputs["readout"]["signal_e_final"] / fwc * 100.0,
        "qe": qe_at(T),
        "dark_shot": nd.get("dark_shot", 0.0),
        "signal_shot": nd.get("signal_shot", 0.0),
        "read_noise": nd.get("read_noise", 0.0),
        "nearfield_shot": nd.get("nearfield_shot", 0.0),
        "quantization": nd.get("quantization", 0.0),
        "total_noise": math.sqrt(sum(v**2 for v in nd.values())),
    })

regime = build_sensor(77.0, 5e4, 0.78).evaluate().stage_outputs["optics"]["regime"]
print(f"\n=== Radiometric regime ===")
print(f"  {regime} — the 300 K chamber shroud fills the aperture. Extended")
print(f"  regime, so the background photon term is skipped (Decision #13);")
print(f"  the noise budget below is signal shot + dark shot + read + nearfield.")

# ---------------------------------------------------------------------------
# Step 4: SNR / NEDT / noise budget vs temperature
# ---------------------------------------------------------------------------

print(f"\n{'=' * 95}")
print(f"  PERFORMANCE vs FPA TEMPERATURE (t_int = {t_int_s * 1e3:.1f} ms)")
print(f"{'=' * 95}")
print(f"  {'T [K]':>6s} {'QE':>5s} {'well%':>6s} {'signal_shot':>12s} "
      f"{'dark_shot':>10s} {'read':>6s} {'σ_tot':>8s} {'SNR':>8s} {'NEDT [mK]':>10s}")
print(f"  {'-' * 6} {'-' * 5} {'-' * 6} {'-' * 12} {'-' * 10} {'-' * 6} "
      f"{'-' * 8} {'-' * 8} {'-' * 10}")
for row in rows_out:
    print(f"  {row['T']:>6.0f} {row['qe']:>5.2f} {row['well_pct']:>6.1f} "
          f"{row['signal_shot']:>12.1f} {row['dark_shot']:>10.1f} "
          f"{row['read_noise']:>6.1f} {row['total_noise']:>8.1f} "
          f"{row['snr']:>8.1f} {row['nedt_mK']:>10.2f}")

# ---------------------------------------------------------------------------
# Step 5: QE(T) impact vs dark-current impact
# ---------------------------------------------------------------------------

print(f"\n=== QE(T) impact isolated (NEDT with QE(T) vs QE frozen at 77 K) ===")
print(f"  {'T [K]':>6s}  {'NEDT QE(T) [mK]':>16s}  {'NEDT QE@77K [mK]':>17s}  "
      f"{'ΔNEDT from QE(T) [%]':>21s}")
print(f"  {'-' * 6}  {'-' * 16}  {'-' * 17}  {'-' * 21}")
for row in rows_out:
    d = (row["nedt_mK"] / row["nedt_fixed_qe_mK"] - 1.0) * 100.0
    print(f"  {row['T']:>6.0f}  {row['nedt_mK']:>16.2f}  "
          f"{row['nedt_fixed_qe_mK']:>17.2f}  {d:>+21.1f}")
qe_swing = (qe_at(temps_K[0]) - qe_at(temps_K[-1])) / qe_at(temps_K[0]) * 100.0
dark_swing = dark_meas[-1] / dark_meas[0]
print(f"  Over {temps_K[0]:.0f}–{temps_K[-1]:.0f} K, QE falls {qe_swing:.0f}% while dark")
print(f"  current rises {dark_swing:,.0f}× — dark current is the dominant")
print(f"  temperature effect on NEDT; QE(T) is a second-order correction.")

# ---------------------------------------------------------------------------
# Step 6: Spec compliance and recommendation
# ---------------------------------------------------------------------------

min_snr = float(spec["SPEC: min SNR"])
max_nedt = float(spec["SPEC: max NEDT"])

print(f"\n{'=' * 95}")
print(f"  SPEC COMPLIANCE  (SNR ≥ {min_snr:.0f}, NEDT ≤ {max_nedt:.0f} mK)")
print(f"{'=' * 95}")
print(f"  {'T [K]':>6s}  {'SNR':>8s}  {'NEDT [mK]':>10s}  {'SNR OK':>7s}  "
      f"{'NEDT OK':>8s}  {'PASS':>6s}")
print(f"  {'-' * 6}  {'-' * 8}  {'-' * 10}  {'-' * 7}  {'-' * 8}  {'-' * 6}")
compliant = []
for row in rows_out:
    snr_ok = row["snr"] >= min_snr
    nedt_ok = row["nedt_mK"] <= max_nedt
    ok = snr_ok and nedt_ok
    if ok:
        compliant.append(row["T"])
    print(f"  {row['T']:>6.0f}  {row['snr']:>8.1f}  {row['nedt_mK']:>10.2f}  "
          f"{'YES' if snr_ok else 'no':>7s}  {'YES' if nedt_ok else 'no':>8s}  "
          f"{'PASS' if ok else 'FAIL':>6s}")

if compliant:
    warmest = max(compliant)
    # NEDT margin at the warmest compliant point
    nedt_at_warmest = next(r["nedt_mK"] for r in rows_out if r["T"] == warmest)
    margin_mK = max_nedt - nedt_at_warmest
    recommend = warmest - 3.0  # 3 K guard band below the compliance edge
    print(f"\n  Warmest compliant set point: {warmest:.0f} K "
          f"(NEDT {nedt_at_warmest:.1f} mK, {margin_mK:.1f} mK under the {max_nedt:.0f} mK spec)")
    print(f"  RECOMMENDATION: operate at {recommend:.0f} K — a 3 K guard band below")
    print(f"  the compliance edge protects against cooler drift and the")
    print(f"  super-Arrhenius knee (dark current climbs steeply past {onset:.0f} K).")
else:
    print(f"\n  No temperature meets both specs in the tested range.")
    recommend = temps_K[0]

# ---------------------------------------------------------------------------
# Step 7: Plots
# ---------------------------------------------------------------------------

OUTPUTS.mkdir(parents=True, exist_ok=True)
T_arr = [r["T"] for r in rows_out]

fig1, ax1a = plt.subplots(figsize=(9, 6))
ax1a.plot(T_arr, [r["snr"] for r in rows_out], "bo-", linewidth=2, label="SNR")
ax1a.axhline(min_snr, color="blue", linestyle=":", label=f"SNR spec ≥ {min_snr:.0f}")
ax1a.set_xlabel("FPA Temperature [K]", fontsize=12)
ax1a.set_ylabel("SNR [--]", fontsize=12, color="tab:blue")
ax1a.tick_params(axis="y", labelcolor="tab:blue")
ax1b = ax1a.twinx()
ax1b.plot(T_arr, [r["nedt_mK"] for r in rows_out], "r^-", linewidth=2, label="NEDT")
ax1b.axhline(max_nedt, color="red", linestyle=":", label=f"NEDT spec ≤ {max_nedt:.0f} mK")
ax1b.set_ylabel("NEDT [mK]", fontsize=12, color="tab:red")
ax1b.tick_params(axis="y", labelcolor="tab:red")
if compliant:
    ax1a.axvspan(min(compliant), max(compliant), alpha=0.1, color="green")
ax1a.set_title("SNR and NEDT vs FPA Temperature (green = spec-compliant)", fontsize=13)
lines = ax1a.get_legend_handles_labels()[0] + ax1b.get_legend_handles_labels()[0]
labs = ax1a.get_legend_handles_labels()[1] + ax1b.get_legend_handles_labels()[1]
ax1a.legend(lines, labs, fontsize=9, loc="upper left")
ax1a.grid(True, alpha=0.3)
fig1.tight_layout()
fig1.savefig(OUTPUTS / "fig1_snr_nedt_vs_temperature.png", dpi=150)
print(f"\n  Saved {OUTPUTS / 'fig1_snr_nedt_vs_temperature.png'}")

fig2, ax2 = plt.subplots(figsize=(9, 6))
ax2.semilogy(temps_K, dark_meas, "ko-", linewidth=2, markersize=7, label="Measured")
ax2.semilogy(temps_K, dark_arrhenius, "b--", linewidth=1.5,
             label=f"Arrhenius fit (Ea={Ea_fit_eV:.3f} eV, low-T)")
if onset is not None:
    ax2.axvline(onset, color="orange", linestyle=":",
                label=f"super-Arrhenius onset ~{onset:.0f} K")
ax2.set_xlabel("FPA Temperature [K]", fontsize=12)
ax2.set_ylabel("Dark current [e⁻/s]", fontsize=12)
ax2.set_title("Measured Dark Current vs Arrhenius Extrapolation", fontsize=13)
ax2.legend(fontsize=10)
ax2.grid(True, which="both", alpha=0.3)
fig2.tight_layout()
fig2.savefig(OUTPUTS / "fig2_dark_current_arrhenius.png", dpi=150)
print(f"  Saved {OUTPUTS / 'fig2_dark_current_arrhenius.png'}")

fig3, ax3 = plt.subplots(figsize=(10, 6))
terms = ["signal_shot", "dark_shot", "read_noise", "nearfield_shot", "quantization"]
bottom = np.zeros(len(rows_out))
colors = ["#4472C4", "#ED7D31", "#FFC000", "#A5A5A5", "#5B9BD5"]
for term, color in zip(terms, colors):
    vals = np.array([r[term]**2 for r in rows_out])  # variance stack
    ax3.bar(T_arr, vals, bottom=bottom, width=2.0, color=color,
            edgecolor="black", linewidth=0.3, label=term)
    bottom += vals
ax3.set_xlabel("FPA Temperature [K]", fontsize=12)
ax3.set_ylabel("Noise variance [e⁻²]", fontsize=12)
ax3.set_title("Noise Budget vs FPA Temperature (variance stack — dark grows)",
              fontsize=13)
ax3.legend(fontsize=9)
ax3.grid(True, axis="y", alpha=0.3)
fig3.tight_layout()
fig3.savefig(OUTPUTS / "fig3_noise_budget_vs_temperature.png", dpi=150)
print(f"  Saved {OUTPUTS / 'fig3_noise_budget_vs_temperature.png'}")

# ---------------------------------------------------------------------------
# Step 8: Output workbook
# ---------------------------------------------------------------------------

wb_out = openpyxl.Workbook()
ws1 = wb_out.active
ws1.title = "Temperature Sweep"
hdr_font = Font(bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
pass_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
ws1["A1"] = "Scenario 7.5 — Environmental temperature sweep"
ws1["A1"].font = Font(bold=True, size=14)
headers = ["T [K]", "QE", "Dark meas [e-/s]", "Dark Arrhenius [e-/s]",
           "well [%]", "SNR", "NEDT [mK]", "PASS"]
for col, htext in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=htext)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
for i, row in enumerate(rows_out, 4):
    ok = row["snr"] >= min_snr and row["nedt_mK"] <= max_nedt
    vals = [round(row["T"], 0), round(row["qe"], 3),
            round(float(dark_meas[i - 4]), 0), round(float(dark_arrhenius[i - 4]), 0),
            round(row["well_pct"], 1), round(row["snr"], 1),
            round(row["nedt_mK"], 2), "PASS" if ok else "FAIL"]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=i, column=col, value=v)
        c.border = border
    ws1.cell(row=i, column=8).fill = pass_fill if ok else fail_fill
for col_letter, width in zip("ABCDEFGH", [8, 8, 17, 20, 9, 10, 11, 8]):
    ws1.column_dimensions[col_letter].width = width
wb_out.save(OUTPUT_FILE)
print(f"  Output workbook: {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# Step 9: Summary
# ---------------------------------------------------------------------------

print(f"\n{'=' * 95}")
print(f"  SUMMARY")
print(f"{'=' * 95}")
print(f"\n  MWIR bench, 300 K shroud, {band_min_um:.2f}–{band_max_um:.2f} µm, "
      f"t_int = {t_int_s * 1e3:.1f} ms, spec SNR ≥ {min_snr:.0f} / NEDT ≤ {max_nedt:.0f} mK")
print(f"  Measured J(T): {dark_meas[0]:,.0f} → {dark_meas[-1]:,.0f} e⁻/s over "
      f"{temps_K[0]:.0f}–{temps_K[-1]:.0f} K (super-Arrhenius above ~{onset:.0f} K)")
if compliant:
    print(f"  Compliant range: {min(compliant):.0f}–{max(compliant):.0f} K; "
          f"recommended set point {recommend:.0f} K")
print(f"\n  Key findings:")
print(f"    1. Measured J(T) importer closes the catalog gap: the TVAC curve")
print(f"       drives detector.dark_rate_e_per_s directly (no Arrhenius model")
print(f"       assumed) — and reveals the super-Arrhenius knee an Arrhenius")
print(f"       Rule-07 extrapolation would miss.")
print(f"    2. QE(T) matters far less than dark current: QE falls "
      f"{qe_swing:.0f}% while dark rises {dark_swing:,.0f}× across the range;")
print(f"       NEDT is dark-driven at the warm end.")
print(f"    3. The spec checker turns the sweep into a decision: warmest")
print(f"       compliant T minus a guard band = the recommended set point,")
print(f"       trading cooler power against thermal-sensitivity margin.")
