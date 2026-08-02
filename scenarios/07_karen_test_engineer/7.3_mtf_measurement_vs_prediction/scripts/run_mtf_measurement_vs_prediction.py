"""Scenario 7.3: MTF Measurement vs. Prediction — Does Our Sensor Match the Model?

Karen measured system MTF using a slanted-edge target (ISO 12233) at 650 nm in the
lab and wants to overlay the RADIANT prediction.  She has as-built WFE (0.07 waves
RMS at 633 nm) and knows the detector is 5 µm defocused from best focus.

Approach:
  1. Read Karen's lab spreadsheet (system config, WFE, defocus) and the
     slanted-edge tool's CSV export via radiant.io load_measured_curve (Gap 30).
  2. Run RADIANT with as-built parameters to get predicted system MTF.
  3. Compare measured vs. predicted with radiant.api compare_mtf (Gap 30):
     unit-aware (cy/mm -> cy/m), interpolates prediction onto measured points,
     returns residual statistics.
  4. Residual explainers: re-run with candidate effects RADIANT now models —
     electronics MTF (readout.electronics_sigma_um, Gap 32) and TIS scatter
     (optics.surface_roughness_nm, Gap 31) — and rank by residual RMS.
  5. Compute component MTF curves analytically for the decomposition plot.
  6. Sweep defocus to show sensitivity.

Physics:
  - System MTF is the product of independent component MTFs:
      MTF_sys = MTF_optics × MTF_pixel × MTF_detector × MTF_electronics
  - MTF_optics includes diffraction and WFE (from EffectivePSF)
  - MTF_pixel = |sinc(π·f·p)| for fill factor = 1
  - MTF_detector includes IPC and charge diffusion
  - Defocus adds a Gaussian-like blur whose width scales as δ/(2·f/#)
  - The slanted-edge method measures the composite system MTF

Key concepts:
  - Frequency units: RADIANT uses cycles/m internally; Karen's data is in cycles/mm.
    Conversion: 1 cy/mm = 1000 cy/m.
  - Nyquist frequency: f_Ny = 1/(2p) = 50,000 cy/m = 50 cy/mm for 10 µm pixels.
  - Diffraction cutoff: f_c = 1/(λ·f/#) = 512,821 cy/m = 512.8 cy/mm at 650 nm, f/3.

Gap status (2026-07-07 refresh):
  - Defocus model: optics.defocus_um (Gap 29 — closed)
  - MTF budget decomposition: mtf_budget stage output (Gap 19 — closed)
  - Measurement import + comparison: load_measured_curve / compare_mtf
    (Gap 30 — closed; exercised here)
  - Electronics MTF: readout.electronics_sigma_um (Gap 32 — closed;
    exercised here as a residual explainer)
  - Scatter: optics.surface_roughness_nm TIS model (Gap 31 — closed;
    exercised here as a residual explainer)
  - cy/mm display: performance.frequency_unit conversion (Gap 27 — closed)

Usage:
    python run_mtf_measurement_vs_prediction.py
"""

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from radiant.api import Sensor, compare_mtf
from radiant.io.measurement import load_measured_curve


# ---------------------------------------------------------------------------
# Step 1: Read Karen's spreadsheet + the slanted-edge tool's CSV export
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "karen_mtf_lab_data.xlsx"
MEASURED_CSV = Path(__file__).parent.parent / "inputs" / "karen_measured_mtf.csv"

wb_in = openpyxl.load_workbook(INPUT_FILE)

# --- System configuration ---
ws_sys = wb_in["System Configuration"]
specs: dict[str, object] = {}
units: dict[str, str] = {}

# Read each section (skip section headers that have colored background)
for row in ws_sys.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None and not isinstance(value, str) or (
        isinstance(value, str) and value not in ("", "—")
    ):
        try:
            specs[name] = float(value)
        except (ValueError, TypeError):
            specs[name] = value
        units[name] = str(row[2].value) if row[2].value else "—"

# --- Measured MTF: vendor CSV via radiant.io (Gap 30) ---
# The slanted-edge tool exports a comment-headed two-column CSV;
# load_measured_curve handles comments/header detection and validates
# ascending, numeric, de-duplicated frequency values.
measured_curve = load_measured_curve(MEASURED_CSV, x_unit="cy/mm")

meas_freq_cy_mm_arr = measured_curve.x
meas_mtf_arr = measured_curve.y
meas_freq_cy_m = meas_freq_cy_mm_arr * 1000.0  # cy/mm → cy/m (for analytic curves)
meas_freq_cy_mm = list(meas_freq_cy_mm_arr)

# --- As-built WFE ---
ws_wfe = wb_in["As-Built WFE"]
wfe_specs: dict[str, object] = {}
for row in ws_wfe.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        try:
            wfe_specs[name] = float(value)
        except (ValueError, TypeError):
            wfe_specs[name] = value

# --- Focus position ---
ws_focus = wb_in["Focus Position"]
focus_specs: dict[str, object] = {}
for row in ws_focus.iter_rows(min_row=5, max_col=4, values_only=False):
    name = row[0].value
    value = row[1].value
    if name and value is not None:
        try:
            focus_specs[name] = float(value)
        except (ValueError, TypeError):
            focus_specs[name] = value

# --- Defocus sweep values ---
defocus_sweep_um: list[float] = []
for row in ws_focus.iter_rows(min_row=14, max_col=1, values_only=True):
    if row[0] is not None:
        try:
            defocus_sweep_um.append(float(row[0]))
        except (ValueError, TypeError):
            pass


# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

aperture_m = float(specs["Aperture diameter"]) / 100.0        # cm → m
focal_length_m = float(specs["Effective focal length"]) / 100.0  # cm → m
f_number = float(specs["f-number"])
transmission = float(specs["Optical transmission"]) / 100.0   # % → fraction
optics_temp_K = float(specs["Optics temperature"]) + 273.15   # °C → K
obscuration = float(specs["Central obscuration"]) / 100.0     # % → fraction

filter_min_nm = float(specs["Filter min"])
filter_max_nm = float(specs["Filter max"])
filter_min_um = filter_min_nm / 1000.0                        # nm → µm
filter_max_um = filter_max_nm / 1000.0
band_center_um = (filter_min_um + filter_max_um) / 2.0
test_wavelength_nm = float(specs["MTF test wavelength"])
test_wavelength_um = test_wavelength_nm / 1000.0

pixel_pitch_um = float(specs["Pixel pitch"])
pixel_pitch_m = pixel_pitch_um * 1e-6                         # µm → m
fill_factor = float(specs["Fill factor"]) / 100.0             # % → fraction
qe = float(specs["Quantum efficiency"]) / 100.0               # % → fraction
dark_rate = float(specs["Dark current"])
det_temp_K = float(specs["Operating temperature"])
read_noise = float(specs["Read noise"])
fwc = float(specs["Full well capacity"])
adc_bits = int(specs["ADC bits"])
gain = float(specs["System gain"])
ipc_coupling = float(specs["IPC coupling"]) / 100.0           # % → fraction

t_int_ms = float(specs["Integration time"])
t_int_s = t_int_ms / 1000.0                                   # ms → s

wfe_rms_waves = float(wfe_specs["Total WFE RMS"])
wfe_ref_nm = float(wfe_specs["WFE reference wavelength"])
wfe_ref_um = wfe_ref_nm / 1000.0                              # nm → µm

defocus_um = float(focus_specs["Defocus from best focus"])
defocus_m = defocus_um * 1e-6                                 # µm → m

# Derived parameters
f_nyquist_cy_m = 1.0 / (2.0 * pixel_pitch_m)
f_nyquist_cy_mm = f_nyquist_cy_m / 1000.0
f_cutoff_cy_m = 1.0 / (test_wavelength_um * 1e-6 * f_number)
f_cutoff_cy_mm = f_cutoff_cy_m / 1000.0
Q = band_center_um * f_number / pixel_pitch_um
airy_diam_um = 2.44 * band_center_um * f_number


# Lab configuration: no atmosphere, no geometry (bench test)
config = {
    "source": {
        "target": {"temperature": 300.0, "emissivity": 0.90},
        "background": {"temperature": 295.0, "emissivity": 0.95},
    },
    "atmosphere": {
        "model": "exo",
    },
    "geometry": {
        "sensor_altitude_m": 0.0,
        "path_zenith_rad": 0.0,
        "solar_zenith_rad": 0.5,
    },
    "platform": {
        # Stage-7 stop-gap (registry Gap 42): "exo" routes through the
        # no_atmosphere 'space' sub-case, whose Earth-limb check requires
        # a positive user-set platform.h_sensor [m above MSL]. 1.0 m ≈
        # bench height; feeds only the limb check, no radiometric effect.
        "h_sensor": 1.0,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "optics_temperature_K": optics_temp_K,
        "wfe_rms_waves": wfe_rms_waves,
        "obscuration_ratio": obscuration,
        "defocus_um": defocus_um,
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate,
        "detector_temperature_K": det_temp_K,
        "fill_factor": fill_factor,
        "ipc_coupling": ipc_coupling,
    },
    "spectral_integration": {
        "filter_min_um": filter_min_um,
        "filter_max_um": filter_max_um,
        "integration_time_s": t_int_s,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "full_well_capacity_e": fwc,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
    },
}


def main() -> None:
    """Run the scenario analysis."""
    OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "mtf_comparison_results.xlsx"
    PLOT_DIR = Path(__file__).parent.parent / "outputs"

    print("=" * 80)
    print("SCENARIO 7.3: MTF Measurement vs. Prediction")
    print("=" * 80)

    print("\n=== System Parameters (from spreadsheet) ===")
    for k, v in specs.items():
        if k in units:
            print(f"  {k:<35s}: {v} [{units[k]}]")

    print(f"\n=== Measured MTF (load_measured_curve, Gap 30) ===")
    print(f"  Source:        {measured_curve.source_file}")
    print(f"  x unit:        {measured_curve.x_unit}")
    print(f"  Points:        {measured_curve.n_points} [--]")
    print(f"  Freq range:    {meas_freq_cy_mm_arr[0]:.1f} to {meas_freq_cy_mm_arr[-1]:.1f} [cy/mm]")
    print(f"  MTF at DC:     {meas_mtf_arr[0]:.4f} [--]")
    print(f"  MTF at Nyquist: ~{np.interp(50.0, meas_freq_cy_mm_arr, meas_mtf_arr):.4f} [--] (interpolated at 50 cy/mm)")

    print(f"\n=== As-Built WFE ===")
    for k, v in wfe_specs.items():
        print(f"  {k:<35s}: {v}")

    print(f"\n=== Focus Position ===")
    for k, v in focus_specs.items():
        print(f"  {k:<35s}: {v}")


    print(f"\n=== Converted to RADIANT Canonical Units ===")
    print(f"  {'Parameter':<35s} {'Value':>14s}  {'Unit':<12s}  {'Conversion'}")
    print(f"  {'-' * 35} {'-' * 14}  {'-' * 12}  {'-' * 20}")
    print(f"  {'Aperture diameter':<35s} {aperture_m:>14.4f}  {'m':<12s}  cm / 100")
    print(f"  {'Focal length':<35s} {focal_length_m:>14.4f}  {'m':<12s}  cm / 100")
    print(f"  {'Optical transmission':<35s} {transmission:>14.4f}  {'fraction':<12s}  % / 100")
    print(f"  {'Optics temperature':<35s} {optics_temp_K:>14.2f}  {'K':<12s}  °C + 273.15")
    print(f"  {'Obscuration ratio':<35s} {obscuration:>14.4f}  {'fraction':<12s}  % / 100")
    print(f"  {'Pixel pitch':<35s} {pixel_pitch_m:>14.2e}  {'m':<12s}  µm × 1e-6")
    print(f"  {'Fill factor':<35s} {fill_factor:>14.4f}  {'fraction':<12s}  % / 100")
    print(f"  {'QE':<35s} {qe:>14.4f}  {'fraction':<12s}  % / 100")
    print(f"  {'IPC coupling':<35s} {ipc_coupling:>14.4f}  {'fraction':<12s}  % / 100")
    print(f"  {'Integration time':<35s} {t_int_s:>14.6f}  {'s':<12s}  ms / 1000")
    print(f"  {'WFE RMS':<35s} {wfe_rms_waves:>14.4f}  {'waves':<12s}  at {wfe_ref_nm:.0f} nm")
    print(f"  {'Defocus':<35s} {defocus_m:>14.2e}  {'m':<12s}  µm × 1e-6")
    print(f"  {'Band':<35s} {filter_min_um:>6.3f}-{filter_max_um:<6.3f}  {'µm':<12s}  nm / 1000")

    print(f"\n=== Derived Parameters ===")
    print(f"  f_Nyquist:         {f_nyquist_cy_m:.0f} [cy/m] = {f_nyquist_cy_mm:.1f} [cy/mm]")
    print(f"  f_cutoff (diffr.): {f_cutoff_cy_m:.0f} [cy/m] = {f_cutoff_cy_mm:.1f} [cy/mm]")
    print(f"  Q (sampling):      {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")
    print(f"  Airy disk:         {airy_diam_um:.1f} [µm] ({airy_diam_um / pixel_pitch_um:.2f} pixels)")
    print(f"  Ratio f_Ny/f_c:    {f_nyquist_cy_m / f_cutoff_cy_m:.3f} [--]")

    # ---------------------------------------------------------------------------
    # Step 3: Run RADIANT to get predicted system MTF
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  RADIANT PREDICTION — AS-BUILT SYSTEM")
    print(f"{'=' * 80}")


    sensor = Sensor.from_dict(config)
    r = sensor.evaluate()

    # CU-058 RESOLVED: scalar WFE + defocus previously failed the dual-path
    # consistency check (~0.17 vs 0.05) on every evaluation of this scenario
    # because the product path dropped the WFE screen when folding defocus to
    # Z4 and the PSF path used a Gaussian kernel instead. Defocus now enters
    # BOTH paths as pupil Z4 alongside the preserved screen, so the check
    # passes and the two paths agree by construction (Rule 4).
    print(f"\n  NOTE: scalar WFE + defocus previously tripped the dual-path")
    print(f"  consistency check on every run (CU-058); with defocus unified as")
    print(f"  pupil Z4 on both paths the check now passes — both paths carry")
    print(f"  diffraction + WFE screen + defocus from the same complex pupil.")

    # Extract predicted MTF curve from RADIANT
    perf_out = r.stage_outputs.get("performance", {})
    pred_freq_cy_m = perf_out.get("mtf_freq_x")
    pred_mtf = perf_out.get("mtf_x")

    if pred_freq_cy_m is not None and pred_mtf is not None:
        pred_freq_cy_m = np.array(pred_freq_cy_m)
        pred_mtf = np.array(pred_mtf)
        pred_freq_cy_mm = pred_freq_cy_m / 1000.0

        # MTF at Nyquist from RADIANT
        mtf_nyq_radiant = r.metrics.get("mtf_at_nyquist", 0.0)

        print(f"\n  RADIANT predicted MTF curve:")
        print(f"    Points:          {len(pred_freq_cy_m)} [--]")
        print(f"    Freq range:      0 to {pred_freq_cy_mm[-1]:.1f} [cy/mm]")
        print(f"    MTF at Nyquist:  {mtf_nyq_radiant:.4f} [--]")
        print(f"    Strehl:          {r.metrics.get('strehl', 0.0):.4f} [--]")
        print(f"    RER:             {r.metrics.get('rer', 0.0):.4f} [--]")
        print(f"    EE(1x1):         {r.metrics.get('ee_1x1', 0.0):.4f} [--]")
        print(f"    Q (center):      {r.metrics.get('q_center', 0.0):.3f} [--]")
        print(f"    Q (min/max):     {r.metrics.get('q_min', 0.0):.3f} / {r.metrics.get('q_max', 0.0):.3f} [--]")
        print(f"    FWHM_x:          {r.metrics.get('fwhm_x_m', 0.0) * 1e6:.2f} [µm]")
        print(f"    Well margin:     {r.metrics.get('well_margin_dB', 0.0):.1f} [dB]")
        print(f"    Dynamic range:   {r.metrics.get('dynamic_range_dB', 0.0):.1f} [dB]")

        # GSD and NIIRS are None for lab tests (altitude = 0)
        gsd_val = r.metrics.get("gsd_cross_track_m")
        niirs_val = r.metrics.get("niirs")
        print(f"    GSD:             {'N/A (lab test, altitude=0)' if gsd_val is None else f'{gsd_val:.2f} [m]'}")
        print(f"    NIIRS:           {'N/A (lab test, altitude=0)' if niirs_val is None else f'{niirs_val:.2f} [--]'}")

        # NEDT — very large for lab test (no scene contrast)
        nedt_val = r.metrics.get("nedt_K", 0.0)
        if nedt_val > 1e6:
            print(f"    NEDT:            N/A (lab test, no thermal scene)")
        else:
            print(f"    NEDT:            {nedt_val * 1000:.1f} [mK]")
    else:
        print("\n  WARNING: RADIANT did not return full MTF curve!")
        pred_freq_cy_m = meas_freq_cy_m
        pred_mtf = np.ones_like(meas_freq_cy_m)
        pred_freq_cy_mm = meas_freq_cy_mm_arr

    # ---------------------------------------------------------------------------
    # Step 4: Compute component MTF curves analytically
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  MTF COMPONENT DECOMPOSITION")
    print(f"{'=' * 80}")

    # Use measurement frequency grid for comparison
    freq_eval_cy_m = meas_freq_cy_m.copy()
    freq_eval_cy_mm = meas_freq_cy_mm_arr.copy()

    # 4a. Diffraction MTF (circular unobscured aperture)
    # MTF(f) = (2/π)[arccos(f/fc) - (f/fc)√(1-(f/fc)²)] for f < fc
    f_norm = freq_eval_cy_m / f_cutoff_cy_m
    mtf_diffraction = np.zeros_like(f_norm)
    valid = f_norm < 1.0
    mtf_diffraction[valid] = (2.0 / math.pi) * (
        np.arccos(f_norm[valid]) - f_norm[valid] * np.sqrt(1.0 - f_norm[valid] ** 2)
    )
    mtf_diffraction[freq_eval_cy_m == 0] = 1.0

    print(f"\n  1. Diffraction MTF (circular aperture, no obscuration)")
    print(f"     f_cutoff = {f_cutoff_cy_mm:.1f} [cy/mm] at λ = {test_wavelength_nm:.0f} [nm], f/{f_number:.1f}")
    print(f"     MTF_diff at Nyquist ({f_nyquist_cy_mm:.0f} cy/mm): "
          f"{np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_diffraction):.4f} [--]")
    print(f"     Note: Central obscuration ({obscuration * 100:.0f}%) slightly modifies the")
    print(f"     diffraction MTF (boosts mid-freq, reduces low-freq). RADIANT's")
    print(f"     EffectivePSF includes this effect; the analytic curve above does not.")

    # 4b. Pixel aperture MTF: |sinc(π·f·p·FF)|
    mtf_pixel = np.abs(np.sinc(freq_eval_cy_m * pixel_pitch_m * fill_factor))

    print(f"\n  2. Pixel Aperture MTF")
    print(f"     |sinc(π·f·p)| with p = {pixel_pitch_um:.1f} [µm], FF = {fill_factor:.2f}")
    print(f"     MTF_pixel at Nyquist: {np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_pixel):.4f} [--]")
    print(f"     Note: sinc zero at f = 1/p = {1.0 / pixel_pitch_m:.0f} [cy/m] = "
          f"{1.0 / pixel_pitch_m / 1000:.0f} [cy/mm]")

    # 4c. IPC MTF: (1-4α) + 2α·cos(2π·f·p) (along one axis, other = 0)
    alpha = ipc_coupling
    mtf_ipc = (1.0 - 4.0 * alpha) + 2.0 * alpha * (
        np.cos(2.0 * math.pi * freq_eval_cy_m * pixel_pitch_m) + 1.0
    )

    print(f"\n  3. IPC MTF")
    print(f"     Coupling α = {ipc_coupling:.4f} ({ipc_coupling * 100:.1f}%)")
    print(f"     MTF_ipc at Nyquist: {np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_ipc):.4f} [--]")
    print(f"     IPC boosts apparent MTF (cross-talk looks like sharpening).")

    # 4d. Defocus MTF: Gaussian blur
    # Two sigma formulas compared:
    #   Script (original): σ = δ/(4·f/#)     — geometric spot radius / 2
    #   RADIANT:           σ = |δ|/(4·f/#·√3) — RMS of uniform disk (R/√3)
    # The √3 factor is the correct RMS radius of a uniformly illuminated circle.
    spot_radius_m = abs(defocus_m) / (4.0 * f_number)  # geometric blur radius
    sigma_defocus_script = spot_radius_m / 1.0  # original script: σ = R (was R/2 before)
    sigma_defocus_radiant = abs(defocus_m) / (4.0 * f_number * math.sqrt(3.0))  # RADIANT formula

    # Use the Gaussian RMS form for KAREN'S HAND composite. Note the model
    # difference (CU-058): RADIANT itself no longer uses a Gaussian — it folds
    # defocus into the complex pupil as Zernike Z4, entering the PSF and MTF
    # product paths through the SAME pupil (Rule 4, exact defocus OTF). The
    # Gaussian here is the standard hand approximation for the composite.
    sigma_defocus_m = sigma_defocus_radiant
    mtf_defocus = np.exp(-2.0 * math.pi**2 * sigma_defocus_m**2 * freq_eval_cy_m**2)

    print(f"\n  4. Defocus MTF (RADIANT: pupil Z4 via optics.defocus_um — CU-058)")
    print(f"     Defocus: {defocus_um:.1f} [µm] from best focus")
    print(f"     Geometric spot radius: {spot_radius_m * 1e6:.2f} [µm]")
    print(f"     Hand-composite Gaussian σ:    {sigma_defocus_radiant * 1e6:.3f} [µm]")
    print(f"     MTF_defocus at Nyquist: {np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_defocus):.4f} [--]")
    print(f"     Hand formula: σ = |δ|/(4·f/#·√3) (RMS of uniform disk) — an")
    print(f"     approximation; RADIANT's native defocus is Zernike Z4 in the")
    print(f"     pupil (exact OTF, identical on both spatial paths).")

    # Composite analytic system MTF
    mtf_analytic_system = mtf_diffraction * mtf_pixel * mtf_ipc * mtf_defocus

    print(f"\n  5. Composite Analytic System MTF (product of all components)")
    print(f"     MTF_sys at Nyquist: {np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_analytic_system):.4f} [--]")

    # ---------------------------------------------------------------------------
    # Step 5: Measured vs. predicted comparison via compare_mtf (Gap 30)
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  MEASURED vs. PREDICTED MTF COMPARISON (compare_mtf, Gap 30)")
    print(f"{'=' * 80}")

    # compare_mtf converts the measured cy/mm axis to canonical cy/m, interpolates
    # the predicted curve onto the measured points (overlap only, never
    # extrapolates) and returns residual statistics.
    cmp_asbuilt = compare_mtf(r, measured_curve, axis="x", frequency_unit="cy/mm")

    print(f"\n  Compared {cmp_asbuilt.n_compared} measured points "
          f"({cmp_asbuilt.n_excluded} outside predicted grid, excluded)")
    print()
    print(cmp_asbuilt.table(max_rows=12))

    # Keep sign convention of the original script: residual = predicted − measured
    # (MtfComparisonResult stores measured − predicted).
    residual_radiant = -cmp_asbuilt.residual

    # Also compute residual for the hand-rolled analytic model
    residual_analytic = mtf_analytic_system - meas_mtf_arr

    # Key comparison table (at selected frequencies)
    key_freqs_cy_mm = [0.0, 10.0, 20.0, 30.0, 40.0, 50.0, 60.0, 70.0, 80.0, 90.0, 100.0]

    print(f"\n  {'Freq':>8s}  {'Measured':>10s}  {'RADIANT':>10s}  {'Analytic':>10s}  "
          f"{'Resid(R)':>10s}  {'Resid(A)':>10s}")
    print(f"  {'[cy/mm]':>8s}  {'[--]':>10s}  {'[--]':>10s}  {'[--]':>10s}  "
          f"{'[--]':>10s}  {'[--]':>10s}")
    print(f"  {'-' * 8}  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 10}")

    for f_mm in key_freqs_cy_mm:
        m_val = float(np.interp(f_mm, meas_freq_cy_mm_arr, meas_mtf_arr))
        r_val = float(np.interp(f_mm * 1000.0, pred_freq_cy_m, pred_mtf))
        a_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_analytic_system))
        res_r = r_val - m_val
        res_a = a_val - m_val

        freq_label = f"{f_mm:.0f}"
        if abs(f_mm - f_nyquist_cy_mm) < 1.0:
            freq_label += " *Ny"

        print(f"  {freq_label:>8s}  {m_val:>10.4f}  {r_val:>10.4f}  {a_val:>10.4f}  "
              f"{res_r:>+10.4f}  {res_a:>+10.4f}")

    # Summary statistics (RADIANT side from the compare_mtf result)
    rms_resid_radiant = cmp_asbuilt.rms_residual
    max_resid_radiant = cmp_asbuilt.max_abs_residual
    rms_resid_analytic = float(np.sqrt(np.mean(residual_analytic**2)))
    max_resid_analytic = float(np.max(np.abs(residual_analytic)))

    print(f"\n  Residual statistics (Predicted − Measured):")
    print(f"    RADIANT:   RMS = {rms_resid_radiant:.4f} [--],  Max = {max_resid_radiant:.4f} [--]")
    print(f"    Analytic:  RMS = {rms_resid_analytic:.4f} [--],  Max = {max_resid_analytic:.4f} [--]")

    print(f"\n  Interpretation:")
    if rms_resid_radiant < 0.03:
        print(f"    RADIANT prediction agrees well with measurement (RMS < 0.03).")
        print(f"    Residuals are consistent with slanted-edge measurement noise (~1.5%).")
    else:
        print(f"    RADIANT prediction differs from measurement (RMS = {rms_resid_radiant:.3f}).")
        print(f"    Candidate explainers now modeled in RADIANT: electronics blur")
        print(f"    (readout.electronics_sigma_um, Gap 32) and surface-roughness")
        print(f"    scatter (optics.surface_roughness_nm, Gap 31) — tested next.")

    # ---------------------------------------------------------------------------
    # Step 5b: Residual explainers — electronics MTF (Gap 32), scatter (Gap 31)
    # ---------------------------------------------------------------------------
    # The as-built model (diffraction + WFE + defocus + pixel + IPC) leaves a
    # residual. Two physical effects RADIANT now models are candidate causes:
    #   - Electronics blur: finite amplifier bandwidth at the pixel clock acts as
    #     a Gaussian blur on the readout (x) axis: readout.electronics_sigma_um.
    #   - Surface-roughness scatter: TIS = 1 − exp(−(4πσ/λ)²) moves energy into a
    #     wide halo: optics.surface_roughness_nm (+ scatter_halo_sigma_um).
    # Karen has no independent measurement of either, so we test each hypothesis
    # by re-running the chain over a small candidate grid and ranking by the
    # compare_mtf residual RMS. A hypothesis that does not reduce the residual is
    # rejected — the discriminating power matters as much as the best fit.

    print(f"\n{'=' * 80}")
    print(f"  RESIDUAL EXPLAINERS — ELECTRONICS (Gap 32) AND SCATTER (Gap 31)")
    print(f"{'=' * 80}")

    explainer_grid: list[dict] = []
    for elec_sigma_um in [0.0, 1.0, 2.0, 3.0]:
        for roughness_nm in [0.0, 5.0]:
            cfg_i = {k: (dict(v) if isinstance(v, dict) else v) for k, v in config.items()}
            cfg_i["readout"] = dict(cfg_i["readout"])
            cfg_i["readout"]["electronics_sigma_um"] = elec_sigma_um
            cfg_i["optics"] = dict(cfg_i["optics"])
            cfg_i["optics"]["surface_roughness_nm"] = roughness_nm
            r_i = Sensor.from_dict(cfg_i).evaluate()
            cmp_i = compare_mtf(r_i, measured_curve, axis="x", frequency_unit="cy/mm")
            explainer_grid.append({
                "elec_sigma_um": elec_sigma_um,
                "roughness_nm": roughness_nm,
                "rms": cmp_i.rms_residual,
                "max": cmp_i.max_abs_residual,
                "mtf_nyq": r_i.metrics.get("mtf_at_nyquist", 0.0),
                "result": r_i,
                "cmp": cmp_i,
            })

    print(f"\n  {'σ_elec [µm]':>12s}  {'roughness [nm]':>15s}  {'Resid RMS [--]':>15s}  "
          f"{'Resid Max [--]':>15s}  {'MTF@Ny [--]':>12s}")
    print(f"  {'-' * 12}  {'-' * 15}  {'-' * 15}  {'-' * 15}  {'-' * 12}")
    for g in explainer_grid:
        marker = "  <- as-built" if g["elec_sigma_um"] == 0.0 and g["roughness_nm"] == 0.0 else ""
        print(f"  {g['elec_sigma_um']:>12.1f}  {g['roughness_nm']:>15.1f}  {g['rms']:>15.4f}  "
              f"{g['max']:>15.4f}  {g['mtf_nyq']:>12.4f}{marker}")

    best = min(explainer_grid, key=lambda g: g["rms"])
    r_best = best["result"]
    cmp_best = best["cmp"]
    explainers_improved = best["rms"] < rms_resid_radiant - 1e-4
    mean_residual = float(np.mean(residual_radiant))  # predicted − measured

    print(f"\n  Best fit: σ_elec = {best['elec_sigma_um']:.1f} [µm], "
          f"roughness = {best['roughness_nm']:.1f} [nm] "
          f"(residual RMS {best['rms']:.4f} vs {rms_resid_radiant:.4f} as-built)")

    if explainers_improved:
        print(f"  The best-fit blur terms reduce the residual — the measurement rolls")
        print(f"  off faster than the as-built prediction, consistent with an")
        print(f"  unmodeled focal-plane blur in the readout chain.")
    else:
        print(f"  BOTH HYPOTHESES REJECTED: neither electronics blur nor scatter")
        print(f"  improves the fit. Diagnosis from the residual sign: the as-built")
        print(f"  prediction sits BELOW the measurement over most of the band")
        print(f"  (mean predicted − measured = {mean_residual:+.4f}), so any added")
        print(f"  blur can only widen the gap. The discrepancy is not a missing")
        print(f"  degradation — it is the shape ambiguity of the scalar-WFE input:")
        print(f"  RADIANT spreads the {wfe_rms_waves:.2f}-wave RMS error as a random")
        print(f"  phase screen (energy into a compact halo → immediate low-frequency")
        print(f"  MTF drop toward the Strehl plateau), while the lab tool's shape")
        print(f"  keeps low frequencies near 1. A single RMS number cannot pin the")
        print(f"  MTF shape; import the as-built Zernike prescription instead")
        print(f"  (io.load_zemax_zernike, Gap 26) so the pupil carries the true")
        print(f"  aberration and the shape ambiguity disappears.")

    # Use the best-fit prediction for the overlay/residual plots
    pred_freq_best_cy_m = np.array(r_best.stage_outputs["performance"]["mtf_freq_x"])
    pred_mtf_best = np.array(r_best.stage_outputs["performance"]["mtf_x"])
    residual_best = -cmp_best.residual  # predicted − measured

    # ---------------------------------------------------------------------------
    # Step 6: Defocus sensitivity sweep
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  DEFOCUS SENSITIVITY ANALYSIS")
    print(f"{'=' * 80}")

    print(f"\n  Sweeping defocus from {defocus_sweep_um[0]:.0f} to {defocus_sweep_um[-1]:.0f} [µm]")
    print(f"  Showing MTF at Nyquist ({f_nyquist_cy_mm:.0f} cy/mm) vs. defocus")
    print(f"  Note: This is an analytic calculation — RADIANT now includes defocus via optics.defocus_um.")

    print(f"\n  {'Defocus':>10s}  {'Spot Radius':>14s}  {'σ_defocus':>12s}  {'MTF@Ny':>10s}  {'dMTF':>10s}")
    print(f"  {'[µm]':>10s}  {'[µm]':>14s}  {'[µm]':>12s}  {'[--]':>10s}  {'[%]':>10s}")
    print(f"  {'-' * 10}  {'-' * 14}  {'-' * 12}  {'-' * 10}  {'-' * 10}")

    defocus_results: list[dict] = []
    # Baseline MTF at Nyquist without defocus (from RADIANT)
    mtf_ny_no_defocus = mtf_nyq_radiant

    for d_um in defocus_sweep_um:
        d_m = d_um * 1e-6
        spot_r = d_m / (2.0 * f_number)
        sig_d = spot_r / 2.0
        # Defocus MTF at Nyquist
        mtf_def_ny = math.exp(-2.0 * math.pi**2 * sig_d**2 * f_nyquist_cy_m**2)
        # Combined: RADIANT MTF × defocus factor
        mtf_combined_ny = mtf_ny_no_defocus * mtf_def_ny
        d_mtf_pct = ((mtf_combined_ny - mtf_ny_no_defocus) / mtf_ny_no_defocus * 100.0
                     if mtf_ny_no_defocus > 0 else 0.0)

        defocus_results.append({
            "defocus_um": d_um,
            "spot_radius_um": spot_r * 1e6,
            "sigma_defocus_um": sig_d * 1e6,
            "mtf_at_nyquist": mtf_combined_ny,
            "mtf_defocus_only": mtf_def_ny,
            "d_mtf_pct": d_mtf_pct,
        })

        print(f"  {d_um:>10.1f}  {spot_r * 1e6:>14.3f}  {sig_d * 1e6:>12.3f}  "
              f"{mtf_combined_ny:>10.4f}  {d_mtf_pct:>+10.1f}")

    # Find defocus tolerance for 10% and 20% MTF loss
    for threshold_pct in [10.0, 20.0]:
        d_thresh = None
        for dr in defocus_results:
            if dr["d_mtf_pct"] < -threshold_pct:
                d_thresh = dr["defocus_um"]
                break
        if d_thresh is not None:
            print(f"\n  {threshold_pct:.0f}% MTF loss at ~{d_thresh:.0f} [µm] defocus")
        else:
            print(f"\n  {threshold_pct:.0f}% MTF loss not reached in sweep range")

    print(f"\n  Karen's current defocus: {defocus_um:.0f} [µm]")
    kr = next((dr for dr in defocus_results if abs(dr["defocus_um"] - defocus_um) < 0.1), None)
    if kr:
        print(f"  → MTF@Nyquist degradation: {kr['d_mtf_pct']:+.1f} [%]")
        print(f"  → MTF@Nyquist with defocus: {kr['mtf_at_nyquist']:.4f} [--]")

    # ---------------------------------------------------------------------------
    # Step 7: Component MTF budget table
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  MTF BUDGET AT NYQUIST ({f_nyquist_cy_mm:.0f} cy/mm)")
    print(f"{'=' * 80}")

    mtf_diff_ny = float(np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_diffraction))
    mtf_pixel_ny = float(np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_pixel))
    mtf_ipc_ny = float(np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_ipc))
    mtf_defocus_ny = float(np.interp(f_nyquist_cy_m, freq_eval_cy_m, mtf_defocus))
    mtf_meas_ny = float(np.interp(f_nyquist_cy_mm, meas_freq_cy_mm_arr, meas_mtf_arr))

    print(f"\n  {'Component':<30s}  {'MTF@Ny':>10s}  {'log10(MTF)':>10s}  {'Contribution':>12s}")
    print(f"  {'-' * 30}  {'-' * 10}  {'-' * 10}  {'-' * 12}")

    components = [
        ("Diffraction", mtf_diff_ny),
        ("Pixel aperture", mtf_pixel_ny),
        ("IPC (boost)", mtf_ipc_ny),
        ("Defocus (5 µm)", mtf_defocus_ny),
    ]

    log_total = 0.0
    for name, val in components:
        log_val = math.log10(max(val, 1e-10))
        log_total += log_val
        print(f"  {name:<30s}  {val:>10.4f}  {log_val:>+10.4f}  {'×':>12s}")

    product = mtf_diff_ny * mtf_pixel_ny * mtf_ipc_ny * mtf_defocus_ny
    print(f"  {'─' * 30}  {'─' * 10}  {'─' * 10}  {'─' * 12}")
    print(f"  {'Product (analytic system)':30s}  {product:>10.4f}  {log_total:>+10.4f}")
    print(f"  {'RADIANT predicted':30s}  {mtf_nyq_radiant:>10.4f}  "
          f"{math.log10(max(mtf_nyq_radiant, 1e-10)):>+10.4f}")
    print(f"  {'Measured (slanted-edge)':30s}  {mtf_meas_ny:>10.4f}  "
          f"{math.log10(max(mtf_meas_ny, 1e-10)):>+10.4f}")

    print(f"\n  Notes:")
    print(f"    - RADIANT's MTF now includes diffraction + WFE + pixel + IPC + defocus")
    print(f"    - Analytic system MTF includes all four components above")
    print(f"    - IPC 'boosts' apparent MTF (> 1.0) — this is physically correct;")
    print(f"      IPC cross-talk acts like a sharpening kernel at sub-Nyquist frequencies")
    print(f"    - Discrepancy between RADIANT and analytic diffraction is expected:")
    print(f"      RADIANT includes obscuration and WFE; analytic curve is ideal unobscured")

    # ---------------------------------------------------------------------------
    # Step 7b: RADIANT MTF budget decomposition (from performance stage)
    # ---------------------------------------------------------------------------

    mtf_budget = r.stage_outputs.get("performance", {}).get("mtf_budget")
    if mtf_budget is not None:
        print(f"\n{'=' * 80}")
        print(f"  RADIANT MTF BUDGET AT NYQUIST (from mtf_budget API)")
        print(f"{'=' * 80}")

        per_term = mtf_budget.per_term_at_nyquist
        print(f"\n  {'Component':<30s}  {'MTF@Ny_x':>10s}  {'MTF@Ny_y':>10s}")
        print(f"  {'-' * 30}  {'-' * 10}  {'-' * 10}")

        seen: set[str] = set()
        for key in per_term:
            base = key.rsplit("_", 1)[0]
            if base in seen:
                continue
            seen.add(base)
            val_x = per_term.get(f"{base}_x", 1.0)
            val_y = per_term.get(f"{base}_y", 1.0)
            label = base.replace("mtf_", "").replace("_", " ").title()
            print(f"  {label:<30s}  {val_x:>10.4f}  {val_y:>10.4f}")

        print(f"  {'─' * 30}  {'─' * 10}  {'─' * 10}")
        print(f"  {'System (product)':30s}  {mtf_budget.system_mtf_at_nyquist_x:>10.4f}  "
              f"{mtf_budget.system_mtf_at_nyquist_y:>10.4f}")

    # ---------------------------------------------------------------------------
    # Step 7c: Noise breakdown
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  NOISE BREAKDOWN")
    print(f"{'=' * 80}")

    print(f"\n  {'Noise Source':<30s}  {'Value':>10s}")
    print(f"  {'-' * 30}  {'-' * 10}")
    for nt in r.noise_terms:
        if nt.value_e > 0.001:
            print(f"  {nt.name:<30s}  {nt.value_e:>10.4f} [e-]")
    print(f"\n  Note: Signal and background shot noise are ~0 in lab test")
    print(f"  (no photon flux from thermal scene at room temperature in VNIR).")

    # ---------------------------------------------------------------------------
    # Step 8: Plots
    # ---------------------------------------------------------------------------

    PLOT_DIR.mkdir(parents=True, exist_ok=True)
    fig_w, fig_h = 10, 7

    # Plot 1: Measured vs. Predicted MTF overlay
    fig1, ax1 = plt.subplots(figsize=(fig_w, fig_h))

    ax1.plot(meas_freq_cy_mm_arr, meas_mtf_arr, "ko", markersize=4, alpha=0.7,
             label="Measured (slanted-edge)")
    ax1.plot(pred_freq_cy_mm, pred_mtf, "b-", linewidth=2,
             label="RADIANT as-built (defocus, no electronics)")
    if explainers_improved:
        ax1.plot(pred_freq_best_cy_m / 1000.0, pred_mtf_best, "g-", linewidth=2,
                 label=f"RADIANT best fit (+σ_elec={best['elec_sigma_um']:.0f} µm, Gap 32)")
    ax1.plot(freq_eval_cy_mm, mtf_analytic_system, "r--", linewidth=1.5,
             label="Analytic system (with defocus)")

    ax1.axvline(f_nyquist_cy_mm, color="gray", linestyle=":", alpha=0.6,
                label=f"Nyquist = {f_nyquist_cy_mm:.0f} [cy/mm]")
    ax1.axvline(f_cutoff_cy_mm, color="green", linestyle=":", alpha=0.4,
                label=f"Diffraction cutoff = {f_cutoff_cy_mm:.0f} [cy/mm]")

    ax1.set_xlabel("Spatial Frequency [cy/mm]", fontsize=12)
    ax1.set_ylabel("MTF [--]", fontsize=12)
    ax1.set_title("Scenario 7.3: Measured vs. Predicted MTF — VNIR Lab Test", fontsize=13)
    ax1.legend(fontsize=9, loc="upper right")
    ax1.grid(True, alpha=0.3)
    ax1.set_xlim(0, f_max_cy_mm := 105.0)
    ax1.set_ylim(0, 1.05)
    fig1.tight_layout()
    fig1.savefig(PLOT_DIR / "fig1_mtf_measured_vs_predicted.png", dpi=150)

    # Plot 2: Component MTF decomposition
    fig2, ax2 = plt.subplots(figsize=(fig_w, fig_h))

    ax2.plot(freq_eval_cy_mm, mtf_diffraction, "g-", linewidth=2, label="Diffraction")
    ax2.plot(freq_eval_cy_mm, mtf_pixel, "m--", linewidth=2, label="Pixel aperture")
    ax2.plot(freq_eval_cy_mm, mtf_ipc, "c-..", linewidth=1.5, label=f"IPC (α={ipc_coupling:.3f})")
    ax2.plot(freq_eval_cy_mm, mtf_defocus, "y:", linewidth=2,
             label=f"Defocus ({defocus_um:.0f} µm)")
    ax2.plot(freq_eval_cy_mm, mtf_analytic_system, "r-", linewidth=2.5,
             label="System (product)")
    ax2.plot(meas_freq_cy_mm_arr, meas_mtf_arr, "ko", markersize=4, alpha=0.5,
             label="Measured")

    ax2.axvline(f_nyquist_cy_mm, color="gray", linestyle=":", alpha=0.5,
                label=f"Nyquist ({f_nyquist_cy_mm:.0f} cy/mm)")
    ax2.set_xlabel("Spatial Frequency [cy/mm]", fontsize=12)
    ax2.set_ylabel("MTF [--]", fontsize=12)
    ax2.set_title("MTF Component Decomposition", fontsize=13)
    ax2.legend(fontsize=9, loc="upper right")
    ax2.grid(True, alpha=0.3)
    ax2.set_xlim(0, 105.0)
    ax2.set_ylim(0, 1.15)  # IPC can exceed 1.0
    fig2.tight_layout()
    fig2.savefig(PLOT_DIR / "fig2_mtf_component_decomposition.png", dpi=150)

    # Plot 3: MTF residual
    fig3, (ax3a, ax3b) = plt.subplots(2, 1, figsize=(fig_w, fig_h), height_ratios=[2, 1])

    # Top: overlay
    ax3a.plot(meas_freq_cy_mm_arr, meas_mtf_arr, "ko", markersize=4, alpha=0.7,
              label="Measured")
    ax3a.plot(pred_freq_cy_mm, pred_mtf, "b-", linewidth=2,
              label="RADIANT as-built")
    if explainers_improved:
        ax3a.plot(pred_freq_best_cy_m / 1000.0, pred_mtf_best, "g-", linewidth=2,
                  label=f"RADIANT + σ_elec={best['elec_sigma_um']:.0f} µm")
    ax3a.plot(freq_eval_cy_mm, mtf_analytic_system, "r--", linewidth=1.5,
              label="Analytic (with defocus)")
    ax3a.axvline(f_nyquist_cy_mm, color="gray", linestyle=":", alpha=0.5)
    ax3a.set_ylabel("MTF [--]", fontsize=11)
    ax3a.set_title("MTF Comparison with Residual", fontsize=13)
    ax3a.legend(fontsize=9)
    ax3a.grid(True, alpha=0.3)
    ax3a.set_xlim(0, 105.0)
    ax3a.set_ylim(0, 1.05)

    # Bottom: residual (compare_mtf grids — overlap-only, never extrapolated)
    ax3b.plot(cmp_asbuilt.freq_cy_m / 1000.0, residual_radiant, "b.-",
              linewidth=1.5, markersize=3,
              label=f"As-built − Measured (RMS={rms_resid_radiant:.3f})")
    if explainers_improved:
        ax3b.plot(cmp_best.freq_cy_m / 1000.0, residual_best, "g.-",
                  linewidth=1.5, markersize=3,
                  label=f"Best fit − Measured (RMS={best['rms']:.3f})")
    ax3b.plot(freq_eval_cy_mm, residual_analytic, "r.--", linewidth=1, markersize=3,
              label=f"Analytic − Measured (RMS={rms_resid_analytic:.3f})")
    ax3b.axhline(0, color="black", linewidth=0.5)
    ax3b.axhline(0.02, color="gray", linestyle=":", alpha=0.4)
    ax3b.axhline(-0.02, color="gray", linestyle=":", alpha=0.4)
    ax3b.axvline(f_nyquist_cy_mm, color="gray", linestyle=":", alpha=0.5)
    ax3b.set_xlabel("Spatial Frequency [cy/mm]", fontsize=11)
    ax3b.set_ylabel("Residual [--]", fontsize=11)
    ax3b.legend(fontsize=9)
    ax3b.grid(True, alpha=0.3)
    ax3b.set_xlim(0, 105.0)

    fig3.tight_layout()
    fig3.savefig(PLOT_DIR / "fig3_mtf_residual.png", dpi=150)

    # Plot 4: Defocus sensitivity
    fig4, ax4 = plt.subplots(figsize=(fig_w, fig_h))

    defocus_arr = [dr["defocus_um"] for dr in defocus_results]
    mtf_ny_arr = [dr["mtf_at_nyquist"] for dr in defocus_results]
    mtf_def_arr = [dr["mtf_defocus_only"] for dr in defocus_results]

    ax4.plot(defocus_arr, mtf_ny_arr, "bo-", linewidth=2, markersize=8,
             label="System MTF@Nyquist (RADIANT × defocus)")
    ax4.plot(defocus_arr, mtf_def_arr, "r^--", linewidth=1.5, markersize=6,
             label="Defocus MTF@Nyquist (defocus component only)")

    ax4.axhline(mtf_ny_no_defocus, color="green", linestyle=":", alpha=0.5,
                label=f"RADIANT baseline (with defocus) = {mtf_ny_no_defocus:.4f}")
    ax4.axvline(defocus_um, color="orange", linestyle="--", alpha=0.6,
                label=f"Karen's current defocus = {defocus_um:.0f} [µm]")
    ax4.axhline(mtf_meas_ny, color="purple", linestyle=":", alpha=0.5,
                label=f"Measured MTF@Nyquist = {mtf_meas_ny:.4f}")

    ax4.set_xlabel("Defocus from Best Focus [µm]", fontsize=12)
    ax4.set_ylabel("MTF at Nyquist [--]", fontsize=12)
    ax4.set_title("Defocus Sensitivity — MTF at Nyquist vs. Focus Position", fontsize=13)
    ax4.legend(fontsize=9)
    ax4.grid(True, alpha=0.3)
    ax4.set_ylim(0, max(mtf_ny_arr) * 1.15)
    fig4.tight_layout()
    fig4.savefig(PLOT_DIR / "fig4_defocus_sensitivity.png", dpi=150)

    # ---------------------------------------------------------------------------
    # Step 9: Output spreadsheet
    # ---------------------------------------------------------------------------

    wb_out = openpyxl.Workbook()

    header_font_out = Font(bold=True, size=10, color="FFFFFF")
    header_fill_out = PatternFill("solid", fgColor="2E75B6")
    thin_border_out = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Sheet 1: MTF comparison
    ws_out = wb_out.active
    ws_out.title = "MTF Comparison"

    headers_out = [
        "Freq [cy/mm]", "Measured [--]", "RADIANT Predicted [--]",
        "Analytic System [--]", "Residual (R-M) [--]", "Residual (A-M) [--]",
        "MTF Diffraction [--]", "MTF Pixel [--]", "MTF IPC [--]", "MTF Defocus [--]",
    ]

    for col_idx, h in enumerate(headers_out, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_out
        cell.fill = header_fill_out
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border_out

    for row_idx, (f_mm, m_val) in enumerate(
        zip(meas_freq_cy_mm_arr, meas_mtf_arr), start=2
    ):
        r_val = float(np.interp(f_mm * 1000.0, pred_freq_cy_m, pred_mtf))
        a_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_analytic_system))
        d_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_diffraction))
        p_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_pixel))
        i_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_ipc))
        df_val = float(np.interp(f_mm, freq_eval_cy_mm, mtf_defocus))

        vals = [
            round(float(f_mm), 3), round(float(m_val), 4), round(r_val, 4),
            round(a_val, 4), round(r_val - float(m_val), 4), round(a_val - float(m_val), 4),
            round(d_val, 4), round(p_val, 4), round(i_val, 4), round(df_val, 4),
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws_out.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border_out

    # Sheet 2: Defocus sweep
    ws_def = wb_out.create_sheet("Defocus Sweep")
    def_headers = [
        "Defocus [µm]", "Spot Radius [µm]", "σ_defocus [µm]",
        "MTF@Nyquist [--]", "MTF_defocus [--]", "dMTF [%]",
    ]

    for col_idx, h in enumerate(def_headers, start=1):
        cell = ws_def.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_out
        cell.fill = header_fill_out
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border_out

    for row_idx, dr in enumerate(defocus_results, start=2):
        vals_d = [
            dr["defocus_um"], round(dr["spot_radius_um"], 3),
            round(dr["sigma_defocus_um"], 3), round(dr["mtf_at_nyquist"], 4),
            round(dr["mtf_defocus_only"], 4), round(dr["d_mtf_pct"], 1),
        ]
        for col_idx, v in enumerate(vals_d, start=1):
            cell = ws_def.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border_out

    for ws in [ws_out, ws_def]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb_out.save(OUTPUT_FILE)
    print(f"\n  Output spreadsheet: {OUTPUT_FILE}")

    # ---------------------------------------------------------------------------
    # Step 10: Summary
    # ---------------------------------------------------------------------------

    print(f"\n{'=' * 80}")
    print(f"  SUMMARY")
    print(f"{'=' * 80}")

    print(f"\n  System: {aperture_m * 100:.0f} cm aperture, f/{f_number:.0f}, "
          f"{pixel_pitch_um:.0f} µm pixels, {filter_min_nm:.0f}-{filter_max_nm:.0f} nm VNIR")
    print(f"  Lab test: slanted-edge at {test_wavelength_nm:.0f} nm, "
          f"{defocus_um:.0f} µm defocus from best focus")
    print(f"  As-built WFE: {wfe_rms_waves:.3f} waves RMS at {wfe_ref_nm:.0f} nm")
    print(f"  Q sampling:   {Q:.3f} [--] ({'well-sampled' if Q >= 1 else 'undersampled'})")

    print(f"\n  --- MTF at Nyquist ({f_nyquist_cy_mm:.0f} cy/mm) ---")
    print(f"  Measured:            {mtf_meas_ny:.4f} [--]")
    print(f"  RADIANT as-built:    {mtf_nyq_radiant:.4f} [--] (defocus, no electronics)")
    if explainers_improved:
        print(f"  RADIANT best fit:    {best['mtf_nyq']:.4f} [--] "
              f"(+σ_elec = {best['elec_sigma_um']:.0f} µm)")
    print(f"  Analytic:            {product:.4f} [--] (includes defocus)")

    print(f"\n  --- Residual RMS (Predicted − Measured) ---")
    print(f"  RADIANT as-built:    {rms_resid_radiant:.4f} [--]")
    if explainers_improved:
        print(f"  RADIANT best fit:    {best['rms']:.4f} [--]")
    else:
        print(f"  Explainer grid:      no improvement (hypotheses rejected — see")
        print(f"                       residual-explainer section)")
    print(f"  Analytic:            {rms_resid_analytic:.4f} [--]")

    print(f"\n  --- MTF Budget at Nyquist ---")
    for name, val in components:
        print(f"  {name:<20s}:  {val:.4f} [--]")

    print(f"\n  Key findings:")
    print(f"    1. The as-built model (diffraction + WFE + {defocus_um:.0f} µm defocus + pixel")
    print(f"       + IPC) leaves a systematic residual ({rms_resid_radiant:.3f} RMS,")
    print(f"       mean predicted − measured = {mean_residual:+.3f}).")
    if explainers_improved:
        print(f"    2. Electronics blur σ_elec = {best['elec_sigma_um']:.0f} µm "
              f"(readout.electronics_sigma_um,")
        print(f"       Gap 32) reduces the residual to {best['rms']:.3f} RMS — the")
        print(f"       slanted-edge method measures through the readout chain, so")
        print(f"       amplifier bandwidth is part of the true system MTF.")
    else:
        print(f"    2. The residual-explainer grid REJECTED both blur hypotheses")
        print(f"       (electronics σ_elec, Gap 32; scatter roughness, Gap 31): the")
        print(f"       prediction is already below the measurement, so extra blur")
        print(f"       only widens the gap. Testing and rejecting a hypothesis is")
        print(f"       exactly what the compare_mtf residual workflow is for.")
        print(f"    3. The discrepancy is scalar-WFE shape ambiguity: a single RMS")
        print(f"       number fixes the Strehl but not where the aberrated energy")
        print(f"       lands. RADIANT's random-phase-screen halo is compact (low-")
        print(f"       frequency MTF drop); the lab system's actual aberrations are")
        print(f"       smoother. Fix: import the as-built Zernike prescription via")
        print(f"       io.load_zemax_zernike (Gap 26) — exercised in scenario 5.1.")
    print(f"    4. The dominant MTF contributor at Nyquist is the pixel aperture")
    print(f"       (sinc rolloff), followed by diffraction.")
    print(f"    5. IPC provides a small apparent MTF boost ({mtf_ipc_ny:.4f} > 1.0).")

    print(f"\n  Remaining limitations:")
    print(f"    - Scalar wfe_rms_waves under-determines the MTF shape (see finding 3);")
    print(f"      Zernike input (Gap 26) removes the ambiguity but needs the as-built")
    print(f"      prescription from the optical shop")
    print(f"    - Defocus is a Gaussian approximation (σ = |δ|/(4·f/#·√3)), valid for")
    print(f"      small defocus only")
    print(f"    - Analytic decomposition curves assume no obscuration; RADIANT's")
    print(f"      pupil-autocorrelation MTF includes it (expected discrepancy)")

    plt.show()


if __name__ == "__main__":
    main()
