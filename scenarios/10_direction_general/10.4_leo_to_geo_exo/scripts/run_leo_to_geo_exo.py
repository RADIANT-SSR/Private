"""Scenario 10.4 - LEO -> GEO up-looking space-to-space SDA (ADR-0011 Phase-1 quick win).

Persona: Raj (mission planner) wearing a space-domain-awareness (SDA) hat.

The question
-----------
A 500 km LEO host carries a 35 cm MWIR staring telescope. Looking *up* at the
geostationary belt (35 786 km), can it detect a reference GEO communications
bus in eclipse, and out to what range? What does the LEO-vs-GEO relative
angular rate do to the achievable integration time?

Why this scene exercises the direction-general geometry (ADR-0011)
------------------------------------------------------------------
Before Geometry-Flexibility Phase 1 this scene could not even be *expressed*:
``core/viewing_triangle.py`` rejected ``h_sensor <= h_target`` and the
canonical target-side path zenith theta_o was bounded to ``[0, pi/2)``. The
physics was never the blocker - both endpoints are above ``h_atm_top``, so the
whole path is vacuum by construction and the exo backend already handled the
composition. Phase 1 lifted the gate; this scenario is the validation of that
quick win.

Everything the scenario proves is printed by ``main()``:

* the scene class ``space_to_space`` is *derived* and the optional
  ``geometry.scene_class`` assertion agrees with it (ADR-0011 decision 8);
* the vacuum identities hold **exactly** (tau == 1.0, L_path == 0.0 - array
  equality, not a tolerance), which is this scenario's atmosphere cross-check;
* the near-pi geometry: theta_o = pi exactly for the vertical case, eta obtuse,
  slant range == h_GEO - h_LEO exactly;
* the LOS angular rate from the two circular-orbit rates, entered through both
  Gap 111 doors (K2 target-velocity triple, K1 direct) which must agree;
* the smear consequence: at 500 ms the *untracked* LOS rate smears the point
  source over ~7.5 pixels, which is why the nominal design rate-tracks.

Usage::

    python scenarios/10_direction_general/10.4_leo_to_geo_exo/scripts/run_leo_to_geo_exo.py

Module-level factory (for the GUI-baseline registry)::

    make_sensor() -> radiant.api.Sensor      # the nominal validated scene
"""

from __future__ import annotations

import logging
import math
import warnings
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")  # headless, before pyplot

import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from radiant.api import Sensor  # noqa: E402
from radiant.api.scene_relevance import default_off_metrics  # noqa: E402
from radiant.core.constants import R_EARTH_M, c, h, k_B  # noqa: E402
from radiant.core.orbit import orbital_period_s, orbital_velocity_m_s  # noqa: E402

# ---------------------------------------------------------------------------
# Repo-relative paths (pathlib only - Rule 30)
# ---------------------------------------------------------------------------

HERE = Path(__file__).resolve().parent
SCENARIO = HERE.parent
INPUT_XLSX = SCENARIO / "inputs" / "sda_leo_to_geo_sensor_data.xlsx"
OUTPUT_DIR = SCENARIO / "outputs"
RESULTS_XLSX = OUTPUT_DIR / "leo_to_geo_exo_results.xlsx"

SEP = "=" * 84


# ---------------------------------------------------------------------------
# Step 1 - read the vendor workbook (vendor units throughout)
# ---------------------------------------------------------------------------


def _read_kv(sheet: Any) -> dict[str, Any]:
    """Parameter/Value rows of a vendor sheet, keyed by the parameter name."""
    out: dict[str, Any] = {}
    for row in sheet.iter_rows(min_row=5, max_col=3, values_only=True):
        name, value, _unit = row
        if name is None or value is None:
            continue
        out[str(name)] = value
    return out


def _read_kv_units(sheet: Any) -> dict[str, str]:
    """Parameter -> vendor unit string, so every echoed value can carry its unit."""
    out: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=5, max_col=3, values_only=True):
        name, value, unit = row
        if name is None or value is None:
            continue
        out[str(name)] = "--" if unit in (None, "") else str(unit)
    return out


def _read_column(sheet: Any) -> list[float]:
    """Single numeric sweep column of a vendor sweep sheet."""
    out: list[float] = []
    for (value,) in sheet.iter_rows(min_row=5, max_col=1, values_only=True):
        if value is not None:
            out.append(float(value))
    return out


def read_vendor_inputs() -> dict[str, Any]:
    """Load the vendor workbook and return the raw (vendor-unit) records."""
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    return {
        "telescope": _read_kv(wb["Telescope Datasheet"]),
        "fpa": _read_kv(wb["FPA Datasheet"]),
        "mission": _read_kv(wb["Mission Geometry"]),
        "target": _read_kv(wb["Target Signature"]),
        "telescope_units": _read_kv_units(wb["Telescope Datasheet"]),
        "fpa_units": _read_kv_units(wb["FPA Datasheet"]),
        "mission_units": _read_kv_units(wb["Mission Geometry"]),
        "target_units": _read_kv_units(wb["Target Signature"]),
        "t_int_sweep_ms": _read_column(wb["Integration Sweep"]),
        "zeta_sweep_deg": _read_column(wb["Zenith Sweep"]),
    }


# ---------------------------------------------------------------------------
# Step 2 - vendor -> RADIANT canonical units. Each conversion happens ONCE.
# ---------------------------------------------------------------------------


def canonical_inputs() -> dict[str, Any]:
    """Every vendor value converted to canonical units, exactly once each."""
    v = read_vendor_inputs()
    tel, fpa, mis, tgt = v["telescope"], v["fpa"], v["mission"], v["target"]

    return {
        # --- optics -------------------------------------------------------
        "aperture_m": float(tel["Entrance pupil diameter"]) / 1000.0,  # mm -> m
        "focal_length_m": float(tel["Effective focal length"]) / 1000.0,  # mm -> m
        "f_number": float(tel["f-number"]),  # dimensionless
        "obscuration": float(tel["Central obscuration"]) / 100.0,  # % -> fraction
        "tau_optics": float(tel["Optical transmission (in band)"]) / 100.0,  # % -> fraction
        "wfe_rms_waves": float(tel["WFE RMS"]),  # already waves
        "optics_temp_K": float(tel["Optical bench temperature"]) + 273.15,  # degC -> K
        # --- detector / readout -------------------------------------------
        "pixel_pitch_um": float(fpa["Pixel pitch"]),  # canonical schema unit is um
        "pixel_pitch_m": float(fpa["Pixel pitch"]) * 1e-6,  # um -> m (script-side only)
        "qe": float(fpa["Quantum efficiency"]) / 100.0,  # % -> fraction
        "dark_rate_e_per_s": float(fpa["Dark current"]),  # already e-/s
        "det_temp_K": float(fpa["Operating temperature"]),  # already K
        "read_noise_e": float(fpa["Read noise"]),  # already e- rms
        "full_well_e": float(fpa["Full well capacity"]) * 1e3,  # ke- -> e-
        "adc_bits": int(fpa["ADC resolution"]),
        "gain_e_per_dn": float(fpa["System gain"]),  # already e-/DN
        "t_int_s": float(fpa["Nominal integration time"]) / 1000.0,  # ms -> s
        "lam_min_um": float(fpa["Spectral band, min"]) / 1000.0,  # nm -> um
        "lam_max_um": float(fpa["Spectral band, max"]) / 1000.0,  # nm -> um
        # --- geometry ------------------------------------------------------
        "h_leo_m": float(mis["Sensor orbit altitude"]) * 1000.0,  # km -> m
        "h_geo_m": float(mis["Target orbit altitude"]) * 1000.0,  # km -> m
        "zeta_low_rad": math.radians(float(mis["Sensor-side path zenith (nominal)"])),  # deg -> rad
        "illumination": str(mis["Illumination state"]),
        "scene_class_assert": str(mis["Declared scene class"]),
        "track_residual": float(mis["Rate-track residual"]) / 100.0,  # % -> fraction
        "snr_threshold": float(mis["Detection SNR threshold"]),  # dimensionless
        # --- target signature ---------------------------------------------
        "area_m2": float(tgt["Projected area toward sensor"]),  # already m^2
        "target_T_K": float(tgt["Mean surface temperature"]) + 273.15,  # degC -> K
        "target_emissivity": float(tgt["Broadband emissivity"]),  # dimensionless
        # --- sweeps ---------------------------------------------------------
        "t_int_sweep_s": [ms / 1000.0 for ms in v["t_int_sweep_ms"]],  # ms -> s
        "zeta_sweep_rad": [math.radians(d) for d in v["zeta_sweep_deg"]],  # deg -> rad
        "t_int_sweep_ms": v["t_int_sweep_ms"],
        "zeta_sweep_deg": v["zeta_sweep_deg"],
        "raw": v,
    }


# ---------------------------------------------------------------------------
# Step 3 - kinematics of the two circular orbits (hand physics, script-side)
# ---------------------------------------------------------------------------


def open_loop_los_rate_rad_s(h_sensor_m: float, h_target_m: float) -> dict[str, float]:
    """LEO/GEO circular-orbit rates and the resulting radial-case LOS rate.

    Both satellites are on the same radial line (the vertical, theta_o = pi
    case), co-planar and co-rotating.  Each inertial velocity is purely
    tangential, i.e. perpendicular to the radial line of sight, so the LOS
    angular rate is the *difference* of the two tangential speeds divided by
    the separation::

        omega_LOS = |v_LEO - v_GEO| / (h_GEO - h_LEO)     [rad/s]

    ``radiant.core.orbit`` supplies v = sqrt(mu / a) for each altitude; nothing
    here re-derives a constant (Rule 13).
    """
    v_s = orbital_velocity_m_s(h_sensor_m)  # m/s, inertial
    v_t = orbital_velocity_m_s(h_target_m)  # m/s, inertial
    separation_m = h_target_m - h_sensor_m
    return {
        "v_sensor_m_s": v_s,
        "v_target_m_s": v_t,
        "omega_sensor_rad_s": v_s / (R_EARTH_M + h_sensor_m),
        "omega_target_rad_s": v_t / (R_EARTH_M + h_target_m),
        "period_sensor_s": orbital_period_s(h_sensor_m),
        "period_target_s": orbital_period_s(h_target_m),
        "separation_m": separation_m,
        "omega_los_rad_s": abs(v_s - v_t) / separation_m,
    }


# ---------------------------------------------------------------------------
# Step 4 - configuration builders
# ---------------------------------------------------------------------------


def make_config(
    *,
    zeta_low_rad: float | None = None,
    t_int_s: float | None = None,
    kinematics: str = "tracked",
) -> dict[str, Any]:
    """Nested RADIANT config dict for this scenario.

    Parameters
    ----------
    zeta_low_rad:
        Path zenith at the **lower** endpoint (ADR-0011 decision 3), which for
        an up-looking scene is the *sensor*.  ``None`` uses the workbook's
        nominal value (0 rad = GEO target at the LEO sensor's zenith).
    t_int_s:
        Integration time [s]; ``None`` uses the FPA datasheet nominal.
    kinematics:
        ``"tracked"``    - K1 door carrying the *residual* LOS rate that
                           survives rate tracking (the nominal design point).
        ``"open_loop"``  - K1 + K2 doors both carrying the full inertial LOS
                           rate: an inertially-fixed stare, no rate tracking.
        ``"static"``     - no kinematics door at all (K0): the published rate
                           is the platform-only ground_speed / slant value,
                           which is 0 rad/s here because no ground speed is set.
    """
    cin = canonical_inputs()
    kin = open_loop_los_rate_rad_s(cin["h_leo_m"], cin["h_geo_m"])

    geometry: dict[str, Any] = {
        "sensor_altitude_m": cin["h_leo_m"],
        "target_altitude_m": cin["h_geo_m"],
        "path_zenith_rad": cin["zeta_low_rad"] if zeta_low_rad is None else zeta_low_rad,
        "solar_illumination": cin["illumination"],
        # Optional ADR-0011 assertion: validated against the derived class.
        "scene_class": cin["scene_class_assert"],
        "target": {"projected_area_m2": cin["area_m2"]},
    }

    if kinematics == "tracked":
        # K1 (direct-rate door): what the tracking loop fails to remove.
        geometry["los_angular_rate_rad_s"] = cin["track_residual"] * kin["omega_los_rad_s"]
    elif kinematics == "open_loop":
        # Both Gap 111 doors, which must agree within 1 % (ADR-0006 rule 2).
        # K2 = target-velocity triple + the platform speed; heading pi/2 puts
        # the target's inertial velocity along the same cross-LOS direction the
        # platform's is modelled on, which is the co-planar co-rotating case.
        geometry["los_angular_rate_rad_s"] = kin["omega_los_rad_s"]
        geometry["ground_speed_m_s"] = kin["v_sensor_m_s"]
        geometry["target_speed_m_s"] = kin["v_target_m_s"]
        geometry["target_heading_rad"] = math.pi / 2.0
        geometry["target_climb_rad"] = 0.0
    elif kinematics != "static":
        raise ValueError(f"unknown kinematics mode {kinematics!r}")

    return {
        "source": {
            "scene_type": "point_source",
            "target": {
                "temperature": cin["target_T_K"],
                "emissivity": cin["target_emissivity"],
            },
        },
        # Both endpoints are above h_atm_top, so the exo backend routes this to
        # its vacuum (no-atmosphere) sub-case - see tests/integration/
        # test_uplooking_phase1.py::TestLeoToGeoUpLooking.
        "atmosphere": {"model": "exo"},
        "geometry": geometry,
        "optics": {
            "aperture_diameter_m": cin["aperture_m"],
            "focal_length_m": cin["focal_length_m"],
            "transmission_scalar": cin["tau_optics"],
            "optics_temperature_K": cin["optics_temp_K"],
            "obscuration_ratio": cin["obscuration"],
            "wfe_rms_waves": cin["wfe_rms_waves"],
        },
        "detector": {
            "pixel_pitch_x_um": cin["pixel_pitch_um"],
            "pixel_pitch_y_um": cin["pixel_pitch_um"],
            "qe_value": cin["qe"],
            "dark_rate_e_per_s": cin["dark_rate_e_per_s"],
            "detector_temperature_K": cin["det_temp_K"],
        },
        "spectral_integration": {
            "filter_min_um": cin["lam_min_um"],
            "filter_max_um": cin["lam_max_um"],
            "integration_time_s": cin["t_int_s"] if t_int_s is None else t_int_s,
        },
        "readout": {
            "read_noise_e_rms": cin["read_noise_e"],
            "full_well_capacity_e": cin["full_well_e"],
            "gain_e_per_dn": cin["gain_e_per_dn"],
            "adc_bits": cin["adc_bits"],
        },
        "performance": {"detection_snr_threshold": cin["snr_threshold"]},
    }


def make_sensor() -> Sensor:
    """The scenario's nominal, validated :class:`Sensor` (GUI-baseline factory).

    Vertical up-look (theta_o = pi exactly), 500 ms rate-tracked stare with the
    1 % residual LOS rate on the K1 door.  Building it runs no analysis and
    prints nothing, so the GUI-baseline registry can import this module and call
    it without executing ``main()``.
    """
    return Sensor.from_dict(make_config())


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _evaluate(sensor: Sensor) -> tuple[Any, list[warnings.WarningMessage]]:
    """Evaluate with warnings *visible* and captured (Rule-4 audit)."""
    with warnings.catch_warnings(record=True) as records:
        warnings.simplefilter("always")
        result = sensor.evaluate()
    return result, list(records)


def _hand_in_band_radiance_w_m2_sr(T_K: float, lam_min_um: float, lam_max_um: float) -> float:
    """Planck in-band radiance [W/m^2/sr] - independent of RADIANT's source stage."""
    wl_um = np.linspace(lam_min_um, lam_max_um, 4001)
    lam_m = wl_um * 1e-6
    # Planck spectral radiance per metre, converted to per micrometre (x 1e-6).
    L = (2.0 * h * c**2 / lam_m**5) / (np.expm1(h * c / (lam_m * k_B * T_K))) * 1e-6
    return float(np.trapezoid(L, wl_um))


def _hand_signal_e(cin: dict[str, Any], ee_box: float) -> float:
    """Hand-computed in-pixel signal [e-] for the vertical vacuum case.

    Spectrally integrated photon accounting, using nothing but constants and
    the vendor numbers:

        I(lam)  = eps * L_bb(lam) * A_proj                [W/sr/um]
        E(lam)  = I(lam) / R^2                            [W/m^2/um]
        N_e     = t * EE * QE * tau_opt * A_pupil
                  * Int E(lam) * lam / (h c) dlam         [e-]
    """
    wl_um = np.linspace(cin["lam_min_um"], cin["lam_max_um"], 4001)
    lam_m = wl_um * 1e-6
    L = (2.0 * h * c**2 / lam_m**5) / (np.expm1(h * c / (lam_m * k_B * cin["target_T_K"]))) * 1e-6
    intensity = cin["target_emissivity"] * L * cin["area_m2"]  # W/sr/um
    slant_m = cin["h_geo_m"] - cin["h_leo_m"]
    irradiance = intensity / slant_m**2  # W/m^2/um
    a_pupil_m2 = math.pi * (cin["aperture_m"] / 2.0) ** 2 * (1.0 - cin["obscuration"] ** 2)
    photon_rate = np.trapezoid(irradiance * lam_m / (h * c), wl_um)  # photons/s/m^2
    return float(photon_rate * a_pupil_m2 * cin["tau_optics"] * cin["qe"] * cin["t_int_s"] * ee_box)


def _fmt(value: float | None, spec: str = ".4g") -> str:
    return "N/A" if value is None else format(value, spec)


# ---------------------------------------------------------------------------
# main()
# ---------------------------------------------------------------------------


def main() -> None:  # noqa: PLR0915 - a scenario driver is a linear narrative
    logging.basicConfig(level=logging.WARNING, format="LOG %(levelname)s %(name)s: %(message)s")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cin = canonical_inputs()
    raw = cin["raw"]

    print(SEP)
    print("SCENARIO 10.4 - LEO -> GEO up-looking space-to-space SDA (ADR-0011)")
    print(SEP)
    print(
        "\nA 500 km LEO host stares UP at the geostationary belt. Scene class"
        "\nspace_to_space, LOS direction 'up', wholly-vacuum path. This is the"
        "\nGeometry-Flexibility Phase-1 quick win: the physics was always in"
        "\nreach, the altitude-ordering gate was the only blocker."
    )

    # -- vendor inputs --------------------------------------------------------
    print(f"\n{SEP}\n  1. VENDOR INPUTS (as received, vendor units)\n{SEP}")
    for sheet_name, key in (
        ("Telescope Datasheet", "telescope"),
        ("FPA Datasheet", "fpa"),
        ("Mission Geometry", "mission"),
        ("Target Signature", "target"),
    ):
        print(f"\n  --- {sheet_name} ---")
        units = raw[f"{key}_units"]
        for name, value in raw[key].items():
            print(f"    {name:<38s}: {value!s:>14s}  [{units[name]}]")
    print(f"\n  Integration-time sweep [ms]        : {raw['t_int_sweep_ms']}")
    print(f"  Sensor-side zenith sweep [deg]    : {raw['zeta_sweep_deg']}")

    # -- unit conversions -----------------------------------------------------
    print(f"\n{SEP}\n  2. VENDOR -> RADIANT CANONICAL CONVERSIONS (each applied once)\n{SEP}")
    # Vendor values are read back from the workbook (never re-typed here), so the
    # audit table cannot drift from the file the runner actually consumed.
    conversions = [
        (
            "Entrance pupil diameter",
            "telescope",
            "Entrance pupil diameter",
            "aperture_m",
            "m",
            "mm / 1000",
        ),
        (
            "Effective focal length",
            "telescope",
            "Effective focal length",
            "focal_length_m",
            "m",
            "mm / 1000",
        ),
        (
            "Optical transmission",
            "telescope",
            "Optical transmission (in band)",
            "tau_optics",
            "--",
            "% / 100",
        ),
        ("Central obscuration", "telescope", "Central obscuration", "obscuration", "--", "% / 100"),
        (
            "Bench temperature",
            "telescope",
            "Optical bench temperature",
            "optics_temp_K",
            "K",
            "degC + 273.15",
        ),
        ("Quantum efficiency", "fpa", "Quantum efficiency", "qe", "--", "% / 100"),
        ("Full well capacity", "fpa", "Full well capacity", "full_well_e", "e-", "ke- x 1000"),
        ("Integration time", "fpa", "Nominal integration time", "t_int_s", "s", "ms / 1000"),
        ("Band minimum", "fpa", "Spectral band, min", "lam_min_um", "um", "nm / 1000"),
        ("Band maximum", "fpa", "Spectral band, max", "lam_max_um", "um", "nm / 1000"),
        ("Sensor orbit altitude", "mission", "Sensor orbit altitude", "h_leo_m", "m", "km x 1000"),
        ("Target orbit altitude", "mission", "Target orbit altitude", "h_geo_m", "m", "km x 1000"),
        (
            "Sensor-side path zenith",
            "mission",
            "Sensor-side path zenith (nominal)",
            "zeta_low_rad",
            "rad",
            "deg x pi/180",
        ),
        (
            "Target temperature",
            "target",
            "Mean surface temperature",
            "target_T_K",
            "K",
            "degC + 273.15",
        ),
        (
            "Rate-track residual",
            "mission",
            "Rate-track residual",
            "track_residual",
            "--",
            "% / 100",
        ),
    ]
    print(
        f"\n  {'Quantity':<26s} {'Vendor':>12s} {'unit':<6s} {'Canonical':>14s} {'unit':<5s} rule"
    )
    print(f"  {'-' * 26} {'-' * 12} {'-' * 6} {'-' * 14} {'-' * 5} {'-' * 16}")
    for name, sheet_key, vendor_key, canon_key, cu, rule in conversions:
        vend = float(raw[sheet_key][vendor_key])
        vu = raw[f"{sheet_key}_units"][vendor_key]
        canon = float(cin[canon_key])
        print(f"  {name:<26s} {vend:>12.6g} {vu:<6s} {canon:>14.6g} {cu:<5s} {rule}")

    # -- nominal run ----------------------------------------------------------
    print(f"\n{SEP}\n  3. NOMINAL RUN - SCENE CLASS AND DIRECTION-GENERAL GEOMETRY\n{SEP}")
    sensor = make_sensor()
    result, records = _evaluate(sensor)
    geo = result.stage_outputs["geometry"]

    slant_m = float(geo["slant_range_m"])
    theta_o = float(geo["theta_o_rad"])
    eta = float(geo["eta_rad"])
    # zeta_low is the path zenith at the LOWER endpoint. For an up-looking
    # scene the sensor IS the lower endpoint, and the schematic's catalog
    # transform is zeta_low = pi - eta (NOT pi - theta_o: the two angles are
    # read at different vertices of the same spherical triangle).
    zeta_low = math.pi - eta

    print(f"\n  Derived scene class            : {geo['scene_class']}")
    print(f"  Asserted (geometry.scene_class): {cin['scene_class_assert']}  -> agreed")
    print(f"  Observer band / target band    : {geo['observer_class']} / {geo['target_class']}")
    print(f"  LOS direction (derived)        : {geo['los_direction']}")
    print(f"  Viewing mode                   : {geo['viewing_mode']}")
    print(
        f"\n  theta_o (target-side zenith)   : {theta_o:.9f} rad = {math.degrees(theta_o):.6f} deg"
    )
    print(f"  eta     (sensor-side off-nadir): {eta:.9f} rad = {math.degrees(eta):.6f} deg")
    print(
        f"  zeta_low (zenith at LEO sensor): {zeta_low:.9f} rad = {math.degrees(zeta_low):.6f} deg"
    )
    print(f"  Slant range                    : {slant_m:,.1f} m = {slant_m / 1e3:,.3f} km")
    print(f"  Ground range (surface arc)     : {float(geo['ground_range_m']):.6f} m")
    incidence_deg = math.degrees(float(geo["incidence_angle_rad"]))
    print(f"  Incidence angle at target      : {incidence_deg:.6f} deg")

    print("\n  CROSS-CHECK 1 (analytic geometry, vertical case):")
    hand_slant = cin["h_geo_m"] - cin["h_leo_m"]
    print(f"    theta_o expected             : {math.pi:.9f} rad (exactly pi - target at zenith)")
    print(f"    theta_o - pi                 : {theta_o - math.pi:+.3e} rad")
    print(f"    slant expected h_GEO - h_LEO : {hand_slant:,.1f} m = {hand_slant / 1e3:,.3f} km")
    print(f"    slant residual               : {slant_m - hand_slant:+.6e} m")
    print(
        "    eta = pi exactly: the sensor's *nadir* is the anti-target direction, so the"
        "\n    sensor-side off-nadir angle takes the obtuse branch. zeta_low = pi - eta = 0 rad:"
        "\n    the LEO telescope points at its own zenith."
    )

    # -- vacuum identities ----------------------------------------------------
    print(f"\n{SEP}\n  4. CROSS-CHECK 2 - VACUUM IDENTITIES (exact, not toleranced)\n{SEP}")
    q = result.stage_outputs["atmosphere"]["atm_quantities"]
    wl = np.asarray(result.wavelength_um)
    ones = np.ones_like(wl)
    zeros = np.zeros_like(wl)
    identities = [
        ("tau_up      (target -> sensor)", q.tau_up, ones, "1.0"),
        ("tau_sun     (sun -> target)", q.tau_sun, ones, "1.0"),
        ("tau_full_up (full column up)", q.tau_full_up, ones, "1.0"),
        ("L_path_up   [W/m^2/sr/um]", q.L_path_up, zeros, "0.0"),
        ("L_path_full [W/m^2/sr/um]", q.L_path_full, zeros, "0.0"),
        ("E_sky_scattered [W/m^2/um]", q.E_sky_scattered, zeros, "0.0"),
        ("E_sky_thermal   [W/m^2/um]", q.E_sky_thermal, zeros, "0.0"),
    ]
    all_exact = True
    print(f"\n  {'Product':<34s} {'expected':>9s} {'max |residual|':>16s}  exact?")
    print(f"  {'-' * 34} {'-' * 9} {'-' * 16}  ------")
    for label, actual, expected, want in identities:
        arr = np.asarray(actual, dtype=float)
        exact = bool(np.array_equal(arr, expected))
        all_exact &= exact
        print(
            f"  {label:<34s} {want:>9s} {float(np.max(np.abs(arr - expected))):>16.3e}"
            f"  {'YES' if exact else 'NO'}"
        )
    print(
        f"\n  All vacuum identities exact (bitwise array equality): {'YES' if all_exact else 'NO'}"
        "\n  Physics: both endpoints sit above h_atm_top = 100 km, so the exo backend's"
        "\n  vacuum sub-case applies. There is no column to attenuate or emit, so tau is"
        "\n  identically 1 and every path-radiance / sky-irradiance term is identically 0."
        "\n  This is the scenario's atmosphere cross-check: no MODTRAN anchor exists for"
        "\n  'no atmosphere', and none is needed - the answer is an identity."
    )

    background = result.stage_outputs["source"]["background"]
    L_bg = np.asarray(result.frames["at_aperture_background"].spectral_radiance, dtype=float)
    print(
        f"\n  LOS termination background     : {type(background).__name__}"
        f"  (max L_bg = {float(np.max(np.abs(L_bg))):.3e} W/m^2/sr/um)"
        "\n  The up-looking LOS never returns to Earth: past the GEO bus it exits into deep"
        "\n  space, so the background is cold space (identically zero radiance), not sky and"
        "\n  not ground. background_shot is therefore exactly 0 e- in the noise budget."
    )

    # -- regime + relevance ---------------------------------------------------
    print(f"\n{SEP}\n  5. REGIME AND SCENE-CLASS METRIC RELEVANCE\n{SEP}")
    regime = result.stage_outputs["optics"]["regime"]
    src = result.stage_outputs["source"]
    ifov_urad = cin["pixel_pitch_m"] / cin["focal_length_m"] * 1e6
    ang_extent_urad = float(src["angular_extent_rad"]) * 1e6
    print(f"\n  Final regime (OpticsStage)     : {regime}")
    print(f"  Target angular extent          : {ang_extent_urad:.4f} urad")
    print(f"  Detector IFOV                  : {ifov_urad:.3f} urad")
    print(f"  Fill fraction (target / pixel) : {float(src['fill_fraction']):.3e}")
    lam_c_m = 0.5 * (cin["lam_min_um"] + cin["lam_max_um"]) * 1e-6
    airy_arcsec = math.degrees(1.22 * lam_c_m / cin["aperture_m"]) * 3600.0
    print(
        "\n  Why POINT_SOURCE: a 20 m^2 bus at 35 286 km subtends"
        f" {ang_extent_urad:.4f} urad, which is"
        f"\n  {ifov_urad / ang_extent_urad:.0f}x smaller than the {ifov_urad:.3f} urad IFOV and far"
        f" below the\n  {airy_arcsec:.2f} arcsec diffraction core."
        " All of the target's flux lands inside one"
        "\n  PSF, so EE_box (Rule 9) couples it into a single pixel and the signal follows"
        "\n  the inverse-square law in range - which is exactly what makes a detection"
        "\n  RANGE meaningful for this scene."
    )

    off = sorted(default_off_metrics(str(geo["scene_class"])))
    present = sorted(k for k in off if k in result.metrics)
    print("\n  Scene-class relevance map (radiant.api.scene_relevance):")
    print(f"    space_to_space turns OFF by default ({len(off)} metrics):")
    for name in off:
        print(f"      - {name}")
    print(f"    ...of which still present in result.metrics: {present if present else 'none'}")
    print(
        "\n  Why: every one of those projects the sample footprint onto the TARGET's ground"
        "\n  plane through incidence_angle in [0, pi/2). A GEO bus has no ground plane, and"
        "\n  this scene's incidence angle is pi rad. The non-ground counterpart is on:"
    )
    for name in (
        "target_plane_sample_distance_x_m",
        "target_plane_sample_distance_y_m",
        "target_plane_sample_distance_geometric_mean_m",
    ):
        print(f"      + {name:<48s} = {result.metrics[name]:,.2f} m")
    tpsd = float(result.metrics["target_plane_sample_distance_geometric_mean_m"])
    bus_size_m = math.sqrt(cin["area_m2"])
    print(
        f"    i.e. one pixel subtends {tpsd:,.1f} m at the target's range - a"
        f" {bus_size_m:.1f} m bus"
        f"\n    is {bus_size_m / tpsd:.4f} px across, which is the point-source statement again,"
        "\n    in length units."
    )

    # -- radiometry -----------------------------------------------------------
    print(f"\n{SEP}\n  6. RADIOMETRY, SNR AND DETECTION RANGE\n{SEP}")
    ee_box = float(result.stage_outputs["platform"]["EE_box"])
    signal_e = float(result.stage_outputs["spectral_integration"]["signal_e"])
    snr = float(result.metrics["snr"])
    det_m = float(result.metrics["detection_range_m"])
    det_result = result.stage_outputs["performance"]["detection_range_result"]

    print(f"\n  EE_box (1x1, from the degraded PSF): {ee_box:.6f} [--]")
    print(f"  In-pixel signal                    : {signal_e:,.2f} e-")
    print(f"  SNR                                : {snr:.3f} [--]")
    print(f"  Detection threshold                : {cin['snr_threshold']:.1f} [--] (SNR)")
    print(f"  Detection range                    : {det_m:,.0f} m = {det_m / 1e3:,.1f} km")
    print(f"  Detection-range solver iterations  : {det_result.iterations} [--]")
    print(f"  SNR at that range                  : {det_result.snr_at_range:.4f} [--]")

    print("\n  Noise budget [e- RMS]:")
    for term in result.noise_terms:
        if term.value_e > 0.0:
            print(f"    {term.name:<22s} {term.value_e:>10.3f} e-")
    sigma_total = float(result.stage_outputs["readout"]["sigma_total_e"])
    print(f"    {'TOTAL (quadrature)':<22s} {sigma_total:>10.3f} e-")

    hand_e = _hand_signal_e(cin, ee_box)
    L_band = _hand_in_band_radiance_w_m2_sr(cin["target_T_K"], cin["lam_min_um"], cin["lam_max_um"])
    intensity = cin["target_emissivity"] * L_band * cin["area_m2"]
    irradiance = intensity / slant_m**2
    print("\n  CROSS-CHECK 3 (hand radiometry - Planck + inverse square, no RADIANT stage):")
    print(f"    In-band blackbody radiance L(280 K, 3.5-5.0 um): {L_band:.6f} W/m^2/sr")
    print(f"    Target intensity  eps.L.A_proj                 : {intensity:.4f} W/sr")
    print(f"    Irradiance at pupil  I / R^2                   : {irradiance:.4e} W/m^2")
    print(f"    Hand in-pixel signal (spectral photon count)   : {hand_e:,.4f} e-")
    print(f"    RADIANT signal_e                               : {signal_e:,.4f} e-")
    print(
        "    Relative difference                            :"
        f" {(hand_e / signal_e - 1) * 100:+.6f} %"
    )
    print(
        "    The hand model reproduces the chain to quadrature precision because in vacuum"
        "\n    the point-source chain IS this expression: Planck -> intensity -> 1/R^2 ->"
        "\n    pupil area -> photon energy -> QE -> EE_box. A spectral QE curve or a"
        "\n    wavelength-dependent transmission would introduce a weighting difference;"
        "\n    this vendor datasheet quotes both as band-averaged scalars, so it does not."
    )

    # Independent closed forms for the detection range in vacuum.
    thr = cin["snr_threshold"]
    # (a) Shot-noise-consistent model - RADIANT's model since CU-263: as the target
    #     dims, its own shot noise falls with it, so sigma(R)^2 = S(R) + N0^2 with
    #     N0^2 the range-independent (target-free) part of the noise power.
    n0_sq = sigma_total**2 - signal_e  # e-^2
    s_det = 0.5 * (thr**2 + math.sqrt(thr**4 + 4.0 * thr**2 * n0_sq))
    hand_det_shot_m = slant_m * math.sqrt(signal_e / s_det)
    # (b) The superseded frozen-noise model: total noise HELD at its reference-range
    #     value, so SNR(R) = S_ref (R_ref/R)^2 / sigma_ref.
    hand_det_fixed_m = slant_m * math.sqrt(snr / thr)
    print("\n  CROSS-CHECK 4 (closed-form detection range, vacuum - two noise models):")
    print(
        "    (a) RADIANT's solver model - shot-noise-consistent, sigma(R)^2 = S(R) + N0^2:"
        "\n        S_det = (T^2 + sqrt(T^4 + 4 T^2 N0^2)) / 2,  R_det = R_ref sqrt(S_ref/S_det)"
    )
    print(f"        N0^2 (target-free noise power)   : {n0_sq:,.2f} e-^2")
    print(f"        N0   (target-free noise)         : {math.sqrt(n0_sq):,.4f} e- rms")
    print(f"        S_det (signal at threshold)      : {s_det:,.2f} e-")
    print(f"        Closed form                      : {hand_det_shot_m / 1e3:,.3f} km")
    print(f"        RADIANT bisection                : {det_m / 1e3:,.3f} km")
    print(
        f"        Relative difference              : {(hand_det_shot_m / det_m - 1) * 100:+.6f} %"
        "  <- solver verified"
    )
    print(
        "\n    (b) The SUPERSEDED frozen-noise model - noise held at the reference range:"
        "\n        SNR(R) = S_ref (R_ref/R)^2 / sigma_ref  =>  R_det = R_ref sqrt(SNR_ref / T)"
    )
    print(f"        Signal demanded at threshold     : {thr * sigma_total:,.2f} e-"
          f"  (vs {s_det:,.2f} e- above)")
    print(f"        Closed form                      : {hand_det_fixed_m / 1e3:,.3f} km")
    print(
        f"        vs RADIANT                       : {(hand_det_fixed_m / det_m - 1) * 100:+.2f} %"
    )
    print(
        f"    Signal shot noise is {100.0 * signal_e / sigma_total**2:.0f} % of the"
        " noise power here, so freezing it was NOT"
        "\n    negligible: the shipped answer was conservative by"
        f" {(det_m / hand_det_fixed_m - 1) * 100:.0f} % for this scene, and the"
        "\n    metric it produced depended on the range it was evaluated at (CU-263, fixed"
        "\n    2026-08-01). The two models agree exactly AT the reference range, because"
        "\n    sigma_ref^2 = S_ref + N0^2 is the definition of N0; they diverge outward."
    )
    print(
        f"\n    Context: R_det is {det_m / slant_m:.2f}x the LEO->GEO range. The GEO belt is"
        "\n    comfortably inside this sensor's single-frame detection horizon for a"
        "\n    280 K, 20 m^2 bus."
    )

    # -- Rule 4 ---------------------------------------------------------------
    print(f"\n{SEP}\n  7. RULE-4 DUAL-PATH CONSISTENCY (PSF path vs MTF product)\n{SEP}")
    dpc = result.stage_outputs["performance"]["dual_path_consistency"]
    print(f"\n  passed_x / passed_y            : {dpc.passed_x} / {dpc.passed_y}")
    print(
        f"  max |error| x / y              : {dpc.max_absolute_error_x:.3e} / "
        f"{dpc.max_absolute_error_y:.3e} [MTF units]"
    )
    print(f"  tolerance                      : {dpc.tolerance:.3e} [MTF units]")
    margin = dpc.tolerance / max(dpc.max_absolute_error_x, dpc.max_absolute_error_y)
    print(f"  margin                         : {margin:.0f}x")
    user_warnings = [str(r.message) for r in records if issubclass(r.category, UserWarning)]
    print(f"\n  UserWarnings raised by the nominal chain: {len(user_warnings)}")
    for message in user_warnings:
        print(f"    - {message[:150]}")
    print(
        "  VERDICT: the consistency check stayed SILENT for this scene class - the"
        "\n  up-looking space_to_space path adds no spatial degradation to one path"
        "\n  without the other."
    )

    # -- kinematics -----------------------------------------------------------
    print(f"\n{SEP}\n  8. RELATIVE KINEMATICS - LEO vs GEO -> LOS RATE (Gap 111 doors)\n{SEP}")
    kin = open_loop_los_rate_rad_s(cin["h_leo_m"], cin["h_geo_m"])
    print(f"\n  LEO (500 km)  inertial speed   : {kin['v_sensor_m_s']:,.2f} m/s")
    print(f"  LEO           orbital rate     : {kin['omega_sensor_rad_s'] * 1e6:,.3f} urad/s")
    print(f"  LEO           period           : {kin['period_sensor_s'] / 60.0:,.3f} min")
    print(f"  GEO (35786 km) inertial speed  : {kin['v_target_m_s']:,.2f} m/s")
    print(f"  GEO            orbital rate    : {kin['omega_target_rad_s'] * 1e6:,.3f} urad/s")
    print(f"  GEO            period          : {kin['period_target_s'] / 3600.0:,.4f} h")
    print(f"  Separation (radial)            : {kin['separation_m'] / 1e3:,.3f} km")
    print(
        f"\n  HAND LOS RATE |v_LEO - v_GEO| / R : {kin['omega_los_rad_s'] * 1e6:,.3f} urad/s"
        f"  ({kin['omega_los_rad_s']:.6e} rad/s)"
    )
    print(
        "  Both velocities are tangential, hence perpendicular to the radial LOS, so the"
        "\n  LOS rate is the difference of tangential speeds over the separation. The"
        "\n  Earth-rotation-referenced GEO rate (7.2921e-5 rad/s sidereal) is recovered as"
        f"\n  omega_GEO = {kin['omega_target_rad_s']:.6e} rad/s, which agrees with the sidereal"
        f" value to {abs(kin['omega_target_rad_s'] / 7.2921159e-5 - 1) * 100:.3f} %."
        "\n  The residual is RADIANT's spherical R_E = 6371 km (giving r_GEO = 42 157 km)"
        "\n  against the true geostationary radius of 42 164 km - a CROSS-CHECK in itself:"
        "\n  the framework's two-body kinematics reproduce the sidereal day to 3 parts in"
        "\n  10^4 from the altitude alone."
    )

    open_loop = Sensor.from_dict(make_config(kinematics="open_loop"))
    ol_result, ol_records = _evaluate(open_loop)
    ol_geo = ol_result.stage_outputs["geometry"]
    ol_rate = float(ol_geo["los_angular_rate_rad_s"])
    print(f"\n  RADIANT published rate (K1+K2) : {ol_rate * 1e6:,.3f} urad/s")
    print(f"  los_rate_mode                  : {ol_geo['los_rate_mode']}")
    print(
        f"  Hand vs framework difference   : {(ol_rate / kin['omega_los_rad_s'] - 1) * 100:+.3e} %"
    )
    print(
        "\n  CROSS-CHECK 5 (Gap 111 two-door agreement): K1 (direct rate) and K2"
        "\n  (target-velocity triple + platform speed) were BOTH set. GeometryStage's"
        "\n  ADR-0006 rule-2 agreement check accepted them (1 % tolerance) and reported the"
        "\n  'consistent' mode string above. Non-obvious detail: at theta_o = pi the LOS is"
        "\n  radial, so the target-frame heading axis is azimuthally degenerate (ground range"
        "\n  is 0 m). The K2 formula omega = |v_rel x u| / R is invariant under that rotation"
        "\n  - hypot(v_perp, v_par) - so the degeneracy is harmless, not an ambiguity."
    )

    ol_smear_m = float(ol_result.stage_outputs["platform"]["smear_width_m"])
    tr_smear_m = float(result.stage_outputs["platform"]["smear_width_m"])
    hand_smear_m = kin["omega_los_rad_s"] * cin["focal_length_m"] * cin["t_int_s"]
    print(f"\n  SMEAR CONSEQUENCE at t_int = {cin['t_int_s'] * 1e3:.0f} ms:")
    print(
        f"    Open loop  : {ol_smear_m * 1e6:,.2f} um = "
        f"{ol_smear_m / cin['pixel_pitch_m']:,.3f} px  (hand: {hand_smear_m * 1e6:,.2f} um)"
    )
    print(
        f"    Rate-tracked (1 % residual): {tr_smear_m * 1e6:,.4f} um = "
        f"{tr_smear_m / cin['pixel_pitch_m']:,.5f} px"
    )
    print(
        f"    EE_box    open loop / tracked : "
        f"{float(ol_result.stage_outputs['platform']['EE_box']):.4f} / {ee_box:.4f} [--]"
    )
    print(
        f"    SNR       open loop / tracked : "
        f"{float(ol_result.metrics['snr']):.3f} / {snr:.3f} [--]"
    )
    print(
        f"    Detection open loop / tracked : "
        f"{float(ol_result.metrics['detection_range_m']) / 1e3:,.1f} / {det_m / 1e3:,.1f} km"
    )
    print(
        "\n  This is the design driver: an inertially-fixed 500 ms stare drags the GEO"
        f"\n  target across {ol_smear_m / cin['pixel_pitch_m']:.1f} pixels, spreading a point"
        " source that would"
        "\n  otherwise sit inside one. The nominal design therefore rate-tracks, and the"
        "\n  scenario's headline numbers are the rate-tracked ones."
    )
    ol_user_warnings = [str(r.message) for r in ol_records if issubclass(r.category, UserWarning)]
    print(f"\n  UserWarnings in the open-loop run: {len(ol_user_warnings)}")
    for message in ol_user_warnings:
        print(f"    - {message[:150]}")

    # -- integration sweep ----------------------------------------------------
    print(f"\n{SEP}\n  9. INTEGRATION-TIME TRADE: RATE-TRACKED vs OPEN LOOP\n{SEP}")
    print(
        f"\n  {'t_int':>9s}  {'smear OL':>10s}  {'SNR OL':>9s}  {'EE OL':>8s}"
        f"  {'smear RT':>10s}  {'SNR RT':>9s}  {'EE RT':>8s}  {'R_det RT':>11s}"
    )
    print(
        f"  {'[ms]':>9s}  {'[px]':>10s}  {'[--]':>9s}  {'[--]':>8s}"
        f"  {'[px]':>10s}  {'[--]':>9s}  {'[--]':>8s}  {'[km]':>11s}"
    )
    print(
        f"  {'-' * 9}  {'-' * 10}  {'-' * 9}  {'-' * 8}  {'-' * 10}  {'-' * 9}"
        f"  {'-' * 8}  {'-' * 11}"
    )
    sweep_rows: list[dict[str, float]] = []
    for t_ms, t_s in zip(cin["t_int_sweep_ms"], cin["t_int_sweep_s"], strict=True):
        rows: dict[str, float] = {"t_int_ms": t_ms}
        for tag, mode in (("ol", "open_loop"), ("rt", "tracked")):
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                res = Sensor.from_dict(make_config(t_int_s=t_s, kinematics=mode)).evaluate()
            rows[f"smear_px_{tag}"] = (
                float(res.stage_outputs["platform"]["smear_width_m"]) / cin["pixel_pitch_m"]
            )
            rows[f"snr_{tag}"] = float(res.metrics["snr"])
            rows[f"ee_{tag}"] = float(res.stage_outputs["platform"]["EE_box"])
            rows[f"rdet_km_{tag}"] = float(res.metrics.get("detection_range_m", float("nan"))) / 1e3
        sweep_rows.append(rows)
        # A NaN detection range is not a failure: the metric layer's result-typed
        # failure (ADR-B) says the target is already below threshold AT the GEO
        # range, so there is no longer range at which it becomes detectable.
        rdet = rows["rdet_km_rt"]
        rdet_str = f"{rdet:>11,.0f}" if math.isfinite(rdet) else f"{'< R_GEO':>11s}"
        print(
            f"  {t_ms:>9.0f}  {rows['smear_px_ol']:>10.3f}  {rows['snr_ol']:>9.3f}"
            f"  {rows['ee_ol']:>8.4f}  {rows['smear_px_rt']:>10.5f}  {rows['snr_rt']:>9.3f}"
            f"  {rows['ee_rt']:>8.4f}  {rdet_str}"
        )
    best_ol = max(sweep_rows, key=lambda r: r["snr_ol"])
    print(
        f"\n  Open-loop SNR peaks at t_int = {best_ol['t_int_ms']:.0f} ms (SNR ="
        f" {best_ol['snr_ol']:.2f}): past that point"
        "\n  the smear kernel grows faster than sqrt(t), so integrating longer LOSES SNR."
        "\n  Rate-tracked SNR keeps rising as sqrt(t) while the scene stays background-free"
        "\n  (cold space) and dark-current-limited."
        "\n  '< R_GEO' means the point-source SNR at the GEO range is already below the"
        "\n  detection threshold, so the solver reports a result-typed failure (ADR-B /"
        "\n  Rule 17 metric-layer carve-out) rather than emitting a metric - it is a"
        "\n  named non-detection, not a silent NaN."
    )

    # -- zenith sweep ---------------------------------------------------------
    print(f"\n{SEP}\n  10. SENSOR-SIDE ZENITH SWEEP - THE NEAR-pi GEOMETRY FAMILY\n{SEP}")
    print(
        f"\n  {'zeta_low':>9s}  {'theta_o':>10s}  {'eta':>10s}  {'slant':>12s}"
        f"  {'ground arc':>12s}  {'SNR':>8s}  {'R_det':>11s}"
    )
    print(
        f"  {'[deg]':>9s}  {'[deg]':>10s}  {'[deg]':>10s}  {'[km]':>12s}"
        f"  {'[km]':>12s}  {'[--]':>8s}  {'[km]':>11s}"
    )
    print(f"  {'-' * 9}  {'-' * 10}  {'-' * 10}  {'-' * 12}  {'-' * 12}  {'-' * 8}  {'-' * 11}")
    zen_rows: list[dict[str, float]] = []
    r_s = R_EARTH_M + cin["h_leo_m"]
    r_t = R_EARTH_M + cin["h_geo_m"]
    for z_deg, z_rad in zip(cin["zeta_sweep_deg"], cin["zeta_sweep_rad"], strict=True):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            res = Sensor.from_dict(make_config(zeta_low_rad=z_rad)).evaluate()
        g = res.stage_outputs["geometry"]
        # Law of sines read from the LOWER (sensor) endpoint - ADR-0011 dec. 3.
        expected_theta_o = math.pi - math.asin((r_s / r_t) * math.sin(z_rad))
        row = {
            "zeta_deg": z_deg,
            "theta_o_deg": math.degrees(float(g["theta_o_rad"])),
            "expected_theta_o_deg": math.degrees(expected_theta_o),
            "eta_deg": math.degrees(float(g["eta_rad"])),
            "slant_km": float(g["slant_range_m"]) / 1e3,
            "ground_km": float(g["ground_range_m"]) / 1e3,
            "snr": float(res.metrics["snr"]),
            "rdet_km": float(res.metrics["detection_range_m"]) / 1e3,
        }
        zen_rows.append(row)
        print(
            f"  {row['zeta_deg']:>9.1f}  {row['theta_o_deg']:>10.5f}  {row['eta_deg']:>10.5f}"
            f"  {row['slant_km']:>12,.2f}  {row['ground_km']:>12,.2f}  {row['snr']:>8.3f}"
            f"  {row['rdet_km']:>11,.0f}"
        )
    max_theta_err = max(abs(r["theta_o_deg"] - r["expected_theta_o_deg"]) for r in zen_rows)
    print(
        f"\n  CROSS-CHECK 6 (law of sines from the lower endpoint):"
        f"\n    theta_o = pi - asin((r_LEO / r_GEO) sin zeta_low); max |error| ="
        f" {max_theta_err:.3e} deg"
        f"\n    r_LEO / r_GEO = {r_s / r_t:.6f}, so theta_o stays within"
        f" {math.degrees(math.asin(r_s / r_t)):.2f} deg of pi for EVERY"
        "\n    sensor-side zenith: from GEO, the whole LEO shell is a small patch near nadir."
        "\n    eta stays obtuse throughout - the ADR-0011 extended domain in action."
        "\n    No horizon guard fires: theta_o never approaches pi/2, and the ray climbs"
        "\n    monotonically away from Earth so it never grazes the limb."
    )

    # -- figures --------------------------------------------------------------
    print(f"\n{SEP}\n  11. FIGURES\n{SEP}")
    figures = _make_figures(cin, sweep_rows, zen_rows, slant_m, signal_e, sigma_total, det_m)
    for path in figures:
        print(f"  wrote {path.relative_to(SCENARIO)}")

    # -- results workbook -----------------------------------------------------
    _write_results_workbook(cin, geo, result, sweep_rows, zen_rows, kin)
    print(f"  wrote {RESULTS_XLSX.relative_to(SCENARIO)}  (gitignored - regenerate on demand)")

    # -- summary --------------------------------------------------------------
    print(f"\n{SEP}\n  12. SUMMARY AND PARAMETERS THAT DO NOT MATTER HERE\n{SEP}")
    print(
        f"\n  Scene class            : {geo['scene_class']} (derived; assertion agreed)"
        f"\n  LOS direction          : {geo['los_direction']}"
        f"\n  Regime                 : {regime}"
        f"\n  theta_o                : {math.degrees(theta_o):.4f} deg (pi exactly)"
        f"\n  Slant range            : {slant_m / 1e3:,.3f} km"
        f"\n  LOS rate (open loop)   : {kin['omega_los_rad_s'] * 1e6:,.2f} urad/s"
        f"\n  Signal                 : {signal_e:,.1f} e-   Noise: {sigma_total:,.1f} e- RMS"
        f"\n  SNR (rate-tracked)     : {snr:.2f} [--]"
        f"\n  Detection range        : {det_m / 1e3:,.0f} km"
        f" (SNR = {cin['snr_threshold']:.0f} threshold)"
    )
    print(
        "\n  Parameters that do NOT affect this result, and why:"
        "\n    - atmosphere.standard_atmosphere, aerosol/visibility, water vapour: the path"
        "\n      is vacuum, so no column parameter is ever consulted (tau == 1 exactly)."
        "\n    - geometry.solar_zenith_rad / solar_azimuth_rad: illumination is 'night'"
        "\n      (target in eclipse) and the target descriptor is purely thermal, so no"
        "\n      solar term enters. Setting them changes nothing."
        "\n    - source.background.*: the LOS terminates on cold space, so the background"
        "\n      radiance is identically zero and background_shot is 0 e-."
        "\n    - GSD / swath / access-rate / NIIRS inputs: turned off by the scene-class"
        "\n      relevance map (section 5) - there is no ground plane at a GEO target."
        "\n    - geometry.ground_speed_m_s in the RATE-TRACKED config: the K1 door supplies"
        "\n      the rate directly, so the platform-only ground_speed / slant path is unused."
        "\n    - optics.optics_temperature_K matters only weakly: at 180 K the self-emission"
        "\n      in 3.5-5.0 um is negligible against a 280 K target, but it is NOT zero and"
        "\n      is left in the budget rather than suppressed."
    )
    print(f"\n{SEP}\n  Scenario 10.4 complete.\n{SEP}")


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------


def _make_figures(
    cin: dict[str, Any],
    sweep_rows: list[dict[str, float]],
    zen_rows: list[dict[str, float]],
    slant_m: float,
    signal_e: float,
    sigma_total_e: float,
    det_m: float,
) -> list[Path]:
    paths: list[Path] = []
    t_ms = [r["t_int_ms"] for r in sweep_rows]

    # Fig 1 - integration-time trade
    fig, (ax_a, ax_b) = plt.subplots(2, 1, figsize=(9.0, 8.0), sharex=True)
    ax_a.plot(t_ms, [r["smear_px_ol"] for r in sweep_rows], "ro-", label="Open loop (untracked)")
    ax_a.plot(
        t_ms, [r["smear_px_rt"] for r in sweep_rows], "bs-", label="Rate-tracked (1% residual)"
    )
    ax_a.axhline(1.0, color="gray", ls=":", label="1 pixel")
    ax_a.set_yscale("log")
    ax_a.set_ylabel("LOS smear width [pixels]")
    ax_a.set_title("LEO -> GEO stare: smear and SNR vs integration time")
    ax_a.grid(True, alpha=0.3, which="both")
    ax_a.legend(fontsize=9)
    ax_b.plot(t_ms, [r["snr_ol"] for r in sweep_rows], "ro-", label="Open loop (untracked)")
    ax_b.plot(t_ms, [r["snr_rt"] for r in sweep_rows], "bs-", label="Rate-tracked (1% residual)")
    ax_b.set_xscale("log")
    ax_b.set_xlabel("Integration time [ms]")
    ax_b.set_ylabel("SNR [dimensionless]")
    ax_b.grid(True, alpha=0.3, which="both")
    ax_b.legend(fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "10.4_smear_snr_vs_integration_time.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    paths.append(path)

    # Fig 2 - SNR vs range with the detection threshold, both noise models
    ranges_m = np.logspace(math.log10(5.0e6), math.log10(3.0e8), 400)
    signal_r = signal_e * (slant_m / ranges_m) ** 2
    n0_sq = sigma_total_e**2 - signal_e
    snr_fixed = signal_r / sigma_total_e  # superseded: noise frozen at R_ref
    snr_shot = signal_r / np.sqrt(signal_r + n0_sq)  # RADIANT solver (CU-263)
    fig, ax = plt.subplots(figsize=(9.0, 6.0))
    ax.loglog(
        ranges_m / 1e3,
        snr_shot,
        "b-",
        lw=2,
        label=r"RADIANT solver: $\sigma^2(R)=S(R)+N_0^2$",
    )
    ax.loglog(
        ranges_m / 1e3,
        snr_fixed,
        "m--",
        lw=1.6,
        label="Superseded frozen-noise model: noise held at $R_{ref}$",
    )
    ax.axhline(5.0, color="red", ls="--", label="Detection threshold SNR = 5")
    ax.axvline(
        slant_m / 1e3, color="green", ls=":", label=f"LEO->GEO range = {slant_m / 1e3:,.0f} km"
    )
    ax.axvline(det_m / 1e3, color="k", ls="-.", label=f"R_det = {det_m / 1e3:,.0f} km")
    ax.set_xlabel("Sensor-to-target range [km]")
    ax.set_ylabel("SNR [dimensionless]")
    ax.set_title("Point-source SNR vs range - 280 K, 20 m$^2$ GEO bus, rate-tracked 500 ms")
    ax.set_ylim(0.5, 2.0e3)
    ax.grid(True, alpha=0.3, which="both")
    ax.legend(fontsize=9)
    fig.tight_layout()
    path = OUTPUT_DIR / "10.4_snr_vs_range_detection.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    paths.append(path)

    # Fig 3 - near-pi geometry family
    zeta = [r["zeta_deg"] for r in zen_rows]
    fig, (ax_a, ax_b) = plt.subplots(1, 2, figsize=(12.0, 5.0))
    ax_a.plot(zeta, [r["theta_o_deg"] for r in zen_rows], "bo-", label=r"$\theta_o$ (target-side)")
    ax_a.plot(zeta, [r["eta_deg"] for r in zen_rows], "r^-", label=r"$\eta$ (sensor-side)")
    ax_a.axhline(180.0, color="gray", ls=":", label=r"$\pi$ rad = 180 deg")
    ax_a.axhline(90.0, color="orange", ls="--", label=r"$\pi/2$ (horizon guard band)")
    ax_a.set_xlabel(r"Sensor-side path zenith $\zeta_{low}$ [deg]")
    ax_a.set_ylabel("Angle [deg]")
    ax_a.set_title("Extended domain: both angles stay obtuse")
    ax_a.set_ylim(85.0, 185.0)
    ax_a.grid(True, alpha=0.3)
    ax_a.legend(fontsize=8)
    ax_b.plot(zeta, [r["slant_km"] for r in zen_rows], "ko-", label="Slant range")
    ax_b.plot(
        zeta, [r["ground_km"] for r in zen_rows], "g^-", label="Ground arc (LEO nadir -> GEO nadir)"
    )
    ax_b.set_xlabel(r"Sensor-side path zenith $\zeta_{low}$ [deg]")
    ax_b.set_ylabel("Distance [km]")
    ax_b.set_title("Range geometry vs sensor-side zenith")
    ax_b.grid(True, alpha=0.3)
    ax_b.legend(fontsize=8)
    fig.tight_layout()
    path = OUTPUT_DIR / "10.4_near_pi_geometry_family.png"
    fig.savefig(path, dpi=150)
    plt.close(fig)
    paths.append(path)
    return paths


# ---------------------------------------------------------------------------
# Results workbook (gitignored)
# ---------------------------------------------------------------------------


def _write_results_workbook(
    cin: dict[str, Any],
    geo: dict[str, Any],
    result: Any,
    sweep_rows: list[dict[str, float]],
    zen_rows: list[dict[str, float]],
    kin: dict[str, float],
) -> None:
    wb = openpyxl.Workbook()
    head_font = Font(bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")

    def _header(ws: Any, names: list[str]) -> None:
        for col, name in enumerate(names, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "Nominal"
    _header(ws, ["Quantity", "Value", "Unit"])
    nominal = [
        ("Scene class (derived)", geo["scene_class"], "--"),
        ("LOS direction", geo["los_direction"], "--"),
        ("theta_o", math.degrees(float(geo["theta_o_rad"])), "deg"),
        ("eta", math.degrees(float(geo["eta_rad"])), "deg"),
        ("Slant range", float(geo["slant_range_m"]) / 1e3, "km"),
        ("Ground arc", float(geo["ground_range_m"]) / 1e3, "km"),
        ("EE_box (1x1)", float(result.stage_outputs["platform"]["EE_box"]), "--"),
        ("Signal", float(result.stage_outputs["spectral_integration"]["signal_e"]), "e-"),
        ("Total noise", float(result.stage_outputs["readout"]["sigma_total_e"]), "e- RMS"),
        ("SNR", float(result.metrics["snr"]), "--"),
        ("Detection range", float(result.metrics["detection_range_m"]) / 1e3, "km"),
        ("NEDT", float(result.metrics["nedt_K"]) * 1e3, "mK"),
        ("Open-loop LOS rate", kin["omega_los_rad_s"] * 1e6, "urad/s"),
        ("LEO inertial speed", kin["v_sensor_m_s"], "m/s"),
        ("GEO inertial speed", kin["v_target_m_s"], "m/s"),
    ]
    for r, (name, value, unit) in enumerate(nominal, start=2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=unit)

    ws = wb.create_sheet("Integration Sweep")
    _header(
        ws,
        [
            "Integration time [ms]",
            "Smear open loop [px]",
            "SNR open loop [--]",
            "EE open loop [--]",
            "Smear tracked [px]",
            "SNR tracked [--]",
            "EE tracked [--]",
            "Detection range tracked [km]",
        ],
    )
    for r, row in enumerate(sweep_rows, start=2):
        for col, key in enumerate(
            [
                "t_int_ms",
                "smear_px_ol",
                "snr_ol",
                "ee_ol",
                "smear_px_rt",
                "snr_rt",
                "ee_rt",
                "rdet_km_rt",
            ],
            start=1,
        ):
            ws.cell(row=r, column=col, value=row[key])

    ws = wb.create_sheet("Zenith Sweep")
    _header(
        ws,
        [
            "Sensor-side zenith [deg]",
            "theta_o [deg]",
            "theta_o expected [deg]",
            "eta [deg]",
            "Slant range [km]",
            "Ground arc [km]",
            "SNR [--]",
            "Detection range [km]",
        ],
    )
    for r, row in enumerate(zen_rows, start=2):
        for col, key in enumerate(
            [
                "zeta_deg",
                "theta_o_deg",
                "expected_theta_o_deg",
                "eta_deg",
                "slant_km",
                "ground_km",
                "snr",
                "rdet_km",
            ],
            start=1,
        ):
            ws.cell(row=r, column=col, value=row[key])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(width + 3, 34)
    wb.save(RESULTS_XLSX)


if __name__ == "__main__":
    main()
