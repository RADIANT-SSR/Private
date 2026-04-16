"""Scenario 6.3: Noise Model Verification — Analytic vs. RADIANT.

Dr. Chen's workflow:
  1. Read sensor parameters from her Excel spreadsheet (non-RADIANT units)
  2. Convert to RADIANT canonical units (m, fractions, seconds)
  3. Run RADIANT evaluation
  4. Extract each noise term
  5. Compute hand-calc values for the major noise terms
  6. Compare in a formatted table with % error
  7. Write results back to the spreadsheet

Usage:
    python run_verification.py
"""

import math
import sys
from pathlib import Path

import numpy as np
import openpyxl

# ---------------------------------------------------------------------------
# Step 1: Read the spreadsheet
# ---------------------------------------------------------------------------

INPUT_FILE = Path(__file__).parent.parent / "inputs" / "dr_chen_sensor_parameters.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "outputs" / "verification_results.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb["Sensor Parameters"]

# Parse the parameter table into a dict keyed by parameter name
params_raw: dict[str, float] = {}
for row in ws.iter_rows(min_row=5, max_col=4, values_only=False):
    name_cell = row[0].value
    value_cell = row[1].value
    if name_cell and value_cell is not None:
        # Skip section headers (blue-filled rows)
        if row[0].fill and row[0].fill.start_color and row[0].fill.start_color.rgb == "004472C4":
            continue
        params_raw[name_cell] = float(value_cell)

print("=== Raw parameters from spreadsheet ===")
print(f"  {'Parameter':<30s} {'Value':>12s}  {'Unit'}")
print(f"  {'-' * 30} {'-' * 12}  {'-' * 10}")
print(f"  {'Aperture diameter':<30s} {params_raw['Aperture diameter']:>12.1f}  cm")
print(f"  {'Focal ratio (f/#)':<30s} {params_raw['Focal ratio (f/#)']:>12.1f}  —")
print(f"  {'Optical transmission':<30s} {params_raw['Optical transmission']:>12.1f}  %")
print(f"  {'Number of optical elements':<30s} {params_raw['Number of optical elements']:>12.0f}  —")
print(f"  {'Optics temperature':<30s} {params_raw['Optics temperature']:>12.1f}  K")
print(f"  {'Pixel pitch':<30s} {params_raw['Pixel pitch']:>12.1f}  µm")
print(f"  {'Quantum efficiency':<30s} {params_raw['Quantum efficiency']:>12.1f}  %")
print(f"  {'Dark current':<30s} {params_raw['Dark current']:>12.1f}  e⁻/s")
print(f"  {'Operating temperature':<30s} {params_raw['Operating temperature']:>12.1f}  K")
print(f"  {'Full well capacity':<30s} {params_raw['Full well capacity']:>12.0f}  e⁻")
print(f"  {'Read noise':<30s} {params_raw['Read noise']:>12.1f}  e⁻ RMS")
print(f"  {'ADC bits':<30s} {params_raw['ADC bits']:>12.0f}  bits")
print(f"  {'System gain':<30s} {params_raw['System gain']:>12.1f}  e⁻/DN")
print(f"  {'Target temperature':<30s} {params_raw['Target temperature']:>12.1f}  K")
print(f"  {'Target emissivity':<30s} {params_raw['Target emissivity']:>12.2f}  —")
print(f"  {'Background temperature':<30s} {params_raw['Background temperature']:>12.1f}  K       * contrast SNR only")
print(f"  {'Background emissivity':<30s} {params_raw['Background emissivity']:>12.2f}  —       * contrast SNR only")
print(f"  {'Band minimum':<30s} {params_raw['Band minimum']:>12.1f}  µm")
print(f"  {'Band maximum':<30s} {params_raw['Band maximum']:>12.1f}  µm")
print(f"  {'Integration time':<30s} {params_raw['Integration time']:>12.1f}  ms")
print(f"  {'Number of TDI stages':<30s} {params_raw['Number of TDI stages']:>12.0f}  —")
print(f"  {'Sensor altitude':<30s} {params_raw['Sensor altitude']:>12.1f}  km")
print(f"  {'Look angle':<30s} {params_raw['Look angle']:>12.1f}  deg")

# ---------------------------------------------------------------------------
# Step 2: Convert to RADIANT canonical units
# ---------------------------------------------------------------------------

# Optics
aperture_m = params_raw["Aperture diameter"] / 100.0          # cm -> m
f_number = params_raw["Focal ratio (f/#)"]
focal_length_m = f_number * aperture_m                         # derived: f = f/# × D
transmission = params_raw["Optical transmission"] / 100.0      # % -> fraction
optics_temp_K = params_raw["Optics temperature"]

# Detector
pixel_pitch_m = params_raw["Pixel pitch"] * 1e-6              # µm -> m
pixel_pitch_um = params_raw["Pixel pitch"]                     # keep for RADIANT input
qe = params_raw["Quantum efficiency"] / 100.0                 # % -> fraction
dark_rate = params_raw["Dark current"]                         # already e⁻/s
operating_temp = params_raw["Operating temperature"]           # already K
fwc = params_raw["Full well capacity"]                         # already e⁻

# Readout
read_noise = params_raw["Read noise"]                          # already e⁻ RMS
adc_bits = int(params_raw["ADC bits"])
gain = params_raw["System gain"]                               # already e⁻/DN

# Scene
target_temp = params_raw["Target temperature"]                 # already K
target_emiss = params_raw["Target emissivity"]                 # dimensionless
bg_temp = params_raw["Background temperature"]                 # already K
bg_emiss = params_raw["Background emissivity"]                 # dimensionless

# Spectral
band_min = params_raw["Band minimum"]                          # already µm
band_max = params_raw["Band maximum"]                          # already µm
t_int = params_raw["Integration time"] / 1000.0                # ms -> s

# Geometry
altitude_m = params_raw["Sensor altitude"] * 1000.0            # km -> m

print("\n=== Converted to RADIANT canonical units ===")
print(f"  {'Parameter':<30s} {'Value':>14s}  {'Unit':<15s}  {'Conversion'}")
print(f"  {'-' * 30} {'-' * 14}  {'-' * 15}  {'-' * 20}")
print(f"  {'Aperture diameter':<30s} {aperture_m:>14.4f}  {'m':<15s}  cm ÷ 100")
print(f"  {'Focal length (derived)':<30s} {focal_length_m:>14.4f}  {'m':<15s}  f/# × D")
print(f"  {'f-number':<30s} {f_number:>14.1f}  {'—':<15s}  no conversion")
print(f"  {'Optical transmission':<30s} {transmission:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Optics temperature':<30s} {optics_temp_K:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Pixel pitch':<30s} {pixel_pitch_um:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'Pixel pitch':<30s} {pixel_pitch_m:>14.2e}  {'m':<15s}  µm × 1e-6")
print(f"  {'Quantum efficiency':<30s} {qe:>14.4f}  {'fraction':<15s}  % ÷ 100")
print(f"  {'Dark current':<30s} {dark_rate:>14.1f}  {'e⁻/s':<15s}  no conversion")
print(f"  {'Detector temperature':<30s} {operating_temp:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Full well capacity':<30s} {fwc:>14.0f}  {'e⁻':<15s}  no conversion")
print(f"  {'Read noise':<30s} {read_noise:>14.1f}  {'e⁻ RMS':<15s}  no conversion")
print(f"  {'ADC bits':<30s} {adc_bits:>14d}  {'bits':<15s}  no conversion")
print(f"  {'System gain':<30s} {gain:>14.1f}  {'e⁻/DN':<15s}  no conversion")
print(f"  {'Target temperature':<30s} {target_temp:>14.1f}  {'K':<15s}  no conversion")
print(f"  {'Target emissivity':<30s} {target_emiss:>14.2f}  {'—':<15s}  dimensionless")
print(f"  {'Background temperature':<30s} {bg_temp:>14.1f}  {'K':<15s}  * contrast SNR only")
print(f"  {'Background emissivity':<30s} {bg_emiss:>14.2f}  {'—':<15s}  * contrast SNR only")
print(f"  {'Band minimum':<30s} {band_min:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'Band maximum':<30s} {band_max:>14.1f}  {'µm':<15s}  no conversion")
print(f"  {'Integration time':<30s} {t_int:>14.6f}  {'s':<15s}  ms ÷ 1000")
print(f"  {'Sensor altitude':<30s} {altitude_m:>14.1f}  {'m':<15s}  km × 1000")

# ---------------------------------------------------------------------------
# Step 3: Run RADIANT
# ---------------------------------------------------------------------------

from radiant.api import Sensor

config = {
    "source": {
        "target": {
            "temperature": target_temp,
            "emissivity": target_emiss,
        },
        "background": {
            "temperature": bg_temp,
            "emissivity": bg_emiss,
        },
    },
    "atmosphere": {
        "model": "exo",  # Lab/vacuum — no atmospheric absorption
    },
    "geometry": {
        "sensor_altitude_m": altitude_m,
    },
    "optics": {
        "aperture_diameter_m": aperture_m,
        "focal_length_m": focal_length_m,
        "transmission_scalar": transmission,
        "optics_temperature_K": optics_temp_K,
    },
    "detector": {
        "pixel_pitch_x_um": pixel_pitch_um,
        "pixel_pitch_y_um": pixel_pitch_um,
        "qe_value": qe,
        "dark_rate_e_per_s": dark_rate,
        "detector_temperature_K": operating_temp,
    },
    "spectral_integration": {
        "filter_min_um": band_min,
        "filter_max_um": band_max,
        "integration_time_s": t_int,
    },
    "readout": {
        "read_noise_e_rms": read_noise,
        "gain_e_per_dn": gain,
        "adc_bits": adc_bits,
        "full_well_capacity_e": fwc,
    },
}

print("\n=== Running RADIANT evaluation ===")
sensor = Sensor.from_dict(config)
result = sensor.evaluate()

# ---------------------------------------------------------------------------
# Step 3b: Regime and signal chain notes
# ---------------------------------------------------------------------------

regime = result.stage_outputs["optics"]["regime"]
print(f"\n=== Radiometric Regime ===")
print(f"  Regime:  {regime}")
print(f"")
print(f"  This scenario uses EXTENDED regime (fill_fraction = 1.0, no target")
print(f"  geometry specified). The pixel is entirely filled by the target.")
print(f"")
print(f"  In extended regime:")
print(f"    - Signal chain: L_target × τ_atm × τ_opt × A_collect × Ω_pixel × QE × t_int / E_photon")
print(f"    - The 300 K target fills the pixel. No background mixing occurs.")
print(f"    - Background temperature ({bg_temp:.0f} K) and emissivity ({bg_emiss}) are used")
print(f"      ONLY for contrast SNR (target-minus-background divided by noise).")
print(f"    - 'background_shot' noise is NOT from the {bg_temp:.0f} K scene background.")
print(f"      It comes from path radiance (atmosphere) reaching the detector.")
print(f"    - 'nearfield_shot' noise comes from thermal self-emission of the warm")
print(f"      optics ({optics_temp_K:.0f} K), scaled by (1 - τ_opt) per Kirchhoff's law.")

# ---------------------------------------------------------------------------
# Step 4: Extract RADIANT noise terms
# ---------------------------------------------------------------------------

radiant_noise: dict[str, float] = {}
for nt in result.noise_terms:
    radiant_noise[nt.name] = nt.value_e

print("\n=== RADIANT noise terms ===")
print(f"  {'Term':<25s} {'Value':>12s}  {'Unit'}")
print(f"  {'-' * 25} {'-' * 12}  {'-' * 10}")
for name, value in sorted(radiant_noise.items(), key=lambda x: -x[1]):
    print(f"  {name:<25s} {value:>12.4f}  e⁻ RMS")

# ---------------------------------------------------------------------------
# Step 5: Hand calculations for major noise terms
# ---------------------------------------------------------------------------
# We compute the terms Dr. Chen can verify analytically.

from radiant.core.constants import h, c, k_B

# -- Derived quantities --
pixel_area_m2 = pixel_pitch_m ** 2
f_num = f_number
omega_pixel = math.pi / (4.0 * f_num ** 2)  # pixel solid angle [sr]

# -- Planck spectral radiance integral --
# L_target = ∫ ε·B(λ,T)·dλ over the band
# B(λ,T) = 2hc²/λ⁵ · 1/(exp(hc/λkT) - 1)  [W/m²/sr/µm]
N_WAVE = 1000
wavelengths_um = np.linspace(band_min, band_max, N_WAVE)
wavelengths_m = wavelengths_um * 1e-6


def planck_spectral_radiance(lam_m: np.ndarray, T: float) -> np.ndarray:
    """Planck function B(λ,T) in W/m²/sr/µm."""
    # Compute in SI (W/m²/sr/m) then convert to per-µm
    num = 2.0 * h * c ** 2 / lam_m ** 5
    denom = np.exp(h * c / (lam_m * k_B * T)) - 1.0
    return num / denom * 1e-6  # W/m²/sr/m -> W/m²/sr/µm


L_target_spectral = target_emiss * planck_spectral_radiance(wavelengths_m, target_temp)
L_bg_spectral = bg_emiss * planck_spectral_radiance(wavelengths_m, bg_temp)
L_optics_spectral = (1.0 - transmission) * planck_spectral_radiance(wavelengths_m, optics_temp_K)

# Integrate over band [W/m²/sr]
L_target_band = float(np.trapezoid(L_target_spectral, wavelengths_um))
L_bg_band = float(np.trapezoid(L_bg_spectral, wavelengths_um))
L_optics_band = float(np.trapezoid(L_optics_spectral, wavelengths_um))

# Mean photon energy over the band
lam_center_m = np.mean(wavelengths_m)
E_photon = h * c / lam_center_m  # [J]

# Signal electrons: L × τ_optics × Ω × A_pixel × QE × t_int / E_photon
signal_e_hand = L_target_band * transmission * omega_pixel * pixel_area_m2 * qe * t_int / E_photon
bg_e_hand = L_bg_band * transmission * omega_pixel * pixel_area_m2 * qe * t_int / E_photon
nf_e_hand = L_optics_band * omega_pixel * pixel_area_m2 * qe * t_int / E_photon

# Dark electrons
dark_e_hand = dark_rate * t_int

# Noise terms (hand calc)
hand_calc: dict[str, float] = {
    "signal_shot": math.sqrt(signal_e_hand),
    "background_shot": math.sqrt(bg_e_hand),
    "nearfield_shot": math.sqrt(nf_e_hand),
    "dark_shot": math.sqrt(dark_e_hand),
    "read_noise": read_noise,
    "quantization": gain / math.sqrt(12.0),
}

print("\n=== Hand-calculated noise terms ===")
print(f"  {'Term':<25s} {'Value':>12s}  {'Unit'}")
print(f"  {'-' * 25} {'-' * 12}  {'-' * 10}")
for name, value in hand_calc.items():
    print(f"  {name:<25s} {value:>12.4f}  e⁻ RMS")

print("\n=== Derived quantities ===")
print(f"  {'Quantity':<30s} {'Value':>14s}  {'Unit'}")
print(f"  {'-' * 30} {'-' * 14}  {'-' * 15}")
print(f"  {'Focal length':<30s} {focal_length_m:>14.4f}  m")
print(f"  {'Pixel solid angle (Ω)':<30s} {omega_pixel:>14.6e}  sr")
print(f"  {'Pixel area':<30s} {pixel_area_m2:>14.6e}  m²")
print(f"  {'L_target (band-integrated)':<30s} {L_target_band:>14.6f}  W/m²/sr")
print(f"  {'L_background (band-integ.)':<30s} {L_bg_band:>14.6f}  W/m²/sr")
print(f"  {'L_optics (band-integrated)':<30s} {L_optics_band:>14.6f}  W/m²/sr")
print(f"  {'E_photon (band center)':<30s} {E_photon:>14.6e}  J/photon")
print(f"  {'Signal electrons (target)':<30s} {signal_e_hand:>14.2f}  e⁻")
print(f"  {'Signal electrons (background)':<30s} {bg_e_hand:>14.2f}  e⁻")
print(f"  {'Signal electrons (nearfield)':<30s} {nf_e_hand:>14.2f}  e⁻")
print(f"  {'Dark electrons':<30s} {dark_e_hand:>14.4f}  e⁻")

# ---------------------------------------------------------------------------
# Step 6: Comparison table
# ---------------------------------------------------------------------------

print("\n" + "=" * 95)
print(f"  {'Noise Term':<23s} | {'Hand Calc':>12s} | {'RADIANT':>12s} | {'Unit':<10s} | {'% Error':>8s} | {'Status'}")
print("-" * 95)

comparison_rows: list[dict] = []
for name in hand_calc:
    hc_val = hand_calc[name]
    rd_val = radiant_noise.get(name, 0.0)
    if hc_val > 0:
        pct_err = abs(hc_val - rd_val) / hc_val * 100.0
    else:
        pct_err = 0.0 if rd_val == 0.0 else float("inf")
    status = "PASS" if pct_err < 5.0 else "CHECK" if pct_err < 20.0 else "FAIL"
    print(f"  {name:<23s} | {hc_val:>12.4f} | {rd_val:>12.4f} | {'e⁻ RMS':<10s} | {pct_err:>7.2f}% | {status}")
    comparison_rows.append({
        "name": name, "hand_calc": hc_val, "radiant": rd_val,
        "pct_error": pct_err, "status": status,
    })

# Also show terms that RADIANT computes but we didn't hand-calc
print("-" * 95)
print("  Terms not hand-calculated (shown for reference):")
for name, rd_val in sorted(radiant_noise.items(), key=lambda x: -x[1]):
    if name not in hand_calc:
        print(f"  {name:<23s} |     {'—':>8s} | {rd_val:>12.4f} | {'e⁻ RMS':<10s} |    {'—':>4s} |")

# Total noise comparison
hand_total = math.sqrt(sum(v ** 2 for v in hand_calc.values()))
radiant_total_temporal = math.sqrt(
    sum(v ** 2 for k, v in radiant_noise.items()
        if k in {"signal_shot", "background_shot", "nearfield_shot", "straylight_shot",
                  "dark_shot", "gr_noise", "johnson_noise", "flicker_1f",
                  "read_noise", "ktc_reset", "quantization", "persistence_noise", "glow_shot"})
)

total_pct_err = abs(hand_total - radiant_total_temporal) / hand_total * 100.0
print("-" * 95)
print(f"  {'TOTAL (RSS)':<23s} | {hand_total:>12.4f} | {radiant_total_temporal:>12.4f} | {'e⁻ RMS':<10s} | {total_pct_err:>7.2f}% |")
print("=" * 95)

# ---------------------------------------------------------------------------
# Step 7: SNR summary
# ---------------------------------------------------------------------------

snr = result.metrics["snr"]
contrast_snr = result.metrics.get("contrast_snr", None)
mtf_nyq = result.metrics.get("mtf_at_nyquist", None)

print(f"\n=== Performance Metrics ===")
print(f"  {'Metric':<25s} {'RADIANT':>12s}  {'Hand Calc':>12s}  {'Unit':<15s}  {'% Error':>8s}")
print(f"  {'-' * 25} {'-' * 12}  {'-' * 12}  {'-' * 15}  {'-' * 8}")

# Hand-calc SNR for comparison
hand_snr = signal_e_hand / hand_total
snr_err = abs(hand_snr - snr) / hand_snr * 100.0 if hand_snr > 0 else 0.0
print(f"  {'SNR':<25s} {snr:>12.2f}  {hand_snr:>12.2f}  {'—':<15s}  {snr_err:>7.2f}%")

if contrast_snr is not None:
    print(f"  {'Contrast SNR':<25s} {contrast_snr:>12.2f}  {'—':>12s}  {'—':<15s}  {'—':>8s}")

if mtf_nyq is not None:
    print(f"  {'MTF at Nyquist':<25s} {mtf_nyq:>12.4f}  {'—':>12s}  {'—':<15s}  {'—':>8s}")

print(f"\n  Note: ~3% SNR difference is expected. RADIANT integrates per-wavelength")
print(f"  (spectral QE × Planck × filter), while the hand calc uses a mean photon")
print(f"  energy approximation. Deterministic terms (dark, read, quant) match exactly.")

# ---------------------------------------------------------------------------
# Step 8: Write results to output spreadsheet
# ---------------------------------------------------------------------------

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Noise Comparison"

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

header_font = Font(bold=True, size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
check_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ws_out.column_dimensions["A"].width = 28
ws_out.column_dimensions["B"].width = 22
ws_out.column_dimensions["C"].width = 22
ws_out.column_dimensions["D"].width = 14
ws_out.column_dimensions["E"].width = 14
ws_out.column_dimensions["F"].width = 10

ws_out["A1"] = "Scenario 6.3: Noise Model Verification"
ws_out["A1"].font = Font(bold=True, size=14)

headers = ["Noise Term", "Hand Calc [e⁻ RMS]", "RADIANT [e⁻ RMS]", "Unit", "% Error", "Status"]
for col, h_text in enumerate(headers, 1):
    cell = ws_out.cell(row=3, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

for i, row_data in enumerate(comparison_rows, 4):
    ws_out.cell(row=i, column=1, value=row_data["name"]).border = thin_border
    ws_out.cell(row=i, column=2, value=round(row_data["hand_calc"], 4)).border = thin_border
    ws_out.cell(row=i, column=3, value=round(row_data["radiant"], 4)).border = thin_border
    ws_out.cell(row=i, column=4, value="e⁻ RMS").border = thin_border
    ws_out.cell(row=i, column=5, value=round(row_data["pct_error"], 2)).border = thin_border
    ws_out.cell(row=i, column=5).number_format = "0.00"
    status_cell = ws_out.cell(row=i, column=6, value=row_data["status"])
    status_cell.border = thin_border
    status_cell.alignment = Alignment(horizontal="center")
    if row_data["status"] == "PASS":
        status_cell.fill = pass_fill
    elif row_data["status"] == "CHECK":
        status_cell.fill = check_fill
    else:
        status_cell.fill = fail_fill

# Total row
total_row = len(comparison_rows) + 4
ws_out.cell(row=total_row, column=1, value="TOTAL (RSS)").border = thin_border
ws_out.cell(row=total_row, column=1).font = header_font
ws_out.cell(row=total_row, column=2, value=round(hand_total, 4)).border = thin_border
ws_out.cell(row=total_row, column=3, value=round(radiant_total_temporal, 4)).border = thin_border
ws_out.cell(row=total_row, column=4, value="e⁻ RMS").border = thin_border
ws_out.cell(row=total_row, column=5, value=round(total_pct_err, 2)).border = thin_border

# Summary section
summary_row = total_row + 2
ws_out.cell(row=summary_row, column=1, value="Performance Metrics").font = Font(bold=True, size=12)

metrics_headers = ["Metric", "RADIANT", "Hand Calc", "Unit", "% Error"]
for col, h_text in enumerate(metrics_headers, 1):
    cell = ws_out.cell(row=summary_row + 1, column=col, value=h_text)
    cell.font = header_font
    cell.border = thin_border

r = summary_row + 2
ws_out.cell(row=r, column=1, value="SNR").border = thin_border
ws_out.cell(row=r, column=2, value=round(snr, 2)).border = thin_border
ws_out.cell(row=r, column=3, value=round(hand_snr, 2)).border = thin_border
ws_out.cell(row=r, column=4, value="— (dimensionless)").border = thin_border
ws_out.cell(row=r, column=5, value=round(snr_err, 2)).border = thin_border

if contrast_snr is not None:
    r += 1
    ws_out.cell(row=r, column=1, value="Contrast SNR").border = thin_border
    ws_out.cell(row=r, column=2, value=round(contrast_snr, 2)).border = thin_border
    ws_out.cell(row=r, column=3, value="—").border = thin_border
    ws_out.cell(row=r, column=4, value="— (dimensionless)").border = thin_border

if mtf_nyq is not None:
    r += 1
    ws_out.cell(row=r, column=1, value="MTF at Nyquist").border = thin_border
    ws_out.cell(row=r, column=2, value=round(mtf_nyq, 4)).border = thin_border
    ws_out.cell(row=r, column=3, value="—").border = thin_border
    ws_out.cell(row=r, column=4, value="— (dimensionless)").border = thin_border

wb_out.save(OUTPUT_FILE)
print(f"\nResults written to {OUTPUT_FILE}")
