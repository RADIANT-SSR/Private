#!/usr/bin/env python3
"""Create Sarah's inputs for Scenario 1.5: obscured aperture + spider vanes.

Emits ``sarah_cassegrain.xlsx`` — the Cassegrain telescope configuration
(primary diameter, central obscuration, secondary-support spider arms) and
the VNIR detector/readout parameters Sarah evaluates.

Run:  python create_spreadsheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Telescope"
ws["A1"] = "Scenario 1.5: Cassegrain — obscuration + spider vanes"
ws["A1"].font = Font(bold=True, size=14)
for col, w in zip("ABCD", [28, 14, 12, 46]):
    ws.column_dimensions[col].width = w
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="375623")
for col, text in zip("ABCD", ["Parameter", "Value", "Unit", "Notes"]):
    ws[f"{col}3"] = text
    ws[f"{col}3"].font = hdr_font
    ws[f"{col}3"].fill = hdr_fill
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ROWS = [
    ("Primary diameter", 50.0, "cm", "Clear primary aperture"),
    ("Obscuration ratio", 0.30, "—", "D_secondary / D_primary (Cassegrain)"),
    ("Spider arm count", 4, "—", "Secondary support struts"),
    ("Spider arm width", 3.0, "cm", "Baseline strut width; swept 0–5 cm"),
    ("Focal length", 6.0, "m", "f/12"),
    ("Waveband", "450–700", "nm", "VNIR panchromatic"),
    ("Pixel pitch", 6.5, "µm", ""),
    ("Optical transmission", 85.0, "%", ""),
    ("QE", 85.0, "%", ""),
    ("Platform altitude", 500.0, "km", "LEO"),
    ("Solar zenith", 30.0, "deg", "Daytime illumination"),
    ("Target reflectance", 0.30, "—", ""),
    ("Integration time", 0.5, "ms", ""),
    ("Dark current", 50.0, "e⁻/s", ""),
    ("Detector temperature", 280.0, "K", ""),
    ("Read noise", 20.0, "e⁻ RMS", ""),
    ("Full well", 30000.0, "e⁻", ""),
    ("System gain", 5.0, "e⁻/DN", ""),
    ("ADC resolution", 12, "bits", ""),
]
r = 4
for name, value, unit, note in ROWS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=value)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=note)
    r += 1

out = HERE / "sarah_cassegrain.xlsx"
wb.save(out)
print(f"Wrote {out}")
