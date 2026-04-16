"""Create Mike's detector datasheet for well capacity optimization.

Three sheets:
  1. "Detector Specs"      — MWIR HgCdTe vendor datasheet
  2. "Optics & Scene"      — optical system and scene parameters
  3. "Trade Requirements"  — performance requirements and sweep definition
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=4):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Detector Specs
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Detector Specs"
ws1["A1"] = "MWIR HgCdTe Detector — Vendor Datasheet"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "640x512 MWIR FPA, 15 µm pitch, large-well variant"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 55

row = 4
row = add_section(ws1, row, "Detector Parameters")
row = add_param(ws1, row, "Format", "640x512", "pixels", "Staring FPA")
row = add_param(ws1, row, "Pixel pitch", 15.0, "µm", "Square pixels")
row = add_param(ws1, row, "Cutoff wavelength", 5.3, "µm", "At 77 K")
row = add_param(ws1, row, "Quantum efficiency (in-band avg)", 72.0, "%",
                "Measured at 77 K, 3.5–5.0 µm average")
row = add_param(ws1, row, "Dark current", 100.0, "e⁻/s",
                "At 77 K operating temperature")
row = add_param(ws1, row, "Operating temperature", 77.0, "K",
                "Stirling cooler, ±0.1 K stability")
row = add_param(ws1, row, "Read noise (post-CDS)", 20.0, "e⁻ RMS",
                "CDS mode, 200-frame ensemble")
row = add_param(ws1, row, "Full well capacity", 2.0e6, "e⁻",
                "Large-well ROIC variant, 90% linearity")
row = add_param(ws1, row, "ADC resolution", 14, "bits", "14-bit ADC")
row = add_param(ws1, row, "System gain", 130.0, "e⁻/DN",
                "Set for 107% FWC at ADC max")
row = add_param(ws1, row, "IPC coupling", 1.0, "%",
                "Per-neighbor, measured via single-pixel reset")

# ---------------------------------------------------------------------------
# Sheet 2: Optics & Scene
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Optics & Scene")
ws2["A1"] = "Optical System & Scene Configuration"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Air surveillance MWIR imager, ground-based"
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 55

row = 4
row = add_section(ws2, row, "Optics")
row = add_param(ws2, row, "Aperture diameter", 20.0, "cm",
                "Cassegrain telescope")
row = add_param(ws2, row, "Focal length", 40.0, "cm", "f/2 for high throughput")
row = add_param(ws2, row, "f-number", 2.0, "—", "Fast optics for sensitivity")
row = add_param(ws2, row, "Optical transmission", 80.0, "%",
                "All elements combined")
row = add_param(ws2, row, "Optics temperature", 20.0, "°C", "Ambient ground-based")
row = add_param(ws2, row, "Number of optical elements", 4, "—",
                "Primary, secondary, fold, window")

row += 1
row = add_section(ws2, row, "Spectral Band")
row = add_param(ws2, row, "Filter cut-on", 3500.0, "nm", "Cold filter at 77 K")
row = add_param(ws2, row, "Filter cut-off", 5000.0, "nm", "Cold filter at 77 K")
row = add_param(ws2, row, "Band center", 4250.0, "nm", "Arithmetic center")

row += 1
row = add_section(ws2, row, "Scene Description")
row = add_param(ws2, row, "Cold target temperature", 200.0, "K",
                "Cold sky background at high elevation angle")
row = add_param(ws2, row, "Hot target temperature", 1500.0, "K",
                "Jet exhaust plume")
row = add_param(ws2, row, "Target emissivity", 1.0, "—",
                "Assume blackbody for both")
row = add_param(ws2, row, "Background temperature", 280.0, "K",
                "Average scene background (sky + terrain)")
row = add_param(ws2, row, "Background emissivity", 1.0, "—",
                "Assume blackbody background")

# ---------------------------------------------------------------------------
# Sheet 3: Trade Requirements
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Trade Requirements")
ws3["A1"] = "Integration Time Trade Study — Requirements"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "Find optimal t_int for max dynamic range without saturation"
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 55

row = 4
row = add_section(ws3, row, "Performance Requirements")
row = add_param(ws3, row, "Minimum SNR (cold target)", 10.0, "—",
                "Detection threshold for 200 K target")
row = add_param(ws3, row, "Maximum well fill (hot target)", 70.0, "%",
                "Stay in linear range for hot target")
row = add_param(ws3, row, "Saturation limit", 90.0, "%",
                "Absolute maximum before FWC clipping")

row += 1
row = add_section(ws3, row, "Sweep Definition")
row = add_param(ws3, row, "Integration time min", 0.001, "ms",
                "1 µs — shortest practical")
row = add_param(ws3, row, "Integration time max", 50.0, "ms",
                "50 ms — longest before smear concerns")
row = add_param(ws3, row, "Number of sweep points", 50, "—",
                "Log-spaced sweep")

row += 1
row = add_section(ws3, row, "Scene Temperatures for Comparison")
headers = ["Temperature [K]", "Description", "", ""]
for c, h in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.border = thin_border
row += 1

scene_temps = [
    (200.0, "Cold sky"),
    (250.0, "High-altitude atmosphere"),
    (280.0, "Cool ambient"),
    (300.0, "Standard ambient"),
    (350.0, "Sun-heated surface"),
    (400.0, "Hot exhaust stack"),
    (500.0, "Incipient fire"),
    (700.0, "Moderate fire"),
    (1000.0, "Hot combustion"),
    (1500.0, "Jet exhaust plume"),
]
for temp, desc in scene_temps:
    ws3.cell(row=row, column=1, value=temp).border = thin_border
    ws3.cell(row=row, column=2, value=desc).border = thin_border
    row += 1

from pathlib import Path
out = Path(__file__).parent / "mike_well_capacity_data.xlsx"
wb.save(out)
print(f"Created {out}")
