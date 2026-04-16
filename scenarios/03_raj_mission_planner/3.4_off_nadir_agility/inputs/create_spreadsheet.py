"""Create Raj's off-nadir agility spreadsheet.

Three sheets:
  1. "Sensor Configuration" — VNIR reconnaissance sensor baseline
  2. "Orbit & Geometry"     — 600 km LEO, solar geometry
  3. "Off-Nadir Sweep"      — 0 to 45 deg in 5 deg steps
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="002E75B6", end_color="002E75B6", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_section(ws, row, title, n_cols=4):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col, value=title if col == 1 else "")
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    return row + 1


def add_param(ws, row, name, value, unit, notes=""):
    ws.cell(row=row, column=1, value=name).border = thin_border
    cell_v = ws.cell(row=row, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal="center")
    cell_u = ws.cell(row=row, column=3, value=unit)
    cell_u.border = thin_border
    cell_u.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=notes).border = thin_border
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Sensor Configuration
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Sensor Configuration"
ws1["A1"] = "VNIR Reconnaissance Sensor — Baseline Configuration"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = "600 km SSO, agile pointing, pushbroom imager"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 55

row = 4
row = add_section(ws1, row, "Optics")
row = add_param(ws1, row, "Aperture diameter", 35.0, "cm", "TMA telescope")
row = add_param(ws1, row, "Focal length", 350.0, "cm", "Effective focal length")
row = add_param(ws1, row, "f-number", 10.0, "—", "Derived from D and f")
row = add_param(ws1, row, "Optical transmission", 75.0, "%", "All elements combined")
row = add_param(ws1, row, "Optics temperature", 20.0, "°C", "Orbital mean")
row = add_param(ws1, row, "Central obscuration", 20.0, "%", "Secondary mirror")
row = add_param(ws1, row, "WFE RMS", 0.05, "waves", "At 633 nm reference")
row += 1

row = add_section(ws1, row, "Spectral Band")
row = add_param(ws1, row, "Filter min", 450.0, "nm", "Panchromatic band")
row = add_param(ws1, row, "Filter max", 900.0, "nm", "Panchromatic band")
row = add_param(ws1, row, "Band center", 675.0, "nm", "Arithmetic center")
row += 1

row = add_section(ws1, row, "Detector")
row = add_param(ws1, row, "Pixel pitch", 8.0, "µm", "Square pixels, Si CCD/CMOS")
row = add_param(ws1, row, "Fill factor", 100.0, "%", "Back-illuminated CMOS")
row = add_param(ws1, row, "Quantum efficiency", 80.0, "%", "Broadband PAN average")
row = add_param(ws1, row, "Dark current", 30.0, "e⁻/s", "At 263 K")
row = add_param(ws1, row, "Operating temperature", 263.0, "K", "TEC-cooled")
row = add_param(ws1, row, "Read noise", 6.0, "e⁻ RMS", "CDS readout")
row = add_param(ws1, row, "Full well capacity", 80000.0, "e⁻", "90% linearity")
row = add_param(ws1, row, "ADC bits", 12, "—", "")
row = add_param(ws1, row, "System gain", 5.0, "e⁻/DN", "")
row = add_param(ws1, row, "IPC coupling", 1.5, "%", "Measured")
row = add_param(ws1, row, "Integration time", 0.5, "ms",
                "Short for pushbroom at 7 km/s ground speed")

# ---------------------------------------------------------------------------
# Sheet 2: Orbit & Geometry
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Orbit & Geometry")
ws2["A1"] = "Orbit and Observation Geometry"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "600 km sun-synchronous orbit, 10:30 AM LTAN"
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 55

row = 4
row = add_section(ws2, row, "Orbit")
row = add_param(ws2, row, "Altitude", 600.0, "km", "Circular LEO")
row = add_param(ws2, row, "Inclination", 97.8, "deg", "Sun-synchronous")
row = add_param(ws2, row, "LTAN", "10:30", "—", "Descending node local time")
row = add_param(ws2, row, "Ground speed", 6.9, "km/s", "Approximate at 600 km")
row += 1

row = add_section(ws2, row, "Solar Geometry")
row = add_param(ws2, row, "Solar zenith angle", 30.0, "deg",
                "Midlatitude summer, 10:30 AM local")
row = add_param(ws2, row, "Solar azimuth", 0.0, "deg",
                "Relative to line of sight")
row += 1

row = add_section(ws2, row, "Target")
row = add_param(ws2, row, "Target reflectance", 0.15, "—", "Typical mixed scene")
row = add_param(ws2, row, "Background reflectance", 0.20, "—", "Surrounding terrain")
row = add_param(ws2, row, "Target temperature", 300.0, "K", "Ground-level")
row = add_param(ws2, row, "Background temperature", 295.0, "K", "")

# ---------------------------------------------------------------------------
# Sheet 3: Off-Nadir Sweep
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Off-Nadir Sweep")
ws3["A1"] = "Off-Nadir Angle Sweep Definition"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = "0 to 45 degrees in 5-degree steps"
ws3.column_dimensions["A"].width = 25

row = 4
ws3.cell(row=row, column=1, value="Off-Nadir Angle [deg]").font = header_font
ws3.cell(row=row, column=1).border = thin_border

angles_deg = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45]
for i, angle in enumerate(angles_deg, row + 1):
    ws3.cell(row=i, column=1, value=angle).border = thin_border

# Notes
note_row = row + len(angles_deg) + 2
ws3.cell(row=note_row, column=1, value="Notes:").font = Font(bold=True)
ws3.cell(row=note_row + 1, column=1,
         value="Off-nadir angle = angle between nadir direction and line of sight.")
ws3.cell(row=note_row + 2, column=1,
         value="RADIANT parameter: geometry.path_zenith_rad (convert deg -> rad).")
ws3.cell(row=note_row + 3, column=1,
         value="Slant range increases as 1/cos(theta) for flat-Earth approximation.")
ws3.cell(row=note_row + 4, column=1,
         value="Earth curvature correction applied for angles > 80 deg.")

from pathlib import Path
out = Path(__file__).parent / "raj_off_nadir_data.xlsx"
wb.save(out)
print(f"Created {out}")
